# 请给出上面6步分别干了什么事情，代码逻辑是什么。只给文本
'''
# 表格重构六步流程说明

## 第1步：准备数据
**功能**：传入已加载的OCR和LLM数据
**逻辑**：
- 接收两个参数：`ocr_result`（OCR识别结果）和 `llm_result`（LLM分析结果）
- 这两个数据已在外部加载好，直接传入使用
- 不涉及文件读取，只做数据验证

## 第2步：提取表格数据
**功能**：从原始数据中提取表格信息
**逻辑**：
- 从 `ocr_result` 提取 `tables_result` 字段，得到OCR表格列表
- 从 `llm_result` 提取 `tables_structure.tables` 字段，得到LLM表格结构列表
- 如果结构不同（如 `tables` 在根目录），做兼容性处理
- 返回两个列表：`ocr_tables` 和 `llm_tables`

## 第3步：合并OCR表格数据
**功能**：根据LLM指示合并多个OCR表格
**逻辑**：
1. 读取LLM表格结构的 `ocr_tables` 字段，确定要合并哪些OCR表格
2. 遍历这些OCR表格，提取它们的单元格数据
3. 垂直拼接表格：调整后续表格的行号（加上行偏移量）
4. 表格间留空行分隔（`row_offset = max_row_in_table + 2`）
5. 返回合并后的数据，包含所有单元格、表头信息和表格ID

## 第4步：创建基础数据表格
**功能**：用OCR数据创建基础表格结构
**逻辑**：
1. 从合并数据中获取所有单元格
2. 计算最大行索引和列索引（`max_row`, `max_col`）
3. 创建二维表格：行数 = `max_row + 1`，列数 = `max_col + 1`
4. 填充数据：将每个单元格的值放入对应的 `[row_start][col_start]` 位置
5. 对于合并单元格，只填充左上角位置
6. 返回纯数据表格（无表头）

## 第5步：添加列标题（从右向左）
**功能**：用LLM列标题填充表格顶部
**逻辑**：
1. 确定最终列数 = LLM列标题数（这是最终列数）
2. 调整表格列数：
   - 如果OCR列数 > LLM列数：删除左侧多余列（保留右侧数据列）
   - 如果OCR列数 < LLM列数：在左侧补充空列
3. 在表格顶部插入一行用于列标题
4. 从右向左填充：最后一个列标题填充到最后一列，依次类推
5. 左侧预留的列将用于行标题

## 第6步：智能匹配填充行表头
**功能**：将LLM行表头匹配到正确的数据行
**逻辑**：
1. 分析表格结构：从列标题行确定左侧空列数
2. 提取每个LLM行表头的最后一层文本（`>>` 分割后的最后部分）
3. 在数据行中搜索匹配：
   - 清理文本（去除空格、换行、特殊符号）
   - 计算文本相似度（包含关系、公共字符比例）
   - 优先搜索左侧列和第一个数据列
4. 找到最佳匹配行后：
   - 将完整的LLM行表头放入该行的第0列
   - 清空该行其他左侧列
5. 对于未匹配的行表头，按顺序填充到空行
6. 返回最终表格，包含完整的表头和数据

## 核心原则总结
1. **LLM列数为准**：最终列数以LLM列标题数量为准
2. **右侧数据优先**：调整列数时只动左侧列，右侧数据列绝对保留
3. **从右向左填充**：列标题从右向左确定位置
4. **智能行匹配**：用文本相似度匹配行表头，不按简单顺序填充
5. **保留原始数据**：只添加表头，不覆盖原始数据内容
'''



def extract_table_data(ocr_result, llm_result):
    """
    从数据中提取表格信息
    """
    # 提取OCR表格
    ocr_tables = ocr_result.get('tables_result', [])
    print(f"OCR表格数: {len(ocr_tables)}")

    # 提取LLM表格结构
    llm_tables = []
    if 'tables_structure' in llm_result:
        llm_tables = llm_result['tables_structure'].get('tables', [])
    elif 'tables' in llm_result:
        llm_tables = llm_result['tables']
    print(f"LLM表格结构数: {len(llm_tables)}")

    return ocr_tables, llm_tables


def merge_ocr_table_data(ocr_tables, llm_table_info):
    """
    将多个OCR表格的数据合并成一个表格的数据
    LLM指定哪些OCR表格属于同一个逻辑表格
    """
    # 获取要合并的OCR表格索引
    ocr_table_indices = llm_table_info.get('ocr_tables', [])
    print(f"要合并的OCR表格索引: {ocr_table_indices}")

    # 收集所有单元格
    all_cells = []
    row_offset = 0  # 行偏移量，用于拼接表格

    for idx in ocr_table_indices:
        if idx >= len(ocr_tables):
            print(f"警告: OCR表格索引 {idx} 超出范围")
            continue

        table = ocr_tables[idx]
        cells = table.get('body', [])

        # 调整行号（表格拼接时行号要累加）
        for cell in cells:
            adjusted_cell = cell.copy()
            adjusted_cell['row_start'] += row_offset
            adjusted_cell['row_end'] += row_offset
            all_cells.append(adjusted_cell)

        # 更新行偏移量：找到这个表格的最大行号
        if cells:
            max_row_in_table = max(cell['row_end'] for cell in cells)
            row_offset = max_row_in_table + 2  # +2留一个空行分隔

    return {
        'cells': all_cells,  # 合并后的所有单元格
        'headers': llm_table_info.get('headers', {}),  # 表头信息
        'table_id': llm_table_info.get('id', '1')  # 表格ID
    }



def create_base_data_table(merged_data):
    """
    创建基础数据表格
    检查每行列数是否一致
    """
    cells = merged_data['cells']

    if not cells:
        return []

    # 找出最大行列
    max_row = 0
    max_col = 0
    for cell in cells:
        max_row = max(max_row, cell['row_end'])
        max_col = max(max_col, cell['col_end'])

    num_rows = max_row
    num_cols = max_col

    # 检查每行的列数
    print("检查每行列数:")
    row_col_counts = {}
    for cell in cells:
        row = cell['row_start']
        col_end = cell['col_end']

        if row not in row_col_counts:
            row_col_counts[row] = []
        row_col_counts[row].append(col_end)

    # 找出每行的最大列号
    row_max_cols = {}
    for row, col_ends in row_col_counts.items():
        row_max_cols[row] = max(col_ends)
        print(f"  第{row}行: 最大列索引={row_max_cols[row]}")

    # 检查是否所有行的最大列索引一致
    all_max_cols = list(row_max_cols.values())
    if len(set(all_max_cols)) > 1:
        print(f"警告: 不同行的列数不一致: {all_max_cols}")

    print(f"表格: {num_rows}行 × {num_cols}列")

    # 创建表格
    table = []
    for r in range(num_rows):
        table.append([None] * num_cols)

    # 填充数据
    for cell in cells:
        value = cell['words']
        row_idx = cell['row_start']
        col_idx = cell['col_start']

        if row_idx < num_rows and col_idx < num_cols:
            table[row_idx][col_idx] = value

    return table


def add_column_headers(table, col_headers):
    """
    最终版：LLM列数是最终列数，从右向左确定列
    """
    if not col_headers:
        print("无列标题")
        return table

    if not table:
        print("表格为空")
        return table

    current_cols = len(table[0])  # OCR表格的列数
    target_cols = len(col_headers)  # LLM的最终列数

    print(f"OCR表格列数: {current_cols}")
    print(f"LLM目标列数: {target_cols}")
    print(f"LLM列标题: {col_headers}")

    # 情况1：OCR列数 > LLM列数（需要删除左侧多余的列）
    if current_cols > target_cols:
        excess_cols = current_cols - target_cols
        print(f"需要删除左侧{excess_cols}列")

        # 删除左侧多余的列（从每行删除）
        for i in range(len(table)):
            # 保留右侧的target_cols列，删除左侧的excess_cols列
            table[i] = table[i][excess_cols:]

        current_cols = target_cols  # 更新列数

    # 情况2：OCR列数 < LLM列数（需要补充左侧空列）
    elif current_cols < target_cols:
        needed_cols = target_cols - current_cols
        print(f"需要补充左侧{needed_cols}列")

        # 在每行左侧补充空列
        for i in range(len(table)):
            table[i] = [None] * needed_cols + table[i]

        current_cols = target_cols  # 更新列数

    # 现在表格列数 = LLM列数
    print(f"调整后表格列数: {current_cols}")

    # 在顶部添加一行用于列标题
    table.insert(0, [None] * current_cols)

    # 从右向左填充列标题
    col_headers_copy = col_headers.copy()
    for col in range(current_cols - 1, -1, -1):
        if col_headers_copy:
            table[0][col] = col_headers_copy.pop()

    print("列标题填充完成")
    print(f"第0行（列标题）: {table[0]}")

    return table


def clean_text_for_matching(text):
    """清理文本用于匹配"""
    if not text:
        return ""
    text = str(text)
    # 替换换行、空格
    text = text.replace('\n', '').replace('\r', '').replace(' ', '')
    # 去掉特殊符号，保留中文、数字、字母
    cleaned = ''.join(c for c in text if c.isalnum() or '\u4e00-\u9fff' in c)
    return cleaned


def calculate_similarity(text1, text2):
    """计算两个文本的相似度"""
    t1 = clean_text_for_matching(text1)
    t2 = clean_text_for_matching(text2)

    if not t1 or not t2:
        return 0

    # 完全相等
    if t1 == t2:
        return 1.0

    # 包含关系
    if t1 in t2 or t2 in t1:
        return 0.9

    # 公共字符比例
    common_chars = set(t1) & set(t2)
    if not common_chars:
        return 0

    # 计算Jaccard相似度
    union_chars = set(t1) | set(t2)
    return len(common_chars) / len(union_chars)


def find_best_match_row(target_text, table, start_row=1, search_cols=None):
    """
    在表格中查找与目标文本最匹配的行
    """
    if not target_text or not table:
        return -1, 0.0

    best_row = -1
    best_score = 0.0

    # 确定搜索列范围
    if search_cols is None:
        # 默认搜索所有列
        search_cols = list(range(len(table[0])))

    for row_idx in range(start_row, len(table)):
        row_score = 0.0

        for col_idx in search_cols:
            if col_idx >= len(table[row_idx]):
                continue

            cell_text = table[row_idx][col_idx]
            if not cell_text:
                continue

            # 计算相似度
            similarity = calculate_similarity(target_text, str(cell_text))
            if similarity > row_score:
                row_score = similarity

        # 更新最佳匹配
        if row_score > best_score:
            best_score = row_score
            best_row = row_idx

    return best_row, best_score



def add_row_headers_intelligent(table, row_headers):
    """
    智能匹配LLM行表头到数据行 - 两阶段匹配
    """
    if not table or not row_headers:
        print("表格或行表头为空")
        return table

    print("=== 第6步：智能匹配行表头（两阶段） ===")
    print(f"表格: {len(table)}行 × {len(table[0])}列")
    print(f"LLM行表头数: {len(row_headers)}")

    # 1. 分析表格结构
    left_empty_cols = 0
    for title in table[0]:
        if not title or str(title).strip() == '':
            left_empty_cols += 1
        else:
            break

    print(f"左侧空列数: {left_empty_cols}")

    # 确定搜索列范围
    if left_empty_cols > 0:
        search_columns = list(range(left_empty_cols))
        if left_empty_cols < len(table[0]):
            search_columns.append(left_empty_cols)
    else:
        search_columns = list(range(min(3, len(table[0]))))

    print(f"搜索列范围: {search_columns}")

    # ========== 第一阶段：自由匹配 ==========
    print("\n--- 第一阶段：自由匹配 ---")
    first_pass_results = []
    used_rows = set()

    for header_idx, llm_header in enumerate(row_headers):
        if '>>' in llm_header:
            target_text = llm_header.split('>>')[-1]
        else:
            target_text = llm_header

        print(f"  表头[{header_idx}]: '{target_text}'")

        # 自由匹配：搜索所有未使用的行
        best_row = -1
        best_score = 0.0
        best_cell_text = ""

        for row_idx in range(1, len(table)):
            if row_idx in used_rows:
                continue  # 跳过已使用的行

            # 计算该行的最佳匹配分数
            row_best_score = 0.0
            cell_text = ""

            for col_idx in search_columns:
                if col_idx >= len(table[row_idx]):
                    continue

                cell_val = table[row_idx][col_idx]
                if not cell_val:
                    continue

                score = calculate_similarity(target_text, str(cell_val))
                if score > row_best_score:
                    row_best_score = score
                    cell_text = str(cell_val)

            if row_best_score > best_score:
                best_score = row_best_score
                best_row = row_idx
                best_cell_text = cell_text

        first_pass_results.append({
            'header_idx': header_idx,
            'llm_header': llm_header,
            'target_text': target_text,
            'matched_row': best_row,
            'score': best_score,
            'matched_cell': best_cell_text
        })

        if best_row != -1:
            used_rows.add(best_row)
            print(f"    匹配到行{best_row}, 分数:{best_score:.2f}, 单元格:'{best_cell_text[:20]}...'")
        else:
            print(f"    未匹配")

    # ========== 第二阶段：顺序检查和调整 ==========
    print("\n--- 第二阶段：顺序检查和调整 ---")

    # 检查第一阶段匹配的顺序
    last_matched_row = 0  # 上一行匹配的行号
    adjusted_results = []
    need_adjustments = False

    for i, result in enumerate(first_pass_results):
        current_row = result['matched_row']

        if current_row == -1:
            # 未匹配，标记为需要顺序填充
            adjusted_results.append(result.copy())
            continue

        # 检查顺序：当前行应该 >= 上一行
        if current_row < last_matched_row:
            print(f"  顺序问题: LLM[{i}]匹配到行{current_row}, 但上一行是{last_matched_row}")
            need_adjustments = True

            # 重新搜索：必须在 last_matched_row+1 之后
            new_best_row = -1
            new_best_score = 0.0
            new_cell_text = ""

            target_text = result['target_text']
            for row_idx in range(last_matched_row + 1, len(table)):
                if row_idx in used_rows:
                    continue

                row_best_score = 0.0
                cell_text = ""

                for col_idx in search_columns:
                    if col_idx >= len(table[row_idx]):
                        continue

                    cell_val = table[row_idx][col_idx]
                    if not cell_val:
                        continue

                    score = calculate_similarity(target_text, str(cell_val))
                    if score > row_best_score:
                        row_best_score = score
                        cell_text = str(cell_val)

                if row_best_score > new_best_score:
                    new_best_score = row_best_score
                    new_best_row = row_idx
                    new_cell_text = cell_text

            if new_best_row != -1:
                # 移除旧的行，添加新的
                used_rows.discard(current_row)
                used_rows.add(new_best_row)

                adjusted_result = result.copy()
                adjusted_result['matched_row'] = new_best_row
                adjusted_result['score'] = new_best_score
                adjusted_result['matched_cell'] = new_cell_text
                adjusted_result['adjusted'] = True

                adjusted_results.append(adjusted_result)
                last_matched_row = new_best_row

                print(f"    调整为行{new_best_row}, 分数:{new_best_score:.2f}")
            else:
                # 找不到符合条件的行，保持原样但标记问题
                adjusted_result = result.copy()
                adjusted_result['has_order_issue'] = True
                adjusted_results.append(adjusted_result)
                last_matched_row = current_row
                print(f"    无法调整，保持行{current_row}（顺序问题）")
        else:
            # 顺序正确，保持原样
            adjusted_results.append(result.copy())
            last_matched_row = current_row

    # ========== 第三阶段：填充表格 ==========
    print("\n--- 第三阶段：填充表格 ---")

    # 清空第0列（准备填充行表头）
    for row_idx in range(1, len(table)):
        table[row_idx][0] = None

    # 按顺序填充匹配成功的行表头
    matched_count = 0
    current_data_row = 1  # 当前数据行指针

    for i, result in enumerate(adjusted_results):
        if result['matched_row'] != -1 and result.get('score', 0) > 0.3:
            # 使用匹配到的行
            target_row = result['matched_row']
            table[target_row][0] = result['llm_header']
            matched_count += 1

            # 清空其他左侧列
            for col in range(1, left_empty_cols):
                if col < len(table[target_row]):
                    table[target_row][col] = None

            print(f"  LLM[{i}] → 行{target_row}: '{result['llm_header'][:30]}...'")
            current_data_row = target_row + 1
        else:
            # 未匹配或匹配分数低，按顺序填充
            # 找下一个可用的行
            while current_data_row < len(table) and current_data_row in used_rows:
                current_data_row += 1

            if current_data_row < len(table):
                table[current_data_row][0] = result['llm_header']
                matched_count += 1
                print(f"  LLM[{i}] → 行{current_data_row}: '{result['llm_header'][:30]}...'（顺序填充）")
                current_data_row += 1
            else:
                print(f"  LLM[{i}] 无法填充: 表格行数不足")

    print(f"\n匹配结果: {matched_count}/{len(row_headers)} 个行表头已填充")

    return table


def save_tables_to_excel(tables_data, output_file):
    """
    将多个表格保存到Excel，每个表格一个Sheet
    tables_data: 列表，每个元素是一个表格的完整数据
    """
    import openpyxl
    from openpyxl.styles import Alignment, Font, Border, Side
    from openpyxl.utils import get_column_letter

    print(f"\n=== 第7步：保存到Excel ===")
    print(f"要保存{len(tables_data)}个表格到: {output_file}")

    # 创建新的工作簿
    wb = openpyxl.Workbook()

    # 删除默认创建的Sheet
    if 'Sheet' in wb.sheetnames:
        default_sheet = wb['Sheet']
        wb.remove(default_sheet)

    # 处理每个表格
    for table_idx, table in enumerate(tables_data):
        if not table:
            print(f"表格{table_idx}为空，跳过")
            continue

        # 创建Sheet名称（Excel限制31字符）
        sheet_name = f"Table{table_idx + 1}"
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]

        # 检查Sheet名称是否重复
        original_name = sheet_name
        counter = 1
        while sheet_name in wb.sheetnames:
            sheet_name = f"{original_name}_{counter}"
            counter += 1

        # 创建工作表
        ws = wb.create_sheet(title=sheet_name)

        # 获取表格尺寸
        num_rows = len(table)
        num_cols = len(table[0]) if num_rows > 0 else 0

        print(f"  表格{table_idx + 1}: {num_rows}行 × {num_cols}列 -> Sheet: '{sheet_name}'")

        # 填充数据到Excel
        for r in range(num_rows):
            for c in range(num_cols):
                cell_value = table[r][c]

                # Excel行号从1开始，列号从1开始
                excel_row = r + 1
                excel_col = c + 1

                # 写入值
                ws.cell(row=excel_row, column=excel_col, value=cell_value)

                # 获取单元格对象设置样式
                cell_obj = ws.cell(row=excel_row, column=excel_col)

                # 设置对齐方式
                if r == 0:  # 第0行是表头
                    cell_obj.font = Font(bold=True)
                    cell_obj.alignment = Alignment(horizontal='center', vertical='center')
                    # 浅灰色背景
                    from openpyxl.styles import PatternFill
                    cell_obj.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                elif c == 0:  # 第0列是行表头
                    cell_obj.alignment = Alignment(horizontal='left', vertical='center')
                else:  # 数据单元格
                    cell_obj.alignment = Alignment(horizontal='center', vertical='center')

        # 设置列宽
        for col in range(1, num_cols + 1):
            col_letter = get_column_letter(col)

            # 根据内容调整列宽
            max_length = 0
            for row in range(1, min(num_rows + 1, 50)):  # 只检查前50行
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    # 计算文本长度（中文算2个字符）
                    text_len = 0
                    for char in str(cell_value):
                        if '\u4e00-\u9fff' in char:  # 中文字符
                            text_len += 2
                        else:
                            text_len += 1
                    max_length = max(max_length, text_len)

            # 设置列宽（最小8，最大50）
            width = min(max(max_length + 2, 8), 50)
            ws.column_dimensions[col_letter].width = width

        # 添加边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 为整个表格添加边框
        for row in ws.iter_rows(min_row=1, max_row=num_rows, min_col=1, max_col=num_cols):
            for cell in row:
                cell.border = thin_border

        # 冻结首行和首列（方便查看）
        ws.freeze_panes = 'B2'  # 冻结第1行和第1列

    # 如果没有表格，创建一个空Sheet
    if len(wb.sheetnames) == 0:
        ws = wb.create_sheet(title="空表格")
        ws['A1'] = "无表格数据"

    # 保存文件
    try:
        wb.save(output_file)
        print(f"\n✅ Excel文件保存成功: {output_file}")
        print(f"   共保存了 {len(wb.sheetnames)} 个工作表")

        # 显示Sheet列表
        print("   Sheet列表:")
        for i, sheet_name in enumerate(wb.sheetnames, 1):
            ws = wb[sheet_name]
            max_row = ws.max_row
            max_column = ws.max_column
            print(f"     {i}. {sheet_name}: {max_row}行 × {max_column}列")

        return True

    except Exception as e:
        print(f"\n❌ 保存Excel文件失败: {str(e)}")
        return False




# 使用
if __name__ == '__main__':
    from pprint import  pprint

    ocr_result = {'tables_result': [{'header': [{'location': [{'x': 1106, 'y': 531}, {'x': 1213, 'y': 531}, {'x': 1213, 'y': 573}, {'x': 1106, 'y': 573}], 'words': '2024年'}, {'location': [{'x': 1306, 'y': 531}, {'x': 1403, 'y': 531}, {'x': 1403, 'y': 571}, {'x': 1306, 'y': 571}], 'words': '2023年'}], 'body': [{'col_end': 1, 'row_end': 1, 'cell_location': [{'x': 176, 'y': 581}, {'x': 992, 'y': 581}, {'x': 992, 'y': 700}, {'x': 176, 'y': 700}], 'row_start': 0, 'col_start': 0, 'words': '全年业绩'}, {'col_end': 2, 'row_end': 1, 'cell_location': [{'x': 992, 'y': 581}, {'x': 1257, 'y': 581}, {'x': 1257, 'y': 700}, {'x': 992, 'y': 700}], 'row_start': 0, 'col_start': 1, 'words': ''}, {'col_end': 3, 'row_end': 1, 'cell_location': [{'x': 1257, 'y': 581}, {'x': 1460, 'y': 580}, {'x': 1460, 'y': 700}, {'x': 1257, 'y': 700}], 'row_start': 0, 'col_start': 2, 'words': ''}, {'col_end': 4, 'row_end': 1, 'cell_location': [{'x': 1460, 'y': 580}, {'x': 1642, 'y': 580}, {'x': 1643, 'y': 700}, {'x': 1460, 'y': 700}], 'row_start': 0, 'col_start': 3, 'words': ''}, {'col_end': 5, 'row_end': 1, 'cell_location': [{'x': 1642, 'y': 580}, {'x': 1824, 'y': 580}, {'x': 1824, 'y': 700}, {'x': 1643, 'y': 700}], 'row_start': 0, 'col_start': 4, 'words': ''}, {'col_end': 6, 'row_end': 1, 'cell_location': [{'x': 1824, 'y': 580}, {'x': 2015, 'y': 580}, {'x': 2015, 'y': 700}, {'x': 1824, 'y': 700}], 'row_start': 0, 'col_start': 5, 'words': ''}, {'col_end': 7, 'row_end': 1, 'cell_location': [{'x': 2015, 'y': 580}, {'x': 2191, 'y': 580}, {'x': 2191, 'y': 700}, {'x': 2015, 'y': 700}], 'row_start': 0, 'col_start': 6, 'words': ''}, {'col_end': 1, 'row_end': 2, 'cell_location': [{'x': 176, 'y': 700}, {'x': 992, 'y': 700}, {'x': 992, 'y': 767}, {'x': 176, 'y': 767}], 'row_start': 1, 'col_start': 0, 'words': '营业收入'}, {'col_end': 2, 'row_end': 2, 'cell_location': [{'x': 992, 'y': 700}, {'x': 1257, 'y': 700}, {'x': 1257, 'y': 767}, {'x': 992, 'y': 767}], 'row_start': 1, 'col_start': 1, 'words': '750,151'}, {'col_end': 3, 'row_end': 2, 'cell_location': [{'x': 1257, 'y': 700}, {'x': 1460, 'y': 700}, {'x': 1460, 'y': 767}, {'x': 1257, 'y': 767}], 'row_start': 1, 'col_start': 2, 'words': '769,736'}, {'col_end': 4, 'row_end': 2, 'cell_location': [{'x': 1460, 'y': 700}, {'x': 1643, 'y': 700}, {'x': 1642, 'y': 768}, {'x': 1460, 'y': 767}], 'row_start': 1, 'col_start': 3, 'words': '(2.54)'}, {'col_end': 5, 'row_end': 2, 'cell_location': [{'x': 1643, 'y': 700}, {'x': 1824, 'y': 700}, {'x': 1824, 'y': 768}, {'x': 1642, 'y': 768}], 'row_start': 1, 'col_start': 4, 'words': '783,760'}, {'col_end': 6, 'row_end': 2, 'cell_location': [{'x': 1824, 'y': 700}, {'x': 2015, 'y': 700}, {'x': 2015, 'y': 768}, {'x': 1824, 'y': 768}], 'row_start': 1, 'col_start': 5, 'words': '824,246'}, {'col_end': 7, 'row_end': 2, 'cell_location': [{'x': 2015, 'y': 700}, {'x': 2191, 'y': 700}, {'x': 2191, 'y': 768}, {'x': 2015, 'y': 768}], 'row_start': 1, 'col_start': 6, 'words': '755,858'}, {'col_end': 1, 'row_end': 3, 'cell_location': [{'x': 176, 'y': 767}, {'x': 992, 'y': 767}, {'x': 992, 'y': 820}, {'x': 176, 'y': 820}], 'row_start': 2, 'col_start': 0, 'words': '利息净收入'}, {'col_end': 2, 'row_end': 3, 'cell_location': [{'x': 992, 'y': 767}, {'x': 1257, 'y': 767}, {'x': 1257, 'y': 820}, {'x': 992, 'y': 820}], 'row_start': 2, 'col_start': 1, 'words': '589,882'}, {'col_end': 3, 'row_end': 3, 'cell_location': [{'x': 1257, 'y': 767}, {'x': 1460, 'y': 767}, {'x': 1460, 'y': 820}, {'x': 1257, 'y': 820}], 'row_start': 2, 'col_start': 2, 'words': '617,233'}, {'col_end': 4, 'row_end': 3, 'cell_location': [{'x': 1460, 'y': 767}, {'x': 1642, 'y': 768}, {'x': 1642, 'y': 821}, {'x': 1460, 'y': 820}], 'row_start': 2, 'col_start': 3, 'words': '(4.43)'}, {'col_end': 5, 'row_end': 3, 'cell_location': [{'x': 1642, 'y': 768}, {'x': 1824, 'y': 768}, {'x': 1824, 'y': 821}, {'x': 1642, 'y': 821}], 'row_start': 2, 'col_start': 4, 'words': '643,669'}, {'col_end': 6, 'row_end': 3, 'cell_location': [{'x': 1824, 'y': 768}, {'x': 2015, 'y': 768}, {'x': 2015, 'y': 822}, {'x': 1824, 'y': 821}], 'row_start': 2, 'col_start': 5, 'words': '605,420'}, {'col_end': 7, 'row_end': 3, 'cell_location': [{'x': 2015, 'y': 768}, {'x': 2191, 'y': 768}, {'x': 2191, 'y': 822}, {'x': 2015, 'y': 822}], 'row_start': 2, 'col_start': 6, 'words': '575,909'}, {'col_end': 1, 'row_end': 4, 'cell_location': [{'x': 176, 'y': 820}, {'x': 992, 'y': 820}, {'x': 992, 'y': 872}, {'x': 176, 'y': 872}], 'row_start': 3, 'col_start': 0, 'words': '手续费及佣金净收入'}, {'col_end': 2, 'row_end': 4, 'cell_location': [{'x': 992, 'y': 820}, {'x': 1257, 'y': 820}, {'x': 1257, 'y': 872}, {'x': 992, 'y': 872}], 'row_start': 3, 'col_start': 1, 'words': '104,928'}, {'col_end': 3, 'row_end': 4, 'cell_location': [{'x': 1257, 'y': 820}, {'x': 1460, 'y': 820}, {'x': 1460, 'y': 872}, {'x': 1257, 'y': 872}], 'row_start': 3, 'col_start': 2, 'words': '115,746'}, {'col_end': 4, 'row_end': 4, 'cell_location': [{'x': 1460, 'y': 820}, {'x': 1642, 'y': 821}, {'x': 1642, 'y': 873}, {'x': 1460, 'y': 872}], 'row_start': 3, 'col_start': 3, 'words': '(9.35)'}, {'col_end': 5, 'row_end': 4, 'cell_location': [{'x': 1642, 'y': 821}, {'x': 1824, 'y': 821}, {'x': 1824, 'y': 874}, {'x': 1642, 'y': 873}], 'row_start': 3, 'col_start': 4, 'words': '116,085'}, {'col_end': 6, 'row_end': 4, 'cell_location': [{'x': 1824, 'y': 821}, {'x': 2015, 'y': 822}, {'x': 2015, 'y': 874}, {'x': 1824, 'y': 874}], 'row_start': 3, 'col_start': 5, 'words': '121,492'}, {'col_end': 7, 'row_end': 4, 'cell_location': [{'x': 2015, 'y': 822}, {'x': 2191, 'y': 822}, {'x': 2191, 'y': 874}, {'x': 2015, 'y': 874}], 'row_start': 3, 'col_start': 6, 'words': '114,582'}, {'col_end': 1, 'row_end': 5, 'cell_location': [{'x': 176, 'y': 872}, {'x': 992, 'y': 872}, {'x': 992, 'y': 924}, {'x': 176, 'y': 924}], 'row_start': 4, 'col_start': 0, 'words': '其他非利息收入'}, {'col_end': 2, 'row_end': 5, 'cell_location': [{'x': 992, 'y': 872}, {'x': 1257, 'y': 872}, {'x': 1257, 'y': 924}, {'x': 992, 'y': 924}], 'row_start': 4, 'col_start': 1, 'words': '55,341'}, {'col_end': 3, 'row_end': 5, 'cell_location': [{'x': 1257, 'y': 872}, {'x': 1460, 'y': 872}, {'x': 1460, 'y': 924}, {'x': 1257, 'y': 924}], 'row_start': 4, 'col_start': 2, 'words': '36,757'}, {'col_end': 4, 'row_end': 5, 'cell_location': [{'x': 1460, 'y': 872}, {'x': 1642, 'y': 873}, {'x': 1642, 'y': 925}, {'x': 1460, 'y': 924}], 'row_start': 4, 'col_start': 3, 'words': '50.56'}, {'col_end': 5, 'row_end': 5, 'cell_location': [{'x': 1642, 'y': 873}, {'x': 1824, 'y': 874}, {'x': 1824, 'y': 925}, {'x': 1642, 'y': 925}], 'row_start': 4, 'col_start': 4, 'words': '24,006'}, {'col_end': 6, 'row_end': 5, 'cell_location': [{'x': 1824, 'y': 874}, {'x': 2015, 'y': 874}, {'x': 2015, 'y': 925}, {'x': 1824, 'y': 925}], 'row_start': 4, 'col_start': 5, 'words': '97,334'}, {'col_end': 7, 'row_end': 5, 'cell_location': [{'x': 2015, 'y': 874}, {'x': 2191, 'y': 874}, {'x': 2191, 'y': 925}, {'x': 2015, 'y': 925}], 'row_start': 4, 'col_start': 6, 'words': '65,367'}, {'col_end': 1, 'row_end': 6, 'cell_location': [{'x': 176, 'y': 924}, {'x': 992, 'y': 924}, {'x': 992, 'y': 976}, {'x': 176, 'y': 976}], 'row_start': 5, 'col_start': 0, 'words': '业务及管理费'}, {'col_end': 2, 'row_end': 6, 'cell_location': [{'x': 992, 'y': 924}, {'x': 1257, 'y': 924}, {'x': 1257, 'y': 976}, {'x': 992, 'y': 976}], 'row_start': 5, 'col_start': 1, 'words': '(214,312)'}, {'col_end': 3, 'row_end': 6, 'cell_location': [{'x': 1257, 'y': 924}, {'x': 1460, 'y': 924}, {'x': 1460, 'y': 977}, {'x': 1257, 'y': 976}], 'row_start': 5, 'col_start': 2, 'words': '(210,088)'}, {'col_end': 4, 'row_end': 6, 'cell_location': [{'x': 1460, 'y': 924}, {'x': 1642, 'y': 925}, {'x': 1642, 'y': 977}, {'x': 1460, 'y': 977}], 'row_start': 5, 'col_start': 3, 'words': '2.01'}, {'col_end': 5, 'row_end': 6, 'cell_location': [{'x': 1642, 'y': 925}, {'x': 1824, 'y': 925}, {'x': 1824, 'y': 977}, {'x': 1642, 'y': 977}], 'row_start': 5, 'col_start': 4, 'words': '(210,896)'}, {'col_end': 6, 'row_end': 6, 'cell_location': [{'x': 1824, 'y': 925}, {'x': 2015, 'y': 925}, {'x': 2015, 'y': 978}, {'x': 1824, 'y': 977}], 'row_start': 5, 'col_start': 5, 'words': '(209,864)'}, {'col_end': 7, 'row_end': 6, 'cell_location': [{'x': 2015, 'y': 925}, {'x': 2191, 'y': 925}, {'x': 2191, 'y': 977}, {'x': 2015, 'y': 978}], 'row_start': 5, 'col_start': 6, 'words': '(179,308)'}, {'col_end': 1, 'row_end': 7, 'cell_location': [{'x': 176, 'y': 976}, {'x': 992, 'y': 976}, {'x': 992, 'y': 1029}, {'x': 176, 'y': 1030}], 'row_start': 6, 'col_start': 0, 'words': '信用减值损失'}, {'col_end': 2, 'row_end': 7, 'cell_location': [{'x': 992, 'y': 976}, {'x': 1257, 'y': 976}, {'x': 1257, 'y': 1029}, {'x': 992, 'y': 1029}], 'row_start': 6, 'col_start': 1, 'words': '(120,700)'}, {'col_end': 3, 'row_end': 7, 'cell_location': [{'x': 1257, 'y': 976}, {'x': 1460, 'y': 977}, {'x': 1460, 'y': 1029}, {'x': 1257, 'y': 1029}], 'row_start': 6, 'col_start': 2, 'words': '(136,774)'}, {'col_end': 4, 'row_end': 7, 'cell_location': [{'x': 1460, 'y': 977}, {'x': 1642, 'y': 977}, {'x': 1642, 'y': 1029}, {'x': 1460, 'y': 1029}], 'row_start': 6, 'col_start': 3, 'words': '(11.75)'}, {'col_end': 5, 'row_end': 7, 'cell_location': [{'x': 1642, 'y': 977}, {'x': 1824, 'y': 977}, {'x': 1824, 'y': 1029}, {'x': 1642, 'y': 1029}], 'row_start': 6, 'col_start': 4, 'words': '(154,535)'}, {'col_end': 6, 'row_end': 7, 'cell_location': [{'x': 1824, 'y': 977}, {'x': 2015, 'y': 978}, {'x': 2015, 'y': 1029}, {'x': 1824, 'y': 1029}], 'row_start': 6, 'col_start': 5, 'words': '(167,949)'}, {'col_end': 7, 'row_end': 7, 'cell_location': [{'x': 2015, 'y': 978}, {'x': 2191, 'y': 977}, {'x': 2191, 'y': 1029}, {'x': 2015, 'y': 1029}], 'row_start': 6, 'col_start': 6, 'words': '(193,491)'}, {'col_end': 1, 'row_end': 8, 'cell_location': [{'x': 176, 'y': 1030}, {'x': 992, 'y': 1029}, {'x': 992, 'y': 1082}, {'x': 176, 'y': 1083}], 'row_start': 7, 'col_start': 0, 'words': '其他资产减值损失'}, {'col_end': 2, 'row_end': 8, 'cell_location': [{'x': 992, 'y': 1029}, {'x': 1257, 'y': 1029}, {'x': 1257, 'y': 1082}, {'x': 992, 'y': 1082}], 'row_start': 7, 'col_start': 1, 'words': '(298)'}, {'col_end': 3, 'row_end': 8, 'cell_location': [{'x': 1257, 'y': 1029}, {'x': 1460, 'y': 1029}, {'x': 1460, 'y': 1082}, {'x': 1257, 'y': 1082}], 'row_start': 7, 'col_start': 2, 'words': '(463)'}, {'col_end': 4, 'row_end': 8, 'cell_location': [{'x': 1460, 'y': 1029}, {'x': 1642, 'y': 1029}, {'x': 1642, 'y': 1082}, {'x': 1460, 'y': 1082}], 'row_start': 7, 'col_start': 3, 'words': '(35.64)'}, {'col_end': 5, 'row_end': 8, 'cell_location': [{'x': 1642, 'y': 1029}, {'x': 1824, 'y': 1029}, {'x': 1824, 'y': 1082}, {'x': 1642, 'y': 1082}], 'row_start': 7, 'col_start': 4, 'words': '(479)'}, {'col_end': 6, 'row_end': 8, 'cell_location': [{'x': 1824, 'y': 1029}, {'x': 2015, 'y': 1029}, {'x': 2015, 'y': 1082}, {'x': 1824, 'y': 1082}], 'row_start': 7, 'col_start': 5, 'words': '(766)'}, {'col_end': 7, 'row_end': 8, 'cell_location': [{'x': 2015, 'y': 1029}, {'x': 2191, 'y': 1029}, {'x': 2192, 'y': 1082}, {'x': 2015, 'y': 1082}], 'row_start': 7, 'col_start': 6, 'words': '3,562'}, {'col_end': 1, 'row_end': 9, 'cell_location': [{'x': 176, 'y': 1083}, {'x': 992, 'y': 1082}, {'x': 992, 'y': 1134}, {'x': 176, 'y': 1135}], 'row_start': 8, 'col_start': 0, 'words': '营业利润'}, {'col_end': 2, 'row_end': 9, 'cell_location': [{'x': 992, 'y': 1082}, {'x': 1257, 'y': 1082}, {'x': 1257, 'y': 1134}, {'x': 992, 'y': 1134}], 'row_start': 8, 'col_start': 1, 'words': '384,272'}, {'col_end': 3, 'row_end': 9, 'cell_location': [{'x': 1257, 'y': 1082}, {'x': 1460, 'y': 1082}, {'x': 1460, 'y': 1134}, {'x': 1257, 'y': 1134}], 'row_start': 8, 'col_start': 2, 'words': '389,227'}, {'col_end': 4, 'row_end': 9, 'cell_location': [{'x': 1460, 'y': 1082}, {'x': 1642, 'y': 1082}, {'x': 1642, 'y': 1134}, {'x': 1460, 'y': 1134}], 'row_start': 8, 'col_start': 3, 'words': '(1.27)'}, {'col_end': 5, 'row_end': 9, 'cell_location': [{'x': 1642, 'y': 1082}, {'x': 1824, 'y': 1082}, {'x': 1824, 'y': 1134}, {'x': 1642, 'y': 1134}], 'row_start': 8, 'col_start': 4, 'words': '383,625'}, {'col_end': 6, 'row_end': 9, 'cell_location': [{'x': 1824, 'y': 1082}, {'x': 2015, 'y': 1082}, {'x': 2015, 'y': 1134}, {'x': 1824, 'y': 1134}], 'row_start': 8, 'col_start': 5, 'words': '378,776'}, {'col_end': 7, 'row_end': 9, 'cell_location': [{'x': 2015, 'y': 1082}, {'x': 2192, 'y': 1082}, {'x': 2192, 'y': 1134}, {'x': 2015, 'y': 1134}], 'row_start': 8, 'col_start': 6, 'words': '337,246'}, {'col_end': 1, 'row_end': 10, 'cell_location': [{'x': 176, 'y': 1135}, {'x': 992, 'y': 1134}, {'x': 992, 'y': 1186}, {'x': 176, 'y': 1187}], 'row_start': 9, 'col_start': 0, 'words': '利润总额'}, {'col_end': 2, 'row_end': 10, 'cell_location': [{'x': 992, 'y': 1134}, {'x': 1257, 'y': 1134}, {'x': 1257, 'y': 1186}, {'x': 992, 'y': 1186}], 'row_start': 9, 'col_start': 1, 'words': '384,377'}, {'col_end': 3, 'row_end': 10, 'cell_location': [{'x': 1257, 'y': 1134}, {'x': 1460, 'y': 1134}, {'x': 1460, 'y': 1186}, {'x': 1257, 'y': 1186}], 'row_start': 9, 'col_start': 2, 'words': '389,377'}, {'col_end': 4, 'row_end': 10, 'cell_location': [{'x': 1460, 'y': 1134}, {'x': 1642, 'y': 1134}, {'x': 1642, 'y': 1187}, {'x': 1460, 'y': 1186}], 'row_start': 9, 'col_start': 3, 'words': '(1.28)'}, {'col_end': 5, 'row_end': 10, 'cell_location': [{'x': 1642, 'y': 1134}, {'x': 1824, 'y': 1134}, {'x': 1824, 'y': 1187}, {'x': 1642, 'y': 1187}], 'row_start': 9, 'col_start': 4, 'words': '383,699'}, {'col_end': 6, 'row_end': 10, 'cell_location': [{'x': 1824, 'y': 1134}, {'x': 2015, 'y': 1134}, {'x': 2015, 'y': 1186}, {'x': 1824, 'y': 1187}], 'row_start': 9, 'col_start': 5, 'words': '378,412'}, {'col_end': 7, 'row_end': 10, 'cell_location': [{'x': 2015, 'y': 1134}, {'x': 2192, 'y': 1134}, {'x': 2192, 'y': 1186}, {'x': 2015, 'y': 1186}], 'row_start': 9, 'col_start': 6, 'words': '336,616'}, {'col_end': 1, 'row_end': 11, 'cell_location': [{'x': 176, 'y': 1187}, {'x': 992, 'y': 1186}, {'x': 992, 'y': 1238}, {'x': 176, 'y': 1239}], 'row_start': 10, 'col_start': 0, 'words': '净利润'}, {'col_end': 2, 'row_end': 11, 'cell_location': [{'x': 992, 'y': 1186}, {'x': 1257, 'y': 1186}, {'x': 1257, 'y': 1238}, {'x': 992, 'y': 1238}], 'row_start': 10, 'col_start': 1, 'words': '336,282'}, {'col_end': 3, 'row_end': 11, 'cell_location': [{'x': 1257, 'y': 1186}, {'x': 1460, 'y': 1186}, {'x': 1460, 'y': 1237}, {'x': 1257, 'y': 1238}], 'row_start': 10, 'col_start': 2, 'words': '332,460'}, {'col_end': 4, 'row_end': 11, 'cell_location': [{'x': 1460, 'y': 1186}, {'x': 1642, 'y': 1187}, {'x': 1642, 'y': 1237}, {'x': 1460, 'y': 1237}], 'row_start': 10, 'col_start': 3, 'words': '1.15'}, {'col_end': 5, 'row_end': 11, 'cell_location': [{'x': 1642, 'y': 1187}, {'x': 1824, 'y': 1187}, {'x': 1824, 'y': 1237}, {'x': 1642, 'y': 1237}], 'row_start': 10, 'col_start': 4, 'words': '324,863'}, {'col_end': 6, 'row_end': 11, 'cell_location': [{'x': 1824, 'y': 1187}, {'x': 2015, 'y': 1186}, {'x': 2015, 'y': 1237}, {'x': 1824, 'y': 1237}], 'row_start': 10, 'col_start': 5, 'words': '303,928'}, {'col_end': 7, 'row_end': 11, 'cell_location': [{'x': 2015, 'y': 1186}, {'x': 2192, 'y': 1186}, {'x': 2192, 'y': 1237}, {'x': 2015, 'y': 1237}], 'row_start': 10, 'col_start': 6, 'words': '273,579'}, {'col_end': 1, 'row_end': 12, 'cell_location': [{'x': 176, 'y': 1239}, {'x': 992, 'y': 1238}, {'x': 992, 'y': 1291}, {'x': 176, 'y': 1292}], 'row_start': 11, 'col_start': 0, 'words': '归属于本行股东的净利润'}, {'col_end': 2, 'row_end': 12, 'cell_location': [{'x': 992, 'y': 1238}, {'x': 1257, 'y': 1238}, {'x': 1257, 'y': 1291}, {'x': 992, 'y': 1291}], 'row_start': 11, 'col_start': 1, 'words': '335,577'}, {'col_end': 3, 'row_end': 12, 'cell_location': [{'x': 1257, 'y': 1238}, {'x': 1460, 'y': 1237}, {'x': 1460, 'y': 1290}, {'x': 1257, 'y': 1291}], 'row_start': 11, 'col_start': 2, 'words': '332,653'}, {'col_end': 4, 'row_end': 12, 'cell_location': [{'x': 1460, 'y': 1237}, {'x': 1642, 'y': 1237}, {'x': 1642, 'y': 1290}, {'x': 1460, 'y': 1290}], 'row_start': 11, 'col_start': 3, 'words': '0.88'}, {'col_end': 5, 'row_end': 12, 'cell_location': [{'x': 1642, 'y': 1237}, {'x': 1824, 'y': 1237}, {'x': 1824, 'y': 1290}, {'x': 1642, 'y': 1290}], 'row_start': 11, 'col_start': 4, 'words': '324,727'}, {'col_end': 6, 'row_end': 12, 'cell_location': [{'x': 1824, 'y': 1237}, {'x': 2015, 'y': 1237}, {'x': 2015, 'y': 1290}, {'x': 1824, 'y': 1290}], 'row_start': 11, 'col_start': 5, 'words': '302,513'}, {'col_end': 7, 'row_end': 12, 'cell_location': [{'x': 2015, 'y': 1237}, {'x': 2192, 'y': 1237}, {'x': 2192, 'y': 1290}, {'x': 2015, 'y': 1290}], 'row_start': 11, 'col_start': 6, 'words': '271,050'}, {'col_end': 1, 'row_end': 13, 'cell_location': [{'x': 176, 'y': 1292}, {'x': 992, 'y': 1291}, {'x': 992, 'y': 1343}, {'x': 176, 'y': 1344}], 'row_start': 12, 'col_start': 0, 'words': '归属于本行普通股股东的净利润'}, {'col_end': 2, 'row_end': 13, 'cell_location': [{'x': 992, 'y': 1291}, {'x': 1257, 'y': 1291}, {'x': 1257, 'y': 1343}, {'x': 992, 'y': 1343}], 'row_start': 12, 'col_start': 1, 'words': '328,469'}, {'col_end': 3, 'row_end': 13, 'cell_location': [{'x': 1257, 'y': 1291}, {'x': 1460, 'y': 1290}, {'x': 1460, 'y': 1343}, {'x': 1257, 'y': 1343}], 'row_start': 12, 'col_start': 2, 'words': '327,543'}, {'col_end': 4, 'row_end': 13, 'cell_location': [{'x': 1460, 'y': 1290}, {'x': 1642, 'y': 1290}, {'x': 1642, 'y': 1343}, {'x': 1460, 'y': 1343}], 'row_start': 12, 'col_start': 3, 'words': '0.28'}, {'col_end': 5, 'row_end': 13, 'cell_location': [{'x': 1642, 'y': 1290}, {'x': 1824, 'y': 1290}, {'x': 1824, 'y': 1343}, {'x': 1642, 'y': 1343}], 'row_start': 12, 'col_start': 4, 'words': '320,189'}, {'col_end': 6, 'row_end': 13, 'cell_location': [{'x': 1824, 'y': 1290}, {'x': 2015, 'y': 1290}, {'x': 2015, 'y': 1342}, {'x': 1824, 'y': 1343}], 'row_start': 12, 'col_start': 5, 'words': '297,975'}, {'col_end': 7, 'row_end': 13, 'cell_location': [{'x': 2015, 'y': 1290}, {'x': 2192, 'y': 1290}, {'x': 2192, 'y': 1342}, {'x': 2015, 'y': 1342}], 'row_start': 12, 'col_start': 6, 'words': '265,426'}, {'col_end': 1, 'row_end': 14, 'cell_location': [{'x': 176, 'y': 1344}, {'x': 992, 'y': 1343}, {'x': 992, 'y': 1395}, {'x': 176, 'y': 1396}], 'row_start': 13, 'col_start': 0, 'words': '扣除非经常性损益后归属于本行股东的净利润'}, {'col_end': 2, 'row_end': 14, 'cell_location': [{'x': 992, 'y': 1343}, {'x': 1257, 'y': 1343}, {'x': 1257, 'y': 1395}, {'x': 992, 'y': 1395}], 'row_start': 13, 'col_start': 1, 'words': '335,323'}, {'col_end': 3, 'row_end': 14, 'cell_location': [{'x': 1257, 'y': 1343}, {'x': 1460, 'y': 1343}, {'x': 1460, 'y': 1395}, {'x': 1257, 'y': 1395}], 'row_start': 13, 'col_start': 2, 'words': '332,291'}, {'col_end': 4, 'row_end': 14, 'cell_location': [{'x': 1460, 'y': 1343}, {'x': 1642, 'y': 1343}, {'x': 1642, 'y': 1395}, {'x': 1460, 'y': 1395}], 'row_start': 13, 'col_start': 3, 'words': '0.91'}, {'col_end': 5, 'row_end': 14, 'cell_location': [{'x': 1642, 'y': 1343}, {'x': 1824, 'y': 1343}, {'x': 1824, 'y': 1395}, {'x': 1642, 'y': 1395}], 'row_start': 13, 'col_start': 4, 'words': '324,569'}, {'col_end': 6, 'row_end': 14, 'cell_location': [{'x': 1824, 'y': 1343}, {'x': 2015, 'y': 1342}, {'x': 2015, 'y': 1394}, {'x': 1824, 'y': 1395}], 'row_start': 13, 'col_start': 5, 'words': '302,694'}, {'col_end': 7, 'row_end': 14, 'cell_location': [{'x': 2015, 'y': 1342}, {'x': 2192, 'y': 1342}, {'x': 2192, 'y': 1394}, {'x': 2015, 'y': 1394}], 'row_start': 13, 'col_start': 6, 'words': '271,947'}, {'col_end': 1, 'row_end': 15, 'cell_location': [{'x': 176, 'y': 1396}, {'x': 992, 'y': 1395}, {'x': 992, 'y': 1452}, {'x': 176, 'y': 1453}], 'row_start': 14, 'col_start': 0, 'words': '经营活动产生的现金流量净额'}, {'col_end': 2, 'row_end': 15, 'cell_location': [{'x': 992, 'y': 1395}, {'x': 1257, 'y': 1395}, {'x': 1257, 'y': 1452}, {'x': 992, 'y': 1452}], 'row_start': 14, 'col_start': 1, 'words': '338,023'}, {'col_end': 3, 'row_end': 15, 'cell_location': [{'x': 1257, 'y': 1395}, {'x': 1460, 'y': 1395}, {'x': 1460, 'y': 1452}, {'x': 1257, 'y': 1452}], 'row_start': 14, 'col_start': 2, 'words': '642,850'}, {'col_end': 4, 'row_end': 15, 'cell_location': [{'x': 1460, 'y': 1395}, {'x': 1642, 'y': 1395}, {'x': 1642, 'y': 1452}, {'x': 1460, 'y': 1452}], 'row_start': 14, 'col_start': 3, 'words': '(47.42)'}, {'col_end': 5, 'row_end': 15, 'cell_location': [{'x': 1642, 'y': 1395}, {'x': 1824, 'y': 1395}, {'x': 1824, 'y': 1451}, {'x': 1642, 'y': 1452}], 'row_start': 14, 'col_start': 4, 'words': '978,419'}, {'col_end': 6, 'row_end': 15, 'cell_location': [{'x': 1824, 'y': 1395}, {'x': 2015, 'y': 1394}, {'x': 2015, 'y': 1451}, {'x': 1824, 'y': 1451}], 'row_start': 14, 'col_start': 5, 'words': '436,718'}, {'col_end': 7, 'row_end': 15, 'cell_location': [{'x': 2015, 'y': 1394}, {'x': 2192, 'y': 1394}, {'x': 2192, 'y': 1451}, {'x': 2015, 'y': 1451}], 'row_start': 14, 'col_start': 6, 'words': '580,685'}, {'col_end': 1, 'row_end': 16, 'cell_location': [{'x': 176, 'y': 1453}, {'x': 992, 'y': 1452}, {'x': 992, 'y': 1582}, {'x': 176, 'y': 1583}], 'row_start': 15, 'col_start': 0, 'words': '于12月31日'}, {'col_end': 2, 'row_end': 16, 'cell_location': [{'x': 992, 'y': 1452}, {'x': 1257, 'y': 1452}, {'x': 1257, 'y': 1582}, {'x': 992, 'y': 1582}], 'row_start': 15, 'col_start': 1, 'words': ''}, {'col_end': 3, 'row_end': 16, 'cell_location': [{'x': 1257, 'y': 1452}, {'x': 1460, 'y': 1452}, {'x': 1460, 'y': 1582}, {'x': 1257, 'y': 1582}], 'row_start': 15, 'col_start': 2, 'words': ''}, {'col_end': 4, 'row_end': 16, 'cell_location': [{'x': 1460, 'y': 1452}, {'x': 1642, 'y': 1452}, {'x': 1642, 'y': 1582}, {'x': 1460, 'y': 1582}], 'row_start': 15, 'col_start': 3, 'words': ''}, {'col_end': 5, 'row_end': 16, 'cell_location': [{'x': 1642, 'y': 1452}, {'x': 1824, 'y': 1451}, {'x': 1824, 'y': 1582}, {'x': 1642, 'y': 1582}], 'row_start': 15, 'col_start': 4, 'words': ''}, {'col_end': 6, 'row_end': 16, 'cell_location': [{'x': 1824, 'y': 1451}, {'x': 2015, 'y': 1451}, {'x': 2015, 'y': 1581}, {'x': 1824, 'y': 1582}], 'row_start': 15, 'col_start': 5, 'words': ''}, {'col_end': 7, 'row_end': 16, 'cell_location': [{'x': 2015, 'y': 1451}, {'x': 2192, 'y': 1451}, {'x': 2192, 'y': 1581}, {'x': 2015, 'y': 1581}], 'row_start': 15, 'col_start': 6, 'words': ''}, {'col_end': 1, 'row_end': 17, 'cell_location': [{'x': 176, 'y': 1583}, {'x': 992, 'y': 1582}, {'x': 992, 'y': 1646}, {'x': 176, 'y': 1647}], 'row_start': 16, 'col_start': 0, 'words': '资产总额'}, {'col_end': 2, 'row_end': 17, 'cell_location': [{'x': 992, 'y': 1582}, {'x': 1257, 'y': 1582}, {'x': 1257, 'y': 1646}, {'x': 992, 'y': 1646}], 'row_start': 16, 'col_start': 1, 'words': '40,571,149'}, {'col_end': 3, 'row_end': 17, 'cell_location': [{'x': 1257, 'y': 1582}, {'x': 1460, 'y': 1582}, {'x': 1460, 'y': 1646}, {'x': 1257, 'y': 1646}], 'row_start': 16, 'col_start': 2, 'words': '38,324,826'}, {'col_end': 4, 'row_end': 17, 'cell_location': [{'x': 1460, 'y': 1582}, {'x': 1642, 'y': 1582}, {'x': 1642, 'y': 1646}, {'x': 1460, 'y': 1646}], 'row_start': 16, 'col_start': 3, 'words': '5.86'}, {'col_end': 5, 'row_end': 17, 'cell_location': [{'x': 1642, 'y': 1582}, {'x': 1824, 'y': 1582}, {'x': 1824, 'y': 1646}, {'x': 1642, 'y': 1646}], 'row_start': 16, 'col_start': 4, 'words': '34,600,711'}, {'col_end': 6, 'row_end': 17, 'cell_location': [{'x': 1824, 'y': 1582}, {'x': 2015, 'y': 1581}, {'x': 2015, 'y': 1646}, {'x': 1824, 'y': 1646}], 'row_start': 16, 'col_start': 5, 'words': '30,253,979'}, {'col_end': 7, 'row_end': 17, 'cell_location': [{'x': 2015, 'y': 1581}, {'x': 2192, 'y': 1581}, {'x': 2193, 'y': 1645}, {'x': 2015, 'y': 1646}], 'row_start': 16, 'col_start': 6, 'words': '28,132,254'}, {'col_end': 1, 'row_end': 18, 'cell_location': [{'x': 176, 'y': 1647}, {'x': 992, 'y': 1646}, {'x': 992, 'y': 1699}, {'x': 176, 'y': 1700}], 'row_start': 17, 'col_start': 0, 'words': '发放贷款和垫款净额'}, {'col_end': 2, 'row_end': 18, 'cell_location': [{'x': 992, 'y': 1646}, {'x': 1257, 'y': 1646}, {'x': 1257, 'y': 1699}, {'x': 992, 'y': 1699}], 'row_start': 17, 'col_start': 1, 'words': '25,040,400'}, {'col_end': 3, 'row_end': 18, 'cell_location': [{'x': 1257, 'y': 1646}, {'x': 1460, 'y': 1646}, {'x': 1460, 'y': 1699}, {'x': 1257, 'y': 1699}], 'row_start': 17, 'col_start': 2, 'words': '23,083,377'}, {'col_end': 4, 'row_end': 18, 'cell_location': [{'x': 1460, 'y': 1646}, {'x': 1642, 'y': 1646}, {'x': 1642, 'y': 1699}, {'x': 1460, 'y': 1699}], 'row_start': 17, 'col_start': 3, 'words': '8.48'}, {'col_end': 5, 'row_end': 18, 'cell_location': [{'x': 1642, 'y': 1646}, {'x': 1824, 'y': 1646}, {'x': 1824, 'y': 1698}, {'x': 1642, 'y': 1699}], 'row_start': 17, 'col_start': 4, 'words': '20,493,042'}, {'col_end': 6, 'row_end': 18, 'cell_location': [{'x': 1824, 'y': 1646}, {'x': 2015, 'y': 1646}, {'x': 2015, 'y': 1698}, {'x': 1824, 'y': 1698}], 'row_start': 17, 'col_start': 5, 'words': '18,170,492'}, {'col_end': 7, 'row_end': 18, 'cell_location': [{'x': 2015, 'y': 1646}, {'x': 2193, 'y': 1645}, {'x': 2193, 'y': 1698}, {'x': 2015, 'y': 1698}], 'row_start': 17, 'col_start': 6, 'words': '16,231,369'}, {'col_end': 1, 'row_end': 19, 'cell_location': [{'x': 176, 'y': 1700}, {'x': 992, 'y': 1699}, {'x': 992, 'y': 1751}, {'x': 176, 'y': 1752}], 'row_start': 18, 'col_start': 0, 'words': '负债总额'}, {'col_end': 2, 'row_end': 19, 'cell_location': [{'x': 992, 'y': 1699}, {'x': 1257, 'y': 1699}, {'x': 1256, 'y': 1751}, {'x': 992, 'y': 1751}], 'row_start': 18, 'col_start': 1, 'words': '37,227,184'}, {'col_end': 3, 'row_end': 19, 'cell_location': [{'x': 1257, 'y': 1699}, {'x': 1460, 'y': 1699}, {'x': 1460, 'y': 1751}, {'x': 1256, 'y': 1751}], 'row_start': 18, 'col_start': 2, 'words': '35,152,752'}, {'col_end': 4, 'row_end': 19, 'cell_location': [{'x': 1460, 'y': 1699}, {'x': 1642, 'y': 1699}, {'x': 1642, 'y': 1751}, {'x': 1460, 'y': 1751}], 'row_start': 18, 'col_start': 3, 'words': '5.90'}, {'col_end': 5, 'row_end': 19, 'cell_location': [{'x': 1642, 'y': 1699}, {'x': 1824, 'y': 1698}, {'x': 1824, 'y': 1751}, {'x': 1642, 'y': 1751}], 'row_start': 18, 'col_start': 4, 'words': '31,724,467'}, {'col_end': 6, 'row_end': 19, 'cell_location': [{'x': 1824, 'y': 1698}, {'x': 2015, 'y': 1698}, {'x': 2015, 'y': 1751}, {'x': 1824, 'y': 1751}], 'row_start': 18, 'col_start': 5, 'words': '27,639,857'}, {'col_end': 7, 'row_end': 19, 'cell_location': [{'x': 2015, 'y': 1698}, {'x': 2193, 'y': 1698}, {'x': 2193, 'y': 1750}, {'x': 2015, 'y': 1751}], 'row_start': 18, 'col_start': 6, 'words': '25,742,901'}, {'col_end': 1, 'row_end': 20, 'cell_location': [{'x': 176, 'y': 1752}, {'x': 992, 'y': 1751}, {'x': 992, 'y': 1803}, {'x': 176, 'y': 1804}], 'row_start': 19, 'col_start': 0, 'words': '吸收存款'}, {'col_end': 2, 'row_end': 20, 'cell_location': [{'x': 992, 'y': 1751}, {'x': 1256, 'y': 1751}, {'x': 1257, 'y': 1803}, {'x': 992, 'y': 1803}], 'row_start': 19, 'col_start': 1, 'words': '28,713,870'}, {'col_end': 3, 'row_end': 20, 'cell_location': [{'x': 1256, 'y': 1751}, {'x': 1460, 'y': 1751}, {'x': 1460, 'y': 1803}, {'x': 1257, 'y': 1803}], 'row_start': 19, 'col_start': 2, 'words': '27,654,011'}, {'col_end': 4, 'row_end': 20, 'cell_location': [{'x': 1460, 'y': 1751}, {'x': 1642, 'y': 1751}, {'x': 1642, 'y': 1803}, {'x': 1460, 'y': 1803}], 'row_start': 19, 'col_start': 3, 'words': '3.83'}, {'col_end': 5, 'row_end': 20, 'cell_location': [{'x': 1642, 'y': 1751}, {'x': 1824, 'y': 1751}, {'x': 1824, 'y': 1803}, {'x': 1642, 'y': 1803}], 'row_start': 19, 'col_start': 4, 'words': '25,020,807'}, {'col_end': 6, 'row_end': 20, 'cell_location': [{'x': 1824, 'y': 1751}, {'x': 2015, 'y': 1751}, {'x': 2015, 'y': 1803}, {'x': 1824, 'y': 1803}], 'row_start': 19, 'col_start': 5, 'words': '22,378,814'}, {'col_end': 7, 'row_end': 20, 'cell_location': [{'x': 2015, 'y': 1751}, {'x': 2193, 'y': 1750}, {'x': 2193, 'y': 1803}, {'x': 2015, 'y': 1803}], 'row_start': 19, 'col_start': 6, 'words': '20,614,976'}, {'col_end': 1, 'row_end': 21, 'cell_location': [{'x': 176, 'y': 1804}, {'x': 992, 'y': 1803}, {'x': 992, 'y': 1856}, {'x': 176, 'y': 1857}], 'row_start': 20, 'col_start': 0, 'words': '股东权益'}, {'col_end': 2, 'row_end': 21, 'cell_location': [{'x': 992, 'y': 1803}, {'x': 1257, 'y': 1803}, {'x': 1256, 'y': 1856}, {'x': 992, 'y': 1856}], 'row_start': 20, 'col_start': 1, 'words': '3,343,965'}, {'col_end': 3, 'row_end': 21, 'cell_location': [{'x': 1257, 'y': 1803}, {'x': 1460, 'y': 1803}, {'x': 1460, 'y': 1856}, {'x': 1256, 'y': 1856}], 'row_start': 20, 'col_start': 2, 'words': '3,172,074'}, {'col_end': 4, 'row_end': 21, 'cell_location': [{'x': 1460, 'y': 1803}, {'x': 1642, 'y': 1803}, {'x': 1642, 'y': 1856}, {'x': 1460, 'y': 1856}], 'row_start': 20, 'col_start': 3, 'words': '5.42'}, {'col_end': 5, 'row_end': 21, 'cell_location': [{'x': 1642, 'y': 1803}, {'x': 1824, 'y': 1803}, {'x': 1824, 'y': 1856}, {'x': 1642, 'y': 1856}], 'row_start': 20, 'col_start': 4, 'words': '2,876,244'}, {'col_end': 6, 'row_end': 21, 'cell_location': [{'x': 1824, 'y': 1803}, {'x': 2015, 'y': 1803}, {'x': 2015, 'y': 1856}, {'x': 1824, 'y': 1856}], 'row_start': 20, 'col_start': 5, 'words': '2,614,122'}, {'col_end': 7, 'row_end': 21, 'cell_location': [{'x': 2015, 'y': 1803}, {'x': 2193, 'y': 1803}, {'x': 2193, 'y': 1855}, {'x': 2015, 'y': 1856}], 'row_start': 20, 'col_start': 6, 'words': '2,389,353'}, {'col_end': 1, 'row_end': 22, 'cell_location': [{'x': 176, 'y': 1857}, {'x': 992, 'y': 1856}, {'x': 992, 'y': 1909}, {'x': 176, 'y': 1910}], 'row_start': 21, 'col_start': 0, 'words': '归属于本行股东权益'}, {'col_end': 2, 'row_end': 22, 'cell_location': [{'x': 992, 'y': 1856}, {'x': 1256, 'y': 1856}, {'x': 1256, 'y': 1909}, {'x': 992, 'y': 1909}], 'row_start': 21, 'col_start': 1, 'words': '3,322,127'}, {'col_end': 3, 'row_end': 22, 'cell_location': [{'x': 1256, 'y': 1856}, {'x': 1460, 'y': 1856}, {'x': 1460, 'y': 1909}, {'x': 1256, 'y': 1909}], 'row_start': 21, 'col_start': 2, 'words': '3,150,145'}, {'col_end': 4, 'row_end': 22, 'cell_location': [{'x': 1460, 'y': 1856}, {'x': 1642, 'y': 1856}, {'x': 1642, 'y': 1909}, {'x': 1460, 'y': 1909}], 'row_start': 21, 'col_start': 3, 'words': '5.46'}, {'col_end': 5, 'row_end': 22, 'cell_location': [{'x': 1642, 'y': 1856}, {'x': 1824, 'y': 1856}, {'x': 1824, 'y': 1909}, {'x': 1642, 'y': 1909}], 'row_start': 21, 'col_start': 4, 'words': '2,855,450'}, {'col_end': 6, 'row_end': 22, 'cell_location': [{'x': 1824, 'y': 1856}, {'x': 2015, 'y': 1856}, {'x': 2015, 'y': 1908}, {'x': 1824, 'y': 1909}], 'row_start': 21, 'col_start': 5, 'words': '2,588,231'}, {'col_end': 7, 'row_end': 22, 'cell_location': [{'x': 2015, 'y': 1856}, {'x': 2193, 'y': 1855}, {'x': 2193, 'y': 1908}, {'x': 2015, 'y': 1908}], 'row_start': 21, 'col_start': 6, 'words': '2,364,808'}, {'col_end': 1, 'row_end': 23, 'cell_location': [{'x': 176, 'y': 1910}, {'x': 992, 'y': 1909}, {'x': 992, 'y': 1961}, {'x': 176, 'y': 1962}], 'row_start': 22, 'col_start': 0, 'words': '股本'}, {'col_end': 2, 'row_end': 23, 'cell_location': [{'x': 992, 'y': 1909}, {'x': 1256, 'y': 1909}, {'x': 1256, 'y': 1961}, {'x': 992, 'y': 1961}], 'row_start': 22, 'col_start': 1, 'words': '250,011'}, {'col_end': 3, 'row_end': 23, 'cell_location': [{'x': 1256, 'y': 1909}, {'x': 1460, 'y': 1909}, {'x': 1460, 'y': 1961}, {'x': 1256, 'y': 1961}], 'row_start': 22, 'col_start': 2, 'words': '250,011'}, {'col_end': 4, 'row_end': 23, 'cell_location': [{'x': 1460, 'y': 1909}, {'x': 1642, 'y': 1909}, {'x': 1642, 'y': 1961}, {'x': 1460, 'y': 1961}], 'row_start': 22, 'col_start': 3, 'words': '—'}, {'col_end': 5, 'row_end': 23, 'cell_location': [{'x': 1642, 'y': 1909}, {'x': 1824, 'y': 1909}, {'x': 1824, 'y': 1961}, {'x': 1642, 'y': 1961}], 'row_start': 22, 'col_start': 4, 'words': '250,011'}, {'col_end': 6, 'row_end': 23, 'cell_location': [{'x': 1824, 'y': 1909}, {'x': 2015, 'y': 1908}, {'x': 2015, 'y': 1961}, {'x': 1824, 'y': 1961}], 'row_start': 22, 'col_start': 5, 'words': '250,011'}, {'col_end': 7, 'row_end': 23, 'cell_location': [{'x': 2015, 'y': 1908}, {'x': 2193, 'y': 1908}, {'x': 2193, 'y': 1960}, {'x': 2015, 'y': 1961}], 'row_start': 22, 'col_start': 6, 'words': '250,011'}, {'col_end': 1, 'row_end': 24, 'cell_location': [{'x': 176, 'y': 1962}, {'x': 992, 'y': 1961}, {'x': 992, 'y': 2014}, {'x': 176, 'y': 2014}], 'row_start': 23, 'col_start': 0, 'words': '核心一级资本净额2'}, {'col_end': 2, 'row_end': 24, 'cell_location': [{'x': 992, 'y': 1961}, {'x': 1256, 'y': 1961}, {'x': 1256, 'y': 2013}, {'x': 992, 'y': 2014}], 'row_start': 23, 'col_start': 1, 'words': '3,165,549'}, {'col_end': 3, 'row_end': 24, 'cell_location': [{'x': 1256, 'y': 1961}, {'x': 1460, 'y': 1961}, {'x': 1460, 'y': 2013}, {'x': 1256, 'y': 2013}], 'row_start': 23, 'col_start': 2, 'words': '2,944,386'}, {'col_end': 4, 'row_end': 24, 'cell_location': [{'x': 1460, 'y': 1961}, {'x': 1642, 'y': 1961}, {'x': 1642, 'y': 2013}, {'x': 1460, 'y': 2013}], 'row_start': 23, 'col_start': 3, 'words': '7.51'}, {'col_end': 5, 'row_end': 24, 'cell_location': [{'x': 1642, 'y': 1961}, {'x': 1824, 'y': 1961}, {'x': 1824, 'y': 2013}, {'x': 1642, 'y': 2013}], 'row_start': 23, 'col_start': 4, 'words': '2,706,459'}, {'col_end': 6, 'row_end': 24, 'cell_location': [{'x': 1824, 'y': 1961}, {'x': 2015, 'y': 1961}, {'x': 2015, 'y': 2013}, {'x': 1824, 'y': 2013}], 'row_start': 23, 'col_start': 5, 'words': '2,475,462'}, {'col_end': 7, 'row_end': 24, 'cell_location': [{'x': 2015, 'y': 1961}, {'x': 2193, 'y': 1960}, {'x': 2193, 'y': 2013}, {'x': 2015, 'y': 2013}], 'row_start': 23, 'col_start': 6, 'words': '2,261,449'}, {'col_end': 1, 'row_end': 25, 'cell_location': [{'x': 176, 'y': 2014}, {'x': 992, 'y': 2014}, {'x': 992, 'y': 2066}, {'x': 176, 'y': 2066}], 'row_start': 24, 'col_start': 0, 'words': '其他一级资本净额'}, {'col_end': 2, 'row_end': 25, 'cell_location': [{'x': 992, 'y': 2014}, {'x': 1256, 'y': 2013}, {'x': 1256, 'y': 2065}, {'x': 992, 'y': 2066}], 'row_start': 24, 'col_start': 1, 'words': '158,875'}, {'col_end': 3, 'row_end': 25, 'cell_location': [{'x': 1256, 'y': 2013}, {'x': 1460, 'y': 2013}, {'x': 1460, 'y': 2065}, {'x': 1256, 'y': 2065}], 'row_start': 24, 'col_start': 2, 'words': '200,088'}, {'col_end': 4, 'row_end': 25, 'cell_location': [{'x': 1460, 'y': 2013}, {'x': 1642, 'y': 2013}, {'x': 1642, 'y': 2065}, {'x': 1460, 'y': 2065}], 'row_start': 24, 'col_start': 3, 'words': '(20.60)'}, {'col_end': 5, 'row_end': 25, 'cell_location': [{'x': 1642, 'y': 2013}, {'x': 1824, 'y': 2013}, {'x': 1824, 'y': 2065}, {'x': 1642, 'y': 2065}], 'row_start': 24, 'col_start': 4, 'words': '140,074'}, {'col_end': 6, 'row_end': 25, 'cell_location': [{'x': 1824, 'y': 2013}, {'x': 2015, 'y': 2013}, {'x': 2015, 'y': 2065}, {'x': 1824, 'y': 2065}], 'row_start': 24, 'col_start': 5, 'words': '100,066'}, {'col_end': 7, 'row_end': 25, 'cell_location': [{'x': 2015, 'y': 2013}, {'x': 2193, 'y': 2013}, {'x': 2193, 'y': 2065}, {'x': 2015, 'y': 2065}], 'row_start': 24, 'col_start': 6, 'words': '100,068'}, {'col_end': 1, 'row_end': 26, 'cell_location': [{'x': 176, 'y': 2066}, {'x': 992, 'y': 2066}, {'x': 992, 'y': 2117}, {'x': 176, 'y': 2117}], 'row_start': 25, 'col_start': 0, 'words': '二级资本净额2'}, {'col_end': 2, 'row_end': 26, 'cell_location': [{'x': 992, 'y': 2066}, {'x': 1256, 'y': 2065}, {'x': 1256, 'y': 2117}, {'x': 992, 'y': 2117}], 'row_start': 25, 'col_start': 1, 'words': '978,839'}, {'col_end': 3, 'row_end': 26, 'cell_location': [{'x': 1256, 'y': 2065}, {'x': 1460, 'y': 2065}, {'x': 1460, 'y': 2117}, {'x': 1256, 'y': 2117}], 'row_start': 25, 'col_start': 2, 'words': '876,187'}, {'col_end': 4, 'row_end': 26, 'cell_location': [{'x': 1460, 'y': 2065}, {'x': 1642, 'y': 2065}, {'x': 1642, 'y': 2117}, {'x': 1460, 'y': 2117}], 'row_start': 25, 'col_start': 3, 'words': '11.72'}, {'col_end': 5, 'row_end': 26, 'cell_location': [{'x': 1642, 'y': 2065}, {'x': 1824, 'y': 2065}, {'x': 1824, 'y': 2117}, {'x': 1642, 'y': 2117}], 'row_start': 25, 'col_start': 4, 'words': '793,905'}, {'col_end': 6, 'row_end': 26, 'cell_location': [{'x': 1824, 'y': 2065}, {'x': 2015, 'y': 2065}, {'x': 2015, 'y': 2117}, {'x': 1824, 'y': 2117}], 'row_start': 25, 'col_start': 5, 'words': '676,754'}, {'col_end': 7, 'row_end': 26, 'cell_location': [{'x': 2015, 'y': 2065}, {'x': 2193, 'y': 2065}, {'x': 2193, 'y': 2116}, {'x': 2015, 'y': 2117}], 'row_start': 25, 'col_start': 6, 'words': '471,164'}, {'col_end': 1, 'row_end': 27, 'cell_location': [{'x': 176, 'y': 2117}, {'x': 992, 'y': 2117}, {'x': 991, 'y': 2169}, {'x': 176, 'y': 2169}], 'row_start': 26, 'col_start': 0, 'words': '资本净额2'}, {'col_end': 2, 'row_end': 27, 'cell_location': [{'x': 992, 'y': 2117}, {'x': 1256, 'y': 2117}, {'x': 1256, 'y': 2169}, {'x': 991, 'y': 2169}], 'row_start': 26, 'col_start': 1, 'words': '4,303,263'}, {'col_end': 3, 'row_end': 27, 'cell_location': [{'x': 1256, 'y': 2117}, {'x': 1460, 'y': 2117}, {'x': 1460, 'y': 2168}, {'x': 1256, 'y': 2169}], 'row_start': 26, 'col_start': 2, 'words': '4,020,661'}, {'col_end': 4, 'row_end': 27, 'cell_location': [{'x': 1460, 'y': 2117}, {'x': 1642, 'y': 2117}, {'x': 1642, 'y': 2168}, {'x': 1460, 'y': 2168}], 'row_start': 26, 'col_start': 3, 'words': '7.03'}, {'col_end': 5, 'row_end': 27, 'cell_location': [{'x': 1642, 'y': 2117}, {'x': 1824, 'y': 2117}, {'x': 1824, 'y': 2168}, {'x': 1642, 'y': 2168}], 'row_start': 26, 'col_start': 4, 'words': '3,640,438'}, {'col_end': 6, 'row_end': 27, 'cell_location': [{'x': 1824, 'y': 2117}, {'x': 2015, 'y': 2117}, {'x': 2015, 'y': 2168}, {'x': 1824, 'y': 2168}], 'row_start': 26, 'col_start': 5, 'words': '3,252,282'}, {'col_end': 7, 'row_end': 27, 'cell_location': [{'x': 2015, 'y': 2117}, {'x': 2193, 'y': 2116}, {'x': 2194, 'y': 2168}, {'x': 2015, 'y': 2168}], 'row_start': 26, 'col_start': 6, 'words': '2,832,681'}, {'col_end': 1, 'row_end': 28, 'cell_location': [{'x': 176, 'y': 2169}, {'x': 991, 'y': 2169}, {'x': 991, 'y': 2222}, {'x': 176, 'y': 2222}], 'row_start': 27, 'col_start': 0, 'words': '风险加权资产2'}, {'col_end': 2, 'row_end': 28, 'cell_location': [{'x': 991, 'y': 2169}, {'x': 1256, 'y': 2169}, {'x': 1256, 'y': 2222}, {'x': 991, 'y': 2222}], 'row_start': 27, 'col_start': 1, 'words': '21,854,590'}, {'col_end': 3, 'row_end': 28, 'cell_location': [{'x': 1256, 'y': 2169}, {'x': 1460, 'y': 2168}, {'x': 1460, 'y': 2222}, {'x': 1256, 'y': 2222}], 'row_start': 27, 'col_start': 2, 'words': '22,395,908'}, {'col_end': 4, 'row_end': 28, 'cell_location': [{'x': 1460, 'y': 2168}, {'x': 1642, 'y': 2168}, {'x': 1642, 'y': 2222}, {'x': 1460, 'y': 2222}], 'row_start': 27, 'col_start': 3, 'words': '(2.42)'}, {'col_end': 5, 'row_end': 28, 'cell_location': [{'x': 1642, 'y': 2168}, {'x': 1824, 'y': 2168}, {'x': 1824, 'y': 2222}, {'x': 1642, 'y': 2222}], 'row_start': 27, 'col_start': 4, 'words': '19,767,834'}, {'col_end': 6, 'row_end': 28, 'cell_location': [{'x': 1824, 'y': 2168}, {'x': 2015, 'y': 2168}, {'x': 2015, 'y': 2222}, {'x': 1824, 'y': 2222}], 'row_start': 27, 'col_start': 5, 'words': '18,215,893'}, {'col_end': 7, 'row_end': 28, 'cell_location': [{'x': 2015, 'y': 2168}, {'x': 2194, 'y': 2168}, {'x': 2194, 'y': 2222}, {'x': 2015, 'y': 2222}], 'row_start': 27, 'col_start': 6, 'words': '16,604,591'}], 'table_location': [{'x': 176, 'y': 580}, {'x': 2194, 'y': 580}, {'x': 2194, 'y': 2222}, {'x': 176, 'y': 2222}], 'footer': [{'location': [{'x': 186, 'y': 2299}, {'x': 449, 'y': 2299}, {'x': 449, 'y': 2342}, {'x': 186, 'y': 2342}], 'words': '每股计（人民币元）'}]}, {'header': [{'location': [{'x': 188, 'y': 528}, {'x': 749, 'y': 528}, {'x': 749, 'y': 574}, {'x': 188, 'y': 574}], 'words': '(除特别注明外，以人民币百万元列示)'}], 'body': [{'col_end': 1, 'row_end': 1, 'cell_location': [{'x': 188, 'y': 2361}, {'x': 1028, 'y': 2366}, {'x': 1029, 'y': 2424}, {'x': 188, 'y': 2423}], 'row_start': 0, 'col_start': 0, 'words': '基本和稀释每股收益'}, {'col_end': 2, 'row_end': 1, 'cell_location': [{'x': 1028, 'y': 2366}, {'x': 1281, 'y': 2370}, {'x': 1282, 'y': 2424}, {'x': 1029, 'y': 2424}], 'row_start': 0, 'col_start': 1, 'words': '1.31'}, {'col_end': 3, 'row_end': 1, 'cell_location': [{'x': 1281, 'y': 2370}, {'x': 1466, 'y': 2372}, {'x': 1466, 'y': 2425}, {'x': 1282, 'y': 2424}], 'row_start': 0, 'col_start': 2, 'words': '1.31'}, {'col_end': 4, 'row_end': 1, 'cell_location': [{'x': 1466, 'y': 2372}, {'x': 1649, 'y': 2373}, {'x': 1649, 'y': 2425}, {'x': 1466, 'y': 2425}], 'row_start': 0, 'col_start': 3, 'words': '—'}, {'col_end': 5, 'row_end': 1, 'cell_location': [{'x': 1649, 'y': 2373}, {'x': 1850, 'y': 2375}, {'x': 1850, 'y': 2425}, {'x': 1649, 'y': 2425}], 'row_start': 0, 'col_start': 4, 'words': '1.28'}, {'col_end': 6, 'row_end': 1, 'cell_location': [{'x': 1850, 'y': 2375}, {'x': 2045, 'y': 2376}, {'x': 2045, 'y': 2425}, {'x': 1850, 'y': 2425}], 'row_start': 0, 'col_start': 5, 'words': '1.19'}, {'col_end': 7, 'row_end': 1, 'cell_location': [{'x': 2045, 'y': 2376}, {'x': 2174, 'y': 2376}, {'x': 2175, 'y': 2425}, {'x': 2045, 'y': 2425}], 'row_start': 0, 'col_start': 6, 'words': '1.06'}, {'col_end': 1, 'row_end': 2, 'cell_location': [{'x': 188, 'y': 2423}, {'x': 1029, 'y': 2424}, {'x': 1028, 'y': 2476}, {'x': 189, 'y': 2476}], 'row_start': 1, 'col_start': 0, 'words': '扣除非经常性损益后的基本和稀释每股收益'}, {'col_end': 2, 'row_end': 2, 'cell_location': [{'x': 1029, 'y': 2424}, {'x': 1282, 'y': 2424}, {'x': 1281, 'y': 2476}, {'x': 1028, 'y': 2476}], 'row_start': 1, 'col_start': 1, 'words': '1.31'}, {'col_end': 3, 'row_end': 2, 'cell_location': [{'x': 1282, 'y': 2424}, {'x': 1466, 'y': 2425}, {'x': 1466, 'y': 2476}, {'x': 1281, 'y': 2476}], 'row_start': 1, 'col_start': 2, 'words': '1.31'}, {'col_end': 4, 'row_end': 2, 'cell_location': [{'x': 1466, 'y': 2425}, {'x': 1649, 'y': 2425}, {'x': 1649, 'y': 2477}, {'x': 1466, 'y': 2476}], 'row_start': 1, 'col_start': 3, 'words': '—'}, {'col_end': 5, 'row_end': 2, 'cell_location': [{'x': 1649, 'y': 2425}, {'x': 1850, 'y': 2425}, {'x': 1850, 'y': 2477}, {'x': 1649, 'y': 2477}], 'row_start': 1, 'col_start': 4, 'words': '1.28'}, {'col_end': 6, 'row_end': 2, 'cell_location': [{'x': 1850, 'y': 2425}, {'x': 2045, 'y': 2425}, {'x': 2045, 'y': 2477}, {'x': 1850, 'y': 2477}], 'row_start': 1, 'col_start': 5, 'words': '1.19'}, {'col_end': 7, 'row_end': 2, 'cell_location': [{'x': 2045, 'y': 2425}, {'x': 2175, 'y': 2425}, {'x': 2175, 'y': 2477}, {'x': 2045, 'y': 2477}], 'row_start': 1, 'col_start': 6, 'words': '1.07'}, {'col_end': 1, 'row_end': 3, 'cell_location': [{'x': 189, 'y': 2476}, {'x': 1028, 'y': 2476}, {'x': 1029, 'y': 2528}, {'x': 189, 'y': 2526}], 'row_start': 2, 'col_start': 0, 'words': '归属于本行普通股股东的每股净资产'}, {'col_end': 2, 'row_end': 3, 'cell_location': [{'x': 1028, 'y': 2476}, {'x': 1281, 'y': 2476}, {'x': 1281, 'y': 2528}, {'x': 1029, 'y': 2528}], 'row_start': 2, 'col_start': 1, 'words': '12.65'}, {'col_end': 3, 'row_end': 3, 'cell_location': [{'x': 1281, 'y': 2476}, {'x': 1466, 'y': 2476}, {'x': 1466, 'y': 2528}, {'x': 1281, 'y': 2528}], 'row_start': 2, 'col_start': 2, 'words': '11.80'}, {'col_end': 4, 'row_end': 3, 'cell_location': [{'x': 1466, 'y': 2476}, {'x': 1649, 'y': 2477}, {'x': 1649, 'y': 2529}, {'x': 1466, 'y': 2528}], 'row_start': 2, 'col_start': 3, 'words': '7.20'}, {'col_end': 5, 'row_end': 3, 'cell_location': [{'x': 1649, 'y': 2477}, {'x': 1850, 'y': 2477}, {'x': 1850, 'y': 2529}, {'x': 1649, 'y': 2529}], 'row_start': 2, 'col_start': 4, 'words': '10.86'}, {'col_end': 6, 'row_end': 3, 'cell_location': [{'x': 1850, 'y': 2477}, {'x': 2045, 'y': 2477}, {'x': 2045, 'y': 2530}, {'x': 1850, 'y': 2529}], 'row_start': 2, 'col_start': 5, 'words': '9.95'}, {'col_end': 7, 'row_end': 3, 'cell_location': [{'x': 2045, 'y': 2477}, {'x': 2175, 'y': 2477}, {'x': 2175, 'y': 2530}, {'x': 2045, 'y': 2530}], 'row_start': 2, 'col_start': 6, 'words': '9.06'}, {'col_end': 1, 'row_end': 4, 'cell_location': [{'x': 189, 'y': 2526}, {'x': 1029, 'y': 2528}, {'x': 1029, 'y': 2582}, {'x': 189, 'y': 2578}], 'row_start': 3, 'col_start': 0, 'words': '每股经营活动产生的现金流量净额'}, {'col_end': 2, 'row_end': 4, 'cell_location': [{'x': 1029, 'y': 2528}, {'x': 1281, 'y': 2528}, {'x': 1282, 'y': 2581}, {'x': 1029, 'y': 2582}], 'row_start': 3, 'col_start': 1, 'words': '1.35'}, {'col_end': 3, 'row_end': 4, 'cell_location': [{'x': 1281, 'y': 2528}, {'x': 1466, 'y': 2528}, {'x': 1467, 'y': 2582}, {'x': 1282, 'y': 2581}], 'row_start': 3, 'col_start': 2, 'words': '2.57'}, {'col_end': 4, 'row_end': 4, 'cell_location': [{'x': 1466, 'y': 2528}, {'x': 1649, 'y': 2529}, {'x': 1649, 'y': 2582}, {'x': 1467, 'y': 2582}], 'row_start': 3, 'col_start': 3, 'words': '(47.47)'}, {'col_end': 5, 'row_end': 4, 'cell_location': [{'x': 1649, 'y': 2529}, {'x': 1850, 'y': 2529}, {'x': 1850, 'y': 2583}, {'x': 1649, 'y': 2582}], 'row_start': 3, 'col_start': 4, 'words': '3.91'}, {'col_end': 6, 'row_end': 4, 'cell_location': [{'x': 1850, 'y': 2529}, {'x': 2045, 'y': 2530}, {'x': 2045, 'y': 2583}, {'x': 1850, 'y': 2583}], 'row_start': 3, 'col_start': 5, 'words': '1.75'}, {'col_end': 7, 'row_end': 4, 'cell_location': [{'x': 2045, 'y': 2530}, {'x': 2175, 'y': 2530}, {'x': 2176, 'y': 2583}, {'x': 2045, 'y': 2583}], 'row_start': 3, 'col_start': 6, 'words': '2.32'}], 'table_location': [{'x': 188, 'y': 2361}, {'x': 2176, 'y': 2361}, {'x': 2176, 'y': 2583}, {'x': 188, 'y': 2583}], 'footer': [{'location': [{'x': 268, 'y': 2659}, {'x': 1859, 'y': 2659}, {'x': 1859, 'y': 2701}, {'x': 268, 'y': 2701}], 'words': '2024年数据按照《商业银行资本管理办法》相关规则计量，往期数据按照《商业银行资本管理办法（试行）》相关规则计量。'}, {'location': [{'x': 269, 'y': 2700}, {'x': 2022, 'y': 2700}, {'x': 2022, 'y': 2742}, {'x': 269, 'y': 2742}], 'words': '根据中国证监会《公开发行证券的公司信息披露编报规则第9号一净资产收益率和每股收益的计算及披露》(2010年修订)的规定计算。'}]}], 'table_num': 2, 'log_id': 1995815416027844086, 'image_info': {'image_path': 'E:\\Datas\\base_pros\\DocuVista\\test_codes\\pngs\\7d4a49dd-9b72-4c02-a7ee-d09a0921ca4b_014.png', 'image_id': 'img_bb21228ea47e'}}
    llm_result = {'success': True, 'image_info': {'image_path': 'E:\\Datas\\base_pros\\DocuVista\\test_codes\\pngs\\7d4a49dd-9b72-4c02-a7ee-d09a0921ca4b_014.png', 'image_id': 'img_bb21228ea47e'}, 'tables_structure': {'tables': [{'id': '1', 'ocr_tables': [0, 1], 'headers': {'cols': ['', '2024年', '2023年', '变化(%)', '2022年', '2021年', '2020年'], 'rows': ['全年业绩>>营业收入', '全年业绩>>利息净收入', '全年业绩>>手续费及佣金净收入', '全年业绩>>其他非利息收入', '全年业绩>>业务及管理费', '全年业绩>>信用减值损失', '全年业绩>>其他资产减值损失', '全年业绩>>营业利润', '全年业绩>>利润总额', '全年业绩>>净利润', '全年业绩>>归属于本行股东的净利润', '全年业绩>>归属于本行普通股股东的净利润', '全年业绩>>扣除非经常性损益后归属于本行股东的净利润¹', '全年业绩>>经营活动产生的现金流量净额', '于12月31日>>资产总额', '于12月31日>>发放贷款和垫款净额', '于12月31日>>负债总额', '于12月31日>>吸收存款', '于12月31日>>股东权益', '于12月31日>>归属于本行股东权益', '于12月31日>>股本', '于12月31日>>核心一级资本净额²', '于12月31日>>其他一级资本净额²', '于12月31日>>二级资本净额²', '于12月31日>>资本净额²', '于12月31日>>风险加权资产²', '每股计(人民币元)>>基本和稀释每股收益³', '每股计(人民币元)>>扣除非经常性损益后的基本和稀释每股收益³', '每股计(人民币元)>>归属于本行普通股股东的每股净资产', '每股计(人民币元)>>每股经营活动产生的现金流量净额']}}]}, 'processing_stats': {'analysis_time_sec': 9.71, 'ocr_tables_count': 2, 'visual_tables_count': 1, 'token_usage': {'prompt_tokens': 2032, 'completion_tokens': 522, 'total_tokens': 2554}}}

    # ocr_result= {'tables_result': [{'header': [{'location': [{'x': 90, 'y': 239}, {'x': 275, 'y': 239}, {'x': 275, 'y': 253}, {'x': 90, 'y': 253}], 'words': '表1(KMI):监管并表关键审慎监管指标'}, {'location': [{'x': 86, 'y': 207}, {'x': 466, 'y': 207}, {'x': 466, 'y': 220}, {'x': 86, 'y': 220}], 'words': '关键审慎监管指标包括资本充足率、杠杆率以及流动性风险相关指标。本集团关键审'}], 'body': [{'col_end': 2, 'row_end': 2, 'cell_location': [{'x': 59, 'y': 263}, {'x': 222, 'y': 263}, {'x': 222, 'y': 300}, {'x': 59, 'y': 300}], 'row_start': 0, 'col_start': 0, 'words': '(人民币百万元，百分比除外)'}, {'col_end': 3, 'row_end': 1, 'cell_location': [{'x': 223, 'y': 263}, {'x': 286, 'y': 264}, {'x': 286, 'y': 275}, {'x': 223, 'y': 275}], 'row_start': 0, 'col_start': 2, 'words': 'a'}, {'col_end': 4, 'row_end': 1, 'cell_location': [{'x': 286, 'y': 264}, {'x': 349, 'y': 264}, {'x': 349, 'y': 275}, {'x': 286, 'y': 275}], 'row_start': 0, 'col_start': 3, 'words': 'b'}, {'col_end': 5, 'row_end': 1, 'cell_location': [{'x': 349, 'y': 264}, {'x': 412, 'y': 264}, {'x': 412, 'y': 275}, {'x': 349, 'y': 275}], 'row_start': 0, 'col_start': 4, 'words': ''}, {'col_end': 6, 'row_end': 1, 'cell_location': [{'x': 412, 'y': 264}, {'x': 475, 'y': 264}, {'x': 475, 'y': 275}, {'x': 412, 'y': 275}], 'row_start': 0, 'col_start': 5, 'words': 'd'}, {'col_end': 3, 'row_end': 2, 'cell_location': [{'x': 223, 'y': 275}, {'x': 286, 'y': 275}, {'x': 286, 'y': 300}, {'x': 223, 'y': 300}], 'row_start': 1, 'col_start': 2, 'words': '2024年\n12月31日'}, {'col_end': 4, 'row_end': 2, 'cell_location': [{'x': 286, 'y': 275}, {'x': 349, 'y': 275}, {'x': 349, 'y': 300}, {'x': 286, 'y': 300}], 'row_start': 1, 'col_start': 3, 'words': '2024年\n9月30日'}, {'col_end': 5, 'row_end': 2, 'cell_location': [{'x': 349, 'y': 275}, {'x': 412, 'y': 275}, {'x': 412, 'y': 300}, {'x': 349, 'y': 300}], 'row_start': 1, 'col_start': 4, 'words': '2024年\n6月30日'}, {'col_end': 6, 'row_end': 2, 'cell_location': [{'x': 412, 'y': 275}, {'x': 475, 'y': 275}, {'x': 475, 'y': 300}, {'x': 412, 'y': 300}], 'row_start': 1, 'col_start': 5, 'words': '2024年\n3月31日'}, {'col_end': 2, 'row_end': 3, 'cell_location': [{'x': 59, 'y': 300}, {'x': 222, 'y': 300}, {'x': 222, 'y': 313}, {'x': 59, 'y': 313}], 'row_start': 2, 'col_start': 0, 'words': '可用资本（数额）'}, {'col_end': 3, 'row_end': 3, 'cell_location': [{'x': 223, 'y': 300}, {'x': 286, 'y': 300}, {'x': 286, 'y': 313}, {'x': 223, 'y': 313}], 'row_start': 2, 'col_start': 2, 'words': ''}, {'col_end': 4, 'row_end': 3, 'cell_location': [{'x': 286, 'y': 300}, {'x': 349, 'y': 300}, {'x': 349, 'y': 313}, {'x': 286, 'y': 313}], 'row_start': 2, 'col_start': 3, 'words': ''}, {'col_end': 5, 'row_end': 3, 'cell_location': [{'x': 349, 'y': 300}, {'x': 412, 'y': 300}, {'x': 412, 'y': 313}, {'x': 349, 'y': 313}], 'row_start': 2, 'col_start': 4, 'words': ''}, {'col_end': 6, 'row_end': 3, 'cell_location': [{'x': 412, 'y': 300}, {'x': 475, 'y': 300}, {'x': 475, 'y': 313}, {'x': 412, 'y': 313}], 'row_start': 2, 'col_start': 5, 'words': ''}, {'col_end': 1, 'row_end': 4, 'cell_location': [{'x': 60, 'y': 313}, {'x': 90, 'y': 313}, {'x': 90, 'y': 325}, {'x': 60, 'y': 325}], 'row_start': 3, 'col_start': 0, 'words': '1'}, {'col_end': 2, 'row_end': 4, 'cell_location': [{'x': 90, 'y': 313}, {'x': 223, 'y': 313}, {'x': 223, 'y': 325}, {'x': 90, 'y': 325}], 'row_start': 3, 'col_start': 1, 'words': '核心一级资本净额'}, {'col_end': 3, 'row_end': 4, 'cell_location': [{'x': 223, 'y': 313}, {'x': 286, 'y': 313}, {'x': 286, 'y': 325}, {'x': 223, 'y': 325}], 'row_start': 3, 'col_start': 2, 'words': '3,165,549'}, {'col_end': 4, 'row_end': 4, 'cell_location': [{'x': 286, 'y': 313}, {'x': 349, 'y': 313}, {'x': 349, 'y': 325}, {'x': 286, 'y': 325}], 'row_start': 3, 'col_start': 3, 'words': '3,124,043'}, {'col_end': 5, 'row_end': 4, 'cell_location': [{'x': 349, 'y': 313}, {'x': 412, 'y': 313}, {'x': 412, 'y': 325}, {'x': 349, 'y': 325}], 'row_start': 3, 'col_start': 4, 'words': '3,038,387'}, {'col_end': 6, 'row_end': 4, 'cell_location': [{'x': 412, 'y': 313}, {'x': 475, 'y': 313}, {'x': 475, 'y': 325}, {'x': 412, 'y': 325}], 'row_start': 3, 'col_start': 5, 'words': '3,045,754'}, {'col_end': 1, 'row_end': 5, 'cell_location': [{'x': 60, 'y': 325}, {'x': 90, 'y': 325}, {'x': 90, 'y': 338}, {'x': 60, 'y': 338}], 'row_start': 4, 'col_start': 0, 'words': '2'}, {'col_end': 2, 'row_end': 5, 'cell_location': [{'x': 90, 'y': 325}, {'x': 223, 'y': 325}, {'x': 223, 'y': 338}, {'x': 90, 'y': 338}], 'row_start': 4, 'col_start': 1, 'words': '一级资本净额'}, {'col_end': 3, 'row_end': 5, 'cell_location': [{'x': 223, 'y': 325}, {'x': 286, 'y': 325}, {'x': 286, 'y': 338}, {'x': 223, 'y': 338}], 'row_start': 4, 'col_start': 2, 'words': '3,324,424'}, {'col_end': 4, 'row_end': 5, 'cell_location': [{'x': 286, 'y': 325}, {'x': 349, 'y': 325}, {'x': 349, 'y': 338}, {'x': 286, 'y': 338}], 'row_start': 4, 'col_start': 3, 'words': '3,322,954'}, {'col_end': 5, 'row_end': 5, 'cell_location': [{'x': 349, 'y': 325}, {'x': 412, 'y': 325}, {'x': 412, 'y': 338}, {'x': 349, 'y': 338}], 'row_start': 4, 'col_start': 4, 'words': '3,237,254'}, {'col_end': 6, 'row_end': 5, 'cell_location': [{'x': 412, 'y': 325}, {'x': 475, 'y': 325}, {'x': 475, 'y': 338}, {'x': 412, 'y': 338}], 'row_start': 4, 'col_start': 5, 'words': '3,245,824'}, {'col_end': 1, 'row_end': 6, 'cell_location': [{'x': 60, 'y': 338}, {'x': 90, 'y': 338}, {'x': 90, 'y': 350}, {'x': 60, 'y': 350}], 'row_start': 5, 'col_start': 0, 'words': '3'}, {'col_end': 2, 'row_end': 6, 'cell_location': [{'x': 90, 'y': 338}, {'x': 223, 'y': 338}, {'x': 223, 'y': 350}, {'x': 90, 'y': 350}], 'row_start': 5, 'col_start': 1, 'words': '资本净额'}, {'col_end': 3, 'row_end': 6, 'cell_location': [{'x': 223, 'y': 338}, {'x': 286, 'y': 338}, {'x': 286, 'y': 350}, {'x': 223, 'y': 350}], 'row_start': 5, 'col_start': 2, 'words': '4,303,263'}, {'col_end': 4, 'row_end': 6, 'cell_location': [{'x': 286, 'y': 338}, {'x': 349, 'y': 338}, {'x': 349, 'y': 350}, {'x': 286, 'y': 350}], 'row_start': 5, 'col_start': 3, 'words': '4,285,564'}, {'col_end': 5, 'row_end': 6, 'cell_location': [{'x': 349, 'y': 338}, {'x': 412, 'y': 338}, {'x': 412, 'y': 350}, {'x': 349, 'y': 350}], 'row_start': 5, 'col_start': 4, 'words': '4,175,087'}, {'col_end': 6, 'row_end': 6, 'cell_location': [{'x': 412, 'y': 338}, {'x': 475, 'y': 338}, {'x': 475, 'y': 350}, {'x': 412, 'y': 350}], 'row_start': 5, 'col_start': 5, 'words': '4,175,290'}, {'col_end': 2, 'row_end': 7, 'cell_location': [{'x': 59, 'y': 350}, {'x': 222, 'y': 350}, {'x': 222, 'y': 362}, {'x': 59, 'y': 362}], 'row_start': 6, 'col_start': 0, 'words': '风险加权资产（数额）'}, {'col_end': 3, 'row_end': 7, 'cell_location': [{'x': 223, 'y': 350}, {'x': 286, 'y': 350}, {'x': 286, 'y': 363}, {'x': 223, 'y': 363}], 'row_start': 6, 'col_start': 2, 'words': ''}, {'col_end': 4, 'row_end': 7, 'cell_location': [{'x': 286, 'y': 350}, {'x': 349, 'y': 350}, {'x': 349, 'y': 363}, {'x': 286, 'y': 363}], 'row_start': 6, 'col_start': 3, 'words': ''}, {'col_end': 5, 'row_end': 7, 'cell_location': [{'x': 349, 'y': 350}, {'x': 412, 'y': 350}, {'x': 412, 'y': 363}, {'x': 349, 'y': 363}], 'row_start': 6, 'col_start': 4, 'words': ''}, {'col_end': 6, 'row_end': 7, 'cell_location': [{'x': 412, 'y': 350}, {'x': 475, 'y': 350}, {'x': 475, 'y': 363}, {'x': 412, 'y': 363}], 'row_start': 6, 'col_start': 5, 'words': ''}, {'col_end': 1, 'row_end': 8, 'cell_location': [{'x': 60, 'y': 363}, {'x': 90, 'y': 363}, {'x': 90, 'y': 376}, {'x': 60, 'y': 376}], 'row_start': 7, 'col_start': 0, 'words': '4'}, {'col_end': 2, 'row_end': 8, 'cell_location': [{'x': 90, 'y': 363}, {'x': 223, 'y': 363}, {'x': 223, 'y': 376}, {'x': 90, 'y': 376}], 'row_start': 7, 'col_start': 1, 'words': '风险加权资产合计'}, {'col_end': 3, 'row_end': 8, 'cell_location': [{'x': 223, 'y': 363}, {'x': 286, 'y': 363}, {'x': 286, 'y': 376}, {'x': 223, 'y': 376}], 'row_start': 7, 'col_start': 2, 'words': '21,854,590'}, {'col_end': 4, 'row_end': 8, 'cell_location': [{'x': 286, 'y': 363}, {'x': 349, 'y': 363}, {'x': 349, 'y': 376}, {'x': 286, 'y': 376}], 'row_start': 7, 'col_start': 3, 'words': '22,150,555'}, {'col_end': 5, 'row_end': 8, 'cell_location': [{'x': 349, 'y': 363}, {'x': 412, 'y': 363}, {'x': 412, 'y': 376}, {'x': 349, 'y': 376}], 'row_start': 7, 'col_start': 4, 'words': '21,690,492'}, {'col_end': 6, 'row_end': 8, 'cell_location': [{'x': 412, 'y': 363}, {'x': 475, 'y': 363}, {'x': 475, 'y': 376}, {'x': 412, 'y': 376}], 'row_start': 7, 'col_start': 5, 'words': '21,586,165'}, {'col_end': 1, 'row_end': 9, 'cell_location': [{'x': 60, 'y': 376}, {'x': 90, 'y': 376}, {'x': 90, 'y': 400}, {'x': 60, 'y': 400}], 'row_start': 8, 'col_start': 0, 'words': '4a'}, {'col_end': 2, 'row_end': 9, 'cell_location': [{'x': 90, 'y': 376}, {'x': 223, 'y': 376}, {'x': 223, 'y': 400}, {'x': 90, 'y': 400}], 'row_start': 8, 'col_start': 1, 'words': '风险加权资产合计（应用资本\n底线前)'}, {'col_end': 3, 'row_end': 9, 'cell_location': [{'x': 223, 'y': 376}, {'x': 286, 'y': 376}, {'x': 286, 'y': 400}, {'x': 223, 'y': 400}], 'row_start': 8, 'col_start': 2, 'words': '21,854,590'}, {'col_end': 4, 'row_end': 9, 'cell_location': [{'x': 286, 'y': 376}, {'x': 349, 'y': 376}, {'x': 349, 'y': 400}, {'x': 286, 'y': 400}], 'row_start': 8, 'col_start': 3, 'words': '22,150,555'}, {'col_end': 5, 'row_end': 9, 'cell_location': [{'x': 349, 'y': 376}, {'x': 412, 'y': 376}, {'x': 412, 'y': 400}, {'x': 349, 'y': 400}], 'row_start': 8, 'col_start': 4, 'words': '21,690,492'}, {'col_end': 6, 'row_end': 9, 'cell_location': [{'x': 412, 'y': 376}, {'x': 475, 'y': 376}, {'x': 475, 'y': 400}, {'x': 412, 'y': 400}], 'row_start': 8, 'col_start': 5, 'words': '21,586,165'}, {'col_end': 2, 'row_end': 10, 'cell_location': [{'x': 59, 'y': 400}, {'x': 222, 'y': 400}, {'x': 222, 'y': 412}, {'x': 59, 'y': 412}], 'row_start': 9, 'col_start': 0, 'words': '资本充足率'}, {'col_end': 3, 'row_end': 10, 'cell_location': [{'x': 223, 'y': 400}, {'x': 286, 'y': 400}, {'x': 286, 'y': 413}, {'x': 223, 'y': 413}], 'row_start': 9, 'col_start': 2, 'words': ''}, {'col_end': 4, 'row_end': 10, 'cell_location': [{'x': 286, 'y': 400}, {'x': 349, 'y': 400}, {'x': 349, 'y': 413}, {'x': 286, 'y': 413}], 'row_start': 9, 'col_start': 3, 'words': ''}, {'col_end': 5, 'row_end': 10, 'cell_location': [{'x': 349, 'y': 400}, {'x': 412, 'y': 400}, {'x': 412, 'y': 413}, {'x': 349, 'y': 413}], 'row_start': 9, 'col_start': 4, 'words': ''}, {'col_end': 6, 'row_end': 10, 'cell_location': [{'x': 412, 'y': 400}, {'x': 475, 'y': 400}, {'x': 475, 'y': 413}, {'x': 412, 'y': 413}], 'row_start': 9, 'col_start': 5, 'words': ''}, {'col_end': 1, 'row_end': 11, 'cell_location': [{'x': 60, 'y': 413}, {'x': 90, 'y': 413}, {'x': 90, 'y': 425}, {'x': 60, 'y': 425}], 'row_start': 10, 'col_start': 0, 'words': '5'}, {'col_end': 2, 'row_end': 11, 'cell_location': [{'x': 90, 'y': 413}, {'x': 223, 'y': 413}, {'x': 223, 'y': 425}, {'x': 90, 'y': 425}], 'row_start': 10, 'col_start': 1, 'words': '核心一级资本充足率(%)'}, {'col_end': 3, 'row_end': 11, 'cell_location': [{'x': 223, 'y': 413}, {'x': 286, 'y': 413}, {'x': 286, 'y': 425}, {'x': 223, 'y': 425}], 'row_start': 10, 'col_start': 2, 'words': '14.48'}, {'col_end': 4, 'row_end': 11, 'cell_location': [{'x': 286, 'y': 413}, {'x': 349, 'y': 413}, {'x': 349, 'y': 425}, {'x': 286, 'y': 425}], 'row_start': 10, 'col_start': 3, 'words': '14.10'}, {'col_end': 5, 'row_end': 11, 'cell_location': [{'x': 349, 'y': 413}, {'x': 412, 'y': 413}, {'x': 412, 'y': 425}, {'x': 349, 'y': 425}], 'row_start': 10, 'col_start': 4, 'words': '14.01'}, {'col_end': 6, 'row_end': 11, 'cell_location': [{'x': 412, 'y': 413}, {'x': 475, 'y': 413}, {'x': 475, 'y': 425}, {'x': 412, 'y': 425}], 'row_start': 10, 'col_start': 5, 'words': '14.11'}, {'col_end': 1, 'row_end': 12, 'cell_location': [{'x': 60, 'y': 425}, {'x': 90, 'y': 425}, {'x': 90, 'y': 449}, {'x': 60, 'y': 449}], 'row_start': 11, 'col_start': 0, 'words': '5a'}, {'col_end': 2, 'row_end': 12, 'cell_location': [{'x': 90, 'y': 425}, {'x': 223, 'y': 425}, {'x': 223, 'y': 449}, {'x': 90, 'y': 449}], 'row_start': 11, 'col_start': 1, 'words': '核心一级资本充足率(%)\n(应用资本底线前)'}, {'col_end': 3, 'row_end': 12, 'cell_location': [{'x': 223, 'y': 425}, {'x': 286, 'y': 425}, {'x': 286, 'y': 449}, {'x': 223, 'y': 449}], 'row_start': 11, 'col_start': 2, 'words': '14.48'}, {'col_end': 4, 'row_end': 12, 'cell_location': [{'x': 286, 'y': 425}, {'x': 349, 'y': 425}, {'x': 349, 'y': 449}, {'x': 286, 'y': 449}], 'row_start': 11, 'col_start': 3, 'words': '14.10'}, {'col_end': 5, 'row_end': 12, 'cell_location': [{'x': 349, 'y': 425}, {'x': 412, 'y': 425}, {'x': 412, 'y': 449}, {'x': 349, 'y': 449}], 'row_start': 11, 'col_start': 4, 'words': '14.01'}, {'col_end': 6, 'row_end': 12, 'cell_location': [{'x': 412, 'y': 425}, {'x': 475, 'y': 425}, {'x': 475, 'y': 449}, {'x': 412, 'y': 449}], 'row_start': 11, 'col_start': 5, 'words': '14.11'}, {'col_end': 1, 'row_end': 13, 'cell_location': [{'x': 60, 'y': 449}, {'x': 90, 'y': 449}, {'x': 90, 'y': 463}, {'x': 60, 'y': 463}], 'row_start': 12, 'col_start': 0, 'words': '6'}, {'col_end': 2, 'row_end': 13, 'cell_location': [{'x': 90, 'y': 449}, {'x': 223, 'y': 449}, {'x': 223, 'y': 463}, {'x': 90, 'y': 463}], 'row_start': 12, 'col_start': 1, 'words': '一级资本充足率(%)'}, {'col_end': 3, 'row_end': 13, 'cell_location': [{'x': 223, 'y': 449}, {'x': 286, 'y': 449}, {'x': 286, 'y': 463}, {'x': 223, 'y': 463}], 'row_start': 12, 'col_start': 2, 'words': '15.21'}, {'col_end': 4, 'row_end': 13, 'cell_location': [{'x': 286, 'y': 449}, {'x': 349, 'y': 449}, {'x': 349, 'y': 463}, {'x': 286, 'y': 463}], 'row_start': 12, 'col_start': 3, 'words': '15.00'}, {'col_end': 5, 'row_end': 13, 'cell_location': [{'x': 349, 'y': 449}, {'x': 412, 'y': 449}, {'x': 412, 'y': 463}, {'x': 349, 'y': 463}], 'row_start': 12, 'col_start': 4, 'words': '14.92'}, {'col_end': 6, 'row_end': 13, 'cell_location': [{'x': 412, 'y': 449}, {'x': 475, 'y': 449}, {'x': 475, 'y': 463}, {'x': 412, 'y': 463}], 'row_start': 12, 'col_start': 5, 'words': '15.04'}, {'col_end': 1, 'row_end': 14, 'cell_location': [{'x': 60, 'y': 463}, {'x': 90, 'y': 463}, {'x': 90, 'y': 487}, {'x': 60, 'y': 487}], 'row_start': 13, 'col_start': 0, 'words': '6a'}, {'col_end': 2, 'row_end': 14, 'cell_location': [{'x': 90, 'y': 463}, {'x': 223, 'y': 463}, {'x': 223, 'y': 487}, {'x': 90, 'y': 487}], 'row_start': 13, 'col_start': 1, 'words': '一级资本充足率(%)（应用\n资本底线前)'}, {'col_end': 3, 'row_end': 14, 'cell_location': [{'x': 223, 'y': 463}, {'x': 286, 'y': 463}, {'x': 286, 'y': 487}, {'x': 223, 'y': 487}], 'row_start': 13, 'col_start': 2, 'words': '15.21'}, {'col_end': 4, 'row_end': 14, 'cell_location': [{'x': 286, 'y': 463}, {'x': 349, 'y': 463}, {'x': 349, 'y': 487}, {'x': 286, 'y': 487}], 'row_start': 13, 'col_start': 3, 'words': '15.00'}, {'col_end': 5, 'row_end': 14, 'cell_location': [{'x': 349, 'y': 463}, {'x': 412, 'y': 463}, {'x': 412, 'y': 487}, {'x': 349, 'y': 487}], 'row_start': 13, 'col_start': 4, 'words': '14.92'}, {'col_end': 6, 'row_end': 14, 'cell_location': [{'x': 412, 'y': 463}, {'x': 475, 'y': 463}, {'x': 475, 'y': 487}, {'x': 412, 'y': 487}], 'row_start': 13, 'col_start': 5, 'words': '15.04'}, {'col_end': 1, 'row_end': 15, 'cell_location': [{'x': 60, 'y': 487}, {'x': 90, 'y': 487}, {'x': 90, 'y': 500}, {'x': 60, 'y': 500}], 'row_start': 14, 'col_start': 0, 'words': '7'}, {'col_end': 2, 'row_end': 15, 'cell_location': [{'x': 90, 'y': 487}, {'x': 223, 'y': 487}, {'x': 223, 'y': 500}, {'x': 90, 'y': 500}], 'row_start': 14, 'col_start': 1, 'words': '资本充足率(%)'}, {'col_end': 3, 'row_end': 15, 'cell_location': [{'x': 223, 'y': 487}, {'x': 286, 'y': 487}, {'x': 286, 'y': 500}, {'x': 223, 'y': 500}], 'row_start': 14, 'col_start': 2, 'words': '19.69'}, {'col_end': 4, 'row_end': 15, 'cell_location': [{'x': 286, 'y': 487}, {'x': 349, 'y': 487}, {'x': 349, 'y': 500}, {'x': 286, 'y': 500}], 'row_start': 14, 'col_start': 3, 'words': '19.35'}, {'col_end': 5, 'row_end': 15, 'cell_location': [{'x': 349, 'y': 487}, {'x': 412, 'y': 487}, {'x': 412, 'y': 500}, {'x': 349, 'y': 500}], 'row_start': 14, 'col_start': 4, 'words': '19.25'}, {'col_end': 6, 'row_end': 15, 'cell_location': [{'x': 412, 'y': 487}, {'x': 475, 'y': 487}, {'x': 475, 'y': 500}, {'x': 412, 'y': 500}], 'row_start': 14, 'col_start': 5, 'words': '19.34'}, {'col_end': 1, 'row_end': 16, 'cell_location': [{'x': 60, 'y': 500}, {'x': 90, 'y': 500}, {'x': 90, 'y': 524}, {'x': 60, 'y': 524}], 'row_start': 15, 'col_start': 0, 'words': '7a'}, {'col_end': 2, 'row_end': 16, 'cell_location': [{'x': 90, 'y': 500}, {'x': 223, 'y': 500}, {'x': 223, 'y': 524}, {'x': 90, 'y': 524}], 'row_start': 15, 'col_start': 1, 'words': '资本充足率(%)（应用资本\n底线前)'}, {'col_end': 3, 'row_end': 16, 'cell_location': [{'x': 223, 'y': 500}, {'x': 286, 'y': 500}, {'x': 286, 'y': 524}, {'x': 223, 'y': 524}], 'row_start': 15, 'col_start': 2, 'words': '19.69'}, {'col_end': 4, 'row_end': 16, 'cell_location': [{'x': 286, 'y': 500}, {'x': 349, 'y': 500}, {'x': 349, 'y': 524}, {'x': 286, 'y': 524}], 'row_start': 15, 'col_start': 3, 'words': '19.35'}, {'col_end': 5, 'row_end': 16, 'cell_location': [{'x': 349, 'y': 500}, {'x': 412, 'y': 500}, {'x': 412, 'y': 524}, {'x': 349, 'y': 524}], 'row_start': 15, 'col_start': 4, 'words': '19.25'}, {'col_end': 6, 'row_end': 16, 'cell_location': [{'x': 412, 'y': 500}, {'x': 475, 'y': 500}, {'x': 475, 'y': 524}, {'x': 412, 'y': 524}], 'row_start': 15, 'col_start': 5, 'words': '19.34'}, {'col_end': 2, 'row_end': 17, 'cell_location': [{'x': 59, 'y': 524}, {'x': 222, 'y': 524}, {'x': 222, 'y': 536}, {'x': 59, 'y': 536}], 'row_start': 16, 'col_start': 0, 'words': '其他各级资本要求'}, {'col_end': 3, 'row_end': 17, 'cell_location': [{'x': 223, 'y': 524}, {'x': 286, 'y': 524}, {'x': 286, 'y': 537}, {'x': 223, 'y': 537}], 'row_start': 16, 'col_start': 2, 'words': ''}, {'col_end': 4, 'row_end': 17, 'cell_location': [{'x': 286, 'y': 524}, {'x': 349, 'y': 524}, {'x': 349, 'y': 537}, {'x': 286, 'y': 537}], 'row_start': 16, 'col_start': 3, 'words': ''}, {'col_end': 5, 'row_end': 17, 'cell_location': [{'x': 349, 'y': 524}, {'x': 412, 'y': 524}, {'x': 412, 'y': 537}, {'x': 349, 'y': 537}], 'row_start': 16, 'col_start': 4, 'words': ''}, {'col_end': 6, 'row_end': 17, 'cell_location': [{'x': 412, 'y': 524}, {'x': 475, 'y': 524}, {'x': 475, 'y': 537}, {'x': 412, 'y': 537}], 'row_start': 16, 'col_start': 5, 'words': ''}, {'col_end': 1, 'row_end': 18, 'cell_location': [{'x': 60, 'y': 537}, {'x': 90, 'y': 537}, {'x': 90, 'y': 550}, {'x': 60, 'y': 550}], 'row_start': 17, 'col_start': 0, 'words': '8'}, {'col_end': 2, 'row_end': 18, 'cell_location': [{'x': 90, 'y': 537}, {'x': 223, 'y': 537}, {'x': 223, 'y': 550}, {'x': 90, 'y': 550}], 'row_start': 17, 'col_start': 1, 'words': '储备资本要求(%)'}, {'col_end': 3, 'row_end': 18, 'cell_location': [{'x': 223, 'y': 537}, {'x': 286, 'y': 537}, {'x': 286, 'y': 550}, {'x': 223, 'y': 550}], 'row_start': 17, 'col_start': 2, 'words': '2.50'}, {'col_end': 4, 'row_end': 18, 'cell_location': [{'x': 286, 'y': 537}, {'x': 349, 'y': 537}, {'x': 349, 'y': 550}, {'x': 286, 'y': 550}], 'row_start': 17, 'col_start': 3, 'words': '2.50'}, {'col_end': 5, 'row_end': 18, 'cell_location': [{'x': 349, 'y': 537}, {'x': 412, 'y': 537}, {'x': 412, 'y': 550}, {'x': 349, 'y': 550}], 'row_start': 17, 'col_start': 4, 'words': '2.50'}, {'col_end': 6, 'row_end': 18, 'cell_location': [{'x': 412, 'y': 537}, {'x': 475, 'y': 537}, {'x': 475, 'y': 550}, {'x': 412, 'y': 550}], 'row_start': 17, 'col_start': 5, 'words': '2.50'}, {'col_end': 1, 'row_end': 19, 'cell_location': [{'x': 60, 'y': 550}, {'x': 90, 'y': 550}, {'x': 90, 'y': 562}, {'x': 60, 'y': 562}], 'row_start': 18, 'col_start': 0, 'words': '9'}, {'col_end': 2, 'row_end': 19, 'cell_location': [{'x': 90, 'y': 550}, {'x': 223, 'y': 550}, {'x': 223, 'y': 562}, {'x': 90, 'y': 562}], 'row_start': 18, 'col_start': 1, 'words': '逆周期资本要求(%)'}, {'col_end': 3, 'row_end': 19, 'cell_location': [{'x': 223, 'y': 550}, {'x': 286, 'y': 550}, {'x': 286, 'y': 562}, {'x': 223, 'y': 562}], 'row_start': 18, 'col_start': 2, 'words': '0.00'}, {'col_end': 4, 'row_end': 19, 'cell_location': [{'x': 286, 'y': 550}, {'x': 349, 'y': 550}, {'x': 349, 'y': 562}, {'x': 286, 'y': 562}], 'row_start': 18, 'col_start': 3, 'words': '0.00'}, {'col_end': 5, 'row_end': 19, 'cell_location': [{'x': 349, 'y': 550}, {'x': 412, 'y': 550}, {'x': 412, 'y': 562}, {'x': 349, 'y': 562}], 'row_start': 18, 'col_start': 4, 'words': '0.00'}, {'col_end': 6, 'row_end': 19, 'cell_location': [{'x': 412, 'y': 550}, {'x': 475, 'y': 550}, {'x': 475, 'y': 562}, {'x': 412, 'y': 562}], 'row_start': 18, 'col_start': 5, 'words': '0.00'}, {'col_end': 1, 'row_end': 20, 'cell_location': [{'x': 60, 'y': 562}, {'x': 90, 'y': 562}, {'x': 90, 'y': 599}, {'x': 60, 'y': 599}], 'row_start': 19, 'col_start': 0, 'words': '10'}, {'col_end': 2, 'row_end': 20, 'cell_location': [{'x': 90, 'y': 562}, {'x': 223, 'y': 562}, {'x': 223, 'y': 599}, {'x': 90, 'y': 599}], 'row_start': 19, 'col_start': 1, 'words': '全球系统重要性银行或国内系\n统重要性银行附加资本要求\n(%)'}, {'col_end': 3, 'row_end': 20, 'cell_location': [{'x': 223, 'y': 562}, {'x': 286, 'y': 562}, {'x': 286, 'y': 599}, {'x': 223, 'y': 599}], 'row_start': 19, 'col_start': 2, 'words': '1.50'}, {'col_end': 4, 'row_end': 20, 'cell_location': [{'x': 286, 'y': 562}, {'x': 349, 'y': 562}, {'x': 349, 'y': 599}, {'x': 286, 'y': 599}], 'row_start': 19, 'col_start': 3, 'words': '1.50'}, {'col_end': 5, 'row_end': 20, 'cell_location': [{'x': 349, 'y': 562}, {'x': 412, 'y': 562}, {'x': 412, 'y': 599}, {'x': 349, 'y': 599}], 'row_start': 19, 'col_start': 4, 'words': '1.50'}, {'col_end': 6, 'row_end': 20, 'cell_location': [{'x': 412, 'y': 562}, {'x': 475, 'y': 562}, {'x': 475, 'y': 599}, {'x': 412, 'y': 599}], 'row_start': 19, 'col_start': 5, 'words': '1.50'}, {'col_end': 1, 'row_end': 21, 'cell_location': [{'x': 60, 'y': 599}, {'x': 90, 'y': 599}, {'x': 90, 'y': 624}, {'x': 60, 'y': 624}], 'row_start': 20, 'col_start': 0, 'words': '11'}, {'col_end': 2, 'row_end': 21, 'cell_location': [{'x': 90, 'y': 599}, {'x': 223, 'y': 599}, {'x': 223, 'y': 624}, {'x': 90, 'y': 624}], 'row_start': 20, 'col_start': 1, 'words': '其他各级资本要求(%)\n(8+9+10)'}, {'col_end': 3, 'row_end': 21, 'cell_location': [{'x': 223, 'y': 599}, {'x': 286, 'y': 599}, {'x': 286, 'y': 624}, {'x': 223, 'y': 624}], 'row_start': 20, 'col_start': 2, 'words': '4.00'}, {'col_end': 4, 'row_end': 21, 'cell_location': [{'x': 286, 'y': 599}, {'x': 349, 'y': 599}, {'x': 349, 'y': 624}, {'x': 286, 'y': 624}], 'row_start': 20, 'col_start': 3, 'words': '4.00'}, {'col_end': 5, 'row_end': 21, 'cell_location': [{'x': 349, 'y': 599}, {'x': 412, 'y': 599}, {'x': 412, 'y': 624}, {'x': 349, 'y': 624}], 'row_start': 20, 'col_start': 4, 'words': '4.00'}, {'col_end': 6, 'row_end': 21, 'cell_location': [{'x': 412, 'y': 599}, {'x': 475, 'y': 599}, {'x': 475, 'y': 624}, {'x': 412, 'y': 624}], 'row_start': 20, 'col_start': 5, 'words': '4.00'}, {'col_end': 1, 'row_end': 22, 'cell_location': [{'x': 60, 'y': 624}, {'x': 90, 'y': 624}, {'x': 90, 'y': 660}, {'x': 60, 'y': 660}], 'row_start': 21, 'col_start': 0, 'words': '12\n杠杆率'}, {'col_end': 2, 'row_end': 22, 'cell_location': [{'x': 90, 'y': 624}, {'x': 223, 'y': 624}, {'x': 223, 'y': 660}, {'x': 90, 'y': 660}], 'row_start': 21, 'col_start': 1, 'words': '满足最低资本要求后的可用核\n心一级资本净额占风险加权资\n产的比例(%)'}, {'col_end': 3, 'row_end': 22, 'cell_location': [{'x': 223, 'y': 624}, {'x': 286, 'y': 624}, {'x': 286, 'y': 660}, {'x': 223, 'y': 660}], 'row_start': 21, 'col_start': 2, 'words': '9.21'}, {'col_end': 4, 'row_end': 22, 'cell_location': [{'x': 286, 'y': 624}, {'x': 349, 'y': 624}, {'x': 349, 'y': 660}, {'x': 286, 'y': 660}], 'row_start': 21, 'col_start': 3, 'words': '9.00'}, {'col_end': 5, 'row_end': 22, 'cell_location': [{'x': 349, 'y': 624}, {'x': 412, 'y': 624}, {'x': 412, 'y': 660}, {'x': 349, 'y': 660}], 'row_start': 21, 'col_start': 4, 'words': '8.92'}, {'col_end': 6, 'row_end': 22, 'cell_location': [{'x': 412, 'y': 624}, {'x': 475, 'y': 624}, {'x': 475, 'y': 660}, {'x': 412, 'y': 660}], 'row_start': 21, 'col_start': 5, 'words': '9.04'}], 'table_location': [{'x': 59, 'y': 263}, {'x': 475, 'y': 263}, {'x': 475, 'y': 660}, {'x': 59, 'y': 660}], 'footer': [{'location': [{'x': 261, 'y': 705}, {'x': 268, 'y': 705}, {'x': 268, 'y': 713}, {'x': 261, 'y': 713}], 'words': '4'}]}], 'table_num': 1, 'log_id': 1995835591557159184, 'image_info': {'image_path': 'E:\\Datas\\base_pros\\DocuVista\\test_codes\\pngs\\123.png', 'image_id': 'img_0e6447019e10'}}
    # llm_result= {'success': True, 'image_info': {'image_path': 'E:\\Datas\\base_pros\\DocuVista\\test_codes\\pngs\\123.png', 'image_id': 'img_0e6447019e10'}, 'tables_structure': {'tables': [{'id': '1', 'ocr_tables': [0], 'headers': {'cols': ['', 'b>>2024年12月31日', 'c>>2024年9月30日', 'd>>2024年6月30日', 'e>>2024年3月31日'], 'rows': ['可用资本（数额）>>1核心一级资本净额', '可用资本（数额）>>2一级资本净额', '可用资本（数额）>>3资本净额', '风险加权资产（数额）>>4风险加权资产合计', '风险加权资产（数额）>>4a风险加权资产合计（应用资本底线前）', '资本充足率>>5核心一级资本充足率（%）', '资本充足率>>5a核心一级资本充足率（%）（应用资本底线前）', '资本充足率>>6一级资本充足率（%）', '资本充足率>>6a一级资本充足率（%）（应用资本底线前）', '资本充足率>>7资本充足率（%）', '资本充足率>>7a资本充足率（%）（应用资本底线前）', '其他资本要求>>8储备资本要求（%）', '其他资本要求>>9逆周期资本要求（%）', '其他资本要求>>10系统重要性银行附加资本要求（%）', '其他资本要求>>11其他各级资本要求（%）（8+9+10）', '其他资本要求>>12满足最低资本要求后的可用核心一级资本净额占风险加权资产的比例（%）']}}]}, 'processing_stats': {'analysis_time_sec': 13.35, 'ocr_tables_count': 1, 'visual_tables_count': 1, 'token_usage': {'prompt_tokens': 1155, 'completion_tokens': 421, 'total_tokens': 1576}}}


    ocr_tables, llm_tables = extract_table_data(ocr_result, llm_result)

    # 使用（只处理第一个LLM表格结构）
    # 使用
    merged_data = merge_ocr_table_data(ocr_tables, llm_tables[0])

    # 使用
    base_table = create_base_data_table(merged_data)

    # 从LLM数据获取列标题
    llm_headers = merged_data['headers']
    col_headers = llm_headers.get('cols', [])

    # 调用函数
    table_with_col_headers = add_column_headers(base_table, col_headers)

    # 测试调用
    # ====================================
    # 假设已经有添加了列标题的表格
    # 获取LLM的行标题
    row_headers = llm_headers.get('rows', [])

    # 调用函数
    # 测试调用
    # ====================================
    # 假设 table_with_col_headers 是第5步的结果
    # row_headers 来自LLM

    # 测试调用
    # ====================================
    final_table = add_row_headers_intelligent(table_with_col_headers, row_headers)


    # ====================================
    # 测试调用示例
    # ====================================

    def test_save_to_excel(all_tables):
        """
        测试保存到Excel
        """

        # 保存到Excel
        success = save_tables_to_excel(all_tables, "output_tables1.xlsx")

        if success:
            print("\n测试完成！")
        else:
            print("\n测试失败！")

    # 如果要测试，取消注释下面一行
    all_tables = [final_table]
    test_save_to_excel(all_tables)


