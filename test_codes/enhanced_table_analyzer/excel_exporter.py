# -*- coding:utf-8 -*-
"""
Excel导出模块 - 基于新的分析结果结构生成Excel表格
"""

import os
import time
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from typing import Dict, List, Any, Optional, Tuple


class ExcelTableGenerator:
    """Excel表格生成器 - 适配新的分析结果结构"""

    def __init__(self):
        # 样式定义
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 表头样式
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')

        # 数据样式
        self.data_font = Font(name='微软雅黑', size=10)

        # 错误样式
        self.error_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
        self.error_font = Font(name='微软雅黑', size=10, bold=True, color="FF0000")

    def _parse_header_mapping(self, header_item: Dict[str, Any]) -> Tuple[str, List[Dict[str, Any]]]:
        """
        解析表头映射关系

        Args:
            header_item: 表头项，包含name和source

        Returns:
            (表头名称, 映射位置列表)
        """
        name = header_item.get("name", "")

        # 处理source字段（可能是字典或列表）
        source = header_item.get("source", {})
        if isinstance(source, dict):
            # 单个映射
            return name, [source]
        elif isinstance(source, list):
            # 多个映射
            return name, source
        else:
            return name, []

    def _get_data_from_ocr(self, ocr_data: List[List[str]],
                           start_row: int, start_col: int,
                           num_rows: int, num_cols: int) -> List[List[str]]:
        """
        从OCR数据中提取指定区域的数据

        Args:
            ocr_data: OCR原始数据
            start_row: 起始行索引（0-based）
            start_col: 起始列索引（0-based）
            num_rows: 需要提取的行数
            num_cols: 需要提取的列数

        Returns:
            提取的数据区域
        """
        if not ocr_data or start_row >= len(ocr_data):
            return []

        extracted = []
        for row_idx in range(start_row, min(start_row + num_rows, len(ocr_data))):
            row = ocr_data[row_idx]
            if start_col < len(row):
                end_col = min(start_col + num_cols, len(row))
                extracted.append(row[start_col:end_col])
            else:
                extracted.append([])

        return extracted

    def _build_complete_table(self, column_headers: List[Dict[str, Any]],
                              row_headers: List[Dict[str, Any]],
                              data_region: List[List[str]]) -> List[List[str]]:
        """
        构建完整的表格数据

        Args:
            column_headers: 列表头信息
            row_headers: 行表头信息
            data_region: 数据区域

        Returns:
            完整的表格数据（包含表头和数据）
        """
        table = []

        # 第一行：列表头（左上角空白单元格 + 列表头）
        header_row = [""]  # 左上角空白
        for header in column_headers:
            name = header.get("name", "")
            table.append(name)
        table.append(header_row)

        # 数据行：行表头 + 数据
        for i, row_header in enumerate(row_headers):
            row = []

            # 添加行表头
            row_name = row_header.get("name", "")
            row.append(row_name)

            # 添加数据
            if i < len(data_region):
                data_row = data_region[i]
                # 确保数据列数匹配
                for j in range(len(column_headers)):
                    if j < len(data_row):
                        row.append(data_row[j])
                    else:
                        row.append("")
            else:
                # 没有对应数据，填充空白
                row.extend([""] * len(column_headers))

            table.append(row)

        return table

    def create_summary_sheet(self, analysis_results: Dict[str, Any]) -> List[List[str]]:
        """
        创建详细的汇总表数据 - 修复版本

        Args:
            analysis_results: 完整的分析结果

        Returns:
            汇总表数据
        """
        import time

        summary_data = []

        # 1. 标题
        summary_data.append(["📊 表格分析汇总报告"])
        summary_data.append([])  # 空行

        # 2. 整体统计
        tables_count = analysis_results.get("tables_count", 0)
        image_path = analysis_results.get("image_path", "未知图片")
        image_id = analysis_results.get("image_id", "未知ID")

        summary_data.append(["分析报告基本信息"])
        summary_data.append(["图片文件:", os.path.basename(image_path) if image_path else "未知"])
        summary_data.append(["图片ID:", image_id])
        summary_data.append(["识别表格总数:", tables_count])
        summary_data.append(["分析时间:", time.strftime("%Y-%m-%d %H:%M:%S")])
        summary_data.append([])  # 空行

        # 3. 详细统计表头
        summary_data.append(["详细表格统计"])
        summary_data.append([
            "表格ID", "表格类型", "列数", "行数",
            "OCR表格索引", "数据起始行", "数据起始列",
            "表头层级", "状态", "备注"
        ])

        # 4. 提取表格数据 - 简化版本
        tables = []
        if "tables_analysis" in analysis_results:
            tables_analysis = analysis_results["tables_analysis"]

            # 只取第一个成功的分析
            for table_analysis in tables_analysis:
                if table_analysis.get("success"):
                    analysis_result = table_analysis.get("analysis_result", {})
                    if "tables" in analysis_result:
                        tables = analysis_result["tables"]
                        break

        print(f"找到 {len(tables)} 个表格用于汇总")

        # 5. 表格详情
        processed_count = 0
        for idx, table_info in enumerate(tables):
            table_id = table_info.get("id", idx + 1)  # 使用1-based的ID

            # 获取列数和行数
            column_headers = table_info.get("column_headers", [])
            row_headers = table_info.get("row_headers", [])

            column_count = len(column_headers) if column_headers else 0
            row_count = len(row_headers) if row_headers else 0

            if column_count == 0 or row_count == 0:
                print(f"表格 {table_id} 缺少表头数据，添加为失败状态")
                summary_data.append([
                    table_id,
                    "未知",
                    0,
                    0,
                    "未知",
                    "未知",
                    "未知",
                    "未知",
                    "失败",
                    "缺少表头数据"
                ])
                continue

            # 获取表格类型 - 增强版本（检查header_structure）
            if "header_structure" in table_info:
                header_structure = table_info.get("header_structure", {})
                table_type_from_structure = header_structure.get("type", "")

                if table_type_from_structure:
                    if table_type_from_structure == "simple":
                        table_type = "简单表格"
                    elif table_type_from_structure == "hierarchical":
                        table_type = "分层表格"
                    elif table_type_from_structure == "cross":
                        table_type = "交叉表格"
                    else:
                        table_type = self._infer_table_type(table_info)
                else:
                    table_type = self._infer_table_type(table_info)
            else:
                table_type = self._infer_table_type(table_info)

            # 获取OCR表格索引
            ocr_table_index = "未知"
            if column_headers:
                first_col = column_headers[0]
                source = first_col.get("source", {})
                if isinstance(source, dict):
                    ocr_table_index = source.get("ocr_table", idx)  # 默认使用当前索引
                elif isinstance(source, list) and source:
                    ocr_table_index = source[0].get("ocr_table", idx)
            else:
                ocr_table_index = idx  # 默认使用当前索引

            # 获取数据起始位置
            data_start = table_info.get("data_start", {})
            start_row = data_start.get("row", 1)  # 默认为第1行
            start_col = data_start.get("column", 1)  # 默认为第1列

            # 判断表头层级
            has_hierarchy = False
            for col in column_headers:
                name = col.get("name", "")
                if "|→" in str(name):
                    has_hierarchy = True
                    break

            # 获取状态和备注
            status = "成功"
            remarks = ""

            # 检查识别错误
            error_found = False
            for col in column_headers:
                name = col.get("name", "")
                if self._looks_like_error(str(name)):
                    error_found = True
                    break

            if error_found:
                status = "部分错误"
                remarks = "存在识别错误的表头"

            # 如果有header_structure，添加额外信息
            if "header_structure" in table_info:
                header_structure = table_info["header_structure"]
                common_headers = header_structure.get("common_headers", [])
                if common_headers:
                    if remarks:
                        remarks += f"; 公共表头:{len(common_headers)}个"
                    else:
                        remarks = f"公共表头:{len(common_headers)}个"

            # 添加到汇总数据
            summary_data.append([
                table_id,
                table_type,
                column_count,
                row_count,
                ocr_table_index,
                start_row,
                start_col,
                "多级" if has_hierarchy else "单级",
                status,
                remarks
            ])

            processed_count += 1



        # 6. 如果没有任何表格被处理，添加提示
        if processed_count == 0:
            summary_data.append(["", "未找到有效的表格数据", "", "", "", "", "", "", "", ""])
            summary_data.append(["", "请检查分析结果JSON结构", "", "", "", "", "", "", "", ""])

        # 7. 质量评估
        if processed_count > 0:
            summary_data.append([])
            summary_data.append(["质量评估"])

            # 计算统计信息
            total_cols = 0
            total_rows = 0
            hierarchy_count = 0
            success_count = 0

            for i, table_info in enumerate(tables):
                column_headers = table_info.get("column_headers", [])
                row_headers = table_info.get("row_headers", [])

                # 只统计成功的表格
                if column_headers and row_headers:
                    total_cols += len(column_headers)
                    total_rows += len(row_headers)
                    success_count += 1

                    # 检查分层表头
                    for col in column_headers:
                        name = col.get("name", "")
                        if "|→" in str(name):
                            hierarchy_count += 1
                            break

            if success_count > 0:
                avg_cols = total_cols / success_count
                avg_rows = total_rows / success_count

                summary_data.append([f"成功分析表格: {success_count}/{processed_count}"])
                summary_data.append([f"平均列数: {avg_cols:.1f}"])
                summary_data.append([f"平均行数: {avg_rows:.1f}"])
                summary_data.append([f"分层表头表格: {hierarchy_count}/{success_count}"])
            else:
                summary_data.append([f"成功分析表格: 0/{processed_count}"])

        return summary_data

    def _infer_table_type(self, table_info: Dict[str, Any]) -> str:
        """
        推断表格类型

        Args:
            table_info: 表格信息

        Returns:
            表格类型字符串
        """
        column_headers = table_info.get("column_headers", [])
        row_headers = table_info.get("row_headers", [])

        # 检查表头关键词
        table_keywords = {
            "财务报表": ["利润", "收入", "成本", "资产", "负债", "现金流", "损益"],
            "股东信息": ["股东", "持股", "比例", "金额", "法人", "自然人"],
            "关联方交易": ["关联", "贷款", "余额", "信用证", "承兑汇票", "占资本"],
            "现金流量": ["经营", "投资", "筹资", "现金流量", "调整项目"],
            "贷款信息": ["贷款", "贴现", "五级分类", "承兑汇票", "信用证"]
        }

        # 检查所有表头
        all_headers = []
        for col in column_headers:
            all_headers.append(col.get("name", ""))
        for row in row_headers:
            all_headers.append(row.get("name", ""))

        # 匹配关键词
        for table_type, keywords in table_keywords.items():
            for keyword in keywords:
                for header in all_headers:
                    if keyword in header:
                        return table_type

        return "其他表格"

    def _looks_like_error(self, text: str) -> bool:
        """
        判断文本是否看起来像识别错误

        Args:
            text: 要检查的文本

        Returns:
            是否可能是识别错误
        """
        if not text:
            return False

        # 常见识别错误模式
        error_patterns = [
            "银行现职",  # 应该是"银行职务"
            "利总",  # 应该是"利润"
            "收人",  # 应该是"收入"
            "用用",  # 应该是"费用"
        ]

        return any(pattern in text for pattern in error_patterns)

    def _extract_ocr_data(self, ocr_extract: Dict[str, Any]) -> List[List[str]]:
        """
        从OCR提取数据中获取完整的表格数据

        Args:
            ocr_extract: OCR提取的数据

        Returns:
            二维列表表示的完整表格数据
        """
        if not ocr_extract:
            return []

        # 尝试从不同位置获取OCR数据
        extracted_data = ocr_extract.get("extracted_data", {})

        # 优先使用 top_rows_all_cols（包含所有行的完整数据）
        table_data = extracted_data.get("top_rows_all_cols", [])

        if not table_data:
            # 尝试其他可能的数据源
            table_data = extracted_data.get("full_table", [])

        if not table_data:
            # 尝试从原始OCR结构获取
            if "tables_result" in ocr_extract:
                tables_result = ocr_extract.get("tables_result", [])
                if tables_result and isinstance(tables_result, list) and len(tables_result) > 0:
                    # 解析百度OCR格式
                    return self._parse_baidu_ocr_to_table(tables_result[0])

        return table_data if table_data else []

    def _parse_baidu_ocr_to_table(self, table_data: Dict[str, Any]) -> List[List[str]]:
        """
        解析百度OCR表格数据为二维列表

        Args:
            table_data: 百度OCR表格数据

        Returns:
            二维列表表格数据
        """
        body_cells = table_data.get("body", [])

        if not body_cells:
            return []

        # 计算表格最大行列
        max_row = 0
        max_col = 0
        for cell in body_cells:
            max_row = max(max_row, cell.get("row_end", 0))
            max_col = max(max_col, cell.get("col_end", 0))

        # 创建空表格
        table = [["" for _ in range(max_col + 1)] for _ in range(max_row + 1)]

        # 填充数据
        for cell in body_cells:
            row_start = cell.get("row_start", 0)
            col_start = cell.get("col_start", 0)
            row_end = cell.get("row_end", row_start)
            col_end = cell.get("col_end", col_start)
            words = cell.get("words", "")

            # 处理合并单元格
            for r in range(row_start, row_end + 1):
                for c in range(col_start, col_end + 1):
                    if r < len(table) and c < len(table[0]):
                        table[r][c] = words

        return table

    def _build_complete_table_with_data(self,
                                        column_headers: List[Dict[str, Any]],
                                        row_headers: List[Dict[str, Any]],
                                        ocr_data: List[List[str]]) -> List[List[str]]:
        """
        构建完整的表格数据（包含表头和从OCR提取的数据）

        Args:
            column_headers: 列表头信息（包含source映射）
            row_headers: 行表头信息（包含source映射）
            ocr_data: OCR原始数据

        Returns:
            完整的表格数据
        """
        table = []

        # 1. 第一行：列表头
        header_row = [""]  # 左上角空白单元格
        for col_header in column_headers:
            name = col_header.get("name", "")
            header_row.append(name)
        table.append(header_row)

        # 2. 数据行：行表头 + 数据
        for i, row_header in enumerate(row_headers):
            row = []

            # 添加行表头
            row_name = row_header.get("name", "")
            row.append(row_name)

            # 获取行表头的OCR映射
            row_source = row_header.get("source", {})
            if isinstance(row_source, list) and row_source:
                row_source = row_source[0]  # 取第一个映射

            row_ocr_table = row_source.get("ocr_table", 0)
            row_ocr_row = row_source.get("ocr_row", i)

            # 对每一列填充数据
            for j, col_header in enumerate(column_headers):
                # 获取列表头的OCR映射
                col_source = col_header.get("source", {})
                if isinstance(col_source, list) and col_source:
                    col_source = col_source[0]  # 取第一个映射

                col_ocr_table = col_source.get("ocr_table", 0)
                col_ocr_column = col_source.get("ocr_column", j)

                # 检查是否在同一个OCR表格中
                if row_ocr_table == col_ocr_table:
                    # 从OCR数据中提取值
                    if (row_ocr_row < len(ocr_data) and
                            col_ocr_column < len(ocr_data[row_ocr_row])):
                        cell_value = ocr_data[row_ocr_row][col_ocr_column]
                    else:
                        cell_value = ""
                else:
                    # 不在同一个OCR表格，暂时留空
                    cell_value = ""
                    print(f"警告: 行表头(i={i})和列表头(j={j})不在同一个OCR表格")

                row.append(cell_value)

            table.append(row)

        return table

    def _build_table_from_headers_only(self, table_info: Dict[str, Any]) -> List[List[str]]:
        """
        仅从表头信息构建表格（无OCR数据时使用）- 简化版

        Args:
            table_info: 表格信息

        Returns:
            只包含表头的表格数据
        """
        column_headers = table_info.get("column_headers", [])
        row_headers = table_info.get("row_headers", [])

        # 提取表头名称
        column_names = []
        for col in column_headers:
            name = col.get("name", "")
            column_names.append(name)

        row_names = []
        for row in row_headers:
            name = row.get("name", "")
            row_names.append(name)

        # 构建只有表头的表格
        table = []

        # 列表头行
        header_row = [""] + column_names
        table.append(header_row)

        # 数据行（只有行表头，数据部分为空）
        for row_name in row_names:
            row = [row_name] + [""] * len(column_names)
            table.append(row)

        return table

    def build_table_from_analysis(self, table_info: Dict[str, Any],
                                  ocr_data: List[List[List[str]]]) -> List[List[str]]:
        """
        根据LLM的表格布局和OCR数据构建新表格
        """
        column_headers = table_info.get("column_headers", [])
        row_headers = table_info.get("row_headers", [])

        # 创建空表格（按照LLM的布局）
        # 第一行：列标题（左上角空白+列标题）
        # 第一列：行标题
        table = []

        # 第一行：列标题
        header_row = [""]  # 左上角空白
        for col_header in column_headers:
            header_row.append(col_header.get("name", ""))
        table.append(header_row)

        # 数据行
        for i, row_header in enumerate(row_headers):
            row = [row_header.get("name", "")]  # 行标题

            for j, col_header in enumerate(column_headers):
                # 关键：从source字段找到OCR中的位置
                cell_value = self._get_cell_from_ocr(row_header, col_header, ocr_data, i, j)
                row.append(cell_value)

            table.append(row)

        return table


    def _get_cell_from_ocr(self, row_header: Dict, col_header: Dict,
                           ocr_data: List, row_idx: int, col_idx: int) -> str:
        """
        根据表头的source字段从OCR中提取单元格数据
        """
        # 优先使用列标题的source（因为列标题通常对应数据列）
        source = col_header.get("source", {})
        if not source:
            # 如果没有source，尝试行标题的source
            source = row_header.get("source", {})

        if isinstance(source, list) and source:
            source = source[0]  # 取第一个映射

        # 获取OCR中的位置
        ocr_table_idx = source.get("ocr_table", 0)
        ocr_row = source.get("ocr_row", row_idx)  # 如果没有指定，用当前行索引
        ocr_col = source.get("ocr_column", col_idx)  # 如果没有指定，用当前列索引

        # 从OCR数据中提取
        if (ocr_table_idx < len(ocr_data) and
                ocr_row < len(ocr_data[ocr_table_idx]) and
                ocr_col < len(ocr_data[ocr_table_idx][ocr_row])):
            return ocr_data[ocr_table_idx][ocr_row][ocr_col]

        return ""  # 找不到就返回空

    def _extract_cell(self, row_source, col_source, ocr_extracts):
        """根据映射关系从OCR提取单元格数据"""
        # 优先使用列标题的source（通常更准确）
        source = col_source if col_source else row_source

        if not source:
            return ""  # 没有映射关系

        ocr_table_idx = source.get("ocr_table", 0)
        ocr_row = source.get("ocr_row", 0)
        ocr_col = source.get("ocr_column", 0)

        # 从OCR数据中提取
        if (ocr_table_idx < len(ocr_extracts) and
                ocr_extracts[ocr_table_idx] and
                ocr_row < len(ocr_extracts[ocr_table_idx]) and
                ocr_col < len(ocr_extracts[ocr_table_idx][ocr_row])):
            return ocr_extracts[ocr_table_idx][ocr_row][ocr_col]

        return ""  # 找不到就返回空

    def _extract_cell_from_ocr(self, row_source, col_source, ocr_extracts, row_idx, col_idx):
        """
        根据映射关系从OCR提取单元格数据 - 简化版本
        """
        if not ocr_extracts:
            return ""

        # 优先使用列标题的source
        source = col_source if col_source else row_source
        if not source:
            return ""

        # 处理source可能是字典或列表
        if isinstance(source, list) and source:
            source = source[0]

        # 获取OCR位置
        ocr_table_idx = source.get("ocr_table", 0)
        ocr_row = source.get("ocr_row", row_idx)
        ocr_col = source.get("ocr_column", col_idx)

        # 检查边界
        if ocr_table_idx >= len(ocr_extracts):
            return ""

        ocr_data = ocr_extracts[ocr_table_idx]

        # 尝试不同的OCR数据格式
        if isinstance(ocr_data, dict):
            # 如果是字典格式，尝试提取表格数据
            extracted_data = ocr_data.get("extracted_data", {})
            table_data = extracted_data.get("top_rows_all_cols", [])

            if ocr_row < len(table_data) and ocr_col < len(table_data[ocr_row]):
                return table_data[ocr_row][ocr_col]
        elif isinstance(ocr_data, list):
            # 如果是列表格式
            if ocr_row < len(ocr_data) and ocr_col < len(ocr_data[ocr_row]):
                return ocr_data[ocr_row][ocr_col]

        return ""


    def export_to_excel(self, analysis_results: Dict[str, Any],
                        output_path: str,
                        ocr_extracts: List[Dict[str, Any]] = None) -> str:
        """
        导出分析结果到Excel - 简化版本
        """
        # 创建工作簿
        wb = Workbook()  # 需要这行！
        default_ws = wb.active
        wb.remove(default_ws)

        # 1. 创建汇总表
        print("创建汇总表...")
        summary_data = self.create_summary_sheet(analysis_results)
        ws_summary = wb.create_sheet(title="汇总")

        # 写入汇总数据（简化版）
        for row_idx, row_data in enumerate(summary_data, 1):
            for col_idx, cell_value in enumerate(row_data, 1):
                ws_summary.cell(row=row_idx, column=col_idx, value=cell_value)

        # 2. 提取表格数据
        tables = []
        tables_analysis = analysis_results.get("tables_analysis", [])

        # 找第一个成功的分析
        for table_analysis in tables_analysis:
            if table_analysis.get("success"):
                analysis_result = table_analysis.get("analysis_result", {})
                tables = analysis_result.get("tables", [])
                break

        print(f"找到 {len(tables)} 个表格")

        # 3. 为每个表格创建工作表
        for table_idx, table in enumerate(tables):
            if not table.get("column_headers") or not table.get("row_headers"):
                continue  # 跳过没有表头的

            print(f"创建表格 {table_idx}...")

            # 创建工作表
            sheet_name = f"表格_{table_idx}"
            ws = wb.create_sheet(title=sheet_name[:31])  # Excel限制31字符

            # 写入表头
            column_headers = table["column_headers"]
            row_headers = table["row_headers"]

            # 列标题（第一行）
            ws.cell(row=1, column=1, value="")  # 左上角空白
            for col_idx, col_header in enumerate(column_headers):
                ws.cell(row=1, column=col_idx + 2, value=col_header.get("name", ""))

            # 行标题（第一列）
            for row_idx, row_header in enumerate(row_headers):
                ws.cell(row=row_idx + 2, column=1, value=row_header.get("name", ""))

            # 填充数据
            if ocr_extracts:
                for row_idx, row_header in enumerate(row_headers):
                    for col_idx, col_header in enumerate(column_headers):
                        # 提取单元格数据
                        cell_value = self._extract_cell_from_ocr(
                            row_header.get("source", {}),
                            col_header.get("source", {}),
                            ocr_extracts,
                            row_idx,
                            col_idx
                        )
                        ws.cell(row=row_idx + 2, column=col_idx + 2, value=cell_value)

            print(f"表格 {table_idx} 创建完成: {len(column_headers)}列 × {len(row_headers)}行")

        # 4. 保存工作簿
        try:
            wb.save(output_path)
            print(f"Excel文件已保存到: {output_path}")
            return output_path
        except Exception as e:
            print(f"保存Excel文件失败: {str(e)}")
            raise





    #     -----------------------
    def build_table_from_analysis(self, analysis_result: Dict[str, Any],
                                  ocr_extract: Dict[str, Any]) -> Tuple[List[List[str]], Dict[str, Any]]:
        """
        根据分析结果和OCR数据构建完整的表格（支持关联表头）
        """
        tables = analysis_result.get("tables", [])
        if not tables:
            return [], {"error": "无表格数据"}

        # 取第一个表格
        table_info = tables[0]
        table_id = table_info.get("id", 0)

        # 获取表头结构信息
        header_structure = table_info.get("header_structure", {})
        table_type = header_structure.get("type", "simple")

        # 获取表头信息
        column_headers = table_info.get("column_headers", [])
        row_headers = table_info.get("row_headers", [])
        common_headers = header_structure.get("common_headers", [])

        if not column_headers or not row_headers:
            return [], {"error": "表头数据不完整"}

        # 提取OCR完整数据
        ocr_data = self._extract_ocr_data(ocr_extract)
        if not ocr_data:
            return [], {"error": "无OCR数据"}

        print(
            f"表格 {table_id}: 类型={table_type}, 列头={len(column_headers)}, 行头={len(row_headers)}, 公共表头={len(common_headers)}")

        # 根据表格类型构建表格
        if table_type == "cross" and common_headers:
            # 交叉表格：处理公共表头
            table_data = self._build_cross_table(
                column_headers,
                row_headers,
                common_headers,
                header_structure,
                ocr_data
            )
        elif table_type == "hierarchical":
            # 分层表格
            table_data = self._build_hierarchical_table(
                column_headers,
                row_headers,
                header_structure,
                ocr_data
            )
        else:
            # 简单表格
            table_data = self._build_simple_table(
                column_headers,
                row_headers,
                ocr_data
            )

        # 构建元数据
        metadata = {
            "table_id": table_id,
            "table_type": table_type,
            "data_rows": len(row_headers),
            "data_cols": len(column_headers),
            "common_headers": len(common_headers),
            "success": True
        }

        return table_data, metadata

    def _build_cross_table(self, column_headers, row_headers, common_headers, header_structure, ocr_data):
        """
        构建交叉表格（有公共表头）
        """
        # 确定表格结构
        # 如果有公共表头，表格的第一行第一列可能是公共表头
        # 需要根据实际情况调整

        table = []

        # 获取关联映射
        row_to_column_map = header_structure.get("row_to_column_map", [])
        column_to_row_map = header_structure.get("column_to_row_map", [])

        # 方案1：公共表头放在左上角
        if common_headers:
            # 第一行：公共表头 + 列标题
            header_row = []
            for common in common_headers:
                # 从OCR中提取公共表头值
                common_value = self._extract_header_from_source(common.get("source", {}), ocr_data)
                header_row.append(common_value)

            # 添加列标题
            for col_header in column_headers:
                col_value = self._extract_header_from_source(col_header.get("source", {}), ocr_data)
                header_row.append(col_value)

            table.append(header_row)

        # 数据行
        for row_idx, row_header in enumerate(row_headers):
            row_data = []

            # 行标题
            row_value = self._extract_header_from_source(row_header.get("source", {}), ocr_data)
            row_data.append(row_value)

            # 数据单元格
            # 使用关联映射确定哪些列属于这一行
            column_indices = []
            for mapping in row_to_column_map:
                if mapping.get("row_index") == row_idx:
                    column_indices = mapping.get("column_indices", [])
                    break

            if not column_indices:
                # 如果没有映射，默认所有列
                column_indices = list(range(len(column_headers)))

            for col_idx in column_indices:
                if col_idx < len(column_headers):
                    col_header = column_headers[col_idx]
                    cell_value = self._extract_cell_by_mapping(
                        row_header.get("source", {}),
                        col_header.get("source", {}),
                        ocr_data,
                        row_idx,
                        col_idx
                    )
                    row_data.append(cell_value)

            table.append(row_data)

        return table

    def _build_hierarchical_table(self, column_headers, row_headers, header_structure, ocr_data):
        """
        构建分层表格
        """
        table = []

        # 构建多层表头
        # 这里需要解析 "|→" 分隔的层级关系
        max_column_depth = 1
        max_row_depth = 1

        # 分析列标题层级
        for col in column_headers:
            name = col.get("name", "")
            depth = name.count("|→") + 1
            max_column_depth = max(max_column_depth, depth)

        # 分析行标题层级
        for row in row_headers:
            name = row.get("name", "")
            depth = name.count("|→") + 1
            max_row_depth = max(max_row_depth, depth)

        # 构建表头区域
        # 这里简化处理，只构建一级表头
        header_row = []
        for col_header in column_headers:
            name = col_header.get("name", "")
            # 提取最后一级作为显示名称
            if "|→" in name:
                display_name = name.split("|→")[-1]
            else:
                display_name = name
            header_row.append(display_name)

        table.append(header_row)

        # 数据行
        for row_idx, row_header in enumerate(row_headers):
            row_data = []

            # 行标题
            name = row_header.get("name", "")
            if "|→" in name:
                display_name = name.split("|→")[-1]
            else:
                display_name = name
            row_data.append(display_name)

            # 数据单元格
            for col_idx, col_header in enumerate(column_headers):
                cell_value = self._extract_cell_by_mapping(
                    row_header.get("source", {}),
                    col_header.get("source", {}),
                    ocr_data,
                    row_idx,
                    col_idx
                )
                row_data.append(cell_value)

            table.append(row_data)

        return table

    def _build_simple_table(self, column_headers, row_headers, ocr_data):
        """
        构建简单表格
        """
        table = []

        # 列标题行
        header_row = []
        for col_header in column_headers:
            name = col_header.get("name", "")
            header_row.append(name)

        table.append(header_row)

        # 数据行
        for row_idx, row_header in enumerate(row_headers):
            row_data = []

            # 行标题
            name = row_header.get("name", "")
            row_data.append(name)

            # 数据单元格
            for col_idx, col_header in enumerate(column_headers):
                cell_value = self._extract_cell_by_mapping(
                    row_header.get("source", {}),
                    col_header.get("source", {}),
                    ocr_data,
                    row_idx,
                    col_idx
                )
                row_data.append(cell_value)

            table.append(row_data)

        return table

    def _extract_header_from_source(self, source, ocr_data):
        """
        从source中提取表头数据
        """
        if not source:
            return ""

        if isinstance(source, list) and source:
            source = source[0]

        ocr_table_idx = source.get("ocr_table", 0)
        ocr_row = source.get("ocr_row", 0)
        ocr_col = source.get("ocr_column", 0)

        return self._extract_from_position(ocr_table_idx, ocr_row, ocr_col, ocr_data)

    def _extract_cell_by_mapping(self, row_source, col_source, ocr_data, row_idx, col_idx):
        """
        根据行列source映射提取单元格数据
        """
        # 优先使用数据映射（如果有的话）
        # 否则使用行列source组合

        if isinstance(row_source, list) and row_source:
            row_source = row_source[0]

        if isinstance(col_source, list) and col_source:
            col_source = col_source[0]

        # 组合映射：行source提供行号，列source提供列号
        ocr_table_idx = col_source.get("ocr_table", row_source.get("ocr_table", 0))
        ocr_row = row_source.get("ocr_row", row_idx + 1)  # +1跳过标题行
        ocr_col = col_source.get("ocr_column", col_idx)

        return self._extract_from_position(ocr_table_idx, ocr_row, ocr_col, ocr_data)

    def _extract_from_position(self, ocr_table_idx, ocr_row, ocr_col, ocr_data):
        """
        从指定位置提取数据
        """
        if not ocr_data:
            return ""

        if ocr_table_idx >= len(ocr_data):
            return ""

        table_data = ocr_data[ocr_table_idx]

        if (ocr_row < len(table_data) and
                ocr_col < len(table_data[ocr_row])):
            return table_data[ocr_row][ocr_col]

        return ""

    def _infer_table_type(self, table_info: Dict[str, Any]) -> str:
        """
        推断表格类型 - 增强版本
        """
        column_headers = table_info.get("column_headers", [])
        row_headers = table_info.get("row_headers", [])
        header_structure = table_info.get("header_structure", {})

        # 首先检查header_structure中的类型
        table_type = header_structure.get("type", "")
        if table_type:
            type_map = {
                "simple": "简单表格",
                "hierarchical": "分层表格",
                "cross": "交叉表格"
            }
            return type_map.get(table_type, "其他表格")

        # 如果没有header_structure，根据表头内容推断
        # 检查分层表头
        for col in column_headers:
            name = col.get("name", "")
            if "|→" in str(name):
                return "分层表格"

        for row in row_headers:
            name = row.get("name", "")
            if "|→" in str(name):
                return "分层表格"

        # 检查表头关键词
        table_keywords = {
            "财务报表": ["利润", "收入", "成本", "资产", "负债", "现金流", "损益"],
            "股东信息": ["股东", "持股", "比例", "金额", "法人", "自然人"],
            "关联方交易": ["关联", "贷款", "余额", "信用证", "承兑汇票", "占资本"],
            "现金流量": ["经营", "投资", "筹资", "现金流量", "调整项目"],
            "贷款信息": ["贷款", "贴现", "五级分类", "承兑汇票", "信用证"]
        }

        # 检查所有表头
        all_headers = []
        for col in column_headers:
            all_headers.append(col.get("name", ""))
        for row in row_headers:
            all_headers.append(row.get("name", ""))

        # 匹配关键词
        for table_type, keywords in table_keywords.items():
            for keyword in keywords:
                for header in all_headers:
                    if keyword in header:
                        return table_type

        return "其他表格"





def export_analysis_to_excel(analysis_results: Dict[str, Any],
                             output_path: str,
                             ocr_data: List[Dict[str, Any]] = None) -> str:
    """
    导出分析结果到Excel的主函数

    Args:
        analysis_results: LLM分析结果
        output_path: 输出文件路径
        ocr_data: OCR提取数据（可选）

    Returns:
        输出的文件路径
    """
    generator = ExcelTableGenerator()
    return generator.export_to_excel(analysis_results, output_path, ocr_data)


def export_from_json(json_path: str, output_dir: str = "./output") -> str:
    """
    从JSON文件导出Excel

    Args:
        json_path: JSON文件路径
        output_dir: 输出目录

    Returns:
        输出的文件路径
    """
    os.makedirs(output_dir, exist_ok=True)

    with open(json_path, 'r', encoding='utf-8') as f:
        analysis_results = json.load(f)

    # 生成输出文件名
    base_name = os.path.splitext(os.path.basename(json_path))[0]
    if base_name.endswith('_analysis'):
        base_name = base_name[:-9]

    output_filename = f"{base_name}_tables.xlsx"
    output_path = os.path.join(output_dir, output_filename)

    # 导出Excel
    return export_analysis_to_excel(analysis_results, output_path)


# 主函数
if __name__ == "__main__":
    # 示例用法
    json_path = "analysis_results.json"
    output_dir = "./output"

    try:
        result_file = export_from_json(json_path, output_dir)
        print(f"Excel文件生成成功: {result_file}")
    except Exception as e:
        print(f"生成失败: {str(e)}")