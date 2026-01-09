
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

class TableReconstructor:
    """表格重构器：整合7步流程"""

    def __init__(self):
        self.warnings = []
        self.issues = []

    # ========== 工具函数 ==========
    def log_warning(self, message):
        """记录警告"""
        self.warnings.append(message)
        print(f"⚠️ {message}")

    def log_issue(self, message):
        """记录问题"""
        self.issues.append(message)
        print(f"❓ {message}")

    def clean_text_for_matching(self, text):
        """清理文本用于匹配"""
        if not text:
            return ""
        text = str(text)
        # 替换换行、空格
        text = text.replace('\n', '').replace('\r', '').replace(' ', '')
        # 去掉特殊符号，保留中文、数字、字母
        cleaned = ''.join(c for c in text if c.isalnum() or '\u4e00-\u9fff' in c)
        return cleaned

    def calculate_similarity(self, text1, text2):
        """计算两个文本的相似度"""
        t1 = self.clean_text_for_matching(text1)
        t2 = self.clean_text_for_matching(text2)

        if not t1 or not t2:
            return 0

        # 完全相等
        if t1 == t2:
            return 1.0

        # 包含关系
        if t1 in t2 or t2 in t1:
            return 0.9

        # 公共字符比例
        common_chars = set(t1) & set(t2)
        if not common_chars:
            return 0

        # 计算Jaccard相似度
        union_chars = set(t1) | set(t2)
        return len(common_chars) / len(union_chars)

    def find_best_match_row(self, target_text, table, start_row=1, search_cols=None):
        """
        在表格中查找与目标文本最匹配的行
        """
        if not target_text or not table:
            return -1, 0.0

        best_row = -1
        best_score = 0.0

        # 确定搜索列范围
        if search_cols is None:
            # 默认搜索所有列
            search_cols = list(range(len(table[0])))

        for row_idx in range(start_row, len(table)):
            row_score = 0.0

            for col_idx in search_cols:
                if col_idx >= len(table[row_idx]):
                    continue

                cell_text = table[row_idx][col_idx]
                if not cell_text:
                    continue

                # 计算相似度
                similarity = self.calculate_similarity(target_text, str(cell_text))
                if similarity > row_score:
                    row_score = similarity

            # 更新最佳匹配
            if row_score > best_score:
                best_score = row_score
                best_row = row_idx

        return best_row, best_score

    # ========== 核心7步 ==========

    def step1_prepare_data(self, ocr_result, llm_result):
        """第1步：准备数据"""
        # 直接返回传入的数据
        return ocr_result, llm_result

    def step2_extract_table_data(self, ocr_result, llm_result):
        """
        从数据中提取表格信息
        """
        # 提取OCR表格
        ocr_tables = ocr_result.get('tables_result', [])
        print(f"OCR表格数: {len(ocr_tables)}")

        # 提取LLM表格结构
        llm_tables = []
        if 'tables_structure' in llm_result:
            llm_tables = llm_result['tables_structure'].get('tables', [])
        elif 'tables' in llm_result:
            llm_tables = llm_result['tables']
        print(f"LLM表格结构数: {len(llm_tables)}")

        return ocr_tables, llm_tables

    def step3_merge_ocr_tables(self, ocr_tables, llm_table_info):
        """
        将多个OCR表格的数据合并成一个表格的数据
        LLM指定哪些OCR表格属于同一个逻辑表格
        """
        # 获取要合并的OCR表格索引
        ocr_table_indices = llm_table_info.get('ocr_tables', [])
        print(f"要合并的OCR表格索引: {ocr_table_indices}")

        # 收集所有单元格
        all_cells = []
        row_offset = 0  # 行偏移量，用于拼接表格

        print("**************************ocr_table_indices:")
        print(ocr_table_indices)

        for idx in ocr_table_indices:
            if idx >= len(ocr_tables):
                print(f"警告: OCR表格索引 {idx} 超出范围")
                continue

            table = ocr_tables[idx]
            cells = table.get('body', [])

            # 调整行号（表格拼接时行号要累加）
            for cell in cells:
                adjusted_cell = cell.copy()
                adjusted_cell['row_start'] += row_offset
                adjusted_cell['row_end'] += row_offset
                all_cells.append(adjusted_cell)

            # 更新行偏移量：找到这个表格的最大行号
            if cells:
                max_row_in_table = max(cell['row_end'] for cell in cells)
                row_offset = max_row_in_table + 2  # +2留一个空行分隔

        return {
            'cells': all_cells,  # 合并后的所有单元格
            'headers': llm_table_info.get('headers', {}),  # 表头信息
            'table_id': llm_table_info.get('id', '1')  # 表格ID
        }

    def step4_create_base_data_table(self, merged_data):
        """
            创建基础数据表格
            检查每行列数是否一致
            """
        cells = merged_data['cells']

        if not cells:
            return []

        # 找出最大行列
        max_row = 0
        max_col = 0
        for cell in cells:
            max_row = max(max_row, cell['row_end'])
            max_col = max(max_col, cell['col_end'])

        num_rows = max_row
        num_cols = max_col

        # 检查每行的列数
        print("检查每行列数:")
        row_col_counts = {}
        for cell in cells:
            row = cell['row_start']
            col_end = cell['col_end']

            if row not in row_col_counts:
                row_col_counts[row] = []
            row_col_counts[row].append(col_end)

        # 找出每行的最大列号
        row_max_cols = {}
        for row, col_ends in row_col_counts.items():
            row_max_cols[row] = max(col_ends)

        # 创建表格
        table = []
        for r in range(num_rows):
            table.append([None] * num_cols)

        # 填充数据
        for cell in cells:
            value = cell['words']
            row_idx = cell['row_start']
            col_idx = cell['col_start']

            if row_idx < num_rows and col_idx < num_cols:
                table[row_idx][col_idx] = value

        return table

    def step5_add_column_headers(self, table, col_headers):
        """
        最终版：LLM列数是最终列数，从右向左确定列
        """
        if not col_headers:
            print("无列标题")
            return table

        if not table:
            print("表格为空")
            return table

        current_cols = len(table[0])  # OCR表格的列数
        target_cols = len(col_headers)  # LLM的最终列数

        # 情况1：OCR列数 > LLM列数（需要删除左侧多余的列）
        if current_cols > target_cols:
            excess_cols = current_cols - target_cols
            print(f"需要删除左侧{excess_cols}列")

            # 删除左侧多余的列（从每行删除）
            for i in range(len(table)):
                # 保留右侧的target_cols列，删除左侧的excess_cols列
                table[i] = table[i][excess_cols:]

            current_cols = target_cols  # 更新列数

        # 情况2：OCR列数 < LLM列数（需要补充左侧空列）
        elif current_cols < target_cols:
            needed_cols = target_cols - current_cols
            print(f"需要补充左侧{needed_cols}列")

            # 在每行左侧补充空列
            for i in range(len(table)):
                table[i] = [None] * needed_cols + table[i]

            current_cols = target_cols  # 更新列数

        # 现在表格列数 = LLM列数
        print(f"调整后表格列数: {current_cols}")

        # 在顶部添加一行用于列标题
        table.insert(0, [None] * current_cols)

        # 从右向左填充列标题
        def change_name(name):
            if name == '>>':
                name = ""
            return name
        col_headers = [change_name(col) for col in col_headers]
        col_headers_copy = col_headers.copy()
        for col in range(current_cols - 1, -1, -1):
            if col_headers_copy:
                table[0][col] = col_headers_copy.pop()

        print("列标题填充完成")
        print(f"第0行（列标题）: {table[0]}")

        return table

    def step6_add_row_headers_intelligent(self, table, row_headers):
        """第6步：智能匹配填充行表头（两阶段匹配）"""
        """
            智能匹配LLM行表头到数据行 - 两阶段匹配
            """
        if not table or not row_headers:
            print("表格或行表头为空")
            return table

        print("=== 第6步：智能匹配行表头（两阶段） ===")
        print(f"表格: {len(table)}行 × {len(table[0])}列")
        print(f"LLM行表头数: {len(row_headers)}")

        # 1. 分析表格结构
        left_empty_cols = 0
        for title in table[0]:
            if not title or str(title).strip() == '':
                left_empty_cols += 1
            else:
                break

        print(f"左侧空列数: {left_empty_cols}")

        # 确定搜索列范围
        if left_empty_cols > 0:
            search_columns = list(range(left_empty_cols))
            if left_empty_cols < len(table[0]):
                search_columns.append(left_empty_cols)
        else:
            search_columns = list(range(min(3, len(table[0]))))

        print(f"搜索列范围: {search_columns}")

        # ========== 第一阶段：自由匹配 ==========
        print("\n--- 第一阶段：自由匹配 ---")
        first_pass_results = []
        used_rows = set()

        for header_idx, llm_header in enumerate(row_headers):
            if '>>' in llm_header:
                target_text = llm_header.split('>>')[-1]
            else:
                target_text = llm_header

            print(f"  表头[{header_idx}]: '{target_text}'")

            # 自由匹配：搜索所有未使用的行
            best_row = -1
            best_score = 0.0
            best_cell_text = ""

            for row_idx in range(1, len(table)):
                if row_idx in used_rows:
                    continue  # 跳过已使用的行

                # 计算该行的最佳匹配分数
                row_best_score = 0.0
                cell_text = ""

                for col_idx in search_columns:
                    if col_idx >= len(table[row_idx]):
                        continue

                    cell_val = table[row_idx][col_idx]
                    if not cell_val:
                        continue

                    score = self.calculate_similarity(target_text, str(cell_val))
                    if score > row_best_score:
                        row_best_score = score
                        cell_text = str(cell_val)

                if row_best_score > best_score:
                    best_score = row_best_score
                    best_row = row_idx
                    best_cell_text = cell_text

            first_pass_results.append({
                'header_idx': header_idx,
                'llm_header': llm_header,
                'target_text': target_text,
                'matched_row': best_row,
                'score': best_score,
                'matched_cell': best_cell_text
            })

            if best_row != -1:
                used_rows.add(best_row)
                print(f"    匹配到行{best_row}, 分数:{best_score:.2f}, 单元格:'{best_cell_text[:20]}...'")
            else:
                print(f"    未匹配")

        # ========== 第二阶段：顺序检查和调整 ==========
        print("\n--- 第二阶段：顺序检查和调整 ---")

        # 检查第一阶段匹配的顺序
        last_matched_row = 0  # 上一行匹配的行号
        adjusted_results = []
        need_adjustments = False

        for i, result in enumerate(first_pass_results):
            current_row = result['matched_row']

            if current_row == -1:
                # 未匹配，标记为需要顺序填充
                adjusted_results.append(result.copy())
                continue

            # 检查顺序：当前行应该 >= 上一行
            if current_row < last_matched_row:
                print(f"  顺序问题: LLM[{i}]匹配到行{current_row}, 但上一行是{last_matched_row}")
                need_adjustments = True

                # 重新搜索：必须在 last_matched_row+1 之后
                new_best_row = -1
                new_best_score = 0.0
                new_cell_text = ""

                target_text = result['target_text']
                for row_idx in range(last_matched_row + 1, len(table)):
                    if row_idx in used_rows:
                        continue

                    row_best_score = 0.0
                    cell_text = ""

                    for col_idx in search_columns:
                        if col_idx >= len(table[row_idx]):
                            continue

                        cell_val = table[row_idx][col_idx]
                        if not cell_val:
                            continue

                        score = self.calculate_similarity(target_text, str(cell_val))
                        if score > row_best_score:
                            row_best_score = score
                            cell_text = str(cell_val)

                    if row_best_score > new_best_score:
                        new_best_score = row_best_score
                        new_best_row = row_idx
                        new_cell_text = cell_text

                if new_best_row != -1:
                    # 移除旧的行，添加新的
                    used_rows.discard(current_row)
                    used_rows.add(new_best_row)

                    adjusted_result = result.copy()
                    adjusted_result['matched_row'] = new_best_row
                    adjusted_result['score'] = new_best_score
                    adjusted_result['matched_cell'] = new_cell_text
                    adjusted_result['adjusted'] = True

                    adjusted_results.append(adjusted_result)
                    last_matched_row = new_best_row

                    print(f"    调整为行{new_best_row}, 分数:{new_best_score:.2f}")
                else:
                    # 找不到符合条件的行，保持原样但标记问题
                    adjusted_result = result.copy()
                    adjusted_result['has_order_issue'] = True
                    adjusted_results.append(adjusted_result)
                    last_matched_row = current_row
                    print(f"    无法调整，保持行{current_row}（顺序问题）")
            else:
                # 顺序正确，保持原样
                adjusted_results.append(result.copy())
                last_matched_row = current_row

        # ========== 第三阶段：填充表格 ==========
        print("\n--- 第三阶段：填充表格 ---")

        # 清空第0列（准备填充行表头）
        for row_idx in range(1, len(table)):
            table[row_idx][0] = None

        # 按顺序填充匹配成功的行表头
        matched_count = 0
        current_data_row = 1  # 当前数据行指针

        for i, result in enumerate(adjusted_results):
            if result['matched_row'] != -1 and result.get('score', 0) > 0.3:
                # 使用匹配到的行
                target_row = result['matched_row']
                table[target_row][0] = result['llm_header']
                matched_count += 1

                # 清空其他左侧列
                for col in range(1, left_empty_cols):
                    if col < len(table[target_row]):
                        table[target_row][col] = None

                print(f"  LLM[{i}] → 行{target_row}: '{result['llm_header'][:30]}...'")
                current_data_row = target_row + 1
            else:
                # 未匹配或匹配分数低，按顺序填充
                # 找下一个可用的行
                while current_data_row < len(table) and current_data_row in used_rows:
                    current_data_row += 1

                if current_data_row < len(table):
                    table[current_data_row][0] = result['llm_header']
                    matched_count += 1
                    print(f"  LLM[{i}] → 行{current_data_row}: '{result['llm_header'][:30]}...'（顺序填充）")
                    current_data_row += 1
                else:
                    print(f"  LLM[{i}] 无法填充: 表格行数不足")

        print(f"\n匹配结果: {matched_count}/{len(row_headers)} 个行表头已填充")

        return table

    def step7_save_to_excel(self, tables_data, output_file):
        """
            将多个表格保存到Excel，每个表格一个Sheet
            tables_data: 列表，每个元素是一个表格的完整数据
            """

        print(f"\n=== 第7步：保存到Excel ===")
        print(f"要保存{len(tables_data)}个表格到: {output_file}")

        # 创建新的工作簿
        wb = openpyxl.Workbook()

        # 删除默认创建的Sheet
        if 'Sheet' in wb.sheetnames:
            default_sheet = wb['Sheet']
            wb.remove(default_sheet)

        # 处理每个表格
        for table_idx, table in enumerate(tables_data):
            if not table:
                print(f"表格{table_idx}为空，跳过")
                continue

            # 创建Sheet名称（Excel限制31字符）
            sheet_name = f"Table{table_idx + 1}"
            if len(sheet_name) > 31:
                sheet_name = sheet_name[:31]

            # 检查Sheet名称是否重复
            original_name = sheet_name
            counter = 1
            while sheet_name in wb.sheetnames:
                sheet_name = f"{original_name}_{counter}"
                counter += 1

            # 创建工作表
            ws = wb.create_sheet(title=sheet_name)

            # 获取表格尺寸
            num_rows = len(table)
            num_cols = len(table[0]) if num_rows > 0 else 0

            print(f"  表格{table_idx + 1}: {num_rows}行 × {num_cols}列 -> Sheet: '{sheet_name}'")

            # 填充数据到Excel
            for r in range(num_rows):
                for c in range(num_cols):
                    cell_value = table[r][c]

                    # Excel行号从1开始，列号从1开始
                    excel_row = r + 1
                    excel_col = c + 1

                    # 写入值
                    ws.cell(row=excel_row, column=excel_col, value=cell_value)

                    # 获取单元格对象设置样式
                    cell_obj = ws.cell(row=excel_row, column=excel_col)

                    # 设置对齐方式
                    if r == 0:  # 第0行是表头
                        cell_obj.font = Font(bold=True)
                        cell_obj.alignment = Alignment(horizontal='center', vertical='center')
                        # 浅灰色背景
                        from openpyxl.styles import PatternFill
                        cell_obj.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                    elif c == 0:  # 第0列是行表头
                        cell_obj.alignment = Alignment(horizontal='left', vertical='center')
                    else:  # 数据单元格
                        cell_obj.alignment = Alignment(horizontal='center', vertical='center')

            # 设置列宽
            for col in range(1, num_cols + 1):
                col_letter = get_column_letter(col)

                # 根据内容调整列宽
                max_length = 0
                for row in range(1, min(num_rows + 1, 50)):  # 只检查前50行
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value:
                        # 计算文本长度（中文算2个字符）
                        text_len = 0
                        for char in str(cell_value):
                            if '\u4e00-\u9fff' in char:  # 中文字符
                                text_len += 2
                            else:
                                text_len += 1
                        max_length = max(max_length, text_len)

                # 设置列宽（最小8，最大50）
                width = min(max(max_length + 2, 8), 50)
                ws.column_dimensions[col_letter].width = width

            # 添加边框
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # 为整个表格添加边框
            for row in ws.iter_rows(min_row=1, max_row=num_rows, min_col=1, max_col=num_cols):
                for cell in row:
                    cell.border = thin_border

            # 冻结首行和首列（方便查看）
            ws.freeze_panes = 'B2'  # 冻结第1行和第1列

        # 如果没有表格，创建一个空Sheet
        if len(wb.sheetnames) == 0:
            ws = wb.create_sheet(title="空表格")
            ws['A1'] = "无表格数据"

        # 保存文件
        try:
            wb.save(output_file)
            print(f"\n✅ Excel文件保存成功: {output_file}")
            print(f"   共保存了 {len(wb.sheetnames)} 个工作表")

            # 显示Sheet列表
            print("   Sheet列表:")
            for i, sheet_name in enumerate(wb.sheetnames, 1):
                ws = wb[sheet_name]
                max_row = ws.max_row
                max_column = ws.max_column
                print(f"     {i}. {sheet_name}: {max_row}行 × {max_column}列")

            return True

        except Exception as e:
            print(f"\n❌ 保存Excel文件失败: {str(e)}")
            return False

    # ========== 完整流程 ==========

    def process_single_table(self, ocr_tables, llm_table_info):
        """
        处理单个表格（第3-6步）
        返回：重构后的表格数据
        """
        print(f"\n{'=' * 60}")
        print(f"开始处理表格")
        print(f"{'=' * 60}")

        # 第3步：合并OCR表格
        merged_data = self.step3_merge_ocr_tables(ocr_tables, llm_table_info)
        if not merged_data:
            self.log_issue("表格合并失败")
            return None

        # 第4步：创建基础表格
        base_table = self.step4_create_base_data_table(merged_data)
        if not base_table:
            self.log_issue("创建基础表格失败")
            return None

        # 第5步：添加列标题
        col_headers = merged_data.get('headers', {}).get('cols', [])
        table_with_cols = self.step5_add_column_headers(base_table, col_headers)

        # 第6步：添加行标题
        row_headers = merged_data.get('headers', {}).get('rows', [])
        final_table = self.step6_add_row_headers_intelligent(table_with_cols, row_headers)

        return final_table

    def process_all_tables(self, ocr_result, llm_result, output_file="output.xlsx"):
        """
        完整处理流程：第1-7步
        """
        print("开始表格重构流程...")

        # 第1步：准备数据
        ocr_result, llm_result = self.step1_prepare_data(ocr_result, llm_result)

        # 第2步：提取表格数据
        ocr_tables, llm_tables = self.step2_extract_table_data(ocr_result, llm_result)


        print("ocr_tablesocr_tablesocr_tables")
        print(ocr_tables)

        if not ocr_tables or not llm_tables:
            self.log_issue("提取表格数据失败")
            return False

        print(f"OCR表格数: {len(ocr_tables)}")
        print(f"LLM表格结构数: {len(llm_tables)}")

        # 处理每个表格
        all_final_tables = []

        for table_idx, llm_table_info in enumerate(llm_tables):
            print(f"\n处理表格 {table_idx + 1}/{len(llm_tables)}")

            # 获取对应的OCR表格
            ocr_table_indices = llm_table_info.get('ocr_tables', [table_idx])
            tables_to_process = []

            for idx in ocr_table_indices:
                if idx < len(ocr_tables):
                    tables_to_process.append(ocr_tables[idx])
                else:
                    self.log_warning(f"OCR表格索引{idx}超出范围")

            if not tables_to_process:
                self.log_issue(f"无OCR表格可处理LLM表格{table_idx}")
                continue

            # 处理单个表格
            final_table = self.process_single_table(ocr_tables, llm_table_info)

            if final_table:
                all_final_tables.append(final_table)

        if not all_final_tables:
            self.log_issue("无表格数据生成")
            return False

        # 第7步：保存到Excel
        success = self.step7_save_to_excel(all_final_tables, output_file)

        # 输出统计信息
        print(f"\n{'=' * 60}")
        print(f"处理完成统计:")
        print(f"  成功处理表格: {len(all_final_tables)}个")
        print(f"  警告: {len(self.warnings)}个")
        print(f"  问题: {len(self.issues)}个")
        print(f"  输出文件: {output_file}")
        print(f"{'=' * 60}")

        return success


# ====================================
# 使用示例
# ====================================
def main():
    # 你的数据
    ocr_result = {}
    llm_result= {'success': True,
                 'image_info': {'image_path': 'E:\\Datas\\base_pros\\DocuVista\\test_codes\\pngs\\123.png',
                                'image_id': 'img_0e6447019e10'}, 'tables_structure': {'tables': [
            {'id': '1', 'ocr_tables': [0],
             'headers': {'cols': ['', 'b>>2024年12月31日', 'c>>2024年9月30日', 'd>>2024年6月30日', 'e>>2024年3月31日'],
                         'rows': ['可用资本（数额）>>1核心一级资本净额', '可用资本（数额）>>2一级资本净额',
                                  '可用资本（数额）>>3资本净额', '风险加权资产（数额）>>4风险加权资产合计',
                                  '风险加权资产（数额）>>4a风险加权资产合计（应用资本底线前）',
                                  '资本充足率>>5核心一级资本充足率（%）',
                                  '资本充足率>>5a核心一级资本充足率（%）（应用资本底线前）',
                                  '资本充足率>>6一级资本充足率（%）', '资本充足率>>6a一级资本充足率（%）（应用资本底线前）',
                                  '资本充足率>>7资本充足率（%）', '资本充足率>>7a资本充足率（%）（应用资本底线前）',
                                  '其他资本要求>>8储备资本要求（%）', '其他资本要求>>9逆周期资本要求（%）',
                                  '其他资本要求>>10系统重要性银行附加资本要求（%）',
                                  '其他资本要求>>11其他各级资本要求（%）（8+9+10）',
                                  '其他资本要求>>12满足最低资本要求后的可用核心一级资本净额占风险加权资产的比例（%）']}}]},
                 'processing_stats': {'analysis_time_sec': 13.35, 'ocr_tables_count': 1, 'visual_tables_count': 1,
                                      'token_usage': {'prompt_tokens': 1155, 'completion_tokens': 421,
                                                      'total_tokens': 1576}}}

    # 创建重构器并处理
    reconstructor = TableReconstructor()
    success = reconstructor.process_all_tables(
        ocr_result=ocr_result,
        llm_result=llm_result,
        output_file="../../enhanced_table_analyzer/reconstructed_tables2.xlsx"
    )

    if success:
        print("✅ 表格重构成功！")
    else:
        print("❌ 表格重构失败！")


if __name__ == "__main__":
    main()