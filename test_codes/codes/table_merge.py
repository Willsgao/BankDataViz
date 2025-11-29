# -*- coding:utf-8 -*-

import json
import re
from typing import Dict, List, Any, Tuple, Optional
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
import os


class TableDataAligner:
    """
    表格数据对齐器 - 对齐LLM和百度OCR的表格数据
    重点基于最后一层文本进行匹配
    """

    def __init__(self):
        self.matched_tables = []
        self.unmatched_llm_tables = []
        self.unmatched_ocr_tables = []

    def load_data(self, llm_path: str, ocr_path: str) -> Tuple[Dict, Dict]:
        """
        加载LLM和OCR数据

        Args:
            llm_path: LLM分析结果JSON路径
            ocr_path: 百度OCR结果JSON路径

        Returns:
            tuple: (llm_data, ocr_data)
        """
        print("📥 加载数据...")

        try:
            with open(llm_path, 'r', encoding='utf-8') as f:
                llm_data = json.load(f)

            with open(ocr_path, 'r', encoding='utf-8') as f:
                ocr_data = json.load(f)

            print(f"✅ LLM数据: {len(llm_data.get('image_results', []))} 张图片")
            print(f"✅ OCR数据: {len(ocr_data.get('tables_result', []))} 个表格")

            return llm_data, ocr_data
        except Exception as e:
            print(f"❌ 数据加载失败: {e}")
            raise

    def normalize_text(self, text: str) -> str:
        """
        文本标准化处理

        Args:
            text: 原始文本

        Returns:
            str: 标准化后的文本
        """
        if not text:
            return ""

        # 移除特殊字符、空格，转为小写
        text = re.sub(r'[^\w\u4e00-\u9fff]', '', text)
        return text.lower()

    def calculate_similarity(self, text1: str, text2: str) -> float:
        """
        计算文本相似度（基于编辑距离的简单实现）

        Args:
            text1: 文本1
            text2: 文本2

        Returns:
            float: 相似度分数(0-1)
        """
        text1_norm = self.normalize_text(text1)
        text2_norm = self.normalize_text(text2)

        if not text1_norm and not text2_norm:
            return 0.0

        if text1_norm == text2_norm:
            return 1.0

        # 简单的包含关系检查
        if text1_norm in text2_norm or text2_norm in text1_norm:
            return 0.8

        # 基于编辑距离的相似度
        distance = self.levenshtein_distance(text1_norm, text2_norm)
        max_len = max(len(text1_norm), len(text2_norm))

        if max_len == 0:
            return 0.0

        return 1 - (distance / max_len)

    def levenshtein_distance(self, s1: str, s2: str) -> int:
        """
        计算Levenshtein编辑距离

        Args:
            s1: 字符串1
            s2: 字符串2

        Returns:
            int: 编辑距离
        """
        if len(s1) < len(s2):
            return self.levenshtein_distance(s2, s1)

        if len(s2) == 0:
            return len(s1)

        previous_row = range(len(s2) + 1)
        for i, c1 in enumerate(s1):
            current_row = [i + 1]
            for j, c2 in enumerate(s2):
                insertions = previous_row[j + 1] + 1
                deletions = current_row[j] + 1
                substitutions = previous_row[j] + (c1 != c2)
                current_row.append(min(insertions, deletions, substitutions))
            previous_row = current_row

        return previous_row[-1]

    def extract_ocr_leaf_nodes(self, ocr_table: Dict) -> List[str]:
        """
        从OCR表格中提取叶子节点文本

        Args:
            ocr_table: OCR表格数据

        Returns:
            list: 叶子节点文本列表
        """
        leaves = []

        # 从body单元格中提取文本
        for cell in ocr_table.get('body', []):
            words = cell.get('words', '').strip()
            if words and words not in leaves:  # 去重
                leaves.append(words)

        # 从header中提取文本
        for header in ocr_table.get('header', []):
            words = header.get('words', '').strip()
            if words and words not in leaves:
                leaves.append(words)

        return leaves


    def merge_aligned_data(self, matches: List[Dict]) -> List[Dict]:
        """
        合并对齐后的数据

        Args:
            matches: 匹配结果列表

        Returns:
            list: 合并后的表格数据列表
        """
        print("🔄 合并对齐数据...")

        merged_tables = []

        for match in matches:
            llm_table = match['llm_table']
            ocr_table = match['ocr_table']

            merged_table = {
                'table_id': llm_table.get('table_id'),
                'table_title': llm_table.get('table_title'),
                'is_financial': llm_table.get('is_financial'),
                'currency': llm_table.get('currency'),
                'reporting_period': llm_table.get('reporting_period'),
                'similarity_score': match['similarity_score'],
                'llm_hierarchy': {
                    'horizontal': llm_table.get('horizontal_hierarchy_fields', []),
                    'vertical': llm_table.get('vertical_hierarchy_fields', [])
                },
                'ocr_data': {
                    'body_cells': ocr_table.get('body', []),
                    'header': ocr_table.get('header', []),
                    'footer': ocr_table.get('footer', [])
                },
                'location': llm_table.get('location', {})
            }

            merged_tables.append(merged_table)

        print(f"✅ 数据合并完成: {len(merged_tables)} 个表格")
        return merged_tables

    def save_alignment_results(self, merged_tables: List[Dict], output_path: str):
        """
        保存对齐结果到JSON文件

        Args:
            merged_tables: 合并后的表格数据
            output_path: 输出文件路径
        """
        results = {
            'alignment_summary': {
                'total_aligned_tables': len(merged_tables),
                'unmatched_llm_tables': len(self.unmatched_llm_tables),
                'unmatched_ocr_tables': len(self.unmatched_ocr_tables),
                'average_similarity': sum(t['similarity_score'] for t in merged_tables) / len(
                    merged_tables) if merged_tables else 0
            },
            'aligned_tables': merged_tables,
            'unmatched_tables': {
                'llm': self.unmatched_llm_tables,
                'ocr': self.unmatched_ocr_tables
            }
        }

        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(results, f, ensure_ascii=False, indent=2)

        print(f"💾 对齐结果已保存: {output_path}")


    def _create_summary_sheet(self, wb: openpyxl.Workbook, merged_tables: List[Dict]):
        """
        创建汇总表

        Args:
            wb: 工作簿对象
            merged_tables: 合并后的表格数据
        """
        ws = wb.create_sheet("汇总表", 0)

        # 设置表头
        headers = [
            "表格ID", "表格标题", "是否财务报表", "货币单位",
            "报告期间", "匹配相似度", "水平层级数", "垂直层级数",
            "OCR单元格数", "表头单元格数"
        ]

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # 填充数据
        for row, table in enumerate(merged_tables, 2):
            ws.cell(row=row, column=1, value=table.get('table_id', ''))
            ws.cell(row=row, column=2, value=table.get('table_title', ''))
            ws.cell(row=row, column=3, value='是' if table.get('is_financial') else '否')
            ws.cell(row=row, column=4, value=table.get('currency', ''))
            ws.cell(row=row, column=5, value=table.get('reporting_period', ''))
            ws.cell(row=row, column=6, value=table.get('similarity_score', 0))

            # 层级数量
            horizontal_count = len(table.get('llm_hierarchy', {}).get('horizontal', []))
            vertical_count = len(table.get('llm_hierarchy', {}).get('vertical', []))
            ws.cell(row=row, column=7, value=horizontal_count)
            ws.cell(row=row, column=8, value=vertical_count)

            # OCR数据统计
            ocr_data = table.get('ocr_data', {})
            body_count = len(ocr_data.get('body_cells', []))
            header_count = len(ocr_data.get('header', []))
            ws.cell(row=row, column=9, value=body_count)
            ws.cell(row=row, column=10, value=header_count)

        # 设置列宽和样式
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

        # 设置表头样式
        for cell in ws[1]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = openpyxl.styles.Font(bold=True)

    def save_to_excel(self, merged_tables: List[Dict], excel_path: str):
        """
        将对齐后的数据保存到Excel文件
        基于test_table3的重建逻辑，用LLM表头替换OCR表头
        """
        print("📊 正在生成Excel文件...")

        try:
            wb = openpyxl.Workbook()

            # 创建汇总表
            self._create_summary_sheet(wb, merged_tables)

            # 为每个表格创建重建的工作表
            for i, table in enumerate(merged_tables):
                self._create_enhanced_table_sheet(wb, table, i)

            # 删除默认创建的空工作表
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])

            wb.save(excel_path)
            print(f"💾 Excel文件已保存: {excel_path}")

        except Exception as e:
            print(f"❌ Excel文件保存失败: {e}")
            raise

    def _create_enhanced_table_sheet(self, wb: openpyxl.Workbook, table: Dict, table_index: int):
        """
        创建增强的表格工作表：只包含表格数据，不包含元信息
        """
        sheet_name = f"表格_{table_index + 1}"
        if len(sheet_name) > 31:
            sheet_name = f"Table_{table_index + 1}"

        ws = wb.create_sheet(sheet_name)

        # 直接重建表格结构，跳过表格标题、ID、匹配相似度等元信息
        self._rebuild_table_with_llm_headers(ws, table, start_row=1)  # 从第1行开始

        # 设置列宽
        for col in range(1, 10):
            ws.column_dimensions[get_column_letter(col)].width = 20


    def _extract_ocr_headers(self, body_cells: List[Dict]) -> List[str]:
        """
        从OCR数据中提取原始表头（第0行的内容）
        """
        headers = []
        max_col = 0

        # 找到最大列号
        for cell in body_cells:
            max_col = max(max_col, cell.get('col', 0))

        # 初始化表头数组
        headers = [''] * (max_col + 1)

        # 提取第0行的内容作为表头
        for cell in body_cells:
            row = cell.get('row', 0)
            col = cell.get('col', 0)
            words = cell.get('words', '')

            if row == 0 and col <= max_col:
                headers[col] = words

        return headers

    def _extract_smart_llm_headers(self, table: Dict, ocr_col_count: int) -> List[str]:
        """
        智能提取LLM表头，避免重复，保持列数一致
        """
        horizontal_fields = table.get('llm_hierarchy', {}).get('horizontal', [])
        headers = []
        seen_headers = set()

        print(f"   原始水平字段数: {len(horizontal_fields)}, OCR列数: {ocr_col_count}")

        # 按字段路径排序
        sorted_fields = sorted(horizontal_fields, key=lambda x: x.get('field_path', ''))

        for field in sorted_fields:
            if len(headers) >= ocr_col_count:
                break

            field_path = field.get('field_path', '')
            if not field_path:
                continue

            # 智能处理字段路径
            header = self._process_field_path(field_path)

            # 去重检查
            if header and header not in seen_headers:
                seen_headers.add(header)
                headers.append(header)
                print(f"     添加表头: {header}")

        # 如果LLM表头不够，用默认表头补充
        while len(headers) < ocr_col_count:
            headers.append(f"列{len(headers) + 1}")

        return headers

    def _process_field_path(self, field_path: str) -> str:
        """
        智能处理字段路径，避免重复
        """
        if not field_path:
            return ""

        # 处理多层路径
        if ' > ' in field_path:
            parts = field_path.split(' > ')

            # 对于包含年份的字段，保留完整路径
            if any(year in field_path for year in ['2022', '2023', '2024']):
                return field_path
            else:
                # 其他字段取最后一级，但确保有意义
                final_part = parts[-1]
                if final_part in ['项目', '金额', '数值']:
                    # 如果最后一级是通用词，尝试使用上一级
                    if len(parts) > 1:
                        return parts[-2] + " > " + final_part
                return final_part
        else:
            return field_path




    def _extract_horizontal_headers(self, table: Dict, max_col: int) -> List[str]:
        """
        提取横向表头（列标题）
        """
        horizontal_fields = table.get('llm_hierarchy', {}).get('horizontal', [])
        headers = []
        seen_headers = set()

        print(f"   原始横向字段: {[f.get('field_path', '') for f in horizontal_fields]}")

        # 优先处理包含年份的字段
        year_fields = []
        other_fields = []

        for field in horizontal_fields:
            field_path = field.get('field_path', '')
            if not field_path:
                continue

            if any(year in field_path for year in ['2023', '2022', '2021', '2020']):
                year_fields.append(field_path)
            else:
                other_fields.append(field_path)

        # 时间相关字段在前
        for field_path in year_fields:
            if field_path not in seen_headers and len(headers) < max_col:
                seen_headers.add(field_path)
                headers.append(field_path)

        # 其他字段在后
        for field_path in other_fields:
            if field_path not in seen_headers and len(headers) < max_col:
                # 简化非时间字段
                if ' > ' in field_path:
                    parts = field_path.split(' > ')
                    simple_header = parts[-1]
                    if simple_header not in ['项目', '名称', '科目'] and simple_header not in seen_headers:
                        seen_headers.add(simple_header)
                        headers.append(simple_header)
                else:
                    if field_path not in ['项目', '名称', '科目'] and field_path not in seen_headers:
                        seen_headers.add(field_path)
                        headers.append(field_path)

        # 补充不足的列
        while len(headers) < max_col:
            headers.append(f"列{len(headers) + 1}")

        return headers[:max_col]

    def _extract_vertical_headers(self, table: Dict, max_row: int) -> List[str]:
        """
        提取纵向表头（行标题）
        """
        vertical_fields = table.get('llm_hierarchy', {}).get('vertical', [])
        headers = []

        print(f"   原始纵向字段: {[f.get('field_path', '') for f in vertical_fields]}")

        for field in vertical_fields:
            field_path = field.get('field_path', '')
            if not field_path:
                continue

            # 纵向表头直接使用字段路径
            if ' > ' in field_path:
                # 对于纵向层级，我们通常需要完整的路径信息
                headers.append(field_path)
            else:
                headers.append(field_path)

            # 如果已经达到最大行数，停止提取
            if len(headers) >= max_row:
                break

        # 如果纵向表头不够，用默认值补充
        while len(headers) < max_row:
            headers.append(f"项目{len(headers) + 1}")

        return headers[:max_row]

    def calculate_leaf_similarity(self, llm_leaves: List[str], ocr_leaves: List[str]) -> float:
        """
        修正版：分别计算横向和纵向叶子的相似度
        """
        if not llm_leaves or not ocr_leaves:
            return 0.0

        # 这里假设llm_leaves已经正确分离了横向和纵向
        # 在实际使用中，应该在调用处分别传入横向和纵向的叶子节点
        total_similarity = 0
        matched_pairs = 0

        for llm_leaf in llm_leaves:
            best_leaf_similarity = 0
            for ocr_leaf in ocr_leaves:
                similarity = self.calculate_similarity(llm_leaf, ocr_leaf)
                if similarity > best_leaf_similarity:
                    best_leaf_similarity = similarity

            if best_leaf_similarity > 0.6:
                total_similarity += best_leaf_similarity
                matched_pairs += 1

        if matched_pairs == 0:
            return 0.0

        avg_similarity = total_similarity / matched_pairs
        match_ratio = matched_pairs / len(llm_leaves)

        return avg_similarity * match_ratio

    def _extract_ocr_horizontal_headers(self, cell_map: Dict, max_col: int) -> List[str]:
        """
        提取OCR横向表头（第0行）
        """
        headers = []
        for col in range(max_col + 1):
            if (0, col) in cell_map:
                headers.append(cell_map[(0, col)])
            else:
                headers.append("")
        return headers

    def _extract_ocr_vertical_headers(self, cell_map: Dict, max_row: int) -> List[str]:
        """
        提取OCR纵向表头（第0列，从第1行开始）
        """
        headers = []
        for row in range(1, max_row + 1):  # 从第1行开始，跳过表头行
            if (row, 0) in cell_map:
                headers.append(cell_map[(row, 0)])
            else:
                headers.append("")
        return headers

    def _get_final_horizontal_headers(self, ocr_headers: List[str], llm_headers: List[str], max_col: int) -> List[str]:
        """
        获取最终横向表头
        """
        result = ocr_headers.copy()

        if not llm_headers:
            return result

        print(f"   OCR表头数量: {len(ocr_headers)}, LLM表头数量: {len(llm_headers)}")

        # 如果OCR表头主要是通用词，用LLM表头替换
        generic_headers = ['项目', '科目', '名称', '']
        ocr_generic_count = sum(1 for h in ocr_headers if h in generic_headers)

        if ocr_generic_count >= len(ocr_headers) - 1 and len(llm_headers) >= len(ocr_headers) - 1:
            # 保留第一列，其他列用LLM表头
            for col in range(1, len(ocr_headers)):
                if col - 1 < len(llm_headers):
                    result[col] = llm_headers[col - 1]
                    print(f"     ✅ 替换横向表头列{col}: '{ocr_headers[col]}' -> '{llm_headers[col - 1]}'")

        # 特殊处理：时间相关表头
        for col, ocr_header in enumerate(ocr_headers):
            if any(year in ocr_header for year in ['2023', '2022', '2021']):
                # 在LLM中寻找匹配的时间表头
                for llm_header in llm_headers:
                    if any(year in llm_header for year in ['2023', '2022', '2021']):
                        similarity = self.calculate_similarity(ocr_header, llm_header)
                        if similarity > 0.8:
                            result[col] = llm_header
                            print(f"     ✅ 替换时间表头: '{ocr_header}' -> '{llm_header}'")
                            break

        return result

    def _get_vertical_header_replacement(self, ocr_header: str, llm_headers: List[str], row_index: int) -> str:
        """
        获取纵向表头替换
        """
        if not ocr_header or not llm_headers:
            return ocr_header

        # 如果行索引在LLM表头范围内，直接使用
        if row_index < len(llm_headers):
            similarity = self.calculate_similarity(ocr_header, llm_headers[row_index])
            if similarity > 0.7:
                return llm_headers[row_index]

        # 否则在整个LLM表头中寻找最佳匹配
        best_match = ocr_header
        best_similarity = 0

        for llm_header in llm_headers:
            similarity = self.calculate_similarity(ocr_header, llm_header)
            if similarity > best_similarity and similarity > 0.7:
                best_similarity = similarity
                best_match = llm_header

        if best_match != ocr_header:
            print(f"     🔄 纵向表头匹配: '{ocr_header}' -> '{best_match}' (相似度: {best_similarity:.2f})")

        return best_match


    def match_tables_by_leaf_nodes(self, llm_tables: List[Dict], ocr_tables: List[Dict]) -> List[Dict]:
        """
        修复版表格匹配
        """
        print("🔄 开始表格匹配...")

        matches = []
        matched_ocr_indices = set()

        for llm_idx, llm_table in enumerate(llm_tables):
            print(f"🔍 匹配LLM表格 {llm_idx + 1}: {llm_table.get('table_title', 'Unknown')}")

            best_match = None
            best_score = 0
            best_ocr_idx = -1

            # 分别提取横向和纵向叶子节点
            llm_vertical_leaves = self.extract_leaf_nodes(llm_table.get('vertical_hierarchy_fields', []))
            llm_horizontal_leaves = self.extract_leaf_nodes(llm_table.get('horizontal_hierarchy_fields', []))

            print(f"   横向叶子: {llm_horizontal_leaves}")
            print(f"   纵向叶子: {llm_vertical_leaves[:3]}...")  # 只显示前3个

            for ocr_idx, ocr_table in enumerate(ocr_tables):
                if ocr_idx in matched_ocr_indices:
                    continue

                # 提取OCR表格的所有叶子节点
                ocr_leaves = self.extract_ocr_leaf_nodes(ocr_table)
                print(f"   OCR表格 {ocr_idx + 1} 叶子节点: {ocr_leaves[:5]}...")

                # 分别计算横向和纵向相似度
                horizontal_similarity = self._calculate_directional_similarity(llm_horizontal_leaves, ocr_leaves,
                                                                               "横向")
                vertical_similarity = self._calculate_directional_similarity(llm_vertical_leaves, ocr_leaves, "纵向")

                # 综合相似度
                overall_similarity = (horizontal_similarity + vertical_similarity) / 2

                if overall_similarity > best_score and overall_similarity > 0.3:
                    best_score = overall_similarity
                    best_match = ocr_table
                    best_ocr_idx = ocr_idx

            if best_match:
                matches.append({
                    'llm_table': llm_table,
                    'ocr_table': best_match,
                    'similarity_score': best_score,
                    'llm_index': llm_idx,
                    'ocr_index': best_ocr_idx
                })
                matched_ocr_indices.add(best_ocr_idx)
                print(f"   ✅ 匹配成功! 相似度: {best_score:.2f}")
            else:
                self.unmatched_llm_tables.append(llm_table)
                print(f"   ❌ 未找到匹配")

        # 记录未匹配的OCR表格
        for ocr_idx, ocr_table in enumerate(ocr_tables):
            if ocr_idx not in matched_ocr_indices:
                self.unmatched_ocr_tables.append(ocr_table)

        print(f"✅ 匹配完成: {len(matches)} 个成功匹配")
        print(f"⚠️  未匹配LLM表格: {len(self.unmatched_llm_tables)}")
        print(f"⚠️  未匹配OCR表格: {len(self.unmatched_ocr_tables)}")

        return matches

    def _calculate_directional_similarity(self, directional_leaves: List[str], ocr_leaves: List[str],
                                          direction: str) -> float:
        """
        计算方向性相似度（横向或纵向）
        """
        if not directional_leaves:
            return 0.0

        total_similarity = 0
        matched_count = 0

        for llm_leaf in directional_leaves:
            best_match_score = 0
            for ocr_leaf in ocr_leaves:
                similarity = self.calculate_similarity(llm_leaf, ocr_leaf)
                if similarity > best_match_score:
                    best_match_score = similarity

            if best_match_score > 0.6:  # 匹配阈值
                total_similarity += best_match_score
                matched_count += 1

        if matched_count == 0:
            return 0.0

        avg_similarity = total_similarity / matched_count
        match_ratio = matched_count / len(directional_leaves)

        print(f"     {direction}匹配: {matched_count}/{len(directional_leaves)} 项, 平均相似度: {avg_similarity:.2f}")

        return avg_similarity * match_ratio


    def _smart_replace_vertical_header(self, ocr_header: str, llm_headers: List[str], row_index: int) -> str:
        """
        智能替换纵向表头
        """
        if not ocr_header or not llm_headers:
            return ocr_header

        # 清理OCR表头（去除换行符等）
        cleaned_ocr_header = ocr_header.replace('\n', '')

        # 策略1: 按位置匹配
        if row_index < len(llm_headers):
            llm_header = llm_headers[row_index]
            similarity = self.calculate_similarity(cleaned_ocr_header, llm_header)
            if similarity > 0.8:
                if cleaned_ocr_header != llm_header:
                    print(f"     🔄 按位置替换纵向表头行{row_index + 1}: '{cleaned_ocr_header}' -> '{llm_header}'")
                return llm_header

        # 策略2: 全局最佳匹配
        best_match = cleaned_ocr_header
        best_similarity = 0

        for llm_header in llm_headers:
            similarity = self.calculate_similarity(cleaned_ocr_header, llm_header)
            if similarity > best_similarity and similarity > 0.7:
                best_similarity = similarity
                best_match = llm_header

        if best_match != cleaned_ocr_header:
            print(
                f"     🔄 全局匹配纵向表头行{row_index + 1}: '{cleaned_ocr_header}' -> '{best_match}' (相似度: {best_similarity:.2f})")

        return best_match

    def align_data(self, llm_path: str, ocr_path: str, output_path: str = 'aligned_results.json',
                   excel_path: str = None):
        """
        主对齐函数 - 添加调试信息
        """
        print("🚀 开始数据对齐流程...")

        # 1. 加载数据
        llm_data, ocr_data = self.load_data(llm_path, ocr_path)

        # 2. 提取表格数据（假设第一张图片）
        llm_tables = llm_data.get('image_results', [{}])[0].get('tables', [])
        ocr_tables = ocr_data.get('tables_result', [])

        print(f"📊 待匹配表格: LLM={len(llm_tables)}, OCR={len(ocr_tables)}")

        # 调试：显示LLM表格的完整结构
        print("\n🔍 LLM表格结构调试:")
        for i, llm_table in enumerate(llm_tables):
            print(f"表格 {i + 1}: {llm_table.get('table_title', 'Unknown')}")
            horizontal_fields = llm_table.get('horizontal_hierarchy_fields', [])
            vertical_fields = llm_table.get('vertical_hierarchy_fields', [])

            print(f"  水平字段:")
            for field in horizontal_fields:
                print(f"    - {field}")

            print(f"  垂直字段:")
            for field in vertical_fields:
                print(f"    - {field}")
            print()

        # 3. 基于叶子节点匹配表格
        matches = self.match_tables_by_leaf_nodes(llm_tables, ocr_tables)

        # 4. 合并对齐数据
        merged_tables = self.merge_aligned_data(matches)

        # 5. 保存JSON结果
        self.save_alignment_results(merged_tables, output_path)

        # 6. 保存Excel结果
        if excel_path:
            self.save_to_excel(merged_tables, excel_path)
        else:
            # 如果没有指定Excel路径，使用JSON路径生成
            excel_path = output_path.replace('.json', '.xlsx')
            self.save_to_excel(merged_tables, excel_path)

        print("🎉 数据对齐流程完成!")

        return merged_tables



    # --------------------------
    def _smart_replace_horizontal_headers(self, ocr_headers: List[str], llm_headers: List[str], max_col: int) -> List[
        str]:
        """
        智能替换横向表头 - 修复版：对于复杂表格，完全使用LLM表头
        """
        result = ocr_headers.copy()

        if not llm_headers:
            return result

        print(f"   智能替换横向表头: OCR={ocr_headers}, LLM={llm_headers}")

        # 对于复杂表格（如递延所得税负债），完全使用LLM表头结构，忽略OCR表头
        if len(llm_headers) >= 3 and any("应纳税暂时性差异" in h for h in llm_headers):
            print("   🔧 检测到复杂表格，完全使用LLM表头结构")
            # 完全使用LLM的表头结构，覆盖所有列
            for col in range(len(ocr_headers)):
                if col < len(llm_headers):
                    if result[col] != llm_headers[col]:
                        print(f"     🔄 使用LLM结构替换列{col}: '{result[col]}' -> '{llm_headers[col]}'")
                        result[col] = llm_headers[col]
                else:
                    # 如果LLM表头不够，清空多余的列
                    result[col] = ""
            return result

        # 策略1: 如果OCR表头主要是通用词，完全用LLM表头替换
        generic_words = ['项目', '科目', '名称', '']
        ocr_generic_count = sum(1 for h in ocr_headers if h in generic_words)

        if ocr_generic_count >= len(ocr_headers) - 1:
            # 完全用LLM表头替换，但确保数量匹配
            for col in range(len(ocr_headers)):
                if col < len(llm_headers):
                    if result[col] != llm_headers[col]:
                        print(f"     🔄 完全替换横向表头列{col}: '{result[col]}' -> '{llm_headers[col]}'")
                        result[col] = llm_headers[col]
                elif col == 0 and "项目" not in result[col]:
                    # 确保第一列是"项目"
                    result[col] = "项目"

        # 策略2: 时间相关表头的智能替换
        for col, ocr_header in enumerate(ocr_headers):
            if any(year in ocr_header for year in ['2023', '2022', '2021']):
                # 在LLM中寻找更精确的时间表述
                for llm_header in llm_headers:
                    if any(year in llm_header for year in ['2023', '2022', '2021']):
                        # 如果LLM的表头更精确，就替换
                        if len(llm_header) > len(ocr_header) or "年" in llm_header:
                            if result[col] != llm_header:
                                print(f"     🔄 优化时间表头列{col}: '{result[col]}' -> '{llm_header}'")
                                result[col] = llm_header
                            break

        return result

    def _rebuild_table_with_llm_headers(self, ws, table: Dict, start_row: int):
        """
        修复版：真正用LLM表头替换OCR表头 - 优化复杂表格处理
        """
        ocr_data = table.get('ocr_data', {})
        body_cells = ocr_data.get('body_cells', [])

        if not body_cells:
            return

        # 1. 构建OCR表格结构
        cell_map = {}  # (row, col) -> words
        max_row = 0
        max_col = 0

        for cell in body_cells:
            row = cell.get('row_start', cell.get('row', 0))
            col = cell.get('col_start', cell.get('col', 0))
            words = cell.get('words', '').strip()

            max_row = max(max_row, row)
            max_col = max(max_col, col)
            cell_map[(row, col)] = words

        print(f"📊 OCR表格结构: {max_row + 1}行 x {max_col + 1}列")

        # 2. 分析LLM表头结构
        llm_horizontal = self.extract_leaf_nodes(table.get('llm_hierarchy', {}).get('horizontal', []))
        llm_vertical = self.extract_leaf_nodes(table.get('llm_hierarchy', {}).get('vertical', []))

        print(f"🔍 LLM横向字段: {llm_horizontal}")
        print(f"🔍 LLM纵向字段: {llm_vertical}")

        current_row = start_row

        # 3. 处理横向表头 - 对于复杂表格，只写入一行LLM表头
        print(f"   处理横向表头，共{max_col + 1}列")

        # 提取OCR横向表头（只取第0行）
        ocr_horizontal_headers = []
        for col in range(max_col + 1):
            if (0, col) in cell_map:
                ocr_horizontal_headers.append(cell_map[(0, col)])
            else:
                ocr_horizontal_headers.append("")

        print(f"   OCR横向表头: {ocr_horizontal_headers}")

        # 智能替换横向表头
        final_horizontal_headers = self._smart_replace_horizontal_headers(ocr_horizontal_headers, llm_horizontal,
                                                                          max_col)

        # 写入横向表头（只写一行）
        for col, header in enumerate(final_horizontal_headers):
            ws.cell(row=current_row, column=col + 1, value=header)

        current_row += 1

        # 4. 处理数据行 - 跳过OCR的多余表头行
        print(f"   处理数据行，共{max_row}行数据")

        # 确定数据起始行（跳过OCR的多余表头行）
        data_start_row = 1  # 从第1行开始（跳过第0行表头）

        # 对于复杂表格，可能需要跳过更多表头行
        if len(llm_horizontal) >= 3 and any("应纳税暂时性差异" in h for h in llm_horizontal):
            data_start_row = 2  # 跳过两行表头

        for row in range(data_start_row, max_row + 1):
            # 处理纵向表头
            if (row, 0) in cell_map:
                ocr_vertical_header = cell_map[(row, 0)]
                final_vertical_header = self._smart_replace_vertical_header(ocr_vertical_header, llm_vertical,
                                                                            row - data_start_row)
                ws.cell(row=current_row, column=1, value=final_vertical_header)

            # 写入数据单元格
            for col in range(1, max_col + 1):
                if (row, col) in cell_map:
                    ws.cell(row=current_row, column=col + 1, value=cell_map[(row, col)])

            current_row += 1

        # 5. 设置对齐方式
        for row in range(start_row, current_row):
            for col in range(1, max_col + 2):
                cell_obj = ws.cell(row=row, column=col)
                if col == 1:  # 第一列左对齐
                    cell_obj.alignment = Alignment(horizontal='left', vertical='center')
                else:  # 数据列居中对齐
                    cell_obj.alignment = Alignment(horizontal='center', vertical='center')

    def extract_leaf_nodes(self, hierarchy_fields: List[Dict]) -> List[str]:
        """
        提取层级字段中的叶子节点（修复版）
        """
        leaf_nodes = []

        for field in hierarchy_fields:
            field_path = field.get('field_path', '')
            field_type = field.get('field_type', '')
            field_value = field.get('field_value', '')

            # 如果有层级路径，使用完整路径
            if field_path:
                leaf_nodes.append(field_path)
            # 否则使用字段值
            elif field_value:
                leaf_nodes.append(field_value)
            # 最后使用字段类型
            elif field_type:
                leaf_nodes.append(field_type)

            print(f"     字段详情: path='{field_path}', type='{field_type}', value='{field_value}'")

        return leaf_nodes


# 使用示例
if __name__ == '__main__':
    aligner = TableDataAligner()

    analysis_results_path = r"E:\Datas\base_pros\DocuVista\test_codes\codes/analysis_results.json"
    baidu_path = r"E:\Datas\base_pros\DocuVista\test_codes\data1.json"
    tabl_merge_path = r"E:\Datas\base_pros\DocuVista\test_codes\table_alignment_results.json"
    excel_output_path = r"E:\Datas\base_pros\DocuVista\test_codes\table_alignment_results.xlsx"

    # 执行数据对齐
    aligned_data = aligner.align_data(
        llm_path=analysis_results_path,
        ocr_path=baidu_path,
        output_path=tabl_merge_path,
        excel_path=excel_output_path
    )

    print(f"\n📈 对齐统计:")
    print(f"   成功对齐: {len(aligned_data)} 个表格")
    if aligned_data:
        avg_score = sum(t['similarity_score'] for t in aligned_data) / len(aligned_data)
        print(f"   平均相似度: {avg_score:.2f}")