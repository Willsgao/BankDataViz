# -*- coding:utf-8 -*-

import json
import re
from typing import Dict, List, Any, Tuple, Optional
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
import os


class TableDataAligner:
    """
    表格数据对齐器 - 基于图片ID对齐LLM和百度OCR的表格数据
    """

    def __init__(self):
        self.matched_tables = []
        self.unmatched_llm_tables = []
        self.unmatched_ocr_tables = []


    def normalize_text(self, text: str) -> str:
        """
        文本标准化处理
        """
        if not text:
            return ""

        # 移除特殊字符、空格，转为小写
        text = re.sub(r'[^\w\u4e00-\u9fff]', '', text)
        return text.lower()

    def calculate_similarity(self, text1: str, text2: str) -> float:
        """
        计算文本相似度
        """
        text1_norm = self.normalize_text(text1)
        text2_norm = self.normalize_text(text2)

        if not text1_norm and not text2_norm:
            return 0.0

        if text1_norm == text2_norm:
            return 1.0

        # 简单的包含关系检查
        if text1_norm in text2_norm or text2_norm in text1_norm:
            return 0.8

        # 基于编辑距离的相似度
        distance = self.levenshtein_distance(text1_norm, text2_norm)
        max_len = max(len(text1_norm), len(text2_norm))

        if max_len == 0:
            return 0.0

        return 1 - (distance / max_len)

    def levenshtein_distance(self, s1: str, s2: str) -> int:
        """
        计算Levenshtein编辑距离
        """
        if len(s1) < len(s2):
            return self.levenshtein_distance(s2, s1)

        if len(s2) == 0:
            return len(s1)

        previous_row = range(len(s2) + 1)
        for i, c1 in enumerate(s1):
            current_row = [i + 1]
            for j, c2 in enumerate(s2):
                insertions = previous_row[j + 1] + 1
                deletions = current_row[j] + 1
                substitutions = previous_row[j] + (c1 != c2)
                current_row.append(min(insertions, deletions, substitutions))
            previous_row = current_row

        return previous_row[-1]

    def extract_leaf_nodes(self, hierarchy_fields: List[Dict]) -> List[str]:
        """
        提取层级字段中的叶子节点
        """
        leaf_nodes = []

        for field in hierarchy_fields:
            field_path = field.get('field_path', '')
            field_type = field.get('field_type', '')
            field_value = field.get('field_value', '')

            # 如果有层级路径，使用完整路径
            if field_path:
                leaf_nodes.append(field_path)
            # 否则使用字段值
            elif field_value:
                leaf_nodes.append(field_value)
            # 最后使用字段类型
            elif field_type:
                leaf_nodes.append(field_type)

        return leaf_nodes

    def match_tables_by_leaf_nodes(self, llm_tables: List[Dict], ocr_tables: List[Dict]) -> List[Dict]:
        """
        基于叶子节点匹配表格
        """
        print("🔄 开始表格匹配...")

        matches = []
        matched_ocr_indices = set()

        for llm_idx, llm_table in enumerate(llm_tables):
            print(f"🔍 匹配LLM表格 {llm_idx + 1}: {llm_table.get('table_title', 'Unknown')}")

            best_match = None
            best_score = 0
            best_ocr_idx = -1

            # 分别提取横向和纵向叶子节点
            llm_vertical_leaves = self.extract_leaf_nodes(llm_table.get('vertical_hierarchy_fields', []))
            llm_horizontal_leaves = self.extract_leaf_nodes(llm_table.get('horizontal_hierarchy_fields', []))

            print(f"   横向叶子: {llm_horizontal_leaves}")
            print(f"   纵向叶子: {llm_vertical_leaves[:3]}...")  # 只显示前3个

            for ocr_idx, ocr_table in enumerate(ocr_tables):
                if ocr_idx in matched_ocr_indices:
                    continue

                # 提取OCR表格的所有叶子节点
                ocr_leaves = self.extract_ocr_leaf_nodes(ocr_table)
                print(f"   OCR表格 {ocr_idx + 1} 叶子节点: {ocr_leaves[:5]}...")

                # 分别计算横向和纵向相似度
                horizontal_similarity = self._calculate_directional_similarity(llm_horizontal_leaves, ocr_leaves,
                                                                               "横向")
                vertical_similarity = self._calculate_directional_similarity(llm_vertical_leaves, ocr_leaves, "纵向")

                # 综合相似度
                overall_similarity = (horizontal_similarity + vertical_similarity) / 2

                if overall_similarity > best_score and overall_similarity > 0.3:
                    best_score = overall_similarity
                    best_match = ocr_table
                    best_ocr_idx = ocr_idx

            if best_match:
                matches.append({
                    'llm_table': llm_table,
                    'ocr_table': best_match,
                    'similarity_score': best_score,
                    'llm_index': llm_idx,
                    'ocr_index': best_ocr_idx
                })
                matched_ocr_indices.add(best_ocr_idx)
                print(f"   ✅ 匹配成功! 相似度: {best_score:.2f}")
            else:
                self.unmatched_llm_tables.append(llm_table)
                print(f"   ❌ 未找到匹配")

        # 记录未匹配的OCR表格
        for ocr_idx, ocr_table in enumerate(ocr_tables):
            if ocr_idx not in matched_ocr_indices:
                self.unmatched_ocr_tables.append(ocr_table)

        print(f"✅ 匹配完成: {len(matches)} 个成功匹配")
        print(f"⚠️  未匹配LLM表格: {len(self.unmatched_llm_tables)}")
        print(f"⚠️  未匹配OCR表格: {len(self.unmatched_ocr_tables)}")

        return matches

    def _calculate_directional_similarity(self, directional_leaves: List[str], ocr_leaves: List[str],
                                          direction: str) -> float:
        """
        计算方向性相似度（横向或纵向）
        """
        if not directional_leaves:
            return 0.0

        total_similarity = 0
        matched_count = 0

        for llm_leaf in directional_leaves:
            best_match_score = 0
            for ocr_leaf in ocr_leaves:
                similarity = self.calculate_similarity(llm_leaf, ocr_leaf)
                if similarity > best_match_score:
                    best_match_score = similarity

            if best_match_score > 0.6:  # 匹配阈值
                total_similarity += best_match_score
                matched_count += 1

        if matched_count == 0:
            return 0.0

        avg_similarity = total_similarity / matched_count
        match_ratio = matched_count / len(directional_leaves)

        print(f"     {direction}匹配: {matched_count}/{len(directional_leaves)} 项, 平均相似度: {avg_similarity:.2f}")

        return avg_similarity * match_ratio

    def merge_aligned_data(self, matches: List[Dict]) -> List[Dict]:
        """
        合并对齐后的数据
        """
        print("🔄 合并对齐数据...")

        merged_tables = []

        for match in matches:
            llm_table = match['llm_table']
            ocr_table = match['ocr_table']

            merged_table = {
                'table_id': llm_table.get('table_id'),
                'table_title': llm_table.get('table_title'),
                'is_financial': llm_table.get('is_financial'),
                'currency': llm_table.get('currency'),
                'reporting_period': llm_table.get('reporting_period'),
                'similarity_score': match['similarity_score'],
                'image_id': match.get('image_id'),
                'llm_image_path': match.get('llm_image_path'),
                'ocr_image_path': match.get('ocr_image_path'),
                'llm_hierarchy': {
                    'horizontal': llm_table.get('horizontal_hierarchy_fields', []),
                    'vertical': llm_table.get('vertical_hierarchy_fields', [])
                },
                'ocr_data': {
                    'body_cells': ocr_table.get('body', []),
                    'header': ocr_table.get('header', []),
                    'footer': ocr_table.get('footer', [])
                },
                'location': llm_table.get('location', {})
            }

            merged_tables.append(merged_table)

        print(f"✅ 数据合并完成: {len(merged_tables)} 个表格")
        return merged_tables

    def save_alignment_results(self, merged_tables: List[Dict], output_path: str):
        """
        保存对齐结果到JSON文件
        """
        results = {
            'alignment_summary': {
                'total_aligned_tables': len(merged_tables),
                'unmatched_llm_tables': len(self.unmatched_llm_tables),
                'unmatched_ocr_tables': len(self.unmatched_ocr_tables),
                'average_similarity': sum(t['similarity_score'] for t in merged_tables) / len(
                    merged_tables) if merged_tables else 0
            },
            'aligned_tables': merged_tables,
            'unmatched_tables': {
                'llm': self.unmatched_llm_tables,
                'ocr': self.unmatched_ocr_tables
            }
        }

        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(results, f, ensure_ascii=False, indent=2)

        print(f"💾 对齐结果已保存: {output_path}")

    def save_to_excel(self, merged_tables: List[Dict], excel_path: str, image_name:str = None):
        """
        将对齐后的数据保存到Excel文件
        """
        print("📊 正在生成Excel文件...")
        try:
            wb = openpyxl.Workbook()

            # 如果没有匹配到表格，创建一个空的工作表
            if not merged_tables:
                ws = wb.active
                ws.title = "无匹配表格"
                ws.cell(row=1, column=1, value="提示")
                ws.cell(row=1, column=2, value="未找到匹配的表格")
                print("⚠️  没有匹配的表格，创建空Excel文件")
            else:
                # 删除默认创建的空工作表
                if 'Sheet' in wb.sheetnames:
                    wb.remove(wb['Sheet'])

                # 为每个表格创建简单的工作表
                for i, table in enumerate(merged_tables):
                    self._create_simple_table_sheet(wb, table, i, image_name)

            wb.save(excel_path)
            print(f"💾 Excel文件已保存: {excel_path}")

        except Exception as e:
            print(f"❌ Excel文件保存失败: {e}")
            raise

    def _create_simple_table_sheet(self, wb: openpyxl.Workbook, table: Dict, table_index: int, image_name: str = None):
        """
        创建简单的表格工作表 - 修复OCR数据提取问题
        """
        import os
        from pathlib import Path

        # 使用图片名称作为工作表名称
        image_num = '0'
        if image_name:
            image_last_name = os.path.basename(image_name)
            image_num = image_last_name.split('.')[0].split('_')[-1]

        sheet_name = f"P{image_num}_表格_{table_index + 1}"
        if len(sheet_name) > 31:
            sheet_name = f"P{image_num}_Table_{table_index + 1}"

        ws = wb.create_sheet(sheet_name)

        ocr_data = table.get('ocr_data', {})
        body_cells = ocr_data.get('body_cells', [])
        header_cells = ocr_data.get('header', [])

        if not body_cells:
            ws.cell(row=1, column=1, value="无表格数据")
            return

        # 1. 构建OCR表格结构 - 修复：同时处理body和header
        cell_map = {}
        max_row = 0
        max_col = 0

        # 处理body cells
        for cell in body_cells:
            row = cell.get('row_start', cell.get('row', 0))
            col = cell.get('col_start', cell.get('col', 0))
            words = cell.get('words', '').strip()
            cell_map[(row, col)] = words
            max_row = max(max_row, row)
            max_col = max(max_col, col)

        # 处理header cells（如果有）
        for cell in header_cells:
            row = cell.get('row_start', cell.get('row', 0))
            col = cell.get('col_start', cell.get('col', 0))
            words = cell.get('words', '').strip()
            cell_map[(row, col)] = words
            max_row = max(max_row, row)
            max_col = max(max_col, col)

        print(f"📊 OCR表格结构: {max_row + 1}行 x {max_col + 1}列")
        print(f"   Body cells数量: {len(body_cells)}")
        print(f"   Header cells数量: {len(header_cells)}")

        # 2. 分析LLM表头结构
        llm_horizontal = self.extract_leaf_nodes(table.get('llm_hierarchy', {}).get('horizontal', []))
        llm_vertical = self.extract_leaf_nodes(table.get('llm_hierarchy', {}).get('vertical', []))

        print(f"🔍 LLM横向字段: {llm_horizontal}")
        print(f"🔍 LLM纵向字段: {llm_vertical}")

        current_row = 1

        # 3. 写入横向表头
        print(f"   写入横向表头，共{len(llm_horizontal)}列")
        for col, header in enumerate(llm_horizontal):
            if col <= max_col:
                ws.cell(row=current_row, column=col + 1, value=header)

        current_row += 1

        # 4. 处理数据行 - 修复：如果OCR数据不完整，使用LLM数据填充
        print(f"   处理数据行，OCR有{max_row + 1}行，LLM有{len(llm_vertical)}行纵向数据")

        # 打印OCR纵向文本用于调试
        print(f"   OCR纵向文本:")
        for row in range(1, max_row + 1):
            if (row, 0) in cell_map:
                ocr_text = cell_map[(row, 0)]
                print(f"     行{row}: '{ocr_text}'")
            else:
                print(f"     行{row}: [空]")

        # 修复策略：如果OCR数据不完整，使用LLM纵向数据
        if max_row < len(llm_vertical):
            print(f"   ⚠️  OCR数据不完整，使用LLM纵向数据填充")
            for i, llm_vertical_header in enumerate(llm_vertical):
                if i < max_row:  # 优先使用OCR数据
                    if (i + 1, 0) in cell_map:
                        ocr_text = cell_map[(i + 1, 0)]
                        final_header = self._smart_replace_vertical_header(ocr_text, llm_vertical, i)
                        ws.cell(row=current_row, column=1, value=final_header)
                    else:
                        ws.cell(row=current_row, column=1, value=llm_vertical_header)
                else:  # OCR数据不够，直接使用LLM数据
                    ws.cell(row=current_row, column=1, value=llm_vertical_header)

                # 写入数据列
                for col in range(1, max_col + 1):
                    if (i + 1, col) in cell_map:
                        value = cell_map[(i + 1, col)]
                        if col < len(llm_horizontal):
                            ws.cell(row=current_row, column=col + 1, value=value)
                        else:
                            ws.cell(row=current_row, column=col + 1, value=value)

                current_row += 1
        else:
            # 正常处理：OCR数据完整
            data_start_row = 1
            for row in range(data_start_row, max_row + 1):
                if (row, 0) in cell_map:
                    ocr_vertical_header = cell_map[(row, 0)]
                    final_vertical_header = self._smart_replace_vertical_header(
                        ocr_vertical_header,
                        llm_vertical,
                        row - data_start_row
                    )
                    ws.cell(row=current_row, column=1, value=final_vertical_header)

                for col in range(1, max_col + 1):
                    if (row, col) in cell_map:
                        value = cell_map[(row, col)]
                        if col < len(llm_horizontal):
                            ws.cell(row=current_row, column=col + 1, value=value)
                        else:
                            ws.cell(row=current_row, column=col + 1, value=value)

                current_row += 1

        # 5. 设置列宽和样式
        for col in range(1, max_col + 2):
            ws.column_dimensions[get_column_letter(col)].width = 15

        for row in range(1, current_row):
            for col in range(1, max_col + 2):
                cell_obj = ws.cell(row=row, column=col)
                if col == 1:
                    cell_obj.alignment = Alignment(horizontal='left', vertical='center')
                else:
                    cell_obj.alignment = Alignment(horizontal='center', vertical='center')

        print(f"✅ 表格 {table_index + 1} 处理完成，共{current_row - 1}行")


    def _smart_replace_vertical_header(self, ocr_header: str, llm_headers: List[str], row_index: int) -> str:
        """
        智能替换纵向表头 - 添加详细调试信息
        """
        if not ocr_header or not llm_headers:
            print(f"     ⚠️  无法替换: OCR='{ocr_header}', LLM数量={len(llm_headers)}")
            return ocr_header

        cleaned_ocr_header = ocr_header.replace('\n', '').strip()
        print(f"     🔍 开始匹配: OCR='{cleaned_ocr_header}', 行索引={row_index}, LLM选项={len(llm_headers)}")

        # 策略1: 按位置匹配
        if row_index < len(llm_headers):
            llm_header = llm_headers[row_index]
            similarity = self.calculate_similarity(cleaned_ocr_header, llm_header)
            print(
                f"     📍 位置匹配: OCR='{cleaned_ocr_header}' vs LLM[{row_index}]='{llm_header}', 相似度={similarity:.2f}")

            if similarity > 0.8:
                if cleaned_ocr_header != llm_header:
                    print(f"     ✅ 按位置替换: '{cleaned_ocr_header}' -> '{llm_header}'")
                return llm_header

        # 策略2: 全局最佳匹配
        best_match = cleaned_ocr_header
        best_similarity = 0
        best_llm_index = -1

        for i, llm_header in enumerate(llm_headers):
            similarity = self.calculate_similarity(cleaned_ocr_header, llm_header)
            if similarity > best_similarity:
                best_similarity = similarity
                best_match = llm_header
                best_llm_index = i

        print(f"     🌐 全局匹配: 最佳相似度={best_similarity:.2f}, 最佳匹配='{best_match}' (索引{best_llm_index})")

        if best_match != cleaned_ocr_header and best_similarity > 0.7:
            print(f"     ✅ 全局替换: '{cleaned_ocr_header}' -> '{best_match}' (相似度: {best_similarity:.2f})")
            return best_match
        else:
            print(f"     ❌ 未达到替换阈值: 相似度={best_similarity:.2f}, 保持OCR文本")
            return cleaned_ocr_header



    def align_data(self, llm_path: str, ocr_path: str, output_path: str = 'aligned_results.json',
                   excel_path: str = None, image_name:str = None):
        """
        主对齐函数 - 基于图片ID进行匹配
        """
        print("🚀 开始基于图片ID的数据对齐流程...")

        # 1. 加载数据
        llm_data, ocr_data = self.load_data(llm_path, ocr_path)

        print(f"📊 数据统计:")
        print(f"   LLM图片数: {len(llm_data.get('image_results', []))}")
        print(f"   OCR图片数: {len(ocr_data.get('image_results', []))}")

        # 2. 基于图片ID匹配表格
        matches = self.match_tables_by_image_id(llm_data, ocr_data)

        # 3. 合并对齐数据
        merged_tables = self.merge_aligned_data(matches)

        # 添加调试信息
        print(f"🔍 调试信息:")
        print(f"   合并的表格数量: {len(merged_tables)}")
        for i, table in enumerate(merged_tables):
            ocr_data = table.get('ocr_data', {})
            body_cells = ocr_data.get('body_cells', [])
            print(f"   表格 {i + 1}: {len(body_cells)} 个body_cells")
            if body_cells:
                # 显示前几个单元格的信息
                for j, cell in enumerate(body_cells[:3]):
                    print(
                        f"     单元格 {j + 1}: row_start={cell.get('row_start')}, col_start={cell.get('col_start')}, words='{cell.get('words')}'")

        # 4. 保存JSON结果
        self.save_alignment_results(merged_tables, output_path)

        # 5. 保存Excel结果
        if excel_path:
            self.save_to_excel(merged_tables, excel_path, image_name)
        else:
            excel_path = output_path.replace('.json', '.xlsx')
            self.save_to_excel(merged_tables, excel_path, image_name)

        print("🎉 基于图片ID的数据对齐流程完成!")
        return merged_tables

    def match_tables_by_image_id(self, llm_data: Dict, ocr_data: Dict) -> List[Dict]:
        """
        基于图片ID匹配表格数据 - 增强版，支持多种OCR数据格式
        """
        print("🔄 基于图片ID进行表格匹配...")

        matches = []

        # 构建图片ID到OCR结果的映射（支持多种格式）
        ocr_image_map = {}

        # 格式1: 新格式（image_results数组）
        if "image_results" in ocr_data:
            for ocr_result in ocr_data["image_results"]:
                image_id = ocr_result.get("image_id")
                if image_id:
                    ocr_image_map[image_id] = ocr_result
                else:
                    # 如果没有image_id，尝试从image_info中获取
                    image_info = ocr_result.get("image_info", {})
                    image_id = image_info.get("image_id")
                    if image_id:
                        ocr_image_map[image_id] = ocr_result

        # 格式2: 旧格式（直接包含image_info）
        elif "image_info" in ocr_data:
            image_info = ocr_data["image_info"]
            image_id = image_info.get("image_id")
            if image_id:
                # 将整个ocr_data作为该图片的结果
                ocr_image_map[image_id] = {
                    "image_id": image_id,
                    "image_path": image_info.get("image_path"),
                    "tables_result": ocr_data.get("tables_result", [])
                }

        # 格式3: 最旧格式（只有tables_result）
        elif "tables_result" in ocr_data:
            # 为这种格式生成一个默认的image_id
            default_image_id = "default_ocr_image"
            ocr_image_map[default_image_id] = {
                "image_id": default_image_id,
                "image_path": "unknown",
                "tables_result": ocr_data.get("tables_result", [])
            }
            print(f"⚠️  OCR数据使用默认图片ID: {default_image_id}")

        print(f"🔍 OCR图片映射: {list(ocr_image_map.keys())}")

        # 其余代码保持不变...
        for llm_image_result in llm_data.get("image_results", []):
            llm_image_id = llm_image_result.get("image_id")
            llm_tables = llm_image_result.get("tables", [])

            if not llm_image_id:
                print(f"⚠️  LLM结果缺少图片ID: {llm_image_result.get('image_path')}")
                continue

            # 查找对应的OCR结果
            ocr_image_result = ocr_image_map.get(llm_image_id)
            if not ocr_image_result:
                print(f"⚠️  未找到图片ID {llm_image_id} 对应的OCR结果")
                print(f"   可用的OCR图片ID: {list(ocr_image_map.keys())}")
                continue

            ocr_tables = ocr_image_result.get("tables_result", [])

            print(f"🔍 匹配图片 {llm_image_id}: LLM表格={len(llm_tables)}, OCR表格={len(ocr_tables)}")

            # 对同一图片内的表格进行匹配
            image_matches = self.match_tables_by_leaf_nodes(llm_tables, ocr_tables)

            # 为匹配结果添加图片ID信息
            for match in image_matches:
                match["image_id"] = llm_image_id
                match["llm_image_path"] = llm_image_result.get("image_path")
                match["ocr_image_path"] = ocr_image_result.get("image_path")

            matches.extend(image_matches)

        print(f"✅ 基于图片ID匹配完成: 共匹配 {len(matches)} 个表格")
        return matches

    def load_data(self, llm_path: str, ocr_path: str) -> Tuple[Dict, Dict]:
        """
        加载LLM和OCR数据 - 增强调试信息
        """
        print("📥 加载数据...")

        try:
            with open(llm_path, 'r', encoding='utf-8') as f:
                llm_data = json.load(f)

            with open(ocr_path, 'r', encoding='utf-8') as f:
                ocr_data = json.load(f)

            # 添加详细的调试信息
            print("🔍 LLM数据结构:")
            if 'image_results' in llm_data:
                for i, img_result in enumerate(llm_data['image_results']):
                    print(f"  图片 {i + 1}: path={img_result.get('image_path')}, id={img_result.get('image_id')}")

            print("🔍 OCR数据结构:")
            if 'image_results' in ocr_data:
                for i, img_result in enumerate(ocr_data['image_results']):
                    print(f"  图片 {i + 1}: path={img_result.get('image_path')}, id={img_result.get('image_id')}")
            elif 'image_info' in ocr_data:
                print(
                    f"  旧格式: path={ocr_data['image_info'].get('image_path')}, id={ocr_data['image_info'].get('image_id')}")
            else:
                print("  OCR数据格式未知，完整结构:")
                print(json.dumps(ocr_data, ensure_ascii=False, indent=2)[:500] + "...")

            return llm_data, ocr_data
        except Exception as e:
            print(f"❌ 数据加载失败: {e}")
            raise

    def extract_ocr_leaf_nodes(self, ocr_table: Dict) -> List[str]:
        """
        从OCR表格中提取叶子节点文本 - 增强版
        """
        leaves = []

        # 从body单元格中提取文本
        for cell in ocr_table.get('body', []):
            words = cell.get('words', '').strip()
            if words and words not in leaves:  # 去重
                leaves.append(words)

        # 从header中提取文本
        for header in ocr_table.get('header', []):
            words = header.get('words', '').strip()
            if words and words not in leaves:
                leaves.append(words)

        # 从footer中提取文本（如果有）
        for footer in ocr_table.get('footer', []):
            words = footer.get('words', '').strip()
            if words and words not in leaves:
                leaves.append(words)

        return leaves



# 使用示例
if __name__ == '__main__':
    aligner = TableDataAligner()

    import os

    code_dir = os.getcwd()
    parent_dir = os.path.dirname(code_dir)

    analysis_results_path = fr"{parent_dir}\codes/analysis_results.json"
    baidu_path = fr"{parent_dir}\data3.json"
    tabl_merge_path = fr"{parent_dir}\table_alignment_results.json"
    excel_output_path = fr"{parent_dir}\table_alignment_results.xlsx"
    image_name = fr"{parent_dir}\pngs\514001_158.png"

    # 执行数据对齐
    aligned_data = aligner.align_data(
        llm_path=analysis_results_path,
        ocr_path=baidu_path,
        output_path=tabl_merge_path,
        excel_path=excel_output_path,
        image_name=image_name
    )

    print(f"\n📈 对齐统计:")
    print(f"   成功对齐: {len(aligned_data)} 个表格")
    if aligned_data:
        avg_score = sum(t['similarity_score'] for t in aligned_data) / len(aligned_data)
        print(f"   平均相似度: {avg_score:.2f}")