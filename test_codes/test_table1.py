import json, openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill


def baidu2excel_simple_with_hierarchy11(bd_json, out_file='baidu_table.xlsx'):
    wb = openpyxl.Workbook()

    for idx, tbl in enumerate(bd_json['tables_result']):
        print(f"处理表格 {idx}")
        ws = wb.create_sheet(title=f'Table{idx + 1}') if idx else wb.active
        ws.title = f'Table{idx + 1}'


        # 创建内容映射 - 只让代表单元格写一次，避免重复
        content_map = {}
        for cell in tbl['body']:
            r_s, c_s, r_e, c_e = cell['row_start'], cell['col_start'], cell['row_end'], cell['col_end']
            words = cell['words']

            # 最左列：只在代表单元格（左上角）加缩进，然后铺满合并区域
            if c_s == 0:
                if (r_s, c_s) in hierarchy_levels:
                    level = hierarchy_levels[(r_s, c_s)]
                    indent = "  " * level
                    words = indent + words
                # 整个合并区域写同一份带缩进的内容
                for row in range(r_s, r_e + 1):
                    for col in range(c_s, c_e + 1):
                        content_map[(row, col)] = words
            else:  # 非最左列直接写原值
                for row in range(r_s, r_e + 1):
                    for col in range(c_s, c_e + 1):
                        content_map[(row, col)] = words

        # 分析纵向表头的层次关系
        hierarchy_levels = analyze_vertical_hierarchy(tbl['body'])

        # # 创建内容映射 - 完全避免合并单元格问题
        # content_map = {}
        # for cell in tbl['body']:
        #     r_s, c_s, r_e, c_e = cell['row_start'], cell['col_start'], cell['row_end'], cell['col_end']
        #     words = cell['words']
        #
        #     # 在合并区域的所有位置都写入相同的内容
        #     for row in range(r_s, r_e + 1):
        #         for col in range(c_s, c_e + 1):
        #             # 对于纵向表头，根据层级添加缩进
        #             # if col == 0 and (row, col) in hierarchy_levels:  # 假设第一列是纵向表头
        #             #     level = hierarchy_levels[(row, col)]
        #             #     indent = "  " * level
        #             #     content_map[(row, col)] = indent + words
        #             if col == 0 and (row, col) in hierarchy_levels:
        #                 level = hierarchy_levels[(row, col)]
        #                 indent = "  " * level
        #                 content_map[(row, col)] = indent + words
        #             else:
        #                 content_map[(row, col)] = words

        # 写入所有内容
        print("写入单元格内容")
        for (row, col), content in content_map.items():
            excel_row = row + 1
            excel_col = col + 1
            try:
                cell = ws.cell(row=excel_row, column=excel_col, value=content)

                # 设置基本对齐方式
                if col == 0:  # 纵向表头左对齐
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                else:  # 数据列居中对齐
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            except Exception as e:
                print(f"写入失败: {e}")

        # 设置列宽
        set_column_widths(ws, tbl['body'])

    wb.save(out_file)
    print(f'已生成 {out_file}')


def baidu2excel_simple_with_hierarchy(bd_json, out_file='baidu_table.xlsx'):
    wb = openpyxl.Workbook()

    for idx, tbl in enumerate(bd_json['tables_result']):
        print(f"处理表格 {idx}")
        ws = wb.create_sheet(title=f'Table{idx + 1}') if idx else wb.active
        ws.title = f'Table{idx + 1}'

        hierarchy_levels = analyze_vertical_hierarchy(tbl['body'])

        # 创建内容映射 - 使用原始逻辑
        content_map = {}
        for cell in tbl['body']:
            r_s, c_s, r_e, c_e = cell['row_start'], cell['col_start'], cell['row_end'], cell['col_end']
            words = cell['words']

            # 最左列加缩进
            if c_s == 0 and (r_s, c_s) in hierarchy_levels:
                level = hierarchy_levels[(r_s, c_s)]
                words = "  " * level + words

            # 原始逻辑：为合并区域的所有位置写入内容
            for row in range(r_s, r_e + 1):
                for col in range(c_s, c_e + 1):
                    content_map[(row, col)] = words

        # 修复1：删除第一行的重复（第0行和第1行内容相同）
        rows_to_remove = []
        for (row, col), content in content_map.items():
            if row == 1:  # 检查第1行（Excel中的第2行）
                # 如果第0行和第1行在相同列有相同内容，标记第1行要删除
                if (0, col) in content_map and content_map[(0, col)] == content:
                    rows_to_remove.append((1, col))

        for pos in rows_to_remove:
            if pos in content_map:
                del content_map[pos]

        # 修复2：删除最后一列的重复
        max_col = max(col for row, col in content_map.keys())
        last_col_duplicates = []
        for (row, col), content in content_map.items():
            if col == max_col:
                # 检查前一列是否有相同内容
                if (row, max_col - 1) in content_map and content_map[(row, max_col - 1)] == content:
                    last_col_duplicates.append((row, max_col))

        for pos in last_col_duplicates:
            if pos in content_map:
                del content_map[pos]

        # 一次性写入 Excel
        for (row, col), content in content_map.items():
            cell = ws.cell(row=row + 1, column=col + 1, value=content)
            if col == 0:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        set_column_widths(ws, tbl['body'])

    wb.save(out_file)
    print(f'已生成 {out_file}')



def analyze_vertical_hierarchy(body_cells, thresh=20):
    """
    基于最左侧列的「左边界 x 坐标」计算缩进级别，
    返回 {(row, col): level_path}  例：'资产/流动资产/货币资金'
    """
    if not body_cells:
        return {}

    min_col = min(c['col_start'] for c in body_cells)
    # 只取最左列，并按行号排序
    # left_col = [c for c in body_cells if c['col_start'] == min_col and c['col_end'] == min_col]
    left_col = [c for c in body_cells if c['col_start'] == min_col]   # 放宽条件
    left_col.sort(key=lambda c: c['row_start'])

    lvl_map = {}
    base_x = min(c['cell_location'][0]['x'] for c in left_col)
    last_x, last_lvl = base_x, 0
    stack = []

    for c in left_col:
        cur_x = c['cell_location'][0]['x']
        # 右移 ≥ thresh 像素 → 加深一级
        if cur_x - last_x >= thresh:
            last_lvl += 1
            last_x = cur_x
        elif cur_x < last_x - thresh:   # 回退到外层
            last_lvl = max(last_lvl - 1, 0)
            last_x = cur_x
        # 维护栈
        stack = stack[:last_lvl]
        stack.append(c['words'].strip())
        # lvl_map[(c['row_start'], min_col)] = ' / '.join(stack)
        lvl_map[(c['row_start'], min_col)] = last_lvl

    return lvl_map


def set_column_widths(ws, body_cells):
    """设置列宽"""
    if not body_cells:
        return

    max_col = max(c['col_end'] for c in body_cells) + 1

    for col in range(max_col):
        if col == 0:  # 纵向表头列设置较宽
            ws.column_dimensions[get_column_letter(col + 1)].width = 30
        else:
            ws.column_dimensions[get_column_letter(col + 1)].width = 15


if __name__ == '__main__':
    try:
        with open('data1.json', 'r', encoding='utf-8') as f:
            j = json.load(f)

        save_file = "百度表格-简化层级版.xlsx"
        baidu2excel_simple_with_hierarchy(j, save_file)

    except Exception as e:
        print(f"程序执行出错: {e}")
        import traceback

        traceback.print_exc()