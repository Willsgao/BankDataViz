import json, openpyxl
from openpyxl.utils import get_column_letter

def baidu2excel(bd_json, out_file='baidu_table.xlsx'):
    wb = openpyxl.Workbook()
    # 如果 JSON 里有多张表，每张表新建一个工作表
    for idx, tbl in enumerate(bd_json['tables_result']):
        print("idx->:", idx)
        print("XXXXXXXXXXX", tbl)
        ws = wb.create_sheet(title=f'Table{idx+1}') if idx else wb.active
        ws.title = f'Table{idx+1}'

        # 计算表格总行/列
        max_row = max(c['row_end'] for c in tbl['body'])
        max_col = max(c['col_end'] for c in tbl['body'])

        # 写值
        for c in tbl['body']:
            print("cccccccccc", c)
            r, c_s, r_e, c_e = c['row_start']+1, c['col_start']+1, c['row_end'], c['col_end']
            ws.cell(r, c_s, c['words'])
            # 需要合并？
            if r_e - r > 0 or c_e - c_s > 0:
                ws.merge_cells(start_row=r, start_column=c_s,
                               end_row=r_e, end_column=c_e)

        # 简单美化：自适应列宽
        for col in range(1, max_col+1):
            ws.column_dimensions[get_column_letter(col)].width = 12

    import os
    main_dir = os.getcwd()
    print("main_dir:", main_dir)
    out_file = fr"{main_dir}/百度表格.xlsx"
    wb.save(out_file)
    print('已生成', out_file)

if __name__ == '__main__':
    with open('data.json', 'r', encoding='utf-8') as f:
        j = json.load(f)

    baidu2excel(j)