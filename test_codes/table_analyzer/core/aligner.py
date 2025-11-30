

import json
import re
import os
from typing import Dict, List, Any, Tuple, Optional
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from test_codes.table_analyzer.utils.image_utils import ImageUtils

class TableDataAligner:
    """
    重构后的表格数据对齐器 - 对齐LLM和百度OCR的表格数据
    """

    def __init__(self):
        self.image_utils = ImageUtils()
        self.matched_tables = []
        self.unmatched_llm_tables = []
        self.unmatched_ocr_tables = []

    def load_data(self, llm_path: str, ocr_path: str) -> Tuple[Dict, Dict]:
        """
        加载LLM和OCR数据 - 支持新格式
        """
        print("📥 加载数据...")

        try:
            with open(llm_path, 'r', encoding='utf-8') as f:
                llm_data = json.load(f)

            with open(ocr_path, 'r', encoding='utf-8') as f:
                ocr_data = json.load(f)

            # 统计表格数量
            llm_table_count = sum(len(img.get('tables', [])) for img in llm_data.get('image_results', []))

            # 检查OCR数据格式
            if 'tables_result' in ocr_data:
                ocr_table_count = len(ocr_data.get('tables_result', []))
            else:
                ocr_table_count = 0

            print(f"✅ LLM数据: {len(llm_data.get('image_results', []))} 张图片, {llm_table_count} 个表格")
            print(f"✅ OCR数据: {ocr_table_count} 个表格")

            # 检查是否有image_id信息
            if 'image_info' in ocr_data:
                image_id = ocr_data['image_info'].get('image_id')
                image_path = ocr_data['image_info'].get('image_path')
                print(f"📷 OCR图片ID: {image_id}")
                print(f"📷 OCR图片路径: {image_path}")

            return llm_data, ocr_data
        except Exception as e:
            print(f"❌ 数据加载失败: {e}")
            raise

    def normalize_text(self, text: str) -> str:
        """
        文本标准化处理
        """
        if not text:
            return ""

        # 移除特殊字符、空格，转为小写
        text = re.sub(r'[^\w\u4e00-\u9fff]', '', text)
        return text.lower()

    def calculate_similarity(self, text1: str, text2: str) -> float:
        """
        计算文本相似度（基于编辑距离的简单实现）
        """
        text1_norm = self.normalize_text(text1)
        text2_norm = self.normalize_text(text2)

        if not text1_norm and not text2_norm:
            return 0.0

        if text1_norm == text2_norm:
            return 1.0

        # 简单的包含关系检查
        if text1_norm in text2_norm or text2_norm in text1_norm:
            return 0.8

        # 基于编辑距离的相似度
        distance = self.levenshtein_distance(text1_norm, text2_norm)
        max_len = max(len(text1_norm), len(text2_norm))

        if max_len == 0:
            return 0.0

        return 1 - (distance / max_len)

    def levenshtein_distance(self, s1: str, s2: str) -> int:
        """
        计算Levenshtein编辑距离
        """
        if len(s1) < len(s2):
            return self.levenshtein_distance(s2, s1)

        if len(s2) == 0:
            return len(s1)

        previous_row = range(len(s2) + 1)
        for i, c1 in enumerate(s1):
            current_row = [i + 1]
            for j, c2 in enumerate(s2):
                insertions = previous_row[j + 1] + 1
                deletions = current_row[j] + 1
                substitutions = previous_row[j] + (c1 != c2)
                current_row.append(min(insertions, deletions, substitutions))
            previous_row = current_row

        return previous_row[-1]

    def extract_ocr_leaf_nodes(self, ocr_table: Dict) -> List[str]:
        """
        从OCR表格中提取叶子节点文本
        """
        leaves = []

        # 从body单元格中提取文本
        for cell in ocr_table.get('body', []):
            words = cell.get('words', '').strip()
            if words and words not in leaves:  # 去重
                leaves.append(words)

        # 从header中提取文本
        for header in ocr_table.get('header', []):
            words = header.get('words', '').strip()
            if words and words not in leaves:
                leaves.append(words)

        return leaves

    def extract_leaf_nodes(self, hierarchy_fields: List[Dict]) -> List[Dict]:
        """
        提取层级字段中的叶子节点 - 返回字典包含完整路径和叶子文本
        """
        leaf_nodes = []

        for field in hierarchy_fields:
            field_path = field.get('field_path', '')
            is_statistical = field.get('is_statistical', False)

            # 如果有层级路径，分离完整路径和叶子文本
            if field_path:
                if ' > ' in field_path:
                    parts = field_path.split(' > ')
                    leaf_text = parts[-1]  # 最后一级用于匹配
                    leaf_nodes.append({
                        'full_path': field_path,  # 完整路径用于显示
                        'leaf_text': leaf_text,  # 叶子文本用于匹配
                        'is_statistical': is_statistical,
                        'original_field': field  # 保留原始字段信息
                    })
                else:
                    leaf_nodes.append({
                        'full_path': field_path,
                        'leaf_text': field_path,
                        'is_statistical': is_statistical,
                        'original_field': field
                    })

            print(f"     字段详情: path='{field_path}', leaf_text='{leaf_nodes[-1]['leaf_text'] if leaf_nodes else ''}'")

        return leaf_nodes

    def match_tables_by_leaf_nodes(self, llm_tables: List[Dict], ocr_tables: List[Dict]) -> List[Dict]:
        """
        修复版表格匹配 - 适配字典格式
        """
        print("🔄 开始表格匹配...")

        matches = []
        matched_ocr_indices = set()

        for llm_idx, llm_table in enumerate(llm_tables):
            print(f"🔍 匹配LLM表格 {llm_idx + 1}: {llm_table.get('table_title', 'Unknown')}")

            best_match = None
            best_score = 0
            best_ocr_idx = -1

            # 分别提取横向和纵向叶子节点
            llm_vertical_leaves = self.extract_leaf_nodes(llm_table.get('vertical_hierarchy_fields', []))
            llm_horizontal_leaves = self.extract_leaf_nodes(llm_table.get('horizontal_hierarchy_fields', []))

            # 更新调试输出
            print(f"   横向叶子: {[h.get('leaf_text', '') for h in llm_horizontal_leaves]}")
            print(f"   纵向叶子: {[v.get('leaf_text', '') for v in llm_vertical_leaves[:3]]}...")  # 只显示前3个

            for ocr_idx, ocr_table in enumerate(ocr_tables):
                if ocr_idx in matched_ocr_indices:
                    continue

                # 提取OCR表格的所有叶子节点
                ocr_leaves = self.extract_ocr_leaf_nodes(ocr_table)
                print(f"   OCR表格 {ocr_idx + 1} 叶子节点: {ocr_leaves[:5]}...")

                # 分别计算横向和纵向相似度
                horizontal_similarity = self._calculate_directional_similarity(llm_horizontal_leaves, ocr_leaves, "横向")
                vertical_similarity = self._calculate_directional_similarity(llm_vertical_leaves, ocr_leaves, "纵向")

                # 综合相似度
                overall_similarity = (horizontal_similarity + vertical_similarity) / 2

                if overall_similarity > best_score and overall_similarity > 0.3:
                    best_score = overall_similarity
                    best_match = ocr_table
                    best_ocr_idx = ocr_idx

            if best_match:
                matches.append({
                    'llm_table': llm_table,
                    'ocr_table': best_match,
                    'similarity_score': best_score,
                    'llm_index': llm_idx,
                    'ocr_index': best_ocr_idx
                })
                matched_ocr_indices.add(best_ocr_idx)
                print(f"   ✅ 匹配成功! 相似度: {best_score:.2f}")
            else:
                self.unmatched_llm_tables.append(llm_table)
                print(f"   ❌ 未找到匹配")

        # 记录未匹配的OCR表格
        for ocr_idx, ocr_table in enumerate(ocr_tables):
            if ocr_idx not in matched_ocr_indices:
                self.unmatched_ocr_tables.append(ocr_table)

        print(f"✅ 匹配完成: {len(matches)} 个成功匹配")
        print(f"⚠️  未匹配LLM表格: {len(self.unmatched_llm_tables)}")
        print(f"⚠️  未匹配OCR表格: {len(self.unmatched_ocr_tables)}")

        return matches

    def _calculate_directional_similarity(self, directional_leaves: List[Dict], ocr_leaves: List[str], direction: str) -> float:
        """
        计算方向性相似度（横向或纵向）- 适配字典格式
        """
        if not directional_leaves:
            return 0.0

        total_similarity = 0
        matched_count = 0

        for llm_leaf_info in directional_leaves:
            # 从字典中提取叶子文本进行匹配
            llm_leaf_text = llm_leaf_info.get('leaf_text', '')

            best_match_score = 0
            for ocr_leaf in ocr_leaves:
                similarity = self.calculate_similarity(llm_leaf_text, ocr_leaf)
                if similarity > best_match_score:
                    best_match_score = similarity

            if best_match_score > 0.6:  # 匹配阈值
                total_similarity += best_match_score
                matched_count += 1

        if matched_count == 0:
            return 0.0

        avg_similarity = total_similarity / matched_count
        match_ratio = matched_count / len(directional_leaves)

        print(f"     {direction}匹配: {matched_count}/{len(directional_leaves)} 项, 平均相似度: {avg_similarity:.2f}")

        return avg_similarity * match_ratio

    def merge_aligned_data(self, matches: List[Dict]) -> List[Dict]:
        """
        合并对齐后的数据
        """
        print("🔄 合并对齐数据...")

        merged_tables = []

        for match in matches:
            llm_table = match['llm_table']
            ocr_table = match['ocr_table']

            merged_table = {
                'table_id': llm_table.get('table_id'),
                'table_title': llm_table.get('table_title'),
                'is_financial': llm_table.get('is_financial'),
                'currency': llm_table.get('currency'),
                'reporting_period': llm_table.get('reporting_period'),
                'similarity_score': match['similarity_score'],
                'llm_hierarchy': {
                    'horizontal': llm_table.get('horizontal_hierarchy_fields', []),
                    'vertical': llm_table.get('vertical_hierarchy_fields', [])
                },
                'ocr_data': {
                    'body_cells': ocr_table.get('body', []),
                    'header': ocr_table.get('header', []),
                    'footer': ocr_table.get('footer', [])
                },
                'location': llm_table.get('location', {})
            }

            merged_tables.append(merged_table)

        print(f"✅ 数据合并完成: {len(merged_tables)} 个表格")
        return merged_tables

    def save_alignment_results(self, merged_tables: List[Dict], output_path: str):
        """
        保存对齐结果到JSON文件
        """
        results = {
            'alignment_summary': {
                'total_aligned_tables': len(merged_tables),
                'unmatched_llm_tables': len(self.unmatched_llm_tables),
                'unmatched_ocr_tables': len(self.unmatched_ocr_tables),
                'average_similarity': sum(t['similarity_score'] for t in merged_tables) / len(merged_tables) if merged_tables else 0
            },
            'aligned_tables': merged_tables,
            'unmatched_tables': {
                'llm': self.unmatched_llm_tables,
                'ocr': self.unmatched_ocr_tables
            }
        }

        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(results, f, ensure_ascii=False, indent=2)

        print(f"💾 对齐结果已保存: {output_path}")

    def _create_summary_sheet(self, wb: openpyxl.Workbook, merged_tables: List[Dict]):
        """
        创建汇总表
        """
        ws = wb.create_sheet("汇总表", 0)

        # 设置表头
        headers = [
            "表格ID", "表格标题", "是否财务报表", "货币单位",
            "报告期间", "匹配相似度", "水平层级数", "垂直层级数",
            "OCR单元格数", "表头单元格数"
        ]

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # 填充数据
        for row, table in enumerate(merged_tables, 2):
            ws.cell(row=row, column=1, value=table.get('table_id', ''))
            ws.cell(row=row, column=2, value=table.get('table_title', ''))
            ws.cell(row=row, column=3, value='是' if table.get('is_financial') else '否')
            ws.cell(row=row, column=4, value=table.get('currency', ''))
            ws.cell(row=row, column=5, value=table.get('reporting_period', ''))
            ws.cell(row=row, column=6, value=table.get('similarity_score', 0))

            # 层级数量
            horizontal_count = len(table.get('llm_hierarchy', {}).get('horizontal', []))
            vertical_count = len(table.get('llm_hierarchy', {}).get('vertical', []))
            ws.cell(row=row, column=7, value=horizontal_count)
            ws.cell(row=row, column=8, value=vertical_count)

            # OCR数据统计
            ocr_data = table.get('ocr_data', {})
            body_count = len(ocr_data.get('body_cells', []))
            header_count = len(ocr_data.get('header', []))
            ws.cell(row=row, column=9, value=body_count)
            ws.cell(row=row, column=10, value=header_count)

        # 设置列宽和样式
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

        # 设置表头样式
        for cell in ws[1]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = openpyxl.styles.Font(bold=True)

    def save_to_excel(self, merged_tables: List[Dict], excel_path: str, page_num=0):
        """
        将对齐后的数据保存到Excel文件
        """
        print("📊 正在生成Excel文件...")

        try:
            wb = openpyxl.Workbook()

            # 创建汇总表
            self._create_summary_sheet(wb, merged_tables)

            # 为每个表格创建重建的工作表
            for i, table in enumerate(merged_tables):
                self._create_enhanced_table_sheet(wb, table, i, page_num)

            # 删除默认创建的空工作表
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])

            wb.save(excel_path)
            print(f"💾 Excel文件已保存: {excel_path}")

        except Exception as e:
            print(f"❌ Excel文件保存失败: {e}")
            raise

    def _create_enhanced_table_sheet(self, wb: openpyxl.Workbook, table: Dict, table_index: int, page_num=0):
        """
        创建增强的表格工作表：只包含表格数据，不包含元信息
        """
        sheet_name = f"P{page_num}_表格_{table_index + 1}"
        if len(sheet_name) > 31:
            sheet_name = f"P{page_num}_Table_{table_index + 1}"

        ws = wb.create_sheet(sheet_name)

        # 直接重建表格结构，跳过表格标题、ID、匹配相似度等元信息
        header_positions = self._rebuild_table_with_llm_headers(ws, table, start_row=1)  # 从第1行开始

        # 可以在这里使用header_positions信息进行后续处理
        if header_positions:
            print(f"📊 表格{table_index + 1}的表头位置: {header_positions}")

        # 设置列宽
        for col in range(1, 10):
            ws.column_dimensions[get_column_letter(col)].width = 20

    def _rebuild_table_with_llm_headers(self, ws, table: Dict, start_row: int):
        """修正版：确保完全替换所有OCR表头行，并处理序号列情况"""
        ocr_data = table.get('ocr_data', {})
        body_cells = ocr_data.get('body_cells', [])

        if not body_cells:
            return

        # 1. 构建OCR表格结构
        cell_map = {}
        max_row, max_col = 0, 0
        for cell in body_cells:
            row = cell.get('row_start', cell.get('row', 0))
            col = cell.get('col_start', cell.get('col', 0))
            words = cell.get('words', '').strip()
            max_row, max_col = max(max_row, row), max(max_col, col)
            cell_map[(row, col)] = words

        print(f"📊 OCR表格结构: {max_row + 1}行 x {max_col + 1}列")

        # 2. 分析LLM表头结构
        llm_horizontal = self.extract_leaf_nodes(table.get('llm_hierarchy', {}).get('horizontal', []))
        llm_vertical = self.extract_leaf_nodes(table.get('llm_hierarchy', {}).get('vertical', []))

        print(f"🔍 LLM横向字段: {[h.get('full_path', '') for h in llm_horizontal]}")
        print(f"🔍 LLM纵向字段: {[v.get('full_path', '') for v in llm_vertical]}")

        current_row = start_row

        # 3. 检测是否为多行表头表格
        is_multi_header = self._is_multi_header_table(cell_map, max_row, max_col, llm_horizontal)

        if is_multi_header:
            print("   检测到多行表头表格，使用LLM表头完全替换...")

            # 关键修复：更严格的数据起始行检测
            data_start_row = self._find_real_data_start_row_strict(cell_map, max_row, max_col, llm_horizontal)
            print(f"   数据起始行: {data_start_row}")

            # 右端对齐写入LLM横向表头
            llm_count = len(llm_horizontal)
            ocr_count = max_col + 1

            print(f"   多级表头右端对齐: LLM={llm_count}, OCR={ocr_count}")

            # 从右端开始写入LLM表头
            for i in range(llm_count):
                llm_index = llm_count - 1 - i
                ocr_index = ocr_count - 1 - i

                if llm_index >= 0 and ocr_index >= 0:
                    llm_header = llm_horizontal[llm_index]
                    header_text = llm_header.get('full_path', '')
                    ws.cell(row=current_row, column=ocr_index + 1, value=header_text)
                    print(f"     ✅ 右端写入表头列{ocr_index}: '{header_text}'")

            current_row += 1

            # 处理数据行（跳过所有OCR表头行）
            for row in range(data_start_row, max_row + 1):
                # 关键修复：更严格的表头行检测
                if self._is_header_row_strict(cell_map, row, max_col, llm_horizontal):
                    print(f"     跳过OCR表头行 {row}: {[cell_map.get((row, c), '') for c in range(max_col + 1)]}")
                    continue

                # 🔧 修复：智能识别纵向表头列
                vertical_header_col = self._find_vertical_header_column_enhanced(
                    cell_map, row, max_col, llm_vertical
                )

                # 处理纵向表头
                if vertical_header_col is not None:
                    if (row, vertical_header_col) in cell_map:
                        ocr_vertical_header = cell_map[(row, vertical_header_col)]
                        # 找到对应的LLM纵向字段
                        matched_llm_header = ocr_vertical_header
                        for llm_vertical_item in llm_vertical:
                            llm_leaf = llm_vertical_item.get('leaf_text', '')
                            similarity = self.calculate_similarity(ocr_vertical_header, llm_leaf)
                            if similarity > 0.7:
                                matched_llm_header = llm_vertical_item.get('full_path', '')
                                print(f"     ✅ 替换纵向表头行{row}列{vertical_header_col}: '{ocr_vertical_header}' -> '{matched_llm_header}'")
                                break

                        ws.cell(row=current_row, column=vertical_header_col + 1, value=matched_llm_header)
                    else:
                        # 如果该列没有内容，写入空值
                        ws.cell(row=current_row, column=vertical_header_col + 1, value="")
                else:
                    # 如果没有找到纵向表头列，在第一列写入空值
                    ws.cell(row=current_row, column=1, value="")

                # 处理数据列（跳过纵向表头列）
                for col in range(max_col + 1):
                    if col == vertical_header_col:
                        continue  # 跳过已经处理的纵向表头列

                    if (row, col) in cell_map:
                        cell_value = cell_map[(row, col)]
                        ws.cell(row=current_row, column=col + 1, value=cell_value)

                current_row += 1

        else:
            # 普通表格处理逻辑
            print("   普通表格处理...")
            # 提取OCR横向表头（第0行）
            ocr_headers = []
            for col in range(max_col + 1):
                if (0, col) in cell_map:
                    ocr_headers.append(cell_map[(0, col)])
                else:
                    ocr_headers.append("")

            print(f"   OCR表头: {ocr_headers}")

            # 智能右端对齐替换策略 - 确保第一列"项目"不被替换
            final_headers = ocr_headers.copy()  # 先复制OCR表头

            # 计算需要替换的列数
            llm_count = len(llm_horizontal)
            ocr_count = len(ocr_headers)

            print(f"   LLM表头数量: {llm_count}, OCR表头数量: {ocr_count}")

            # 如果数量相等，直接一一对应替换（但保留第一列）
            if llm_count == ocr_count:
                # 保留第一列"项目"，从第二列开始替换
                for i in range(1, llm_count):  # 从1开始，跳过第一列
                    llm_header = llm_horizontal[i]
                    final_headers[i] = llm_header.get('full_path', '')
                    print(f"     ✅ 替换列{i}: '{ocr_headers[i]}' -> '{final_headers[i]}'")

            # 如果LLM表头少于OCR表头，从右端开始替换（保留第一列）
            elif llm_count < ocr_count:
                replace_count = llm_count
                for i in range(replace_count):
                    llm_index = llm_count - 1 - i
                    ocr_index = ocr_count - 1 - i

                    if llm_index >= 0 and ocr_index >= 1:  # ocr_index >= 1 跳过第一列
                        llm_header = llm_horizontal[llm_index]
                        final_headers[ocr_index] = llm_header.get('full_path', '')
                        print(f"     ✅ 右端替换列{ocr_index}: '{ocr_headers[ocr_index]}' -> '{final_headers[ocr_index]}'")

            # 如果LLM表头多于OCR表头，从左端开始填充（但保留第一列）
            else:  # llm_count > ocr_count
                # 保留第一列"项目"，从第二列开始填充
                for i in range(1, ocr_count):  # 从1开始，跳过第一列
                    if i < llm_count:
                        llm_header = llm_horizontal[i]
                        final_headers[i] = llm_header.get('full_path', '')
                        print(f"     🔄 左端填充列{i}: '{ocr_headers[i]}' -> '{final_headers[i]}'")

            # 写入横向表头
            for col, header in enumerate(final_headers):
                if header:  # 只写入非空表头
                    ws.cell(row=current_row, column=col + 1, value=header)

            current_row += 1

            # 处理数据行（从第1行开始）
            for row in range(1, max_row + 1):
                # 🔧 修复：智能识别纵向表头列
                vertical_header_col = self._find_vertical_header_column_enhanced(
                    cell_map, row, max_col, llm_vertical
                )

                # 处理纵向表头
                if vertical_header_col is not None:
                    if (row, vertical_header_col) in cell_map:
                        ocr_vertical_header = cell_map[(row, vertical_header_col)]

                        # 找到对应的LLM纵向字段
                        matched_llm_header = ocr_vertical_header
                        for llm_vertical_item in llm_vertical:
                            llm_leaf = llm_vertical_item.get('leaf_text', '')
                            similarity = self.calculate_similarity(ocr_vertical_header, llm_leaf)
                            if similarity > 0.7:
                                matched_llm_header = llm_vertical_item.get('full_path', '')
                                print(f"     ✅ 替换纵向表头行{row}列{vertical_header_col}: '{ocr_vertical_header}' -> '{matched_llm_header}'")
                                break

                        ws.cell(row=current_row, column=vertical_header_col + 1, value=matched_llm_header)
                    else:
                        # 如果该列没有内容，写入空值
                        ws.cell(row=current_row, column=vertical_header_col + 1, value="")
                else:
                    # 如果没有找到纵向表头列，在第一列写入空值
                    ws.cell(row=current_row, column=1, value="")

                # 处理数据列（跳过纵向表头列）
                for col in range(max_col + 1):
                    if col == vertical_header_col:
                        continue  # 跳过已经处理的纵向表头列

                    if (row, col) in cell_map:
                        cell_value = cell_map[(row, col)]
                        ws.cell(row=current_row, column=col + 1, value=cell_value)

                current_row += 1

        # 设置对齐方式
        for row in range(start_row, current_row):
            for col in range(1, max_col + 2):
                cell_obj = ws.cell(row=row, column=col)
                # 纵向表头列左对齐，其他列居中对齐
                if col == 1:
                    cell_obj.alignment = Alignment(horizontal='left', vertical='center')
                else:
                    cell_obj.alignment = Alignment(horizontal='center', vertical='center')

    def _find_vertical_header_column_enhanced(self, cell_map: Dict, row: int, max_col: int,
                                              llm_vertical: List[Dict]) -> int:
        """
        增强版：智能识别纵向表头列
        如果第一列是序号，则尝试使用第二列作为纵向表头
        """
        # 检查第一列是否为序号
        first_col_text = cell_map.get((row, 0), '').strip()

        # 如果是纯数字序号，则尝试第二列
        if first_col_text.isdigit():
            print(f"     第{row}行第1列为序号 '{first_col_text}'，尝试第2列作为纵向表头")

            # 检查第二列是否有内容且与LLM纵向字段匹配
            if (row, 1) in cell_map:
                second_col_text = cell_map.get((row, 1), '').strip()
                if second_col_text:
                    # 检查第二列内容是否与LLM纵向字段匹配
                    for llm_vertical_item in llm_vertical:
                        llm_leaf = llm_vertical_item.get('leaf_text', '')
                        similarity = self.calculate_similarity(second_col_text, llm_leaf)
                        if similarity > 0.6:  # 降低匹配阈值以增加匹配机会
                            print(f"     ✅ 第{row}行使用第2列作为纵向表头: '{second_col_text}'")
                            return 1  # 返回第二列索引

        # 如果第一列不是序号或第二列不匹配，使用原来的逻辑
        # 检查第一列是否有内容且与LLM纵向字段匹配
        if first_col_text and not first_col_text.isdigit():
            for llm_vertical_item in llm_vertical:
                llm_leaf = llm_vertical_item.get('leaf_text', '')
                similarity = self.calculate_similarity(first_col_text, llm_leaf)
                if similarity > 0.7:
                    print(f"     ✅ 第{row}行使用第1列作为纵向表头: '{first_col_text}'")
                    return 0  # 返回第一列索引

        # 如果前两列都不匹配，尝试在整个行中寻找最佳匹配
        best_col = 0
        best_similarity = 0

        for col in range(max_col + 1):
            cell_text = cell_map.get((row, col), '').strip()
            if cell_text and not cell_text.isdigit():  # 排除纯数字
                for llm_vertical_item in llm_vertical:
                    llm_leaf = llm_vertical_item.get('leaf_text', '')
                    similarity = self.calculate_similarity(cell_text, llm_leaf)
                    if similarity > best_similarity:
                        best_similarity = similarity
                        best_col = col

        if best_similarity > 0.6:
            best_text = cell_map.get((row, best_col), '')
            print(f"     🔄 第{row}行使用第{best_col + 1}列作为纵向表头: '{best_text}' (相似度: {best_similarity:.2f})")
            return best_col

        # 如果没有找到匹配的列，返回None，让调用方处理
        print(f"     ⚠️ 第{row}行未找到合适的纵向表头列")
        return None

    def _find_real_data_start_row_strict(self, cell_map: Dict, max_row: int, max_col: int,
                                         llm_horizontal: List[Dict]) -> int:
        """严格版：找到真正的数据起始行"""
        for row in range(max_row + 1):
            # 检查该行是否包含数字数据
            has_numeric_data = False
            row_contents = []

            for col in range(max_col + 1):
                text = cell_map.get((row, col), '')
                row_contents.append(text)
                # 🔧 更严格的数字检测：必须包含逗号分隔的数字（财务数据特征）
                if any(char.isdigit() for char in text) and ',' in text and len(text) > 6:
                    has_numeric_data = True
                    break

            print(f"     检查行{row}: {row_contents} -> 有财务数据: {has_numeric_data}")

            if has_numeric_data:
                return row

        # 如果没有找到明确的财务数据行，使用保守策略
        print("     ⚠️ 未找到明确的财务数据行，使用保守策略从第2行开始")
        return 2

    def _is_header_row_strict(self, cell_map: Dict, row: int, max_col: int, llm_horizontal: List[Dict]) -> bool:
        """严格版：检测指定行是否为表头行"""
        row_contents = []
        for col in range(max_col + 1):
            text = cell_map.get((row, col), '')
            row_contents.append(text)

            # 🔧 严格检测：如果包含典型的表头文本
            header_keywords = ['应纳税暂时性差异', '递延所得税负债', '项目', '差异', '负债']
            if any(keyword in text for keyword in header_keywords):
                print(f"       检测到表头关键词: '{text}'")
                return True

            # 检查是否与LLM表头匹配
            for llm_header in llm_horizontal:
                llm_leaf = llm_header.get('leaf_text', '')
                similarity = self.calculate_similarity(text, llm_leaf)
                if similarity > 0.7:
                    print(f"       与LLM表头匹配: '{text}' -> '{llm_leaf}'")
                    return True

        # 🔧 额外检查：如果该行不包含任何数字，很可能是表头行
        has_digits = any(any(char.isdigit() for char in text) for text in row_contents if text)
        if not has_digits:
            print(f"       行{row}不包含数字，判定为表头行")
            return True

        return False

    def _is_multi_header_table(self, cell_map: Dict, max_row: int, max_col: int, llm_horizontal: List[Dict]) -> bool:
        """修复版：检测是否为多行表头表格"""
        # 🔧 修复1：如果LLM横向字段包含多级路径（有">"符号），直接判定为多级表头
        if any(' > ' in h.get('full_path', '') for h in llm_horizontal):
            print("    检测到LLM多级路径，判定为多级表头表格")
            return True

        # 🔧 修复2：检查OCR表格前两行是否都包含表头内容
        header_row_count = 0
        for row in range(min(2, max_row + 1)):  # 只检查前2行
            is_header_row = False
            for col in range(max_col + 1):
                if (row, col) in cell_map:
                    cell_text = cell_map[(row, col)].strip()
                    # 检查是否为表头内容（不包含数字）
                    if cell_text and not any(char.isdigit() for char in cell_text):
                        # 检查是否与LLM表头匹配
                        for llm_header in llm_horizontal:
                            llm_leaf = llm_header.get('leaf_text', '')
                            similarity = self.calculate_similarity(cell_text, llm_leaf)
                            if similarity > 0.7:
                                is_header_row = True
                                break
                        if is_header_row:
                            break
            if is_header_row:
                header_row_count += 1

        result = header_row_count >= 2
        print(f"    多级表头检测结果: {result} (表头行数: {header_row_count})")
        return result

    def align_data(self, llm_path: str, ocr_path: str, output_path: str = 'aligned_results.json',
                   excel_path: str = None, image_n=0, use_image_id=False):
        """
        主对齐函数 - 支持通过image_id或索引匹配
        """
        print("🚀 开始数据对齐流程...")

        # 1. 加载数据
        llm_data, ocr_data = self.load_data(llm_path, ocr_path)

        # 2. 根据image_id或索引提取表格数据
        if use_image_id and 'image_info' in ocr_data:
            # 使用image_id匹配
            ocr_image_id = ocr_data.get('image_info', {}).get('image_id')
            llm_tables = []
            ocr_tables = ocr_data.get('tables_result', [])
            page_name = ""

            # 在LLM数据中查找匹配的image_id
            for image_result in llm_data.get('image_results', []):
                if image_result.get('image_id') == ocr_image_id:
                    llm_tables = image_result.get('tables', [])
                    page_name = image_result.get("image_path", "")
                    break

            if not llm_tables:
                print(f"❌ 未找到与OCR image_id '{ocr_image_id}' 匹配的LLM数据")
                return []
        else:
            # 使用索引匹配（原有逻辑）
            image_llm_tables = llm_data.get('image_results', [{}])[image_n]
            llm_tables = image_llm_tables.get('tables', [])
            page_name = image_llm_tables.get("image_path", "")
            ocr_tables = ocr_data.get('tables_result', [])

        # 提取页码信息
        page_num = '0'
        if page_name:
            image_last_name = os.path.basename(page_name)
            page_num = image_last_name.split('.')[0].split('_')[-1]

        print(f"📊 待匹配表格: LLM={len(llm_tables)}, OCR={len(ocr_tables)}")

        # 3. 基于叶子节点匹配表格
        matches = self.match_tables_by_leaf_nodes(llm_tables, ocr_tables)

        # 4. 合并对齐数据
        merged_tables = self.merge_aligned_data(matches)

        # 5. 保存JSON结果
        self.save_alignment_results(merged_tables, output_path)

        # 6. 保存Excel结果
        if excel_path:
            self.save_to_excel(merged_tables, excel_path, page_num)
        else:
            excel_path = output_path.replace('.json', '.xlsx')
            self.save_to_excel(merged_tables, excel_path, page_num)

        print("🎉 数据对齐流程完成!")
        return merged_tables