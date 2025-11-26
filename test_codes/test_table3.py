import json
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment


def baidu2excel_fixed(bd_json, out_file='baidu_table_fixed.xlsx'):
    """
    修复版的百度表格JSON转Excel
    """
    wb = openpyxl.Workbook()

    # 先删除默认创建的sheet
    wb.remove(wb.active)

    # 检查数据结构
    if 'tables_result' not in bd_json:
        print("❌ 错误：JSON中缺少 'tables_result' 字段")
        print("✅ 可用字段:", bd_json.keys())
        return

    for idx, tbl in enumerate(bd_json['tables_result']):
        print(f"📊 处理表格 {idx + 1}")

        # 创建新的工作表
        ws = wb.create_sheet(title=f'Table{idx + 1}')

        # 检查表格结构
        if 'body' not in tbl:
            print(f"❌ 表格 {idx} 缺少 'body' 字段")
            continue

        # 简单的单元格映射 - 直接写入，不处理复杂合并
        for cell in tbl['body']:
            try:
                row_start = cell.get('row_start', 0)
                col_start = cell.get('col_start', 0)
                words = cell.get('words', '')

                # 直接写入单元格
                excel_row = row_start + 1
                excel_col = col_start + 1

                ws.cell(row=excel_row, column=excel_col, value=words)

                # 设置对齐方式
                cell_obj = ws.cell(row=excel_row, column=excel_col)
                if col_start == 0:  # 第一列左对齐
                    cell_obj.alignment = Alignment(horizontal='left', vertical='center')
                else:  # 其他列居中对齐
                    cell_obj.alignment = Alignment(horizontal='center', vertical='center')

            except Exception as e:
                print(f"⚠️ 写入单元格失败: {e}")
                continue

        # 设置基本列宽
        for col in range(1, 10):  # 假设最多10列
            ws.column_dimensions[get_column_letter(col)].width = 15

        print(f"✅ 表格 {idx + 1} 完成: {len(tbl['body'])} 个单元格")

    # 保存文件
    wb.save(out_file)
    print(f"🎉 文件已生成: {out_file}")


def debug_json_structure(bd_json):
    """
    调试JSON结构
    """
    print("🔍 调试JSON结构:")
    print(f"根级字段: {list(bd_json.keys())}")

    if 'tables_result' in bd_json:
        print(f"表格数量: {len(bd_json['tables_result'])}")

        for i, tbl in enumerate(bd_json['tables_result']):
            print(f"表格 {i} 字段: {list(tbl.keys())}")

            if 'body' in tbl and len(tbl['body']) > 0:
                sample_cell = tbl['body'][0]
                print(f"单元格示例字段: {list(sample_cell.keys())}")
                break


if __name__ == '__main__':
    try:
        with open('data1.json', 'r', encoding='utf-8') as f:
            j = json.load(f)

        # 先调试结构
        debug_json_structure(j)

        # 生成Excel
        save_file = "百度表格-修复版1.xlsx"
        baidu2excel_fixed(j, save_file)

    except FileNotFoundError:
        print("❌ 找不到 data1.json 文件")
    except Exception as e:
        print(f"❌ 程序执行出错: {e}")
        import traceback

        traceback.print_exc()