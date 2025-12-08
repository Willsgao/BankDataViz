
from backend.table_processor.final_data import FinalDataConverter

class TableReconstructor:
    """表格重构器：整合7步流程"""

    def __init__(self):
        self.warnings = []
        self.issues = []
        # 新增：记忆前一个表格的列结构
        self.prev_table_header_structure = None
        self.final_data_converter = FinalDataConverter()

    def _unify_headers_across_tables(self, llm_tables):
        """
        统一相似表格的列标题结构 - 简单实现版
        对于结构相似的相邻表格，如果存在列标题缺失，用前面表格的结构填充
        """
        print("\n=== 列标题统一化处理 ===")

        if len(llm_tables) <= 1:
            print("表格数量≤1，无需统一处理")
            return llm_tables

        for i in range(1, len(llm_tables)):
            current_table = llm_tables[i]
            prev_table = llm_tables[i - 1]

            # 获取列标题
            current_cols = current_table.get('headers', {}).get('cols', [])
            prev_cols = prev_table.get('headers', {}).get('cols', [])

            # 只有列数相同时才考虑统一
            if len(current_cols) == len(prev_cols):
                # 检查是否存在空列标题
                new_cols = []
                need_unify = False

                for idx, (prev_col, curr_col) in enumerate(zip(prev_cols, current_cols)):
                    # 如果当前列标题为空，而前一表格的对应列有内容
                    if (not curr_col or str(curr_col).strip() == '') and prev_col and str(prev_col).strip() != '':
                        new_cols.append(prev_col)
                        need_unify = True
                        print(f"  表格{i}第{idx}列: 空白 → 填充为 '{prev_col}'")
                    else:
                        new_cols.append(curr_col)

                # 如果需要统一，更新表格
                if need_unify:
                    llm_tables[i]['headers']['cols'] = new_cols
                    print(f"  表格{i}统一化完成")

        return llm_tables

    def _analyze_column_structure(self, col_headers):
        """
        分析列标题的结构特征
        返回: 结构特征字典
        """
        if not col_headers:
            return {"empty": True}

        structure = {
            "count": len(col_headers),
            "first_empty": col_headers[0] == '' if col_headers else False,
            "year_count": sum(
                1 for col in col_headers if col and any(year in str(col) for year in ['年', '202', '201'])),
            "change_col": any(col and '变化' in str(col) for col in col_headers),
            "pattern": []
        }

        # 分析列标题模式
        for col in col_headers:
            if not col or str(col).strip() == '':
                structure["pattern"].append("empty")
            elif any(year in str(col) for year in ['2024', '2023', '2022']):
                structure["pattern"].append("year")
            elif '变化' in str(col) or 'chang' in str(col).lower():
                structure["pattern"].append("change")
            else:
                structure["pattern"].append("other")

        return structure

    def _are_similar_column_structures(self, struct1, struct2):
        """
        判断两个列结构是否相似
        """
        if not struct1 or not struct2:
            return False

        # 如果都是空结构
        if struct1.get("empty") and struct2.get("empty"):
            return False

        # 列数相同是关键
        if struct1.get("count") != struct2.get("count"):
            return False

        # 年份列数量相似
        if abs(struct1.get("year_count", 0) - struct2.get("year_count", 0)) > 1:
            return False

        # 是否都有变化列
        if struct1.get("change_col") != struct2.get("change_col"):
            return False

        # 模式相似度（至少60%相同）
        pattern1 = struct1.get("pattern", [])
        pattern2 = struct2.get("pattern", [])

        if len(pattern1) != len(pattern2):
            return False

        same_count = sum(1 for p1, p2 in zip(pattern1, pattern2) if p1 == p2)
        similarity = same_count / len(pattern1) if pattern1 else 0

        return similarity >= 0.6


    # ========== 工具函数 ==========
    def log_warning(self, message):
        """记录警告"""
        self.warnings.append(message)
        print(f"⚠️ {message}")

    def log_issue(self, message):
        """记录问题"""
        self.issues.append(message)
        print(f"❓ {message}")

    def clean_text_for_matching(self, text):
        """清理文本用于匹配"""
        if not text:
            return ""
        text = str(text)
        # 替换换行、空格
        text = text.replace('\n', '').replace('\r', '').replace(' ', '')
        # 去掉特殊符号，保留中文、数字、字母
        cleaned = ''.join(c for c in text if c.isalnum() or '\u4e00-\u9fff' in c)
        return cleaned

    def calculate_similarity(self, text1, text2):
        """计算两个文本的相似度"""
        t1 = self.clean_text_for_matching(text1)
        t2 = self.clean_text_for_matching(text2)

        if not t1 or not t2:
            return 0

        # 完全相等
        if t1 == t2:
            return 1.0

        # 包含关系
        if t1 in t2 or t2 in t1:
            return 0.9

        # 公共字符比例
        common_chars = set(t1) & set(t2)
        if not common_chars:
            return 0

        # 计算Jaccard相似度
        union_chars = set(t1) | set(t2)
        return len(common_chars) / len(union_chars)

    def calculate_similarity_v2(self, text1, text2):
        """增强版相似度计算（考虑包含关系的覆盖率）"""
        t1 = self.clean_text_for_matching(text1)
        t2 = self.clean_text_for_matching(text2)

        if not t1 or not t2:
            return 0

        # 1. 完全相等
        if t1 == t2:
            return 1.0

        # 2. 包含关系，但要求高覆盖率
        if t1 in t2:
            coverage = len(t1) / len(t2)
            # 覆盖度>80%才给高分，否则给低分
            # 避免"经营活动"匹配"每股经营活动"
            return 0.9 if coverage > 0.8 else 0.2

        if t2 in t1:
            coverage = len(t2) / len(t1)
            return 0.9 if coverage > 0.8 else 0.2

        # 3. 公共字符比例（原有逻辑）
        common_chars = set(t1) & set(t2)
        if not common_chars:
            return 0

        # 计算Jaccard相似度
        union_chars = set(t1) | set(t2)
        similarity = len(common_chars) / len(union_chars)

        return similarity

    def find_best_match_row(self, target_text, table, start_row=1, search_cols=None):
        """
        在表格中查找与目标文本最匹配的行
        """
        if not target_text or not table:
            return -1, 0.0

        best_row = -1
        best_score = 0.0

        # 确定搜索列范围
        if search_cols is None:
            # 默认搜索所有列
            search_cols = list(range(len(table[0])))

        for row_idx in range(start_row, len(table)):
            row_score = 0.0

            for col_idx in search_cols:
                if col_idx >= len(table[row_idx]):
                    continue

                cell_text = table[row_idx][col_idx]
                if not cell_text:
                    continue

                # 计算相似度
                similarity = self.calculate_similarity(target_text, str(cell_text))
                if similarity > row_score:
                    row_score = similarity

            # 更新最佳匹配
            if row_score > best_score:
                best_score = row_score
                best_row = row_idx

        return best_row, best_score

    def _detect_tables_to_merge(self, llm_tables):
        """
        检测哪些LLM表格应该合并处理
        返回：需要合并的表格组列表，每个组包含应该合并的表格索引
        """
        print("\n=== 检测可合并的表格结构 ===")

        if len(llm_tables) <= 1:
            print("表格数量≤1，无需合并检测")
            return []

        # 检查每个表格的列结构
        table_col_structures = []

        for i, table in enumerate(llm_tables):
            col_headers = table.get('headers', {}).get('cols', [])
            # 清理列标题（去掉空值和特殊字符）
            cleaned_cols = []
            for col in col_headers:
                if col and str(col).strip():
                    # 提取文本内容（去掉>>等层级标记）
                    text = str(col).split('>>')[-1].strip()
                    if text:
                        cleaned_cols.append(text)

            table_info = {
                'index': i,
                'col_count': len(col_headers),
                'cleaned_cols': cleaned_cols,
                'original_cols': col_headers,
                'row_count': len(table.get('headers', {}).get('rows', []))
            }
            table_col_structures.append(table_info)

            print(f"表格{i}: {table_info['row_count']}行 × {table_info['col_count']}列")
            print(f"  列标题: {cleaned_cols[:5]}...")

        # 检测相似的列结构（用于合并）
        merge_groups = []
        processed_indices = set()

        for i in range(len(table_col_structures)):
            if i in processed_indices:
                continue

            current_table = table_col_structures[i]
            merge_group = [i]

            # 寻找与当前表格列结构相似的表格
            for j in range(i + 1, len(table_col_structures)):
                if j in processed_indices:
                    continue

                next_table = table_col_structures[j]

                # 合并条件1：列数相同
                if current_table['col_count'] != next_table['col_count']:
                    continue

                # 合并条件2：列标题相似度检查（简化版）
                is_similar = self._check_columns_similarity(
                    current_table['cleaned_cols'],
                    next_table['cleaned_cols']
                )

                if is_similar:
                    merge_group.append(j)
                    processed_indices.add(j)

            if len(merge_group) > 1:
                merge_groups.append(merge_group)
                processed_indices.add(i)

        # 输出检测结果
        if merge_groups:
            print("\n检测到可合并的表格组:")
            for group_idx, group in enumerate(merge_groups):
                print(f"  合并组{group_idx}: 表格索引 {group}")
                print(f"    包含表格: ", end="")
                for table_idx in group:
                    table_info = table_col_structures[table_idx]
                    print(f"表格{table_idx}({table_info['row_count']}行) ", end="")
                print()
        else:
            print("未检测到需要合并的表格")

        return merge_groups

    def _check_columns_similarity(self, cols1, cols2):
        """
        检查两个列标题列表是否相似
        简化版：检查前几列是否包含相同的年份/数字模式
        """
        if not cols1 or not cols2:
            return False

        # 如果列数不同，肯定不相似
        if len(cols1) != len(cols2):
            return False

        # 检查是否包含典型的年份模式
        year_patterns = ['年', '202', '201', '月', '季度', 'Q']

        def contains_year_pattern(text):
            if not text:
                return False
            text_str = str(text)
            for pattern in year_patterns:
                if pattern in text_str:
                    return True
            return False

        # 统计两个列表中包含年份模式的列数
        cols1_year_count = sum(1 for col in cols1 if contains_year_pattern(col))
        cols2_year_count = sum(1 for col in cols2 if contains_year_pattern(col))

        # 如果都有年份列，认为可能是相似的表格结构
        if cols1_year_count > 0 and cols2_year_count > 0:
            return True

        # 检查前两列是否相似（对于财务报表，通常第一列是项目，后面是年份）
        similarity_threshold = 0.7
        similar_count = 0

        for idx in range(min(len(cols1), len(cols2))):
            if idx == 0:
                # 第一列通常是项目名称，允许不同
                continue

            sim = self.calculate_similarity(cols1[idx], cols2[idx])
            if sim > similarity_threshold:
                similar_count += 1

        # 如果有超过一半的非首列相似，认为表格结构相似
        if len(cols1) > 1:
            return similar_count / (len(cols1) - 1) > 0.5

        return False

    def _match_with_table_boundaries(self, table, row_headers, table_boundaries):
        """基于表格边界的匹配算法"""
        print(f"表格: {len(table)}行 × {len(table[0])}列")
        print(f"LLM行表头数: {len(row_headers)}")

        # 1. 为每个OCR表格建立搜索空间
        search_spaces = []
        for boundary in table_boundaries:
            start_row = boundary['start_row']
            end_row = boundary['end_row']
            # 转换为1-based索引（表格数据是0-based）
            search_space = list(range(start_row, end_row + 1))
            search_spaces.append({
                'table_idx': boundary['table_idx'],
                'rows': search_space,
                'next_row_idx': 0,  # 在这个表格内下一个可用的行索引
                'used_rows': set()  # 已使用的行
            })
            print(f"  表格{boundary['table_idx']}: 行{start_row}-{end_row}")

        # 2. 确定搜索列（与原有逻辑一致）
        left_empty_cols = 0
        for title in table[0]:
            if not title or str(title).strip() == '':
                left_empty_cols += 1
            else:
                break

        search_columns = list(range(min(3, len(table[0]))))
        if left_empty_cols > 0:
            search_columns = list(range(left_empty_cols))

        print(f"搜索列: {search_columns}")

        # 3. 按LLM顺序匹配
        assignments = {}

        for header_idx, llm_header in enumerate(row_headers):
            target_text = llm_header.split('>>')[-1] if '>>' in llm_header else llm_header

            print(f"  表头[{header_idx}]: '{target_text[:30]}...'")

            best_match = None
            best_score = 0
            best_table_idx = -1

            # 在每个表格的搜索空间中寻找最佳匹配
            for table_info in search_spaces:
                for row_idx in table_info['rows']:
                    if row_idx in table_info['used_rows']:
                        continue

                    # 计算匹配分数
                    row_score = 0
                    for col_idx in search_columns:
                        if col_idx < len(table[row_idx]):
                            cell_val = table[row_idx][col_idx]
                            if cell_val:
                                score = self.calculate_similarity_v2(target_text, str(cell_val))
                                if score > row_score:
                                    row_score = score

                    if row_score > best_score:
                        best_score = row_score
                        best_match = row_idx
                        best_table_idx = table_info['table_idx']

            # 4. 如果找到高质量匹配，分配该行
            if best_match and best_score > 0.6:
                assignments[header_idx] = best_match
                # 标记该行为已使用
                for table_info in search_spaces:
                    if best_match in table_info['rows']:
                        table_info['used_rows'].add(best_match)
                print(f"    → 匹配到表格{best_table_idx}的行{best_match} (分数:{best_score:.2f})")
            else:
                # 5. 按表格顺序分配下一个可用行
                assigned = False
                for table_info in search_spaces:
                    # 在这个表格内按顺序找下一个可用行
                    while table_info['next_row_idx'] < len(table_info['rows']):
                        row_idx = table_info['rows'][table_info['next_row_idx']]
                        if row_idx not in table_info['used_rows']:
                            assignments[header_idx] = row_idx
                            table_info['used_rows'].add(row_idx)
                            table_info['next_row_idx'] += 1
                            assigned = True
                            print(f"    → 顺序分配到表格{table_info['table_idx']}的行{row_idx}")
                            break
                        table_info['next_row_idx'] += 1

                    if assigned:
                        break

                if not assigned:
                    print(f"    → 无法分配")
                    assignments[header_idx] = -1

        # 6. 填充表格（与原有逻辑一致）
        for row_idx in range(1, len(table)):
            table[row_idx][0] = None

        filled_count = 0
        for header_idx, row_idx in assignments.items():
            if row_idx != -1 and row_idx < len(table):
                table[row_idx][0] = row_headers[header_idx]
                filled_count += 1

        print(f"\n匹配结果: {filled_count}/{len(row_headers)} 个行表头已填充")
        return table

    def _validate_table_data(self, table):
        """
        验证表格数据，标记可疑的行和列
        返回: (marked_table, row_checks, col_checks)
        其中 row_checks 和 col_checks 是标记数组
        """
        if not table or len(table) == 0:
            return table, [], []

        num_rows = len(table)
        num_cols = len(table[0]) if table else 0

        if num_rows == 0 or num_cols == 0:
            return table, [], []

        print(f"\n=== 开始表格数据验证 ===")
        print(f"表格尺寸: {num_rows}行 × {num_cols}列")

        # 初始化标记数组
        row_checks = [0] * num_rows  # 0=正常, 1=可疑
        col_checks = [0] * num_cols  # 0=正常, 1=可疑

        # 先分析每列的数据特征
        col_features = self._analyze_column_features(table)

        # 1. 检查孤立数据问题（主要问题）
        self._check_isolated_data(table, row_checks, col_checks)

        # 2. 检查数据类型不一致
        self._check_data_type_inconsistency(table, row_checks, col_checks, col_features)

        # 3. 检查空值模式异常
        self._check_null_patterns(table, row_checks, col_checks)

        # 4. 检查格式不规范
        self._check_format_inconsistency(table, row_checks, col_checks)

        # 创建带标记的表格（不修改原数据）
        marked_table = self._create_marked_table(table, row_checks, col_checks)

        # 输出验证结果
        self._print_validation_results(row_checks, col_checks)

        return marked_table, row_checks, col_checks

    def _analyze_column_features(self, table):
        """分析每列的数据特征"""
        if not table:
            return {}

        num_rows = len(table)
        num_cols = len(table[0]) if table else 0

        col_features = {}

        for col in range(num_cols):
            # 统计该列的数据类型
            numeric_count = 0
            text_count = 0
            empty_count = 0
            has_percent = False
            has_currency = False

            for row in range(num_rows):
                value = table[row][col]
                if value is None or str(value).strip() == '':
                    empty_count += 1
                elif self._is_numeric_value(value):
                    numeric_count += 1
                    val_str = str(value)
                    if '%' in val_str:
                        has_percent = True
                    if any(symbol in val_str for symbol in ['¥', '$', '€', '￡', 'HK$']):
                        has_currency = True
                else:
                    text_count += 1

            total_non_empty = num_rows - empty_count

            # 确定主要数据类型
            main_data_type = "unknown"
            if total_non_empty > 0:
                if numeric_count / total_non_empty > 0.7:
                    main_data_type = "numeric"
                elif text_count / total_non_empty > 0.7:
                    main_data_type = "text"
                else:
                    main_data_type = "mixed"

            col_features[col] = {
                "main_data_type": main_data_type,
                "numeric_count": numeric_count,
                "text_count": text_count,
                "empty_count": empty_count,
                "has_percent": has_percent,
                "has_currency": has_currency,
                "total_non_empty": total_non_empty
            }

        return col_features

    def _check_isolated_data(self, table, row_checks, col_checks):
        """检查孤立数据问题"""
        num_rows = len(table)
        num_cols = len(table[0]) if table else 0

        for row in range(num_rows):
            for col in range(1, num_cols):  # 从第1列开始（第0列是行表头）
                value = table[row][col]

                # 如果有数据
                if value and str(value).strip() != '':
                    # 检查这一行的第一列是否为空
                    row_header = table[row][0] if 0 < num_cols else None

                    if row_header is None or str(row_header).strip() == '':
                        # 检查上一行的第一列是否有内容
                        prev_row_header = table[row - 1][0] if row > 0 and 0 < num_cols else None

                        if prev_row_header and str(prev_row_header).strip() != '':
                            # 检查上一行的同列是否为空
                            prev_row_same_col = table[row - 1][col] if row > 0 else None

                            if prev_row_same_col is None or str(prev_row_same_col).strip() == '':
                                # 发现孤立数据！
                                row_checks[row] = 1
                                col_checks[col] = 1

                                # 也标记上一行（可能需要合并）
                                if row > 0:
                                    row_checks[row - 1] = 1

                                print(f"  ⚠️ 发现孤立数据: 行{row}列{col}='{value}'")
                                print(f"     当前行表头为空，上一行表头为'{prev_row_header}'")

    def _check_data_type_inconsistency(self, table, row_checks, col_checks, col_features):
        """检查数据类型不一致"""
        num_rows = len(table)
        num_cols = len(table[0]) if table else 0

        for col in range(num_cols):
            if col not in col_features:
                continue

            features = col_features[col]
            main_type = features["main_data_type"]

            if main_type in ["numeric", "text"]:
                for row in range(num_rows):
                    value = table[row][col]
                    if value is None or str(value).strip() == '':
                        continue

                    # 检查是否与主要类型不一致
                    is_numeric = self._is_numeric_value(value)

                    if main_type == "numeric" and not is_numeric:
                        # 应该是数字但不是数字
                        row_checks[row] = 1
                        col_checks[col] = 1
                        print(f"  ⚠️ 数据类型不一致: 行{row}列{col}='{value}'")
                        print(f"     该列主要类型为数字，但此单元格为文本")

                    elif main_type == "text" and is_numeric:
                        # 应该是文本但是数字
                        row_checks[row] = 1
                        col_checks[col] = 1
                        print(f"  ⚠️ 数据类型不一致: 行{row}列{col}='{value}'")
                        print(f"     该列主要类型为文本，但此单元格为数字")

    def _check_null_patterns(self, table, row_checks, col_checks):
        """检查空值模式异常"""
        num_rows = len(table)
        num_cols = len(table[0]) if table else 0

        # 检查列的空值模式
        for col in range(num_cols):
            empty_positions = []
            for row in range(num_rows):
                value = table[row][col]
                if value is None or str(value).strip() == '':
                    empty_positions.append(row)

            # 如果该列只有少数几个空值，标记这些行
            if 0 < len(empty_positions) < num_rows * 0.3:  # 少于30%为空
                for row in empty_positions:
                    # 检查该行的其他列是否有数据
                    has_other_data = False
                    for c in range(num_cols):
                        if c != col:
                            val = table[row][c]
                            if val and str(val).strip() != '':
                                has_other_data = True
                                break

                    if has_other_data:
                        row_checks[row] = 1
                        col_checks[col] = 1
                        print(f"  ⚠️ 异常空值: 行{row}列{col}为空，但该行其他列有数据")

    def _check_format_inconsistency(self, table, row_checks, col_checks):
        """检查格式不规范"""
        num_rows = len(table)
        num_cols = len(table[0]) if table else 0

        for col in range(num_cols):
            # 收集该列所有数值的格式特征
            has_comma = False
            has_dot = False
            has_percent_in_any = False

            for row in range(num_rows):
                value = table[row][col]
                if value and str(value).strip() != '':
                    val_str = str(value)
                    if ',' in val_str:
                        has_comma = True
                    if '.' in val_str and any(ch.isdigit() for ch in val_str):
                        has_dot = True
                    if '%' in val_str:
                        has_percent_in_any = True

            # 检查格式一致性
            for row in range(num_rows):
                value = table[row][col]
                if value and str(value).strip() != '':
                    val_str = str(value)

                    # 检查千分位格式不一致
                    if has_comma and ',' not in val_str and self._is_numeric_value(value):
                        # 其他单元格有逗号，但这个没有
                        row_checks[row] = 1
                        col_checks[col] = 1
                        print(f"  ⚠️ 格式不一致: 行{row}列{col}='{value}'")
                        print(f"     其他单元格使用千分位逗号，但此单元格没有")

                    # 检查百分比格式不一致
                    if has_percent_in_any and '%' not in val_str and self._looks_like_percentage(value):
                        # 看起来像百分比但没有%符号
                        row_checks[row] = 1
                        col_checks[col] = 1
                        print(f"  ⚠️ 格式不一致: 行{row}列{col}='{value}'")
                        print(f"     其他单元格有%符号，但此单元格没有")

    def _create_marked_table(self, table, row_checks, col_checks):
        """创建带标记的表格"""
        if not table:
            return []

        num_rows = len(table)
        num_cols = len(table[0]) if table else 0

        # 创建新表格（增加一行一列用于标记）
        marked_table = []
        for r in range(num_rows + 1):
            marked_table.append([None] * (num_cols + 1))

        # 复制原始数据
        for r in range(num_rows):
            for c in range(num_cols):
                marked_table[r][c] = table[r][c]

        # 添加行标记（在最后一列）
        for r in range(num_rows):
            marked_table[r][num_cols] = row_checks[r]

        # 添加列标记（在最后一行）
        for c in range(num_cols):
            marked_table[num_rows][c] = col_checks[c]

        # 右下角单元格标记总数
        total_issues = sum(row_checks) + sum(col_checks)
        marked_table[num_rows][num_cols] = f"问题:{total_issues}"

        return marked_table

    def _print_validation_results(self, row_checks, col_checks):
        """打印验证结果"""
        suspicious_rows = [i for i, check in enumerate(row_checks) if check == 1]
        suspicious_cols = [i for i, check in enumerate(col_checks) if check == 1]

        print(f"\n=== 验证结果 ===")
        print(f"可疑行: {len(suspicious_rows)}个")
        if suspicious_rows:
            print(f"  行索引: {suspicious_rows}")

        print(f"可疑列: {len(suspicious_cols)}个")
        if suspicious_cols:
            print(f"  列索引: {suspicious_cols}")

        total_cells = len(row_checks) * len(col_checks) if row_checks and col_checks else 0
        suspicious_cells = sum(row_checks) * sum(col_checks)  # 粗略估计

        print(f"总检查点: {total_cells}")
        print(f"可疑单元格估计: {suspicious_cells}")
        print("=" * 40)

    # 工具函数
    def _looks_like_percentage(self, value):
        """判断是否看起来像百分比"""
        if not value:
            return False

        value_str = str(value).strip()

        # 已经包含%符号
        if '%' in value_str:
            return True

        # 检查是否为0-100之间的数字（可能是百分比）
        try:
            cleaned = value_str.replace(',', '').replace(' ', '')
            num = float(cleaned)
            return 0 <= num <= 100
        except:
            return False

    def _check_cell_validity(self, cell, row_idx, col_idx):
        """
        检查单元格数据的有效性
        """
        if cell is None or cell == "":
            return False

        # 检查是否为数值型数据
        if isinstance(cell, (int, float)):
            return True

        # 检查是否为可转换的数值字符串
        cell_str = str(cell).strip()

        # 移除括号（表示负数的括号表示法）
        if cell_str.startswith('(') and cell_str.endswith(')'):
            cell_str = '-' + cell_str[1:-1]

        # 检查是否为数字（包括负数和小数）
        try:
            # 移除百分比符号等
            clean_str = cell_str.replace('%', '').replace(',', '')
            float(clean_str)
            return True
        except ValueError:
            # 可能是文本或混合内容
            return cell_str not in ['N/A', 'null', '--', '---']

    def _generate_row_feature(self, score):
        """
        根据行数据质量生成特征标记
        """
        if score >= 0.9:
            return "1"  # 高质量行
        elif score >= 0.7:
            return "0"  # 中等质量行
        else:
            return "问题"  # 低质量行，需要检查

    def _generate_col_feature(self, score):
        """
        根据列数据质量生成特征标记
        """
        if score >= 0.9:
            return "1"  # 高质量列
        elif score >= 0.7:
            return "0"  # 中等质量列
        else:
            return "检查"  # 低质量列

    def _generate_overall_feature(self, score):
        """
        生成整体表格的特征标记
        """
        if score >= 0.9:
            return "表格完整"
        elif score >= 0.7:
            return "部分问题"
        else:
            return f"问题:{int((1 - score) * 100)}"

    def step8_add_feature_marks11(self, validated_table, validation_marks):
        """
        添加特征标记：
        - 最后一列：行标记（第一行写"行标记"作为列标题）
        - 最后一行：列标记（第一列写"列标记"作为行标题）
        - 最后一行最后一列：行标记和列标记的最大值
        """
        marked_table = [row[:] for row in validated_table]

        # 添加调试信息
        print(f"\n=== step8调试信息 ===")
        print(f"输入表格行数: {len(validated_table)}")
        print(f"行标记数量: {len(validation_marks['row_marks'])} - {validation_marks['row_marks']}")
        print(f"列标记数量: {len(validation_marks['col_marks'])} - {validation_marks['col_marks']}")

        if not marked_table:
            return marked_table

        row_marks = validation_marks["row_marks"]
        col_marks = validation_marks["col_marks"]

        # 1. 添加最后一列（行标记列）
        print(f"\n添加行标记列:")
        for i in range(len(marked_table)):
            if i == 0:
                # 第一行：写"行标记"作为列标题
                marked_table[i].append("行标记")
                print(f"  行{i}: 添加'行标记'（列标题）")
            else:
                # 其他行：直接使用对应的行标记
                # row_marks对应validated_table的每一行，所以索引应该是i
                mark_idx = i if i < len(row_marks) else len(row_marks) - 1
                mark_value = str(row_marks[mark_idx]) if mark_idx >= 0 and mark_idx < len(row_marks) else "0"
                marked_table[i].append(mark_value)
                print(f"  行{i}: 添加行标记{mark_value} (使用row_marks[{mark_idx}])")

        # 2. 添加最后一行（列标记行）
        last_row = []
        print(f"\n创建列标记行:")
        for j in range(len(marked_table[0])):  # 包括新添加的最后一列
            if j == 0:
                # 第一列：写"列标记"作为行标题
                last_row.append("列标记")
                print(f"  列{j}: 添加'列标记'（行标题）")
            else:
                # 其他列：直接使用对应的列标记
                # col_marks对应validated_table的每一列，所以索引应该是j-1（因为j=0是表头列）
                mark_idx = j - 1 if (j - 1) < len(col_marks) else len(col_marks) - 1
                mark_value = str(col_marks[mark_idx]) if mark_idx >= 0 and mark_idx < len(col_marks) else "0"
                last_row.append(mark_value)
                print(f"  列{j}: 添加列标记{mark_value} (使用col_marks[{mark_idx}])")

        # 3. 修正最后一列最后一行的值（取行标记和列标记的最大值）
        max_row_mark = max(row_marks) if row_marks else 0
        max_col_mark = max(col_marks) if col_marks else 0
        overall_mark = max(max_row_mark, max_col_mark)
        last_row[-1] = str(overall_mark)  # 最后一行最后一列
        print(f"\n最后一行最后一列: {overall_mark} (max({max_row_mark}, {max_col_mark}))")

        marked_table.append(last_row)
        print(f"最终表格: {len(marked_table)}行 × {len(marked_table[0])}列")

        return marked_table

    def step8_add_feature_marks(self, validated_table, validation_marks):
        """
        添加特征标记 - 修复最后一列表头问题
        """
        marked_table = [row[:] for row in validated_table]  # 深拷贝

        print(f"\n=== step8调试信息 ===")
        print(f"输入表格行数: {len(validated_table)}")
        print(f"行标记数量: {len(validation_marks['row_marks'])}")
        print(f"列标记数量: {len(validation_marks['col_marks'])}")

        if not marked_table:
            return marked_table

        row_marks = validation_marks["row_marks"]
        col_marks = validation_marks["col_marks"]

        # 1. 添加最后一列（行标记列）
        print(f"\n添加行标记列:")

        for i in range(len(marked_table)):
            if i == 0:
                # 第一行：添加"行标记"作为列标题
                marked_table[i].append("行标记")
                print(f"  行{i}: 添加'行标记'（列标题）")
            else:
                # 其他行：添加对应的行标记值
                mark_idx = i - 1 if (i - 1) < len(row_marks) else len(row_marks) - 1
                mark_value = str(row_marks[mark_idx]) if mark_idx >= 0 and mark_idx < len(row_marks) else "0"
                marked_table[i].append(mark_value)
                print(f"  行{i}: 添加行标记值{mark_value}")

        # 2. 添加最后一行（列标记行）
        last_row = []
        print(f"\n创建列标记行:")

        # 第一列：写"列标记"作为行标题
        last_row.append("列标记")
        print(f"  列0: 添加'列标记'（行标题）")

        # 其他列：添加列标记（注意索引对齐）
        # 这里要包括新增的行标记列，所以从第1列到最后一列
        for j in range(1, len(marked_table[0])):
            mark_idx = j - 1  # 因为第0列是行表头列

            if mark_idx < len(col_marks):
                mark_value = str(col_marks[mark_idx])
            else:
                # 对于行标记列（最后一列），给它一个特殊的列标记值
                if j == len(marked_table[0]) - 1:
                    mark_value = "1"  # 行标记列的列标记
                else:
                    mark_value = "0"

            last_row.append(mark_value)
            print(f"  列{j}: 添加列标记值{mark_value}")

        marked_table.append(last_row)

        # 输出最终表格信息
        print(f"最终表格: {len(marked_table)}行 × {len(marked_table[0])}列")

        # 打印第一行看看
        print(f"第一行（应包含'行标记'表头）: {marked_table[0]}")

        return marked_table

    def _is_numeric_value(self, cell):
        """
        判断是否为数值
        """
        if cell is None:
            return False

        cell_str = str(cell).strip()

        # 如果已经是标记，不算数值
        if cell_str in ["0", "1", "2", "3"]:
            return False

        # 尝试转换为数值
        try:
            # 处理括号表示法
            if cell_str.startswith('(') and cell_str.endswith(')'):
                cell_str = '-' + cell_str[1:-1]

            # 清理字符
            clean_str = cell_str.replace(',', '').replace(' ', '')

            float(clean_str)
            return True
        except ValueError:
            # 检查是否包含数字（可能是文本数字混合）
            import re
            if re.search(r'\d', cell_str):
                # 包含数字但不完全是数字，算是文本
                return False
            # 纯文本
            return False

    def _contains_digits(self, cell):
        """
        判断是否包含数字
        """
        if cell is None:
            return False

        import re
        cell_str = str(cell)
        return bool(re.search(r'\d', cell_str))

    def step7_validate_and_mark_preserve_headers(self, final_table):
        """
        验证数据并记录标记（保留表头，只检查数据区域）
        """
        validated_table = [row[:] for row in final_table]

        if not validated_table:
            return validated_table, {"row_marks": [], "col_marks": []}

        # 记录每行的标记（基于数据区域）
        row_marks = []

        for i, row in enumerate(validated_table):
            if not row:  # 空行
                row_marks.append(0)
                continue

            # 检查该行是否有数据（跳过第一列表头）
            has_data = False
            for j in range(1, len(row)):  # 从第二列开始检查
                cell = row[j]
                if cell is not None and cell != "":
                    has_data = True
                    break

            if not has_data:  # 这行没有数据
                row_marks.append(0)
                continue

            # 分析数据区域的类型（跳过第一列）
            numeric_count = 0
            text_count = 0
            empty_count = 0  # 新增：记录空单元格数量

            for j in range(1, len(row)):  # 跳过第一列（表头）
                cell = row[j]
                if cell is None or cell == "":
                    empty_count += 1  # 记录空单元格
                    continue

                if self._is_pure_numeric(cell):
                    numeric_count += 1
                else:
                    text_count += 1

            total_count = numeric_count + text_count + empty_count  # 包括空单元格

            # 打印调试信息
            print(f"行 {i}: numeric={numeric_count}, text={text_count}, empty={empty_count}, total={total_count}")

            if total_count == 0:
                row_marks.append(0)
                continue

            # 判断标记
            mark = 0

            # 首先检查空单元格
            if empty_count > 0:
                # 有空单元格，至少标记1
                mark = 1

                # 打印调试信息
                print(f"  行 {i}: 有空单元格 {empty_count}个，标记至少为1")

                # 如果只有很少的数值，标记2
                if numeric_count > 0 and numeric_count <= 1:
                    mark = 2
                    print(f"  行 {i}: 只有{numeric_count}个数值，标记为2")

            # 条件2：最需要检查 - 数值很少（没有空单元格的情况）
            elif numeric_count > 0:
                numeric_ratio = numeric_count / total_count if total_count > 0 else 0
                if numeric_ratio < 0.2 or numeric_count <= 1:
                    mark = 2
                    print(f"  行 {i}: 数值很少({numeric_count}/{total_count})，标记为2")

            # 条件1：需要检查 - 混合类型
            if numeric_count > 0 and text_count > 0 and mark < 2:  # 如果还没标记2
                mark = 1
                print(f"  行 {i}: 混合类型({numeric_count}数值+{text_count}文本)，标记为1")

            # 如果没有标记任何问题，标记0
            if mark == 0:
                print(f"  行 {i}: 格式一致，标记为0")

            row_marks.append(mark)

        # 记录每列的标记（基于数据区域）
        col_marks = []

        if validated_table and validated_table[0]:
            num_cols = len(validated_table[0])
            for j in range(num_cols):
                # 检查该列是否有数据（跳过第一行表头）
                has_data = False
                for i in range(1, len(validated_table)):  # 从第二行开始检查
                    if j < len(validated_table[i]):
                        cell = validated_table[i][j]
                        if cell is not None and cell != "":
                            has_data = True
                            break

                if not has_data:  # 这列没有数据
                    col_marks.append(0)
                    continue

                # 分析数据区域的类型（跳过第一行）
                numeric_count = 0
                text_count = 0
                empty_count = 0  # 新增

                for i in range(1, len(validated_table)):  # 跳过第一行（表头）
                    if j >= len(validated_table[i]):
                        empty_count += 1  # 缺失的单元格
                        continue

                    cell = validated_table[i][j]
                    if cell is None or cell == "":
                        empty_count += 1
                        continue

                    if self._is_pure_numeric(cell):
                        numeric_count += 1
                    else:
                        text_count += 1

                total_count = numeric_count + text_count + empty_count

                # 打印列调试信息
                print(f"列 {j}: numeric={numeric_count}, text={text_count}, empty={empty_count}, total={total_count}")

                if total_count == 0:
                    col_marks.append(0)
                    continue

                # 判断标记
                mark = 0

                # 首先检查空单元格
                if empty_count > 0:
                    mark = 1
                    print(f"  列 {j}: 有空单元格 {empty_count}个，标记至少为1")

                    if numeric_count > 0 and numeric_count <= 1:
                        mark = 2
                        print(f"  列 {j}: 只有{numeric_count}个数值，标记为2")

                # 条件2：最需要检查 - 数值很少
                elif numeric_count > 0 and numeric_count < total_count / 3:
                    mark = 2
                    print(f"  列 {j}: 数值很少({numeric_count}/{total_count})，标记为2")

                # 条件1：需要检查 - 混合类型
                elif numeric_count > 0 and text_count > 0:
                    mark = 1
                    print(f"  列 {j}: 混合类型({numeric_count}数值+{text_count}文本)，标记为1")

                if mark == 0:
                    print(f"  列 {j}: 格式一致，标记为0")

                col_marks.append(mark)

        print(f"最终行标记: {row_marks}")
        print(f"最终列标记: {col_marks}")

        return validated_table, {"row_marks": row_marks, "col_marks": col_marks}

    def _is_pure_numeric(self, cell):
        """
        判断是否为纯数值
        """
        if cell is None:
            return False

        cell_str = str(cell).strip()

        # 如果已经是标记，不算数值
        if cell_str in ["0", "1", "2"]:
            return False

        # 尝试转换为数值
        try:
            # 处理括号表示法
            if cell_str.startswith('(') and cell_str.endswith(')'):
                cell_str = '-' + cell_str[1:-1]

            # 清理字符
            clean_str = cell_str.replace(',', '').replace(' ', '')

            float(clean_str)
            return True
        except ValueError:
            return False

    # ========== 核心7步 ==========
    def step1_prepare_data(self, ocr_result, llm_result):
        """第1步：准备数据"""
        # 直接返回传入的数据
        return ocr_result, llm_result

    def step2_extract_table_data(self, ocr_result, llm_result):
        """
        从数据中提取表格信息
        """
        # 提取OCR表格
        ocr_tables = ocr_result.get('tables_result', [])
        print(f"OCR表格数: {len(ocr_tables)}")

        # 提取LLM表格结构
        llm_tables = []
        if 'tables_structure' in llm_result:
            llm_tables = llm_result['tables_structure'].get('tables', [])
        elif 'tables' in llm_result:
            llm_tables = llm_result['tables']
        print(f"LLM表格结构数: {len(llm_tables)}")

        return ocr_tables, llm_tables

    def step3_merge_ocr_tables(self, ocr_tables, llm_table_info):
        """合并OCR表格 - 修正版"""
        ocr_table_indices = llm_table_info.get('ocr_tables', [])
        print(f"要合并的OCR表格索引: {ocr_table_indices}")

        all_cells = []
        row_offset = 0
        table_boundaries = []

        for idx in ocr_table_indices:
            if idx >= len(ocr_tables):
                print(f"警告: OCR表格索引 {idx} 超出范围")
                continue

            table = ocr_tables[idx]
            cells = table.get('body', [])

            print(f"处理OCR表格{idx}: {len(cells)}个单元格")

            if not cells:
                continue

            # 找出这个表格的实际最大行号
            max_row_in_table = 0
            for cell in cells:
                max_row_in_table = max(max_row_in_table, cell['row_end'])

            # 记录表格边界（在调整行号之前）
            table_boundaries.append({
                'table_idx': idx,
                'start_row': row_offset,  # 调整后的起始行
                'end_row': row_offset + max_row_in_table,  # 调整后的结束行
                'row_count': max_row_in_table + 1,
                'original_max_row': max_row_in_table
            })

            print(f"  表格{idx}: 原始行0-{max_row_in_table} → 合并后行{row_offset}-{row_offset + max_row_in_table}")

            # 调整行号并收集单元格
            for cell in cells:
                adjusted_cell = cell.copy()
                adjusted_cell['row_start'] += row_offset
                adjusted_cell['row_end'] += row_offset
                adjusted_cell['source_table_idx'] = idx
                adjusted_cell['original_row'] = cell['row_start']
                all_cells.append(adjusted_cell)

            # 更新行偏移量：当前表格的最大行号 + 分隔空行
            row_offset += (max_row_in_table + 2)  # +2留一个空行分隔

        print(f"\n表格边界信息（修正后）:")
        for boundary in table_boundaries:
            print(
                f"  表格{boundary['table_idx']}: 行{boundary['start_row']}-{boundary['end_row']} ({boundary['row_count']}行)")

        return {
            'cells': all_cells,
            'headers': llm_table_info.get('headers', {}),
            'table_id': llm_table_info.get('id', '1'),
            'name': llm_table_info.get('name', ''),
            'table_boundaries': table_boundaries
        }

    def step4_create_base_data_table(self, merged_data):
        """
            创建基础数据表格
            检查每行列数是否一致
            """
        cells = merged_data['cells']

        if not cells:
            return []

        # 找出最大行列
        max_row = 0
        max_col = 0
        for cell in cells:
            max_row = max(max_row, cell['row_end'])
            max_col = max(max_col, cell['col_end'])

        num_rows = max_row
        num_cols = max_col

        # 检查每行的列数
        print("检查每行列数:")
        row_col_counts = {}
        for cell in cells:
            row = cell['row_start']
            col_end = cell['col_end']

            if row not in row_col_counts:
                row_col_counts[row] = []
            row_col_counts[row].append(col_end)

        # 找出每行的最大列号
        row_max_cols = {}
        for row, col_ends in row_col_counts.items():
            row_max_cols[row] = max(col_ends)
            print(f"  第{row}行: 最大列索引={row_max_cols[row]}")

        # 检查是否所有行的最大列索引一致
        all_max_cols = list(row_max_cols.values())
        if len(set(all_max_cols)) > 1:
            print(f"警告: 不同行的列数不一致: {all_max_cols}")

        print(f"表格: {num_rows}行 × {num_cols}列")

        # 创建表格
        table = []
        for r in range(num_rows):
            table.append([None] * num_cols)

        # 填充数据
        for cell in cells:
            value = cell['words']
            row_idx = cell['row_start']
            col_idx = cell['col_start']

            if row_idx < num_rows and col_idx < num_cols:
                table[row_idx][col_idx] = value

        return table

    def step5_add_column_headers(self, table, col_headers):
        """
        最终版：LLM列数是最终列数，从右向左确定列
        """
        if not col_headers:
            print("无列标题")
            return table

        if not table:
            print("表格为空")
            return table

        current_cols = len(table[0])  # OCR表格的列数
        target_cols = len(col_headers)  # LLM的最终列数

        # 情况1：OCR列数 > LLM列数（需要删除左侧多余的列）
        if current_cols > target_cols:
            excess_cols = current_cols - target_cols
            print(f"需要删除左侧{excess_cols}列")

            # 删除左侧多余的列（从每行删除）
            for i in range(len(table)):
                # 保留右侧的target_cols列，删除左侧的excess_cols列
                table[i] = table[i][excess_cols:]

            current_cols = target_cols  # 更新列数

        # 情况2：OCR列数 < LLM列数（需要补充左侧空列）
        elif current_cols < target_cols:
            needed_cols = target_cols - current_cols
            print(f"需要补充左侧{needed_cols}列")

            # 在每行左侧补充空列
            for i in range(len(table)):
                table[i] = [None] * needed_cols + table[i]

            current_cols = target_cols  # 更新列数

        # 现在表格列数 = LLM列数
        print(f"调整后表格列数: {current_cols}")

        # 在顶部添加一行用于列标题
        table.insert(0, [None] * current_cols)

        # 从右向左填充列标题
        def change_name(name):
            if name == '>>':
                name = ""
            return name
        col_headers = [change_name(col) for col in col_headers]
        col_headers_copy = col_headers.copy()
        for col in range(current_cols - 1, -1, -1):
            if col_headers_copy:
                table[0][col] = col_headers_copy.pop()

        print("列标题填充完成")
        print(f"第0行（列标题）: {table[0]}")

        return table

    def step6_add_row_headers_intelligent(self, table, row_headers, table_boundaries=None):
        """第6步：智能匹配 - 考虑OCR表格内部顺序"""
        if not table or not row_headers:
            print("表格或行表头为空")
            return table

        print(f"=== 第6步：智能匹配行表头（考虑表格边界） ===")

        # 如果有表格边界信息，使用它
        if table_boundaries:
            print(f"表格边界: {table_boundaries}")
            # 这里添加基于表格边界的匹配逻辑
            return self._match_with_table_boundaries(table, row_headers, table_boundaries)
        else:
            # 原有逻辑（向后兼容）
            return self._match_without_boundaries(table, row_headers)

    def step7_save_to_excel(self, tables_data, output_file):
        """
            将多个表格保存到Excel，每个表格一个Sheet
            tables_data: 列表，每个元素是一个表格的完整数据
            """
        import openpyxl
        from openpyxl.styles import Alignment, Font, Border, Side
        from openpyxl.utils import get_column_letter

        print(f"\n=== 第7步：保存到Excel ===")
        print(f"要保存{len(tables_data)}个表格到: {output_file}")

        # 创建新的工作簿
        wb = openpyxl.Workbook()

        # 删除默认创建的Sheet
        if 'Sheet' in wb.sheetnames:
            default_sheet = wb['Sheet']
            wb.remove(default_sheet)

        # 处理每个表格
        for table_idx, table in enumerate(tables_data):
            if not table:
                print(f"表格{table_idx}为空，跳过")
                continue

            # 创建Sheet名称（Excel限制31字符）
            sheet_name = f"Table{table_idx + 1}"
            if len(sheet_name) > 31:
                sheet_name = sheet_name[:31]

            # 检查Sheet名称是否重复
            original_name = sheet_name
            counter = 1
            while sheet_name in wb.sheetnames:
                sheet_name = f"{original_name}_{counter}"
                counter += 1

            # 创建工作表
            ws = wb.create_sheet(title=sheet_name)

            # 获取表格尺寸
            num_rows = len(table)
            num_cols = len(table[0]) if num_rows > 0 else 0

            print(f"  表格{table_idx + 1}: {num_rows}行 × {num_cols}列 -> Sheet: '{sheet_name}'")

            # 填充数据到Excel
            for r in range(num_rows):
                for c in range(num_cols):
                    cell_value = table[r][c]

                    # Excel行号从1开始，列号从1开始
                    excel_row = r + 1
                    excel_col = c + 1

                    # 写入值
                    ws.cell(row=excel_row, column=excel_col, value=cell_value)

                    # 获取单元格对象设置样式
                    cell_obj = ws.cell(row=excel_row, column=excel_col)

                    # 设置对齐方式
                    if r == 0:  # 第0行是表头
                        cell_obj.font = Font(bold=True)
                        cell_obj.alignment = Alignment(horizontal='center', vertical='center')
                        # 浅灰色背景
                        from openpyxl.styles import PatternFill
                        cell_obj.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                    elif c == 0:  # 第0列是行表头
                        cell_obj.alignment = Alignment(horizontal='left', vertical='center')
                    else:  # 数据单元格
                        cell_obj.alignment = Alignment(horizontal='center', vertical='center')

            # 设置列宽
            for col in range(1, num_cols + 1):
                col_letter = get_column_letter(col)

                # 根据内容调整列宽
                max_length = 0
                for row in range(1, min(num_rows + 1, 50)):  # 只检查前50行
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value:
                        # 计算文本长度（中文算2个字符）
                        text_len = 0
                        for char in str(cell_value):
                            if '\u4e00-\u9fff' in char:  # 中文字符
                                text_len += 2
                            else:
                                text_len += 1
                        max_length = max(max_length, text_len)

                # 设置列宽（最小8，最大50）
                width = min(max(max_length + 2, 8), 50)
                ws.column_dimensions[col_letter].width = width

            # 添加边框
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # 为整个表格添加边框
            for row in ws.iter_rows(min_row=1, max_row=num_rows, min_col=1, max_col=num_cols):
                for cell in row:
                    cell.border = thin_border

            # 冻结首行和首列（方便查看）
            ws.freeze_panes = 'B2'  # 冻结第1行和第1列

        # 如果没有表格，创建一个空Sheet
        if len(wb.sheetnames) == 0:
            ws = wb.create_sheet(title="空表格")
            ws['A1'] = "无表格数据"

        # 保存文件
        try:
            wb.save(output_file)
            print(f"\n✅ Excel文件保存成功: {output_file}")
            print(f"   共保存了 {len(wb.sheetnames)} 个工作表")

            # 显示Sheet列表
            print("   Sheet列表:")
            for i, sheet_name in enumerate(wb.sheetnames, 1):
                ws = wb[sheet_name]
                max_row = ws.max_row
                max_column = ws.max_column
                print(f"     {i}. {sheet_name}: {max_row}行 × {max_column}列")

            return True

        except Exception as e:
            print(f"\n❌ 保存Excel文件失败: {str(e)}")
            return False

    # ========== 完整流程 ==========
    def process_all_tables11(self, ocr_result, llm_result, output_file="output.xlsx", final_output_file="final.xlsx"):
        """
        完整处理流程：第1-7步 - 增强版（支持表格合并）
        """
        print("开始表格重构流程...")

        # 第1步：准备数据
        ocr_result, llm_result = self.step1_prepare_data(ocr_result, llm_result)

        # 第2步：提取表格数据
        ocr_tables, llm_tables = self.step2_extract_table_data(ocr_result, llm_result)

        print("ocr_tables数量:", len(ocr_tables))
        print("llm_tables数量:", len(llm_tables))

        if not ocr_tables or not llm_tables:
            self.log_issue("提取表格数据失败")
            return False

        print(f"OCR表格数: {len(ocr_tables)}")
        print(f"LLM表格结构数: {len(llm_tables)}")

        # === 新增：列标题统一化处理 ===
        print("\n=== 开始列标题统一化处理 ===")
        llm_tables = self._unify_headers_across_tables(llm_tables)


        # === 新增：检测需要合并的表格 ===
        merge_groups = self._detect_tables_to_merge(llm_tables)

        # 分离需要合并的表格和独立处理的表格
        all_indices = set(range(len(llm_tables)))
        merged_indices = set()

        for group in merge_groups:
            for idx in group:
                merged_indices.add(idx)

        independent_indices = all_indices - merged_indices

        print(f"\n表格处理策略:")
        print(f"  需要合并的表格: {merged_indices}")
        print(f"  独立处理的表格: {independent_indices}")

        # 处理每个表格（先处理合并的，再处理独立的）
        all_final_tables = []

        # 1. 处理合并的表格组
        for group_idx, group in enumerate(merge_groups):
            print(f"\n{'=' * 60}")
            print(f"处理合并表格组 {group_idx + 1}/{len(merge_groups)}")
            print(f"包含表格索引: {group}")
            print(f"{'=' * 60}")

            # 暂时跳过合并逻辑（下一步实现）
            # 先按独立表格处理（保持原有行为）
            for table_idx in group:
                llm_table_info = llm_tables[table_idx]
                print(f"\n处理表格 {table_idx + 1}/{len(llm_tables)} (合并组)")

                # 原有独立处理逻辑（暂时保持）
                final_table = self.process_single_table(ocr_tables, llm_table_info)

                if final_table:
                    all_final_tables.append(final_table)

        # 2. 处理独立的表格
        for table_idx in independent_indices:
            print(f"\n处理表格 {table_idx + 1}/{len(llm_tables)} (独立)")

            llm_table_info = llm_tables[table_idx]

            # 原有独立处理逻辑
            final_table = self.process_single_table(ocr_tables, llm_table_info)

            if final_table:
                all_final_tables.append(final_table)

        if not all_final_tables:
            self.log_issue("无表格数据生成")
            return False

        # 第7步：保存到Excel
        if all_final_tables:
            success = self.step7_save_to_excel(all_final_tables, output_file)

            # 新增：生成最终数据Excel
            final_success = self.final_data_converter.batch_convert_tables(
                all_tables_data=all_final_tables,
                all_llm_tables=llm_result.get('tables_structure', {}).get('tables', []),
                output_excel_path=final_output_file,
                bank_name="中国建设银行"  # 可以从图片文件名或其他方式获取
            )

            return success and final_success


        # 输出统计信息
        print(f"\n{'=' * 60}")
        print(f"处理完成统计:")
        print(f"  成功处理表格: {len(all_final_tables)}个")
        print(f"  表格合并组: {len(merge_groups)}组")
        print(f"  警告: {len(self.warnings)}个")
        print(f"  问题: {len(self.issues)}个")
        print(f"  输出文件: {output_file}")
        print(f"{'=' * 60}")

        return success

    def process_single_table(self, ocr_tables, llm_table_info):
        """
        处理单个表格的完整流程
        """
        print(f"\n{'=' * 60}")
        print(f"开始处理表格")
        print(f"{'=' * 60}")

        # 第3步：合并OCR表格
        merged_data = self.step3_merge_ocr_tables(ocr_tables, llm_table_info)
        if not merged_data:
            self.log_issue("表格合并失败")
            return None

        # 第4步：创建基础表格
        base_table = self.step4_create_base_data_table(merged_data)
        if not base_table:
            self.log_issue("创建基础表格失败")
            return None

        # 第5步：添加列标题
        col_headers = merged_data.get('headers', {}).get('cols', [])
        table_with_cols = self.step5_add_column_headers(base_table, col_headers)

        # 第6步：添加行标题
        row_headers = merged_data.get('headers', {}).get('rows', [])
        final_table = self.step6_add_row_headers_intelligent(
            table_with_cols,
            row_headers,
            table_boundaries=merged_data.get('table_boundaries')
        )

        # 第7步：数据验证（只检查数据区域，保留表头）
        validated_table, validation_marks = self.step7_validate_and_mark_preserve_headers(final_table)

        # 第8步：添加特征值标记
        table_with_features = self.step8_add_feature_marks(validated_table, validation_marks)

        return table_with_features


    def process_all_tables(self, ocr_result, llm_result, output_file="output.xlsx", final_output_file="final.xlsx"):
        """
                完整处理流程：第1-7步 - 增强版（支持表格合并）
                """
        print("开始表格重构流程...")

        # 第1步：准备数据
        ocr_result, llm_result = self.step1_prepare_data(ocr_result, llm_result)

        # 第2步：提取表格数据
        ocr_tables, llm_tables = self.step2_extract_table_data(ocr_result, llm_result)

        print("ocr_tables数量:", len(ocr_tables))
        print("llm_tables数量:", len(llm_tables))

        if not ocr_tables or not llm_tables:
            self.log_issue("提取表格数据失败")
            return False

        print(f"OCR表格数: {len(ocr_tables)}")
        print(f"LLM表格结构数: {len(llm_tables)}")

        # === 新增：列标题统一化处理 ===
        print("\n=== 开始列标题统一化处理 ===")
        llm_tables = self._unify_headers_across_tables(llm_tables)

        # === 新增：检测需要合并的表格 ===
        merge_groups = self._detect_tables_to_merge(llm_tables)

        # 分离需要合并的表格和独立处理的表格
        all_indices = set(range(len(llm_tables)))
        merged_indices = set()

        for group in merge_groups:
            for idx in group:
                merged_indices.add(idx)

        independent_indices = all_indices - merged_indices

        print(f"\n表格处理策略:")
        print(f"  需要合并的表格: {merged_indices}")
        print(f"  独立处理的表格: {independent_indices}")

        # 处理每个表格（先处理合并的，再处理独立的）
        all_final_tables = []

        # 1. 处理合并的表格组
        for group_idx, group in enumerate(merge_groups):
            print(f"\n{'=' * 60}")
            print(f"处理合并表格组 {group_idx + 1}/{len(merge_groups)}")
            print(f"包含表格索引: {group}")
            print(f"{'=' * 60}")

            # 暂时跳过合并逻辑（下一步实现）
            # 先按独立表格处理（保持原有行为）
            for table_idx in group:
                llm_table_info = llm_tables[table_idx]
                print(f"\n处理表格 {table_idx + 1}/{len(llm_tables)} (合并组)")

                # 原有独立处理逻辑（暂时保持）
                final_table = self.process_single_table(ocr_tables, llm_table_info)

                if final_table:
                    all_final_tables.append(final_table)

        # 2. 处理独立的表格
        for table_idx in independent_indices:
            print(f"\n处理表格 {table_idx + 1}/{len(llm_tables)} (独立)")

            llm_table_info = llm_tables[table_idx]

            # 原有独立处理逻辑
            final_table = self.process_single_table(ocr_tables, llm_table_info)

            if final_table:
                all_final_tables.append(final_table)

        if not all_final_tables:
            self.log_issue("无表格数据生成")
            return False

        if not all_final_tables:
            self.log_issue("无表格数据生成")
            return False

        # 第7步：保存到Excel
        success = self.step7_save_to_excel(all_final_tables, output_file)

        if not success:
            self.log_issue("保存原始Excel失败")
            return False

        # 新增：生成最终数据Excel
        try:
            print(f"\n开始生成最终数据Excel...")
            print(f"最终数据输出路径: {final_output_file}")

            # 确保输出路径不为空
            if not final_output_file or final_output_file == "final.xlsx":
                # 自动生成最终数据文件名
                import os
                base_name = os.path.splitext(output_file)[0]
                final_output_file = f"{base_name}_final_data.xlsx"
                print(f"自动生成最终数据文件名: {final_output_file}")

            # 获取LLM表格数据
            llm_table_list = []
            if 'tables_structure' in llm_result:
                llm_table_list = llm_result['tables_structure'].get('tables', [])
            elif 'tables' in llm_result:
                llm_table_list = llm_result['tables']

            print(f"要转换的表格数: {len(all_final_tables)}")
            print(f"LLM表格元数据数: {len(llm_table_list)}")

            # 调用转换器
            final_success = self.final_data_converter.batch_convert_tables(
                all_tables_data=all_final_tables,
                all_llm_tables=llm_table_list,
                output_excel_path=final_output_file,
                bank_name="中国建设银行"
            )

            if final_success:
                print(f"✅ 最终数据Excel生成成功: {final_output_file}")
            else:
                print(f"⚠️ 最终数据Excel生成失败")

            # 只要原始表格保存成功就返回True
            return success

        except Exception as e:
            print(f"⚠️ 最终数据转换异常: {str(e)}")
            print(f"原始表格已保存: {output_file}")
            import traceback
            traceback.print_exc()
            return success  # 即使最终数据转换失败，也返回原始表格保存成功

        # 输出统计信息
        print(f"\n{'=' * 60}")
        print(f"处理完成统计:")
        print(f"  成功处理表格: {len(all_final_tables)}个")
        print(f"  表格合并组: {len(merge_groups)}组")
        print(f"  警告: {len(self.warnings)}个")
        print(f"  问题: {len(self.issues)}个")
        print(f"  原始输出文件: {output_file}")
        if final_success:
            print(f"  最终数据文件: {final_output_file}")
        print(f"{'=' * 60}")





# ====================================
# 使用示例
# ====================================

def main():
    # 你的数据
    ocr_result = {}
    llm_result= {'success': True,
                 'image_info': {'image_path': 'E:\\Datas\\base_pros\\DocuVista\\test_codes\\pngs\\123.png',
                                'image_id': 'img_0e6447019e10'}, 'tables_structure': {'tables': [
            {'id': '1', 'ocr_tables': [0],
             'headers': {'cols': ['', 'b>>2024年12月31日', 'c>>2024年9月30日', 'd>>2024年6月30日', 'e>>2024年3月31日'],
                         'rows': ['可用资本（数额）>>1核心一级资本净额', '可用资本（数额）>>2一级资本净额',
                                  '可用资本（数额）>>3资本净额', '风险加权资产（数额）>>4风险加权资产合计',
                                  '风险加权资产（数额）>>4a风险加权资产合计（应用资本底线前）',
                                  '资本充足率>>5核心一级资本充足率（%）',
                                  '资本充足率>>5a核心一级资本充足率（%）（应用资本底线前）',
                                  '资本充足率>>6一级资本充足率（%）', '资本充足率>>6a一级资本充足率（%）（应用资本底线前）',
                                  '资本充足率>>7资本充足率（%）', '资本充足率>>7a资本充足率（%）（应用资本底线前）',
                                  '其他资本要求>>8储备资本要求（%）', '其他资本要求>>9逆周期资本要求（%）',
                                  '其他资本要求>>10系统重要性银行附加资本要求（%）',
                                  '其他资本要求>>11其他各级资本要求（%）（8+9+10）',
                                  '其他资本要求>>12满足最低资本要求后的可用核心一级资本净额占风险加权资产的比例（%）']}}]},
                 'processing_stats': {'analysis_time_sec': 13.35, 'ocr_tables_count': 1, 'visual_tables_count': 1,
                                      'token_usage': {'prompt_tokens': 1155, 'completion_tokens': 421,
                                                      'total_tokens': 1576}}}

    # 创建重构器并处理
    reconstructor = TableReconstructor()
    success = reconstructor.process_all_tables(
        ocr_result=ocr_result,
        llm_result=llm_result,
        output_file="../../test_codes/enhanced_table_analyzer/reconstructed_tables2.xlsx"
    )

    if success:
        print("✅ 表格重构成功！")
    else:
        print("❌ 表格重构失败！")


if __name__ == "__main__":
    main()