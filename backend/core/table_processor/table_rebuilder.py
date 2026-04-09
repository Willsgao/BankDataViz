

import re
import os
from backend.core.table_processor.long_format_converter import FinalDataConverter
from backend.core.table_processor.marked_table_processor import MarkedTableProcessor


# ========== 审核状态常量 ==========
class ReviewStatus:
    """表格审核状态枚举"""
    AUTO = "auto"           # 自动处理完成（无异常）
    PENDING_REVIEW = "pending_review"  # 需要人工审核
    REVIEWED = "reviewed"   # 已人工审核通过
    NEEDS_REPROCESS = "needs_reprocess"  # 需要重新处理


class TableReconstructor:
    """表格重构器：整合7步流程"""

    def __init__(self):
        self.warnings = []
        self.issues = []
        # 新增：记忆前一个表格的列结构
        self.prev_table_header_structure = None
        self.final_data_converter = FinalDataConverter()

        # 新增：独立的标记表格处理器
        self.marked_table_processor = MarkedTableProcessor()

    def _unify_headers_across_tables(self, llm_tables):
        """
        统一相似表格的列标题结构 - 简单实现版
        对于结构相似的相邻表格，如果存在列标题缺失，用前面表格的结构填充
        """
        print("\n=== 列标题统一化处理 ===")

        if len(llm_tables) <= 1:
            print("表格数量≤1，无需统一处理")
            return llm_tables

        for i in range(1, len(llm_tables)):
            current_table = llm_tables[i]
            prev_table = llm_tables[i - 1]

            # 获取列标题
            current_cols = current_table.get('headers', {}).get('cols', [])
            prev_cols = prev_table.get('headers', {}).get('cols', [])

            # 只有列数相同时才考虑统一
            if len(current_cols) == len(prev_cols):
                # 检查是否存在空列标题
                new_cols = []
                need_unify = False

                for idx, (prev_col, curr_col) in enumerate(zip(prev_cols, current_cols)):
                    # 如果当前列标题为空，而前一表格的对应列有内容
                    if (not curr_col or str(curr_col).strip() == '') and prev_col and str(prev_col).strip() != '':
                        new_cols.append(prev_col)
                        need_unify = True
                        print(f"  表格{i}第{idx}列: 空白 → 填充为 '{prev_col}'")
                    else:
                        new_cols.append(curr_col)

                # 如果需要统一，更新表格
                if need_unify:
                    llm_tables[i]['headers']['cols'] = new_cols
                    print(f"  表格{i}统一化完成")

        return llm_tables

    def _analyze_column_structure(self, col_headers):
        """
        分析列标题的结构特征
        返回: 结构特征字典
        """
        if not col_headers:
            return {"empty": True}

        structure = {
            "count": len(col_headers),
            "first_empty": col_headers[0] == '' if col_headers else False,
            "year_count": sum(
                1 for col in col_headers if col and any(year in str(col) for year in ['年', '202', '201'])),
            "change_col": any(col and '变化' in str(col) for col in col_headers),
            "pattern": []
        }

        # 分析列标题模式
        for col in col_headers:
            if not col or str(col).strip() == '':
                structure["pattern"].append("empty")
            elif any(year in str(col) for year in ['2024', '2023', '2022']):
                structure["pattern"].append("year")
            elif '变化' in str(col) or 'chang' in str(col).lower():
                structure["pattern"].append("change")
            else:
                structure["pattern"].append("other")

        return structure

    def _are_similar_column_structures(self, struct1, struct2):
        """
        判断两个列结构是否相似
        """
        if not struct1 or not struct2:
            return False

        # 如果都是空结构
        if struct1.get("empty") and struct2.get("empty"):
            return False

        # 列数相同是关键
        if struct1.get("count") != struct2.get("count"):
            return False

        # 年份列数量相似
        if abs(struct1.get("year_count", 0) - struct2.get("year_count", 0)) > 1:
            return False

        # 是否都有变化列
        if struct1.get("change_col") != struct2.get("change_col"):
            return False

        # 模式相似度（至少60%相同）
        pattern1 = struct1.get("pattern", [])
        pattern2 = struct2.get("pattern", [])

        if len(pattern1) != len(pattern2):
            return False

        same_count = sum(1 for p1, p2 in zip(pattern1, pattern2) if p1 == p2)
        similarity = same_count / len(pattern1) if pattern1 else 0

        return similarity >= 0.6

    # ========== 工具函数 ==========
    def log_warning(self, message):
        """记录警告"""
        self.warnings.append(message)
        print(f"⚠️ {message}")

    def log_issue(self, message):
        """记录问题"""
        self.issues.append(message)
        print(f"❓ {message}")

    def clean_text_for_matching(self, text):
        """清理文本用于匹配"""
        if not text:
            return ""
        text = str(text)
        # 替换换行、空格
        text = text.replace('\n', '').replace('\r', '').replace(' ', '')
        # 去掉特殊符号，保留中文、数字、字母
        cleaned = ''.join(c for c in text if c.isalnum() or '\u4e00-\u9fff' in c)
        return cleaned

    def calculate_similarity(self, text1, text2):
        """计算两个文本的相似度"""
        t1 = self.clean_text_for_matching(text1)
        t2 = self.clean_text_for_matching(text2)

        if not t1 or not t2:
            return 0

        # 完全相等
        if t1 == t2:
            return 1.0

        # 包含关系
        if t1 in t2 or t2 in t1:
            return 0.9

        # 公共字符比例
        common_chars = set(t1) & set(t2)
        if not common_chars:
            return 0

        # 计算Jaccard相似度
        union_chars = set(t1) | set(t2)
        return len(common_chars) / len(union_chars)

    def calculate_similarity_v2(self, text1, text2):
        """增强版相似度计算（考虑包含关系的覆盖率）"""
        t1 = self.clean_text_for_matching(text1)
        t2 = self.clean_text_for_matching(text2)

        if not t1 or not t2:
            return 0

        # 1. 完全相等
        if t1 == t2:
            return 1.0

        # 2. 包含关系，但要求高覆盖率
        if t1 in t2:
            coverage = len(t1) / len(t2)
            # 覆盖度>80%才给高分，否则给低分
            # 避免"经营活动"匹配"每股经营活动"
            return 0.9 if coverage > 0.8 else 0.2

        if t2 in t1:
            coverage = len(t2) / len(t1)
            return 0.9 if coverage > 0.8 else 0.2

        # 3. 公共字符比例（原有逻辑）
        common_chars = set(t1) & set(t2)
        if not common_chars:
            return 0

        # 计算Jaccard相似度
        union_chars = set(t1) | set(t2)
        similarity = len(common_chars) / len(union_chars)

        return similarity

    def _detect_tables_to_merge(self, llm_tables):
        """
        检测哪些LLM表格应该合并处理
        返回：需要合并的表格组列表，每个组包含应该合并的表格索引
        """
        print("\n=== 检测可合并的表格结构 ===")

        if len(llm_tables) <= 1:
            return []

        # 检查每个表格的列结构
        table_col_structures = []

        for i, table in enumerate(llm_tables):
            col_headers = table.get('headers', {}).get('cols', [])
            # 清理列标题（去掉空值和特殊字符）
            cleaned_cols = []
            for col in col_headers:
                if col and str(col).strip():
                    # 提取文本内容（去掉>>等层级标记）
                    text = str(col).split('>>')[-1].strip()
                    if text:
                        cleaned_cols.append(text)

            table_info = {
                'index': i,
                'col_count': len(col_headers),
                'cleaned_cols': cleaned_cols,
                'original_cols': col_headers,
                'row_count': len(table.get('headers', {}).get('rows', []))
            }
            table_col_structures.append(table_info)

            print(f"表格{i}: {table_info['row_count']}行 × {table_info['col_count']}列")
            print(f"  列标题: {cleaned_cols[:5]}...")

        # 检测相似的列结构（用于合并）
        merge_groups = []
        processed_indices = set()

        for i in range(len(table_col_structures)):
            if i in processed_indices:
                continue

            current_table = table_col_structures[i]
            merge_group = [i]

            # 寻找与当前表格列结构相似的表格
            for j in range(i + 1, len(table_col_structures)):
                if j in processed_indices:
                    continue

                next_table = table_col_structures[j]

                # 合并条件1：列数相同
                if current_table['col_count'] != next_table['col_count']:
                    continue

                # 合并条件2：列标题相似度检查（简化版）
                is_similar = self._check_columns_similarity(
                    current_table['cleaned_cols'],
                    next_table['cleaned_cols']
                )

                if is_similar:
                    merge_group.append(j)
                    processed_indices.add(j)

            if len(merge_group) > 1:
                merge_groups.append(merge_group)
                processed_indices.add(i)

        return merge_groups

    def _check_columns_similarity(self, cols1, cols2):
        """
        检查两个列标题列表是否相似
        简化版：检查前几列是否包含相同的年份/数字模式
        """
        if not cols1 or not cols2:
            return False

        # 如果列数不同，肯定不相似
        if len(cols1) != len(cols2):
            return False

        # 检查是否包含典型的年份模式
        year_patterns = ['年', '202', '201', '月', '季度', 'Q']

        def contains_year_pattern(text):
            if not text:
                return False
            text_str = str(text)
            for pattern in year_patterns:
                if pattern in text_str:
                    return True
            return False

        # 统计两个列表中包含年份模式的列数
        cols1_year_count = sum(1 for col in cols1 if contains_year_pattern(col))
        cols2_year_count = sum(1 for col in cols2 if contains_year_pattern(col))

        # 如果都有年份列，认为可能是相似的表格结构
        if cols1_year_count > 0 and cols2_year_count > 0:
            return True

        # 检查前两列是否相似（对于财务报表，通常第一列是项目，后面是年份）
        similarity_threshold = 0.7
        similar_count = 0

        for idx in range(min(len(cols1), len(cols2))):
            if idx == 0:
                # 第一列通常是项目名称，允许不同
                continue

            sim = self.calculate_similarity(cols1[idx], cols2[idx])
            if sim > similarity_threshold:
                similar_count += 1

        # 如果有超过一半的非首列相似，认为表格结构相似
        if len(cols1) > 1:
            return similar_count / (len(cols1) - 1) > 0.5

        return False

    def _match_with_table_boundaries(self, table, row_headers, table_boundaries):
        """基于表格边界的匹配算法"""

        # 1. 为每个OCR表格建立搜索空间
        search_spaces = []
        for boundary in table_boundaries:
            start_row = boundary['start_row']
            end_row = boundary['end_row']
            # 转换为1-based索引（表格数据是0-based）
            search_space = list(range(start_row, end_row + 1))
            search_spaces.append({
                'table_idx': boundary['table_idx'],
                'rows': search_space,
                'next_row_idx': 0,  # 在这个表格内下一个可用的行索引
                'used_rows': set()  # 已使用的行
            })
            print(f"  表格{boundary['table_idx']}: 行{start_row}-{end_row}")

        # 2. 确定搜索列（与原有逻辑一致）
        left_empty_cols = 0
        for title in table[0]:
            if not title or str(title).strip() == '':
                left_empty_cols += 1
            else:
                break

        search_columns = list(range(min(3, len(table[0]))))
        if left_empty_cols > 0:
            search_columns = list(range(left_empty_cols))

        print(f"搜索列: {search_columns}")

        # 3. 按LLM顺序匹配
        assignments = {}

        for header_idx, llm_header in enumerate(row_headers):
            target_text = llm_header.split('>>')[-1] if '>>' in llm_header else llm_header

            print(f"  表头[{header_idx}]: '{target_text[:30]}...'")

            best_match = None
            best_score = 0
            best_table_idx = -1

            # 在每个表格的搜索空间中寻找最佳匹配
            for table_info in search_spaces:
                for row_idx in table_info['rows']:
                    if row_idx in table_info['used_rows']:
                        continue

                    # 计算匹配分数
                    row_score = 0
                    for col_idx in search_columns:
                        if col_idx < len(table[row_idx]):
                            cell_val = table[row_idx][col_idx]
                            if cell_val:
                                score = self.calculate_similarity_v2(target_text, str(cell_val))
                                if score > row_score:
                                    row_score = score

                    if row_score > best_score:
                        best_score = row_score
                        best_match = row_idx
                        best_table_idx = table_info['table_idx']

            # 4. 如果找到高质量匹配，分配该行
            if best_match and best_score > 0.6:
                assignments[header_idx] = best_match
                # 标记该行为已使用
                for table_info in search_spaces:
                    if best_match in table_info['rows']:
                        table_info['used_rows'].add(best_match)
                print(f"    → 匹配到表格{best_table_idx}的行{best_match} (分数:{best_score:.2f})")
            else:
                # 5. 按表格顺序分配下一个可用行
                assigned = False
                for table_info in search_spaces:
                    # 在这个表格内按顺序找下一个可用行
                    while table_info['next_row_idx'] < len(table_info['rows']):
                        row_idx = table_info['rows'][table_info['next_row_idx']]
                        if row_idx not in table_info['used_rows']:
                            assignments[header_idx] = row_idx
                            table_info['used_rows'].add(row_idx)
                            table_info['next_row_idx'] += 1
                            assigned = True
                            print(f"    → 顺序分配到表格{table_info['table_idx']}的行{row_idx}")
                            break
                        table_info['next_row_idx'] += 1

                    if assigned:
                        break

                if not assigned:
                    print(f"    → 无法分配")
                    assignments[header_idx] = -1

        # 6. 填充表格（与原有逻辑一致）
        for row_idx in range(1, len(table)):
            table[row_idx][0] = None

        filled_count = 0
        for header_idx, row_idx in assignments.items():
            if row_idx != -1 and row_idx < len(table):
                table[row_idx][0] = row_headers[header_idx]
                filled_count += 1

        print(f"\n匹配结果: {filled_count}/{len(row_headers)} 个行表头已填充")
        return table


    def _analyze_cell_type(self, cell_value):
        """
        分析单元格类型
        返回: "blank", "text", "std_num", "minor_num", "error_num"
        """
        if cell_value is None:
            return "blank"

        text = str(cell_value).strip()

        # 1. 空白单元格
        if text == "":
            return "blank"

        # 2. 检查是否为纯文本（不包含数字）
        if not any(c.isdigit() for c in text):
            return "text"

        # ========== 新增：检查%号前是否有逗号 ==========
        if '%' in text:
            # 找到%号的位置
            percent_pos = text.find('%')
            # 检查%号前面是否有逗号
            if percent_pos > 0 and ',' in text[:percent_pos]:
                # %号前面有逗号，这是错误格式（如"0,82%"）
                # 但是要排除类似"1,234.56%"的情况
                # 检查逗号是否在正确的位置
                before_percent = text[:percent_pos]
                if '.' in before_percent:
                    # 有小数点，检查逗号和小数点的关系
                    dot_pos = before_percent.find('.')
                    for i, char in enumerate(before_percent):
                        if char == ',' and i > dot_pos:
                            # 逗号在小数点后面，这是错误格式
                            return "error_num"
                else:
                    # 没有小数点，%号前不应该有逗号
                    return "error_num"

        # 3. 检查负值格式
        cleaned = text

        # 处理括号负数
        is_parenthesis_negative = (
                cleaned.startswith('(') and
                cleaned.endswith(')') and
                '(' not in cleaned[1:-1] and
                ')' not in cleaned[1:-1]
        )

        if is_parenthesis_negative:
            # 检查括号内是否还有负号
            inside = cleaned[1:-1].strip()
            if inside.startswith('-') or inside.endswith('-'):
                return "error_num"  # 如"(-450)"或"(450-)"，错误格式
            cleaned = '-' + inside

        # 4. 检查多个负号或负号位置错误
        if cleaned.count('-') > 1:
            return "error_num"  # 如"--450"

        if '-' in cleaned and not cleaned.startswith('-'):
            return "error_num"  # 如"450-"

        # 5. 移除负号、逗号、小数点、%号用于数值转换
        test_str = cleaned
        if '-' in test_str:
            test_str = test_str.replace('-', '', 1)

        # 记录原始符号用于格式检查
        has_percent = '%' in test_str
        dot_count = test_str.count('.')
        comma_count = test_str.count(',')

        # 用于数值转换的字符串
        numeric_test = test_str.replace(',', '').replace('.', '').replace('%', '')

        # 6. 尝试转换为数值
        try:
            # 如果转换失败，说明不是有效数值
            float(numeric_test)
        except ValueError:
            return "text"  # 包含数字但不是有效数值

        # 7. 格式检查
        # 检查%号
        if has_percent:
            if not cleaned.endswith('%'):
                return "error_num"  # %不在末尾
            if cleaned.count('%') > 1:
                return "error_num"  # 多个%

        # 检查小数点
        if dot_count > 1:
            return "error_num"  # 多个小数点

        # 检查逗号位置（如果有小数点）
        if dot_count == 1 and comma_count > 0:
            dot_pos = cleaned.find('.')
            # 检查逗号是否都在小数点前面
            for i, char in enumerate(cleaned):
                if char == ',' and i > dot_pos:
                    return "error_num"  # 逗号在小数点后面

        # 8. 检查是否为标准格式
        # 标准格式：有逗号时必须在正确位置，%在末尾等
        if self._is_standard_format(cleaned):
            return "std_num"
        else:
            return "minor_num"

    def _is_standard_format(self, text):
        """
        判断是否为标准数值格式
        标准格式要求：
        1. 千分位逗号格式正确（每3位一个逗号）
        2. 如果有小数点，后面通常有数字
        3. 没有多余的空格或其他字符
        """
        if not text:
            return False

        # 移除负号和%号
        test_text = text
        if test_text.startswith('-'):
            test_text = test_text[1:]
        if test_text.endswith('%'):
            test_text = test_text[:-1]

        # 检查逗号格式
        if ',' in test_text:
            # 分割整数和小数部分
            if '.' in test_text:
                int_part, dec_part = test_text.split('.', 1)
            else:
                int_part, dec_part = test_text, ""

            # 检查整数部分的逗号格式
            groups = int_part.split(',')
            # 第一组可以是1-3位
            if not (1 <= len(groups[0]) <= 3):
                return False

            # 后面的每组必须正好是3位
            for group in groups[1:]:
                if len(group) != 3:
                    return False

        return True

    def _fill_blank_cells(self, cell_types):
        """
        填充空白单元格的类型（根据同行/同列的众数类型）
        """
        if not cell_types:
            return

        num_rows = len(cell_types)
        num_cols = len(cell_types[0]) if num_rows > 0 else 0

        for r in range(num_rows):
            for c in range(num_cols):
                if cell_types[r][c] == "blank":
                    # 查找同行和同列的非空白类型
                    row_types = []
                    col_types = []

                    # 同行
                    for cc in range(num_cols):
                        if cc != c and cell_types[r][cc] != "blank":
                            row_types.append(cell_types[r][cc])

                    # 同列
                    for rr in range(num_rows):
                        if rr != r and cell_types[rr][c] != "blank":
                            col_types.append(cell_types[rr][c])

                    # 合并所有类型
                    all_types = row_types + col_types

                    if all_types:
                        # 找出众数类型
                        from collections import Counter
                        type_counts = Counter(all_types)
                        most_common = type_counts.most_common(1)

                        if most_common:
                            cell_types[r][c] = most_common[0][0]

    def _determine_row_column_mark(self, cell_types, label):
        """
        根据单元格类型列表确定行或列的标记
        """
        # 过滤掉空白单元格
        non_blank_types = [t for t in cell_types if t != "blank"]

        if not non_blank_types:
            return 0  # 全空白，视为纯文本

        # 检查是否有文本
        has_text = "text" in non_blank_types
        has_std_num = "std_num" in non_blank_types
        has_minor_num = "minor_num" in non_blank_types
        has_error_num = "error_num" in non_blank_types

        # 4. 混合类型（包含文本和任何数值）
        if has_text and (has_std_num or has_minor_num or has_error_num):
            return 4

        # 3. 很可能错误的数值
        if has_error_num:
            return 3

        # 现在只包含数值类型
        # 2. 有小问题的数值（有minor_num，可能也有std_num）
        if has_minor_num:
            return 2

        # 1. 完全正确的数值（全是std_num）
        if has_std_num and not has_minor_num and not has_error_num:
            return 1

        # 0. 纯文本
        if not has_std_num and not has_minor_num and not has_error_num:
            return 0

        # 默认返回0
        return 0

    def _count_marks(self, marks_list):
        """
        统计标记数量
        """
        from collections import Counter
        counter = Counter(marks_list)

        result = []
        for mark in range(5):
            count = counter.get(mark, 0)
            if count > 0:
                result.append(f"标记{mark}:{count}个")

        return ", ".join(result)

    def _fix_llm_table_references(self, ocr_result, llm_result):
        """修正LLM引用的表格索引，确保引用有数据的表格"""
        ocr_tables = ocr_result.get('tables_result', [])

        # 找出有数据的OCR表格
        valid_ocr_indices = []
        for i, table in enumerate(ocr_tables):
            if table.get('body') and len(table['body']) > 0:
                valid_ocr_indices.append(i)

        print(f"有效OCR表格索引: {valid_ocr_indices}")

        if not valid_ocr_indices:
            print("⚠️ 没有找到有数据的OCR表格")
            return

        # 修正每个LLM表格的引用
        llm_tables = llm_result.get('tables_structure', {}).get('tables', [])
        for llm_table in llm_tables:
            original_refs = llm_table.get('ocr_tables', [])

            # 检查引用的表格是否有数据
            valid_refs = []
            for ref_idx in original_refs:
                if ref_idx in valid_ocr_indices:
                    valid_refs.append(ref_idx)

            # 如果没有有效的引用，使用第一个有效表格
            if not valid_refs and valid_ocr_indices:
                valid_refs = [valid_ocr_indices[0]]
                print(f"⚠️ 表格'{llm_table.get('name')}'的引用无效，修正为: {valid_refs}")

            llm_table['ocr_tables'] = valid_refs

    # ========== 核心7步 ==========
    def step1_prepare_data(self, ocr_result, llm_result):
        """第1步：准备数据"""
        # 直接返回传入的数据
        return ocr_result, llm_result

    def step2_extract_table_data(self, ocr_result, llm_result):
        """
        从数据中提取表格信息
        """
        # 提取OCR表格
        ocr_tables = ocr_result.get('tables_result', [])
        print(f"OCR表格数: {len(ocr_tables)}")

        # 提取LLM表格结构
        llm_tables = []
        if 'tables_structure' in llm_result:
            llm_tables = llm_result['tables_structure'].get('tables', [])
        elif 'tables' in llm_result:
            llm_tables = llm_result['tables']
        print(f"LLM表格结构数: {len(llm_tables)}")

        return ocr_tables, llm_tables

    def step3_merge_ocr_tables(self, ocr_tables, llm_table_info):
        """合并OCR表格 - 修正版"""
        ocr_table_indices = llm_table_info.get('ocr_tables', [])
        print(f"要合并的OCR表格索引: {ocr_table_indices}")

        # 🔥 添加调试：查看整个OCR数据结构
        print(f"\n🔍 OCR表格数据结构调试:")
        for i, table in enumerate(ocr_tables):
            print(f"OCR表格{i} 类型: {type(table)}")
            if isinstance(table, dict):
                print(f"  键: {list(table.keys())}")
                if 'body' in table:
                    print(f"  body类型: {type(table['body'])}, 长度: {len(table['body']) if table['body'] else 0}")
                if 'cell_set' in table:
                    print(f"  cell_set长度: {len(table['cell_set']) if table['cell_set'] else 0}")



        all_cells = []
        row_offset = 0
        table_boundaries = []

        for idx in ocr_table_indices:
            if idx >= len(ocr_tables):
                print(f"警告: OCR表格索引 {idx} 超出范围")
                continue

            table = ocr_tables[idx]
            cells = table.get('body', [])

            print(f"处理OCR表格{idx}: {len(cells)}个单元格")

            if not cells:
                continue

            # 找出这个表格的实际最大行号
            max_row_in_table = 0
            for cell in cells:
                max_row_in_table = max(max_row_in_table, cell['row_end'])

            # 记录表格边界（在调整行号之前）
            table_boundaries.append({
                'table_idx': idx,
                'start_row': row_offset,  # 调整后的起始行
                'end_row': row_offset + max_row_in_table,  # 调整后的结束行
                'row_count': max_row_in_table + 1,
                'original_max_row': max_row_in_table
            })

            print(f"  表格{idx}: 原始行0-{max_row_in_table} → 合并后行{row_offset}-{row_offset + max_row_in_table}")

            # 调整行号并收集单元格
            for cell in cells:
                adjusted_cell = cell.copy()
                adjusted_cell['row_start'] += row_offset
                adjusted_cell['row_end'] += row_offset
                adjusted_cell['source_table_idx'] = idx
                adjusted_cell['original_row'] = cell['row_start']
                all_cells.append(adjusted_cell)

            # 更新行偏移量：当前表格的最大行号 + 分隔空行
            row_offset += (max_row_in_table + 2)  # +2留一个空行分隔

        print(f"\n表格边界信息（修正后）:")
        for boundary in table_boundaries:
            print(
                f"  表格{boundary['table_idx']}: 行{boundary['start_row']}-{boundary['end_row']} ({boundary['row_count']}行)")

        return {
            'cells': all_cells,
            'headers': llm_table_info.get('headers', {}),
            'table_id': llm_table_info.get('id', '1'),
            'name': llm_table_info.get('name', ''),
            'table_boundaries': table_boundaries
        }

    def step4_create_base_data_table(self, merged_data):
        """
            创建基础数据表格
            检查每行列数是否一致
            """
        cells = merged_data['cells']

        if not cells:
            return []

        # 找出最大行列
        max_row = 0
        max_col = 0
        for cell in cells:
            max_row = max(max_row, cell['row_end'])
            max_col = max(max_col, cell['col_end'])

        num_rows = max_row
        num_cols = max_col

        # 检查每行的列数
        row_col_counts = {}
        for cell in cells:
            row = cell['row_start']
            col_end = cell['col_end']

            if row not in row_col_counts:
                row_col_counts[row] = []
            row_col_counts[row].append(col_end)

        # 找出每行的最大列号
        row_max_cols = {}
        for row, col_ends in row_col_counts.items():
            row_max_cols[row] = max(col_ends)

        # 检查是否所有行的最大列索引一致
        all_max_cols = list(row_max_cols.values())

        # 创建表格
        table = []
        for r in range(num_rows):
            table.append([None] * num_cols)

        # 填充数据
        for cell in cells:
            value = cell['words']
            row_idx = cell['row_start']
            col_idx = cell['col_start']

            if row_idx < num_rows and col_idx < num_cols:
                table[row_idx][col_idx] = value

        return table


    def step5_add_column_headers(self, table, col_headers, ocr_cells=None):
        """
        优化版：清理和验证列标题，调整列数匹配
        """
        if not col_headers:
            print("无列标题")
            return table

        if not table:
            print("表格为空")
            return table

        current_cols = len(table[0])  # OCR表格的列数
        target_cols = len(col_headers)  # LLM的最终列数

        print(f"OCR列数: {current_cols}, LLM列数: {target_cols}")

        # 1. 清理列标题
        cleaned_headers = []
        for header in col_headers:
            # 移除空白和空值
            if not header or str(header).strip() == '':
                cleaned_headers.append("")
                continue

            header_str = str(header).strip()

            # 标准化分隔符
            header_str = header_str.replace('><', '>>')
            header_str = header_str.replace('＞＞', '>>')

            # 修复可能的分隔符错误
            if '>' in header_str and '>' in header_str[1:]:
                # 处理类似 "b>2024年12月31日" 的情况
                parts = header_str.split('>')
                if len(parts) == 2:
                    header_str = f"{parts[0]}>>{parts[1]}"

            # 移除多余空格
            header_str = '>>'.join([part.strip() for part in header_str.split('>>')])

            # 特别处理：如果标题就是">>"，设为空
            if header_str == '>>' or header_str == '>':
                header_str = ""

            cleaned_headers.append(header_str)

        # 2. 验证列标题有效性
        self._validate_cleaned_headers(cleaned_headers)

        # 3. 调整列数匹配
        # 情况1：OCR列数 > LLM列数（需要删除多余的空列/分隔列）
        if current_cols > target_cols:
            excess_cols = current_cols - target_cols
            print(f"需要删除{excess_cols}个多余列")

            # [重构-20260403] 智能空列检测：
            # 遍历所有列，找到真正完全为空的列（仅含None/空字符串/纯空白）
            # 注意：含中文的文本（年份"2024年1-6月"、中文标签）不是"空"，不能删
            empty_col_indices = []
            for col_idx in range(current_cols):
                is_empty = True
                for row_idx in range(1, min(len(table), 20)):
                    if col_idx < len(table[row_idx]):
                        cell = table[row_idx][col_idx]
                        if cell is not None and str(cell).strip() != '':
                            # 检查是否含中文，含中文 = 标签/年份文本，不是空
                            if any('\u4e00' <= ch <= '\u9fff' for ch in str(cell)):
                                is_empty = False  # 含中文，是有效列
                                break
                            # 纯数字/带逗号的数字 = 有效数据列
                            clean = str(cell).replace(',', '').replace(' ', '')
                            if clean.replace('.', '').replace('-', '').replace('(', '').replace(')', '').replace(' ', '').isdigit():
                                is_empty = False
                                break
                            # 其他非空字符串（不含中文非数字）—— 也视为有效
                            is_empty = False
                            break
                if is_empty:
                    empty_col_indices.append(col_idx)

            print(f"  [step5_safety] 检测到空列索引: {empty_col_indices} (共{len(empty_col_indices)}个)")

            if len(empty_col_indices) >= excess_cols:
                # 有足够的空列可以删除
                # [修复-20260404] 优先删除右侧空列（OCR span 产生的冗余列通常在右侧）
                cols_to_delete = sorted(empty_col_indices, reverse=True)[:excess_cols]
                print(f"  [step5_safety] 删除空列索引(优先右侧): {cols_to_delete}")
                for i in range(len(table)):
                    # 从右向左删除，避免索引偏移
                    for col_idx in sorted(cols_to_delete, reverse=True):
                        if col_idx < len(table[i]):
                            del table[i][col_idx]
                current_cols = target_cols
                print(f"  [step5_safety] 空列删除完成，剩余{current_cols}列")
            elif len(empty_col_indices) > 0 and len(empty_col_indices) < excess_cols:
                # [修复-20260404] 部分空列：有但不够，尝试用 OCR span 信息找到更多冗余列
                still_need = excess_cols - len(empty_col_indices)
                print(f"  [step5_safety] 部分空列不足（找到{len(empty_col_indices)}个 < 需删除{excess_cols}个），尝试OCR span补充...")

                span_redundant = []
                if ocr_cells:
                    # 找到被 span 覆盖的冗余列：该列在数据行(row 2+)全为空，
                    # 但被某个 span>1 的 OCR cell 覆盖（即该列是另一个数据列的 span 尾部）
                    for col_idx in range(current_cols):
                        if col_idx in empty_col_indices:
                            continue  # 已经标记为空列
                        # 检查数据行是否全空
                        data_rows_empty = True
                        for row_idx in range(2, min(len(table), 20)):
                            if col_idx < len(table[row_idx]):
                                cell = table[row_idx][col_idx]
                                if cell is not None and str(cell).strip() != '':
                                    data_rows_empty = False
                                    break
                        if not data_rows_empty:
                            continue  # 数据行有值，不是冗余列
                        # 检查该列是否被某个 span>1 的 cell 覆盖
                        covered_by_span = False
                        for cell in ocr_cells:
                            cs = cell['col_start']
                            ce = cell['col_end']
                            if ce - cs > 1 and cs <= col_idx < ce:
                                covered_by_span = True
                                break
                        if covered_by_span:
                            span_redundant.append(col_idx)
                            print(f"    col {col_idx}: 数据行全空，被OCR span覆盖 -> 冗余列")

                    if span_redundant and len(span_redundant) >= still_need:
                        # 优先删除右侧冗余列
                        extra_delete = sorted(span_redundant, reverse=True)[:still_need]
                        all_delete = sorted(empty_col_indices + extra_delete, reverse=True)
                        print(f"  [step5_safety] 补充删除span冗余列: {extra_delete}")
                        print(f"  [step5_safety] 最终删除索引: {all_delete}")
                        for i in range(len(table)):
                            for col_idx in all_delete:
                                if col_idx < len(table[i]):
                                    del table[i][col_idx]
                        current_cols = target_cols
                        print(f"  [step5_safety] 空列+span冗余列删除完成，剩余{current_cols}列")
                    else:
                        print(f"  [WARN step5_safety] span冗余列也不足（找到{len(span_redundant)}个 < 需{still_need}个），跳过删除")
                else:
                    print(f"  [WARN step5_safety] 部分空列不足（找到{len(empty_col_indices)}个 < 需删除{excess_cols}个），跳过删除")
            elif len(empty_col_indices) == 0 and excess_cols > 0:
                # [重构-20260404] Primary 完全失败（找到0个空列），尝试 fallback
                # 策略1（有 OCR cells）：利用 span 信息找到被同一 cell 覆盖的冗余列
                # 策略2：查找所有行都完全为 None 的列
                # 策略3：查找内容完全重复的列对
                print(f"  [WARN step5_safety] 空列为0（primary 失败），执行 fallback...")

                col_to_delete = None

                # 策略1：利用 OCR span 信息
                if ocr_cells and col_to_delete is None:
                    # 统计每列被不同 cell 的 span 覆盖的次数
                    # 关键洞察：真正需要删除的冗余列，是被不同数据源（不同 cell）的 span
                    # 同时覆盖的列（比如 col2 被 2025-header span 和 2024-data span 覆盖）
                    span_covering_cells = {}  # col_idx -> set of (row_start, col_start, col_end) tuples
                    for col_idx in range(current_cols):
                        span_covering_cells[col_idx] = set()
                    for cell in ocr_cells:
                        col_start = cell['col_start']
                        col_end = cell['col_end']
                        if col_end > col_start:  # span 宽度 >= 1
                            cell_key = (cell['row_start'], col_start, col_end)
                            for c in range(col_start, col_end):
                                if c in span_covering_cells:
                                    span_covering_cells[c].add(cell_key)

                    # 统计每列被不同 cell span 覆盖的次数
                    coverage_count = {c: len(cells) for c, cells in span_covering_cells.items()}
                    print(f"  [fallback] 各列被不同span覆盖数: {coverage_count}")

                    # 被覆盖次数最多的列最可能是冗余列
                    max_cov = max(coverage_count.values())
                    candidates = [c for c, v in coverage_count.items() if v == max_cov]
                    print(f"  [fallback] 最大覆盖列(候选): {candidates}")

                    # 进一步筛选：在候选列中，找到在数据行中被覆盖但单列cell最少的列
                    # 即该列的数据主要来自其他列的span溢出，而非独立数据
                    best_candidate = None
                    best_score = -1  # 越小越可能是冗余列
                    independent_values_map = {}
                    span_overlap_map = {}
                    for col_idx in candidates:
                        # 计算该列有多少独立（非span溢出）的有值cell
                        independent_values = 0
                        for cell in ocr_cells:
                            if cell['col_start'] == col_idx and cell['col_end'] == col_idx + 1:
                                # 单列 cell
                                if cell['words'] and str(cell['words']).strip():
                                    independent_values += 1
                        # 该列的总有值cell数
                        total_values = 0
                        for row_idx in range(len(table)):
                            if col_idx < len(table[row_idx]):
                                cell = table[row_idx][col_idx]
                                if cell is not None and str(cell).strip():
                                    total_values += 1
                        # 差值 = 独立值 - 来自其他span的值，越小越可能是冗余
                        overlap_values = total_values - independent_values
                        score = independent_values  # 独立值越少，越可能冗余
                        independent_values_map[col_idx] = independent_values
                        span_overlap_map[col_idx] = overlap_values
                        print(f"  [fallback] col {col_idx}: 独立值={independent_values}, 总值={total_values}, span溢出值={overlap_values}")
                        if score < best_score or best_candidate is None:
                            best_score = score
                            best_candidate = col_idx

                    if best_candidate is not None:
                        # 在独立值同为最少的候选列中，选择span溢出值最少的
                        # （该列本身数据最少，更可能是冗余列）
                        min_independent = min(independent_values_map.values())
                        lowest_candidates = [c for c in candidates
                                           if independent_values_map.get(c, 999) == min_independent]
                        best = min(lowest_candidates, key=lambda c: span_overlap_map.get(c, 0))
                        col_to_delete = best

                # 策略2：查找所有行都完全为 None 的列
                if col_to_delete is None:
                    for col_idx in range(current_cols):
                        all_none = True
                        for row_idx in range(len(table)):
                            cell = table[row_idx][col_idx] if col_idx < len(table[row_idx]) else None
                            if cell is not None and str(cell).strip() != '':
                                all_none = False
                                break
                        if all_none:
                            col_to_delete = col_idx
                            print(f"  [fallback] 发现全空列: col {col_idx}")
                            break

                # 策略3：查找内容完全重复的列对
                if col_to_delete is None:
                    for col_a in range(current_cols):
                        if col_to_delete is not None:
                            break
                        for col_b in range(col_a + 1, current_cols):
                            is_duplicate = True
                            for row_idx in range(len(table)):
                                val_a = table[row_idx][col_a] if col_a < len(table[row_idx]) else None
                                val_b = table[row_idx][col_b] if col_b < len(table[row_idx]) else None
                                va = str(val_a).strip() if val_a is not None else ''
                                vb = str(val_b).strip() if val_b is not None else ''
                                if va != vb:
                                    is_duplicate = False
                                    break
                            if is_duplicate:
                                print(f"  [fallback] 发现重复列: col {col_a} 和 col {col_b}")
                                col_to_delete = col_b
                                break

                if col_to_delete is not None:
                    print(f"  [fallback] 删除列: col {col_to_delete}")
                    for i in range(len(table)):
                        if col_to_delete < len(table[i]):
                            del table[i][col_to_delete]
                    current_cols = target_cols
                    print(f"  [fallback] 列删除完成，剩余{current_cols}列")
                else:
                    print(f"  [WARN step5_safety] fallback 也无法确定，跳过删除")
            else:
                # primary 部分成功（找到一些但不够），跳过删除
                print(f"  [WARN step5_safety] 部分空列不足（找到{len(empty_col_indices)}个 < 需删除{excess_cols}个），跳过删除")

        # 情况2：OCR列数 < LLM列数（需要补充左侧空列）
        elif current_cols < target_cols:
            needed_cols = target_cols - current_cols
            print(f"需要补充左侧{needed_cols}列")

            for i in range(len(table)):
                table[i] = [None] * needed_cols + table[i]

            current_cols = target_cols
        else:
            print("列数已匹配，无需调整")

        print(f"调整后表格列数: {current_cols}")

        # 4. 在顶部添加一行用于列标题
        table.insert(0, [None] * current_cols)

        # 5. 从右向左填充清理后的列标题
        cleaned_headers_copy = cleaned_headers.copy()
        for col in range(current_cols - 1, -1, -1):
            if cleaned_headers_copy:
                table[0][col] = cleaned_headers_copy.pop()

        return table

    def _validate_cleaned_headers(self, cleaned_headers):
        """
        验证清理后的列标题
        """

        # 统计空列标题
        empty_count = sum(1 for h in cleaned_headers if not h)
        if empty_count > 0:
            print(f"⚠️ 有{empty_count}个空列标题")

        # 检查是否有非空的列标题
        non_empty_count = len(cleaned_headers) - empty_count
        if non_empty_count == 0:
            print("⚠️ 所有列标题都为空，请检查LLM输出")

        # 检查列标题格式（可选，仅用于信息提示）
        for i, header in enumerate(cleaned_headers):
            if header:
                # 检查是否有分隔符但格式异常
                if '>>' in header:
                    parts = header.split('>>')
                    if len(parts) > 2:
                        print(f"  列{i}: 包含多层分隔符 -> {header}")
                elif '>' in header:
                    print(f"  列{i}: 使用单层分隔符 -> {header}")

    def _looks_like_numeric_data(self, text):
        """判断文本是否看起来像数值型数据"""
        if not text:
            return False

        text_str = str(text).strip()
        clean_text = text_str.replace(',', '').replace(' ', '').replace('¥', '').replace('$', '').replace('€', '')

        if clean_text.startswith('(') and clean_text.endswith(')'):
            clean_text = '-' + clean_text[1:-1]

        if not any(c.isdigit() for c in clean_text):
            return False

        try:
            if '%' in clean_text:
                clean_text = clean_text.replace('%', '')
                float(clean_text)
                return True
            else:
                float(clean_text)
                return True
        except ValueError:
            digit_count = sum(1 for c in clean_text if c.isdigit())
            total_len = len(clean_text)
            if digit_count / total_len > 0.7:
                return True

        return False

    def step6_add_row_headers_intelligent(self, table, row_headers, table_boundaries=None):
        """第6步：优化版 - 利用层级关系和相邻位置插入"""

        # [1] 首先尝试匹配所有LLM表头
        llm_to_ocr_match = {}
        ocr_to_llm_match = {}

        for llm_idx, llm_header in enumerate(row_headers):
            header_text = str(llm_header)

            # 使用更灵活的匹配策略
            best_match_row = -1
            best_match_score = 0.0

            for ocr_row in range(1, len(table)):
                if ocr_row in ocr_to_llm_match:
                    continue

                # 计算匹配分数
                row_score = self._calculate_header_match_score(header_text, table[ocr_row])

                if row_score > best_match_score:
                    best_match_score = row_score
                    best_match_row = ocr_row

            # 动态调整匹配阈值
            threshold = self._get_dynamic_threshold(header_text)
            if best_match_score >= threshold:
                llm_to_ocr_match[llm_idx] = (best_match_row, best_match_score)
                ocr_to_llm_match[best_match_row] = llm_idx

        # [2] 分析LLM表头的层级关系
        header_hierarchy = self._analyze_header_hierarchy(row_headers)

        # [3] 为每个LLM表头确定插入位置
        llm_insert_positions = {}

        for llm_idx, llm_header in enumerate(row_headers):
            if llm_idx in llm_to_ocr_match:
                # 有精确匹配
                ocr_row, score = llm_to_ocr_match[llm_idx]


                llm_insert_positions[llm_idx] = ('exact', ocr_row)
            else:
                # 无匹配，需要找插入位置
                insert_info = self._find_best_insert_position(
                    llm_idx, llm_header, row_headers,
                    llm_to_ocr_match, header_hierarchy, table
                )
                llm_insert_positions[llm_idx] = insert_info

        # [4] 按LLM顺序构建表格
        new_table = []
        if len(table) > 0:
            new_table.append(table[0].copy())

        # 记录已处理的OCR行
        processed_ocr_rows = set()

        # 按LLM顺序处理
        for llm_idx in range(len(row_headers)):
            insert_type, ref_info = llm_insert_positions[llm_idx]

            if insert_type == 'exact':
                # 处理精确匹配
                ocr_row = ref_info
                if ocr_row not in processed_ocr_rows:
                    new_row = table[ocr_row].copy()
                    new_row[0] = row_headers[llm_idx]
                    new_table.append(new_row)
                    processed_ocr_rows.add(ocr_row)
            elif insert_type == 'before':
                # 在参考行之前插入
                ref_row = ref_info
                if ref_row in processed_ocr_rows:
                    # 找到参考行在新表格中的位置
                    ref_position = self._find_row_in_new_table(new_table, ref_row)
                    if ref_position != -1:
                        blank_row = [None] * len(table[0]) if table else []
                        blank_row[0] = row_headers[llm_idx]
                        new_table.insert(ref_position, blank_row)
            elif insert_type == 'after':
                # 在参考行之后插入
                ref_row = ref_info
                if ref_row in processed_ocr_rows:
                    ref_position = self._find_row_in_new_table(new_table, ref_row)
                    if ref_position != -1:
                        blank_row = [None] * len(table[0]) if table else []
                        blank_row[0] = row_headers[llm_idx]
                        new_table.insert(ref_position + 1, blank_row)

        # 假设new_table中的行顺序和原始table中匹配的行顺序一致
        complete_table = []

        # 添加表头行
        if len(new_table) > 0:
            complete_table.append(new_table[0].copy())

        # 遍历原始table的行
        new_table_index = 1  # 跳过表头
        for ocr_row in range(1, len(table)):
            if ocr_row in ocr_to_llm_match:
                # 这行应该有LLM匹配，应该在new_table中
                if new_table_index < len(new_table):
                    complete_table.append(new_table[new_table_index].copy())
                    new_table_index += 1
            else:
                # 这行没有LLM匹配，直接使用原始行
                complete_table.append(table[ocr_row].copy())

        # 处理new_table中可能剩余的行
        while new_table_index < len(new_table):
            complete_table.append(new_table[new_table_index].copy())
            new_table_index += 1

        table[:] = complete_table

        return table

    def _find_best_insert_position(self, llm_idx, llm_header, row_headers,
                                   llm_to_ocr_match, header_hierarchy, table):
        """
        为无匹配的LLM表头找到最佳插入位置
        """

        # 1. 检查层级关系：如果是分类标题，应该在其子项之前
        if llm_idx in header_hierarchy['parents']:
            children = header_hierarchy['parents'][llm_idx]
            # 找到第一个有匹配的子项
            for child_idx in children:
                if child_idx in llm_to_ocr_match:
                    ocr_row, score = llm_to_ocr_match[child_idx]
                    return ('before', ocr_row)  # 在子项之前插入

        # 2. 向后查找相邻的有匹配表头
        next_idx = llm_idx + 1
        while next_idx < len(row_headers):
            if next_idx in llm_to_ocr_match:
                ocr_row, score = llm_to_ocr_match[next_idx]
                return ('before', ocr_row)  # 在后一个匹配表头之前插入
            next_idx += 1

        # 3. 向前查找相邻的有匹配表头
        prev_idx = llm_idx - 1
        while prev_idx >= 0:
            if prev_idx in llm_to_ocr_match:
                ocr_row, score = llm_to_ocr_match[prev_idx]
                return ('after', ocr_row)  # 在前一个匹配表头之后插入
            prev_idx -= 1

        # 4. 都没有找到，使用默认位置
        return ('end', -1)

    def _calculate_header_match_score(self, header_text, table_row):
        """
        计算表头与表格行的匹配分数
        不使用硬编码的关键词
        """
        # 提取表头关键词（最后一个>>后的部分）
        if '>>' in header_text:
            match_text = header_text.split('>>')[-1].strip()
        else:
            match_text = header_text.strip()

        if not match_text:
            return 0.0

        best_score = 0.0
        for cell in table_row:
            if not cell:
                continue

            cell_text = str(cell).strip()

            # 使用模糊匹配
            similarity = self.calculate_similarity_v2(match_text, cell_text)

            # 考虑长度比例：短文本完全匹配更重要
            len_ratio = min(len(match_text), len(cell_text)) / max(len(match_text), len(cell_text), 1)
            adjusted_score = similarity * len_ratio

            if adjusted_score > best_score:
                best_score = adjusted_score

        return best_score

    def _analyze_header_hierarchy(self, row_headers):
        """
        分析表头的层级关系
        """
        hierarchy = {
            'parents': {},  # parent_idx -> [child_idx1, child_idx2, ...]
            'children': {},  # child_idx -> parent_idx
            'depths': {}  # idx -> depth
        }

        for i, header in enumerate(row_headers):
            header_text = str(header)

            # 计算深度（>>的数量）
            depth = header_text.count('>>') + 1
            hierarchy['depths'][i] = depth

            # 如果是子项，寻找父级
            if '>>' in header_text:
                # 提取父级路径
                parts = header_text.split('>>')
                parent_path = '>>'.join(parts[:-1])

                # 寻找父级索引
                for j in range(i - 1, -1, -1):
                    if str(row_headers[j]) == parent_path:
                        hierarchy['children'][i] = j
                        if j not in hierarchy['parents']:
                            hierarchy['parents'][j] = []
                        hierarchy['parents'][j].append(i)
                        break

        return hierarchy

    def _get_dynamic_threshold(self, header_text):
        """
        动态调整匹配阈值
        """
        # 短文本需要更高的匹配度
        text_len = len(header_text)

        if text_len <= 3:
            return 0.9  # 短文本要求高匹配度
        elif text_len <= 10:
            return 0.7
        else:
            return 0.6  # 长文本允许较低的匹配度

    def _find_row_in_new_table(self, new_table, ocr_row):
        """
        在新表格中查找原始OCR行的位置
        """
        for i, row in enumerate(new_table):
            # 这里需要根据实际情况实现
            # 可能需要额外的标记来追踪原始行号
            pass
        return -1

    def step7_merge_and_remove_columns(self, table):
        """
        合并行表头并删除列：
        1. 如果列标题行左侧有2个及以上连续空表头
        2. 对每一行：
           - 如果第一列不为空，则将后面连续空表头列的数据合并到第一列
           - 如果第一列为空，则不合并
        3. 删除后面连续的空表头列（保留第一列）
        """

        if not table or len(table) < 2:
            return table

        # 获取列标题行（第0行）
        header_row = table[0]

        # 统计左侧连续空表头的数量
        empty_header_count = 0
        for header in header_row:
            if not header or str(header).strip() == "":
                empty_header_count += 1
            else:
                break

        # 如果左侧连续空表头少于2个，不处理
        if empty_header_count < 2:
            return table

        # 第一步：逐行处理，条件性合并
        for i in range(1, len(table)):
            row = table[i]

            # 跳过完全空的行
            if not any(cell for cell in row if cell and str(cell).strip() != ""):
                continue

            # 如果第一列不为空，才进行合并
            if row[0] and str(row[0]).strip() != "":

                # 构建合并后的内容
                merged_parts = [str(row[0]).strip()]

                # 添加后续连续空表头列的内容
                for j in range(1, empty_header_count):
                    if j < len(row) and row[j] and str(row[j]).strip() != "":
                        merged_parts.append(str(row[j]).strip())

                # 如果有多部分内容，用>>连接
                if len(merged_parts) > 1:
                    table[i][0] = ">>".join(merged_parts)
            else:
                pass

        # 第二步：无论如何都删除后面连续的空表头列（保留第一列）
        # 要删除的列索引：1 到 empty_header_count-1
        columns_to_remove = list(range(1, empty_header_count))

        if columns_to_remove:

            # 对每一行，删除指定列
            for i in range(len(table)):
                # 从右向左删除，避免索引错位
                for col_idx in sorted(columns_to_remove, reverse=True):
                    if col_idx < len(table[i]):
                        del table[i][col_idx]

        return table

    # 新增：标记表格的代理方法
    def step8_create_marked_table(self, table, row_checks=None, col_checks=None):
        """代理到独立的标记表格处理器"""
        return self.marked_table_processor.create_marked_table(table, row_checks, col_checks)

    def step8_add_feature_marks(self, validated_table, validation_marks):
        """代理到独立的标记表格处理器"""
        return self.marked_table_processor.add_feature_marks(validated_table, validation_marks)

    def is_numeric_value(self, cell_value):
        """代理到独立的标记表格处理器"""
        return self.marked_table_processor.is_numeric_value(cell_value)

    def _is_pure_numeric(self, cell):
        """代理到独立的标记表格处理器"""
        return self.marked_table_processor._is_pure_numeric(cell)

    # ========== 完整流程 ==========
    def process_single_table(self, ocr_tables, llm_table_info):
        """
        处理单个表格的完整流程
        """
        print(f"\n{'=' * 60}")
        print(f"开始处理表格")
        print(f"{'=' * 60}")

        # 第3步：合并OCR表格
        merged_data = self.step3_merge_ocr_tables(ocr_tables, llm_table_info)
        if not merged_data:
            self.log_issue("表格合并失败")
            return None

        # 第4步：创建基础表格
        base_table = self.step4_create_base_data_table(merged_data)
        if not base_table:
            self.log_issue("创建基础表格失败")
            return None

        # 第5步：添加列标题
        col_headers = merged_data.get('headers', {}).get('cols', [])
        table_with_cols = self.step5_add_column_headers(base_table, col_headers, merged_data.get('cells', []))

        # 第6步：添加行标题
        row_headers = merged_data.get('headers', {}).get('rows', [])
        final_table = self.step6_add_row_headers_intelligent(
            table_with_cols,
            row_headers,
            table_boundaries=merged_data.get('table_boundaries')
        )

        # 第7步：合并行表头并删除列
        final_table = self.step7_merge_and_remove_columns(final_table)

        # 第8步：添加行列标记（根据数据类型）
        marked_table = self.step8_create_marked_table(final_table)

        return marked_table

    def _clean_sheet_name(self, name):
        """
        清理Sheet名称，确保符合Excel要求
        Excel Sheet名称限制：
        1. 不超过31个字符
        2. 不能包含字符：: \ / ? * [ ]
        3. 不能为空
        4. 不能以'开头
        """
        if not name or not isinstance(name, str):
            return "未命名表格"

        # 替换非法字符
        illegal_chars = [':', '\\', '/', '?', '*', '[', ']']
        for char in illegal_chars:
            name = name.replace(char, '_')

        # 移除首尾空格
        name = name.strip()

        # 如果以'开头，移除
        if name.startswith("'"):
            name = name[1:]

        # 截断到31个字符
        if len(name) > 31:
            name = name[:31]

        # 确保不为空
        if not name:
            name = "未命名表格"

        return name

    def step9_save_to_excel_optimized_old(self, tables_data, output_file, table_names, metadata_list=None):
        """
        保存到 Excel（优化版）- 增加元数据支持
        强制使用外部传入的 table_names，不再自己拼名字

        Args:
            tables_data: 表格数据列表
            output_file: 输出文件路径
            table_names: 表格名称列表
            metadata_list: 元数据列表（可选，新增参数）
        """
        from openpyxl import Workbook
        from pathlib import Path

        wb = Workbook()
        wb.remove(wb.active)  # 删默认 Sheet

        print(f"📊📊 开始保存Excel到: {output_file}")

        for idx, (table, name) in enumerate(zip(tables_data, table_names)):
            ws = wb.create_sheet(title=name)  # 直接用外部名字

            # 保存表格数据（原有逻辑保持不变）
            for r, row in enumerate(table, 1):
                for c, val in enumerate(row, 1):
                    ws.cell(row=r, column=c, value=val)

            # ========== 新增：保存元数据到表格末尾 ==========
            if metadata_list and idx < len(metadata_list):
                metadata = metadata_list[idx]

                # 计算数据行数
                data_row_count = len(table)

                # 在数据之后空一行，然后添加元数据
                metadata_start_row = data_row_count + 2

                # 添加元数据标记和内容
                if any(metadata.values()):  # 只有存在有效元数据时才保存
                    ws.cell(row=metadata_start_row, column=1, value="")

                    row_offset = 1
                    if metadata.get('default_currency'):
                        ws.cell(row=metadata_start_row + row_offset, column=1,
                                value=f"currency:{metadata['default_currency']}")
                        print(f"      币种: {metadata['default_currency']}")
                        row_offset += 1

                    if metadata.get('default_report_period'):
                        ws.cell(row=metadata_start_row + row_offset, column=1,
                                value=f"report_period:{metadata['default_report_period']}")
                        print(f"      报告期: {metadata['default_report_period']}")
                        row_offset += 1

                    if metadata.get('default_unit'):
                        ws.cell(row=metadata_start_row + row_offset, column=1,
                                value=f"unit:{metadata['default_unit']}")
                        print(f"      单位: {metadata['default_unit']}")
                        row_offset += 1

                    if metadata.get('original_table_name'):
                        ws.cell(row=metadata_start_row + row_offset, column=1,
                                value=f"table_name:{metadata['original_table_name']}")
                        print(f"      原始表名: {metadata['original_table_name']}")
                        row_offset += 1

                    if metadata.get('ocr_table_id') != -1:
                        ws.cell(row=metadata_start_row + row_offset, column=1,
                                value=f"ocr_table_id:{metadata['ocr_table_id']}")
                        print(f"      OCR表ID: {metadata['ocr_table_id']}")
                        row_offset += 1

                    ws.cell(row=metadata_start_row + row_offset, column=1, value="")

                    print(f"    元数据已保存到第{metadata_start_row}行之后")
            else:
                print(f"    无元数据或元数据索引超出范围")

        Path(output_file).parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_file)
        wb.close()

        print(f"✅ Excel保存完成: {output_file}")
        return True

    def step9_save_to_excel_optimized(self, tables_data, output_file, table_names, metadata_list=None):
        from openpyxl import Workbook
        from pathlib import Path

        wb = Workbook()
        wb.remove(wb.active)

        print(f"📊📊📊📊 开始保存Excel到: {output_file}")

        for idx, (table, name) in enumerate(zip(tables_data, table_names)):
            ws = wb.create_sheet(title=name)

            # 🔥🔥🔥🔥 只处理第2行（索引为1，因为第1行是表头）
            filtered_table = table.copy()  # 先复制整个表格

            # 检查是否有第2行
            if len(table) >= 2:
                second_row = table[1]  # 第2行（索引1）

                # 检查第2行的行标记
                row_marker = self._extract_row_marker(second_row)

                # 如果行标记是0或'0'，删除第2行
                if str(row_marker).strip() in ['0', '0']:
                    del filtered_table[1]  # 删除第2行


            # 写入过滤后的数据
            for r, row in enumerate(filtered_table, 1):
                # 第1列：插入"项目0"（表头行）或空值（数据行）
                if r == 1:
                    ws.cell(row=r, column=1, value="项目0")
                else:
                    ws.cell(row=r, column=1, value="")

                # 原有数据从第2列开始写入
                for c, val in enumerate(row, 2):
                    ws.cell(row=r, column=c, value=val)

            # ========== 保存元数据到表格末尾 ==========
            if metadata_list and idx < len(metadata_list):
                metadata = metadata_list[idx]
                data_row_count = len(filtered_table)
                metadata_start_row = data_row_count + 2

                # 🔥 定义所有需要保存的元数据字段
                valid_keys = ["bankname", "currency", "report_period", "unit", "table_name", "ocr_table_id", "entity"]

                # 🔥 映射字典：将valid_keys中的字段名映射到metadata中的实际键名
                # 注意：这里假设metadata中的键名与valid_keys不完全一致
                field_mapping = {
                    "bankname": "bank_name",  # 假设元数据中可能是bank_name而不是bankname
                    "currency": "default_currency",
                    "report_period": "default_report_period",
                    "unit": "default_unit",
                    "table_name": "original_table_name",
                    "ocr_table_id": "ocr_table_id",
                    "entity": "entity"
                }

                ws.cell(row=metadata_start_row, column=1, value="")

                row_offset = 1
                # 🔥 强制保存所有valid_keys字段
                for key in valid_keys:
                    # 获取元数据中对应的键名
                    metadata_key = field_mapping.get(key, key)

                    # 获取值，如果不存在则使用空字符串
                    value = metadata.get(metadata_key, "")

                    # 保存到Excel
                    ws.cell(row=metadata_start_row + row_offset, column=1,
                            value=f"{key}:{value}")

                    row_offset += 1

                ws.cell(row=metadata_start_row + row_offset, column=1, value="")


        Path(output_file).parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_file)
        wb.close()

        print(f"✅ Excel保存完成: {output_file}")
        return True


    def _extract_row_marker(self, row):
        """
        从行数据中提取行标记
        根据你的数据结构，行标记可能在特定位置
        """
        if not row or len(row) == 0:
            return None

        last_val = row[-1]
        return last_val

    def detect_table_anomalies(self, table_data, table_name, ocr_result=None, llm_result=None):
        """
        检测表格是否存在需要人工审核的异常

        Args:
            table_data: 表格数据（二维数组）
            table_name: 表格名称
            ocr_result: OCR识别结果（可选）
            llm_result: LLM分析结果（可选）

        Returns:
            dict: {
                'status': ReviewStatus,
                'issues': [异常描述列表],
                'severity': 'warning' | 'error'
            }
        """
        issues = []
        severity = 'warning'

        if not table_data or len(table_data) == 0:
            return {
                'status': ReviewStatus.AUTO,
                'issues': [],
                'severity': 'warning'
            }

        # 1. 检测列数异常（表头列数与数据列数不匹配）
        header_row = table_data[0] if len(table_data) > 0 else []
        col_count = len(header_row) if header_row else 0

        # 检查是否有明显的重复表头（同一表头出现多次）
        if header_row:
            header_values = [str(h).strip() if h else '' for h in header_row]
            non_empty_headers = [h for h in header_values if h and h != '项目0']

            # 检查是否有重复的年份表头
            seen_headers = {}
            for h in non_empty_headers:
                # 提取年份部分进行比较
                year_match = re.search(r'(\d{4})年', h)
                if year_match:
                    year = year_match.group(1)
                    if year not in seen_headers:
                        seen_headers[year] = []
                    seen_headers[year].append(h)

            # 如果同一个年份出现多次（不同的期间如"1-6月"和"12月31日"），标记为需要审核
            for year, headers in seen_headers.items():
                if len(headers) > 1:
                    # 检查是否是不同期间
                    periods = set()
                    for h in headers:
                        period_match = re.search(r'1-6月|12月31日|6月30日', h)
                        if period_match:
                            periods.add(period_match.group(0))
                    if len(periods) > 1:
                        issues.append(f"检测到多个时间期间的列（{year}年），可能是LLM将多个表格合并识别")
                        severity = 'warning'

        # 2. 检测空表格
        if len(table_data) <= 1:
            issues.append("表格行数过少，可能是空表格")
            severity = 'error'

        # 3. 检测列数过多或过少
        expected_cols_range = (4, 15)  # 合理的列数范围
        if col_count < expected_cols_range[0] or col_count > expected_cols_range[1]:
            issues.append(f"表格列数异常（{col_count}列），可能在{expected_cols_range[0]}-{expected_cols_range[1]}列之间")
            severity = 'warning'

        # 4. 检测数据行是否为空
        data_rows = table_data[1:] if len(table_data) > 1 else []
        empty_rows = 0
        for row in data_rows:
            if row and all(not cell for cell in row):
                empty_rows += 1

        if data_rows and empty_rows / len(data_rows) > 0.5:
            issues.append(f"超过50%的数据行为空，可能存在数据缺失")
            severity = 'warning'

        # 5. 如果LLM返回了列信息，检查是否与OCR不一致
        if llm_result and 'tables' in llm_result:
            for llm_table in llm_result['tables']:
                llm_cols = llm_table.get('headers', {}).get('cols', [])
                if llm_cols and len(llm_cols) != col_count:
                    diff = abs(len(llm_cols) - col_count)
                    if diff >= 2:
                        issues.append(f"LLM识别的列数({len(llm_cols)})与实际列数({col_count})差异较大，可能存在表格合并问题")
                        severity = 'warning'

        # 6. 检测表头缺失（跳过前两列，跳过行标题列和多级表头）
        # 第一列是"项目0"（行标题），第二列通常是"项目"或子表头，不检查
        if header_row and len(table_data) > 2:
            missing_header_cols = []
            for col_idx, header_val in enumerate(header_row):
                # 跳过前两列（行标题列和多级表头列）
                if col_idx < 2:
                    continue
                    
                header_str = str(header_val).strip() if header_val else ''
                # 认为是无效表头：空值、纯数字、通用占位符
                is_valid_header = (
                    header_str and 
                    header_str not in ['nan', 'None', '', '项目0', '项目'] and
                    not re.match(r'^[\d.]+$', header_str)  # 不是纯数字
                )
                if not is_valid_header:
                    # 检查这一列的数据填充率
                    col_values = [table_data[row_idx][col_idx] for row_idx in range(1, len(table_data)) if col_idx < len(table_data[row_idx])]
                    if col_values:
                        non_empty_count = sum(1 for v in col_values if v and str(v).strip() and str(v).strip() not in ['nan', 'None', ''])
                        fill_rate = non_empty_count / len(col_values) if col_values else 0
                        # 只有数据填充率超过70%，才说明这列确实重要但表头没识别出来
                        if fill_rate > 0.7:
                            missing_header_cols.append(col_idx + 1)  # 1-based

            if missing_header_cols:
                issues.append(f"第 {', '.join(map(str, missing_header_cols))} 列缺少有效表头，但数据填充率>70%，可能存在表头识别错误")
                severity = 'warning'

        # 7. 检测数据缺失（跳过前两列，某列有大量数据缺失）
        # 只有填充率<15%且行数>15的列才认为是严重缺失
        if len(table_data) > 3 and col_count > 2:
            missing_data_cols = []
            for col_idx in range(2, col_count):  # 跳过前两列
                col_values = [table_data[row_idx][col_idx] for row_idx in range(1, len(table_data)) if col_idx < len(table_data[row_idx])]
                if col_values:
                    non_empty_count = sum(1 for v in col_values if v and str(v).strip() and str(v).strip() not in ['nan', 'None', ''])
                    fill_rate = non_empty_count / len(col_values) if col_values else 0
                    # 只有填充率<15%且行数>15，才认为是严重缺失
                    if fill_rate < 0.15 and len(col_values) > 15:
                        missing_data_cols.append((col_idx + 1, fill_rate))

            if missing_data_cols:
                col_info = [f"第{col}列({int(rate*100)}%)" for col, rate in missing_data_cols[:5]]  # 只显示前5个
                if len(missing_data_cols) > 5:
                    col_info.append(f"等共{len(missing_data_cols)}列")
                issues.append(f"以下列数据缺失严重(填充率<15%): {', '.join(col_info)}")
                severity = 'warning'

        # 8. 返回审核状态
        if issues:
            print(f"⚠️ 表格 '{table_name}' 检测到异常:")
            for issue in issues:
                print(f"   - {issue}")

            return {
                'status': ReviewStatus.PENDING_REVIEW,
                'issues': issues,
                'severity': severity
            }
        else:
            return {
                'status': ReviewStatus.AUTO,
                'issues': [],
                'severity': 'warning'
            }

    def detect_all_tables_anomalies(self, tables_data, table_names, ocr_results=None, llm_results=None):
        """
        批量检测所有表格的异常

        Args:
            tables_data: 表格数据列表
            table_names: 表格名称列表
            ocr_results: OCR识别结果列表（可选）
            llm_results: LLM分析结果列表（可选）

        Returns:
            list: 每个表格的审核状态信息列表
        """
        review_results = []

        for idx, (table, name) in enumerate(zip(tables_data, table_names)):
            ocr = ocr_results[idx] if ocr_results and idx < len(ocr_results) else None
            llm = llm_results[idx] if llm_results and idx < len(llm_results) else None

            result = self.detect_table_anomalies(table, name, ocr, llm)
            result['table_name'] = name
            result['table_index'] = idx
            review_results.append(result)

            # 打印审核状态
            status_icon = "⚠️" if result['status'] == ReviewStatus.PENDING_REVIEW else "✅"
            print(f"  {status_icon} [{idx+1}] {name}: {result['status']}")

        # 统计汇总
        pending_count = sum(1 for r in review_results if r['status'] == ReviewStatus.PENDING_REVIEW)
        if pending_count > 0:
            print(f"\n📋 审核汇总: {pending_count}/{len(review_results)} 个表格需要人工审核")

        return review_results


    def _extract_page_number_from_image_path(self, image_path):
        """
        从图片路径中提取页码信息
        例如: "XXX_015.png" -> "P015"
              "514001_152.png" -> "P152"  # 假设152是页码
              "document_123_page_045.jpg" -> "P045"
              "img_001.png" -> "P001"
        """
        if not image_path:
            print("  图片路径为空")
            return ""

        # 获取文件名（不含路径和扩展名）
        file_name = os.path.basename(image_path)
        base_name = os.path.splitext(file_name)[0]

        print(f"解析图片文件名: '{base_name}'")

        # 先尝试从文件名中直接提取可能的页码（通常是最后一段数字）
        # 对于类似"514001_152"的文件名，152可能是页码
        parts = base_name.split('_')
        if len(parts) >= 2:
            last_part = parts[-1]
            if last_part.isdigit() and len(last_part) >= 2:
                # 假设最后一部分数字是页码
                page_num = last_part.zfill(3)  # 格式化为3位
                result = f"P{page_num}"
                print(f"  从最后一部分提取页码: {last_part} -> {result}")
                return result

        # 尝试匹配各种页码格式
        patterns = [
            r'_(\d{3})$',  # 匹配_015（3位数字结尾）
            r'_(\d{2})$',  # 匹配_15（2位数字结尾）
            r'[pP]age_?(\d+)',  # 匹配page_15, Page45等
            r'p(\d+)',  # 匹配p15, P123等
            r'(\d{3})[._]',  # 匹配数字后跟点或下划线
            r'\D(\d{3})\D',  # 匹配被非数字包围的3位数字
        ]

        for pattern in patterns:
            match = re.search(pattern, base_name)
            if match:
                page_num = match.group(1)
                # 格式化页码，确保有前导零
                page_num_formatted = page_num.zfill(3)  # 至少3位，如015
                result = f"P{page_num_formatted}"
                print(f"  正则匹配到页码: {page_num} -> {result} (模式: {pattern})")
                return result

        # 如果没有匹配到，尝试提取所有数字
        all_digits = re.findall(r'\d+', base_name)
        if all_digits:
            # 取最长的一组数字（通常页码较长）
            longest_digits = max(all_digits, key=len)
            if len(longest_digits) >= 2:  # 至少2位才认为是页码
                page_num_formatted = longest_digits.zfill(3)
                result = f"P{page_num_formatted}"
                print(f"  提取最长数字作为页码: {longest_digits} -> {result}")
                return result

        print(f"  未找到页码信息，返回空字符串")
        return ""

    def _process_single_table_to_memory(self, ocr_tables, llm_table_info):
        """
        处理单个表格到内存（不保存文件）

        Args:
            ocr_tables: OCR表格列表
            llm_table_info: LLM表格信息

        Returns:
            list: 表格数据（二维列表）或 None
        """
        try:
            # 第3步：合并OCR表格
            merged_data = self.step3_merge_ocr_tables(ocr_tables, llm_table_info)
            if not merged_data:
                self.log_issue("表格合并失败")
                return None

            # 第4步：创建基础表格
            base_table = self.step4_create_base_data_table(merged_data)
            if not base_table:
                self.log_issue("创建基础表格失败")
                return None

            # 第5步：添加列标题
            col_headers = merged_data.get('headers', {}).get('cols', [])
            table_with_cols = self.step5_add_column_headers(base_table, col_headers, merged_data.get('cells', []))

            # 第6步：添加行标题
            row_headers = merged_data.get('headers', {}).get('rows', [])
            final_table = self.step6_add_row_headers_intelligent(
                table_with_cols,
                row_headers,
                table_boundaries=merged_data.get('table_boundaries')
            )

            # 第7步：合并行表头并删除列
            final_table = self.step7_merge_and_remove_columns(final_table)

            # 第8步：添加行列标记（根据数据类型）
            marked_table = self.step8_create_marked_table(final_table)

            return marked_table

        except Exception as e:
            print(f"❌ 内存处理单个表格失败: {e}")
            import traceback
            traceback.print_exc()
            return None

    def process_all_tables_to_memory(self, ocr_result, llm_result, image_path=None, bank_name="未知银行"):
        """
        处理表格并返回内存中的数据（不保存文件）- 改进版
        确保inner_idx按照表格在图片中的实际顺序分配，并为最后一个表格添加_T_后缀
        """
        print(f"\n🧠🧠🧠🧠 内存模式处理表格...")

        # ========== 第0步：数据预处理和验证 ==========
        self._fix_llm_table_references(ocr_result, llm_result)

        # ========== 第1步：准备数据 ==========
        ocr_result, llm_result = self.step1_prepare_data(ocr_result, llm_result)

        # ========== 第2步：提取表格数据 ==========
        ocr_tables, llm_tables = self.step2_extract_table_data(ocr_result, llm_result)

        if not ocr_tables or not llm_tables:
            self.log_issue("提取表格数据失败")
            return [], [], []

        print(f"OCR表格数量: {len(ocr_tables)}")
        print(f"LLM表格结构数量: {len(llm_tables)}")

        # ========== 第3步：列标题统一化处理 ==========
        llm_tables = self._unify_headers_across_tables(llm_tables)

        # ========== 第4步：提取页码前缀 ==========
        page_prefix = ""
        if image_path:
            page_prefix = self._extract_page_number_from_image_path(image_path)
            print(f"从图片路径提取页码前缀: {page_prefix}")

        # ========== 第5步：按顺序处理所有表格 ==========
        all_final_tables = []
        all_table_names = []
        all_metadata_list = []

        total_tables = len(llm_tables)

        # 直接按照LLM表格的原始顺序处理，确保顺序正确
        for inner_idx, llm_table_info in enumerate(llm_tables, 1):
            original_idx = inner_idx - 1  # 原始索引（0-based）

            print(f"\n处理表格 {inner_idx}/{total_tables} (原始索引:{original_idx})")

            # 提取表格名称
            table_name = llm_table_info.get('name', f'表格{inner_idx}')

            # ========== 关键修改：判断是否为最后一个表格 ==========
            is_last_table = (inner_idx == total_tables)

            # 构建完整Sheet名称（带页码前缀、内部序号和最后一个表格标记）
            if page_prefix:
                if is_last_table:
                    full_table_name = f"{page_prefix}_{inner_idx}_T_{table_name}"
                    print(f"  📍 这是本页最后一个表格，添加_T_后缀")
                else:
                    full_table_name = f"{page_prefix}_{inner_idx}_{table_name}"
            else:
                if is_last_table:
                    full_table_name = f"{inner_idx}_T_{table_name}"
                    print(f"  📍 这是本页最后一个表格，添加_T_后缀")
                else:
                    full_table_name = f"{inner_idx}_{table_name}"

            print(f"  Sheet名称: {full_table_name}", "bank_nam:::", bank_name)

            # 提取元数据
            table_metadata = {
                'bank_name': bank_name,
                'default_currency': llm_table_info.get('default_currency', ''),
                'default_report_period': llm_table_info.get('default_report_period', ''),
                'default_unit': llm_table_info.get('default_unit', ''),
                'original_table_name': llm_table_info.get('name', ''),
                'ocr_table_id': llm_table_info.get('ocr_tables', [])[0] if llm_table_info.get('ocr_tables') else -1,
                'inner_index': inner_idx,  # 内部序号（1-based）
                'original_index': original_idx,  # 原始索引（0-based）
                'page_prefix': page_prefix,
                'total_tables_in_image': total_tables,  # 图片中表格总数
                'is_last_table': is_last_table  # 新增：标记是否为最后一个表格
            }
            all_metadata_list.append(table_metadata)

            # 处理单个表格
            final_table = self._process_single_table_to_memory(ocr_tables, llm_table_info)

            if final_table:
                all_final_tables.append(final_table)
                all_table_names.append(full_table_name)
                print(f"  ✅ 表格处理成功")
            else:
                print(f"  ❌ 表格处理失败")
                # 即使处理失败，也保留名称和元数据记录
                all_table_names.append(full_table_name)

        # 验证数据一致性
        success_count = len(all_final_tables)
        total_count = len(llm_tables)

        if success_count != total_count:
            print(f"⚠️ 警告：部分表格处理失败 - 成功:{success_count}, 总数:{total_count}")

            # 确保三个列表长度一致
            min_count = min(len(all_final_tables), len(all_table_names), len(all_metadata_list))
            all_final_tables = all_final_tables[:min_count]
            all_table_names = all_table_names[:min_count]
            all_metadata_list = all_metadata_list[:min_count]

        print(f"\n✅ 内存处理完成: {success_count}/{total_count} 个表格成功")

        # 打印生成的Sheet名称列表
        print(f"📋 生成的Sheet名称:")
        for i, name in enumerate(all_table_names, 1):
            is_last = (i == total_count)
            last_marker = " 📍(最后一个)" if is_last else ""
            print(f"  {i}. {name}{last_marker}")

        return all_final_tables, all_table_names, all_metadata_list


    def process_all_tables(self, ocr_result, llm_result, output_file, final_output_file=None, image_path=None,
                           bank_name=""):
        """
        处理所有表格并保存到Excel文件

        Args:
            ocr_result: OCR识别结果
            llm_result: LLM分析结果
            output_file: 输出Excel文件路径
            final_output_file: 最终输出文件路径（可选）
            image_path: 图片路径（用于提取页码）
            bank_name: 银行名称

        Returns:
            dict: {
                'success': bool,
                'review_results': list (审核状态列表) - 新增
            }
        """
        print(f"\n📊📊 处理所有表格...")

        try:
            # 1. 处理表格到内存（现在返回三个值）
            tables_data, table_names, metadata_list = self.process_all_tables_to_memory(
                ocr_result,
                llm_result,
                image_path=image_path,
                bank_name=bank_name
            )

            if not tables_data:
                print("❌❌ 没有表格数据生成")
                return {'success': False, 'review_results': []}

            print(f"✅ 生成 {len(tables_data)} 个表格")

            # 2. 清理表格名称
            cleaned_table_names = []
            for name in table_names:
                cleaned_name = self._clean_sheet_name(name)
                cleaned_table_names.append(cleaned_name)

            # 3. 保存到Excel（现在传递metadata_list参数）
            success = self.step9_save_to_excel_optimized(
                tables_data,
                output_file,
                cleaned_table_names,
                metadata_list  # 新增参数
            )

            # 4. 检测表格异常（新增）
            review_results = []
            if success:
                print("\n🔍 开始审核表格质量...")
                review_results = self.detect_all_tables_anomalies(
                    tables_data,
                    cleaned_table_names,
                    llm_results=[llm_result] if llm_result else None
                )

            if success:
                print(f"✅ 表格已保存到: {output_file}")

                # 5. 如果有final_output_file，调用final_data_converter
                if final_output_file:
                    print(f"📋📋 生成最终数据文件: {final_output_file}")
                    try:
                        # 确保final_data_converter已初始化
                        if not hasattr(self, 'final_data_converter'):
                            self.final_data_converter = FinalDataConverter()

                        # 检查final_data_converter是否有process_to_final_format方法
                        if hasattr(self.final_data_converter, 'process_to_final_format'):
                            # 调用final_data_converter处理
                            final_success = self.final_data_converter.process_to_final_format(
                                tables_data=tables_data,
                                table_names=cleaned_table_names,
                                bank_name=bank_name,
                                output_path=final_output_file
                            )

                            if final_success:
                                print(f"✅ 最终数据文件已生成: {final_output_file}")
                            else:
                                print(f"⚠️ 最终数据文件生成失败")
                        else:
                            # 如果没有process_to_final_format方法，使用其他方法或直接跳过
                            print(f"⚠️ FinalDataConverter没有process_to_final_format方法，跳过最终数据生成")
                            if hasattr(self.final_data_converter, 'convert'):
                                # 尝试使用convert方法
                                try:
                                    final_success = self.final_data_converter.convert(
                                        input_path=output_file,
                                        output_path=final_output_file,
                                        bank_name=bank_name
                                    )
                                    if final_success:
                                        print(f"✅ 通过convert方法生成最终数据文件: {final_output_file}")
                                except Exception as e:
                                    print(f"⚠️ convert方法失败: {e}")
                            else:
                                print(f"⚠️ FinalDataConverter没有可用的转换方法，跳过")

                    except Exception as e:
                        print(f"⚠️ 最终数据转换失败: {e}")
                        import traceback
                        traceback.print_exc()

                return {'success': True, 'review_results': review_results}
            else:
                print("❌❌ 保存到Excel失败")
                return False

        except Exception as e:
            print(f"❌❌ 处理所有表格失败: {e}")
            import traceback
            traceback.print_exc()
            return False



# ====================================
# 使用示例
# ====================================
def main():
    # 你的数据
    ocr_result = {}
    llm_result= {'success': True,
                 'image_info': {'image_path': 'E:\\Datas\\base_pros\\DocuVista\\test_codes\\pngs\\123.png',
                                'image_id': 'img_0e6447019e10'}, 'tables_structure': {'tables': [
            {'id': '1', 'ocr_tables': [0],
             'headers': {'cols': ['', 'b>>2024年12月31日', 'c>>2024年9月30日', 'd>>2024年6月30日', 'e>>2024年3月31日'],
                         'rows': ['可用资本（数额）>>1核心一级资本净额', '可用资本（数额）>>2一级资本净额',
                                  '可用资本（数额）>>3资本净额', '风险加权资产（数额）>>4风险加权资产合计',
                                  '风险加权资产（数额）>>4a风险加权资产合计（应用资本底线前）',
                                  '资本充足率>>5核心一级资本充足率（%）',
                                  '资本充足率>>5a核心一级资本充足率（%）（应用资本底线前）',
                                  '资本充足率>>6一级资本充足率（%）', '资本充足率>>6a一级资本充足率（%）（应用资本底线前）',
                                  '资本充足率>>7资本充足率（%）', '资本充足率>>7a资本充足率（%）（应用资本底线前）',
                                  '其他资本要求>>8储备资本要求（%）', '其他资本要求>>9逆周期资本要求（%）',
                                  '其他资本要求>>10系统重要性银行附加资本要求（%）',
                                  '其他资本要求>>11其他各级资本要求（%）（8+9+10）',
                                  '其他资本要求>>12满足最低资本要求后的可用核心一级资本净额占风险加权资产的比例（%）']}}]},
                 'processing_stats': {'analysis_time_sec': 13.35, 'ocr_tables_count': 1, 'visual_tables_count': 1,
                                      'token_usage': {'prompt_tokens': 1155, 'completion_tokens': 421,
                                                      'total_tokens': 1576}}}

    # 创建重构器并处理
    reconstructor = TableReconstructor()
    success = reconstructor.process_all_tables123(
        ocr_result=ocr_result,
        llm_result=llm_result,
        output_file="../../../../test_codes/enhanced_table_analyzer/reconstructed_tables2.xlsx",
        bank_name="中国建设银行"
    )

    if success:
        print("✅ 表格重构成功！")
    else:
        print("❌ 表格重构失败！")


if __name__ == "__main__":
    main()