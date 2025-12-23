
import os
import pandas as pd
import time
import asyncio
from openpyxl import load_workbook
from pathlib import Path
from flask import jsonify
from typing import List, Dict

from backend.llm_services.state_manager import state_manager
from backend.llm_services.utils import validate_required_params, convert_to_excel_url
from backend.llm_services.task_management_service import _send_websocket_notification, logger
from backend.llm_services.single_table_service import _process_single_non_financial_table

from backend.utils.constants import MAIN_ROOT


# 异步批量处理图片
def _batch_process_images_sync(data, task_id):
    """同步批量处理函数"""
    # 预置最终状态，保证 finally 里一定有值
    final_status = "error"
    final_data   = {"error": "未知处理结果"}

    # ⬇️ 任务一创建就存进去，确保轮询能查到
    state_manager.set_task_result(task_id, {
        "status": "running",
        "data": {},
        "completed_at": None
    })

    try:
        # 1. 设置初始状态
        state_manager.set_processing_status(task_id, {
            "status": "processing",
            "progress": 0,
            "message": "开始处理",
            "total": len(data.get('image_paths', [])),
            "processed": 0,
            "start_time": time.time()
        })

        print(f"🔄 开始同步批量处理 - 任务ID: {task_id}")
        _send_websocket_notification(task_id, 'started', {
            'message': '任务已开始处理',
            'total': len(data.get('image_paths', []))
        })

        # 2. 执行业务逻辑
        result = asyncio.run(_batch_process_images(data, task_id))
        print(f"🔄 批量处理完成，结果: {result.get('success')}")

        # 3. 根据结果分支更新状态
        if result.get('processing_completed', False) or result.get('success', False):
            final_status = "completed"
            final_data   = result.get('data', {})

            excel_url = final_data.get('excel_url')
            print("data_payload.get('excel_url')::::", excel_url)

            state_manager.set_processing_status(task_id, {
                "status": "completed",
                "progress": 100,
                "message": final_data.get('message', '处理完成'),
                "excel_url": excel_url,
                "success_count": final_data.get('success', 0),
                "failed_count": final_data.get('failed', 0),
                "total": final_data.get('total', 0),
                "table_type": final_data.get('table_type', 'non_financial'),
                "completion_time": time.time()
            })
            _send_websocket_notification(task_id, 'completed', final_data)

        else:
            final_status = "error"
            final_data   = {"error": result.get('error', '处理失败')}
            state_manager.set_processing_status(task_id, {
                "status": "error",
                "message": final_data["error"],
                "error_time": time.time()
            })
            _send_websocket_notification(task_id, 'error', error_message=final_data["error"])

    except Exception as e:
        logger.error(f"批量处理异常: {str(e)}")
        final_status = "error"
        final_data   = {"error": str(e)}
        state_manager.set_processing_status(task_id, {
            "status": "error",
            "message": str(e),
            "error_time": time.time()
        })
        _send_websocket_notification(task_id, 'error', error_message=str(e))

    finally:
        # 🔚 无论成功、失败、抛异常，都存结果供轮询接口查询
        state_manager.set_task_result(task_id, {
            "status": final_status,
            "data": final_data,
            "completed_at": time.time()
        })
        print(f"✅ 任务结果已存入状态管理器 - 任务ID: {task_id}, 状态: {final_status}")

def _handle_batch_process(data, task_id):
    """统一的批量处理入口"""
    try:
        if not data:
            return jsonify({
                "success": False,
                "error": "请求体不能为空"
            }), 400

        # 添加调试信息
        table_type = data.get('table_type', 'unknown')
        print(f"🔄 批量处理入口 - 表格类型: {table_type}")

        result = asyncio.run(_batch_process_images(data, task_id))
        return jsonify(result)
    except Exception as e:
        logger.error(f"批量处理错误: {str(e)}")
        return jsonify({
            "success": False,
            "error": f"批量处理错误: {str(e)}"
        }), 500



async def _validate_batch_processing_input(data, task_id):
    """验证批量处理输入参数 - 添加详细调试"""
    table_type = data.get('table_type', 'financial')
    print(f"🔄 批量处理开始 - 类型: {table_type}")

    # 打印前端传递的所有数据
    print(f"📋 前端传递的数据: {data}")

    # 获取处理器
    processor = state_manager.get_appropriate_processor(table_type)
    print(f"🔍 处理器检查: {processor}, llm_client: {getattr(processor, 'llm_client', None)}")

    # 验证处理器
    if processor is None:
        error_msg = "处理器未初始化"
        return _create_error_response(task_id, error_msg)

    if not hasattr(processor, 'llm_client') or processor.llm_client is None:
        error_msg = "LLM客户端未配置，请先调用配置接口"
        return _create_error_response(task_id, error_msg)

    # 验证必要参数
    is_valid, error_msg = validate_required_params(data, ['image_paths', 'output_dir'])
    if not is_valid:
        return _create_error_response(task_id, error_msg)

    # 验证图片文件
    image_paths = data.get('image_paths', [])
    print(f"🖼️ 前端传递的图片路径列表: {image_paths}")

    valid_image_paths, missing_images = _validate_image_paths(image_paths)

    if not valid_image_paths:
        error_msg = f"以下图片文件不存在: {missing_images}"
        return _create_error_response(task_id, error_msg)

    print(f"✅ 处理器检查通过 - llm_client: {processor.llm_client is not None}")

    return {
        "success": True,
        "processor": processor,
        "table_type": table_type,
        "image_paths": valid_image_paths
    }


# 在 _validate_image_paths 函数中添加目录结构检查
def _validate_image_paths(image_paths):
    """验证图片路径是否存在 - 使用常量路径"""
    valid_image_paths = []
    missing_images = []

    # 从常量导入路径
    from backend.utils.constants import MAIN_ROOT, JOINED_TABLES_ROOT

    for img_path in image_paths:
        # 清理路径
        img_path = img_path.lstrip('/')

        print("iiiiiiiiiiiiiiiiiiimg_path:", img_path)

        # 使用常量构建正确的路径
        if img_path.startswith('static/joined_tables/'):
            # 提取相对路径部分
            relative_path = img_path.replace('static/joined_tables/', '')
            full_path = Path(MAIN_ROOT) / JOINED_TABLES_ROOT / relative_path
        elif img_path.startswith('joined_tables/'):
            # 如果已经是 joined_tables/ 开头
            relative_path = img_path.replace('joined_tables/', '')
            full_path = Path(MAIN_ROOT) / JOINED_TABLES_ROOT / relative_path
        else:
            # 其他情况，直接使用完整路径
            full_path = Path(MAIN_ROOT) / JOINED_TABLES_ROOT / img_path

        print(f"🔍 原始路径: {img_path}")
        print(f"🔍 完整路径: {full_path}")
        print(f"🔍 路径是否存在: {full_path.exists()}")

        if full_path.exists():
            valid_image_paths.append(str(full_path))
            print(f"✅ 找到图片: {full_path}")
        else:
            missing_images.append(img_path)
            print(f"❌ 图片文件不存在: {full_path}")

    return valid_image_paths, missing_images


async def _prepare_output_file(data, image_paths, table_type):
    """准备输出文件路径"""
    # 提取文件夹名
    folder_name = _extract_folder_name(image_paths)

    # 创建输出目录
    excel_base_dir = Path("static/excel_data")
    excel_dir = excel_base_dir / folder_name
    excel_dir.mkdir(parents=True, exist_ok=True)

    # 构建文件名
    if table_type == 'non_financial':
        excel_filename = f"non_financial_batch_{folder_name}.xlsx"
    else:
        excel_filename = f"financial_batch_{folder_name}.xlsx"

    new_output_file = excel_dir / excel_filename
    output_file = str(new_output_file)

    print(f"📁 批量处理Excel文件将保存到: {output_file}")
    return output_file


def _extract_folder_name(image_paths):
    """从图片路径中提取文件夹名"""
    if not image_paths:
        return "unknown_batch"

    first_image_path = Path(image_paths[0])
    print(f"🔍 第一个图片路径: {first_image_path}")

    if 'joined_tables' in str(first_image_path):
        folder_name = first_image_path.parent.name
        print(f"📁 从joined_tables提取文件夹名: {folder_name}")
    else:
        folder_name = first_image_path.stem.split('_')[0]
        print(f"📁 从文件名提取文件夹名: {folder_name}")

    return folder_name


def _update_processing_status_start(task_id, image_paths, table_type, output_file):
    """更新处理状态为开始"""
    state_manager.set_processing_status(task_id, {
        "status": "processing",
        "progress": 0,
        "message": f"开始处理 {len(image_paths)} 张图片",
        "total": len(image_paths),
        "processed": 0,
        "table_type": table_type,
        "output_file": output_file,
        "start_time": time.time()
    })

    # 发送任务开始通知
    _send_websocket_notification(task_id, 'started', {
        'message': '任务已开始处理',
        'total': len(image_paths),
        'table_type': table_type
    })




async def _update_processing_progress(task_id, current_index, total_count, image_path):
    """更新处理进度"""
    progress = int((current_index / total_count) * 100)
    state_manager.set_processing_status(task_id, {
        "progress": progress,
        "processed": current_index,
        "message": f"正在处理第 {current_index + 1}/{total_count} 张图片: {Path(image_path).name}"
    })

    # 发送进度通知
    _send_websocket_notification(task_id, 'progress', {
        'progress': progress,
        'message': f'正在处理第 {current_index + 1}/{total_count} 张图片',
        'current': current_index + 1,
        'total': total_count
    })




def _create_single_image_error_result(image_path, error_msg):
    """创建单张图片处理错误结果"""
    return {
        "image_path": image_path,
        "status": "error",
        "error": error_msg,
        "success": False
    }


# 添加辅助函数
def _create_basic_worksheet(output_file, sheet_name, table_name, image_path, result):
    """创建基础信息工作表"""
    try:
        wb = load_workbook(output_file)
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

        ws = wb.create_sheet(sheet_name)

        # 添加基本信息
        ws['A1'] = "表格信息"
        ws['A2'] = f"工作表名称: {sheet_name}"
        ws['A3'] = f"表格名称: {table_name}"
        ws['A4'] = f"图片路径: {image_path}"
        ws['A5'] = f"处理状态: {result.get('status', '未知')}"
        ws['A6'] = f"处理时间: {time.strftime('%Y-%m-%d %H:%M:%S')}"

        # 添加处理结果信息
        ws['A8'] = "处理结果详情"
        ws['A9'] = f"复杂度: {result.get('complexity', '未知')}"
        ws['A10'] = f"评估原因: {result.get('assessment_reason', '无')}"
        ws['A11'] = f"错误信息: {result.get('error_message', '无')}"

        wb.save(output_file)
        print(f"✅ 基础工作表 {sheet_name} 创建完成")
    except Exception as e:
        print(f"❌ 创建基础工作表失败: {e}")


def _create_basic_dataframe(sheet_name, table_name, image_path, result):
    """创建基础信息的DataFrame"""
    import pandas as pd

    data = {
        '字段': ['工作表名称', '表格名称', '图片路径', '处理状态', '处理时间', '复杂度', '评估原因', '错误信息'],
        '值': [
            sheet_name,
            table_name,
            image_path,
            result.get('status', '未知'),
            time.strftime('%Y-%m-%d %H:%M:%S'),
            result.get('complexity', '未知'),
            result.get('assessment_reason', '无'),
            result.get('error_message', '无')
        ]
    }

    return pd.DataFrame(data)


def _clean_data_for_storage(data):
    """清理数据，移除不可序列化的对象"""
    if isinstance(data, dict):
        cleaned = {}
        for key, value in data.items():
            if key == 'df':
                # 跳过 DataFrame 对象
                continue
            elif isinstance(value, (dict, list)):
                cleaned[key] = _clean_data_for_storage(value)
            else:
                cleaned[key] = value
        return cleaned
    elif isinstance(data, list):
        return [_clean_data_for_storage(item) for item in data]
    else:
        return data


async def _generate_processing_report(task_id, processing_results, output_file, table_type):
    """生成处理结果报告"""
    results = processing_results["results"]
    success_count = processing_results["success_count"]
    failed_count = processing_results["failed_count"]
    total_count = processing_results["total_count"]

    # 更新最终状态
    state_manager.set_processing_status(task_id, {
        "progress": 100,
        "processed": total_count,
        "message": f"批量处理完成，成功: {success_count}, 失败: {failed_count}"
    })

    logger.info(f"批量处理完成 - 类型: {table_type}, 总数: {total_count}, 成功: {success_count}, 失败: {failed_count}")

    # ⭐⭐⭐ 简化：不再需要复杂的Excel保存逻辑，因为数据已经在处理过程中保存了 ⭐⭐⭐
    excel_saved = Path(output_file).exists()

    # 构建Excel URL
    excel_url = convert_to_excel_url(output_file)
    print(f"🔗 生成的Excel URL: {excel_url}")

    # 生成完成消息
    completion_message = _generate_completion_message(total_count, success_count, failed_count, table_type, results)

    # ⭐⭐⭐ 清理结果，移除DataFrame对象 ⭐⭐⭐
    cleaned_results = []
    for result in results:
        cleaned_result = result.copy()
        # 移除 DataFrame 对象
        if 'df' in cleaned_result:
            # 只保留 DataFrame 的基本信息用于调试
            df_info = None
            if cleaned_result['df'] is not None:
                df_info = {
                    'has_data': True,
                    'shape': getattr(cleaned_result['df'], 'shape', None),
                    'columns': list(cleaned_result['df'].columns) if hasattr(cleaned_result['df'], 'columns') else None
                }
            else:
                df_info = {'has_data': False}

            cleaned_result['df_info'] = df_info
            del cleaned_result['df']  # 移除 DataFrame 对象

        cleaned_results.append(cleaned_result)

    # 构建最终数据
    final_data = {
        "total": total_count,
        "success": success_count,
        "failed": failed_count,
        "non_financial": sum(1 for r in cleaned_results if r.get('is_non_financial')),
        "results": cleaned_results,
        "output_file": output_file,
        "excel_url": excel_url,
        "table_type": table_type,
        "processing_completed": True,
        "message": completion_message,
        "excel_saved": excel_saved
    }

    print(f"📊 最终处理统计 - 总数: {total_count}, 成功: {success_count}, 失败: {failed_count}")
    print(f"💾 Excel文件生成: {final_data['excel_saved']}")

    final_data = _clean_data_for_storage(final_data)

    # 存储任务结果
    state_manager.set_task_result(task_id, {
        "status": "completed",
        "data": final_data,
        "processing_completed": True,
        "completed_at": time.time()
    })

    # 发送完成通知
    _send_websocket_notification(task_id, 'completed', final_data)

    print(f"✅ 任务结果已存储到状态管理器: {task_id}")

    return {
        "success": True,
        "data": final_data,
        "task_id": task_id,
        "processing_completed": True
    }


def _fallback_to_batch_format(output_file, processing_results, table_type):
    """回退到批量处理格式保存"""
    try:
        from backend.llm_services.utils import export_processing_results_to_excel

        processing_info = {
            'total': processing_results["total_count"],
            'success': processing_results["success_count"],
            'failed': processing_results["failed_count"],
            'table_type': table_type,
            'results': processing_results["results"],
            'message': f"批量处理完成，成功: {processing_results['success_count']}, 失败: {processing_results['failed_count']}"
        }

        excel_saved = export_processing_results_to_excel(output_file, processing_info)

        if not excel_saved:
            # 如果详细导出失败，至少确保基础文件存在
            from backend.llm_services.utils import ensure_excel_file_exists
            ensure_excel_file_exists(output_file, processing_info)
            print("⚠️ 详细Excel导出失败，已创建基础文件")
            return False

        return True

    except Exception as e:
        print(f"❌ 批量处理格式保存失败: {e}")
        return False


def _generate_completion_message(total_count, success_count, failed_count, table_type, results=None):
    """生成完成消息"""
    non_financial_count = sum(1 for r in (results or []) if r.get('is_non_financial'))

    if non_financial_count > 0:
        return f"批量处理完成，总数: {total_count}, 成功识别: {success_count}, 非金融表格: {non_financial_count}, 失败: {failed_count}"
    elif success_count > 0:
        return f"批量处理完成，成功: {success_count}, 失败: {failed_count}"
    else:
        return f"批量处理完成，总数: {total_count}, 成功: {success_count}, 失败: {failed_count}"


def _create_error_response(task_id, error_msg):
    """创建错误响应"""
    logger.error(error_msg)
    state_manager.set_task_result(task_id, {
        "status": "error",
        "error": error_msg,
        "completed_at": time.time()
    })

    state_manager.set_processing_status(task_id, {
        "status": "error",
        "message": error_msg,
        "error_time": time.time()
    })

    _send_websocket_notification(task_id, 'error', error_message=error_msg)

    return {
        "success": False,
        "error": error_msg,
        "task_id": task_id
    }


async def _handle_batch_processing_error(task_id, error):
    """处理批量处理错误"""
    logger.error(f"批量处理失败: {str(error)}")
    import traceback
    traceback.print_exc()

    error_msg = f"批量处理失败: {str(error)}"

    state_manager.set_task_result(task_id, {
        "status": "error",
        "error": error_msg,
        "completed_at": time.time()
    })

    state_manager.set_processing_status(task_id, {
        "status": "error",
        "message": error_msg,
        "error_time": time.time(),
        "error": str(error)
    })

    _send_websocket_notification(task_id, 'error', error_message=str(error))

    print(f"❌ 错误结果已存储到状态管理器: {task_id}")

    return {
        "success": False,
        "error": error_msg,
        "task_id": task_id
    }


async def _analyze_processed_data(self, task_id: str, processing_results: Dict, output_file: str):
    """分析处理后的表格数据"""
    try:
        from backend.service.table_analysis_service import TableAnalysisService

        # 获取LLM客户端
        processor = state_manager.get_appropriate_processor('financial')
        analysis_service = TableAnalysisService(
            llm_client=processor.llm_client if processor else None,
            model_id=processor.model_id if processor else None
        )

        analysis_results = []

        # 分析每个成功处理的表格
        for result in processing_results["results"]:
            if result.get("success") and result.get("df") is not None:
                df = result["df"]
                table_name = result.get("table_name", "未知表格")

                # 进行分析
                analysis_result = await analysis_service.analyze_table_data(df, table_name)
                analysis_results.append({
                    "image_path": result["image_path"],
                    "table_name": table_name,
                    "analysis": analysis_result
                })

        # 保存分析结果
        analysis_file = await self._save_analysis_results(analysis_results, output_file)

        # 更新任务状态
        state_manager.set_analysis_results(task_id, {
            "analysis_completed": True,
            "analysis_file": analysis_file,
            "analyzed_tables": len(analysis_results),
            "results": analysis_results
        })

        return analysis_results

    except Exception as e:
        logger.error(f"数据分析失败: {e}")
        return []


async def _save_analysis_results(self, analysis_results: List[Dict], output_file: str) -> str:
    """保存分析结果到文件"""
    try:
        import json
        from datetime import datetime

        # 创建分析结果文件路径
        output_path = Path(output_file)
        analysis_file = output_path.parent / f"{output_path.stem}_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"

        # 简化分析结果（移除base64图片数据以减小文件大小）
        simplified_results = []
        for result in analysis_results:
            simplified = {
                "image_path": result["image_path"],
                "table_name": result["table_name"],
                "analysis": {
                    "basic_statistics": result["analysis"].get("basic_statistics", {}),
                    "data_quality": result["analysis"].get("data_quality", {}),
                    "llm_insights": result["analysis"].get("llm_insights", {}),
                    "summary": result["analysis"].get("summary", {})
                }
            }
            simplified_results.append(simplified)

        # 保存到JSON文件
        with open(analysis_file, 'w', encoding='utf-8') as f:
            json.dump(simplified_results, f, ensure_ascii=False, indent=2)

        return str(analysis_file)

    except Exception as e:
        logger.error(f"保存分析结果失败: {e}")
        return ""





async def _batch_process_images(data, task_id):
    """批量处理主逻辑"""
    try:
        # 初始化和验证
        validation_result = await _validate_batch_processing_input(data, task_id)
        print("validation_result:", validation_result)
        if not validation_result["success"]:
            return validation_result

        processor = validation_result["processor"]
        table_type = validation_result["table_type"]
        image_paths = validation_result["image_paths"]

        # ⭐⭐⭐ 关键修复：确保缓存检查被正确调用 ⭐⭐⭐
        print("🔄 开始批量缓存检查...")
        cache_check_result = await _check_batch_cache(data, image_paths, table_type, task_id)

        print("*******cache_check_result*********")
        print(cache_check_result)

        if cache_check_result and cache_check_result.get("success"):
            print("✅ 从批量缓存返回数据，跳过LLM调用")
            return cache_check_result
        else:
            print("🔄 没有找到完整缓存，继续处理")

        # 准备输出文件
        output_file = await _prepare_output_file(data, image_paths, table_type)

        print("output_fileoutput_file:", output_file, table_type)

        # 更新任务状态
        _update_processing_status_start(task_id, image_paths, table_type, output_file)

        # 执行批量处理
        processing_results = await _execute_batch_processing(
            task_id, image_paths, output_file, table_type, processor, data
        )

        # 生成结果报告
        final_result = await _generate_processing_report(
            task_id, processing_results, output_file, table_type
        )

        return final_result

    except Exception as e:
        return await _handle_batch_processing_error(task_id, e)


async def _check_batch_cache(data, image_paths, table_type, task_id):
    """批量缓存检查 - 超简版"""
    try:
        output_file = await _prepare_output_file(data, image_paths, table_type)
        output_path = Path(output_file)

        if not output_path.exists():
            return None

        # ⭐⭐⭐ 只要文件存在且大于1KB就认为缓存有效 ⭐⭐⭐
        if output_path.stat().st_size > 1024:
            excel_url = convert_to_excel_url(output_file)
            return {
                "success": True,
                "data": {
                    "total": len(image_paths),
                    "success": len(image_paths),
                    "failed": 0,
                    "non_financial": len(image_paths),
                    "results": [
                        {
                            "image_path": image_path,
                            "status": "success",
                            "status_text": "从缓存加载",
                            "table_name": f"表格_{i + 1}",
                            "sheet_name": f"表格_{i + 1}",
                            "success": True,
                            "is_non_financial": True,
                            "excel_url": excel_url
                        }
                        for i, image_path in enumerate(image_paths)
                    ],
                    "output_file": output_file,
                    "excel_url": excel_url,
                    "table_type": "non_financial",
                    "processing_completed": True,
                    "message": f"从缓存加载 {len(image_paths)} 个表格",
                    "excel_saved": True,
                    "from_cache": True
                },
                "task_id": task_id,
                "processing_completed": True,
                "from_cache": True
            }
        return None

    except Exception:
        return None


async def _execute_batch_processing(task_id, image_paths, output_file, table_type, processor, data):
    """执行批量处理 - 修复文件写入问题"""
    results = []
    success_count = 0
    failed_count = 0
    bank_name = data.get('bank_name', '未知银行')

    # ⭐⭐⭐ 关键修复：确保输出目录存在 ⭐⭐⭐
    output_path = Path(output_file)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # ⭐⭐⭐ 关键修复：如果输出文件已存在，先备份再处理 ⭐⭐⭐
    if output_path.exists():
        try:
            import shutil
            backup_file = output_path.with_suffix('.backup.xlsx')
            shutil.copy2(output_path, backup_file)
            print(f"✅ 已备份现有文件: {backup_file}")

            # 删除原文件，重新创建
            output_path.unlink()
            print(f"✅ 已删除原文件，准备重新创建: {output_file}")
        except Exception as e:
            print(f"❌ 备份文件失败: {e}")

    # 创建新的Excel文件
    try:
        from backend.service.excel_storage_service import ExcelStorageService
        excel_service = ExcelStorageService()
        excel_service.create_new_excel(output_file)
        print(f"✅ 创建新Excel文件: {output_file}")
    except Exception as e:
        print(f"❌ 创建Excel文件失败: {e}")
        return {
            "results": [],
            "success_count": 0,
            "failed_count": len(image_paths),
            "total_count": len(image_paths)
        }

    for i, image_path in enumerate(image_paths):
        try:
            # 更新进度
            await _update_processing_progress(task_id, i, len(image_paths), image_path)

            # 处理单张图片
            result = await _process_single_image_in_batch(
                image_path, output_file, table_type, processor, bank_name, i
            )

            if result["success"]:
                success_count += 1
                print(f"✅ 处理成功: {Path(image_path).name}")
            else:
                failed_count += 1
                print(f"❌ 处理失败: {Path(image_path).name} - {result.get('error_message', '')}")

            results.append(result)

        except Exception as e:
            error_result = _create_single_image_error_result(image_path, str(e))
            results.append(error_result)
            failed_count += 1
            logger.error(f"处理图片失败 {image_path}: {str(e)}")

    # 最终检查Excel文件
    if output_path.exists():
        print(f"✅ 批量处理完成，Excel文件已生成: {output_file}")
        # 验证Excel文件内容
        try:
            import pandas as pd
            excel_data = pd.read_excel(output_file, sheet_name=None)
            sheet_count = len(excel_data)
            print(f"📊 Excel文件包含 {sheet_count} 个工作表")
            for sheet_name, df in excel_data.items():
                print(f"   - {sheet_name}: {df.shape[0]}行 x {df.shape[1]}列")
        except Exception as e:
            print(f"⚠️ 读取Excel文件失败: {e}")
    else:
        print(f"❌ 批量处理完成，但Excel文件未生成: {output_file}")

    return {
        "results": results,
        "success_count": success_count,
        "failed_count": failed_count,
        "total_count": len(image_paths)
    }


async def _process_single_image_in_batch(image_path, output_file, table_type, processor, bank_name, index):
    """处理批量中的单张图片"""
    sheet_name = f"表格_{index + 1}"
    file_name = Path(image_path).stem

    print(f"🔄 处理第 {index + 1} 张图片: {Path(image_path).name}")

    if table_type == 'non_financial':
        # ⭐⭐⭐ 关键修复：传递索引信息用于调试 ⭐⭐⭐
        result = await _process_single_non_financial_table({
            'image_path': image_path,
            'output_path': output_file,
            'sheet_name': sheet_name,
            'bank_name': bank_name,
            'file_name': file_name,
            'table_index': index
        })

        # ⭐⭐⭐ 添加详细的结果检查 ⭐⭐⭐
        print(f"🔍 批量处理结果检查 - 图片{index + 1}:")
        print(f"   - 成功: {result.get('success')}")
        print(f"   - 表格名: {result.get('table_name', '未获取')}")
        print(f"   - 状态: {result.get('status', '未知')}")

        return result
    else:
        return await _process_financial_table(processor, image_path, output_file, sheet_name, bank_name, file_name)


def _ensure_dataframe_saved_to_excel(df, output_file, sheet_name, table_name):
    """确保 DataFrame 数据被保存到 Excel 文件"""
    try:
        if df is None or df.empty:
            print(f"⚠️ DataFrame 为空，跳过保存: {sheet_name}")
            return False

        output_path = Path(output_file)

        # 使用ExcelStorageService进行保存
        from backend.service.excel_storage_service import ExcelStorageService
        excel_service = ExcelStorageService()

        # 如果文件不存在，先创建
        if not output_path.exists():
            excel_service.create_new_excel(output_file)
            print(f"✅ 创建新Excel文件: {output_file}")

        # 保存数据
        save_success = excel_service.save_dataframe(
            df=df,
            excel_path=output_file,
            sheet_name=sheet_name,
            map_name=table_name
        )

        if save_success:
            print(f"✅ 成功保存工作表: {sheet_name}")
            return True
        else:
            print(f"❌ 保存工作表失败: {sheet_name}")
            return False

    except Exception as e:
        print(f"❌ 保存DataFrame到Excel失败: {e}")
        return False


async def _process_financial_table(processor, image_path, output_file, sheet_name, bank_name, file_name):
    """处理金融表格"""
    try:
        result = await processor.process_table_pipeline(
            image_path=image_path,
            out_file=output_file,
            sheet_name=sheet_name,
            bank_name=bank_name,
            file_name=file_name
        )

        print(f"🔍 金融表格处理结果: {result}")
        print(f"🔍 处理状态: {getattr(result, 'status', '无状态')}")

        status = getattr(result, 'status', 'unknown')
        df_data = getattr(result, 'df', None)

        # ⭐⭐⭐ 关键修复：确保DataFrame被保存到Excel ⭐⭐⭐
        if status == "success" and df_data is not None and not df_data.empty:
            print(f"✅ 获取到金融表格 DataFrame 数据，形状: {df_data.shape}")
            # 立即保存DataFrame到Excel
            save_success = _ensure_dataframe_saved_to_excel(df_data, output_file, sheet_name,
                                                            getattr(result, 'table_name', sheet_name))
            if not save_success:
                print(f"❌ 保存DataFrame到Excel失败: {sheet_name}")
        else:
            print(f"⚠️ 没有获取到DataFrame数据: {sheet_name}")

        # 返回结果
        return {
            "image_path": image_path,
            "status": status,
            "status_text": "成功" if status == "success" else "失败",
            "complexity": getattr(result, 'complexity', 'unknown'),
            "table_name": getattr(result, 'table_name', '未知'),
            "sheet_name": sheet_name,
            "error_message": getattr(result, 'error_message', ''),
            "assessment_reason": getattr(result, 'assessment_reason', ''),
            "success": status == "success",
            "is_non_financial": status == "non_financial",
            "df": df_data
        }

    except Exception as process_error:
        print(f"💥 金融表格处理异常: {str(process_error)}")
        import traceback
        traceback.print_exc()
        return _create_single_image_error_result(image_path, f"处理异常: {str(process_error)}")






