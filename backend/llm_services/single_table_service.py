
import os
import time
import asyncio
import pandas as pd

from pathlib import Path
from flask import jsonify, request

from backend.schemas.table_schemas import ExcelSaveConfig
from backend.llm_services.task_management_service import logger

from backend.llm_services.utils import (
    validate_required_params,
    convert_to_excel_url)

from backend.llm_services.core_service import  (
    get_non_financial_table_service,
    get_table_processor)




async def _process_single_non_financial_table(process_data):
    """处理单个普通表格 - 返回扁平化结果"""
    try:
        image_path = process_data['image_path']
        output_path = process_data['output_path']
        sheet_name = process_data['sheet_name']
        bank_name = process_data['bank_name']
        file_name = process_data['file_name']
        table_index = process_data.get('table_index', 0)

        print(f"处理普通表格 #{table_index}: {image_path}")

        # 调用普通表格识别的核心逻辑
        result = await _process_non_financial_table(process_data)

        print("XCCCCCCCCCCCCCCCCCCCCCZZZZZZZZZZZZprocess_table_pipeline", result)
        print(type(result))

        # ⭐⭐⭐ 关键修复：扁平化结果，避免嵌套结构 ⭐⭐⭐
        if result.get('success') and result.get('data'):
            # 提取内部结果
            inner_data = result['data']
            if inner_data.get('results') and len(inner_data['results']) > 0:
                inner_result = inner_data['results'][0]

                # 返回扁平化结果
                return {
                    "success": True,
                    "image_path": image_path,
                    "status": inner_result.get('status', 'success'),
                    "status_text": inner_result.get('status_text', '成功'),
                    "complexity": inner_result.get('complexity', '普通表格'),
                    "table_name": inner_result.get('table_name', f'表格_{table_index + 1}'),
                    "sheet_name": sheet_name,
                    "error_message": inner_result.get('error_message', ''),
                    "assessment_reason": inner_result.get('assessment_reason', '普通表格模式'),
                    "is_non_financial": inner_result.get('is_non_financial', True),
                    "df": inner_result.get('df'),  # ⭐⭐⭐ 确保包含 DataFrame 数据 ⭐⭐⭐
                    "excel_url": inner_result.get('excel_url', '')
                }

        # 如果处理失败，返回错误
        return {
            "success": False,
            "error": result.get('error', '处理失败'),
            "errorCode": result.get('errorCode', 'PROCESS_FAILED')
        }

    except Exception as e:
        return {
            "success": False,
            "error": f"处理失败: {str(e)}",
            "errorCode": "SYSTEM_ERROR"
        }


def _check_excel_cache(image_path, output_path, sheet_name, table_type="financial"):
    """检查Excel缓存是否存在，如果存在则返回缓存数据"""
    try:
        excel_path = Path(output_path)
        if excel_path.exists():
            print(f"✅ Excel文件已存在，直接读取数据: {output_path}")

            # 读取现有Excel文件
            import pandas as pd
            df = pd.read_excel(output_path, sheet_name=sheet_name)

            # 构建前端可访问的URL
            excel_url = convert_to_excel_url(output_path)

            # 转换DataFrame为前端需要的格式
            headers = list(df.columns) if not df.empty else []
            data_list = []

            if not df.empty:
                for _, row in df.iterrows():
                    row_data = {}
                    for header in headers:
                        cell_value = row[header]
                        row_data[header] = str(cell_value) if pd.notna(cell_value) else ""
                    data_list.append(row_data)

            is_non_financial = table_type == "non_financial"

            # 构建返回结果
            return {
                "success": True,
                "data": {
                    "total": 1,
                    "success": 1,
                    "failed": 0,
                    "non_financial": 1 if is_non_financial else 0,
                    "results": [
                        {
                            "image_path": image_path,
                            "status": "success",
                            "status_text": "从缓存加载",
                            "complexity": "unknown",
                            "table_name": sheet_name,
                            "sheet_name": sheet_name,
                            "error_message": "",
                            "assessment_reason": "从已存在的Excel文件加载",
                            "success": True,
                            "is_non_financial": is_non_financial,
                            "output_file": output_path,
                            "excel_url": excel_url
                        }
                    ],
                    "output_file": output_path,
                    "excel_url": excel_url,
                    "table_type": table_type,
                    "processing_completed": True,
                    "message": "从已存在的Excel文件加载数据",
                    "excel_saved": True,
                    "from_cache": True
                },
                "recognizedData": {
                    "headers": headers,
                    "data": data_list,
                    "tableName": sheet_name,
                    "excelPath": output_path,
                    "fromCache": True
                },
                "task_id": f"cached_{table_type}_{int(time.time())}",
                "processing_completed": True,
                "from_cache": True
            }
        return None  # 表示缓存不存在
    except Exception as e:
        print(f"❌ 读取缓存Excel文件失败: {e}")
        return None  # 读取失败，返回None表示需要重新识别


def recognize_table_internal():
    """识别表格并保存到指定路径 - 先检查Excel是否存在"""
    try:
        data = request.get_json()
        print(f"收到识别请求: {data}")

        if not data:
            return jsonify({
                "success": False,
                "error": "请求体不能为空"
            }), 400

        # 验证必要参数
        required_fields = ['imageUrl', 'excelPath']
        missing_fields = [field for field in required_fields if not data.get(field)]
        if missing_fields:
            return jsonify({
                "success": False,
                "error": f"缺少必要参数: {', '.join(missing_fields)}"
            }), 400

        image_url = data.get('imageUrl')
        excel_path = data.get('excelPath')
        table_name = data.get('tableName', '识别结果')
        index = data.get('index', 0)

        print(f"识别参数 - imageUrl: {image_url}, excelPath: {excel_path}")

        # 构建完整Excel路径
        excel_full_path = Path(excel_path)
        if not excel_full_path.is_absolute():
            excel_full_path = Path.cwd() / excel_path

        print(f"检查Excel文件是否存在: {excel_full_path}")

        # 1. 首先检查Excel文件是否已经存在
        if excel_full_path.exists():
            print("✅ Excel文件已存在，直接读取数据")
            try:
                import openpyxl
                workbook = openpyxl.load_workbook(excel_full_path)
                sheet_name = workbook.sheetnames[0]
                worksheet = workbook[sheet_name]

                # 转换为JSON数据
                headers = []
                data_list = []

                # 读取表头（第一行）
                for col in range(1, worksheet.max_column + 1):
                    cell_value = worksheet.cell(row=1, column=col).value
                    headers.append(str(cell_value) if cell_value is not None else f"列{col}")

                # 读取数据行
                for row in range(2, worksheet.max_row + 1):
                    row_data = {}
                    for col, header in enumerate(headers, 1):
                        cell_value = worksheet.cell(row=row, column=col).value
                        row_data[header] = str(cell_value) if cell_value is not None else ""
                    if any(row_data.values()):  # 只添加非空行
                        data_list.append(row_data)

                # ⭐⭐⭐ 统一返回格式 ⭐⭐⭐
                excel_url = convert_to_excel_url(str(excel_full_path))

                return jsonify({
                    "success": True,
                    "data": {
                        "excelUrl": excel_url,
                        "tableName": table_name,
                        "tableType": "single_table",
                        "fromCache": True,
                        "fileName": excel_full_path.name,
                        "sheetName": sheet_name
                    },
                    "recognizedData": {
                        "headers": headers,
                        "data": data_list,
                        "tableName": table_name,
                        "excelPath": str(excel_full_path),
                        "fromCache": True  # 添加缓存标记
                    },
                    "message": "已加载现有表格数据",
                    "from_cache": True  # 添加缓存标记
                })
            except Exception as e:
                logger.error(f"读取现有Excel失败: {str(e)}")
                print("❌ 读取现有Excel失败，继续走LLM识别")

        # 2. 如果Excel不存在，进行LLM识别
        print("🔄 Excel文件不存在，开始LLM识别流程")

        # 图片路径提取逻辑保持不变...
        image_path = image_url

        if image_url.startswith('http://'):
            from urllib.parse import urlparse
            parsed_url = urlparse(image_url)
            image_path = parsed_url.path
            if image_path.startswith('/'):
                image_path = image_path[1:]
            print(f"✅ 从URL提取路径: {image_path}")

        elif image_url.startswith('/'):
            image_path = image_url[1:]
            print(f"✅ 处理绝对路径: {image_path}")

        print(f"处理图片路径: {image_path}")

        # 检查图片文件是否存在
        image_full_path = Path(image_path)
        if not image_full_path.exists():
            static_path = Path("static") / image_path
            if static_path.exists():
                image_full_path = static_path
                print(f"✅ 在static目录找到图片: {static_path}")
            else:
                current_dir_path = Path.cwd() / image_path
                if current_dir_path.exists():
                    image_full_path = current_dir_path
                    print(f"✅ 在当前目录找到图片: {current_dir_path}")
                else:
                    print(f"❌ 图片文件不存在，尝试的路径:")
                    print(f"   - {image_full_path}")
                    print(f"   - {static_path}")
                    print(f"   - {current_dir_path}")
                    return jsonify({
                        "success": False,
                        "error": f"图片文件不存在: {image_path}"
                    }), 404

        print(f"✅ 最终使用的图片路径: {image_full_path}")

        # 确保Excel目录存在
        excel_full_path.parent.mkdir(parents=True, exist_ok=True)

        print(f"Excel保存路径: {excel_full_path}")

        # 调用 _process_single_image 异步函数进行LLM识别
        process_data = {
            'image_path': str(image_full_path),
            'output_path': str(excel_full_path),
            'sheet_name': table_name,
            'bank_name': '未知银行',
            'file_name': f"table_{index + 1}"
        }

        # ⭐⭐⭐ 修复异步调用 ⭐⭐⭐
        result = asyncio.run(_process_single_image(process_data))
        print(f"LLM识别结果: {result}")

        if result.get('success'):
            # ⭐⭐⭐ 直接返回统一格式的结果 ⭐⭐⭐
            return jsonify(result)
        else:
            return jsonify({
                "success": False,
                "error": result.get('error', '识别失败'),
                "errorCode": result.get('errorCode', 'UNKNOWN_ERROR')
            }), 500

    except Exception as e:
        logger.error(f"表格识别失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": f"表格识别失败: {str(e)}",
            "errorCode": "SYSTEM_ERROR"
        }), 500


async def _process_single_image_uni_1(data):
    """处理单张图片的异步函数"""
    try:
        # 验证必要参数
        is_valid, error_msg = validate_required_params(
            data, ['image_path', 'output_path']
        )

        print("###################")
        print(is_valid, error_msg)

        if not is_valid:
            return {
                "success": False,
                "error": error_msg
            }

        image_path = data.get('image_path')
        output_path = data.get('output_path')
        sheet_name = data.get('sheet_name', '识别结果')
        bank_name = data.get('bank_name', '未知银行')
        file_name = data.get('file_name', Path(image_path).stem)

        print("************image_path****output_path****************")
        print(image_path)
        print(output_path)
        print(sheet_name)
        print(bank_name)
        print(file_name)

        # 检查图片文件是否存在
        if not Path(image_path).exists():
            return {
                "success": False,
                "error": f"图片文件不存在: {image_path}"
            }

        processor = get_table_processor()

        print("****************processor.llm_client******************")
        print(processor.llm_client)

        if not processor.llm_client:
            return {
                "success": False,
                "error": "请先配置LLM客户端"
            }

        # 使用已有的标识创建Excel存储文件夹
        folder_name = Path(image_path).stem.split('_')[0]

        excel_base_dir = Path("static/excel_data")
        excel_dir = excel_base_dir / folder_name
        excel_dir.mkdir(parents=True, exist_ok=True)

        # 重新构建输出路径
        excel_filename = Path(output_path).name
        new_output_path = excel_dir / excel_filename
        output_path = str(new_output_path)

        print(f"Excel文件将保存到: {output_path}")

        # 配置Excel保存参数
        excel_config = ExcelSaveConfig(
            anchor_cell=data.get('anchor_cell', 'R2'),
            width_px=data.get('width_px', 768),
            mode=data.get('mode', 'overwrite')
        )

        # 处理图片
        result = await processor.process_table_pipeline(
            image_path=image_path,
            out_file=output_path,
            sheet_name=sheet_name,
            bank_name=bank_name,
            file_name=file_name,
            excel_config=excel_config
        )

        print(f"单张图片result: {result}")

        # ⭐⭐⭐ 关键修复：确保生成有效的Excel文件 ⭐⭐⭐
        excel_exists = Path(output_path).exists()

        # 如果Excel文件不存在，但处理成功，创建一个基础Excel文件
        if not excel_exists and (result.status == "success" or result.status == "non_financial"):
            try:
                import pandas as pd
                from backend.service.excel_storage_service import ExcelStorageService

                excel_service = ExcelStorageService()
                # 创建基础Excel文件
                excel_service.create_new_excel(output_path)

                # 添加基础信息工作表
                basic_data = {
                    '信息类型': ['处理状态', '表格类型', '复杂度', '评估原因', '图片路径'],
                    '值': [
                        result.status,
                        '非金融表格' if result.status == 'non_financial' else '金融表格',
                        result.complexity,
                        result.assessment_reason,
                        image_path
                    ]
                }
                basic_df = pd.DataFrame(basic_data)

                excel_service.save_dataframe(
                    df=basic_df,
                    excel_path=output_path,
                    sheet_name=sheet_name,
                    map_name=result.table_name or f"表格_{data.get('file_name', 'unknown')}",
                    image_data=image_path if Path(image_path).exists() else None
                )

                excel_exists = True
                print("✅ 已创建基础Excel文件")

            except Exception as create_error:
                print(f"❌ 创建基础Excel文件失败: {create_error}")

        # ⭐⭐⭐ 统一返回格式 ⭐⭐⭐
        if (result.status == "success" or result.status == "non_financial") and excel_exists:
            # 构建前端可访问的URL
            excel_url = convert_to_excel_url(output_path)

            # 判断是否为非金融表格
            is_non_financial = result.status == "non_financial"

            # 构建与批量处理一致的返回格式
            return {
                "success": True,
                "data": {
                    # 基础信息（与批量处理统一）
                    "total": 1,
                    "success": 1,
                    "failed": 0,
                    "non_financial": 1 if is_non_financial else 0,
                    "results": [
                        {
                            "image_path": image_path,
                            "status": result.status,
                            "status_text": "成功" if result.status == "success" else "非金融表格",
                            "complexity": result.complexity,
                            "table_name": result.table_name or f"表格_{data.get('file_name', 'unknown')}",
                            "sheet_name": sheet_name,
                            "error_message": result.error_message,
                            "assessment_reason": result.assessment_reason,
                            "success": True,
                            "is_non_financial": is_non_financial,
                            "output_file": output_path,  # ⭐⭐⭐ 确保包含输出文件路径 ⭐⭐⭐
                            "excel_url": excel_url
                        }
                    ],
                    "output_file": output_path,
                    "excel_url": excel_url,
                    "table_type": "financial" if not is_non_financial else "non_financial",
                    "processing_completed": True,
                    "message": "单张图片处理完成",
                    "excel_saved": True
                },
                "task_id": f"single_{int(time.time())}",
                "processing_completed": True
            }
        else:
            return {
                "success": False,
                "error": result.error_message or "表格处理失败",
                "errorCode": "PROCESS_FAILED"
            }

    except Exception as e:
        return {
            "success": False,
            "error": f"处理失败: {str(e)}",
            "errorCode": "SYSTEM_ERROR"
        }


async def _process_non_financial_table_uni_1(data):
    """处理普通表格的异步函数"""
    try:
        # 验证必要参数
        is_valid, error_msg = validate_required_params(
            data, ['image_path', 'output_path']
        )

        if not is_valid:
            return {
                "success": False,
                "error": error_msg
            }

        image_path = data.get('image_path')
        output_path = data.get('output_path')
        sheet_name = data.get('sheet_name', '普通表格识别结果')
        bank_name = data.get('bank_name', '未知机构')
        file_name = data.get('file_name', Path(image_path).stem)

        print("普通表格处理参数:")
        print(f"image_path: {image_path}")
        print(f"output_path: {output_path}")
        print(f"sheet_name: {sheet_name}")
        print(f"bank_name: {bank_name}")
        print(f"file_name: {file_name}")

        # 检查图片文件是否存在
        if not Path(image_path).exists():
            return {
                "success": False,
                "error": f"图片文件不存在: {image_path}"
            }

        service = get_non_financial_table_service()

        if not service.llm_client:
            return {
                "success": False,
                "error": "请先配置LLM客户端"
            }

        # 使用已有的标识创建Excel存储文件夹
        folder_name = Path(image_path).stem.split('_')[0]
        excel_base_dir = Path("static/excel_data")
        excel_dir = excel_base_dir / folder_name
        excel_dir.mkdir(parents=True, exist_ok=True)

        # 重新构建输出路径
        excel_filename = Path(output_path).name
        new_output_path = excel_dir / excel_filename
        output_path = str(new_output_path)

        print(f"普通表格Excel文件将保存到: {output_path}")

        # 配置Excel保存参数
        excel_config = ExcelSaveConfig(
            anchor_cell=data.get('anchor_cell', 'R2'),
            width_px=data.get('width_px', 768),
            mode=data.get('mode', 'overwrite')
        )

        # 处理普通表格
        result = await service.process_table_pipeline(
            image_path=image_path,
            out_file=output_path,
            sheet_name=sheet_name,
            bank_name=bank_name,
            file_name=file_name,
            excel_config=excel_config
        )

        # ⭐⭐⭐ 关键修复：确保获取 DataFrame 数据但不直接返回 ⭐⭐⭐
        df_data = None
        if hasattr(result, 'df') and result.df is not None:
            df_data = result.df
            print(f"✅ 普通表格获取到 DataFrame 数据，形状: {df_data.shape}")
        else:
            print("⚠️ 普通表格没有获取到 DataFrame 数据，尝试从Excel文件读取")
            # 如果result中没有df，尝试从生成的Excel文件读取
            try:
                if Path(output_path).exists():
                    import pandas as pd
                    df_data = pd.read_excel(output_path, sheet_name=sheet_name)
                    print(f"✅ 从Excel文件读取到 DataFrame 数据，形状: {df_data.shape}")
            except Exception as e:
                print(f"❌ 从Excel文件读取数据失败: {e}")

        # ⭐⭐⭐ 统一返回格式 - 改为批量处理模式 ⭐⭐⭐
        if result.status == "success" and Path(output_path).exists():
            # 构建前端可访问的URL
            excel_url = convert_to_excel_url(output_path)

            # ⭐⭐⭐ 关键修复：不直接返回 DataFrame 对象，而是保存到文件 ⭐⭐⭐
            # 如果有 DataFrame 数据，先保存到 Excel
            if df_data is not None and not df_data.empty:
                try:
                    # 使用 excel_storage_service 保存数据
                    from backend.service.excel_storage_service import ExcelStorageService
                    excel_service = ExcelStorageService()
                    excel_service.save_dataframe(
                        df=df_data,
                        excel_path=output_path,
                        sheet_name=sheet_name,
                        map_name=result.table_name,
                        image_data=image_path if image_path and os.path.exists(image_path) else None
                    )
                    print(f"✅ 已保存 DataFrame 数据到 Excel 文件")
                except Exception as save_error:
                    print(f"❌ 保存 DataFrame 数据失败: {save_error}")

            # 构建与批量处理一致的返回格式
            batch_format_result = {
                "success": True,
                "data": {
                    # 基础信息（与批量处理统一）
                    "total": 1,
                    "success": 1,
                    "failed": 0,
                    "non_financial": 1,  # 普通表格都是非金融表格
                    "results": [
                        {
                            "image_path": image_path,
                            "status": result.status,
                            "status_text": "成功",
                            "complexity": result.complexity,
                            "table_name": result.table_name,
                            "sheet_name": sheet_name,
                            "error_message": result.error_message,
                            "assessment_reason": result.assessment_reason,
                            "success": True,
                            "is_non_financial": True,
                            "excel_url": excel_url
                            # ⭐⭐⭐ 不再直接返回 df 字段，避免序列化问题 ⭐⭐⭐
                        }
                    ],
                    "output_file": output_path,
                    "excel_url": excel_url,
                    "table_type": "non_financial",
                    "processing_completed": True,
                    "message": "普通表格处理完成",
                    "excel_saved": True
                },
                "task_id": f"non_financial_{int(time.time())}",
                "processing_completed": True
            }

            # ⭐⭐⭐ 同时提供旧格式的excel_url字段以保持兼容性 ⭐⭐⭐
            batch_format_result["excel_url"] = excel_url
            batch_format_result["from_cache"] = False
            batch_format_result["tableName"] = f"{result.table_name} - {folder_name}"

            return batch_format_result
        else:
            return {
                "success": False,
                "error": result.error_message or "表格处理失败",
                "errorCode": "PROCESS_FAILED"
            }

    except Exception as e:
        return {
            "success": False,
            "error": f"处理失败: {str(e)}",
            "errorCode": "SYSTEM_ERROR"
        }


async def _process_single_image_2(data):
    """处理单张图片的异步函数 - 添加Excel存在性检查"""
    try:
        # 验证必要参数
        is_valid, error_msg = validate_required_params(
            data, ['image_path', 'output_path']
        )

        print("###################")
        print(is_valid, error_msg)

        if not is_valid:
            return {
                "success": False,
                "error": error_msg
            }

        image_path = data.get('image_path')
        output_path = data.get('output_path')
        sheet_name = data.get('sheet_name', '识别结果')
        bank_name = data.get('bank_name', '未知银行')
        file_name = data.get('file_name', Path(image_path).stem)

        print("************image_path****output_path****************")
        print(image_path)
        print(output_path)
        print(sheet_name)
        print(bank_name)
        print(file_name)

        # 检查图片文件是否存在
        if not Path(image_path).exists():
            return {
                "success": False,
                "error": f"图片文件不存在: {image_path}"
            }

        # ⭐⭐⭐ 新增：首先检查Excel文件是否已经存在 ⭐⭐⭐
        excel_path = Path(output_path)
        if excel_path.exists():
            print(f"✅ Excel文件已存在，直接读取数据: {output_path}")
            try:
                # 读取现有Excel文件
                import pandas as pd
                df = pd.read_excel(output_path, sheet_name=sheet_name)

                # 构建前端可访问的URL
                excel_url = convert_to_excel_url(output_path)

                # ⭐⭐⭐ 关键修复：构建完整的数据结构，包含 recognizedData ⭐⭐⭐
                # 转换DataFrame为前端需要的格式
                headers = list(df.columns) if not df.empty else []
                data_list = []

                if not df.empty:
                    for _, row in df.iterrows():
                        row_data = {}
                        for header in headers:
                            cell_value = row[header]
                            row_data[header] = str(cell_value) if pd.notna(cell_value) else ""
                        data_list.append(row_data)

                # 返回现有数据 - 包含完整的 recognizedData
                return {
                    "success": True,
                    "data": {
                        "total": 1,
                        "success": 1,
                        "failed": 0,
                        "non_financial": 0,
                        "results": [
                            {
                                "image_path": image_path,
                                "status": "success",
                                "status_text": "从缓存加载",
                                "complexity": "unknown",
                                "table_name": sheet_name,
                                "sheet_name": sheet_name,
                                "error_message": "",
                                "assessment_reason": "从已存在的Excel文件加载",
                                "success": True,
                                "is_non_financial": False,
                                "output_file": output_path,
                                "excel_url": excel_url
                            }
                        ],
                        "output_file": output_path,
                        "excel_url": excel_url,
                        "table_type": "financial",
                        "processing_completed": True,
                        "message": "从已存在的Excel文件加载数据",
                        "excel_saved": True,
                        "from_cache": True
                    },
                    # ⭐⭐⭐ 关键修复：添加 recognizedData 字段 ⭐⭐⭐
                    "recognizedData": {
                        "headers": headers,
                        "data": data_list,
                        "tableName": sheet_name,
                        "excelPath": output_path,
                        "fromCache": True
                    },
                    "task_id": f"cached_{int(time.time())}",
                    "processing_completed": True,
                    "from_cache": True
                }
            except Exception as e:
                print(f"❌ 读取现有Excel文件失败，将重新识别: {e}")


        # ⭐⭐⭐ 如果Excel不存在，继续原有的LLM识别流程 ⭐⭐⭐
        processor = get_table_processor()

        print("****************processor.llm_client******************")
        print(processor.llm_client)

        if not processor.llm_client:
            return {
                "success": False,
                "error": "请先配置LLM客户端"
            }

        # 使用已有的标识创建Excel存储文件夹
        folder_name = Path(image_path).stem.split('_')[0]

        excel_base_dir = Path("static/excel_data")
        excel_dir = excel_base_dir / folder_name
        excel_dir.mkdir(parents=True, exist_ok=True)

        # 重新构建输出路径
        excel_filename = Path(output_path).name
        new_output_path = excel_dir / excel_filename
        output_path = str(new_output_path)

        print(f"Excel文件将保存到: {output_path}")

        # 配置Excel保存参数
        excel_config = ExcelSaveConfig(
            anchor_cell=data.get('anchor_cell', 'R2'),
            width_px=data.get('width_px', 768),
            mode=data.get('mode', 'overwrite')
        )

        # 处理图片
        result = await processor.process_table_pipeline(
            image_path=image_path,
            out_file=output_path,
            sheet_name=sheet_name,
            bank_name=bank_name,
            file_name=file_name,
            excel_config=excel_config
        )

        print(f"单张图片result: {result}")

        # ⭐⭐⭐ 关键修复：确保生成有效的Excel文件 ⭐⭐⭐
        excel_exists = Path(output_path).exists()

        # 如果Excel文件不存在，但处理成功，创建一个基础Excel文件
        if not excel_exists and (result.status == "success" or result.status == "non_financial"):
            try:
                import pandas as pd
                from backend.service.excel_storage_service import ExcelStorageService

                excel_service = ExcelStorageService()
                # 创建基础Excel文件
                excel_service.create_new_excel(output_path)

                # 添加基础信息工作表
                basic_data = {
                    '信息类型': ['处理状态', '表格类型', '复杂度', '评估原因', '图片路径'],
                    '值': [
                        result.status,
                        '非金融表格' if result.status == 'non_financial' else '金融表格',
                        result.complexity,
                        result.assessment_reason,
                        image_path
                    ]
                }
                basic_df = pd.DataFrame(basic_data)

                excel_service.save_dataframe(
                    df=basic_df,
                    excel_path=output_path,
                    sheet_name=sheet_name,
                    map_name=result.table_name or f"表格_{data.get('file_name', 'unknown')}",
                    image_data=image_path if Path(image_path).exists() else None
                )

                excel_exists = True
                print("✅ 已创建基础Excel文件")

            except Exception as create_error:
                print(f"❌ 创建基础Excel文件失败: {create_error}")

        # ⭐⭐⭐ 统一返回格式 ⭐⭐⭐
        if (result.status == "success" or result.status == "non_financial") and excel_exists:
            # 构建前端可访问的URL
            excel_url = convert_to_excel_url(output_path)

            # 判断是否为非金融表格
            is_non_financial = result.status == "non_financial"

            # 构建与批量处理一致的返回格式
            return {
                "success": True,
                "data": {
                    # 基础信息（与批量处理统一）
                    "total": 1,
                    "success": 1,
                    "failed": 0,
                    "non_financial": 1 if is_non_financial else 0,
                    "results": [
                        {
                            "image_path": image_path,
                            "status": result.status,
                            "status_text": "成功" if result.status == "success" else "非金融表格",
                            "complexity": result.complexity,
                            "table_name": result.table_name or f"表格_{data.get('file_name', 'unknown')}",
                            "sheet_name": sheet_name,
                            "error_message": result.error_message,
                            "assessment_reason": result.assessment_reason,
                            "success": True,
                            "is_non_financial": is_non_financial,
                            "output_file": output_path,  # ⭐⭐⭐ 确保包含输出文件路径 ⭐⭐⭐
                            "excel_url": excel_url
                        }
                    ],
                    "output_file": output_path,
                    "excel_url": excel_url,
                    "table_type": "financial" if not is_non_financial else "non_financial",
                    "processing_completed": True,
                    "message": "单张图片处理完成",
                    "excel_saved": True,
                    "from_cache": False  # 新增标记，表示重新识别
                },
                "task_id": f"single_{int(time.time())}",
                "processing_completed": True,
                "from_cache": False  # 新增标记
            }
        else:
            return {
                "success": False,
                "error": result.error_message or "表格处理失败",
                "errorCode": "PROCESS_FAILED"
            }

    except Exception as e:
        return {
            "success": False,
            "error": f"处理失败: {str(e)}",
            "errorCode": "SYSTEM_ERROR"
        }


async def _process_single_image(data):
    """处理单张图片的异步函数 - 添加Excel存在性检查"""
    try:
        # 验证必要参数
        is_valid, error_msg = validate_required_params(
            data, ['image_path', 'output_path']
        )

        print("###################")
        print(is_valid, error_msg)

        if not is_valid:
            return {
                "success": False,
                "error": error_msg
            }

        image_path = data.get('image_path')
        output_path = data.get('output_path')
        sheet_name = data.get('sheet_name', '识别结果')
        bank_name = data.get('bank_name', '未知银行')
        file_name = data.get('file_name', Path(image_path).stem)

        print("************image_path****output_path****************")
        print(image_path)
        print(output_path)
        print(sheet_name)
        print(bank_name)
        print(file_name)

        # 检查图片文件是否存在
        if not Path(image_path).exists():
            return {
                "success": False,
                "error": f"图片文件不存在: {image_path}"
            }

        # ⭐⭐⭐ 首先检查Excel缓存是否存在 ⭐⭐⭐
        cache_result = _check_excel_cache(image_path, output_path, sheet_name, "financial")
        if cache_result:
            return cache_result

        # ⭐⭐⭐ 如果Excel不存在，继续原有的LLM识别流程 ⭐⭐⭐
        processor = get_table_processor()

        print("****************processor.llm_client******************")
        print(processor.llm_client)

        if not processor.llm_client:
            return {
                "success": False,
                "error": "请先配置LLM客户端"
            }

        # 使用已有的标识创建Excel存储文件夹
        folder_name = Path(image_path).stem.split('_')[0]

        excel_base_dir = Path("static/excel_data")
        excel_dir = excel_base_dir / folder_name
        excel_dir.mkdir(parents=True, exist_ok=True)

        # 重新构建输出路径
        excel_filename = Path(output_path).name
        new_output_path = excel_dir / excel_filename
        output_path = str(new_output_path)

        print(f"Excel文件将保存到: {output_path}")

        # 配置Excel保存参数
        excel_config = ExcelSaveConfig(
            anchor_cell=data.get('anchor_cell', 'R2'),
            width_px=data.get('width_px', 768),
            mode=data.get('mode', 'overwrite')
        )

        # 处理图片
        result = await processor.process_table_pipeline(
            image_path=image_path,
            out_file=output_path,
            sheet_name=sheet_name,
            bank_name=bank_name,
            file_name=file_name,
            excel_config=excel_config
        )

        print(f"单张图片result: {result}")

        # ⭐⭐⭐ 关键修复：确保生成有效的Excel文件 ⭐⭐⭐
        excel_exists = Path(output_path).exists()

        # 如果Excel文件不存在，但处理成功，创建一个基础Excel文件
        if not excel_exists and (result.status == "success" or result.status == "non_financial"):
            try:
                import pandas as pd
                from backend.service.excel_storage_service import ExcelStorageService

                excel_service = ExcelStorageService()
                # 创建基础Excel文件
                excel_service.create_new_excel(output_path)

                # 添加基础信息工作表
                basic_data = {
                    '信息类型': ['处理状态', '表格类型', '复杂度', '评估原因', '图片路径'],
                    '值': [
                        result.status,
                        '非金融表格' if result.status == 'non_financial' else '金融表格',
                        result.complexity,
                        result.assessment_reason,
                        image_path
                    ]
                }
                basic_df = pd.DataFrame(basic_data)

                excel_service.save_dataframe(
                    df=basic_df,
                    excel_path=output_path,
                    sheet_name=sheet_name,
                    map_name=result.table_name or f"表格_{data.get('file_name', 'unknown')}",
                    image_data=image_path if Path(image_path).exists() else None
                )

                excel_exists = True
                print("✅ 已创建基础Excel文件")

            except Exception as create_error:
                print(f"❌ 创建基础Excel文件失败: {create_error}")

        # ⭐⭐⭐ 统一返回格式 ⭐⭐⭐
        if (result.status == "success" or result.status == "non_financial") and excel_exists:
            # 构建前端可访问的URL
            excel_url = convert_to_excel_url(output_path)

            # 判断是否为非金融表格
            is_non_financial = result.status == "non_financial"

            # 构建与批量处理一致的返回格式
            return {
                "success": True,
                "data": {
                    # 基础信息（与批量处理统一）
                    "total": 1,
                    "success": 1,
                    "failed": 0,
                    "non_financial": 1 if is_non_financial else 0,
                    "results": [
                        {
                            "image_path": image_path,
                            "status": result.status,
                            "status_text": "成功" if result.status == "success" else "非金融表格",
                            "complexity": result.complexity,
                            "table_name": result.table_name or f"表格_{data.get('file_name', 'unknown')}",
                            "sheet_name": sheet_name,
                            "error_message": result.error_message,
                            "assessment_reason": result.assessment_reason,
                            "success": True,
                            "is_non_financial": is_non_financial,
                            "output_file": output_path,
                            "excel_url": excel_url
                        }
                    ],
                    "output_file": output_path,
                    "excel_url": excel_url,
                    "table_type": "financial" if not is_non_financial else "non_financial",
                    "processing_completed": True,
                    "message": "单张图片处理完成",
                    "excel_saved": True,
                    "from_cache": False  # 明确标记不是从缓存加载
                },
                "task_id": f"single_{int(time.time())}",
                "processing_completed": True,
                "from_cache": False
            }
        else:
            return {
                "success": False,
                "error": result.error_message or "表格处理失败",
                "errorCode": "PROCESS_FAILED"
            }

    except Exception as e:
        return {
            "success": False,
            "error": f"处理失败: {str(e)}",
            "errorCode": "SYSTEM_ERROR"
        }


async def _process_non_financial_table1(data):
    """处理普通表格的异步函数 - 添加Excel存在性检查"""
    try:
        # 验证必要参数
        is_valid, error_msg = validate_required_params(
            data, ['image_path', 'output_path']
        )

        if not is_valid:
            return {
                "success": False,
                "error": error_msg
            }

        image_path = data.get('image_path')
        output_path = data.get('output_path')
        sheet_name = data.get('sheet_name', '普通表格识别结果')
        bank_name = data.get('bank_name', '未知机构')
        file_name = data.get('file_name', Path(image_path).stem)

        print("普通表格处理参数:")
        print(f"image_path: {image_path}")
        print(f"output_path: {output_path}")
        print(f"sheet_name: {sheet_name}")
        print(f"bank_name: {bank_name}")
        print(f"file_name: {file_name}")

        # 检查图片文件是否存在
        if not Path(image_path).exists():
            return {
                "success": False,
                "error": f"图片文件不存在: {image_path}"
            }

        # ⭐⭐⭐ 新增：首先检查Excel文件是否已经存在 ⭐⭐⭐
        excel_path = Path(output_path)
        if excel_path.exists():
            print(f"✅ 普通表格Excel文件已存在，直接读取数据: {output_path}")
            try:
                # 读取现有Excel文件
                import pandas as pd
                df = pd.read_excel(output_path, sheet_name=sheet_name)

                # 构建前端可访问的URL
                excel_url = convert_to_excel_url(output_path)

                # ⭐⭐⭐ 关键修复：构建完整的数据结构 ⭐⭐⭐
                # 转换DataFrame为前端需要的格式
                headers = list(df.columns) if not df.empty else []
                data_list = []

                if not df.empty:
                    for _, row in df.iterrows():
                        row_data = {}
                        for header in headers:
                            cell_value = row[header]
                            row_data[header] = str(cell_value) if pd.notna(cell_value) else ""
                        data_list.append(row_data)

                # 返回现有数据 - 包含完整的 recognizedData
                return {
                    "success": True,
                    "data": {
                        "total": 1,
                        "success": 1,
                        "failed": 0,
                        "non_financial": 1,
                        "results": [
                            {
                                "image_path": image_path,
                                "status": "success",
                                "status_text": "从缓存加载",
                                "complexity": "unknown",
                                "table_name": sheet_name,
                                "sheet_name": sheet_name,
                                "error_message": "",
                                "assessment_reason": "从已存在的Excel文件加载",
                                "success": True,
                                "is_non_financial": True,
                                "excel_url": excel_url
                            }
                        ],
                        "output_file": output_path,
                        "excel_url": excel_url,
                        "table_type": "non_financial",
                        "processing_completed": True,
                        "message": "从已存在的Excel文件加载普通表格数据",
                        "excel_saved": True,
                        "from_cache": True
                    },
                    # ⭐⭐⭐ 关键修复：添加 recognizedData 字段 ⭐⭐⭐
                    "recognizedData": {
                        "headers": headers,
                        "data": data_list,
                        "tableName": sheet_name,
                        "excelPath": output_path,
                        "fromCache": True
                    },
                    "task_id": f"non_financial_cached_{int(time.time())}",
                    "processing_completed": True,
                    "from_cache": True
                }
            except Exception as e:
                print(f"❌ 读取现有普通表格Excel文件失败，将重新识别: {e}")
                # 如果读取失败，继续执行LLM识别流程

        # ⭐⭐⭐ 如果Excel不存在，继续原有的LLM识别流程 ⭐⭐⭐
        service = get_non_financial_table_service()

        if not service.llm_client:
            return {
                "success": False,
                "error": "请先配置LLM客户端"
            }

        # 使用已有的标识创建Excel存储文件夹
        folder_name = Path(image_path).stem.split('_')[0]
        excel_base_dir = Path("static/excel_data")
        excel_dir = excel_base_dir / folder_name
        excel_dir.mkdir(parents=True, exist_ok=True)

        # 重新构建输出路径
        excel_filename = Path(output_path).name
        new_output_path = excel_dir / excel_filename
        output_path = str(new_output_path)

        print(f"普通表格Excel文件将保存到: {output_path}")

        # 配置Excel保存参数
        excel_config = ExcelSaveConfig(
            anchor_cell=data.get('anchor_cell', 'R2'),
            width_px=data.get('width_px', 768),
            mode=data.get('mode', 'overwrite')
        )

        # 处理普通表格
        result = await service.process_table_pipeline(
            image_path=image_path,
            out_file=output_path,
            sheet_name=sheet_name,
            bank_name=bank_name,
            file_name=file_name,
            excel_config=excel_config
        )

        # ⭐⭐⭐ 关键修复：确保获取 DataFrame 数据但不直接返回 ⭐⭐⭐
        df_data = None
        if hasattr(result, 'df') and result.df is not None:
            df_data = result.df
            print(f"✅ 普通表格获取到 DataFrame 数据，形状: {df_data.shape}")
        else:
            print("⚠️ 普通表格没有获取到 DataFrame 数据，尝试从Excel文件读取")
            # 如果result中没有df，尝试从生成的Excel文件读取
            try:
                if Path(output_path).exists():
                    import pandas as pd
                    df_data = pd.read_excel(output_path, sheet_name=sheet_name)
                    print(f"✅ 从Excel文件读取到 DataFrame 数据，形状: {df_data.shape}")
            except Exception as e:
                print(f"❌ 从Excel文件读取数据失败: {e}")

        # ⭐⭐⭐ 统一返回格式 - 改为批量处理模式 ⭐⭐⭐
        if result.status == "success" and Path(output_path).exists():
            # 构建前端可访问的URL
            excel_url = convert_to_excel_url(output_path)

            # ⭐⭐⭐ 关键修复：不直接返回 DataFrame 对象，而是保存到文件 ⭐⭐⭐
            # 如果有 DataFrame 数据，先保存到 Excel
            if df_data is not None and not df_data.empty:
                try:
                    # 使用 excel_storage_service 保存数据
                    from backend.service.excel_storage_service import ExcelStorageService
                    excel_service = ExcelStorageService()
                    excel_service.save_dataframe(
                        df=df_data,
                        excel_path=output_path,
                        sheet_name=sheet_name,
                        map_name=result.table_name,
                        image_data=image_path if image_path and os.path.exists(image_path) else None
                    )
                    print(f"✅ 已保存 DataFrame 数据到 Excel 文件")
                except Exception as save_error:
                    print(f"❌ 保存 DataFrame 数据失败: {save_error}")

            # 构建与批量处理一致的返回格式
            batch_format_result = {
                "success": True,
                "data": {
                    # 基础信息（与批量处理统一）
                    "total": 1,
                    "success": 1,
                    "failed": 0,
                    "non_financial": 1,  # 普通表格都是非金融表格
                    "results": [
                        {
                            "image_path": image_path,
                            "status": result.status,
                            "status_text": "成功",
                            "complexity": result.complexity,
                            "table_name": result.table_name,
                            "sheet_name": sheet_name,
                            "error_message": result.error_message,
                            "assessment_reason": result.assessment_reason,
                            "success": True,
                            "is_non_financial": True,
                            "excel_url": excel_url
                            # ⭐⭐⭐ 不再直接返回 df 字段，避免序列化问题 ⭐⭐⭐
                        }
                    ],
                    "output_file": output_path,
                    "excel_url": excel_url,
                    "table_type": "non_financial",
                    "processing_completed": True,
                    "message": "普通表格处理完成",
                    "excel_saved": True,
                    "from_cache": False  # 新增标记
                },
                "task_id": f"non_financial_{int(time.time())}",
                "processing_completed": True,
                "from_cache": False  # 新增标记
            }

            # ⭐⭐⭐ 同时提供旧格式的excel_url字段以保持兼容性 ⭐⭐⭐
            batch_format_result["excel_url"] = excel_url
            batch_format_result["from_cache"] = False
            batch_format_result["tableName"] = f"{result.table_name} - {folder_name}"

            return batch_format_result
        else:
            return {
                "success": False,
                "error": result.error_message or "表格处理失败",
                "errorCode": "PROCESS_FAILED"
            }

    except Exception as e:
        return {
            "success": False,
            "error": f"处理失败: {str(e)}",
            "errorCode": "SYSTEM_ERROR"
        }


async def _process_non_financial_table(data):
    """处理普通表格的异步函数 - 添加Excel存在性检查"""
    try:
        # 验证必要参数
        is_valid, error_msg = validate_required_params(
            data, ['image_path', 'output_path']
        )

        if not is_valid:
            return {
                "success": False,
                "error": error_msg
            }

        image_path = data.get('image_path')
        output_path = data.get('output_path')
        sheet_name = data.get('sheet_name', '普通表格识别结果')
        bank_name = data.get('bank_name', '未知机构')
        file_name = data.get('file_name', Path(image_path).stem)

        print("普通表格处理参数:")
        print(f"image_path: {image_path}")
        print(f"output_path: {output_path}")
        print(f"sheet_name: {sheet_name}")
        print(f"bank_name: {bank_name}")
        print(f"file_name: {file_name}")

        # 检查图片文件是否存在
        if not Path(image_path).exists():
            return {
                "success": False,
                "error": f"图片文件不存在: {image_path}"
            }

        # ⭐⭐⭐ 关键修复：统一Excel路径处理 ⭐⭐⭐
        # 使用已有的标识创建Excel存储文件夹
        folder_name = Path(image_path).stem.split('_')[0]
        excel_base_dir = Path("static/excel_data")
        excel_dir = excel_base_dir / folder_name
        excel_dir.mkdir(parents=True, exist_ok=True)

        # 重新构建输出路径（确保路径一致性）
        excel_filename = Path(output_path).name
        new_output_path = excel_dir / excel_filename
        final_output_path = str(new_output_path)

        print(f"普通表格Excel文件将保存到: {final_output_path}")

        # ⭐⭐⭐ 首先检查Excel文件是否已经存在（使用最终路径）⭐⭐⭐
        excel_path = Path(final_output_path)
        if excel_path.exists():
            print(f"✅ 普通表格Excel文件已存在，直接读取数据: {final_output_path}")
            try:
                # 读取现有Excel文件
                import pandas as pd
                df = pd.read_excel(final_output_path, sheet_name=sheet_name)

                # 构建前端可访问的URL
                excel_url = convert_to_excel_url(final_output_path)

                # 构建完整的数据结构
                cache_result = _build_cache_response(
                    image_path=image_path,
                    output_path=final_output_path,
                    sheet_name=sheet_name,
                    excel_url=excel_url,
                    df=df,
                    table_type="non_financial"
                )
                return cache_result

            except Exception as e:
                print(f"❌ 读取现有普通表格Excel文件失败，将重新识别: {e}")
                # 如果读取失败，继续执行LLM识别流程

        # ⭐⭐⭐ 如果Excel不存在，继续原有的LLM识别流程 ⭐⭐⭐
        service = get_non_financial_table_service()

        if not service.llm_client:
            return {
                "success": False,
                "error": "请先配置LLM客户端"
            }

        # 配置Excel保存参数
        excel_config = ExcelSaveConfig(
            anchor_cell=data.get('anchor_cell', 'R2'),
            width_px=data.get('width_px', 768),
            mode=data.get('mode', 'overwrite')
        )

        # 处理普通表格
        result = await service.process_table_pipeline(
            image_path=image_path,
            out_file=final_output_path,
            sheet_name=sheet_name,
            bank_name=bank_name,
            file_name=file_name,
            excel_config=excel_config
        )

        # ⭐⭐⭐ 关键修复：确保获取 DataFrame 数据但不直接返回 ⭐⭐⭐
        df_data = None
        if hasattr(result, 'df') and result.df is not None:
            df_data = result.df
            print(f"✅ 普通表格获取到 DataFrame 数据，形状: {df_data.shape}")
        else:
            print("⚠️ 普通表格没有获取到 DataFrame 数据，尝试从Excel文件读取")
            # 如果result中没有df，尝试从生成的Excel文件读取
            try:
                if Path(final_output_path).exists():
                    import pandas as pd
                    df_data = pd.read_excel(final_output_path, sheet_name=sheet_name)
                    print(f"✅ 从Excel文件读取到 DataFrame 数据，形状: {df_data.shape}")
            except Exception as e:
                print(f"❌ 从Excel文件读取数据失败: {e}")

        # ⭐⭐⭐ 统一返回格式 - 改为批量处理模式 ⭐⭐⭐
        if result.status == "success" and Path(final_output_path).exists():
            # 构建前端可访问的URL
            excel_url = convert_to_excel_url(final_output_path)

            # ⭐⭐⭐ 关键修复：不直接返回 DataFrame 对象，而是保存到文件 ⭐⭐⭐
            # 如果有 DataFrame 数据，先保存到 Excel
            if df_data is not None and not df_data.empty:
                try:
                    # 使用 excel_storage_service 保存数据
                    from backend.service.excel_storage_service import ExcelStorageService
                    excel_service = ExcelStorageService()
                    excel_service.save_dataframe(
                        df=df_data,
                        excel_path=final_output_path,
                        sheet_name=sheet_name,
                        map_name=result.table_name,
                        image_data=image_path if image_path and os.path.exists(image_path) else None
                    )
                    print(f"✅ 已保存 DataFrame 数据到 Excel 文件")
                except Exception as save_error:
                    print(f"❌ 保存 DataFrame 数据失败: {save_error}")

            # ⭐⭐⭐ 关键修复：构建完整的返回结果，包含 recognizedData ⭐⭐⭐
            return _build_processing_response(
                image_path=image_path,
                output_path=final_output_path,
                sheet_name=sheet_name,
                excel_url=excel_url,
                result=result,
                df_data=df_data,
                table_type="non_financial",
                folder_name=folder_name,
                from_cache=False
            )
        else:
            return {
                "success": False,
                "error": result.error_message or "表格处理失败",
                "errorCode": "PROCESS_FAILED"
            }

    except Exception as e:
        return {
            "success": False,
            "error": f"处理失败: {str(e)}",
            "errorCode": "SYSTEM_ERROR"
        }


def _build_cache_response(image_path, output_path, sheet_name, excel_url, df, table_type="financial"):
    """构建缓存响应"""
    # 转换DataFrame为前端需要的格式
    headers = list(df.columns) if not df.empty else []
    data_list = []

    if not df.empty:
        for _, row in df.iterrows():
            row_data = {}
            for header in headers:
                cell_value = row[header]
                row_data[header] = str(cell_value) if pd.notna(cell_value) else ""
            data_list.append(row_data)

    is_non_financial = table_type == "non_financial"

    # 返回现有数据 - 包含完整的 recognizedData
    return {
        "success": True,
        "data": {
            "total": 1,
            "success": 1,
            "failed": 0,
            "non_financial": 1 if is_non_financial else 0,
            "results": [
                {
                    "image_path": image_path,
                    "status": "success",
                    "status_text": "从缓存加载",
                    "complexity": "unknown",
                    "table_name": sheet_name,
                    "sheet_name": sheet_name,
                    "error_message": "",
                    "assessment_reason": "从已存在的Excel文件加载",
                    "success": True,
                    "is_non_financial": is_non_financial,
                    "output_file": output_path,
                    "excel_url": excel_url
                }
            ],
            "output_file": output_path,
            "excel_url": excel_url,
            "table_type": table_type,
            "processing_completed": True,
            "message": "从已存在的Excel文件加载数据",
            "excel_saved": True,
            "from_cache": True
        },
        # ⭐⭐⭐ 关键修复：添加 recognizedData 字段 ⭐⭐⭐
        "recognizedData": {
            "headers": headers,
            "data": data_list,
            "tableName": sheet_name,
            "excelPath": output_path,
            "fromCache": True
        },
        "task_id": f"cached_{table_type}_{int(time.time())}",
        "processing_completed": True,
        "from_cache": True
    }


def _build_processing_response(image_path, output_path, sheet_name, excel_url, result, df_data,
                               table_type="financial", folder_name="", from_cache=False):
    """构建处理响应"""
    is_non_financial = table_type == "non_financial"

    # 构建 recognizedData
    headers = []
    data_list = []
    if df_data is not None and not df_data.empty:
        headers = list(df_data.columns)
        for _, row in df_data.iterrows():
            row_data = {}
            for header in headers:
                cell_value = row[header]
                row_data[header] = str(cell_value) if pd.notna(cell_value) else ""
            data_list.append(row_data)

    # 构建响应
    response = {
        "success": True,
        "data": {
            "total": 1,
            "success": 1,
            "failed": 0,
            "non_financial": 1 if is_non_financial else 0,
            "results": [
                {
                    "image_path": image_path,
                    "status": result.status,
                    "status_text": "成功" if result.status == "success" else "非金融表格",
                    "complexity": getattr(result, 'complexity', 'unknown'),
                    "table_name": getattr(result, 'table_name', sheet_name),
                    "sheet_name": sheet_name,
                    "error_message": getattr(result, 'error_message', ''),
                    "assessment_reason": getattr(result, 'assessment_reason', ''),
                    "success": True,
                    "is_non_financial": is_non_financial,
                    "output_file": output_path,
                    "excel_url": excel_url
                }
            ],
            "output_file": output_path,
            "excel_url": excel_url,
            "table_type": table_type,
            "processing_completed": True,
            "message": "普通表格处理完成" if is_non_financial else "表格处理完成",
            "excel_saved": True,
            "from_cache": from_cache
        },
        # ⭐⭐⭐ 关键修复：添加 recognizedData 字段 ⭐⭐⭐
        "recognizedData": {
            "headers": headers,
            "data": data_list,
            "tableName": getattr(result, 'table_name', sheet_name),
            "excelPath": output_path,
            "fromCache": from_cache
        },
        "task_id": f"{table_type}_{int(time.time())}",
        "processing_completed": True,
        "from_cache": from_cache
    }

    # 保持兼容性
    response["excel_url"] = excel_url
    response["from_cache"] = from_cache
    if folder_name:
        response["tableName"] = f"{getattr(result, 'table_name', sheet_name)} - {folder_name}"

    return response

