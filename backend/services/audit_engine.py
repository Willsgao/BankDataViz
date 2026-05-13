# -*- coding:utf-8 -*-
"""
会计勾稽规则引擎

从 Excel 文件读取数据，按规则执行勾稽校验，返回结构化结果。
不依赖数据库，直接读 Excel 文件。

支持两种数据源模式：
1. 传统模式：传入 excel_path，直接读取文件
2. DAL模式：传入 data_source 和 file_id，使用数据访问层
"""
import json
import os
import re
import openpyxl
from datetime import datetime
from typing import Any, Optional, TYPE_CHECKING

if TYPE_CHECKING:
    from backend.services.dal import ExcelDataSource


# =============================================================================
# 工具函数
# =============================================================================

def parse_number(val: Any) -> Optional[float]:
    """将字符串/数字转为 float，处理千分位逗号和括号负数"""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s or s == '-' or s == '—':
        return None
    if s.startswith('(') and s.endswith(')'):
        s = '-' + s[1:-1]
    s = s.replace(',', '')
    s = s.replace('%', '')
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def fuzzy_match(text: str, keywords: list[str]) -> bool:
    """判断文本是否包含任意一个关键词（模糊匹配）"""
    if not text or not keywords:
        return False
    text_lower = text.lower()
    for kw in keywords:
        if kw.lower() in text_lower:
            return True
    return False


def _clean_header(val: str) -> str:
    """清理表头：移除 a>>/b>> 等前缀和多余空白"""
    if not val:
        return ''
    # 移除前缀 a>> b>> c>> d>>
    s = re.sub(r'^[a-d]>>', '', str(val), flags=re.IGNORECASE)
    # 移除多余空白
    s = re.sub(r'\s+', '', s)
    return s


def _match_hint(cell_val: Any, hint: str) -> bool:
    """判断单元格值是否匹配 hint"""
    if not cell_val or not hint:
        return False
    cell_str = str(cell_val)
    hint_lower = hint.lower().replace(' ', '')
    cell_clean = cell_str.lower().replace(' ', '').replace('\n', '').replace('\r', '')
    # 移除 a>> b>> 等前缀后比对
    cell_clean = _clean_header(cell_clean)
    hint_clean = _clean_header(hint_lower)
    return hint_clean in cell_clean or cell_clean in hint_clean


def find_sheet_by_hint(wb, sheet_hint: str):
    """根据 sheet_hint 找到匹配的 sheet"""
    for ws in wb.worksheets:
        if sheet_hint.lower() in ws.title.lower():
            return ws
    return None


def read_cell(ws, row_idx: int, col_idx: int):
    """安全读取单元格"""
    try:
        return ws.cell(row=row_idx, column=col_idx).value
    except Exception:
        return None


def _normalize_date_str(s: str) -> str:
    """标准化日期字符串以便比对（提取纯日期部分）"""
    if not s:
        return ''
    s = str(s)
    # 移除所有 a>>/b>>/c>>/d>> 前缀和中间的分隔符
    s = re.sub(r'[a-z]>>', '', s, flags=re.IGNORECASE)
    # 移除多余空白
    s = re.sub(r'\s+', '', s)
    # 标准化格式：年.月.日 → 年月日
    s = s.replace('年', '').replace('月', '').replace('日', '')
    s = s.replace('.', '').replace('/', '')
    return s


def _is_date_hint(hint: str) -> bool:
    """判断 hint 是否是日期类"""
    return bool(re.search(r'202\d|20\d{2}|年月|季度|期间', hint))


def _find_period_date_col(ws, date_hint: str) -> Optional[int]:
    """
    智能查找含日期的列，处理多种 Excel 表头结构。

    关键判断逻辑（对每个候选列独立判断）：
    - Row3 无数据（None）→ 需要偏移 +1（如 P005）
    - Row3 有"年"但无"日"（如"2024年第四季度"）→ 需要偏移 +1（如 P036_2）
    - Row3 有完整日期（含"日"，如"2024年12月31日"）→ 不偏移，用 Row3 日期匹配

    策略：
    1. 在 Row1 找所有含 date_hint 的列 → candidates
    2. 对每个候选列独立检查 Row3 是否为完整日期
    3. Row3 有完整日期 → 用 Row3 日期匹配该候选列
    4. Row3 无完整日期 → 该候选列需要偏移 +1
    """
    def _has_full_date(val):
        """判断 Row3 值是否是完整的年月日日期"""
        if val is None:
            return False
        s = str(val).replace(' ', '').replace('\n', '')
        return '年' in s and '日' in s

    def _is_period_label(val):
        """判断 Row3 值是否是期间标签（只有"年"，没有"日"）"""
        if val is None:
            return False
        s = str(val).replace(' ', '').replace('\n', '')
        return '年' in s and '日' not in s

    # 1. 在 Row1 找所有匹配列
    candidates = []
    for col_idx in range(1, min(ws.max_column + 1, 20)):
        cell_val = ws.cell(1, col_idx).value
        if cell_val and _match_hint(cell_val, date_hint):
            candidates.append(col_idx)

    if not candidates:
        return None

    # 2. 对每个候选列，检查其 Row3 是否为完整日期
    for col_idx in candidates:
        row3_val = ws.cell(3, col_idx).value
        if row3_val and _has_full_date(row3_val):
            # Row3 有完整日期 → 检查是否与 hint 匹配
            if _match_hint(row3_val, date_hint):
                return col_idx  # Row3 日期精确匹配该列 → 无偏移

    # 3. 所有候选列 Row3 无完整日期 → 对每个候选列判断偏移
    for col_idx in candidates:
        row3_val = ws.cell(3, col_idx).value
        if row3_val and _is_period_label(row3_val):
            # Row3 有期间标签（有"年"无"日"）→ 偏移 +1
            return col_idx + 1
        elif row3_val is None:
            # Row3 无数据 → 偏移 +1
            return col_idx + 1

    # 4. 兜底：返回最左候选 +1
    return min(candidates) + 1








def find_col_by_hint(ws, hint: str, max_rows: int = 3) -> Optional[int]:
    """
    根据列 hint 在前几行找到匹配的列索引。
    智能处理：
    - 日期类 hint：使用 _find_period_date_col 处理列偏移
    - 标签类 hint：在前 max_rows 行模糊搜索
    """
    # 日期类 hint 使用专用方法
    if _is_date_hint(hint):
        return _find_period_date_col(ws, hint)

    # 非日期 hint：前几行模糊搜索
    for row_idx in range(1, max_rows + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val and _match_hint(cell_val, hint):
                return col_idx
    return None


def find_row_by_keywords(ws, keywords: list[str], start_row: int = 1) -> Optional[int]:
    """
    根据行关键词查找数据行。

    优先级策略：
    1. 精确匹配：>>段内关键词完全相等 → 最高优先级
    2. 包含匹配：关键词是单元格值的一部分 → 次优先级
    3. 数值验证：匹配行必须有数值型数据，否则跳过（排除标题行）

    从第2/3列开始找，避免误匹配期号列。
    """
    best_row = None
    best_score = 0  # 精确匹配=100, 包含=10

    for row_idx in range(start_row, ws.max_row + 1):
        has_numeric = False
        row_score = 0

        for col_idx in range(2, min(ws.max_column + 1, 10)):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val:
                cell_str = str(cell_val)
                cell_clean = _clean_header(cell_str)
                cell_lower = cell_clean.lower()

                for kw in keywords:
                    kw_clean = kw.lower().strip()

                    # 精确匹配：kw 完全等于某个 >> 段
                    cell_parts = [p.strip() for p in cell_clean.split('>>') if p.strip()]
                    for part in cell_parts:
                        if part.lower() == kw_clean:
                            row_score = max(row_score, 100)  # 精确匹配得分 100
                            break

                    # 包含匹配：kw 是 cell_clean 的一部分
                    if kw_clean in cell_lower:
                        row_score = max(row_score, 10)  # 包含匹配得分 10

                # 检查是否有数值数据（排除标题行）
                if not isinstance(cell_val, str):
                    has_numeric = True
                else:
                    stripped = cell_val.strip().replace(',', '').replace('(', '-').replace(')', '').replace(' ', '')
                    try:
                        float(stripped)
                        has_numeric = True
                    except (ValueError, TypeError):
                        pass

        if row_score > 0:
            # 精确匹配优先于包含匹配
            if row_score > best_score:
                best_score = row_score
                best_row = row_idx

    return best_row


def get_indicator_value(ws, field_keywords: list[str], col_hint: str) -> Optional[float]:
    """
    从 sheet 中提取某指标在指定列的值。

    策略：先找行，再找列。
    - 目标行该列：优先尝试 parse_number（数字和含逗号字符串都能解析）
    - 如果解析失败，向右偏移找实际数据
    - 如果向右都找不到，尝试下一行（同列向下回退）
    """
    row_idx = find_row_by_keywords(ws, field_keywords)
    if row_idx is None:
        return None

    col_idx = find_col_by_hint(ws, col_hint)
    if col_idx is None:
        return None

    def _try_parse(ws, row, col):
        """尝试解析指定单元格数值，排除行标记列的数字0"""
        v = read_cell(ws, row, col)
        if v is None:
            return None
        p = parse_number(v)
        # 排除纯数字 0（通常是列末的行标记）
        if p is not None and p == 0 and isinstance(v, (int, float)):
            return None
        return p

    # 1. 目标行该列
    val = _try_parse(ws, row_idx, col_idx)
    if val is not None:
        return val

    # 2. 向下回退（标题行偏移，如 Row4 无数据但 Row5 有）
    for row_offset in range(1, 5):
        adj_row = row_idx + row_offset
        if adj_row <= ws.max_row:
            val = _try_parse(ws, adj_row, col_idx)
            if val is not None:
                return val

    # 3. 向右偏移（处理表头偏移情况）
    for offset in range(1, 5):
        adj = col_idx + offset
        if adj <= ws.max_column:
            val = _try_parse(ws, row_idx, adj)
            if val is not None:
                return val

    return None


# =============================================================================
# 规则执行器
# =============================================================================

def run_formula_rule(rule: dict, ws) -> dict:
    """执行 formula 类型的规则"""
    formula = rule.get('formula', {})
    num_field = formula.get('numerator', {}).get('field', '')
    num_hint = formula.get('numerator', {}).get('col_hint', '')
    den_field = formula.get('denominator', {}).get('field', '')
    den_hint = formula.get('denominator', {}).get('col_hint', '')
    result_field = formula.get('result_field', '')
    result_hint = formula.get('result_col_hint', '')
    multiplier = formula.get('multiplier', 1)

    num_val = get_indicator_value(ws, [num_field], num_hint)
    den_val = get_indicator_value(ws, [den_field], den_hint)

    if num_val is None or den_val is None or den_val == 0:
        return {
            'status': 'fail',
            'actual_value': None,
            'expected_value': None,
            'diff': None,
            'detail': f'无法提取数据（{num_field}={num_val}, {den_field}={den_val}）'
        }

    calculated = (num_val / den_val) * multiplier

    # 读报告中的实际值
    reported = None
    if result_field:
        reported = get_indicator_value(ws, [result_field], result_hint)

    tolerance = rule.get('tolerance', 0.01)
    unit = rule.get('unit', '%')
    tol_type = rule.get('tolerance_type', 'percent')

    if reported is None:
        return {
            'status': 'warn',
            'actual_value': f'{calculated:.4f}{unit}',
            'expected_value': '报告值未找到',
            'diff': None,
            'detail': f'计算值={calculated:.4f}{unit}，但报告中的{result_field}未找到'
        }

    diff_abs = abs(calculated - reported)

    if tol_type == 'percent':
        diff_pct = (diff_abs / abs(reported) * 100) if reported != 0 else 0
        is_pass = diff_pct <= tolerance
        detail = f'计算={calculated:.4f}{unit}，报告={reported:.4f}{unit}，误差={diff_pct:.4f}%'
        diff_out = diff_pct
    else:
        diff_pct = None
        is_pass = diff_abs <= tolerance
        detail = f'计算={calculated:.4f}{unit}，报告={reported:.4f}{unit}，差值={diff_abs:.4f}'
        diff_out = diff_abs

    status = 'pass' if is_pass else 'fail'
    return {
        'status': status,
        'actual_value': f'{reported:.4f}{unit}',
        'expected_value': f'{calculated:.4f}{unit}',
        'diff': diff_out,
        'diff_percent': diff_pct,
        'detail': detail
    }


def run_sum_check_rule(rule: dict, ws) -> dict:
    """执行 sum_check 类型的规则"""
    sum_check = rule.get('sum_check', {})
    target_field = sum_check.get('target_field', '')
    target_hint = sum_check.get('target_col_hint', '')
    components = sum_check.get('components', [])
    # 支持多行求和: sum_col_hint + start_row + end_row
    sum_col_hint = sum_check.get('sum_col_hint', '')
    sum_start_row = sum_check.get('sum_start_row')
    sum_end_row = sum_check.get('sum_end_row')

    target_val = get_indicator_value(ws, [target_field], target_hint)

    # 多行求和模式（用于 LCR 等：多个子项在同列不同行）
    unit = rule.get('unit', '')
    if sum_col_hint and sum_start_row and sum_end_row:
        col_idx = find_col_by_hint(ws, sum_col_hint)
        col_sum = 0.0
        cnt = 0
        for r in range(sum_start_row, sum_end_row + 1):
            v = parse_number(ws.cell(r, col_idx).value)
            if v is not None:
                col_sum += v
                cnt += 1
        if target_val is not None:
            diff_abs = abs(target_val - col_sum)
            tolerance = rule.get('tolerance', 1000)
            is_pass = diff_abs <= tolerance
            status = 'pass' if is_pass else 'fail'
            return {
                'status': status,
                'actual_value': f'{target_val:.0f}{unit}',
                'expected_value': f'{col_sum:.0f}{unit}',
                'diff': diff_abs,
                'detail': f'目标={target_val:.0f}，多行求和({cnt}行)={col_sum:.0f}，差值={diff_abs:.0f}'
            }
        else:
            return {
                'status': 'warn',
                'actual_value': '未找到目标',
                'expected_value': f'求和={col_sum:.0f}',
                'diff': None,
                'detail': f'求和={col_sum:.0f}（{cnt}行），但目标字段 [{target_field}] 未找到'
            }

    # 单行分量模式
    component_vals = []
    component_names = []
    for comp in components:
        comp_field = comp.get('field', '')
        comp_hint = comp.get('col_hint', '')
        sign = comp.get('sign', 1)
        val = get_indicator_value(ws, [comp_field], comp_hint)
        if val is not None:
            component_vals.append(val * sign)
            component_names.append(comp_field)

    tolerance = rule.get('tolerance', 1000)
    unit = rule.get('unit', '')

    if target_val is not None and component_vals:
        calculated_sum = sum(component_vals)
        diff_abs = abs(target_val - calculated_sum)
        is_pass = diff_abs <= tolerance
        status = 'pass' if is_pass else 'fail'
        parts_str = ' + '.join([f'{v:.0f}' if v >= 0 else f'({abs(v):.0f})' for v in component_vals])
        detail = f'目标={target_val:.0f}，分项=({parts_str})={calculated_sum:.0f}，差值={diff_abs:.0f}'
        return {
            'status': status,
            'actual_value': f'{target_val:.0f}{unit}',
            'expected_value': f'{calculated_sum:.0f}{unit}',
            'diff': diff_abs,
            'detail': detail
        }
    elif target_val is None and component_vals:
        calculated_sum = sum(component_vals)
        parts_str = ' + '.join([f'{v:.0f}' if v >= 0 else f'({abs(v):.0f})' for v in component_vals])
        return {
            'status': 'warn',
            'actual_value': '未找到合计值',
            'expected_value': f'分项计算={calculated_sum:.0f}{unit}',
            'diff': None,
            'detail': f'分项之和=({parts_str})={calculated_sum:.0f}，目标行 [{target_field}] 未找到'
        }
    elif target_val is not None and not component_vals:
        return {
            'status': 'warn',
            'actual_value': f'{target_val:.0f}{unit}',
            'expected_value': '未找到任何分项',
            'diff': None,
            'detail': f'目标={target_val:.0f}，但未找到任何分项数据'
        }
    else:
        return {
            'status': 'fail',
            'actual_value': None,
            'expected_value': None,
            'diff': None,
            'detail': f'无法提取目标字段 [{target_field}] 和任何分项数据'
        }


def run_periodicity_rule(rule: dict, ws) -> dict:
    """执行 periodicity 类型的跨期一致性规则"""
    periodicity = rule.get('periodicity', {})
    field_keywords = [periodicity.get('field', '')]
    periods = periodicity.get('periods', [])
    tolerance = rule.get('tolerance', 0.5)
    unit = rule.get('unit', '%')

    period_vals = []
    period_labels = []
    for p in periods:
        hint = p.get('col_hint', '')
        label = p.get('label', hint)
        val = get_indicator_value(ws, field_keywords, hint)
        period_vals.append(val)
        period_labels.append(label)

    valid_vals = [(l, v) for l, v in zip(period_labels, period_vals) if v is not None]

    if len(valid_vals) < 2:
        return {
            'status': 'warn',
            'actual_value': ', '.join([f'{v}' if v else 'N/A' for v in period_vals]),
            'expected_value': '需要至少2期数据',
            'diff': None,
            'detail': f'仅找到 {len(valid_vals)} 期数据，无法进行跨期比对'
        }

    issues = []
    for i in range(1, len(valid_vals)):
        prev_label, prev_val = valid_vals[i - 1]
        curr_label, curr_val = valid_vals[i]
        diff = abs(curr_val - prev_val)
        if diff > tolerance:
            issues.append(f'{prev_label}→{curr_label}: {prev_val:.4f}→{curr_val:.4f} (变化{diff:.4f}{unit})')

    if issues:
        status = 'fail'
        detail = '跨期波动超过容差: ' + '; '.join(issues)
    else:
        status = 'pass'
        detail = '各期数据连续性正常: ' + ', '.join([f'{l}={v:.4f}' for l, v in valid_vals])

    return {
        'status': status,
        'actual_value': ', '.join([f'{v:.4f}' if v else 'N/A' for _, v in valid_vals]),
        'expected_value': '跨期一致性',
        'diff': None,
        'detail': detail
    }


def run_rule(rule: dict, excel_path: str = None, sheet_mapping: dict = None,
              file_id: str = None, data_source = None) -> dict:
    """
    对一条规则执行校验。
    
    支持两种模式：
    1. 传统模式（excel_path）：直接传入Excel文件路径
    2. DAL模式（data_source + file_id）：使用数据访问层
    
    Args:
        rule: 规则配置 dict
        excel_path: Excel 文件路径（传统模式）
        sheet_mapping: 可选，{rule_id: sheet_name} 映射。
                      如果传入，则直接使用指定 sheet，不做自动匹配。
        file_id: 档案ID（DAL模式）
        data_source: 数据源（DAL模式）
    """
    rule_id = rule.get('id', '')
    rule_type = rule.get('rule_type', '')
    rule_name = rule.get('name', '')

    # 确定使用哪种模式
    use_dal = data_source is not None and file_id is not None
    wb = None
    
    try:
        # DAL模式：使用数据访问层
        if use_dal:
            # 获取所有Sheet数据用于规则匹配
            all_sheets = data_source.get_all_sheets_data(file_id)
            if not all_sheets:
                return {
                    'rule_id': rule_id, 'rule_name': rule_name,
                    'sheet_name': None, 'period': None,
                    'status': 'fail', 'actual_value': None,
                    'expected_value': None, 'diff': None,
                    'diff_percent': None,
                    'detail': f'未找到任何 Sheet（档案ID: {file_id}）'
                }
            
            # 确定使用哪个 sheet
            target_sheet_name = None
            if sheet_mapping and rule_id in sheet_mapping:
                # 人工指定的 sheet
                target_sheet_name = sheet_mapping[rule_id]
            else:
                # 自动匹配：用内容关键词找 sheet
                keywords = _extract_rule_keywords(rule)
                target_sheet_name, match_score = _find_sheet_by_name_keywords_dal(
                    all_sheets, keywords
                )
                if target_sheet_name is None:
                    return {
                        'rule_id': rule_id, 'rule_name': rule_name,
                        'sheet_name': None, 'period': None,
                        'status': 'fail', 'actual_value': None,
                        'expected_value': None, 'diff': None,
                        'diff_percent': None,
                        'detail': f'未找到匹配的 Sheet（关键词={keywords}）'
                    }
            
            # 获取指定Sheet的完整数据
            sheet_data = None
            for sd in all_sheets:
                if sd.name == target_sheet_name:
                    sheet_data = sd
                    break
            
            if sheet_data is None:
                return {
                    'rule_id': rule_id, 'rule_name': rule_name,
                    'sheet_name': target_sheet_name, 'period': None,
                    'status': 'fail', 'actual_value': None,
                    'expected_value': None, 'diff': None,
                    'diff_percent': None,
                    'detail': f'指定的 Sheet 不存在: {target_sheet_name}'
                }
            
            # 提取报告期
            period = _extract_period_from_headers_dal(sheet_data.headers)
            
            # 使用DAL模式执行规则
            if rule_type == 'formula':
                result = run_formula_rule_dal(rule, sheet_data)
            elif rule_type == 'sum_check':
                result = run_sum_check_rule_dal(rule, sheet_data)
            elif rule_type == 'periodicity':
                result = run_periodicity_rule_dal(rule, sheet_data)
            else:
                result = {'status': 'fail', 'detail': f'未知规则类型: {rule_type}'}

            return {
                'rule_id': rule_id,
                'rule_name': rule_name,
                'sheet_name': sheet_data.name,
                'period': period or rule.get('sheet_hint', ''),
                **result
            }
        
        # 传统模式：直接读取Excel文件
        elif excel_path:
            wb = openpyxl.load_workbook(excel_path, data_only=True)

            # 确定使用哪个 sheet
            ws = None
            if sheet_mapping and rule_id in sheet_mapping:
                # 人工指定的 sheet，直接按名称查找
                target_sheet = sheet_mapping[rule_id]
                for ws_candidate in wb.worksheets:
                    if ws_candidate.title == target_sheet:
                        ws = ws_candidate
                        break
                if ws is None:
                    return {
                        'rule_id': rule_id, 'rule_name': rule_name,
                        'sheet_name': target_sheet, 'period': None,
                        'status': 'fail', 'actual_value': None,
                        'expected_value': None, 'diff': None,
                        'diff_percent': None,
                        'detail': f'指定的 Sheet 不存在: {target_sheet}'
                    }
            else:
                # 自动匹配：用内容关键词找 sheet
                keywords = _extract_rule_keywords(rule)
                ws, match_score = _find_sheet_by_keywords(wb, keywords)
                if ws is None:
                    return {
                        'rule_id': rule_id, 'rule_name': rule_name,
                        'sheet_name': None, 'period': None,
                        'status': 'fail', 'actual_value': None,
                        'expected_value': None, 'diff': None,
                        'diff_percent': None,
                        'detail': f'未找到匹配的 Sheet（关键词={keywords}）'
                    }

            # 提取报告期
            period = None
            for row_idx in range(1, 4):
                for col_idx in range(1, min(ws.max_column + 1, 12)):
                    cell_val = ws.cell(row_idx, col_idx).value
                    if cell_val and re.search(r'202[0-9]年\d{1,2}月\d{1,2}日', str(cell_val)):
                        period = str(cell_val)
                        break
                if period:
                    break

            if rule_type == 'formula':
                result = run_formula_rule(rule, ws)
            elif rule_type == 'sum_check':
                result = run_sum_check_rule(rule, ws)
            elif rule_type == 'periodicity':
                result = run_periodicity_rule(rule, ws)
            else:
                result = {'status': 'fail', 'detail': f'未知规则类型: {rule_type}'}

            return {
                'rule_id': rule_id,
                'rule_name': rule_name,
                'sheet_name': ws.title,
                'period': period or rule.get('sheet_hint', ''),
                **result
            }
        else:
            return {
                'rule_id': rule_id, 'rule_name': rule_name,
                'sheet_name': None, 'period': None,
                'status': 'fail', 'actual_value': None,
                'expected_value': None, 'diff': None,
                'diff_percent': None,
                'detail': '必须提供 excel_path 或 (data_source + file_id)'
            }

    except FileNotFoundError:
        return {
            'rule_id': rule_id, 'rule_name': rule_name,
            'sheet_name': None, 'period': None,
            'status': 'fail', 'actual_value': None,
            'expected_value': None, 'diff': None,
            'diff_percent': None,
            'detail': f'Excel 文件未找到: {excel_path}'
        }
    except Exception as e:
        import traceback
        return {
            'rule_id': rule_id, 'rule_name': rule_name,
            'sheet_name': None, 'period': None,
            'status': 'fail', 'actual_value': None,
            'expected_value': None, 'diff': None,
            'diff_percent': None,
            'detail': f'规则执行异常: {str(e)}\n{traceback.format_exc()}'
        }
    finally:
        if wb:
            wb.close()


def _extract_rule_keywords(rule: dict) -> list[str]:
    """从规则配置中提取关键词，用于自动匹配 Sheet"""
    keywords = []
    name = rule.get('name', '')
    # 从名称提取，去除勾稽/校验等后缀
    for kw in name.replace('勾稽', '').replace('校验', '').split('、'):
        kw = kw.strip()
        if len(kw) >= 2 and isinstance(kw, str):
            keywords.append(kw)
    # 从 description 提取中文关键词
    import re
    desc = rule.get('description', '')
    for m in re.findall(r'[\u4e00-\u9fa5]{2,10}', desc):
        if m not in ['分项', '目标', '计算', '报告', '数值', '目标行']:
            keywords.append(m)
    # 从 formula/sum_check 的 field 提取
    if 'formula' in rule:
        for part in ['numerator', 'denominator', 'result_field']:
            v = rule['formula'].get(part, '')
            # 处理嵌套字典情况 (如 numerator: {field: ..., col_hint: ...})
            if isinstance(v, dict):
                v = v.get('field', '')
            if v and isinstance(v, str):
                keywords.append(v)
    if 'sum_check' in rule:
        target = rule['sum_check'].get('target_field', '')
        if target and isinstance(target, str):
            keywords.append(target)
        for comp in rule['sum_check'].get('components', []):
            field = comp.get('field', '')
            if field and isinstance(field, str):
                keywords.append(field)
    # 确保所有元素都是字符串
    keywords = [k for k in keywords if isinstance(k, str)]
    return list(set(keywords))


def _find_sheet_by_keywords(wb, keywords: list[str]):
    """
    根据关键词列表，在所有 sheet 中找最佳匹配。
    返回 (worksheet, score)。
    """
    best_ws = None
    best_score = 0

    for ws in wb.worksheets:
        score = 0
        sheet_name = ws.title.lower()
        for kw in keywords:
            kw_lower = kw.lower()
            # Sheet 名匹配权重最高
            if kw_lower in sheet_name:
                score += 20
            # 检查前几行表头
            for row_idx in range(1, 4):
                for col_idx in range(1, min(ws.max_column + 1, 10)):
                    cell_val = ws.cell(row_idx, col_idx).value
                    if cell_val:
                        cell_str = str(cell_val).lower().replace('\n', ' ')
                        if kw_lower in cell_str:
                            score += 3
            # 检查前几列表头
            for col_idx in range(1, 4):
                for row_idx in range(1, min(ws.max_row + 1, 10)):
                    cell_val = ws.cell(row_idx, col_idx).value
                    if cell_val:
                        cell_str = str(cell_val).lower().replace('\n', ' ')
                        if kw_lower in cell_str:
                            score += 2

        if score > best_score:
            best_score = score
            best_ws = ws

    return best_ws, best_score


def load_rules(config_path: str = None) -> list[dict]:
    """加载勾稽规则配置"""
    if config_path is None:
        # __file__ = backend/services/audit_engine.py
        # backend_dir = backend/  → DocuVista 项目根目录是 parent of backend
        backend_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        # 实际路径: {project_root}/data/backend/config/audit_rules.json
        # 而 backend_dir = backend/ (即 DocuVista/backend/)
        # 修正：project_root = backend_dir.parent
        project_root = os.path.dirname(backend_dir)
        config_path = os.path.join(project_root, 'data', 'backend', 'config', 'audit_rules.json')

    if not os.path.exists(config_path):
        return []

    with open(config_path, 'r', encoding='utf-8') as f:
        rules = json.load(f)

    return [r for r in rules if r.get('enabled', True)]


# =============================================================================
# DAL 模式辅助函数
# =============================================================================

def _find_sheet_by_name_keywords_dal(all_sheets, keywords: list[str]):
    """
    DAL模式：根据关键词列表，在所有 sheet 中找最佳匹配。
    返回 (sheet_name, score)。
    """
    from backend.services.data_access_layer import SheetData
    
    best_sheet_name = None
    best_score = 0

    for sd in all_sheets:
        if not isinstance(sd, SheetData):
            continue
            
        score = 0
        sheet_name = sd.name.lower()
        
        for kw in keywords:
            kw_lower = kw.lower()
            # Sheet 名匹配权重最高
            if kw_lower in sheet_name:
                score += 20
            
            # 检查表头内容
            for header_row in sd.headers:
                for cell_val in header_row:
                    if cell_val:
                        cell_str = str(cell_val).lower().replace('\n', ' ')
                        if kw_lower in cell_str:
                            score += 3
                            break

        if score > best_score:
            best_score = score
            best_sheet_name = sd.name

    return best_sheet_name, best_score


def _extract_period_from_headers_dal(headers: list) -> str:
    """从表头中提取报告期"""
    for header_row in headers:
        for cell_val in header_row:
            if cell_val and re.search(r'202[0-9]年\d{1,2}月\d{1,2}日', str(cell_val)):
                return str(cell_val)
    return None


def _find_row_by_keywords_dal(sheet_data, keywords: list[str]) -> Optional[int]:
    """
    DAL模式：根据行关键词查找数据行。
    返回行索引（从1开始）。
    
    查找策略：
    1. 优先检查前5列（纵向表头通常在前面几列）
    2. 精确匹配优先于包含匹配
    3. 支持部分匹配（如"风险加权资产"可以匹配"风险加权资产合计"）
    """
    best_row_idx = None
    best_score = 0
    best_match_info = ""

    # 检查前5列作为纵向表头列
    header_col_keys = [f"col_{i}" for i in range(1, 6)]

    for row_idx, row_data in enumerate(sheet_data.rows, start=len(sheet_data.headers) + 1):
        row_score = 0
        match_info = ""

        for col_key in header_col_keys:
            cell_val = row_data.get(col_key)
            if cell_val is None:
                continue
            
            cell_str = str(cell_val)
            cell_clean = _clean_header(cell_str)
            cell_lower = cell_clean.lower()

            for kw in keywords:
                kw_clean = kw.lower().strip()

                # 精确匹配：kw 完全等于某个 >> 段
                cell_parts = [p.strip() for p in cell_clean.split('>>') if p.strip()]
                for part in cell_parts:
                    if part.lower() == kw_clean:
                        row_score = max(row_score, 100)
                        match_info = f"精确匹配: '{kw}' == '{part}'"
                        break

                # 包含匹配：cell 包含 kw
                if kw_clean in cell_lower:
                    row_score = max(row_score, 20)
                    if not match_info:
                        match_info = f"包含匹配: '{kw}' in '{cell_clean[:30]}'"
                
                # 反向包含：kw 包含 cell（部分匹配）
                elif cell_lower and cell_lower in kw_clean and len(cell_lower) >= 4:
                    row_score = max(row_score, 15)
                    if not match_info:
                        match_info = f"部分匹配: '{cell_clean[:30]}' in '{kw}'"

        if row_score > 0 and row_score > best_score:
            best_score = row_score
            best_row_idx = row_idx
            best_match_info = match_info

    # 调试日志
    if best_row_idx:
        print(f"[DAL查找] 关键词 {keywords} -> 行{best_row_idx}, 得分{best_score}, {best_match_info}")
    else:
        print(f"[DAL查找] 关键词 {keywords} -> 未找到匹配")

    return best_row_idx


def _find_col_by_hint_dal(sheet_data, hint: str, row_idx: int = None) -> Optional[int]:
    """DAL模式：根据列 hint 查找列索引"""
    if row_idx is None:
        # 在表头中查找
        for col_idx in range(1, 100):
            for header_row in sheet_data.headers:
                if col_idx <= len(header_row):
                    cell_val = header_row[col_idx - 1]
                    if cell_val and _match_hint(cell_val, hint):
                        return col_idx
        return None
    else:
        # 在指定行查找
        row_data = sheet_data.rows[row_idx - len(sheet_data.headers) - 1]
        if row_data:
            for col_idx in range(1, 100):
                col_key = f"col_{col_idx}"
                if col_key in row_data:
                    cell_val = row_data[col_key]
                    if cell_val and _match_hint(cell_val, hint):
                        return col_idx
        return None


def _get_cell_value_dal(sheet_data, row_idx: int, col_idx: int):
    """DAL模式：获取单元格值"""
    if row_idx <= len(sheet_data.headers):
        # 表头行
        header_row = sheet_data.headers[row_idx - 1]
        if col_idx <= len(header_row):
            return header_row[col_idx - 1]
    else:
        # 数据行
        data_row_idx = row_idx - len(sheet_data.headers) - 1
        if 0 <= data_row_idx < len(sheet_data.rows):
            row_data = sheet_data.rows[data_row_idx]
            col_key = f"col_{col_idx}"
            return row_data.get(col_key)
    return None


def _get_indicator_value_dal(sheet_data, field_keywords: list[str], col_hint: str) -> Optional[float]:
    """DAL模式：从 sheet 中提取某指标在指定列的值"""
    row_idx = _find_row_by_keywords_dal(sheet_data, field_keywords)
    if row_idx is None:
        return None

    col_idx = _find_col_by_hint_dal(sheet_data, col_hint, row_idx)
    if col_idx is None:
        return None

    # 尝试解析数值
    val = _get_cell_value_dal(sheet_data, row_idx, col_idx)
    if val is None:
        return None
    
    p = parse_number(val)
    if p is not None and not (p == 0 and isinstance(val, (int, float))):
        return p
    
    # 向下回退
    for row_offset in range(1, 5):
        adj_row = row_idx + row_offset
        if adj_row > len(sheet_data.headers) + len(sheet_data.rows):
            break
        val = _get_cell_value_dal(sheet_data, adj_row, col_idx)
        p = parse_number(val)
        if p is not None and not (p == 0 and isinstance(val, (int, float))):
            return p
    
    # 向右偏移
    for offset in range(1, 5):
        adj_col = col_idx + offset
        val = _get_cell_value_dal(sheet_data, row_idx, adj_col)
        p = parse_number(val)
        if p is not None and not (p == 0 and isinstance(val, (int, float))):
            return p
    
    return None


# =============================================================================
# DAL 模式规则执行器
# =============================================================================

def run_formula_rule_dal(rule: dict, sheet_data) -> dict:
    """DAL模式：执行 formula 类型的规则"""
    formula = rule.get('formula', {})
    num_field = formula.get('numerator', {}).get('field', '')
    num_hint = formula.get('numerator', {}).get('col_hint', '')
    den_field = formula.get('denominator', {}).get('field', '')
    den_hint = formula.get('denominator', {}).get('col_hint', '')
    result_field = formula.get('result_field', '')
    result_hint = formula.get('result_col_hint', '')
    multiplier = formula.get('multiplier', 1)

    # 调试：打印表头行内容
    print(f"\n=== [DAL Formula Rule: {rule.get('name', rule.get('id'))}] ===")
    print(f"Sheet: {sheet_data.name}, 数据行数: {len(sheet_data.rows)}")
    print(f"表头行 ({len(sheet_data.headers)} 行):")
    for i, header_row in enumerate(sheet_data.headers):
        print(f"  Row {i+1}: {header_row[:8]}")  # 只打印前8列

    # 调试：打印第一行数据的所有列（了解数据结构）
    if sheet_data.rows:
        first_row = sheet_data.rows[0]
        all_cols = {k: v for k, v in first_row.items() if v is not None}
        print(f"第一行数据前10列: {dict(list(all_cols.items())[:10])}")

    num_val = _get_indicator_value_dal(sheet_data, [num_field], num_hint)
    den_val = _get_indicator_value_dal(sheet_data, [den_field], den_hint)

    if num_val is None or den_val is None or den_val == 0:
        return {
            'status': 'fail',
            'actual_value': None,
            'expected_value': None,
            'diff': None,
            'detail': f'无法提取数据（{num_field}={num_val}, {den_field}={den_val}）'
        }

    calculated = (num_val / den_val) * multiplier

    reported = None
    if result_field:
        reported = _get_indicator_value_dal(sheet_data, [result_field], result_hint)

    tolerance = rule.get('tolerance', 0.01)
    unit = rule.get('unit', '%')
    tol_type = rule.get('tolerance_type', 'percent')

    if reported is None:
        return {
            'status': 'warn',
            'actual_value': f'{calculated:.4f}{unit}',
            'expected_value': '报告值未找到',
            'diff': None,
            'detail': f'计算值={calculated:.4f}{unit}，但报告中的{result_field}未找到'
        }

    diff_abs = abs(calculated - reported)

    if tol_type == 'percent':
        diff_pct = (diff_abs / abs(reported) * 100) if reported != 0 else 0
        is_pass = diff_pct <= tolerance
        detail = f'计算={calculated:.4f}{unit}，报告={reported:.4f}{unit}，误差={diff_pct:.4f}%'
        diff_out = diff_pct
    else:
        diff_pct = None
        is_pass = diff_abs <= tolerance
        detail = f'计算={calculated:.4f}{unit}，报告={reported:.4f}{unit}，差值={diff_abs:.4f}'
        diff_out = diff_abs

    status = 'pass' if is_pass else 'fail'
    return {
        'status': status,
        'actual_value': f'{reported:.4f}{unit}',
        'expected_value': f'{calculated:.4f}{unit}',
        'diff': diff_out,
        'diff_percent': diff_pct,
        'detail': detail
    }


def run_sum_check_rule_dal(rule: dict, sheet_data) -> dict:
    """DAL模式：执行 sum_check 类型的规则"""
    sum_check = rule.get('sum_check', {})
    target_field = sum_check.get('target_field', '')
    target_hint = sum_check.get('target_col_hint', '')
    components = sum_check.get('components', [])
    sum_col_hint = sum_check.get('sum_col_hint', '')
    sum_start_row = sum_check.get('sum_start_row')
    sum_end_row = sum_check.get('sum_end_row')

    target_val = _get_indicator_value_dal(sheet_data, [target_field], target_hint)
    unit = rule.get('unit', '')
    
    # 多行求和模式
    if sum_col_hint and sum_start_row and sum_end_row:
        col_idx = _find_col_by_hint_dal(sheet_data, sum_col_hint)
        col_sum = 0.0
        cnt = 0
        for r in range(sum_start_row, sum_end_row + 1):
            v = _get_cell_value_dal(sheet_data, r, col_idx)
            p = parse_number(v)
            if p is not None:
                col_sum += p
                cnt += 1
        
        if target_val is not None:
            diff_abs = abs(target_val - col_sum)
            tolerance = rule.get('tolerance', 1000)
            is_pass = diff_abs <= tolerance
            status = 'pass' if is_pass else 'fail'
            return {
                'status': status,
                'actual_value': f'{target_val:.0f}{unit}',
                'expected_value': f'{col_sum:.0f}{unit}',
                'diff': diff_abs,
                'detail': f'目标={target_val:.0f}，多行求和({cnt}行)={col_sum:.0f}，差值={diff_abs:.0f}'
            }
        else:
            return {
                'status': 'warn',
                'actual_value': '未找到目标',
                'expected_value': f'求和={col_sum:.0f}',
                'diff': None,
                'detail': f'求和={col_sum:.0f}（{cnt}行），但目标字段 [{target_field}] 未找到'
            }

    # 单行分量模式
    component_vals = []
    for comp in components:
        comp_field = comp.get('field', '')
        comp_hint = comp.get('col_hint', '')
        sign = comp.get('sign', 1)
        val = _get_indicator_value_dal(sheet_data, [comp_field], comp_hint)
        if val is not None:
            component_vals.append(val * sign)

    tolerance = rule.get('tolerance', 1000)

    if target_val is not None and component_vals:
        calculated_sum = sum(component_vals)
        diff_abs = abs(target_val - calculated_sum)
        is_pass = diff_abs <= tolerance
        status = 'pass' if is_pass else 'fail'
        parts_str = ' + '.join([f'{v:.0f}' if v >= 0 else f'({abs(v):.0f})' for v in component_vals])
        detail = f'目标={target_val:.0f}，分项=({parts_str})={calculated_sum:.0f}，差值={diff_abs:.0f}'
        return {
            'status': status,
            'actual_value': f'{target_val:.0f}{unit}',
            'expected_value': f'{calculated_sum:.0f}{unit}',
            'diff': diff_abs,
            'detail': detail
        }
    elif target_val is None and component_vals:
        calculated_sum = sum(component_vals)
        parts_str = ' + '.join([f'{v:.0f}' if v >= 0 else f'({abs(v):.0f})' for v in component_vals])
        return {
            'status': 'warn',
            'actual_value': '未找到合计值',
            'expected_value': f'分项计算={calculated_sum:.0f}{unit}',
            'diff': None,
            'detail': f'分项之和=({parts_str})={calculated_sum:.0f}，目标行 [{target_field}] 未找到'
        }
    else:
        return {
            'status': 'fail',
            'actual_value': None,
            'expected_value': None,
            'diff': None,
            'detail': f'无法提取目标字段 [{target_field}] 和任何分项数据'
        }


def run_periodicity_rule_dal(rule: dict, sheet_data) -> dict:
    """DAL模式：执行 periodicity 类型的跨期一致性规则"""
    periodicity = rule.get('periodicity', {})
    field_keywords = [periodicity.get('field', '')]
    periods = periodicity.get('periods', [])
    tolerance = rule.get('tolerance', 0.5)
    unit = rule.get('unit', '%')

    period_vals = []
    period_labels = []
    for p in periods:
        hint = p.get('col_hint', '')
        label = p.get('label', hint)
        val = _get_indicator_value_dal(sheet_data, field_keywords, hint)
        period_vals.append(val)
        period_labels.append(label)

    valid_vals = [(l, v) for l, v in zip(period_labels, period_vals) if v is not None]

    if len(valid_vals) < 2:
        return {
            'status': 'warn',
            'actual_value': ', '.join([f'{v}' if v else 'N/A' for v in period_vals]),
            'expected_value': '需要至少2期数据',
            'diff': None,
            'detail': f'仅找到 {len(valid_vals)} 期数据，无法进行跨期比对'
        }

    issues = []
    for i in range(1, len(valid_vals)):
        prev_label, prev_val = valid_vals[i - 1]
        curr_label, curr_val = valid_vals[i]
        diff = abs(curr_val - prev_val)
        if diff > tolerance:
            issues.append(f'{prev_label}->{curr_label}: {prev_val:.4f}->{curr_val:.4f} (diff={diff:.4f}{unit})')

    if issues:
        status = 'fail'
        detail = '跨期波动超过容差: ' + '; '.join(issues)
    else:
        status = 'pass'
        detail = '各期数据连续性正常: ' + ', '.join([f'{l}={v:.4f}' for l, v in valid_vals])

    return {
        'status': status,
        'actual_value': ', '.join([f'{v:.4f}' if v else 'N/A' for _, v in valid_vals]),
        'expected_value': '跨期一致性',
        'diff': None,
        'detail': detail
    }


def run_audit(file_id: str, file_name: str, excel_path: str = None,
               rule_ids: list[str] = None, sheet_mapping: dict = None,
               data_source = None) -> dict:
    """
    对一份 Excel 执行所有勾稽规则。
    
    支持两种模式:
    1. 传统模式(excel_path): 直接传入Excel文件路径
    2. DAL模式(data_source + file_id): 使用数据访问层
    
    Args:
        file_id: 档案ID
        file_name: 档案名称
        excel_path: Excel 文件路径
        rule_ids: 可选, 要执行的规则ID列表
        sheet_mapping: 可选, {rule_id: sheet_name} 映射, 用于人工指定规则对应的 Sheet
        data_source: 数据源实例(DAL模式)
    """
    import uuid as uuid_lib

    run_uuid = str(uuid_lib.uuid4())[:8]
    started_at = datetime.now().isoformat()

    rules = load_rules()
    if rule_ids:
        rules = [r for r in rules if r['id'] in rule_ids]

    results = []
    for rule in rules:
        # 根据模式选择调用方式
        if data_source is not None:
            result = run_rule(rule, excel_path=excel_path, sheet_mapping=sheet_mapping,
                            file_id=file_id, data_source=data_source)
        else:
            result = run_rule(rule, excel_path=excel_path, sheet_mapping=sheet_mapping)
        results.append(result)

    completed_at = datetime.now().isoformat()

    pass_count = sum(1 for r in results if r['status'] == 'pass')
    warn_count = sum(1 for r in results if r['status'] == 'warn')
    fail_count = sum(1 for r in results if r['status'] == 'fail')

    return {
        'run_uuid': run_uuid,
        'file_id': file_id,
        'file_name': file_name,
        'started_at': started_at,
        'completed_at': completed_at,
        'status': 'completed',
        'total': len(results),
        'pass_count': pass_count,
        'warn_count': warn_count,
        'fail_count': fail_count,
        'results': results
    }
