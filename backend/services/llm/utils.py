

import time
from pathlib import Path

# ⭐⭐⭐ 新增：导入路径常量 ⭐⭐⭐
from backend.utils.constants import EXCEL_DATA_URL_PREFIX, EXCEL_DATA_RELATIVE_PATH, EXCEL_DATA_DIR

def validate_required_params(data, required_fields):
    """验证必要参数 - 函数体不变"""
    missing_fields = [field for field in required_fields if not data.get(field)]
    if missing_fields:
        return False, f"缺少必要参数: {', '.join(missing_fields)}"
    return True, None


def convert_to_excel_url(file_path):
    """将文件路径转换为Excel URL"""
    file_path = file_path.replace('\\', '/')
    print(f"🔧 转换Excel路径: {file_path}")

    # ⭐⭐⭐ 修复：使用常量路径 ⭐⭐⭐
    # 初始化相对路径
    relative_path = ""

    # 处理各种可能的路径格式
    excel_data_marker = f"{EXCEL_DATA_RELATIVE_PATH}/"
    if excel_data_marker in file_path:
        # 提取相对路径部分
        parts = file_path.split(excel_data_marker)
        if len(parts) > 1:
            relative_path = parts[1]

    # 如果路径已经是相对路径
    elif file_path.startswith(excel_data_marker):
        relative_path = file_path.replace(excel_data_marker, '')

    else:
        # 从路径对象中提取
        path_obj = Path(file_path)

        # 查找 excel_data 在路径中的位置
        try:
            if 'excel_data' in path_obj.parts:
                excel_data_index = path_obj.parts.index('excel_data')
                if excel_data_index + 1 < len(path_obj.parts):
                    relative_parts = path_obj.parts[excel_data_index + 1:]
                    relative_path = '/'.join(relative_parts)
        except (ValueError, IndexError):
            pass

        # 如果上述方法都没找到，使用回退方案
        if not relative_path:
            # ⭐⭐⭐ 修复：使用常量路径进行相对路径计算 ⭐⭐⭐
            try:
                relative_path = str(Path(file_path).relative_to(EXCEL_DATA_DIR))
                relative_path = relative_path.replace('\\', '/')
            except ValueError:
                # 如果不在EXCEL_DATA_DIR下，使用文件名
                file_name = Path(file_path).name
                folder_name = Path(file_path).parent.name
                relative_path = f"{folder_name}/{file_name}"

    # 清理路径（移除开头的斜杠）
    relative_path = relative_path.lstrip('/')

    # ⭐⭐⭐ 修复：使用常量URL前缀 ⭐⭐⭐
    excel_url = f"{EXCEL_DATA_URL_PREFIX}/{relative_path}"
    print(f"✅ 生成的Excel URL: {excel_url}")

    return excel_url


def ensure_excel_file_exists(file_path, processing_results=None):
    """
    确保Excel文件存在，如果不存在则创建空文件

    Args:
        file_path: Excel文件路径
        processing_results: 处理结果信息

    Returns:
        bool: 文件是否存在或创建成功
    """
    file_path = Path(file_path)

    if file_path.exists():
        print(f"✅ Excel文件已存在: {file_path}")
        return True
    else:
        print(f"🔄 Excel文件不存在，创建空文件: {file_path}")
        return create_empty_excel_file(file_path, processing_results)


def create_empty_excel_file(file_path, processing_results=None):
    """
    创建空的Excel文件（基础版本）

    Args:
        file_path: Excel文件路径
        processing_results: 处理结果信息
    """
    try:
        import openpyxl

        # 确保目录存在
        file_path = Path(file_path)
        file_path.parent.mkdir(parents=True, exist_ok=True)

        # 创建工作簿
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "处理结果"

        # 添加基本信息
        worksheet['A1'] = "批量处理完成"
        worksheet['A2'] = f"总图片数: {processing_results.get('total', 0) if processing_results else 0}"
        worksheet['A3'] = f"成功: {processing_results.get('success', 0) if processing_results else 0}"
        worksheet['A4'] = f"失败: {processing_results.get('failed', 0) if processing_results else 0}"
        worksheet['A5'] = f"生成时间: {time.strftime('%Y-%m-%d %H:%M:%S')}"

        workbook.save(file_path)
        print(f"✅ 已创建空Excel文件: {file_path}")
        return True

    except Exception as e:
        print(f"❌ 创建空Excel文件失败: {str(e)}")
        return False








def export_processing_results_to_excel(file_path, results_data):
    """
    将处理结果导出到Excel文件（修复版本）
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment

        # 确保目录存在
        file_path = Path(file_path)
        file_path.parent.mkdir(parents=True, exist_ok=True)

        # 创建工作簿
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "批量处理结果"

        # 设置标题
        title_cell = worksheet['A1']
        title_cell.value = "表格识别批量处理报告"
        title_cell.font = Font(bold=True, size=16)
        worksheet.merge_cells('A1:F1')  # 修复：合并更多列以适应标题
        title_cell.alignment = Alignment(horizontal='center')

        # 基本信息
        worksheet['A3'] = "处理概览"
        worksheet['A3'].font = Font(bold=True, size=12)

        overview_data = [
            ("处理时间", time.strftime('%Y-%m-%d %H:%M:%S')),
            ("总文件数", results_data.get('total', 0)),
            ("成功识别", results_data.get('success', 0)),
            ("识别失败", results_data.get('failed', 0)),
            ("成功率", f"{results_data.get('success', 0) / max(results_data.get('total', 1), 1) * 100:.1f}%"),
            ("表格类型", results_data.get('table_type', '未知')),
            ("处理状态", "完成" if results_data.get('success', 0) > 0 else "部分失败")
        ]

        for i, (label, value) in enumerate(overview_data, start=4):
            worksheet[f'A{i}'] = label
            worksheet[f'B{i}'] = value
            worksheet[f'A{i}'].font = Font(bold=True)

        # 详细结果
        start_row = len(overview_data) + 6

        if results_data.get('results'):
            # 表头
            headers = ["序号", "图片路径", "表格名称", "处理状态", "复杂度", "错误信息"]
            for col, header in enumerate(headers, start=1):
                cell = worksheet.cell(row=start_row, column=col)
                cell.value = header
                cell.font = Font(bold=True)

            # 数据行
            for idx, result in enumerate(results_data.get('results', []), start=1):
                row = start_row + idx
                worksheet.cell(row=row, column=1).value = idx
                worksheet.cell(row=row, column=2).value = result.get('image_path', '')
                worksheet.cell(row=row, column=3).value = result.get('table_name', '')

                # 处理状态
                status = result.get('status', '')
                if status == 'success':
                    status_text = "成功"
                elif status == 'non_financial':
                    status_text = "非金融表格"
                else:
                    status_text = "失败"
                worksheet.cell(row=row, column=4).value = status_text

                worksheet.cell(row=row, column=5).value = result.get('complexity', '')
                worksheet.cell(row=row, column=6).value = result.get('error', result.get('error_message', ''))

        # 修复：安全的列宽调整
        _adjust_column_widths_safely(worksheet)

        # 保存文件
        workbook.save(file_path)
        print(f"✅ 处理结果已导出到: {file_path}")
        return True

    except Exception as e:
        print(f"❌ 导出处理结果失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


def _adjust_column_widths_safely(worksheet):
    """
    安全地调整列宽，避免合并单元格的问题
    """
    try:
        for column in worksheet.columns:
            # 跳过合并的单元格
            if hasattr(column[0], 'column_letter'):
                column_letter = column[0].column_letter
                max_length = 0

                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
    except Exception as e:
        print(f"⚠️ 调整列宽时出现警告: {str(e)}")
        # 忽略列宽调整的错误，不影响主要功能