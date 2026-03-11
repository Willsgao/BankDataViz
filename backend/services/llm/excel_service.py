

from pathlib import Path
from flask import request, jsonify, send_from_directory
from backend.utils.constants import EXCEL_DATA_DIR

from backend.services.llm.task_management_service import logger

def serve_excel_data(excel_path):
    """提供Excel文件数据访问"""
    try:
        # ⭐⭐⭐ 修复：使用常量路径 ⭐⭐⭐
        file_path = EXCEL_DATA_DIR / excel_path

        if not file_path.exists():
            return jsonify({
                "success": False,
                "error": f"Excel文件不存在: {excel_path}"
            }), 404

        # ⭐⭐⭐ 修复：使用正确的目录和文件名 ⭐⭐⭐
        return send_from_directory(EXCEL_DATA_DIR, excel_path)

    except Exception as e:
        logger.error(f"提供Excel文件失败: {str(e)}")
        return jsonify({
            "success": False,
            "error": f"文件访问失败: {str(e)}"
        }), 500


def check_excel_internal():
    """检查Excel文件是否存在"""
    try:
        file_path = request.args.get('path')
        print(f"🔍 检查Excel文件请求 - path: {file_path}")

        if not file_path:
            return jsonify({
                "success": False,
                "error": "缺少path参数"
            }), 400

        # 构建完整路径
        full_path = Path(file_path)
        if not full_path.is_absolute():
            full_path = Path.cwd() / file_path

        print(f"🔍 检查Excel完整路径: {full_path}")
        print(f"🔍 文件是否存在: {full_path.exists()}")

        if full_path.exists():
            try:
                import openpyxl
                workbook = openpyxl.load_workbook(full_path)
                sheet_name = workbook.sheetnames[0]
                worksheet = workbook[sheet_name]

                # 转换为JSON数据
                data = []
                headers = []

                # 读取表头（第一行）
                for col in range(1, worksheet.max_column + 1):
                    cell_value = worksheet.cell(row=1, column=col).value
                    headers.append(str(cell_value) if cell_value is not None else f"列{col}")

                # 读取数据行
                for row in range(2, worksheet.max_row + 1):
                    row_data = {}
                    for col, header in enumerate(headers, 1):
                        cell_value = worksheet.cell(row=row, column=col).value
                        row_data[header] = str(cell_value) if cell_value is not None else ""
                    if any(row_data.values()):  # 只添加非空行
                        data.append(row_data)

                print(f"🔍 Excel数据读取成功 - 表头: {headers}, 数据行数: {len(data)}")

                return jsonify({
                    "success": True,
                    "exists": True,
                    "excelData": {
                        "headers": headers,
                        "data": data,
                        "sheet_name": sheet_name,
                        "file_path": str(full_path)
                    }
                })
            except Exception as e:
                logger.error(f"读取Excel文件失败: {str(e)}")
                return jsonify({
                    "success": False,
                    "error": f"读取Excel文件失败: {str(e)}"
                }), 500
        else:
            print(f"🔍 Excel文件不存在: {file_path}")
            return jsonify({
                "success": True,
                "exists": False,
                "message": f"Excel文件不存在: {file_path}"
            })

    except Exception as e:
        logger.error(f"检查Excel失败: {str(e)}")
        return jsonify({
            "success": False,
            "error": f"检查Excel失败: {str(e)}"
        }), 500


def get_excel_data_internal():
    """读取Excel文件内容并返回给前端"""
    try:
        excel_url = request.args.get('url')
        if not excel_url:
            return jsonify({
                "success": False,
                "error": "缺少url参数"
            }), 400

        # 将URL转换为文件路径
        if excel_url.startswith('/static/'):
            file_path = excel_url[1:]  # 去掉开头的斜杠
        else:
            file_path = excel_url

        full_path = Path(file_path)
        if not full_path.is_absolute():
            full_path = Path.cwd() / file_path

        print(f"读取Excel文件: {full_path}")

        if not full_path.exists():
            return jsonify({
                "success": False,
                "error": f"Excel文件不存在: {file_path}"
            }), 404

        # 读取Excel文件
        import openpyxl
        workbook = openpyxl.load_workbook(full_path)
        sheet_name = workbook.sheetnames[0]
        worksheet = workbook[sheet_name]

        # 转换为JSON数据
        headers = []
        data_rows = []

        # 读取表头（第一行）
        for col in range(1, worksheet.max_column + 1):
            cell_value = worksheet.cell(row=1, column=col).value
            headers.append(str(cell_value) if cell_value is not None else f"列{col}")

        # 读取数据行
        for row in range(2, worksheet.max_row + 1):
            row_data = {}
            for col, header in enumerate(headers, 1):
                cell_value = worksheet.cell(row=row, column=col).value
                row_data[header] = str(cell_value) if cell_value is not None else ""
            if any(row_data.values()):  # 只添加非空行
                data_rows.append(row_data)

        return jsonify({
            "success": True,
            "data": {
                "headers": headers,
                "data": data_rows,
                "tableName": sheet_name,
                "excelPath": str(full_path)
            }
        })

    except Exception as e:
        logger.error(f"读取Excel数据失败: {str(e)}")
        return jsonify({
            "success": False,
            "error": f"读取Excel数据失败: {str(e)}"
        }), 500


def get_excel_content_internal(excel_url):
    """读取Excel文件内容并返回结构化数据 - 修复版本"""
    try:
        print(f"📖 读取Excel内容: {excel_url}")

        if not excel_url:
            return {
                "success": False,
                "error": "缺少excel_url参数"
            }

        # 处理URL编码
        import urllib.parse
        excel_url = urllib.parse.unquote(excel_url)
        print(f"🔍 解码后excel_url: {excel_url}")

        file_path = None

        if excel_url.startswith('/api/excel-data/'):
            relative_path = excel_url.replace('/api/excel-data/', '')
            file_path = EXCEL_DATA_DIR / relative_path
        elif excel_url.startswith('/static/excel_output/'):
            relative_path = excel_url.replace('/static/excel_output/', '')
            backend_dir = Path(__file__).parent.parent
            file_path = backend_dir / 'static' / 'excel_output' / relative_path
        elif excel_url.startswith('/static/excel_data/'):
            relative_path = excel_url.replace('/static/excel_data/', '')
            file_path = EXCEL_DATA_DIR / relative_path
        else:
            file_path = Path(excel_url)

        print(f"🔍 最终文件路径: {file_path}")

        if not file_path.exists():
            return {
                "success": False,
                "error": f"Excel文件不存在: {file_path}"
            }

        # 读取Excel文件内容
        import openpyxl
        workbook = openpyxl.load_workbook(file_path)

        # 获取所有工作表
        sheet_data = []

        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]

            # 转换为JSON数据
            headers = []
            data_rows = []

            # 查找数据开始行（跳过空行和标题行）
            data_start_row = 1
            max_col = worksheet.max_column

            # 读取表头（第一行有数据的行）
            for row in range(1, worksheet.max_row + 1):
                row_has_data = False
                for col in range(1, max_col + 1):
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value and str(cell_value).strip():
                        row_has_data = True
                        break

                if row_has_data:
                    # 这一行有数据，作为表头
                    for col in range(1, max_col + 1):
                        cell_value = worksheet.cell(row=row, column=col).value
                        headers.append(str(cell_value) if cell_value is not None else f"列{col}")
                    data_start_row = row + 1
                    break

            # 读取数据行
            for row in range(data_start_row, worksheet.max_row + 1):
                row_data = {}
                has_data = False

                for col, header in enumerate(headers, 1):
                    if col > max_col:
                        break
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value is not None and str(cell_value).strip():
                        has_data = True
                    row_data[header] = str(cell_value) if cell_value is not None else ""

                if has_data:  # 只添加有数据的行
                    data_rows.append(row_data)

            sheet_data.append({
                "sheetName": sheet_name,
                "headers": headers,
                "data": data_rows,
                "rowCount": len(data_rows),
                "colCount": len(headers)
            })

        # ⭐⭐⭐ 返回正确的数据格式 ⭐⭐⭐
        result = {
            "success": True,
            "data": {
                "filePath": str(file_path),
                "sheets": sheet_data,
                "totalSheets": len(sheet_data),
                # 为了兼容性，也返回平铺的数据
                "headers": sheet_data[0]["headers"] if sheet_data else [],
                "data": sheet_data[0]["data"] if sheet_data else []
            }
        }

        print(f"✅ 返回数据: 共{len(sheet_data)}个工作表")
        return result

    except Exception as e:
        logger.error(f"读取Excel内容失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return {
            "success": False,
            "error": f"读取Excel内容失败: {str(e)}"
        }