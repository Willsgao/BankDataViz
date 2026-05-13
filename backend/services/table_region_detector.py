# -*- coding: UTF-8 -*-
"""
智能识别 - 表格区域检测服务
使用 PyMuPDF find_tables() 检测 PDF 中每个表格的精确边界框。
"""
import base64
import io
import logging
from typing import List, Optional, Dict, Any
from pathlib import Path

logger = logging.getLogger(__name__)

# 默认渲染 DPI
DEFAULT_DPI = 150


def _detect_tables_in_pdf(pdf_path: Path, dpi: int = DEFAULT_DPI) -> List[Dict[str, Any]]:
    """
    使用 PyMuPDF find_tables() 检测 PDF 中的所有表格区域。

    Returns:
        List[{
            "page_idx": int,         # 0-based 页码
            "page_width": float,     # PDF 页面宽度（points）
            "page_height": float,    # PDF 页面高度（points）
            "tables": [{
                "id": int,            # 区域内表格序号
                "bbox": [x0, y0, x1, y1],  # PDF 坐标
                "rows": int,
                "cols": int,
            }]
        }]
    """
    import fitz

    results = []

    try:
        doc = fitz.open(str(pdf_path))
        for page_idx in range(len(doc)):
            page = doc[page_idx]
            tabs = page.find_tables()

            if not tabs.tables:
                continue

            tables = []
            for i, tab in enumerate(tabs.tables):
                bbox = tab.bbox  # fitz.Rect -> tuple (x0, y0, x1, y1)
                tables.append({
                    "id": i,
                    "bbox": [round(float(v), 1) for v in bbox],
                    "rows": tab.row_count if hasattr(tab, "row_count") else 0,
                    "cols": tab.col_count if hasattr(tab, "col_count") else 0,
                })

            results.append({
                "page_idx": page_idx,
                "page_width": round(page.rect.width, 1),
                "page_height": round(page.rect.height, 1),
                "tables": tables,
            })

        # 补充没有表格的页（只渲染，不检测表格）
        found_pages = {r["page_idx"] for r in results}
        for page_idx in range(len(doc)):
            if page_idx not in found_pages:
                results.append({
                    "page_idx": page_idx,
                    "page_width": round(doc[page_idx].rect.width, 1),
                    "page_height": round(doc[page_idx].rect.height, 1),
                    "tables": [],
                })

        results.sort(key=lambda x: x["page_idx"])

        doc.close()
        logger.info(f"检测到 {len(results)} 页含表格")
        return results

    except Exception as e:
        logger.exception(f"PDF 表格检测失败: {pdf_path}")
        return []


def _render_pdf_page(
    pdf_path: Path,
    page_idx: int,
    dpi: int = DEFAULT_DPI
) -> Optional[Dict[str, Any]]:
    """
    渲染指定 PDF 页为 PNG 图片（带坐标缩放系数）。

    Returns:
        {
            "image_base64": str,    # data:image/png;base64,...
            "scale_x": float,       # 图片宽度 / PDF 宽度
            "scale_y": float,       # 图片高度 / PDF 高度
            "image_width": int,
            "image_height": int,
            "pdf_width": float,
            "pdf_height": float,
        }
    """
    import fitz
    from PIL import Image

    try:
        doc = fitz.open(str(pdf_path))
        page = doc[page_idx]

        zoom = dpi / 72.0
        mat = fitz.Matrix(zoom, zoom)

        # 渲染为 pixmap
        pix = page.get_pixmap(matrix=mat)

        # 转为 PIL Image
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)

        # 转 base64（必须在 doc.close() 之前做）
        buf = io.BytesIO()
        img.save(buf, format="PNG", optimize=False)
        img_bytes = buf.getvalue()
        b64 = "data:image/png;base64," + base64.b64encode(img_bytes).decode()

        # 缩放系数（必须在 doc.close() 之前访问 page.rect）
        scale_x = pix.width / page.rect.width
        scale_y = pix.height / page.rect.height
        pdf_width = round(page.rect.width, 1)
        pdf_height = round(page.rect.height, 1)

        doc.close()

        return {
            "image_base64": b64,
            "scale_x": round(scale_x, 4),
            "scale_y": round(scale_y, 4),
            "image_width": pix.width,
            "image_height": pix.height,
            "pdf_width": pdf_width,
            "pdf_height": pdf_height,
        }

    except Exception as e:
        logger.exception(f"PDF 页面渲染失败: {pdf_path} page={page_idx}")
        return None


def _detect_tables_in_excel(file_path: Path) -> List[Dict[str, Any]]:
    """
    检测 Excel 文件中每个 sheet 的表格区域（used range）。

    Returns:
        List[{
            "sheet_name": str,
            "bbox": [x0, y0, x1, y1],   # 列/行索引 (0-based)
            "rows": int,
            "cols": int,
        }]
    """
    try:
        from openpyxl import load_workbook

        results = []
        wb = load_workbook(str(file_path), read_only=True, data_only=True)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws.max_row and ws.max_column:
                results.append({
                    "sheet_name": sheet_name,
                    "bbox": [1, 1, ws.max_column, ws.max_row],  # Excel-style 1-based
                    "rows": ws.max_row,
                    "cols": ws.max_column,
                })

        wb.close()
        return results
    except Exception as e:
        logger.exception(f"Excel 表格检测失败: {file_path}")
        return []


def detect_tables(file_path: Path, dpi: int = DEFAULT_DPI) -> Dict[str, Any]:
    """
    统一入口：根据文件类型自动选择检测策略。

    Args:
        file_path: 文件路径
        dpi: PDF 渲染 DPI

    Returns:
        {
            "success": bool,
            "file_type": "pdf" | "excel" | "image",
            "total_pages": int,          # PDF/Excel sheet 数
            "pages": [
                {
                    "page_idx": int,
                    "page_name": str,     # PDF: "page N" / Excel: sheet 名
                    "table_count": int,   # 检测到的表格数
                    "tables": [...],       # 同 _detect_tables_in_pdf/excel
                    "render": {...} | None, # PDF: 渲染图信息 / Excel: None
                },
                ...
            ],
            "total_tables": int,
            "error": str | None,
        }
    """
    suffix = file_path.suffix.lower()
    pages_result = []
    total_tables = 0

    if suffix == ".pdf":
        tables_by_page = _detect_tables_in_pdf(file_path, dpi)

        for page_info in tables_by_page:
            page_idx = page_info["page_idx"]
            render_info = _render_pdf_page(file_path, page_idx, dpi)

            # 将 PDF 坐标转为图片像素坐标
            tables_pixel = []
            if render_info:
                sx = render_info["scale_x"]
                sy = render_info["scale_y"]
                for t in page_info["tables"]:
                    bbox_pdf = t["bbox"]
                    bbox_pixel = [
                        round(bbox_pdf[0] * sx, 1),
                        round(bbox_pdf[1] * sy, 1),
                        round(bbox_pdf[2] * sx, 1),
                        round(bbox_pdf[3] * sy, 1),
                    ]
                    tables_pixel.append({
                        "id": t["id"],
                        "bbox_pdf": bbox_pdf,
                        "bbox_pixel": bbox_pixel,
                        "rows": t["rows"],
                        "cols": t["cols"],
                    })

            pages_result.append({
                "page_idx": page_idx,
                "page_name": f"第 {page_idx + 1} 页",
                "table_count": len(tables_pixel),
                "tables": tables_pixel,
                "render": render_info,
            })
            total_tables += len(tables_pixel)

        # 获取 PDF 总页数
        import fitz
        doc = fitz.open(str(file_path))
        total_pages = len(doc)
        doc.close()

        return {
            "success": True,
            "file_type": "pdf",
            "total_pages": total_pages,
            "pages": pages_result,
            "total_tables": total_tables,
            "error": None,
        }

    elif suffix in (".xlsx", ".xls"):
        tables_by_sheet = _detect_tables_in_excel(file_path)
        for info in tables_by_sheet:
            pages_result.append({
                "page_idx": 0,
                "page_name": info["sheet_name"],
                "table_count": 1,
                "tables": [info],
                "render": None,
            })
            total_tables += 1

        return {
            "success": True,
            "file_type": "excel",
            "total_pages": len(pages_result),
            "pages": pages_result,
            "total_tables": total_tables,
            "error": None,
        }

    elif suffix in (".png", ".jpg", ".jpeg", ".bmp", ".gif"):
        # 图片类型暂不支持自动检测，返回空，由用户手动框选
        return {
            "success": True,
            "file_type": "image",
            "total_pages": 1,
            "pages": [{
                "page_idx": 0,
                "page_name": "图片",
                "table_count": 0,
                "tables": [],
                "render": None,
            }],
            "total_tables": 0,
            "error": None,
        }

    else:
        return {
            "success": False,
            "file_type": "unknown",
            "total_pages": 0,
            "pages": [],
            "total_tables": 0,
            "error": f"不支持的文件类型: {suffix}",
        }
