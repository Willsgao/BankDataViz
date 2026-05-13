# -*- coding:utf-8 -*-
"""
Excel 数据源实现

从 Excel 文件读取数据，实现 DataSource 接口。
"""
import os
import re
import openpyxl
from datetime import datetime
from typing import List, Dict, Any, Optional

from backend.services.data_access_layer import (
    DataSource, DataSourceFactory,
    FileInfo, SheetSummary, SheetData
)


def _get_project_root() -> str:
    """获取项目根目录"""
    # backend/services/dal/excel_source.py
    # backend/services/dal/ -> backend/services/ -> backend/ -> project_root/
    current = os.path.dirname(os.path.abspath(__file__))
    # 向上3层: dal -> services -> backend -> project_root
    for _ in range(3):
        current = os.path.dirname(current)
    return current


class ExcelDataSource(DataSource):
    """
    Excel 文件数据源
    
    从 Excel 文件读取数据，兼容现有 PDF 解析系统生成的文件。
    
    文件存储结构：
    - Excel数据根目录: {project_root}/data/backend/static/excel_data/
    - 单个档案: {project_root}/data/backend/static/excel_data/{file_id}/
    - Excel文件: {project_root}/data/backend/static/excel_data/{file_id}/*_合并.xlsx
    """
    
    def __init__(self, base_path: str = None, db_path: str = None):
        """
        Args:
            base_path: 数据文件根目录
            db_path: SQLite 数据库路径（用于获取文件列表）
        """
        # 获取默认路径
        if base_path is None:
            project_root = _get_project_root()
            base_path = os.path.join(project_root, 'data', 'backend', 'static')
        
        self.base_path = base_path
        self.excel_data_root = os.path.join(base_path, 'excel_data')
        
        # 数据库路径
        if db_path is None:
            project_root = _get_project_root()
            db_path = os.path.join(project_root, 'data', 'database.db')
        self.db_path = db_path
    
    def _get_excel_path(self, file_id: str) -> str:
        """根据 file_id 获取 Excel 文件路径"""
        folder = os.path.join(self.excel_data_root, file_id)
        
        if not os.path.isdir(folder):
            raise FileNotFoundError(f"档案目录不存在: {folder}")
        
        # 优先查找 _合并.xlsx 文件
        for fname in os.listdir(folder):
            if fname.endswith('.xlsx') and '_合并' in fname:
                return os.path.join(folder, fname)
        
        # 兜底：取第一个 xlsx
        for fname in os.listdir(folder):
            if fname.endswith('.xlsx'):
                return os.path.join(folder, fname)
        
        raise FileNotFoundError(f"档案目录中没有 Excel 文件: {folder}")
    
    def _file_id_to_db_id(self, file_id: str) -> int:
        """将文件路径ID转换为数据库ID"""
        try:
            return int(file_id)
        except:
            return None
    
    def get_file_list(self) -> List[FileInfo]:
        """获取档案列表"""
        files = []
        
        # 方式1：从数据库获取档案信息
        if os.path.exists(self.db_path):
            try:
                import sqlite3
                conn = sqlite3.connect(self.db_path)
                cursor = conn.cursor()
                
                # 查询已处理的档案
                cursor.execute("""
                    SELECT id, filename, bank_name, upload_time, file_type
                    FROM files 
                    WHERE deleted = 0
                    ORDER BY upload_time DESC
                """)
                
                for row in cursor.fetchall():
                    file_id, filename, bank_name, upload_time, file_type = row
                    
                    # 检查 Excel 文件是否存在
                    excel_folder = os.path.join(self.excel_data_root, str(file_id))
                    if os.path.isdir(excel_folder):
                        excel_files = [f for f in os.listdir(excel_folder) 
                                       if f.endswith('.xlsx')]
                        if excel_files:
                            files.append(FileInfo(
                                id=str(file_id),
                                name=filename or bank_name or f"档案{file_id}",
                                source_path=excel_folder,
                                source_type='excel',
                                created_at=str(upload_time) if upload_time else None
                            ))
                
                conn.close()
            except Exception as e:
                print(f"数据库查询失败: {e}")
        
        # 方式2：从目录扫描（兜底）
        if not files and os.path.isdir(self.excel_data_root):
            for folder_name in os.listdir(self.excel_data_root):
                folder_path = os.path.join(self.excel_data_root, folder_name)
                if os.path.isdir(folder_path):
                    excel_files = [f for f in os.listdir(folder_path) 
                                   if f.endswith('.xlsx')]
                    if excel_files:
                        files.append(FileInfo(
                            id=folder_name,
                            name=excel_files[0],
                            source_path=folder_path,
                            source_type='excel'
                        ))
        
        return files
    
    def get_file_info(self, file_id: str) -> Optional[FileInfo]:
        """获取指定档案信息"""
        files = self.get_file_list()
        for f in files:
            if f.id == file_id:
                return f
        return None
    
    def get_sheet_names(self, file_id: str) -> List[str]:
        """获取某个档案的所有Sheet名称"""
        try:
            excel_path = self._get_excel_path(file_id)
        except FileNotFoundError:
            return []
        
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
            return sheet_names
        except Exception as e:
            print(f"读取Excel失败: {e}")
            return []
    
    def get_sheet_summary(self, file_id: str, sheet_name: str) -> SheetSummary:
        """获取Sheet摘要（用于前端预览）"""
        excel_path = self._get_excel_path(file_id)
        
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb[sheet_name]
        
        # 表头预览（前3行）
        header_preview = []
        for row_idx in range(1, min(4, ws.max_row + 1)):
            row_cells = []
            for col_idx in range(1, min(ws.max_column + 1, 10)):
                val = ws.cell(row_idx, col_idx).value
                if val is not None:
                    val_str = str(val).replace('\n', ' ').strip()[:50]
                    row_cells.append(val_str)
                else:
                    row_cells.append('')
            if any(row_cells):  # 只保留有内容的行
                header_preview.append(row_cells)
        
        # 数据行预览（前3行数据）
        row_preview = []
        data_start_row = 3  # 假设数据从第4行开始
        for row_idx in range(data_start_row, min(data_start_row + 3, ws.max_row + 1)):
            row_data = {}
            for col_idx in range(1, min(ws.max_column + 1, 10)):
                val = ws.cell(row_idx, col_idx).value
                row_data[f"col_{col_idx}"] = str(val) if val is not None else ''
            if any(row_data.values()):
                row_preview.append(row_data)
        
        wb.close()
        
        return SheetSummary(
            name=sheet_name,
            row_count=ws.max_row,
            col_count=ws.max_column,
            header_preview=header_preview,
            row_preview=row_preview
        )
    
    def get_sheet_data(self, file_id: str, sheet_name: str) -> SheetData:
        """获取Sheet完整数据（用于规则引擎）
        
        数据格式：
        - headers: 前N行表头，用于列定位
        - rows: 数据行，每行包含纵向表头（col_1）和其他列的数据
          例: {"col_1": "核心一级资本净额", "col_2": 123456789, "col_3": 987654321}
        """
        excel_path = self._get_excel_path(file_id)
        
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb[sheet_name]
        
        # 读取表头（前3行）
        headers = []
        header_row_count = 0
        for row_idx in range(1, min(4, ws.max_row + 1)):
            row_cells = []
            has_content = False
            for col_idx in range(1, ws.max_column + 1):
                val = ws.cell(row_idx, col_idx).value
                if val is not None:
                    val_str = str(val).replace('\n', ' ').strip()
                    row_cells.append(val_str)
                    has_content = True
                else:
                    row_cells.append('')
            if has_content:
                headers.append(row_cells)
                header_row_count += 1
        
        # 读取数据行（保留第一列的纵向表头）
        rows = []
        data_start_row = header_row_count + 1
        for row_idx in range(data_start_row, ws.max_row + 1):
            row_data = {}
            has_content = False
            
            for col_idx in range(1, ws.max_column + 1):
                val = ws.cell(row_idx, col_idx).value
                col_key = f"col_{col_idx}"
                if val is not None:
                    # 保留原始值，包括纵向表头的文本
                    row_data[col_key] = val
                    has_content = True
                else:
                    row_data[col_key] = None
            
            if has_content:
                rows.append(row_data)
        
        wb.close()
        
        return SheetData(
            name=sheet_name,
            headers=headers,
            rows=rows,
            raw_ws=None
        )
    
    def get_all_sheets_data(self, file_id: str) -> List[SheetData]:
        """获取档案所有Sheet数据（批量获取）"""
        sheet_names = self.get_sheet_names(file_id)
        sheets_data = []
        
        for sheet_name in sheet_names:
            try:
                sheet_data = self.get_sheet_data(file_id, sheet_name)
                sheets_data.append(sheet_data)
            except Exception as e:
                print(f"读取Sheet失败 {sheet_name}: {e}")
                continue
        
        return sheets_data
    
    def get_excel_path(self, file_id: str) -> str:
        """获取Excel文件路径（用于兼容现有代码）"""
        return self._get_excel_path(file_id)


# 注册到工厂
DataSourceFactory.register('excel', ExcelDataSource)
