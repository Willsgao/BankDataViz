#!/usr/bin/env python3
"""
表格处理后台Worker
独立进程，从Redis队列获取任务并处理
保留原有的防重复处理逻辑
"""

import os
import sys
import json
import time
import redis
import signal
import sqlite3
import threading
from pathlib import Path
from datetime import datetime
from typing import Dict, Any, Optional

# 添加项目根目录到Python路径
project_root = Path(__file__).parent.parent.parent
sys.path.insert(0, str(project_root))
# ==================================================

from backend.utils.redis_util import redis_hset_compatible

try:
    # 使用工厂模式创建Flask应用
    from backend.app_factory import create_app

    app = create_app()
    print("✅ 使用工厂模式创建Flask应用")
except ImportError as e:
    print(f"❌ 无法导入app_factory: {e}")
    sys.exit(1)

# 导入原有的处理函数
try:
    from backend.api.convert.table_processor import TableProcessingService, process_table_images_real, process_images_with_real_time_updates,pdf_aggregator_manager
    from backend.api.convert_apis import progress_tracker
    from backend.src.incremental_processor import incremental_processor

    print("✅ 导入业务模块成功")
except ImportError as e:
    print(f"❌ 无法导入业务模块: {e}")
    sys.exit(1)

from backend.utils.constants import PROJECT_ROOT_STR, EXCEL_DATA_DIR, DATABASE_PATH

class TableProcessingWorker:
    """表格处理Worker类"""

    def __init__(self, worker_id: str = None):
        self.worker_id = worker_id or f"worker_{os.getpid()}_{int(time.time())}"
        self.running = False
        self.redis_client = None
        self.current_job = None

        # 初始化信号处理
        signal.signal(signal.SIGINT, self.signal_handler)
        signal.signal(signal.SIGTERM, self.signal_handler)

        # 添加增量处理器引用
        try:
            self.incremental_processor = incremental_processor
        except ImportError as e:
            print(f"⚠️ 无法导入增量处理器: {e}")
            self.incremental_processor = None

    def signal_handler(self, signum, frame):
        """处理退出信号"""
        print(f"\n🛑 收到退出信号，停止Worker: {self.worker_id}")
        self.running = False

        # 如果正在处理任务，标记为中断
        if self.current_job:
            # 使用兼容函数
            redis_hset_compatible(self.redis_client, f"table:job:{self.current_job}", {
                "status": "interrupted",
                "message": f"Worker {self.worker_id} 被中断",
                "interrupted_at": datetime.now().isoformat()
            })

    def connect_redis(self) -> bool:
        """连接到Redis - 优化版本，增加重试机制"""
        try:
            self.redis_client = redis.Redis(
                host='localhost',
                port=6379,
                db=0,
                decode_responses=False,  # 保持bytes以便JSON序列化
                socket_connect_timeout=5,
                socket_timeout=15,  # 增加套接字超时时间
                retry_on_timeout=True,  # 新增：超时重试
                health_check_interval=30  # 新增：健康检查间隔
            )

            # 测试连接
            self.redis_client.ping()
            print(f"✅ Redis连接成功: {self.worker_id}")
            return True

        except redis.exceptions.ConnectionError as e:
            print(f"❌ Redis连接失败: {e}")
            return False
        except Exception as e:
            print(f"❌ Redis连接异常: {e}")
            return False

    def filter_processed_images(self, pdf_folder, image_names, output_dir=None):
        """
        过滤已处理的图片 - 使用极简增量处理器
        直接从原有代码复制过来
        """
        print("=" * 60)
        print(f"🔍 增量处理检查: {pdf_folder}")
        print("=" * 60)

        # 1. 使用增量处理器过滤图片
        images_to_process = self.incremental_processor.filter_processed_images(pdf_folder, image_names)

        # 2. 计算跳过的图片
        skipped_images = [img for img in image_names if img not in images_to_process]

        # 3. 获取处理统计
        stats = self.incremental_processor.get_processing_stats(pdf_folder, image_names)

        # 4. 输出详细结果
        print("\n📊 增量处理统计:")
        print(f"  ├─ 总图片数: {stats['total_images']}")
        print(f"  ├─ 已处理: {stats['processed_images']}")
        print(f"  ├─ 未处理: {stats['unprocessed_images']}")
        print(f"  └─ 进度: {stats['progress_percentage']:.1f}%")

        if stats['is_completed']:
            print(f"  🎯 处理完成: 所有图片都已处理")

        if skipped_images:
            print(f"\n⏭️ 跳过的图片 ({len(skipped_images)}张):")
            for i, img in enumerate(skipped_images[:5]):
                print(f"  {i + 1}. {img}")
            if len(skipped_images) > 5:
                print(f"  ... 等 {len(skipped_images) - 5} 张")

        if images_to_process:
            print(f"\n🆕 需要处理的图片 ({len(images_to_process)}张):")
            for i, img in enumerate(images_to_process[:5]):
                print(f"  {i + 1}. {img}")
            if len(images_to_process) > 5:
                print(f"  ... 等 {len(images_to_process) - 5} 张")

        print("=" * 60)

        # 返回格式保持兼容
        processed_set = set(self.incremental_processor.records.get(pdf_folder, []))
        return images_to_process, skipped_images, processed_set

    def update_job_status_old(self, job_id: str, status_updates: Dict[str, Any]):
        """更新任务状态到Redis - 优化版本"""
        try:
            # 确保所有值是字符串
            updates = {k: str(v) for k, v in status_updates.items()}
            updates["updated_at"] = datetime.now().isoformat()

            # 添加时间戳，用于排序
            updates["timestamp"] = str(time.time())

            # 使用兼容函数更新Redis
            success = redis_hset_compatible(self.redis_client, f"table:job:{job_id}", updates)

            if success:
                # 发布进度消息到频道（用于SSE实时推送）
                try:
                    progress_message = {
                        "job_id": job_id,
                        "type": "progress",
                        "data": updates,
                        "timestamp": time.time(),
                        "worker_id": self.worker_id
                    }

                    message_json = json.dumps(progress_message, ensure_ascii=False)
                    self.redis_client.publish(f"table:progress:{job_id}", message_json)

                    # 调试日志
                    if updates.get("progress") and int(updates["progress"]) % 20 == 0:
                        print(f"📤 发布进度更新: {updates.get('progress')}% - {updates.get('message', '')}")

                except Exception as pub_error:
                    print(f"⚠️ Redis发布消息失败: {pub_error}")

            else:
                print(f"⚠️ Redis HSET操作失败: {job_id}")

        except Exception as e:
            print(f"⚠️ 更新任务状态失败: {e}")

    def generate_excel_directly_000(self, job_id, pdf_folder, bank_name="", aggregator=None):
        """
        直接生成Excel文件，不依赖process_table_images_real的复杂逻辑

        参数:
            job_id: 任务ID
            pdf_folder: PDF文件夹名称
            bank_name: 银行名称
            aggregator: PDF聚合器对象（如果已有）

        返回:
            tuple: (success, excel_path, error_msg)
        """
        from pathlib import Path
        from datetime import datetime

        print(f"\n{'=' * 60}")
        print(f"🆕 开始直接生成Excel文件")
        print(f"📁 PDF文件夹: {pdf_folder}")
        print(f"🏦 银行名称: {bank_name}")
        print(f"{'=' * 60}")

        try:
            # 1. 定义Excel输出路径
            # EXCEL_DATA_DIR = "data/backend/static/excel_data"
            excel_dir = os.path.join(EXCEL_DATA_DIR, pdf_folder)
            excel_path = os.path.join(excel_dir, f"{pdf_folder}_合并.xlsx")

            # 确保目录存在
            os.makedirs(excel_dir, exist_ok=True)
            print(f"📁 Excel输出路径: {excel_path}")
            print(f"📁 目录已创建: {os.path.exists(excel_dir)}")

            # 2. 获取或创建聚合器
            if aggregator is None:
                print(f"🔄 获取PDF聚合器...")
                aggregator = pdf_aggregator_manager.get_aggregator(pdf_folder, bank_name)

            print(f"🔍 聚合器状态:")
            print(f"  - 聚合器对象: {aggregator}")
            print(f"  - 表格数量: {len(aggregator)}")

            if hasattr(aggregator, 'tables'):
                print(f"  - 表格列表: {list(aggregator.tables.keys())}")

            # 3. 检查是否已有Excel文件
            existing_files = list(Path(excel_dir).glob("*.xlsx"))
            if existing_files:
                print(f"🔍 发现现有Excel文件: {existing_files[0]}")

            # 4. 如果聚合器为空，检查是否已有文件
            if len(aggregator) == 0:
                if existing_files:
                    excel_path = str(existing_files[0])
                    print(f"✅ 聚合器为空，但已有Excel文件存在: {excel_path}")
                    return True, excel_path, "使用已有Excel文件"
                else:
                    print(f"⚠️ 聚合器为空，也没有现有Excel文件")

                    # 创建空的Excel文件
                    try:
                        import pandas as pd
                        # 创建一个包含说明的工作表
                        empty_df = pd.DataFrame({
                            "说明": ["这是一个空的Excel文件，因为没有找到表格数据"],
                            "生成时间": [datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
                            "PDF文件夹": [pdf_folder],
                            "银行名称": [bank_name or "未知"]
                        })
                        empty_df.to_excel(excel_path, index=False, engine='openpyxl')
                        print(f"✅ 已创建空的Excel文件: {excel_path}")
                        return True, excel_path, "创建了空的Excel文件"
                    except Exception as e:
                        print(f"❌ 创建空Excel文件失败: {e}")
                        return False, None, f"创建空Excel文件失败: {str(e)}"

            # 5. 聚合器不为空，保存Excel
            print(f"🔄 开始保存Excel文件...")
            print(f"  - 表格数量: {len(aggregator)}")
            print(f"  - 输出路径: {excel_path}")

            try:
                # 调用聚合器的保存方法
                success = aggregator.save_to_excel(str(excel_path), metadata_list=None)

                if success:
                    print(f"✅ Excel文件保存成功: {excel_path}")

                    # 验证文件确实存在
                    if os.path.exists(excel_path):
                        file_size = os.path.getsize(excel_path)
                        print(f"✅ 文件已确认存在，大小: {file_size} 字节")
                        return True, excel_path, None
                    else:
                        print(f"❌ 文件保存成功但文件不存在: {excel_path}")
                        return False, None, "文件保存成功但文件不存在"
                else:
                    print(f"❌ 聚合器save_to_excel返回失败")
                    return False, None, "聚合器保存失败"

            except Exception as e:
                print(f"❌ 保存Excel文件异常: {e}")
                import traceback
                traceback.print_exc()
                return False, None, f"保存Excel异常: {str(e)}"

        except Exception as e:
            print(f"❌ 直接生成Excel函数异常: {e}")
            import traceback
            traceback.print_exc()
            return False, None, f"生成Excel函数异常: {str(e)}"

    def generate_excel_directly(self, job_id, pdf_folder, bank_name="", aggregator=None):
        """
        直接生成Excel文件，不依赖process_table_images_real的复杂逻辑

        参数:
            job_id: 任务ID
            pdf_folder: PDF文件夹名称
            bank_name: 银行名称
            aggregator: PDF聚合器对象（如果已有）

        返回:
            tuple: (success, excel_path, error_msg)
        """
        import os
        from pathlib import Path
        from datetime import datetime

        print(f"\n{'=' * 60}")
        print(f"🆕 开始直接生成Excel文件")
        print(f"📁 PDF文件夹: {pdf_folder}")
        print(f"🏦 银行名称: {bank_name}")
        print(f"{'=' * 60}")

        try:
            # ✅ 修复1：确保使用了正确导入的 EXCEL_DATA_DIR 常量
            # 如果 EXCEL_DATA_DIR 是绝对路径，直接使用；如果是相对路径，基于 PROJECT_ROOT_STR 构建
            try:
                # 假设 EXCEL_DATA_DIR 已经是从 backend.utils.constants 导入的常量
                excel_base_dir = Path(EXCEL_DATA_DIR)
                print(f"🔍 使用导入的 EXCEL_DATA_DIR: {excel_base_dir}")
            except NameError:
                # 如果 EXCEL_DATA_DIR 未导入，基于 PROJECT_ROOT_STR 构建
                print(f"⚠️ EXCEL_DATA_DIR 未导入，使用 PROJECT_ROOT_STR 构建路径")
                excel_base_dir = Path(PROJECT_ROOT_STR) / "data" / "backend" / "static" / "excel_data"

            # ✅ 修复2：使用 Path 对象进行路径操作，更安全
            excel_dir = excel_base_dir / pdf_folder
            excel_path = excel_dir / f"{pdf_folder}_合并.xlsx"

            # ✅ 修复3：确保输出路径的绝对性
            excel_dir = excel_dir.resolve()
            excel_path = excel_path.resolve()

            # 确保目录存在
            excel_dir.mkdir(parents=True, exist_ok=True)
            print(f"📁 Excel输出绝对路径: {excel_path}")
            print(f"📁 目录已创建: {excel_dir.exists()}")

            # 2. 获取或创建聚合器
            if aggregator is None:
                print(f"🔄 获取PDF聚合器...")
                aggregator = pdf_aggregator_manager.get_aggregator(pdf_folder, bank_name)

            print(f"🔍 聚合器状态:")
            print(f"  - 聚合器对象: {aggregator}")
            print(f"  - 表格数量: {len(aggregator)}")

            if hasattr(aggregator, 'tables'):
                print(f"  - 表格列表: {list(aggregator.tables.keys())}")

            # 3. 检查是否已有Excel文件
            existing_files = list(excel_dir.glob("*.xlsx"))
            if existing_files:
                print(f"🔍 发现现有Excel文件: {existing_files[0]}")

            # 4. 如果聚合器为空，检查是否已有文件
            if len(aggregator) == 0:
                if existing_files:
                    excel_path = str(existing_files[0])
                    print(f"✅ 聚合器为空，但已有Excel文件存在: {excel_path}")
                    return True, str(excel_path), "使用已有Excel文件"
                else:
                    print(f"⚠️ 聚合器为空，也没有现有Excel文件")

                    # 创建空的Excel文件
                    try:
                        import pandas as pd
                        # 创建一个包含说明的工作表
                        empty_df = pd.DataFrame({
                            "说明": ["这是一个空的Excel文件，因为没有找到表格数据"],
                            "生成时间": [datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
                            "PDF文件夹": [pdf_folder],
                            "银行名称": [bank_name or "未知"]
                        })
                        empty_df.to_excel(str(excel_path), index=False, engine='openpyxl')
                        print(f"✅ 已创建空的Excel文件: {excel_path}")
                        return True, str(excel_path), "创建了空的Excel文件"
                    except Exception as e:
                        print(f"❌ 创建空Excel文件失败: {e}")
                        return False, None, f"创建空Excel文件失败: {str(e)}"

            # 5. 聚合器不为空，保存Excel
            print(f"🔄 开始保存Excel文件...")
            print(f"  - 表格数量: {len(aggregator)}")
            print(f"  - 输出路径: {excel_path}")

            try:
                # 调用聚合器的保存方法
                success = aggregator.save_to_excel(str(excel_path), metadata_list=None)

                if success:
                    print(f"✅ Excel文件保存成功: {excel_path}")

                    # 验证文件确实存在
                    if excel_path.exists():
                        file_size = excel_path.stat().st_size
                        print(f"✅ 文件已确认存在，大小: {file_size} 字节")
                        return True, str(excel_path), None
                    else:
                        print(f"❌ 文件保存成功但文件不存在: {excel_path}")
                        return False, None, "文件保存成功但文件不存在"
                else:
                    print(f"❌ 聚合器save_to_excel返回失败")
                    return False, None, "聚合器保存失败"

            except Exception as e:
                print(f"❌ 保存Excel文件异常: {e}")
                import traceback
                traceback.print_exc()
                return False, None, f"保存Excel异常: {str(e)}"

        except Exception as e:
            print(f"❌ 直接生成Excel函数异常: {e}")
            import traceback
            traceback.print_exc()
            return False, None, f"生成Excel函数异常: {str(e)}"

    def _diagnose_image_lists(self, all_images, to_process_images, skipped_images):
        """诊断图片列表的详细函数"""
        print(f"\n🖼️ 图片列表详细信息:")
        print(f"  - 原始图片列表 ({len(all_images)}张):")
        for i, img_path in enumerate(all_images[:5]):
            print(f"    {i + 1}. {os.path.basename(img_path)}")
        if len(all_images) > 5:
            print(f"    ... 等 {len(all_images) - 5} 张")

        print(f"  - 需要处理图片 ({len(to_process_images)}张):")
        if to_process_images:
            for i, img_path in enumerate(to_process_images[:5]):
                print(f"    {i + 1}. {os.path.basename(img_path)}")
            if len(to_process_images) > 5:
                print(f"    ... 等 {len(to_process_images) - 5} 张")
        else:
            print(f"    ⚠️ 空列表 - 没有需要处理的新图片")

        print(f"  - 跳过图片 ({len(skipped_images)}张):")
        if skipped_images:
            for i, img_path in enumerate(skipped_images[:5]):
                print(f"    {i + 1}. {os.path.basename(img_path)}")
            if len(skipped_images) > 5:
                print(f"    ... 等 {len(skipped_images) - 5} 张")
        else:
            print(f"    ✓ 没有跳过的图片")

    def _diagnose_aggregator_state(self, stage_name, pdf_folder, bank_name, images_to_process):
        """诊断聚合器状态的详细函数"""
        print(f"\n{'=' * 60}")
        print(f"🔍🔍🔍 {stage_name}聚合器状态诊断")
        print(f"{'=' * 60}")

        try:
            # 1. 获取聚合器实例
            aggregator = pdf_aggregator_manager.get_aggregator(pdf_folder, bank_name)
            print(f"📊 聚合器实例信息:")
            print(f"  - 对象ID: {id(aggregator)}")
            print(f"  - 类型: {type(aggregator)}")
            print(f"  - PDF文件夹: {aggregator.pdf_folder}")
            print(f"  - 银行名称: {aggregator.bank_name}")

            # 2. 检查聚合器表格数量
            tables_count = len(aggregator)
            print(f"📊 表格数量统计:")
            print(f"  - 表格数量(len): {tables_count}")

            # 3. 检查聚合器内部数据结构
            if hasattr(aggregator, 'tables_data'):
                data_count = len(aggregator.tables_data) if aggregator.tables_data else 0
                print(f"  - tables_data长度: {data_count}")

                if data_count > 0 and data_count <= 3:
                    print(f"  - 表格数据形状:")
                    for i, table in enumerate(aggregator.tables_data[:3]):
                        if isinstance(table, list):
                            rows = len(table)
                            cols = len(table[0]) if rows > 0 and isinstance(table[0], list) else 0
                            print(f"    - 表格{i + 1}: {rows}行×{cols}列")

            if hasattr(aggregator, 'table_names'):
                names_count = len(aggregator.table_names) if aggregator.table_names else 0
                print(f"  - table_names长度: {names_count}")

                if names_count > 0 and names_count <= 3:
                    print(f"  - 前{min(3, names_count)}个Sheet名称:")
                    for i, name in enumerate(aggregator.table_names[:3]):
                        print(f"    - {i + 1}. '{name}'")

            if hasattr(aggregator, 'image_refs'):
                refs_count = len(aggregator.image_refs) if aggregator.image_refs else 0
                print(f"  - image_refs长度: {refs_count}")

            # 4. 检查aggregator_manager状态
            print(f"📊 PDF聚合器管理器状态:")
            print(f"  - 管理的PDF数量: {len(pdf_aggregator_manager)}")

            if hasattr(pdf_aggregator_manager, '_aggregators'):
                all_aggregators = pdf_aggregator_manager._aggregators
                print(f"  - 聚合器键列表: {list(all_aggregators.keys())}")

                if pdf_folder in all_aggregators:
                    manager_aggregator = all_aggregators[pdf_folder]
                    print(f"  - 管理器中的聚合器ID: {id(manager_aggregator)}")
                    print(f"  - 与当前聚合器相同: {id(aggregator) == id(manager_aggregator)}")

            # 5. 检查待处理图片与聚合器的关系
            print(f"📊 待处理图片与聚合器关系:")
            print(f"  - 待处理图片数量: {len(images_to_process)}")

            if hasattr(aggregator, 'image_refs') and aggregator.image_refs:
                processed_images = [ref.get('image_name', '') for ref in aggregator.image_refs]
                print(f"  - 聚合器中已处理的图片: {len(processed_images)}张")
                if processed_images and len(processed_images) <= 5:
                    print(f"  - 前{min(5, len(processed_images))}张: {processed_images[:5]}")

        except Exception as e:
            print(f"❌ 聚合器状态诊断失败: {e}")
            import traceback
            traceback.print_exc()

    def _process_images_with_validation(self, job_id, pdf_folder, images_to_process, table_type, bank_name,
                                        progress_tracker):
        """调用图片处理函数并进行验证"""
        print(f"\n{'=' * 60}")
        print(f"🔄 开始调用 process_images_with_real_time_updates")
        print(f"{'=' * 60}")

        print(f"📤 调用参数详情:")
        print(f"  - job_id: {job_id}")
        print(f"  - pdf_folder: {pdf_folder}")
        print(f"  - 图片数量: {len(images_to_process)}")
        print(f"  - 银行名称: {bank_name}")
        print(f"  - 图片列表: {[os.path.basename(p) for p in images_to_process]}")

        try:
            # 调用前的聚合器状态
            before_aggregator = pdf_aggregator_manager.get_aggregator(pdf_folder, bank_name)
            before_tables = len(before_aggregator)
            print(f"📊 调用前状态:")
            print(f"  - 表格数量: {before_tables}")

            # 调用处理函数
            result = process_images_with_real_time_updates(
                job_id=job_id,
                pdf_folder=pdf_folder,
                image_paths=images_to_process,
                table_type=table_type,
                bank_name=bank_name,
                progress_tracker=progress_tracker
            )

            # 调用后的聚合器状态
            after_aggregator = pdf_aggregator_manager.get_aggregator(pdf_folder, bank_name)
            after_tables = len(after_aggregator)
            print(f"📊 调用后状态:")
            print(f"  - 表格数量: {after_tables}")
            print(f"  - 新增表格: {after_tables - before_tables}")

            # 打印处理结果
            print(f"📊 处理结果:")
            print(f"  - 成功: {result.get('success', False)}")
            print(f"  - 错误: {result.get('error', '无')}")
            print(f"  - 处理时间: {result.get('processing_time', 0)}")

            if result.get('success'):
                # 检查聚合器中的表格数据
                if hasattr(after_aggregator, 'tables_data'):
                    tables_count = len(after_aggregator.tables_data) if after_aggregator.tables_data else 0
                    print(f"  📊 聚合器中的表格数据:")
                    print(f"    - 表格数据数量: {tables_count}")

                    if tables_count > 0:
                        # 打印前3个表格的详细信息
                        for i, table in enumerate(after_aggregator.tables_data[:3]):
                            if isinstance(table, list):
                                rows = len(table)
                                cols = len(table[0]) if rows > 0 and isinstance(table[0], list) else 0
                                print(f"    - 表格{i + 1}: {rows}行×{cols}列")

                        # 检查表格数据质量
                        if tables_count > 0:
                            sample_table = after_aggregator.tables_data[0]
                            if isinstance(sample_table, list) and len(sample_table) > 0:
                                print(
                                    f"    - 第一个表格的第一行: {sample_table[0][:5] if len(sample_table[0]) > 5 else sample_table[0]}")

                # 检查table_names
                if hasattr(after_aggregator, 'table_names'):
                    names_count = len(after_aggregator.table_names) if after_aggregator.table_names else 0
                    print(f"    - Sheet名称数量: {names_count}")
                    if names_count > 0:
                        print(f"    - 前{min(3, names_count)}个Sheet: {after_aggregator.table_names[:3]}")

            return result

        except Exception as e:
            print(f"❌ 调用 process_images_with_real_time_updates 失败: {e}")
            import traceback
            traceback.print_exc()

            return {
                "success": False,
                "error": str(e),
                "processing_time": 0
            }

    def _final_check_before_excel_generation(self, pdf_folder, bank_name, images_to_process):
        """Excel生成前的最终检查"""
        print(f"\n{'=' * 60}")
        print(f"🔍 调用 process_table_images_real 前的状态检查")
        print(f"{'=' * 60}")

        try:
            # 获取最终的聚合器
            final_aggregator = pdf_aggregator_manager.get_aggregator(pdf_folder, bank_name)
            final_tables_count = len(final_aggregator)

            print(f"📊 最终聚合器状态:")
            print(f"  - 表格数量: {final_tables_count}")
            print(f"  - 聚合器ID: {id(final_aggregator)}")

            if final_tables_count == 0:
                print(f"⚠️ 警告: 聚合器为空，没有表格数据")
                print(f"🔍 可能的原因:")
                print(f"  1. images_to_process 为空: {len(images_to_process) == 0}")
                print(f"  2. process_images_with_real_time_updates 没有添加数据到聚合器")
                print(f"  3. 聚合器实例在不同调用中不一致")

                # 检查聚合器内部数据
                if hasattr(final_aggregator, 'tables_data'):
                    data = final_aggregator.tables_data
                    print(f"  4. tables_data 类型: {type(data)}")
                    if data is not None:
                        print(f"  5. tables_data 长度: {len(data)}")
                    else:
                        print(f"  5. tables_data 为 None")

                # 检查PDF聚合器管理器
                if hasattr(pdf_aggregator_manager, '_aggregators'):
                    all_aggregators = pdf_aggregator_manager._aggregators
                    if pdf_folder in all_aggregators:
                        mgr_aggregator = all_aggregators[pdf_folder]
                        print(f"  6. 管理器中的聚合器ID: {id(mgr_aggregator)}")
                        print(f"  7. 与当前聚合器相同: {id(final_aggregator) == id(mgr_aggregator)}")

            else:
                print(f"✅ 聚合器中有 {final_tables_count} 个表格，可以生成Excel")

                # 输出表格详细信息
                if hasattr(final_aggregator, 'tables_data') and final_aggregator.tables_data:
                    print(f"📊 表格详细信息:")
                    for i, table in enumerate(final_aggregator.tables_data[:3]):
                        if isinstance(table, list):
                            rows = len(table)
                            cols = len(table[0]) if rows > 0 and isinstance(table[0], list) else 0
                            print(f"  - 表格{i + 1}: {rows}行×{cols}列")

        except Exception as e:
            print(f"❌ 最终检查失败: {e}")


    def _fallback_generate_excel(self, job_id, pdf_folder, bank_name):
        """备选方案：直接生成Excel文件"""
        print(f"\n{'=' * 60}")
        print(f"🆕 尝试备选方案：直接生成Excel")
        print(f"{'=' * 60}")

        try:
            success, excel_path, error_msg = self.generate_excel_directly(
                job_id, pdf_folder, bank_name, aggregator=None
            )

            if success:
                print(f"✅ 备选方案成功: {excel_path}")

                # 更新任务状态
                self.update_job_status(job_id, {
                    "status": "completed",
                    "progress": "100",
                    "message": f"通过备选方案生成Excel: {excel_path}",
                    "excel_path": excel_path,
                    "fallback_used": "true"
                })

                return True
            else:
                print(f"❌ 备选方案失败: {error_msg}")

                # 更新任务状态
                self.update_job_status(job_id, {
                    "status": "failed",
                    "error": f"备选方案失败: {error_msg}",
                    "fallback_failed": "true"
                })

                return False

        except Exception as e:
            print(f"❌ 备选方案异常: {e}")
            return False

    def _process_images_with_validation(self, job_id, pdf_folder, images_to_process, table_type, bank_name,
                                        progress_tracker):
        """调用重构后的图片处理函数"""

        print(f"\n{'=' * 60}")
        print(f"🔄 调用重构版图片处理函数")
        print(f"{'=' * 60}")

        # 调用重构后的函数
        result = process_images_with_real_time_updates(
            job_id=job_id,
            pdf_folder=pdf_folder,
            image_paths=images_to_process,
            table_type=table_type,
            bank_name=bank_name,
            progress_tracker=progress_tracker
        )

        if result.get('excel_path'):
            print(f"  - Excel路径: {result.get('excel_path')}")

        return result

    def _safe_generate_excel(self, job_id, pdf_folder, image_paths, table_type, bank_name, progress_tracker,
                             skipped_images):
        """安全的Excel生成函数 - 修正参数传递问题"""

        # 获取原始的所有图片（从任务数据中获取）
        all_image_paths = []  # 这里应该从任务数据获取
        # 但由于我们不知道原始的任务数据，需要重构这个函数

        # 临时方案：如果image_paths为空，尝试从其他地方获取
        if not image_paths:
            print(f"⚠️ 警告：image_paths为空列表")
            print(f"🔍 尝试从增量处理器获取已处理的图片列表")

            try:
                # 尝试从增量处理器获取已处理的图片
                # from backend.src.incremental_processor import incremental_processor

                # 获取已处理的图片名称
                if hasattr(incremental_processor, 'records'):
                    if pdf_folder in incremental_processor.records:
                        processed_images = incremental_processor.records[pdf_folder]
                        print(f"🔍 从增量处理器获取到 {len(processed_images)} 个已处理的图片名称")

                        # 尝试构建图片路径
                        for img_name in processed_images:
                            # 这里需要知道图片的实际存储路径
                            # 假设图片在filtered_tables_dir/pdf_folder/tables/目录下
                            possible_paths = [
                                f"data/backend/static/filtered_tables/{pdf_folder}/tables/{img_name}",
                                f"data/filtered_tables/{pdf_folder}/tables/{img_name}",
                                f"filtered_tables/{pdf_folder}/tables/{img_name}",
                            ]

                            for possible_path in possible_paths:
                                if os.path.exists(possible_path):
                                    all_image_paths.append(possible_path)
                                    break

                        if all_image_paths:
                            print(f"✅ 构建了 {len(all_image_paths)} 个图片路径")
                            image_paths = all_image_paths
            except Exception as e:
                print(f"⚠️ 获取图片路径失败: {e}")

        # ✅ 关键修复2：在调用处理函数前验证参数
        print(f"🔍 调用参数验证:")
        print(f"  - pdf_folder: {pdf_folder}")
        print(f"  - 图片列表长度: {len(image_paths)}")
        print(f"  - 图片示例: {[os.path.basename(p) for p in image_paths[:3]] if image_paths else '空列表'}")

        # 1. 先调用图片处理函数
        processing_result = self._process_images_with_validation(
            job_id=job_id,
            pdf_folder=pdf_folder,
            images_to_process=image_paths,  # ✅ 使用修正后的图片列表
            table_type=table_type,
            bank_name=bank_name,
            progress_tracker=progress_tracker
        )

        # 2. 根据返回结果决定是否生成Excel
        need_generate_excel = processing_result.get("need_generate_excel", False)
        excel_exists = processing_result.get("excel_exists", False)
        existing_excel_path = processing_result.get("excel_path")

        print(f"\n📊 决策分析:")
        print(f"  - 需要生成Excel: {need_generate_excel}")
        print(f"  - Excel文件存在: {excel_exists}")
        print(f"  - 现有Excel路径: {existing_excel_path}")

        if not need_generate_excel and excel_exists and existing_excel_path:
            # 情况1：有现有Excel文件，无需生成
            print(f"✅ 决策: 使用现有Excel文件: {existing_excel_path}")

            # 更新任务状态
            self.update_job_status(job_id, {
                "status": "completed",
                "progress": 100,
                "message": f"使用现有Excel文件: {existing_excel_path}",
                "excel_path": existing_excel_path,
                "existing_excel_used": "true"
            })

            return True
        elif need_generate_excel:
            # 情况2：需要生成Excel
            print(f"\n{'=' * 60}")
            print(f"🔄 调用 process_table_images_real 生成Excel")
            print(f"{'=' * 60}")

            print(f"📤 调用参数:")
            print(f"  - job_id: {job_id}")
            print(f"  - pdf_folder: {pdf_folder}")
            print(f"  - image_paths: {len(image_paths)}张图片")
            print(f"  - skipped_images: {len(skipped_images)}张图片")

            try:
                # 调用原有的Excel生成函数
                process_table_images_real(
                    job_id=job_id,
                    pdf_folder=pdf_folder,
                    image_paths=image_paths,
                    table_type=table_type,
                    bank_name=bank_name,
                    progress_tracker=progress_tracker,
                    skipped_images=skipped_images,
                    existing_sheets=None
                )

                print(f"✅ process_table_images_real 调用完成")
                return True

            except Exception as e:
                print(f"❌ 调用 process_table_images_real 失败: {e}")
                import traceback
                traceback.print_exc()

                # 尝试备选方案
                return self._fallback_generate_excel(job_id, pdf_folder, bank_name)
        else:
            # 情况3：既不需要生成，也没有现有文件
            print(f"⚠️ 决策: 既不需要生成Excel，也没有现有Excel文件")

            # 尝试创建空的Excel文件
            return self._fallback_generate_excel(job_id, pdf_folder, bank_name)

    def _verify_excel_file_generated(self, pdf_folder):
        """使用EXCEL_DATA_DIR验证Excel文件是否生成"""
        print(f"\n{'=' * 60}")
        print(f"🔍 使用EXCEL_DATA_DIR验证Excel文件")
        print(f"{'=' * 60}")

        try:
            from pathlib import Path

            # ✅ 使用导入的EXCEL_DATA_DIR常量
            excel_dir = Path(EXCEL_DATA_DIR) / pdf_folder

            print(f"📁 检查目录: {excel_dir}")
            print(f"📁 绝对路径: {excel_dir.absolute()}")
            print(f"📁 目录存在: {excel_dir.exists()}")

            if not excel_dir.exists():
                print(f"❌ EXCEL_DATA_DIR目录不存在: {excel_dir}")
                return

            # 列出目录内容
            files = list(excel_dir.iterdir())
            print(f"📁 目录内容 ({len(files)}个):")

            for item in files[:10]:  # 只显示前10个
                if item.is_file():
                    size = item.stat().st_size
                    mtime = datetime.fromtimestamp(item.stat().st_mtime)
                    print(f"  📄 {item.name} ({size} 字节, 修改于: {mtime})")
                else:
                    print(f"  📁 {item.name}/")

            if len(files) > 10:
                print(f"  ... 等 {len(files) - 10} 个")

            # 检查Excel文件
            excel_files = list(excel_dir.glob("*.xlsx"))
            print(f"\n📊 Excel文件统计:")
            print(f"  - 找到 {len(excel_files)} 个Excel文件")

            if excel_files:
                for excel_file in excel_files:
                    file_path = str(excel_file.absolute())
                    size = excel_file.stat().st_size
                    mtime = datetime.fromtimestamp(excel_file.stat().st_mtime)

                    print(f"  ✅ Excel文件: {excel_file.name}")
                    print(f"    - 绝对路径: {file_path}")
                    print(f"    - 大小: {size} 字节")
                    print(f"    - 修改时间: {mtime}")

                    # 验证文件可读性
                    try:
                        import openpyxl
                        wb = openpyxl.load_workbook(file_path, read_only=True)
                        sheet_count = len(wb.sheetnames)
                        print(f"    - Sheet数量: {sheet_count}")
                        print(f"    - Sheet名称: {wb.sheetnames[:3] if sheet_count > 3 else wb.sheetnames}")
                        wb.close()
                    except Exception as e:
                        print(f"    ⚠️ 无法读取Excel: {e}")
            else:
                print(f"  ⚠️ 没有找到Excel文件 (.xlsx)")

        except Exception as e:
            print(f"❌ Excel文件验证失败: {e}")
            import traceback
            traceback.print_exc()

    def _update_job_status_original(self, job_id: str, status_data: Dict[str, Any]):
        """原始的Redis更新方法 - 修复布尔值问题，包含所有字段"""
        from datetime import datetime
        import json

        redis_key = f"table:job:{job_id}"

        # ✅ 修复1：转换所有值为字符串
        processed_data = {}
        for key, value in status_data.items():
            if value is None:
                processed_data[key] = ""
            elif isinstance(value, bool):
                processed_data[key] = str(value).lower()  # True -> "true", False -> "false"
            elif isinstance(value, (int, float)):
                processed_data[key] = str(value)  # 数字也转为字符串
            elif isinstance(value, dict) or isinstance(value, list):
                processed_data[key] = json.dumps(value, ensure_ascii=False)  # 复杂对象转为JSON
            else:
                processed_data[key] = str(value)

        # 设置默认字段
        processed_data.setdefault("last_updated", datetime.now().isoformat())

        # ✅ 确保有原始文件名
        if "original_filename" not in processed_data:
            pdf_folder = processed_data.get("pdf_folder", "")
            if pdf_folder:
                processed_data["original_filename"] = f"{pdf_folder}.pdf"
            else:
                processed_data["original_filename"] = job_id

        # 如果状态是processing，添加开始时间
        if processed_data.get("status") == "processing" and "started_at" not in processed_data:
            processed_data["started_at"] = datetime.now().isoformat()

        # 如果状态是completed或failed，添加结束时间
        if processed_data.get("status") in ["completed", "failed"] and "completed_at" not in processed_data:
            processed_data["completed_at"] = datetime.now().isoformat()

        print(f"🔍 写入Redis的数据:")
        print(f"  - 字段数量: {len(processed_data)}")
        print(f"  - 原始文件名: {processed_data.get('original_filename')}")
        print(f"  - 开始时间: {processed_data.get('started_at')}")
        print(f"  - PDF文件夹: {processed_data.get('pdf_folder')}")

        # 更新到Redis
        self.redis_client.hmset(redis_key, processed_data)
        self.redis_client.expire(redis_key, 604800)  # 7天过期

        # ✅ 修复2：SSE推送包含所有关键字段
        sse_message = {
            "job_id": job_id,
            "status": processed_data.get("status", ""),
            "progress": processed_data.get("progress", 0),
            "message": processed_data.get("message", ""),
            "original_filename": processed_data.get("original_filename", ""),
            "started_at": processed_data.get("started_at", ""),
            "pdf_folder": processed_data.get("pdf_folder", ""),
            "skipped_images": processed_data.get("skipped_images", 0),
            "processed_images": processed_data.get("processed_images", 0),
            "total_images": processed_data.get("total_images", 0),
            "timestamp": datetime.now().isoformat()
        }

        # 发布SSE消息
        message_json = json.dumps(sse_message, ensure_ascii=False)
        self.redis_client.publish(f"table:progress:{job_id}", message_json)

        print(f"📤 SSE推送消息: {json.dumps(sse_message, ensure_ascii=False, indent=2)}")

        # ✅ 修复3：同时更新按PDF名称索引的状态
        pdf_folder = processed_data.get("pdf_folder")
        if not pdf_folder and "pdf_folder" in status_data:
            pdf_folder = status_data.get("pdf_folder")

        if pdf_folder:
            pdf_key = f"pdf:{pdf_folder}:current_status"
            pdf_data = {
                "job_id": job_id,
                "status": processed_data.get("status", "unknown"),
                "progress": processed_data.get("progress", 0),
                "processed_images": processed_data.get("processed_images", 0),
                "total_images": processed_data.get("total_images", 0),
                "message": processed_data.get("message", ""),
                "original_filename": processed_data.get("original_filename", ""),
                "started_at": processed_data.get("started_at", ""),
                "last_updated": datetime.now().isoformat()
            }
            self.redis_client.hmset(pdf_key, pdf_data)
            self.redis_client.expire(pdf_key, 604800)
            print(f"  📁 同时更新PDF状态: {pdf_folder}")

    def _get_database_path(self):
        """获取数据库路径"""
        # 这里需要根据您的项目结构调整
        try:
            from backend.utils.constants import DATABASE_PATH
            return DATABASE_PATH
        except ImportError:
            # 默认路径
            import os
            from pathlib import Path
            project_root = Path(__file__).parent.parent.parent.parent
            return os.path.join(project_root, "data", "database.db")

    def update_table_job(self, job_id: str, updates: Dict[str, Any]):
        """
        更新表格任务状态 - 与 table_processor.py 的进度更新兼容
        此方法会被 process_images_with_real_time_updates 调用

        参数:
            job_id: 任务ID
            updates: 更新数据字典，包含 status, progress, message, processed_images, total_images, current_image 等字段
        """
        try:
            # 提取关键字段
            status = updates.get("status", "processing")
            progress = updates.get("progress", 0)
            message = updates.get("message", "")
            processed_images = updates.get("processed_images", 0)
            total_images = updates.get("total_images", 0)
            current_image = updates.get("current_image", "")

            # 计算进度百分比
            progress_percent = 0
            if total_images > 0 and processed_images >= 0:
                progress_percent = int((processed_images / total_images) * 100)
                progress_percent = min(100, max(0, progress_percent))  # 限制在0-100之间

            # 使用现有的 update_job_status 方法
            self.update_job_status(job_id, {
                "status": status,
                "progress": str(progress_percent),
                "progress_percentage": progress_percent,
                "message": message,
                "processed_images": str(processed_images),
                "total_images": str(total_images),
                "skipped_images": updates.get("skipped_images", "0"),
                "current_image": current_image,
                "current_image_index": str(processed_images),
                "current_image_name": current_image
            })

            print(f"✅ update_table_job 完成: {job_id} - {progress_percent}%")

        except Exception as e:
            print(f"❌ update_table_job 异常: {e}")
            import traceback
            traceback.print_exc()

    def update_job_status(self, job_id: str, status_data: Dict[str, Any]):
        """更新任务状态到Redis - 修复时间字段问题，防止前端Invalid time value错误"""

        try:
            # ✅ 1. 获取当前有效时间戳
            current_iso_time = datetime.now().isoformat()
            current_timestamp = str(time.time())

            # ✅ 2. 确保关键字段存在
            if status_data.get("status") in ["completed", "success"]:
                # 确保进度是100%
                status_data["progress"] = "100"
                status_data["progress_percentage"] = 100

                # 确保有图片数量信息
                if "processed_images" not in status_data:
                    status_data["processed_images"] = "0"
                if "skipped_images" not in status_data:
                    status_data["skipped_images"] = "0"
                if "total_images" not in status_data:
                    # 计算总数
                    processed = int(status_data.get("processed_images", "0") or 0)
                    skipped = int(status_data.get("skipped_images", "0") or 0)
                    status_data["total_images"] = str(processed + skipped)

            processed_data = {}

            # 自动填充缺失的关键字段
            field_defaults = {
                "status": "unknown",
                "progress": "0",
                "progress_percentage": 0,
                "message": "",
                "original_filename": "",
                "pdf_folder": "",
                "started_at": "",  # 先设为空，后面处理
                "completed_at": "",  # 先设为空，后面处理
                "worker_id": self.worker_id,
                "total_images": "0",
                "processed_images": "0",
                "skipped_images": "0",
                "current_image": "",
                "current_image_index": "0",
                "current_image_name": "",
                "last_updated": current_iso_time,  # 使用当前有效时间
                "timestamp": current_timestamp
            }

            # ✅ 3. 合并默认值和传入数据
            for field, default_value in field_defaults.items():
                if field in status_data and status_data[field] is not None:
                    processed_data[field] = status_data[field]
                else:
                    processed_data[field] = default_value

            # ✅ 4. 关键修复：确保时间字段是有效的ISO格式字符串
            def ensure_valid_time(time_str, default_time=current_iso_time):
                """确保时间字符串有效"""
                if not time_str or time_str in ["", "None", "null", "undefined", "Invalid Date"]:
                    return default_time

                # 尝试解析时间字符串
                try:
                    # 如果是时间戳字符串
                    if time_str.replace('.', '', 1).isdigit():
                        timestamp = float(time_str)
                        if timestamp > 0:
                            return datetime.fromtimestamp(timestamp).isoformat()

                    # 如果是ISO格式字符串
                    datetime.fromisoformat(time_str.replace('Z', '+00:00').replace('z', '+00:00'))
                    return time_str
                except (ValueError, TypeError, OverflowError):
                    return default_time

            # ✅ 5. 修复时间字段
            time_fields_to_fix = [
                ("started_at", processed_data["status"] in ["queued", "processing", "starting", "pending"]),
                ("completed_at", processed_data["status"] in ["completed", "failed", "cancelled"]),
                ("task_start_time", "task_start_time" in processed_data),
                ("last_updated", True)  # 总是更新
            ]

            for field_name, should_fix in time_fields_to_fix:
                if should_fix and field_name in processed_data:
                    if not processed_data[field_name] or processed_data[field_name] in ["", "None", "null"]:
                        processed_data[field_name] = current_iso_time
                    else:
                        processed_data[field_name] = ensure_valid_time(processed_data[field_name])

            # ✅ 6. 特殊处理：原始文件名推断
            if not processed_data["original_filename"] and processed_data["pdf_folder"]:
                # 从pdf_folder推断原始文件名
                pdf_folder = processed_data["pdf_folder"]
                if pdf_folder.endswith(".pdf"):
                    processed_data["original_filename"] = pdf_folder
                else:
                    processed_data["original_filename"] = f"{pdf_folder}.pdf"

            # ✅ 7. 特殊处理：开始时间
            if processed_data["status"] in ["processing", "starting"] and not processed_data["started_at"]:
                processed_data["started_at"] = current_iso_time

            # ✅ 8. 特殊处理：结束时间
            if processed_data["status"] in ["completed", "failed"] and not processed_data["completed_at"]:
                processed_data["completed_at"] = current_iso_time

            # ✅ 9. 修复进度计算百分比
            if ("processed_images" in processed_data and "total_images" in processed_data and
                    processed_data["total_images"] and int(processed_data["total_images"]) > 0):

                try:
                    processed = int(processed_data["processed_images"])
                    skipped = int(processed_data["skipped_images"])
                    total = int(processed_data["total_images"])

                    if total > 0:
                        # ✅ 关键修复：已处理总数 = 新处理 + 跳过
                        total_processed = processed + skipped
                        progress_percent = int((total_processed / total) * 100)

                        # 确保进度在0-100之间
                        progress_percent = max(0, min(100, progress_percent))
                        processed_data["progress"] = str(progress_percent)
                        processed_data["progress_percentage"] = progress_percent

                        # ✅ 关键修复2：任务完成时进度强制为100%
                        if processed_data["status"] in ["completed", "success"]:
                            processed_data["progress"] = "100"
                            processed_data["progress_percentage"] = 100
                            # 更新消息
                            processed_data[
                                "message"] = f"任务完成。处理 {processed} 张新图片，跳过 {skipped} 张已处理图片"

                        # 如果提供了message但没有进度信息，更新message
                        if "正在处理" in processed_data.get("message", ""):
                            current_image = processed_data.get("current_image", "")
                            processed_data["message"] = f"正在处理第 {processed}/{total} 张图片: {current_image}"
                except (ValueError, ZeroDivisionError) as e:
                    print(f"⚠️ 进度计算错误: {e}")
                    processed_data["progress"] = "0"
                    processed_data["progress_percentage"] = 0

            # ✅ 10. 转换数据类型为字符串
            redis_data = {}
            for key, value in processed_data.items():
                if value is None:
                    redis_data[key] = ""
                elif isinstance(value, bool):
                    redis_data[key] = str(value).lower()  # True -> "true"
                elif isinstance(value, (int, float)):
                    # 确保数字不会太大导致溢出
                    try:
                        redis_data[key] = str(value)
                    except Exception:
                        redis_data[key] = "0"
                elif isinstance(value, (dict, list)):
                    try:
                        redis_data[key] = json.dumps(value, ensure_ascii=False)
                    except Exception:
                        redis_data[key] = "{}"
                else:
                    redis_data[key] = str(value)

            # ✅ 11. 额外安全处理：确保时间字段不为空
            time_fields = ["started_at", "completed_at", "last_updated", "task_start_time", "created_at"]
            for field in time_fields:
                if field in redis_data and (not redis_data[field] or redis_data[field] in ["", "None", "null"]):
                    redis_data[field] = "1970-01-01T00:00:00.000Z"

            # ✅ 13. 更新到Redis
            redis_key = f"table:job:{job_id}"

            try:
                # 使用兼容函数更新Redis
                from backend.utils.redis_util import redis_hset_compatible
                success = redis_hset_compatible(self.redis_client, redis_key, redis_data)

                if success:
                    # 设置过期时间（7天）
                    self.redis_client.expire(redis_key, 7 * 24 * 60 * 60)
                else:
                    print(f"⚠️ Redis HSET操作失败")

            except Exception as redis_error:
                print(f"❌ Redis更新异常: {redis_error}")
                import traceback
                traceback.print_exc()

            # ✅ 在SSE消息部分之前添加状态转换函数
            def get_frontend_status(backend_status):
                """将后端状态转换为前端支持的status值"""
                status_map = {
                    "queued": "",  # 排队中 -> 空状态
                    "processing": "",  # 处理中 -> 空状态
                    "generating_excel": "",  # 生成Excel -> 空状态
                    "completed": "success",  # 已完成 -> 成功
                    "success": "success",  # 成功 -> 成功
                    "failed": "exception",  # 失败 -> 异常
                    "exception": "exception"  # 异常 -> 异常
                }
                return status_map.get(backend_status, "")

            def get_safe_progress(progress_str):
                """确保进度在0-100范围内"""
                try:
                    progress = int(progress_str) if progress_str else 0
                    return max(0, min(100, progress))  # 限制在0-100
                except (ValueError, TypeError):
                    return 0


            # ✅ 14. 发布SSE消息 - 修复：确保立即发布
            try:
                # 确保SSE消息中的时间字段有效
                # 获取并转换状态
                backend_status = redis_data.get("status", "")
                frontend_status = get_frontend_status(backend_status)
                safe_progress = get_safe_progress(redis_data.get("progress", "0"))

                # 确保SSE消息中的时间字段有效
                sse_data = {
                    "job_id": job_id,
                    "status": frontend_status,  # ✅ 使用转换后的状态
                    "progress": str(safe_progress),
                    "progress_percentage": safe_progress,
                    "message": redis_data.get("message", ""),
                    "original_filename": redis_data.get("original_filename", ""),
                    "pdf_folder": redis_data.get("pdf_folder", ""),
                    "started_at": redis_data.get("started_at", datetime.now().isoformat()),
                    "processed_images": int(redis_data.get("processed_images", "0") or 0),
                    "skipped_images": int(redis_data.get("skipped_images", "0") or 0),
                    "total_images": int(redis_data.get("total_images", "0") or 0),
                    "current_image": redis_data.get("current_image", ""),
                    "current_image_index": int(redis_data.get("current_image_index", "0") or 0),
                    "current_image_name": redis_data.get("current_image_name", ""),
                    "worker_id": redis_data.get("worker_id", ""),
                    "timestamp": current_iso_time
                }

                # ✅ 关键修复：任务完成时强制设置进度为100%
                if backend_status in ["completed", "success"]:
                    sse_data["progress"] = "100"
                    sse_data["progress_percentage"] = 100
                    sse_data["status"] = "success"  # ✅ 确保前端显示为success

                    # 计算已处理总数用于显示
                    processed = sse_data["processed_images"]
                    skipped = sse_data["skipped_images"]
                    total = sse_data["total_images"] if sse_data["total_images"] > 0 else (processed + skipped)

                    # 更新消息
                    sse_data["message"] = f"任务完成。处理 {processed} 张新图片，跳过 {skipped} 张已处理图片"

                sse_message = {
                    "job_id": job_id,
                    "type": "progress_update",
                    "event_type": "table_progress",
                    "data": sse_data
                }

                # ✅ 关键修复1：立即发布到Redis频道，不要延迟
                message_json = json.dumps(sse_message, ensure_ascii=False)

                # ✅ 关键修复2：使用publish命令，这是Redis的发布-订阅机制
                published_count = self.redis_client.publish(f"table:progress:{job_id}", message_json)

                # ✅ 关键修复3：强制刷新缓冲区，确保消息立即发送
                try:
                    # 使用ping保持连接活跃
                    self.redis_client.ping()
                except:
                    pass

                # 调试信息
                print(
                    f"📤 SSE推送[实时]: 进度={sse_data['progress']}%, 状态={sse_data['status']}, 消息={sse_data['message']}")
                print(f"📤 频道: table:progress:{job_id}, 订阅者: {published_count}")

                # ✅ 关键修复4：如果状态是processing，额外发送一个心跳消息
                if sse_data["status"] == "processing":
                    heartbeat_msg = {
                        "job_id": job_id,
                        "type": "heartbeat",
                        "event_type": "table_progress",
                        "data": {"status": "processing", "heartbeat": datetime.now().isoformat()}
                    }
                    self.redis_client.publish(f"table:progress:{job_id}", json.dumps(heartbeat_msg, ensure_ascii=False))
                    print(f"💓 发送心跳消息")

            except Exception as sse_error:
                print(f"⚠️ SSE推送失败: {sse_error}")
                import traceback
                traceback.print_exc()

            # ✅ 15. 同时更新PDF级别的状态
            pdf_folder = redis_data.get("pdf_folder")
            if pdf_folder:
                try:
                    pdf_key = f"pdf:{pdf_folder}:current_status"
                    pdf_data = {
                        "job_id": job_id,
                        "status": redis_data.get("status", "unknown"),
                        "progress": redis_data.get("progress", "0"),
                        "processed_images": redis_data.get("processed_images", "0"),
                        "total_images": redis_data.get("total_images", "0"),
                        "message": redis_data.get("message", ""),
                        "original_filename": redis_data.get("original_filename", ""),
                        "started_at": redis_data.get("started_at", datetime.now().isoformat()),
                        "current_image": redis_data.get("current_image", ""),
                        "last_updated": current_iso_time
                    }

                    self.redis_client.hmset(pdf_key, pdf_data)
                    self.redis_client.expire(pdf_key, 7 * 24 * 60 * 60)

                    print(f"  📁 同时更新PDF状态: {pdf_folder}")

                except Exception as pdf_error:
                    print(f"⚠️ 更新PDF状态失败: {pdf_error}")

            print(f"✅ 状态更新完成")

        except Exception as e:
            print(f"❌ update_job_status 异常: {e}")
            import traceback
            traceback.print_exc()

    def check_and_update_progress(self, job_id: str, check_interval: float = 0.5):
        """
        定时检查并更新进度状态
        这个方法可以在单独的线程中运行
        """

        while self.running:
            try:
                # 从Redis获取当前状态
                redis_key = f"table:job:{job_id}"
                current_data = self.redis_client.hgetall(redis_key)

                if current_data:
                    # 解析数据
                    current_data = {k.decode('utf-8'): v.decode('utf-8') for k, v in current_data.items()}

                    status = current_data.get("status", "")
                    processed = int(current_data.get("processed_images", "0") or 0)
                    total = int(current_data.get("total_images", "0") or 0)
                    message = current_data.get("message", "")

                    # 如果还在处理中，发送心跳消息
                    if status == "processing" and total > 0:
                        progress = int((processed / total) * 100) if total > 0 else 0

                        # 发送心跳消息
                        heartbeat_msg = {
                            "job_id": job_id,
                            "type": "progress_update",
                            "event_type": "table_progress",
                            "data": {
                                "status": "processing",
                                "progress": str(progress),
                                "message": f"心跳检测: {message}",
                                "processed_images": processed,
                                "total_images": total,
                                "timestamp": datetime.now().isoformat()
                            }
                        }

                        self.redis_client.publish(f"table:progress:{job_id}",
                                                  json.dumps(heartbeat_msg, ensure_ascii=False))


                    # 如果完成，退出循环
                    elif status in ["completed", "failed", "success"]:
                        print(f"✅ 任务 {job_id} 完成，停止进度检查")
                        break

            except Exception as e:
                print(f"⚠️ 进度检查异常: {e}")

            # 短暂等待后继续检查
            time.sleep(check_interval)

    def process_single_task(self, task_data: Dict[str, Any]) -> bool:
        """处理单个表格处理任务 - 增强版：支持图片级实时进度更新"""

        # 提取任务参数
        job_id = task_data.get("job_id", "")
        pdf_folder = task_data.get("pdf_folder", "")
        image_paths = task_data.get("image_paths", [])
        png_names = task_data.get("png_names", [])
        table_type = task_data.get("table_type", "financial")
        bank_name = task_data.get("bank_name", "")
        filtered_tables_dir = task_data.get("filtered_tables_dir", "")
        use_ocr = task_data.get("use_ocr", True)

        if not job_id or not pdf_folder:
            print(f"❌ 任务数据不完整: job_id={job_id}, pdf_folder={pdf_folder}")
            self.update_job_status(job_id, {
                "status": "failed",
                "progress": "0",
                "message": f"任务数据不完整: job_id={job_id}, pdf_folder={pdf_folder}"
            })
            return False

        self.current_job = job_id
        task_start_time = time.time()

        try:
            # ✅ 1. 获取原始文件名
            print(f"🔍 查询原始文件名: {pdf_folder}")
            original_filename = pdf_folder
            db_connection_success = False

            try:
                from backend.models.unified_db import UnifiedDatabaseManager
                db_manager = UnifiedDatabaseManager()
                if hasattr(db_manager, 'db_path'):
                    db_path = self._get_database_path()
                    print(f"✅ 数据库路径: {db_path}")

                    if os.path.exists(db_path):
                        conn = sqlite3.connect(db_path)
                        conn.row_factory = sqlite3.Row
                        cursor = conn.cursor()

                        # 查询原始文件名
                        cursor.execute("SELECT raw_filename FROM files WHERE filename = ?", (pdf_folder + ".pdf",))
                        result = cursor.fetchone()

                        if result and result['raw_filename']:
                            original_filename = result['raw_filename']
                            print(f"✅ 查询到原始文件名: {original_filename}")
                            db_connection_success = True
                        else:
                            # 模糊匹配
                            cursor.execute("SELECT filename, raw_filename FROM files WHERE filename LIKE ?",
                                           (f"%{pdf_folder}%",))
                            all_results = cursor.fetchall()
                            if all_results:
                                original_filename = all_results[0]['raw_filename']
                                db_connection_success = True
                                print(f"✅ 模糊匹配到原始文件名: {original_filename}")
                            else:
                                print(f"⚠️ 数据库中未找到记录")
                                original_filename = pdf_folder

                        conn.close()
                    else:
                        print(f"❌ 数据库文件不存在: {db_path}")
                else:
                    print(f"⚠️ DatabaseManager没有db_path属性")
            except Exception as e:
                print(f"❌ 查询数据库异常: {e}")

            # ✅ 2. 增量处理：过滤已处理的图片
            print(f"\n🔍 增量处理检查...")
            image_names = [os.path.basename(img_path) for img_path in image_paths]

            # 导入增量处理器
            try:
                from backend.src.incremental_processor import incremental_processor
                images_to_process_names = incremental_processor.filter_processed_images(pdf_folder, image_names)
                skipped_images_names = [img for img in image_names if img not in images_to_process_names]

                # 筛选出需要处理的图片路径
                images_to_process = [
                    img_path for img_path in image_paths
                    if os.path.basename(img_path) in images_to_process_names
                ]

                total_new_images = len(images_to_process)
                total_skipped_images = len(skipped_images_names)
                total_processed = total_new_images + total_skipped_images

                print(f"📊 增量处理结果:")
                print(f"  - 总图片: {len(image_paths)}")
                print(f"  - 新图片: {total_new_images}")
                print(f"  - 跳过图片: {total_skipped_images}")

                if skipped_images_names:
                    print(f"⏭️ 跳过的图片 (前5张):")
                    for i, img_name in enumerate(skipped_images_names[:5]):
                        print(f"    {i + 1}. {img_name}")
                    if len(skipped_images_names) > 5:
                        print(f"    ... 等 {len(skipped_images_names) - 5} 张")

            except Exception as e:
                print(f"⚠️ 增量处理器导入失败，处理所有图片: {e}")
                images_to_process = image_paths
                total_new_images = len(image_paths)
                total_skipped_images = 0
                total_processed = len(image_paths)
                skipped_images_names = []

            # ✅ 3. 初始化任务状态 - 包含完整的进度信息
            job_start_time = datetime.now().isoformat()
            self.update_job_status(job_id, {
                "status": "processing",
                "progress": "0",
                "progress_percentage": 0,
                "message": f"任务开始处理: {total_new_images}张新图片 + {total_skipped_images}张已处理图片",
                "started_at": job_start_time,
                "worker_id": self.worker_id,
                "total_images": str(total_processed),  # 总图片数
                "total_to_process": str(total_new_images),  # 需要处理的新图片数
                "processed_images": "0",  # 初始已处理数为0
                "skipped_images": str(total_skipped_images),  # 跳过图片数
                "current_image": "",  # 当前处理的图片
                "current_image_index": "0",  # 当前索引
                "current_image_name": "",  # 当前图片名称
                "pdf_folder": pdf_folder,
                "original_filename": original_filename,
                "db_filename": pdf_folder,
                "task_start_time": job_start_time,
                "db_query_success": str(db_connection_success)
            })


            # ✅ 4. 核心：使用优化后的图片处理函数（带实时进度更新）
            if total_new_images > 0:

                images_processed = 0
                tables_added = 0
                success = True
                errors = []

                # ✅ 关键修复：遍历每张图片，但通过定时检查控制流程
                for i, image_path in enumerate(images_to_process):
                    try:
                        image_name = os.path.basename(image_path)
                        current_processed = i + 1
                        progress_percentage = int((current_processed / total_new_images) * 100)

                        # ✅ 1. 立即更新状态
                        self.update_job_status(job_id, {
                            "status": "processing",
                            "progress": str(progress_percentage),
                            "progress_percentage": progress_percentage,
                            "message": f"正在处理第 {current_processed}/{total_new_images} 张图片: {image_name}",
                            "processed_images": str(current_processed),
                            "total_images": str(total_new_images),
                            "skipped_images": str(total_skipped_images),
                            "current_image": image_name,
                            "current_image_index": str(current_processed),
                            "current_image_name": image_name
                        })

                        # ✅ 2. 处理图片，不延迟
                        try:
                            from backend.api.convert.table_processor import process_single_table_image

                            # 异步处理图片（不阻塞主线程）
                            result = process_single_table_image(
                                pdf_folder=pdf_folder,
                                image_path=image_path,
                                table_type=table_type,
                                bank_name=bank_name
                            )

                            if result.get("success", False):
                                images_processed += 1
                                if "tables" in result and len(result["tables"]) > 0:
                                    tables_added += len(result["tables"])
                                print(f"  ✅ 处理成功")
                            else:
                                error_msg = result.get("error", "未知错误")
                                errors.append(f"{image_name}: {error_msg}")
                                print(f"  ⚠️ 处理失败: {error_msg}")

                        except Exception as img_error:
                            error_msg = f"图片处理异常: {str(img_error)}"
                            errors.append(f"{image_name}: {error_msg}")
                            print(f"  ❌ 处理异常: {img_error}")

                    except Exception as loop_error:
                        print(f"❌ 图片处理循环异常: {loop_error}")
                        import traceback
                        traceback.print_exc()
                        continue

                # 处理结果汇总
                processing_result = {
                    "success": success and len(errors) == 0,
                    "images_processed": images_processed,
                    "tables_added": tables_added,
                    "need_generate_excel": images_processed > 0 or total_skipped_images > 0,
                    "excel_exists": False,
                    "next_action": "generate_excel" if images_processed > 0 else "skip",
                    "errors": errors
                }
            else:
                print(f"ℹ️ 没有新图片需要处理")
                # 直接进入Excel生成阶段
                processing_result = {
                    "success": True,
                    "images_processed": 0,
                    "tables_added": 0,
                    "need_generate_excel": True,  # 即使没有新图片，也可能需要生成Excel
                    "excel_exists": False
                }

            # ✅ 5. 生成Excel文件
            excel_generated = False
            excel_path = ""

            try:
                # 更新状态：开始生成Excel
                self.update_job_status(job_id, {
                    "status": "generating_excel",
                    "progress": "95",
                    "message": "正在生成Excel文件...",
                    "processed_images": str(total_new_images),
                    "current_image": "正在生成Excel"
                })

                # ✅ 关键：获取PDF聚合器并生成Excel
                aggregator = pdf_aggregator_manager.get_aggregator(pdf_folder, bank_name)

                if len(aggregator) > 0:
                    print(f"📊 聚合器中有 {len(aggregator)} 个表格需要保存")

                    # 生成Excel
                    success, excel_path, error_msg = pdf_aggregator_manager.finalize_pdf(
                        pdf_folder=pdf_folder,
                        output_dir=EXCEL_DATA_DIR,
                        force=False,
                        metadata_list=[]
                    )

                    if success and excel_path and os.path.exists(excel_path):
                        excel_generated = True
                        file_size = os.path.getsize(excel_path)
                    else:
                        print(f"❌ Excel文件生成失败: {error_msg}")
                else:
                    print(f"⚠️ 聚合器中没有表格数据")

                    # 检查是否已有Excel文件
                    import glob
                    excel_dir = os.path.join(EXCEL_DATA_DIR, pdf_folder)
                    if os.path.exists(excel_dir):
                        excel_files = glob.glob(os.path.join(excel_dir, "*.xlsx"))
                        if excel_files:
                            existing_excel_path = excel_files[0]
                            if os.path.getsize(existing_excel_path) > 0:
                                excel_generated = True
                                excel_path = existing_excel_path
                                print(f"✅ 使用现有Excel文件: {excel_path}")
                            else:
                                print(f"⚠️ 现有Excel文件为空: {existing_excel_path}")
                        else:
                            print(f"⚠️ 没有找到Excel文件")
                    else:
                        print(f"⚠️ Excel目录不存在: {excel_dir}")

            except Exception as excel_error:
                print(f"❌ 生成Excel异常: {excel_error}")
                import traceback
                traceback.print_exc()
                excel_generated = False

            # ✅ 6. 最终状态更新
            duration = time.time() - task_start_time
            final_status = "completed" if excel_generated else "failed"
            final_message = (
                f"任务完成。处理 {total_new_images} 张新图片，跳过 {total_skipped_images} 张已处理图片"
                if excel_generated else
                f"任务失败。处理了 {total_new_images} 张图片，但生成Excel失败"
            )

            # 记录任务详细信息
            task_details = {
                "status": final_status,
                "progress": "100" if excel_generated else "0",
                "progress_percentage": 100 if excel_generated else 0,
                "message": final_message,
                "completed_at": datetime.now().isoformat(),
                "duration": f"{duration:.2f}秒",
                "total_processed": str(total_processed),
                "total_images": str(total_processed),
                "processed_images": str(total_new_images),
                "skipped_images": str(total_skipped_images),
                "excel_generated": str(excel_generated).lower(),
                "excel_path": excel_path if excel_generated else "",
                "original_filename": original_filename,
                "db_filename": pdf_folder,
                "task_start_time": job_start_time,
                "processing_summary": {
                    "total_images": len(image_paths),
                    "new_images_processed": total_new_images,
                    "skipped_images": total_skipped_images,
                    "aggregator_tables": len(pdf_aggregator_manager.get_aggregator(pdf_folder,
                                                                                   bank_name)) if 'pdf_aggregator_manager' in locals() else 0,
                    "excel_created": excel_generated,
                    "processing_time": f"{duration:.2f}秒",
                    "images_per_second": f"{total_new_images / duration:.2f}" if duration > 0 else "0"
                }
            }

            # 如果有错误信息，也记录下来
            if 'errors' in locals() and errors:
                task_details["errors"] = errors

            self.update_job_status(job_id, task_details)

            return excel_generated

        except Exception as e:
            print(f"❌ 任务处理异常: {e}")
            import traceback
            traceback.print_exc()

            # 异常时更新状态
            try:
                self.update_job_status(job_id, {
                    "status": "failed",
                    "progress": "0",
                    "message": f"任务处理异常: {str(e)}",
                    "error": str(e),
                    "completed_at": datetime.now().isoformat(),
                    "original_filename": original_filename if 'original_filename' in locals() else pdf_folder,
                    "db_filename": pdf_folder
                })
            except:
                pass

            return False


    def run(self):
        """运行Worker主循环 - 优化版本，增加超时错误处理"""
        print(f"\n{'=' * 60}")
        print(f"🚀 启动表格处理Worker")
        print(f"👷 Worker ID: {self.worker_id}")
        print(f"📅 启动时间: {datetime.now().isoformat()}")
        print(f"{'=' * 60}")

        # 连接Redis
        if not self.connect_redis():
            print("❌ Worker启动失败: Redis连接失败")
            return

        self.running = True
        empty_queue_count = 0
        connection_errors = 0  # 新增：记录连接错误次数

        while self.running:
            try:
                # 阻塞式从队列获取任务（等待30秒）
                result = self.redis_client.brpop("table_parse_queue", timeout=30)

                if result:
                    # 重置计数
                    empty_queue_count = 0
                    connection_errors = 0  # 重置连接错误计数

                    # 解析任务数据
                    queue_name, json_bytes = result
                    task_data = json.loads(json_bytes.decode('utf-8'))

                    # 处理任务
                    success = self.process_single_task(task_data)

                    if not success:
                        # 任务失败，可以考虑重试或记录
                        print(f"⚠️ 任务处理失败: {task_data.get('job_id')}")

                else:
                    # 队列为空
                    empty_queue_count += 1
                    if empty_queue_count % 10 == 0:  # 每5分钟报告一次
                        print(f"⏳ Worker {self.worker_id} 等待任务中... ({empty_queue_count * 30}秒)")

                    # 长时间空闲时检查连接
                    if empty_queue_count % 20 == 0:
                        try:
                            self.redis_client.ping()
                        except:
                            print("🔌 Redis连接检查失败，尝试重连...")
                            self.connect_redis()

            # ✅ 新增：处理TimeoutError
            except redis.exceptions.TimeoutError as e:
                connection_errors += 1
                print(f"⏱️ Redis读取超时 ({connection_errors}/3): {e}")

                if connection_errors >= 3:
                    print("🔌 Redis多次超时，尝试重连...")
                    self.connect_redis()
                    connection_errors = 0
                continue

            except redis.exceptions.ConnectionError as e:
                print(f"🔌 Redis连接错误，5秒后重试: {e}")
                time.sleep(5)
                self.connect_redis()
                continue

            except Exception as e:
                print(f"💥 Worker主循环异常: {e}")
                import traceback
                traceback.print_exc()
                time.sleep(10)

        print(f"🛑 Worker {self.worker_id} 已停止")



def main():
    """主函数"""
    import argparse

    parser = argparse.ArgumentParser(description='表格处理Worker')
    parser.add_argument('--id', type=str, help='Worker ID')
    parser.add_argument('--count', type=int, default=1, help='启动Worker数量')
    args = parser.parse_args()

    workers = []

    # 启动指定数量的Worker
    for i in range(args.count):
        worker_id = args.id or f"worker_{i + 1}_{int(time.time())}"
        worker = TableProcessingWorker(worker_id)

        if args.count > 1:
            # 启动多个Worker线程
            thread = threading.Thread(
                target=worker.run,
                name=f"TableWorker-{worker_id}",
                daemon=True
            )

            thread.start()
            workers.append((worker, thread))
        else:
            # 单Worker直接运行
            worker.run()

    # 等待所有Worker线程
    if args.count > 1:
        try:
            for worker, thread in workers:
                thread.join()
        except KeyboardInterrupt:
            print("\n🛑 收到中断信号，停止所有Worker...")
            for worker, _ in workers:
                worker.running = False



import sys
import subprocess


def main_with_reload():
    """带热更新的主函数"""
    import argparse

    parser = argparse.ArgumentParser(description='表格处理Worker')
    parser.add_argument('--id', type=str, help='Worker ID')
    parser.add_argument('--count', type=int, default=1, help='启动Worker数量')
    parser.add_argument('--reload', action='store_true', help='启用热更新')
    args = parser.parse_args()

    if args.reload:
        print("🔄 热更新模式已启用，修改文件后会自动重启")

        cmd = [sys.executable, __file__]
        if args.id:
            cmd.extend(['--id', args.id])
        if args.count:
            cmd.extend(['--count', str(args.count)])

        while True:
            process = subprocess.Popen(cmd)
            try:
                process.wait()
            except KeyboardInterrupt:
                print("\n🛑 停止进程...")
                process.terminate()
                break
            print("🔄 重启中...")
            time.sleep(1)
    else:
        # 您原有的main函数逻辑
        workers = []

        for i in range(args.count):
            worker_id = args.id or f"worker_{i + 1}_{int(time.time())}"
            worker = TableProcessingWorker(worker_id)

            if args.count > 1:
                thread = threading.Thread(
                    target=worker.run,
                    name=f"TableWorker-{worker_id}",
                    daemon=True
                )
                thread.start()
                workers.append((worker, thread))
            else:
                worker.run()

        if args.count > 1:
            try:
                for worker, thread in workers:
                    thread.join()
            except KeyboardInterrupt:
                print("\n🛑 收到中断信号，停止所有Worker...")
                for worker, _ in workers:
                    worker.running = False


if __name__ == "__main__":
    main_with_reload()