"""
文件相关蓝图 - 重构版本（只重构Excel转PDF功能）
"""
from flask import Blueprint, request, jsonify, send_from_directory, make_response
from backend.models.unified_db import DatabaseManager as OldDatabaseManager
from backend.utils.constants import UPLOAD_FOLDER, MAIN_ROOT, DATABASE, EXCEL_OUTPUT_ROOT
from pathlib import Path
import datetime

# 新增导入
from backend.models.safe_unified_db import SafeDatabaseManager  # 使用安全的数据库管理器
from backend.service.file_mapping_service import file_mapping_service


file_bp = Blueprint('file', __name__)

db = OldDatabaseManager(DATABASE)


# 尝试导入转换器，提供多种导入路径
CONVERTER_AVAILABLE = False
FinalDataConverter = None

try:
    # 尝试从 backend.src.services.table_processor 导入
    from backend.src.services.table_processor.long_format_converter import FinalDataConverter as FC
    FinalDataConverter = FC
    CONVERTER_AVAILABLE = True
    print("✅ long_format_converter 从标准路径导入成功")
except ImportError as e:
    print(f"⚠️ 标准路径导入失败: {e}")
    try:
        # 尝试从当前目录导入
        from long_format_converter import FinalDataConverter as FC
        FinalDataConverter = FC
        CONVERTER_AVAILABLE = True
        print("✅ long_format_converter 从当前目录导入成功")
    except ImportError as e2:
        print(f"❌ 所有导入尝试都失败: {e2}")
        CONVERTER_AVAILABLE = False



# ========== 导入重构后的Excel转PDF功能模块 ==========
from .file_handlers.validators import (
    validate_excel_file,
    validate_conversion_params,
    ValidationError
)
from .file_handlers.processors import (
    save_uploaded_file,
    convert_excel_to_pdf,
    save_and_return_result
)


# ========== 原有的所有接口保持不变 ==========
# 只添加新的Excel转PDF接口，其他所有接口保持原样
@file_bp.get('/files')
def list_files():
    print("🔍 文件列表API被调用")
    print(f"📁 FILES - 数据库路径: {DATABASE}")

    conn = db.connect()
    if not conn:
        print("❌ 数据库连接失败")
        return jsonify({"error": "数据库连接失败"}), 500

    c = conn.cursor()

    # 先查询所有文件（包括已删除的）
    c.execute("SELECT COUNT(*) as total FROM files")
    total_count = c.fetchone()[0]
    print(f"📊 数据库中总文件数（包括已删除）: {total_count}")

    c.execute("SELECT COUNT(*) as active FROM files WHERE deleted = 0")
    active_count = c.fetchone()[0]
    print(f"📊 未删除的文件数: {active_count}")

    # 执行原始查询
    c.execute(
        "SELECT id, filename, raw_filename, file_type, created_at, deleted "
        "FROM files WHERE deleted = 0 ORDER BY created_at DESC"
    )
    rows = c.fetchall()

    print(f"📋 查询结果行数: {len(rows)}")

    # 打印所有查询到的行（包括已删除的）
    c.execute("SELECT id, filename, raw_filename, deleted FROM files")
    all_rows = c.fetchall()
    print(f"📋 数据库中所有文件记录: {all_rows}")

    files_list = []
    upload_dir = Path(MAIN_ROOT) / UPLOAD_FOLDER
    missing_files = []  # 记录不存在的文件ID

    for row in rows:
        file_id = row[0]
        disk_name = row[1]  # UUID文件名
        raw_name = row[2]  # 原始中文名
        file_type = row[3]

        # 检查物理文件是否存在
        file_path = upload_dir / disk_name
        file_exists = file_path.exists()

        if file_exists:
            file_info = {
                "id": file_id,
                "filename": raw_name or disk_name,
                "disk_name": disk_name,
                "file_type": file_type,
                "created_at": row[4],
                "file_id": disk_name.split('.')[0] if '.' in disk_name else disk_name
            }
            files_list.append(file_info)
            print(f"✅ 包含文件: {raw_name} (磁盘名: {disk_name})")
        else:
            print(f"❌ 文件不存在，跳过: {raw_name} (磁盘名: {disk_name})")
            missing_files.append(file_id)

    # 自动删除数据库中不存在的文件记录（可选）
    if missing_files:
        print(f"🗑️ 自动删除 {len(missing_files)} 个不存在的文件记录")
        placeholders = ','.join('?' * len(missing_files))
        c.execute(f"DELETE FROM files WHERE id IN ({placeholders})", missing_files)
        conn.commit()

    conn.close()

    print(f"✅ 返回给前端的有效文件数: {len(files_list)}")
    return jsonify(files_list)



# ---------- 2. 下载/预览（不返回已软删） ----------
@file_bp.get('/file/<path:filename>')
def get_file(filename):
    """
    filename 可能是中文原始名，也可能是磁盘 UUID 名；
    一律先查库映射到真实磁盘名，且只返回未删除的。
    """
    print(f"🔍 查找文件: {filename}")

    conn = db.connect()
    if not conn:
        return jsonify({"error": "数据库连接失败"}), 500

    try:
        c = conn.cursor()
        # 查询文件信息（包括ID和磁盘文件名）
        c.execute(
            "SELECT id, filename, raw_filename FROM files "
            "WHERE (raw_filename = ? OR filename = ?) AND deleted = 0",
            (filename, filename)
        )
        row = c.fetchone()

        if row is None:
            print(f"❌ 文件不存在或已隐藏: {filename}")
            return jsonify({"error": "文件不存在或已隐藏"}), 404

        file_id = row["id"]
        real_name = row["filename"]  # 磁盘 UUID 文件名
        raw_name = row["raw_filename"]  # 原始中文名

        print(f"✅ 找到文件: ID={file_id}, 磁盘名={real_name}, 原始名={raw_name}")

        # 构建文件路径
        upload_dir = Path(MAIN_ROOT) / UPLOAD_FOLDER
        file_path = upload_dir / real_name
        print(f"📁 文件路径: {file_path}")
        print(f"📁 上传目录: {upload_dir}")

        if not file_path.exists():
            print(f"❌ 物理文件不存在: {file_path}")
            return jsonify({"error": "物理文件不存在"}), 404

        # 返回文件 - 使用正确的目录路径
        print(f"📤 返回文件: {real_name}")
        return send_from_directory(str(upload_dir), real_name)

    except Exception as e:
        print(f"❌ 文件查找错误: {e}")
        return jsonify({"error": "文件查找失败"}), 500
    finally:
        conn.close()




# ---------- 2.2 获取文件信息 ----------
@file_bp.get('/file-info/<path:filename>')
def get_file_info(filename):
    """获取文件详细信息"""
    print(f"🔍 获取文件信息: {filename}")

    conn = db.connect()
    if not conn:
        return jsonify({"error": "数据库连接失败"}), 500

    try:
        c = conn.cursor()
        c.execute(
            "SELECT id, filename, raw_filename, file_type, created_at FROM files "
            "WHERE (raw_filename = ? OR filename = ?) AND deleted = 0",
            (filename, filename)
        )
        row = c.fetchone()

        if row is None:
            return jsonify({"error": "文件不存在或已隐藏"}), 404

        # 构建文件路径信息
        upload_dir = Path(MAIN_ROOT) / UPLOAD_FOLDER
        file_path = upload_dir / row["filename"]
        file_exists = file_path.exists()

        return jsonify({
            "id": row["id"],
            "disk_name": row["filename"],
            "original_name": row["raw_filename"],
            "file_type": row["file_type"],
            "created_at": row["created_at"],
            "file_exists": file_exists,
            "file_path": str(file_path)
        })

    except Exception as e:
        print(f"❌ 获取文件信息错误: {e}")
        return jsonify({"error": "获取文件信息失败"}), 500
    finally:
        conn.close()




# ---------- 3. 软删除（增强版：物理文件不存在时直接删除） ----------
@file_bp.delete('/file/<path:filename>')
def delete_file(filename):
    """
    删除文件逻辑：
    1. 如果物理文件不存在，直接删除数据库记录
    2. 如果物理文件存在，只进行软删除 (deleted = 1)
    """
    conn = db.connect()
    if not conn:
        return jsonify({"error": "数据库连接失败"}), 500

    c = conn.cursor()

    try:
        # 先找真实磁盘名和文件信息
        c.execute(
            "SELECT id, filename, raw_filename FROM files "
            "WHERE (raw_filename = ? OR filename = ?) AND deleted = 0",
            (filename, filename)
        )
        row = c.fetchone()
        if row is None:
            conn.close()
            return jsonify({"error": "文件不存在"}), 404

        file_id = row["id"]
        real_name = row["filename"]
        raw_name = row["raw_filename"]

        print(f"🔍 删除文件: ID={file_id}, 磁盘名={real_name}, 原始名={raw_name}")

        # 检查物理文件是否存在
        upload_dir = Path(MAIN_ROOT) / UPLOAD_FOLDER
        file_path = upload_dir / real_name

        if not file_path.exists():
            print(f"📁 物理文件不存在，直接删除数据库记录: {file_path}")
            # 物理文件不存在，直接删除数据库记录
            c.execute("DELETE FROM files WHERE id = ?", (file_id,))
            conn.commit()
            conn.close()
            return jsonify({"message": "文件已彻底删除（物理文件不存在）"}), 200
        else:
            print(f"📁 物理文件存在，进行软删除: {file_path}")
            # 物理文件存在，进行软删除
            c.execute("UPDATE files SET deleted = 1 WHERE filename = ?", (real_name,))
            conn.commit()
            conn.close()
            return jsonify({"message": "已隐藏（软删除）"}), 200

    except Exception as e:
        print(f"❌ 删除文件时出错: {e}")
        if conn:
            conn.rollback()
            conn.close()
        return jsonify({"error": "删除文件失败"}), 500




# ---------- 4. 搜索PDF文件（新增接口） ----------
@file_bp.get('/search-pdf')
def search_pdf():
    """
    搜索PDF文件名称 - 基于文件映射服务
    参数: keyword - 搜索关键词
    返回: 匹配的PDF文件列表（显示原始中文名）
    """
    keyword = request.args.get('keyword', '').strip()
    print(f"搜索关键词: '{keyword}'")

    if not keyword:
        return jsonify({"files": []})

    try:
        # 使用文件映射服务搜索PDF文件
        search_results = file_mapping_service.search_files(keyword, 'pdf')

        return jsonify({"files": search_results})

    except Exception as e:
        print(f"搜索PDF失败: {e}")
        return jsonify({"error": "搜索失败"}), 500




@file_bp.get('/file-by-id/<file_id>')
def get_file_by_id(file_id):
    try:
        print(f"🔍🔍 文件下载请求 file_id: {file_id}")

        # 直接使用 file_id（UUID格式）
        file_uuid = file_id.split('.')[0] if '.' in file_id else file_id
        print(f"🔍🔍 直接使用UUID: {file_uuid}")

        # 检查文件是否存在映射中
        file_info = file_mapping_service.get_file_info(file_uuid)
        if not file_info:
            print(f"❌ 文件映射服务中没有找到UUID: {file_uuid}")
            return jsonify({"error": "文件不存在"}), 404

        disk_name = file_info["disk_name"]
        PDF_DIR = Path(MAIN_ROOT) / UPLOAD_FOLDER
        print(f"✅ 准备返回文件: {disk_name}")

        return send_from_directory(PDF_DIR, disk_name)

    except Exception as e:
        print(f"❌❌ 文件下载失败: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": "文件下载失败"}), 500

@file_bp.get('/excel-sheets/<file_id>')
def get_excel_sheets(file_id):
    """
    根据PDF文件ID获取对应的Excel sheet列表（支持数字ID和UUID）
    Excel文件存储在 backend/static/excel_data/<file_id>/ 目录中
    """
    try:
        print(f"🔍🔍 获取Excel sheets请求 file_id: {file_id}")

        # 如果file_id是数字，需要先查询数据库获取真实的UUID
        if file_id.isdigit():
            print(f"🔍🔍 检测到数字ID，查询数据库: {file_id}")
            # 通过数字ID查询真实文件信息
            conn = db.connect()
            c = conn.cursor()
            c.execute("SELECT filename FROM files WHERE id = ? AND deleted = 0", (file_id,))
            row = c.fetchone()
            conn.close()

            if row:
                real_file_id = row["filename"]  # 获取真实的UUID
                print(f"✅ 找到对应文件 UUID: {real_file_id}")
            else:
                print(f"❌ 数据库中没有找到ID为 {file_id} 的文件")
                return jsonify({"error": "PDF文件不存在"}), 404
        else:
            real_file_id = file_id
            print(f"🔍🔍 直接使用UUID: {real_file_id}")

        # 1. 验证PDF文件是否存在
        file_info = file_mapping_service.get_file_info(real_file_id)
        print("file_info:", file_info)
        if not file_info:
            print(f"❌ 文件映射服务中没有找到UUID: {real_file_id}")
            return jsonify({"error": "PDF文件不存在"}), 404

        # 2. 构建Excel文件目录路径 - 修正：统一使用Path对象
        excel_dir = Path(MAIN_ROOT) / EXCEL_OUTPUT_ROOT / real_file_id
        print(f"🔍🔍 Excel目录路径: {excel_dir}")

        if not excel_dir.exists():
            print(f"⚠️ Excel目录不存在: {excel_dir}")
            return jsonify({"excel_files": []})

        # 3. 查找目录中的所有Excel文件
        excel_files = []
        supported_extensions = ['.xlsx', '.xls']

        for ext in supported_extensions:
            print(f"🔍🔍 查找扩展名: {ext}")
            for excel_file in excel_dir.glob(f"*{ext}"):
                print(f"🔍🔍 找到Excel文件: {excel_file}")
                if excel_file.is_file():
                    print(f"✅ 添加Excel文件: {excel_file.name}")
                    excel_files.append({
                        "file_name": excel_file.name,
                        "file_path": str(excel_file),
                        "file_size": excel_file.stat().st_size
                    })

        if not excel_files:
            print("⚠️ 目录中没有找到Excel文件")
            return jsonify({"excel_files": []})

        print(f"✅ 找到 {len(excel_files)} 个Excel文件")

        # 4. 读取每个Excel文件的sheet列表
        import pandas as pd
        result = []

        for excel_file in excel_files:
            try:
                print(f"🔍🔍 读取Excel文件: {excel_file['file_name']}")
                # 读取Excel文件的所有sheet
                excel_file_obj = pd.ExcelFile(excel_file["file_path"])
                sheet_names = excel_file_obj.sheet_names

                file_sheets = []
                for sheet_name in sheet_names:
                    file_sheets.append({
                        "name": sheet_name,
                        "excel_file": excel_file["file_name"]
                    })

                result.append({
                    "excel_file": excel_file["file_name"],
                    "sheets": file_sheets,
                    "total_sheets": len(sheet_names)
                })
                print(f"✅ 成功读取 {len(sheet_names)} 个sheet")

            except Exception as e:
                print(f"❌ 读取Excel文件失败 {excel_file['file_name']}: {e}")
                # 即使读取失败，也返回文件信息
                result.append({
                    "excel_file": excel_file["file_name"],
                    "sheets": [],
                    "total_sheets": 0,
                    "error": str(e)
                })
                continue

        print(f"✅ 返回结果: {len(result)} 个Excel文件信息")
        return jsonify({
            "excel_files": result,
            "pdf_id": real_file_id,
            "pdf_name": file_info["original_name"],
            "total_excel_files": len(result)
        })

    except Exception as e:
        print(f"❌❌ 获取Excel sheet列表失败: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": "获取表格列表失败"}), 500



@file_bp.get('/excel-data/<file_id>/<path:excel_file_name>/<sheet_name>')
def get_excel_data(file_id, excel_file_name, sheet_name):
    """
    根据文件ID、Excel文件名和sheet名称获取Excel数据
    """
    try:
        print("🎯 直接读取Excel文件数据（跳过快照）")
        print(f"📋 请求参数: file_id={file_id}, excel_file={excel_file_name}, sheet={sheet_name}")

        # 🔥🔥🔥 新增：如果file_id是数字，转换为真实UUID
        if file_id.isdigit():
            print(f"🔍 检测到数字ID，查询真实UUID: {file_id}")

            # 1. 先尝试从 table_processing_records 查询
            conn = db.connect()
            c = conn.cursor()

            # 🔥 先检查表结构
            c.execute("PRAGMA table_info(table_processing_records)")
            columns = c.fetchall()
            column_names = [col[1] for col in columns]
            print(f"🔍 table_processing_records 表结构: {column_names}")

            # 根据实际表结构构建查询
            if 'pdf_folder' in column_names:
                # 如果有 pdf_folder 列
                c.execute("""
                    SELECT pdf_folder 
                    FROM table_processing_records 
                    WHERE id = ? OR CAST(id AS TEXT) = ?
                """, (file_id, file_id))
                row = c.fetchone()

                if row:
                    real_file_id = row["pdf_folder"]
                    print(f"✅ 从 table_processing_records 找到UUID: {real_file_id}")
                else:
                    # 如果没有找到，尝试查询 files 表
                    c.execute("""
                        SELECT filename 
                        FROM files 
                        WHERE id = ? AND deleted = 0
                    """, (file_id,))
                    row = c.fetchone()

                    if row:
                        real_file_id = row["filename"]
                        print(f"✅ 从 files 表找到UUID: {real_file_id}")
                    else:
                        print(f"❌ 未找到ID为 {file_id} 的记录")
                        conn.close()
                        return jsonify({"error": "文件不存在"}), 404
            else:
                # 如果没有 pdf_folder 列，直接查询 files 表
                c.execute("""
                    SELECT filename 
                    FROM files 
                    WHERE id = ? AND deleted = 0
                """, (file_id,))
                row = c.fetchone()

                if row:
                    real_file_id = row["filename"]
                    print(f"✅ 从 files 表找到UUID: {real_file_id}")
                else:
                    print(f"❌ 未找到ID为 {file_id} 的记录")
                    conn.close()
                    return jsonify({"error": "文件不存在"}), 404

            conn.close()

            if 'real_file_id' in locals():
                file_id = real_file_id  # 替换为真实的UUID
            else:
                print(f"❌ 未找到ID为 {file_id} 的记录")
                return jsonify({"error": "文件不存在"}), 404
        else:
            print(f"🔍 已经是UUID格式: {file_id}")

        # 1. 构建Excel文件路径
        excel_dir = Path(MAIN_ROOT) / EXCEL_OUTPUT_ROOT / file_id
        excel_path = excel_dir / excel_file_name

        print("📁 Excel文件路径:", excel_path)

        if not excel_path.exists():
            print("❌ Excel文件不存在")
            return jsonify({"error": "Excel文件不存在"}), 404

        # 2. 读取指定sheet，不把任何行当作表头
        import pandas as pd
        df = pd.read_excel(excel_path, sheet_name=sheet_name, header=None)
        df = df.fillna('')

        print("✅ Excel文件读取成功")
        print("📊 Excel原始数据形状:", df.shape)
        print("🔢 列数:", df.shape[1], "行数:", df.shape[0])

        # 获取表头
        horizontal_headers = []
        vertical_headers = []
        data_rows = []
        top_left_cell = ""

        # 提取横向表头（第一行，从第二列开始）
        if df.shape[0] > 0:
            # 第一列第一行可能是左上角单元格
            top_left_cell = str(df.iloc[0, 0]) if df.iloc[0, 0] != '' else ""

            # 横向表头（第一行，从第二列开始）
            for col in range(1, df.shape[1]):
                header = str(df.iloc[0, col]) if df.iloc[0, col] != '' else f""
                horizontal_headers.append(header)

        # 提取纵向表头（第一列，从第二行开始）
        if df.shape[1] > 0:
            for row in range(1, df.shape[0]):
                header = str(df.iloc[row, 0]) if df.iloc[row, 0] != '' else f""
                vertical_headers.append(header)

        # 提取数据（从第二行第二列开始）
        for row in range(1, df.shape[0]):
            data_row = []
            for col in range(1, df.shape[1]):
                value = df.iloc[row, col]
                # 尝试保持数值类型
                try:
                    if isinstance(value, (int, float)):
                        data_row.append(value)
                    elif str(value).replace(',', '').replace('.', '').isdigit():
                        clean_value = str(value).replace(',', '')
                        if '.' in str(value):
                            data_row.append(float(clean_value))
                        else:
                            data_row.append(int(clean_value))
                    else:
                        data_row.append(str(value) if value != '' else "")
                except:
                    data_row.append(str(value) if value != '' else "")
            data_rows.append(data_row)

        print("📈 数据结构分析:")
        print(f"   - 左上角单元格: '{top_left_cell}'")
        print(f"   - 横向表头数: {len(horizontal_headers)}")
        print(f"   - 纵向表头数: {len(vertical_headers)}")
        print(f"   - 数据行数: {len(data_rows)}")
        print(f"   - 数据列数: {len(data_rows[0]) if data_rows else 0}")
        print(f"   - 横向表头样本: {horizontal_headers[:3]}")
        print(f"   - 纵向表头样本: {vertical_headers[:3]}")
        print(f"   - 数据样本: {data_rows[0][:3] if data_rows else '无'}")

        # 4. 构建前端友好的数据结构
        frontend_data = []

        # 添加元数据行
        metadata_row = {
            "__metadata": {
                "has_dual_headers": True,
                "top_left_cell": top_left_cell,
                "horizontal_headers": horizontal_headers,
                "vertical_headers": vertical_headers
            }
        }
        frontend_data.append(metadata_row)

        # 添加表头数据行（第一行：左上角 + 横向表头）
        header_row = {
            "__is_first_row": True,
            "__top_left_cell": top_left_cell
        }
        for i, header in enumerate(horizontal_headers, 1):
            header_row[f"H_{i}"] = header

        frontend_data.append(header_row)

        # 添加数据行（纵向表头 + 数据）
        for i in range(len(data_rows)):
            row_data = data_rows[i]
            vertical_header = vertical_headers[i] if i < len(vertical_headers) else f""

            row_obj = {
                "__is_data_row": True,
                "__vertical_header": vertical_header
            }

            for j, value in enumerate(row_data, 1):
                row_obj[f"H_{j}"] = value

            frontend_data.append(row_obj)

        print("✅ 前端数据构建完成:")
        print(f"   - 总行数: {len(frontend_data)}")
        print(f"   - 总列数: {len(horizontal_headers)}")
        print(f"   - 数据样本: {frontend_data[:2]}")

        result = {
            "rows": frontend_data,
            "total_rows": len(frontend_data),
            "total_columns": len(horizontal_headers),
            "sheet_name": sheet_name,
            "excel_file": excel_file_name,
            "pdf_id": file_id,
            "has_dual_headers": True,
            "source": "excel_file"  # 标记数据来源为Excel文件
        }

        print("++++++++++++++++result+++++++++++++++++++")
        print(result)

        # 5. 返回给前端
        return jsonify(result)

    except Exception as e:
        print(f"❌ 处理Excel数据请求失败: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"处理请求失败: {str(e)}"}), 500


def get_correct_pdf_id(pdf_id):
    """获取正确的PDF ID（UUID格式）"""
    # 如果已经是UUID格式，直接返回
    if not pdf_id.isdigit():
        return pdf_id

    # 如果是数字ID，查询对应的UUID
    try:
        conn = db.connect()
        c = conn.cursor()

        # 🔥🔥🔥 修复：使用正确的列名 filename
        query = "SELECT filename FROM files WHERE id = ? AND deleted = 0"
        params = (int(pdf_id),)  # 确保是整数

        print(f"🔍 查询UUID: {query} 参数: {params}")
        c.execute(query, params)
        row = c.fetchone()
        conn.close()

        if row:
            uuid = row["filename"]
            print(f"✅ 数字ID {pdf_id} 对应的UUID: {uuid}")
            return uuid
        else:
            print(f"⚠️ 未找到数字ID {pdf_id} 对应的UUID，使用原ID")
            return pdf_id

    except Exception as e:
        print(f"❌ 查询UUID失败: {e}，使用原ID: {pdf_id}")
        import traceback
        traceback.print_exc()
        return pdf_id


@file_bp.route('/excel/save-final', methods=['POST'])
def save_final_excel():
    """统一保存整个表格数据 - 保护其他Sheet"""
    data = request.json
    print("******************** 保存整个表格（保护其他Sheet） ******************")

    # 基础校验
    required_fields = ['pdf_id', 'excel_file', 'sheet_name', 'table_type', 'data']
    for field in required_fields:
        if field not in data:
            return jsonify({'error': f'缺少必要字段: {field}'}), 400

    try:
        pdf_id = data['pdf_id']
        excel_file = data['excel_file']
        sheet_name = data['sheet_name']
        table_type = data['table_type']
        table_data = data['data']

        print(f"💾 保存数据: PDF={pdf_id}, 文件={excel_file}, Sheet={sheet_name}, 类型={table_type}")
        print(f"📊 接收数据: {len(table_data)}行 × {len(table_data[0]) if table_data else 0}列")

        # 根据表类型选择保存方式
        if table_type == 'original':
            result = save_complete_table_data(pdf_id, excel_file, sheet_name, table_data, table_type)
        elif table_type == 'flattened':
            result = save_flattened_table_data(pdf_id, excel_file, sheet_name, table_data, table_type)
        else:
            return jsonify({'error': f'不支持的表类型: {table_type}'}), 400

        if not result['success']:
            return jsonify({'success': False, 'error': result['error']}), 500

        return jsonify({
            'success': True,
            'message': '表格数据保存成功',
            'saved_count': result.get('saved_rows', 0),
            'data_dimensions': result.get('data_dimensions', '未知'),
            'excel_updated': result.get('excel_updated', False),
            'sheets_protected': result.get('sheets_protected', False),
            'protected_sheets_count': result.get('protected_sheets_count', 0),
            'file_created': result.get('file_created', False)
        }), 200

    except Exception as e:
        print(f"❌ 保存失败: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': f'保存失败: {str(e)}'}), 500


def save_complete_table_data(pdf_id, excel_file, sheet_name, table_data, table_type):
    """保存完整表格数据 - 只覆盖目标Sheet"""
    print("📊 保存完整表格数据（保护其他Sheet）...")

    # 🔥 在函数开头初始化变量
    protected_sheets = []

    try:
        # 1. 获取Excel文件路径
        

        from pathlib import Path
        pdf_id = get_correct_pdf_id(pdf_id)
        excel_dir = Path(MAIN_ROOT) / EXCEL_OUTPUT_ROOT / pdf_id
        excel_path = excel_dir / excel_file

        print(f"📁 Excel文件路径: {excel_path}")

        if not excel_path.exists():
            return {'success': False, 'error': f'Excel文件不存在: {excel_path}'}

        # 2. 加载工作簿
        from openpyxl import load_workbook
        workbook = load_workbook(excel_path)

        # 🔥 记录所有Sheet，确保保护其他Sheet
        all_sheets = workbook.sheetnames.copy()
        print(f"📋 工作簿包含 {len(all_sheets)} 个Sheet: {all_sheets}")
        print(f"🎯 目标Sheet: {sheet_name}")

        if sheet_name not in all_sheets:
            workbook.close()
            return {'success': False, 'error': f'Sheet不存在: {sheet_name}'}

        # 3. 🔥 只操作目标Sheet，保护其他Sheet
        worksheet = workbook[sheet_name]
        print(f"📈 目标Sheet原维度: {worksheet.max_row}行 × {worksheet.max_column}列")
        print(f"📦 新数据维度: {len(table_data)}行 × {len(table_data[0]) if table_data else 0}列")

        # 4. 🔥 只清空目标Sheet的数据（保留表头结构）
        if worksheet.max_row > 0:
            # 删除所有行（包括表头）
            worksheet.delete_rows(1, worksheet.max_row)
            print(f"🗑️ 清空目标Sheet: 删除了{worksheet.max_row}行")

        # 5. 🔥 写入完整数据到目标Sheet
        for row_idx, row_data in enumerate(table_data, 1):  # 从第1行开始
            for col_idx, cell_value in enumerate(row_data, 1):
                worksheet.cell(row=row_idx, column=col_idx, value=cell_value)

        print(f"📝 写入目标Sheet: {len(table_data)}行 × {len(table_data[0])}列")

        # 6. 🔥 保存工作簿（其他Sheet自动保留）
        workbook.save(excel_path)
        workbook.close()

        # 7. 🔥 验证其他Sheet是否被保护
        workbook_after = load_workbook(excel_path)
        sheets_after = workbook_after.sheetnames
        workbook_after.close()

        # 🔥 计算受保护的Sheet
        protected_sheets = [s for s in all_sheets if s != sheet_name]

        print(f"✅ 保存完成，验证Sheet保护:")
        print(f"  保存前Sheet数: {len(all_sheets)}")
        print(f"  保存后Sheet数: {len(sheets_after)}")
        print(f"  Sheet保持一致: {set(all_sheets) == set(sheets_after)}")

        if protected_sheets:
            print(f"  🛡️ 受保护的Sheet: {protected_sheets}")

        # 8. 🔥 保存快照
        print("📸 开始保存数据快照...")
        snapshot_result = save_data_snapshot(pdf_id, excel_file, sheet_name, table_data, table_type)
        print(f"✅ 快照保存结果: {snapshot_result.get('success', False)}")

        print("✅ 完整表格数据保存成功（其他Sheet已保护）")

        return {
            'success': True,
            'saved_rows': len(table_data),
            'saved_columns': len(table_data[0]) if table_data else 0,
            'data_dimensions': f'{len(table_data)}行 × {len(table_data[0])}列',
            'excel_updated': True,
            'sheets_protected': True,
            'protected_sheets_count': len(protected_sheets),  # 🔥 现在这个变量已定义
            'snapshot_saved': snapshot_result.get('success', False),
            'snapshot_path': snapshot_result.get('path', '')
        }

    except Exception as e:
        print(f"❌ 完整表格保存失败: {e}")
        import traceback
        traceback.print_exc()
        return {'success': False, 'error': f'完整表格保存失败: {str(e)}'}


def save_flattened_table_data(pdf_id, excel_file, sheet_name, table_data, table_type):
    """保存扁平化表格数据 - 只覆盖目标Sheet"""
    print("📊 保存扁平化表格数据（保护其他Sheet）...")

    try:
        

        from pathlib import Path
        pdf_id = get_correct_pdf_id(pdf_id)
        excel_dir = Path(MAIN_ROOT) / EXCEL_OUTPUT_ROOT / pdf_id
        excel_dir.mkdir(parents=True, exist_ok=True)
        excel_path = excel_dir / excel_file

        print(f"📁 Excel文件路径: {excel_path}")

        from openpyxl import Workbook, load_workbook

        file_exists = excel_path.exists()
        all_sheets = []

        if file_exists:
            # 🔥 文件存在，加载现有工作簿（保护其他Sheet）
            workbook = load_workbook(excel_path)
            all_sheets = workbook.sheetnames.copy()
            print(f"📄 加载现有Excel文件，包含 {len(all_sheets)} 个Sheet: {all_sheets}")
        else:
            # 文件不存在，创建新工作簿
            workbook = Workbook()
            # 删除默认Sheet
            default_sheet = workbook.active
            workbook.remove(default_sheet)
            print("📄 创建新Excel文件")

        # 🔥 处理目标Sheet（不影响其他Sheet）
        if sheet_name in workbook.sheetnames:
            print(f"📋 Sheet已存在，删除重写: {sheet_name}")
            del workbook[sheet_name]

        # 创建/重写目标Sheet
        worksheet = workbook.create_sheet(sheet_name)
        print(f"✅ 创建/重写目标Sheet: {sheet_name}")

        # 写入完整数据到目标Sheet
        for row_idx, row_data in enumerate(table_data, 1):
            for col_idx, cell_value in enumerate(row_data, 1):
                worksheet.cell(row=row_idx, column=col_idx, value=cell_value)

        print(f"📝 写入目标Sheet: {len(table_data)}行 × {len(table_data[0])}列")

        # 保存工作簿
        workbook.save(excel_path)
        workbook.close()

        # 验证Sheet保护
        if file_exists and all_sheets:
            workbook_after = load_workbook(excel_path)
            sheets_after = workbook_after.sheetnames
            workbook_after.close()

            protected_sheets = [s for s in all_sheets if s != sheet_name]
            print(f"✅ Sheet保护验证:")
            print(f"  保存前Sheet: {len(all_sheets)}个")
            print(f"  保存后Sheet: {len(sheets_after)}个")
            print(f"  受保护Sheet: {len(protected_sheets)}个")
            print(f"  Sheet列表: {protected_sheets}")

        action = "创建" if not file_exists else "更新"
        print(f"✅ 扁平化数据{action}成功（其他Sheet已保护）")

        return {
            'success': True,
            'file_created': not file_exists,
            'saved_rows': len(table_data),
            'saved_columns': len(table_data[0]) if table_data else 0,
            'excel_updated': True,
            'sheets_protected': file_exists,  # 只有文件存在时才有其他Sheet需要保护
            'protected_sheets_count': len(all_sheets) - 1 if file_exists else 0
        }

    except Exception as e:
        print(f"❌ 扁平化数据保存失败: {e}")
        import traceback
        traceback.print_exc()
        return {'success': False, 'error': f'扁平化数据保存失败: {str(e)}'}


def handle_frontend_data_format(data):
    """处理前端发送的数据格式"""
    print("🎯 处理前端数据格式...")

    pdf_id = data['pdf_id']
    excel_file = data['excel_file']
    sheet_name = data['sheet_name']
    table_type = data['table_type']

    # 🔥 关键：检查前端发送的数据字段
    modifications = data.get('modifications', [])
    current_data = data.get('current_data') or data.get('data')  # 兼容两种字段名
    total_changes = data.get('total_changes', 0)

    print(f"📊 前端数据详情:")
    print(f"  📝 修改记录数: {len(modifications)}")
    print(f"  📦 当前数据行数: {len(current_data) if current_data else 0}")
    print(f"  🔢 总修改数: {total_changes}")

    # 🔥 策略1：优先使用修改记录（最精确）
    if modifications and len(modifications) > 0:
        print("🔧 策略1: 使用修改记录进行精确更新")
        result = save_with_modifications(pdf_id, excel_file, sheet_name, modifications, table_type)
        if result['success']:
            return result
        else:
            print(f"⚠️ 修改记录保存失败，回退到完整数据: {result['error']}")

    # 🔥 策略2：使用当前数据
    if current_data and len(current_data) > 0:
        print("📊 策略2: 使用当前完整数据")
        return save_with_current_data(pdf_id, excel_file, sheet_name, current_data, table_type)

    # 🔥 策略3：没有有效数据
    print("❌ 策略3: 没有接收到有效数据")
    return {'success': False, 'error': '没有接收到有效数据'}


def save_with_modifications(pdf_id, excel_file, sheet_name, modifications, table_type):
    """使用修改记录精确更新"""
    print(f"🔧 使用修改记录更新: {len(modifications)} 个修改")

    try:
        # 1. 获取Excel文件路径
        

        from pathlib import Path
        pdf_id = get_correct_pdf_id(pdf_id)
        excel_dir = Path(MAIN_ROOT) / EXCEL_OUTPUT_ROOT / pdf_id
        excel_path = excel_dir / excel_file

        print(f"📁 Excel文件路径: {excel_path}")

        if not excel_path.exists():
            return {'success': False, 'error': f'Excel文件不存在: {excel_path}'}

        # 2. 加载工作簿
        from openpyxl import load_workbook
        workbook = load_workbook(excel_path)

        if sheet_name not in workbook.sheetnames:
            workbook.close()
            return {'success': False, 'error': f'Sheet不存在: {sheet_name}'}

        worksheet = workbook[sheet_name]
        print(f"📊 目标工作表: {worksheet.max_row}行 × {worksheet.max_column}列")

        # 3. 🔥 精确解析前端修改记录格式
        applied_count = 0
        failed_count = 0

        for i, mod in enumerate(modifications):
            try:
                # 🔥 前端格式：{ row, col, oldValue, newValue, saved, timestamp, tableType }
                row_idx = mod.get('row')
                col_idx = mod.get('col')
                new_value = mod.get('newValue') or mod.get('new_value')  # 兼容两种字段名

                print(f"  🔍 处理修改 {i + 1}: row={row_idx}, col={col_idx}, value={new_value}")

                # 验证修改记录
                if (row_idx is not None and col_idx is not None and
                        new_value is not None and
                        isinstance(row_idx, (int, float)) and
                        isinstance(col_idx, (int, float))):

                    # 转换为整数（前端从0开始，Excel从1开始）
                    excel_row = int(row_idx) + 1
                    excel_col = int(col_idx) + 1

                    # 检查行列是否在有效范围内
                    if (excel_row >= 1 and excel_row <= worksheet.max_row and
                            excel_col >= 1 and excel_col <= worksheet.max_column):

                        # 应用修改
                        worksheet.cell(row=excel_row, column=excel_col, value=new_value)
                        applied_count += 1
                        print(f"    ✅ 应用修改: [{row_idx},{col_idx}] = '{new_value}'")
                    else:
                        print(f"    ⚠️ 坐标超出范围: [{row_idx},{col_idx}] -> Excel[{excel_row},{excel_col}]")
                        failed_count += 1
                else:
                    print(f"    ⚠️ 无效修改格式: {mod}")
                    failed_count += 1

            except Exception as e:
                print(f"    ❌ 修改处理失败: {e}")
                failed_count += 1
                continue

        # 4. 保存文件
        workbook.save(excel_path)
        workbook.close()

        print(f"✅ 修改应用完成: {applied_count}成功, {failed_count}失败")

        return {
            'success': True,
            'strategy_used': 'modifications',
            'saved_count': applied_count,
            'failed_count': failed_count,
            'excel_updated': applied_count > 0,
            'data_dimensions': f'{applied_count}处修改'
        }

    except Exception as e:
        print(f"❌ 修改记录保存失败: {e}")
        return {'success': False, 'error': f'修改记录保存失败: {str(e)}'}


def save_with_current_data(pdf_id, excel_file, sheet_name, current_data, table_type):
    """使用当前完整数据覆盖"""
    print("📊 使用当前完整数据覆盖")
    print(f"🔍 接收数据格式: {type(current_data)}, 长度: {len(current_data)}")

    try:
        # 1. 获取Excel文件路径
        

        from pathlib import Path
        pdf_id = get_correct_pdf_id(pdf_id)
        excel_dir = Path(MAIN_ROOT) / EXCEL_OUTPUT_ROOT / pdf_id
        excel_path = excel_dir / excel_file

        print(f"📁 Excel文件路径: {excel_path}")

        if not excel_path.exists():
            return {'success': False, 'error': f'Excel文件不存在: {excel_path}'}

        # 2. 🔥 转换前端数据格式为二维数组
        print("🔄 转换前端数据格式...")
        backend_data = convert_frontend_to_backend_format(current_data)

        if not backend_data or len(backend_data) == 0:
            return {'success': False, 'error': '转换后数据为空'}

        print(f"✅ 转换后数据: {len(backend_data)}行 × {len(backend_data[0]) if backend_data else 0}列")

        # 3. 加载工作簿
        from openpyxl import load_workbook
        workbook = load_workbook(excel_path)

        if sheet_name not in workbook.sheetnames:
            workbook.close()
            return {'success': False, 'error': f'Sheet不存在: {sheet_name}'}

        worksheet = workbook[sheet_name]
        expected_columns = worksheet.max_column

        print(f"📈 Excel工作表列数: {expected_columns}")
        print(f"📊 转换后数据列数: {len(backend_data[0])}")

        # 4. 修复列数不匹配
        if len(backend_data[0]) != expected_columns:
            print(f"⚠️ 列数不匹配! 数据{len(backend_data[0])}列, Excel{expected_columns}列")
            backend_data = fix_column_mismatch(backend_data, expected_columns)
            print(f"✅ 修复后数据: {len(backend_data)}行 × {len(backend_data[0])}列")

        # 5. 清空数据行（保留表头）
        if worksheet.max_row > 1:
            rows_to_delete = worksheet.max_row - 1
            worksheet.delete_rows(2, rows_to_delete)
            print(f"🗑️ 清空数据行: 删除了{rows_to_delete}行")

        # 6. 写入新数据
        if backend_data and len(backend_data) > 0:
            for row_idx, row_data in enumerate(backend_data, 2):  # 从第2行开始（保留表头）
                for col_idx, cell_value in enumerate(row_data, 1):
                    if col_idx <= worksheet.max_column:
                        worksheet.cell(row=row_idx, column=col_idx, value=cell_value)

            print(f"📝 写入数据: {len(backend_data)}行 × {len(backend_data[0])}列")

        # 7. 保存文件
        workbook.save(excel_path)
        workbook.close()

        print("✅ 完整数据覆盖完成")

        return {
            'success': True,
            'strategy_used': 'current_data',
            'saved_count': len(backend_data),
            'excel_updated': True,
            'data_dimensions': f'{len(backend_data)}行 × {len(backend_data[0])}列'
        }

    except Exception as e:
        print(f"❌ 完整数据保存失败: {e}")
        return {'success': False, 'error': f'完整数据保存失败: {str(e)}'}


def convert_frontend_to_backend_format(frontend_data):
    """🔥 精确转换前端数据格式为二维数组"""
    if not frontend_data:
        print("⚠️ 前端数据为空")
        return []

    print("🔄 转换前端数据格式...")
    print(f"📊 原始数据: {len(frontend_data)}行")

    backend_data = []

    for i, row in enumerate(frontend_data):
        # 🔥 处理前端可能的多种格式
        if isinstance(row, list):
            # 已经是数组格式，直接使用
            backend_data.append(row)
            if i < 2:  # 只打印前2行样本
                print(f"  ✅ 行{i}: 数组格式 ({len(row)}列)")

        elif isinstance(row, dict):
            # 🔥 处理对象格式：{ H_1: '值1', H_2: '值2', ... }
            if any(key.startswith('H_') for key in row.keys()):
                # 提取 H_1, H_2, H_3, ... 字段
                row_values = []
                col_idx = 1

                while f'H_{col_idx}' in row:
                    value = row[f'H_{col_idx}']
                    row_values.append(value)
                    col_idx += 1

                if row_values:
                    backend_data.append(row_values)
                    if i < 2:
                        print(f"  ✅ 行{i}: H_*对象格式 ({len(row_values)}列)")
                else:
                    print(f"  ⚠️ 行{i}: H_*对象但无有效值")

            # 🔥 跳过元数据行
            elif row.get('__metadata') or row.get('__is_first_row'):
                if i < 2:
                    print(f"  ⏭️ 行{i}: 跳过元数据行")
                continue

            else:
                # 其他对象格式，尝试提取所有值（跳过内部字段）
                row_values = []
                for key, value in row.items():
                    if not key.startswith('__'):  # 跳过内部字段
                        row_values.append(value)

                if row_values:
                    backend_data.append(row_values)
                    if i < 2:
                        print(f"  ✅ 行{i}: 普通对象格式 ({len(row_values)}列)")
                else:
                    print(f"  ⚠️ 行{i}: 无法处理的对象格式")

        else:
            # 其他格式（字符串、数字等），包装成数组
            backend_data.append([row])
            if i < 2:
                print(f"  ✅ 行{i}: 简单值转数组")

    print(f"📈 转换完成: {len(backend_data)}行有效数据")

    if backend_data and len(backend_data) > 0:
        print(f"📏 数据维度: {len(backend_data)}行 × {len(backend_data[0])}列")
        # 显示样本数据
        for i, row in enumerate(backend_data[:2]):
            sample = row[:3] if len(row) > 3 else row
            print(f"    行{i}样本: {sample}")

    return backend_data


def fix_column_mismatch(data, expected_columns):
    """修复列数不匹配"""
    print(f"🔧 修复列数不匹配: {len(data[0])}列 -> {expected_columns}列")

    fixed_data = []
    for i, row in enumerate(data):
        if len(row) < expected_columns:
            # 补全缺失列
            fixed_row = row + [''] * (expected_columns - len(row))
            if i < 2:
                print(f"  ✅ 行{i}: 补全{expected_columns - len(row)}列")
        elif len(row) > expected_columns:
            # 截断多余列
            fixed_row = row[:expected_columns]
            if i < 2:
                print(f"  ✅ 行{i}: 截断{len(row) - expected_columns}列")
        else:
            fixed_row = row
        fixed_data.append(fixed_row)

    return fixed_data


def save_data_snapshot(pdf_id, excel_file, sheet_name, data, table_type):
    """保存数据快照"""
    from pathlib import Path
    import json
    import time

    try:
        # 🔥 使用默认路径
        SNAPSHOT_ROOT = "data/backend/static/modify_data"

        # 创建快照目录
        snap_dir = Path(MAIN_ROOT) / SNAPSHOT_ROOT / pdf_id
        snap_dir.mkdir(parents=True, exist_ok=True)

        # 生成时间戳文件名
        ts = int(time.time())
        snap_file = f"{pdf_id}_{sheet_name}_{table_type}_{ts}.json"
        snap_path = snap_dir / snap_file

        # 准备快照数据
        snapshot_data = {
            'pdf_id': pdf_id,
            'excel_file': excel_file,
            'sheet_name': sheet_name,
            'table_type': table_type,
            'data': data,  # 保存原始数据
            'saved_at': ts,
            'data_dimensions': {
                'rows': len(data),
                'columns': len(data[0]) if data else 0
            }
        }

        # 保存快照
        with open(snap_path, 'w', encoding='utf-8') as f:
            json.dump(snapshot_data, f, ensure_ascii=False, indent=2)

        print(f"✅ 数据快照已保存: {snap_path}")

        return {
            'success': True,
            'path': str(snap_path),
            'file': snap_file
        }

    except Exception as e:
        print(f"⚠️ 快照保存失败: {e}")
        return {
            'success': False,
            'error': str(e)
        }


@file_bp.route('/excel/save-flattened', methods=['POST', 'OPTIONS'])
def save_flattened_data():
    """保存扁平化数据到独立的Excel文件（通过文件名前缀映射）"""
    if request.method == 'OPTIONS':
        # CORS预检请求处理
        response = make_response()
        response.headers.add('Access-Control-Allow-Origin', '*')
        response.headers.add('Access-Control-Allow-Headers', 'Content-Type,Authorization')
        response.headers.add('Access-Control-Allow-Methods', 'GET,PUT,POST,DELETE,OPTIONS')
        return response

    data = request.json
    print("******************** 保存扁平化数据到独立文件 ******************")

    # 基础校验
    required_fields = ['pdf_id', 'excel_file', 'sheet_name', 'table_type', 'flattened_data']
    for field in required_fields:
        if field not in data:
            return jsonify({'error': f'缺少必要字段: {field}'}), 400

    try:
        pdf_id = data['pdf_id']
        original_excel_file = data['excel_file']  # 原Excel文件名
        sheet_name = data['sheet_name']
        table_type = data['table_type']
        table_data = data['flattened_data']

        print(f"💾 保存扁平化数据: PDF={pdf_id}, 原文件={original_excel_file}, Sheet={sheet_name}")

        # 通过固定前缀生成扁平化Excel文件名
        flattened_excel_file = generate_flattened_filename(original_excel_file)
        print(f"📁 扁平化文件名: {flattened_excel_file}")
        print("&&&&&&&&&&&&扁平化文件名&&&&&&&&&&&&&")
        print(pdf_id,
            original_excel_file,
            flattened_excel_file,
            sheet_name,)

        # 保存到独立文件
        result = save_to_flattened_excel(
            pdf_id,
            original_excel_file,
            flattened_excel_file,
            sheet_name,
            table_data
        )

        if not result['success']:
            return jsonify({'success': False, 'error': result['error']}), 500

        return jsonify({
            'success': True,
            'message': '扁平化数据保存成功',
            'flattened_file': flattened_excel_file,
            'original_file': original_excel_file,
            'file_created': result.get('file_created', False),
            'sheet_created': result.get('sheet_created', False),
            'saved_rows': result.get('saved_rows', 0),
            'saved_columns': result.get('saved_columns', 0),
            'data_dimensions': result.get('data_dimensions', '未知')
        }), 200

    except Exception as e:
        print(f"❌ 扁平化数据保存失败: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': f'扁平化数据保存失败: {str(e)}'}), 500


def generate_flattened_filename(original_excel_file):
    """通过固定前缀生成扁平化Excel文件名"""
    from pathlib import Path

    original_path = Path(original_excel_file)

    # 使用固定前缀 "flattened_" 来建立映射关系
    flattened_name = f"flattened_{original_path.name}"

    print(f"📝 文件名映射: {original_excel_file} -> {flattened_name}")
    return flattened_name


def save_to_flattened_excel(pdf_id, original_excel_file, flattened_excel_file, sheet_name, table_data):
    """保存数据到独立的扁平化Excel文件"""
    print("📊 保存到独立扁平化Excel文件...")

    try:
        from pathlib import Path
        import os

        # 使用相同的目录结构，通过文件名区分
        pdf_id = get_correct_pdf_id(pdf_id)
        excel_dir = Path(MAIN_ROOT) / EXCEL_OUTPUT_ROOT / pdf_id
        excel_dir.mkdir(parents=True, exist_ok=True)

        flattened_excel_path = excel_dir / flattened_excel_file
        print(f"📁 扁平化文件路径: {flattened_excel_path}")

        from openpyxl import Workbook, load_workbook

        file_exists = flattened_excel_path.exists()
        file_created = not file_exists
        sheet_created = False

        if file_exists:
            # 文件存在，加载现有工作簿
            workbook = load_workbook(flattened_excel_path)
            existing_sheets = workbook.sheetnames.copy()
            print(f"📄 加载现有扁平化文件，包含 {len(existing_sheets)} 个Sheet: {existing_sheets}")
        else:
            # 文件不存在，创建新工作簿
            workbook = Workbook()
            # 删除默认Sheet
            default_sheet = workbook.active
            workbook.remove(default_sheet)
            print("📄 创建新的扁平化Excel文件")

        # 处理目标Sheet
        if sheet_name in workbook.sheetnames:
            print(f"📋 Sheet已存在，覆盖: {sheet_name}")
            # 删除现有Sheet
            del workbook[sheet_name]
            sheet_created = False
        else:
            sheet_created = True

        # 创建/重写目标Sheet
        worksheet = workbook.create_sheet(sheet_name)
        print(f"✅ 创建/覆盖Sheet: {sheet_name}")

        # 转换前端数据格式
        backend_data = convert_frontend_to_backend_format(table_data)
        if not backend_data or len(backend_data) == 0:
            workbook.close()
            return {'success': False, 'error': '转换后数据为空'}

        print(f"📊 转换后数据: {len(backend_data)}行 × {len(backend_data[0])}列")

        # 写入完整数据到目标Sheet
        for row_idx, row_data in enumerate(backend_data, 1):
            for col_idx, cell_value in enumerate(row_data, 1):
                worksheet.cell(row=row_idx, column=col_idx, value=cell_value)

        print(f"📝 写入数据: {len(backend_data)}行 × {len(backend_data[0])}列")

        # 保存工作簿
        workbook.save(flattened_excel_path)
        workbook.close()

        # 验证原文件是否存在（可选，用于调试）
        original_excel_path = excel_dir / original_excel_file
        print(f"🔍 验证原文件存在: {original_excel_path.exists()}")

        print("✅ 扁平化文件保存成功")

        return {
            'success': True,
            'file_created': file_created,
            'sheet_created': sheet_created,
            'saved_rows': len(backend_data),
            'saved_columns': len(backend_data[0]) if backend_data else 0,
            'data_dimensions': f'{len(backend_data)}行 × {len(backend_data[0])}列',
            'flattened_path': str(flattened_excel_path),
            'original_path': str(original_excel_path)
        }

    except Exception as e:
        print(f"❌ 扁平化文件保存失败: {e}")
        import traceback
        traceback.print_exc()
        return {'success': False, 'error': f'扁平化文件保存失败: {str(e)}'}


def get_flattened_filename(original_excel_file):
    """根据原文件名获取对应的扁平化文件名"""
    from pathlib import Path

    original_path = Path(original_excel_file)
    return f"flattened_{original_path.name}"


@file_bp.route('/excel-data/<path:filename>')
def serve_excel_file(filename):
    """提供 Excel 文件下载"""
    # 修正：统一使用Path对象构建路径

    file_path = Path(MAIN_ROOT) / EXCEL_OUTPUT_ROOT / filename
    if not file_path.exists():
        return "文件不存在", 404
    return send_from_directory(file_path.parent, file_path.name)



# ---------- 7. Excel数据扁平化处理（支持Excel标准格式） ----------
@file_bp.route('/excel-flatten', methods=['POST', 'OPTIONS'])
def excel_flatten_from_excel():
    """
    处理从Excel直接提取的标准格式数据
    使用long_format_converter.py的逻辑进行扁平化转换
    """
    if request.method == 'OPTIONS':
        return jsonify({"status": "ok"}), 200

    try:
        # 检查转换器是否可用
        if not CONVERTER_AVAILABLE:
            print("❌ 转换器不可用")
            return jsonify({
                "success": False,
                "error": "数据转换器模块不可用"
            }), 500

        # 获取请求数据
        data = request.get_json()

        if not data:
            print("❌ 请求数据为空")
            return jsonify({
                "success": False,
                "error": "请求数据为空"
            }), 400

        table_data = data.get('table_data', [])
        source_info = data.get('source_info', {})

        if not table_data or len(table_data) < 2:
            print("❌ 表格数据至少需要表头行和一个数据行")
            return jsonify({
                "success": False,
                "error": "表格数据至少需要表头行和一个数据行"
            }), 400

        print(f"📊 开始处理Excel表格数据:")
        print(f"  - 原始表格尺寸: {len(table_data)}行 × {len(table_data[0]) if table_data else 0}列")

        # 🔥 修复1：直接使用原始数据，不自动添加表头
        processed_table_data = table_data
        print(f"📊 使用原始数据，不添加表头行")

        # 创建转换器实例
        converter = FinalDataConverter()

        print("source_infosource_infosource_info:", source_info)
        ori_table_metadata = data.get("table_metadata", {})
        print("table_metadata:", ori_table_metadata)
        print("data:", data)

        # 准备表格元数据
        table_metadata = {
            'name': ori_table_metadata.get('name', ''),
            'default_unit': ori_table_metadata.get('default_unit', ''),
            'default_currency': ori_table_metadata.get('default_currency', '人民币'),
            'default_report_period': ori_table_metadata.get('default_report_period', ''),
            'headers': {
                'rows': [],
                'cols': []
            }
        }

        print(f"📊 提取原始标记信息...")

        # 1. 找行标记列
        row_mark_col_index = -1
        if len(processed_table_data) > 0:
            first_row = processed_table_data[0]
            for j in range(len(first_row)):
                header = str(first_row[j]).strip() if first_row[j] else ""
                if header == "行标记":
                    row_mark_col_index = j
                    print(f"✅ '行标记'列位置: 第{j}列")
                    break

        # 2. 找列标记行
        col_mark_row_index = -1
        for i in range(len(processed_table_data)):
            if len(processed_table_data[i]) > 0:
                cell_value = str(processed_table_data[i][0]).strip() if processed_table_data[i][0] else ""
                if cell_value == "列标记":
                    col_mark_row_index = i
                    print(f"✅ '列标记'行位置: 第{i}行")
                    break

        # 3. 读取行标记
        row_marks = []
        if row_mark_col_index >= 0:
            print(f"🔍 读取行标记...")
            for i in range(1, len(processed_table_data)):
                if row_mark_col_index < len(processed_table_data[i]):
                    mark_value = processed_table_data[i][row_mark_col_index]
                    try:
                        if mark_value is None or mark_value == "":
                            row_mark = 1
                        else:
                            if isinstance(mark_value, (int, float)):
                                row_mark = int(mark_value)
                            else:
                                row_mark = int(float(mark_value)) if '.' in str(mark_value) else int(mark_value)
                    except:
                        row_mark = 1
                    row_marks.append(row_mark)
                else:
                    row_marks.append(1)
            print(f"✅ 读取完成: {len(row_marks)}个行标记")

        # 4. 读取列标记
        col_marks = []
        if col_mark_row_index >= 0:
            print(f"🔍 读取列标记...")
            col_mark_row = processed_table_data[col_mark_row_index]  # 🔥 修复变量名
            for j in range(1, len(col_mark_row)):
                mark_value = col_mark_row[j]
                try:
                    if mark_value is None or mark_value == "":
                        col_mark = 1
                    else:
                        if isinstance(mark_value, (int, float)):
                            col_mark = int(mark_value)
                        else:
                            col_mark = int(float(mark_value)) if '.' in str(mark_value) else int(mark_value)
                except:
                    col_mark = 1
                col_marks.append(col_mark)
            print(f"✅ 读取完成: {len(col_marks)}个列标记")

        marks_info = {
            'row_marks': row_marks,
            'col_marks': col_marks,
            'row_mark_col_index': row_mark_col_index,
            'col_mark_row_index': col_mark_row_index
        }

        # 执行转换
        print(f"🔄 开始转换表格数据...")

        long_format_data = converter.convert_table_to_long_format(
            table_data=processed_table_data,
            table_metadata=table_metadata,
            marks_info=marks_info,
            bank_name=source_info.get('bank_name', '中国建设银行'),
            entity=source_info.get('entity', '本集团')
        )

        print(f"✅ 转换完成: {len(long_format_data)} 条标准格式记录")

        # 将长格式数据转换为前端需要的双表头格式
        print(f"🔄 将长格式数据转换为前端双表头格式...")

        if not long_format_data or len(long_format_data) == 0:
            print("⚠️ 长格式数据为空，返回空结构")
            frontend_rows = []
            field_names = []
        else:
            # 提取所有字段名作为表头
            field_names = list(long_format_data[0].keys())
            print(f"📊 字段名（表头）: {field_names}")

            # 构建前端需要的rows数组
            frontend_rows = []

            # 添加元数据行
            metadata_row = {
                "__metadata": {
                    "has_dual_headers": True,
                    "top_left_cell": "",
                    "horizontal_headers": field_names,
                    "vertical_headers": []
                }
            }
            frontend_rows.append(metadata_row)

            # 添加表头行
            header_row = {
                "__is_first_row": True,
                "__top_left_cell": "字段名"
            }
            for i, field_name in enumerate(field_names, 1):
                header_row[f"H_{i}"] = field_name
            frontend_rows.append(header_row)

            # 添加数据行
            for record_idx, record in enumerate(long_format_data):
                data_row = {
                    "__is_data_row": True,
                    "__vertical_header": f"记录{record_idx + 1}"
                }

                for i, field_name in enumerate(field_names, 1):
                    value = record.get(field_name, "")
                    if field_name == '行标记' and isinstance(value, (int, float)):
                        data_row[f"H_{i}"] = value
                    else:
                        data_row[f"H_{i}"] = str(value) if value is not None else ""

                frontend_rows.append(data_row)

            print(f"✅ 转换完成: {len(frontend_rows)} 行前端格式数据")

        # 返回结果
        result = {
            "rows": frontend_rows,
            "total_rows": len(frontend_rows),
            "total_columns": len(field_names) if long_format_data else 0,
            "sheet_name": source_info.get('table_name', '扁平化数据'),
            "excel_file": source_info.get('excel_file', ''),
            "pdf_id": source_info.get('pdf_id', ''),
            "has_dual_headers": True,
            "success": True,
            "source_info": source_info,
            "timestamp": datetime.datetime.now().isoformat(),
            "stats": {
                "original_rows": len(table_data),
                "converted_records": len(long_format_data),
                "has_data": len(long_format_data) > 0
            }
        }

        print("✅ API处理完成，返回结果")
        return jsonify(result)

    except Exception as e:
        print(f"❌ Excel数据处理失败: {str(e)}")
        import traceback
        traceback.print_exc()

        return jsonify({
            "success": False,
            "error": f"处理失败: {str(e)}"
        }), 500




# 后端API示例（Python Flask）
# ========== 新增的Excel转PDF接口 ==========
@file_bp.route('/convert/excel-to-pdf', methods=['POST'])
def convert_excel_to_pdf_api():
    """Excel转PDF接口 - 新增功能"""
    try:
        # 1. 获取文件和参数
        if 'file' not in request.files:
            return jsonify({'error': '没有上传文件'}), 400

        file = request.files['file']
        pages = request.form.get('pages', 'all')
        orientation = request.form.get('orientation', 'portrait')
        dpi = int(request.form.get('dpi', 200))

        # 2. 验证文件
        validate_excel_file(file)

        # 3. 验证参数
        orientation, page_range = validate_conversion_params(orientation, pages)

        # 4. 保存上传的文件
        temp_filepath = save_uploaded_file(file)

        # 5. 执行转换
        output_filepath = convert_excel_to_pdf(
            temp_filepath,
            page_range=page_range,
            orientation=orientation,
            dpi=dpi
        )

        # 6. 返回结果
        return save_and_return_result(output_filepath)

    except ValidationError as e:
        return jsonify({'error': e.message}), e.status_code
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'转换失败: {str(e)}'}), 500


# 在 file.py 中添加一个通用的OPTIONS处理
@file_bp.route('/save-excel-modifications', methods=['OPTIONS'])
def handle_save_options():
    """处理CORS预检请求"""
    response = jsonify({'status': 'ok'})
    response.headers.add('Access-Control-Allow-Origin', 'http://localhost:8080')
    response.headers.add('Access-Control-Allow-Methods', 'POST, OPTIONS')
    response.headers.add('Access-Control-Allow-Headers', 'Content-Type')
    response.headers.add('Access-Control-Allow-Credentials', 'true')
    return response

@file_bp.route('/save-excel-modifications', methods=['POST'])
def save_excel_modifications():
    """保存Excel修改"""
    # 添加CORS头
    response = jsonify({'success': True, 'message': '保存成功'})
    response.headers.add('Access-Control-Allow-Origin', 'http://localhost:8080')
    response.headers.add('Access-Control-Allow-Credentials', 'true')
    return response



# 在Python中直接查看数据库内容
@file_bp.route('/debug-files')
def debug_files():
    """调试接口：查看数据库中的实际文件名"""
    conn = db.connect()
    c = conn.cursor()

    c.execute("""
        SELECT id, filename, raw_filename, file_type
        FROM files 
        WHERE deleted = 0 AND file_type = 'pdf'
        LIMIT 10
    """)

    rows = c.fetchall()
    conn.close()

    print("📋 数据库中的文件:")
    for row in rows:
        print(f"  ID: {row['id']}")
        print(f"  文件名: {row['filename']}")
        print(f"  原始名: {row['raw_filename']}")
        print(f"  类型: {row['file_type']}")
        print("  ---")

    return jsonify({
        "files": [dict(row) for row in rows]
    })



@file_bp.route('/debug-table-data')
def debug_table_data():
    """调试接口：查看表数据"""
    try:
        table_name = request.args.get('table', 'table_processing_records')
        limit = request.args.get('limit', 10, type=int)

        db_manager = SafeDatabaseManager()
        db_manager._connect()

        # 获取表数据
        db_manager.cursor.execute(f"SELECT * FROM {table_name} LIMIT ?", (limit,))
        rows = db_manager.cursor.fetchall()

        # 获取表结构
        db_manager.cursor.execute(f"PRAGMA table_info({table_name})")
        columns = [dict(row) for row in db_manager.cursor.fetchall()]

        db_manager._close()

        return jsonify({
            "table": table_name,
            "columns": columns,
            "data": [dict(row) for row in rows],
            "count": len(rows)
        })

    except Exception as e:
        return jsonify({"error": str(e)})


@file_bp.route('/debug-search-query')
def debug_search_query():
    """调试搜索查询"""
    try:
        keyword = request.args.get('keyword', '建设银行')

        db_manager = SafeDatabaseManager()
        db_manager._connect()

        # 执行搜索查询
        query = """
            SELECT * FROM table_processing_records 
            WHERE bank_name LIKE ? OR pdf_folder LIKE ? 
            ORDER BY created_at DESC LIMIT 10
        """
        params = (f'%{keyword}%', f'%{keyword}%')

        print(f"🔍🔍 调试搜索查询:")
        print(f"SQL: {query}")
        print(f"参数: {params}")

        db_manager.cursor.execute(query, params)
        rows = db_manager.cursor.fetchall()

        db_manager._close()

        return jsonify({
            "query": query,
            "params": params,
            "results": [dict(row) for row in rows],
            "count": len(rows)
        })

    except Exception as e:
        return jsonify({"error": str(e)})



@file_bp.route('/debug-search-complete')
def debug_search_complete():
    """完整调试搜索流程"""
    try:
        keyword = request.args.get('keyword', '建设银行')

        print(f"🔍🔍 完整调试搜索流程: 关键词='{keyword}'")

        # 1. 直接执行SQL查询
        db_manager = SafeDatabaseManager()
        db_manager._connect()

        query = "SELECT * FROM table_processing_records WHERE bank_name LIKE ? LIMIT 5"
        params = (f'%{keyword}%',)

        db_manager.cursor.execute(query, params)
        rows = db_manager.cursor.fetchall()

        print(f"  📊 直接SQL查询结果: {len(rows)} 条")

        # 2. 转换为字典
        dict_results = []
        for i, row in enumerate(rows):
            row_dict = dict(row)
            dict_results.append(row_dict)
            print(f"    第{i + 1}条: {row_dict.get('bank_name', '无银行名')}")

        db_manager._close()

        # 3. 测试SafeDatabaseManager方法
        manager_results = db_manager.search_pdf_files(keyword, 5)
        print(f"  📊 SafeDatabaseManager结果: {len(manager_results)} 条")

        return jsonify({
            "direct_sql_count": len(rows),
            "direct_sql_results": dict_results,
            "manager_results_count": len(manager_results),
            "manager_results_sample": manager_results[:2] if manager_results else []
        })

    except Exception as e:
        return jsonify({"error": str(e)})



@file_bp.route('/search-pdf-compatible1')
def search_pdf_compatible1():
    """搜索PDF文件 - 带完整调试信息"""
    import traceback

    print("\n" + "=" * 60)
    print("🔥🔥🔥 /api/search-pdf-compatible 被调用")
    print("=" * 60)

    try:
        keyword = request.args.get('keyword', '').strip()
        limit = request.args.get('limit', 100, type=int)

        print(f"🔍 参数: keyword='{keyword}', limit={limit}")
        print(f"🔍 完整URL: {request.url}")

        # 🔥 1. 先硬编码返回测试数据
        print("🧪 阶段1: 硬编码测试数据")
        test_data = {
            "files": [
                {
                    "id": "hardcoded-test-1",
                    "file_id": "hardcoded-1",
                    "disk_name": "hardcoded.pdf",
                    "file_type": "pdf",
                    "filename": "硬编码测试文件.pdf",
                    "name": "硬编码测试文件.pdf",
                    "matchType": "硬编码测试"
                }
            ],
            "count": 1
        }

        print("✅ 硬编码数据:", test_data)
        return jsonify(test_data)

    except Exception as e:
        print(f"❌❌❌ 严重错误: {e}")
        traceback.print_exc()
        return jsonify({
            "files": [],
            "count": 0,
            "error": str(e)
        })

print("🔥🔥🔥🔥🔥 file.py 文件被加载了！")



@file_bp.route('/search-pdf-compatible-test')
def search_pdf_compatible_test():
    """测试函数 - 验证路由是否工作"""
    print("🔥🔥🔥🔥🔥 search_pdf_compatible_test 函数被调用了！")

    return jsonify({
        "message": "测试函数正常工作",
        "count": 1
    })


@file_bp.route('/search-pdf-compatible')
def search_pdf_compatible():
    """搜索PDF文件 - 实际搜索数据库版本"""
    print("🔥🔥🔥🔥🔥 search_pdf_compatible 函数被调用了！")

    try:
        keyword = request.args.get('keyword', '').strip()
        limit = request.args.get('limit', 100, type=int)

        print(f"🔍 搜索关键词: '{keyword}'")

        # 连接数据库
        db_manager = SafeDatabaseManager()

        # 使用 search_pdf_files 方法
        results = db_manager.search_pdf_files(keyword, limit)

        print(f"📊 数据库返回 {len(results)} 条结果")

        if not results:
            return jsonify({
                "files": [],
                "count": 0
            })

        # 转换为前端需要的格式
        files = []
        for row in results:
            # 确保 row 是字典
            if not isinstance(row, dict):
                row = dict(row)

            # 构建文件信息
            file_info = {
                "id": str(row.get("id", "")),
                "file_id": str(row.get("pdf_folder", "")),  # 重复id字段确保兼容
                "disk_name": row.get("pdf_folder", ""),
                "file_type": "pdf",
                "filename": row.get("bank_name", "未知银行"),
                "name": row.get("bank_name", "未知银行"),
                "matchType": "数据库匹配",
                "status": row.get("status", ""),
                "created_at": row.get("created_at", ""),
                "raw_filename": row.get("bank_name", "未知银行"),
            }

            files.append(file_info)

        print(f"✅ 转换完成，返回 {len(files)} 个文件")

        return jsonify({
            "files": files,
            "count": len(files)
        })

    except Exception as e:
        print(f"❌❌ 搜索失败: {e}")
        import traceback
        traceback.print_exc()

        # 出错时返回空数组
        return jsonify({
            "files": [],
            "count": 0
        })


