"""
搜索和保存服务模块 - 处理搜索和保存相关操作
"""

from pathlib import Path
from backend.models.safe_unified_db import SafeDatabaseManager
from backend.models.unified_db import DatabaseManager as OldDatabaseManager
from backend.utils.constants import MAIN_ROOT, EXCEL_OUTPUT_ROOT, DATABASE
from backend.service.file_mapping_service import file_mapping_service


db = OldDatabaseManager(DATABASE)
safe_db = SafeDatabaseManager()


class SearchService:
    """搜索服务类"""

    @staticmethod
    def search_pdf_files(keyword):
        """搜索PDF文件"""
        print(f"搜索关键词: '{keyword}'")

        if not keyword:
            return {"files": []}

        try:
            search_results = file_mapping_service.search_files(keyword, 'pdf')
            return {"files": search_results}
        except Exception as e:
            print(f"搜索PDF失败: {e}")
            return {"error": "搜索失败"}, 500

    @staticmethod
    def search_pdf_compatible(keyword, limit=100):
        """兼容性搜索PDF文件"""
        print(f"🔍🔍 搜索关键词: '{keyword}'")

        try:
            results = safe_db.search_pdf_files(keyword, limit)
            print(f"📊📊 数据库返回 {len(results)} 条结果")

            if not results:
                return {"files": [], "count": 0}

            # 转换为前端需要的格式
            files = []
            for row in results:
                if not isinstance(row, dict):
                    row = dict(row)

                file_info = {
                    "id": str(row.get("id", "")),
                    "file_id": str(row.get("pdf_folder", "")),
                    "disk_name": row.get("pdf_folder", ""),
                    "file_type": "pdf",
                    "filename": row.get("bank_name", "未知银行"),
                    "name": row.get("bank_name", "未知银行"),
                    "matchType": "数据库匹配",
                    "status": row.get("status", ""),
                    "created_at": row.get("created_at", ""),
                    "raw_filename": row.get("bank_name", "未知银行"),
                }
                files.append(file_info)

            return {"files": files, "count": len(files)}

        except Exception as e:
            print(f"❌❌❌❌ 搜索失败: {e}")
            return {"files": [], "count": 0}


class SaveService:
    """保存服务类"""

    @staticmethod
    def save_final_excel(data):
        """保存Excel数据"""
        required_fields = ['pdf_id', 'excel_file', 'sheet_name', 'table_type', 'data']
        for field in required_fields:
            if field not in data:
                return {'error': f'缺少必要字段: {field}'}, 400

        try:
            pdf_id = data['pdf_id']
            excel_file = data['excel_file']
            sheet_name = data['sheet_name']
            table_type = data['table_type']
            table_data = data['data']

            print(f"💾💾 保存数据: PDF={pdf_id}, 文件={excel_file}, Sheet={sheet_name}, 类型={table_type}")

            # 根据表类型选择保存方式
            if table_type == 'original':
                result = SaveService._save_complete_table_data(pdf_id, excel_file, sheet_name, table_data, table_type)
            elif table_type == 'flattened':
                result = SaveService._save_flattened_table_data(pdf_id, excel_file, sheet_name, table_data, table_type)
            else:
                return {'error': f'不支持的表类型: {table_type}'}, 400

            if not result['success']:
                return {'success': False, 'error': result['error']}, 500

            return {
                'success': True,
                'message': '表格数据保存成功',
                'saved_count': result.get('saved_rows', 0),
                'data_dimensions': result.get('data_dimensions', '未知'),
                'excel_updated': result.get('excel_updated', False),
                'sheets_protected': result.get('sheets_protected', False),
                'protected_sheets_count': result.get('protected_sheets_count', 0),
                'file_created': result.get('file_created', False)
            }

        except Exception as e:
            print(f"❌❌ 保存失败: {e}")
            return {'success': False, 'error': f'保存失败: {str(e)}'}, 500

    @staticmethod
    def _save_complete_table_data(pdf_id, excel_file, sheet_name, table_data, table_type):
        """保存完整表格数据"""
        try:
            from openpyxl import load_workbook

            # 获取正确的PDF ID
            if pdf_id.isdigit():
                conn = db.connect()
                c = conn.cursor()
                c.execute("SELECT filename FROM files WHERE id = ? AND deleted = 0", (pdf_id,))
                row = c.fetchone()
                conn.close()
                real_pdf_id = row["filename"] if row else pdf_id
            else:
                real_pdf_id = pdf_id

            # 构建文件路径
            excel_dir = Path(MAIN_ROOT) / EXCEL_OUTPUT_ROOT / real_pdf_id
            excel_path = excel_dir / excel_file

            if not excel_path.exists():
                return {'success': False, 'error': f'Excel文件不存在: {excel_path}'}

            # 加载工作簿
            workbook = load_workbook(excel_path)
            if sheet_name not in workbook.sheetnames:
                workbook.close()
                return {'success': False, 'error': f'Sheet不存在: {sheet_name}'}

            worksheet = workbook[sheet_name]

            # 清空数据
            if worksheet.max_row > 0:
                worksheet.delete_rows(1, worksheet.max_row)

            # 写入新数据
            for row_idx, row_data in enumerate(table_data, 1):
                for col_idx, cell_value in enumerate(row_data, 1):
                    worksheet.cell(row=row_idx, column=col_idx, value=cell_value)

            workbook.save(excel_path)
            workbook.close()

            return {
                'success': True,
                'saved_rows': len(table_data),
                'saved_columns': len(table_data[0]) if table_data else 0,
                'excel_updated': True
            }

        except Exception as e:
            print(f"❌❌ 完整表格保存失败: {e}")
            return {'success': False, 'error': f'完整表格保存失败: {str(e)}'}

    @staticmethod
    def _save_flattened_table_data(pdf_id, excel_file, sheet_name, table_data, table_type):
        """保存扁平化表格数据"""
        try:
            from openpyxl import Workbook, load_workbook

            # 获取正确的PDF ID
            if pdf_id.isdigit():
                conn = db.connect()
                c = conn.cursor()
                c.execute("SELECT filename FROM files WHERE id = ? AND deleted = 0", (pdf_id,))
                row = c.fetchone()
                conn.close()
                real_pdf_id = row["filename"] if row else pdf_id
            else:
                real_pdf_id = pdf_id

            # 构建文件路径
            excel_dir = Path(MAIN_ROOT) / EXCEL_OUTPUT_ROOT / real_pdf_id
            excel_dir.mkdir(parents=True, exist_ok=True)
            excel_path = excel_dir / excel_file

            file_exists = excel_path.exists()

            if file_exists:
                workbook = load_workbook(excel_path)
            else:
                workbook = Workbook()
                default_sheet = workbook.active
                workbook.remove(default_sheet)

            # 处理目标Sheet
            if sheet_name in workbook.sheetnames:
                del workbook[sheet_name]

            worksheet = workbook.create_sheet(sheet_name)

            # 写入数据
            for row_idx, row_data in enumerate(table_data, 1):
                for col_idx, cell_value in enumerate(row_data, 1):
                    worksheet.cell(row=row_idx, column=col_idx, value=cell_value)

            workbook.save(excel_path)
            workbook.close()

            return {
                'success': True,
                'file_created': not file_exists,
                'saved_rows': len(table_data),
                'saved_columns': len(table_data[0]) if table_data else 0,
                'excel_updated': True
            }

        except Exception as e:
            print(f"❌❌ 扁平化数据保存失败: {e}")
            return {'success': False, 'error': f'扁平化数据保存失败: {str(e)}'}