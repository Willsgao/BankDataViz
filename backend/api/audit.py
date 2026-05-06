# -*- coding:utf-8 -*-
"""
会计勾稽验证 API

提供档案列表、规则配置、运行校验、查询结果等接口。
"""
import os
import json
import sqlite3
import uuid
from datetime import datetime
from flask import Blueprint, request, jsonify

# =============================================================================
# 蓝图与常量
# =============================================================================
audit_bp = Blueprint('audit', __name__, url_prefix='/api/audit')

# 数据库路径（与现有系统保持一致）
DATABASE_PATH = r'F:\wills\codes\DocuVista\data\database.db'
EXCEL_DATA_ROOT = r'F:\wills\codes\DocuVista\data\backend\static\excel_data'
RULES_CONFIG_PATH = r'F:\wills\codes\DocuVista\data\backend\config\audit_rules.json'


# =============================================================================
# 数据库工具
# =============================================================================
def get_db():
    conn = sqlite3.connect(DATABASE_PATH, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    return conn


def row_to_dict(row):
    return dict(row) if row else None


# =============================================================================
# 工具函数
# =============================================================================
def get_excel_path_for_file(file_id: str) -> str:
    """根据 file_id 找到对应的合并 Excel 路径"""
    folder = os.path.join(EXCEL_DATA_ROOT, file_id)
    if not os.path.isdir(folder):
        return None
    for fname in os.listdir(folder):
        if fname.endswith('.xlsx') and '_合并' in fname:
            return os.path.join(folder, fname)
    # 兜底：取第一个 xlsx
    for fname in os.listdir(folder):
        if fname.endswith('.xlsx'):
            return os.path.join(folder, fname)
    return None


def load_rules_from_config(enabled_only: bool = False) -> list:
    """加载勾稽规则配置"""
    if not os.path.exists(RULES_CONFIG_PATH):
        return []
    with open(RULES_CONFIG_PATH, 'r', encoding='utf-8') as f:
        rules = json.load(f)
    if enabled_only:
        rules = [r for r in rules if r.get('enabled', True)]
    return rules


# =============================================================================
# 路由定义
# =============================================================================

@audit_bp.route('/rules', methods=['GET'])
def get_rules():
    """获取所有勾稽规则配置"""
    enabled_only = request.args.get('enabled_only', 'false').lower() == 'true'
    rules = load_rules_from_config(enabled_only=enabled_only)
    return jsonify({
        'success': True,
        'rules': rules,
        'total': len(rules)
    })


@audit_bp.route('/files', methods=['GET'])
def list_files():
    """获取已解析的可校验档案列表（从 reports 表）"""
    conn = get_db()
    cur = conn.cursor()
    try:
        # 从 reports 表获取有 Excel 的档案
        cur.execute("""
            SELECT r.id, r.bank_id, r.report_type, r.period,
                   r.fiscal_year, r.pdf_filename,
                   r.excel_output_path, r.status,
                   b.bank_name
            FROM reports r
            LEFT JOIN banks b ON r.bank_id = b.id
            WHERE r.excel_output_path IS NOT NULL AND r.excel_output_path != ''
            ORDER BY r.created_at DESC
        """)
        rows = cur.fetchall()

        files = []
        for row in rows:
            r = dict(row)
            # 从 excel_output_path 提取 folder_id（UUID）
            excel_path = r.get('excel_output_path', '')
            folder_id = ''
            if excel_path:
                parts = excel_path.replace('\\\\', '/').split('/')
                for i, p in enumerate(parts):
                    if p.endswith('.xlsx'):
                        folder_id = parts[i - 1] if i > 0 else ''
                        break
                if not folder_id and '_合并' in excel_path:
                    folder_id = excel_path.split('_合并')[0].replace('\\\\', '').replace('\\', '').split('/')[-1]

            files.append({
                'id': folder_id or str(r['id']),
                'report_id': r['id'],
                'bank_name': r.get('bank_name', '未知银行'),
                'report_type': r.get('report_type', ''),
                'period': r.get('period', r.get('fiscal_year', '')),
                'excel_path': excel_path,
                'status': r.get('status', ''),
                'pdf_filename': r.get('pdf_filename', '')
            })
        return jsonify({'success': True, 'files': files, 'total': len(files)})
    finally:
        conn.close()


@audit_bp.route('/files/available', methods=['GET'])
def list_available_files():
    """获取目录下所有有合并 Excel 的档案（兜底方案）"""
    folders = []
    if not os.path.isdir(EXCEL_DATA_ROOT):
        return jsonify({'success': True, 'files': [], 'total': 0})

    for folder_name in os.listdir(EXCEL_DATA_ROOT):
        folder_path = os.path.join(EXCEL_DATA_ROOT, folder_name)
        if not os.path.isdir(folder_path):
            continue
        excel_file = None
        for fname in os.listdir(folder_path):
            if fname.endswith('.xlsx') and '_合并' in fname:
                excel_file = fname
                break
        if excel_file:
            # 获取最近一次校验结果
            conn = get_db()
            cur = conn.cursor()
            cur.execute("""
                SELECT status, pass_count, warn_count, fail_count, completed_at
                FROM audit_runs
                WHERE file_id = ?
                ORDER BY created_at DESC LIMIT 1
            """, (folder_name,))
            last_run = cur.fetchone()
            conn.close()

            files_count = len([f for f in os.listdir(folder_path) if f.endswith('.xlsx')])
            folders.append({
                'id': folder_name,
                'name': excel_file,
                'sheet_count': files_count,
                'excel_path': os.path.join(folder_path, excel_file),
                'last_status': dict(last_run).get('status') if last_run else None,
                'last_pass': dict(last_run).get('pass_count', 0) if last_run else 0,
                'last_warn': dict(last_run).get('warn_count', 0) if last_run else 0,
                'last_fail': dict(last_run).get('fail_count', 0) if last_run else 0,
            })

    return jsonify({'success': True, 'files': folders, 'total': len(folders)})


@audit_bp.route('/sheets/<file_id>', methods=['GET'])
def get_sheet_info(file_id: str):
    """获取指定档案的所有 Sheet 摘要信息（表头预览）"""
    excel_path = get_excel_path_for_file(file_id)
    if not excel_path or not os.path.exists(excel_path):
        return jsonify({'success': False, 'error': f'Excel not found: {file_id}'}), 404

    import openpyxl
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        return jsonify({'success': False, 'error': f'Cannot open Excel: {str(e)}'}), 500

    sheets = []
    for ws in wb.worksheets:
        # 横向表头预览：前3行 x 前8列
        row_headers = []
        for row_idx in range(1, 4):
            row_data = []
            for col_idx in range(1, min(ws.max_column + 1, 10)):
                v = ws.cell(row_idx, col_idx).value
                # 清理并截断
                s = str(v) if v is not None else ''
                s = s.replace('\n', ' ').replace('\r', '')
                if len(s) > 30:
                    s = s[:30] + '...'
                row_data.append(s)
            row_headers.append(row_data)

        # 纵向表头预览：前3列 x 前6行
        col_previews = []
        for col_idx in range(1, 4):
            col_data = []
            for row_idx in range(1, min(ws.max_row + 1, 8)):
                v = ws.cell(row_idx, col_idx).value
                s = str(v) if v is not None else ''
                s = s.replace('\n', ' ').replace('\r', '')
                if len(s) > 20:
                    s = s[:20] + '...'
                col_data.append(s)
            col_previews.append(col_data)

        sheets.append({
            'name': ws.title,
            'max_row': ws.max_row,
            'max_col': ws.max_column,
            'row_headers': row_headers,      # 横向表头预览
            'col_previews': col_previews    # 纵向表头预览
        })

    wb.close()
    return jsonify({
        'success': True,
        'file_id': file_id,
        'excel_path': excel_path,
        'sheets': sheets,
        'total': len(sheets)
    })


@audit_bp.route('/mapping/suggest', methods=['POST'])
def suggest_mapping():
    """根据 Sheet 摘要信息，自动推荐规则→Sheet 映射"""
    data = request.get_json() or {}
    sheets = data.get('sheets', [])  # [{name, row_headers, col_previews}, ...]
    rules = load_rules_from_config(enabled_only=True)

    # 规则关键词（从规则配置中提取）
    suggestions = []
    for rule in rules:
        rule_id = rule.get('id', '')
        rule_name = rule.get('name', '')
        # 用规则名称 + description + sheet_hint 中的关键词进行匹配
        keywords = _extract_keywords(rule)

        best_match = None
        best_score = 0

        for sheet in sheets:
            score = _score_sheet(sheet, keywords)
            if score > best_score:
                best_score = score
                best_match = {
                    'sheet_name': sheet['name'],
                    'score': score
                }

        suggestions.append({
            'rule_id': rule_id,
            'rule_name': rule_name,
            'suggested_sheet': best_match['sheet_name'] if best_match else None,
            'score': best_score
        })

    # 同时生成 sheet→rules 的映射
    sheet_rules = {s['name']: [] for s in sheets}
    for sug in suggestions:
        if sug['suggested_sheet']:
            sheet_rules[sug['suggested_sheet']].append({
                'rule_id': sug['rule_id'],
                'rule_name': sug['rule_name'],
                'score': sug['score']
            })

    return jsonify({
        'success': True,
        'rule_suggestions': suggestions,   # 每条规则推荐的 sheet
        'sheet_rules': sheet_rules          # 每个 sheet 对应的规则
    })


def _extract_keywords(rule: dict) -> list[str]:
    """从规则配置中提取关键词"""
    keywords = []
    # 从规则名称提取
    name = rule.get('name', '')
    # 去除勾稽、校验、校验等后缀
    for kw in name.replace('勾稽', '').replace('校验', '').split('、'):
        kw = kw.strip()
        if kw:
            keywords.append(kw)
    # 从 description 提取
    desc = rule.get('description', '')
    # 提取括号内容或关键指标名
    import re
    # 匹配中文关键词（2字以上）
    for m in re.findall(r'[\u4e00-\u9fa5]{2,10}', desc):
        if m not in ['分项', '目标', '计算', '报告', '数值']:
            keywords.append(m)
    return list(set(keywords))


def _score_sheet(sheet: dict, keywords: list[str]) -> int:
    """计算 Sheet 与关键词的匹配得分
    
    支持多种字段名格式（兼容前端发送的数据结构）
    
    header_preview: List[List[str]] - 每行是一个字符串列表
    row_preview: List[Dict[str, str]] - 每行是 {"col_1": "值1", ...}
    """
    score = 0
    sheet_name = sheet.get('name', '').lower()
    
    # 收集所有表头内容
    all_headers = []
    
    # header_preview: List[List[str]]
    for row in sheet.get('header_preview', []):
        if isinstance(row, list):
            for cell in row:
                if cell:
                    all_headers.append(str(cell).lower())
    
    # row_preview: List[Dict[str, str]]
    for row in sheet.get('row_preview', []):
        if isinstance(row, dict):
            for val in row.values():
                if val:
                    all_headers.append(str(val).lower())
    
    for kw in keywords:
        kw_lower = kw.lower()
        # Sheet 名匹配，权重最高
        if kw_lower in sheet_name:
            score += 10
        # 表头内容匹配
        for header in all_headers:
            if kw_lower in header:
                score += 2
    return score


@audit_bp.route('/run', methods=['POST'])
def run_audit():
    """执行勾稽校验（支持 sheet→规则映射）"""
    data = request.get_json() or {}
    file_id = data.get('file_id', '')
    file_name = data.get('file_name', '')
    rule_ids = data.get('rule_ids', [])
    sheet_mapping = data.get('sheet_mapping', {})  # {rule_id: sheet_name}

    if not file_id:
        return jsonify({'success': False, 'error': '缺少 file_id'}), 400

    # 找到 Excel 路径
    excel_path = get_excel_path_for_file(file_id)
    if not excel_path or not os.path.exists(excel_path):
        return jsonify({'success': False, 'error': f'Excel 文件未找到 (file_id={file_id})'}), 404

    # 生成 run_uuid
    run_uuid = str(uuid.uuid4())[:12]
    started_at = datetime.now().isoformat()

    # 保存 run 记录（running 状态）
    conn = get_db()
    cur = conn.cursor()
    try:
        cur.execute("""
            INSERT INTO audit_runs
                (run_uuid, file_id, file_name, started_at, status, total_rules, created_at)
            VALUES (?, ?, ?, ?, 'running', 0, ?)
        """, (run_uuid, file_id, file_name, started_at, started_at))
        run_db_id = cur.lastrowid

        # 执行规则引擎（传入 sheet_mapping）
        from backend.services.audit_engine import run_audit as engine_run
        audit_result = engine_run(
            file_id, file_name, excel_path,
            rule_ids=rule_ids,
            sheet_mapping=sheet_mapping
        )
        results = audit_result['results']
        total = len(results)
        pass_count = audit_result['pass_count']
        warn_count = audit_result['warn_count']
        fail_count = audit_result['fail_count']

        # 更新 run 记录
        cur.execute("""
            UPDATE audit_runs
            SET status = 'completed',
                completed_at = ?,
                total_rules = ?,
                pass_count = ?,
                warn_count = ?,
                fail_count = ?
            WHERE id = ?
        """, (audit_result['completed_at'], total, pass_count, warn_count, fail_count, run_db_id))

        # 写入详情
        for r in results:
            cur.execute("""
                INSERT INTO audit_details
                    (run_id, rule_id, rule_name, sheet_name, period, status,
                     actual_value, expected_value, diff, diff_percent, detail, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                run_db_id,
                r['rule_id'], r['rule_name'], r.get('sheet_name'), r.get('period'),
                r['status'],
                r.get('actual_value'), r.get('expected_value'),
                r.get('diff'), r.get('diff_percent'),
                r.get('detail', ''),
                started_at
            ))

        conn.commit()

        return jsonify({
            'success': True,
            'run_id': run_uuid,
            'run_db_id': run_db_id,
            'file_id': file_id,
            'file_name': file_name,
            'total': total,
            'pass_count': pass_count,
            'warn_count': warn_count,
            'fail_count': fail_count,
            'results': results
        })

    except Exception as e:
        import traceback
        conn.rollback()
        return jsonify({'success': False, 'error': str(e), 'detail': traceback.format_exc()}), 500
    finally:
        conn.close()


@audit_bp.route('/run-dal', methods=['POST'])
def run_audit_dal():
    """
    执行勾稽校验 - DAL模式
    
    使用数据访问层（DAL）读取数据，支持从Excel或未来从数据库读取。
    支持字段映射（人工确认字段对应关系）。
    """
    data = request.get_json() or {}
    file_id = data.get('file_id', '')
    file_name = data.get('file_name', '')
    rule_ids = data.get('rule_ids', [])
    sheet_mapping = data.get('sheet_mapping', {})  # {rule_id: sheet_name}
    field_mappings = data.get('field_mappings', {})  # {rule_id: {field: mapped_field}}

    if not file_id:
        return jsonify({'success': False, 'error': 'missing file_id'}), 400

    # 生成 run_uuid
    run_uuid = str(uuid.uuid4())[:12]
    started_at = datetime.now().isoformat()

    # 使用 DAL 模式
    try:
        from backend.services.dal import ExcelDataSource
        ds = ExcelDataSource()
        
        # 检查文件是否存在
        try:
            excel_path = ds.get_excel_path(file_id)
        except FileNotFoundError:
            return jsonify({'success': False, 'error': f'File not found: {file_id}'}), 404

        # 执行规则引擎（DAL模式）
        from backend.services.audit_engine import run_audit as engine_run
        audit_result = engine_run(
            file_id=file_id,
            file_name=file_name,
            rule_ids=rule_ids,
            sheet_mapping=sheet_mapping,
            data_source=ds
        )
        
        results = audit_result['results']
        total = len(results)
        pass_count = audit_result['pass_count']
        warn_count = audit_result['warn_count']
        fail_count = audit_result['fail_count']
        completed_at = audit_result['completed_at']

        return jsonify({
            'success': True,
            'run_id': run_uuid,
            'file_id': file_id,
            'file_name': file_name,
            'total': total,
            'pass_count': pass_count,
            'warn_count': warn_count,
            'fail_count': fail_count,
            'results': results,
            'started_at': started_at,
            'completed_at': completed_at
        })

    except Exception as e:
        import traceback
        return jsonify({
            'success': False, 
            'error': str(e),
            'detail': traceback.format_exc()
        }), 500


@audit_bp.route('/files/dal', methods=['GET'])
def list_files_dal():
    """
    获取档案列表 - DAL模式
    
    使用数据访问层获取档案列表，格式与原有接口兼容。
    """
    try:
        from backend.services.dal import ExcelDataSource
        ds = ExcelDataSource()
        files = ds.get_file_list()
        
        result = []
        for f in files:
            result.append({
                'id': f.id,
                'name': f.name,
                'source_type': f.source_type,
                'created_at': f.created_at,
                'status': f.status
            })
        
        return jsonify({'success': True, 'files': result, 'total': len(result)})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@audit_bp.route('/sheets/dal/<file_id>', methods=['GET'])
def get_sheets_dal(file_id: str):
    """
    获取档案的所有Sheet - DAL模式
    
    返回每个Sheet的摘要信息，用于前端预览和字段映射确认。
    """
    try:
        from backend.services.dal import ExcelDataSource
        ds = ExcelDataSource()
        
        sheets = ds.get_sheet_names(file_id)
        sheet_info_list = []
        
        for sheet_name in sheets:
            summary = ds.get_sheet_summary(file_id, sheet_name)
            sheet_info_list.append({
                'name': summary.name,
                'row_count': summary.row_count,
                'col_count': summary.col_count,
                'header_preview': summary.header_preview,
                'row_preview': summary.row_preview
            })
        
        return jsonify({
            'success': True,
            'file_id': file_id,
            'sheets': sheet_info_list,
            'total': len(sheet_info_list)
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@audit_bp.route('/fields/analyze', methods=['POST'])
def analyze_fields():
    """
    分析Sheet中的字段，提供字段映射建议
    
    根据规则配置，分析Sheet中的字段，给出哪些字段可能需要人工确认。
    """
    data = request.get_json() or {}
    file_id = data.get('file_id', '')
    sheet_name = data.get('sheet_name', '')
    rules = data.get('rules', [])  # 规则配置

    if not file_id or not sheet_name:
        return jsonify({'success': False, 'error': 'missing parameters'}), 400

    try:
        from backend.services.dal import ExcelDataSource
        ds = ExcelDataSource()
        
        sheet_data = ds.get_sheet_data(file_id, sheet_name)
        
        # 分析每个规则需要的字段
        field_analysis = []
        for rule in rules:
            rule_id = rule.get('id', '')
            rule_name = rule.get('name', '')
            
            # 提取规则需要的字段
            needed_fields = []
            if 'formula' in rule:
                formula = rule['formula']
                for part in ['numerator', 'denominator', 'result_field']:
                    f = formula.get(part, {})
                    if isinstance(f, dict):
                        needed_fields.append({
                            'role': part,
                            'field': f.get('field', ''),
                            'col_hint': f.get('col_hint', '')
                        })
                    elif f:
                        needed_fields.append({
                            'role': part,
                            'field': str(f),
                            'col_hint': ''
                        })
            
            # 在Sheet中查找匹配
            field_matches = []
            for nf in needed_fields:
                field_name = nf['field']
                matched = False
                for row in sheet_data.rows[:10]:  # 只检查前10行
                    for col_key, val in row.items():
                        if val and field_name in str(val):
                            matched = True
                            break
                field_matches.append({
                    'role': nf['role'],
                    'field': field_name,
                    'col_hint': nf['col_hint'],
                    'found_in_sheet': matched,
                    'confidence': 'high' if matched else 'low'
                })
            
            # 判断是否需要人工确认
            needs_confirmation = any(m['confidence'] == 'low' for m in field_matches)
            
            field_analysis.append({
                'rule_id': rule_id,
                'rule_name': rule_name,
                'fields': field_matches,
                'needs_confirmation': needs_confirmation
            })
        
        # 找出需要确认的字段
        uncertain_fields = []
        for fa in field_analysis:
            for f in fa['fields']:
                if f['confidence'] == 'low':
                    uncertain_fields.append({
                        'rule_id': fa['rule_id'],
                        'rule_name': fa['rule_name'],
                        'field': f['field'],
                        'role': f['role']
                    })
        
        return jsonify({
            'success': True,
            'file_id': file_id,
            'sheet_name': sheet_name,
            'field_analysis': field_analysis,
            'uncertain_fields': uncertain_fields,
            'total_uncertain': len(uncertain_fields)
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@audit_bp.route('/results/<run_uuid>', methods=['GET'])
def get_audit_result(run_uuid: str):
    """根据 run_uuid 查询校验结果"""
    conn = get_db()
    cur = conn.cursor()
    try:
        cur.execute("SELECT * FROM audit_runs WHERE run_uuid = ?", (run_uuid,))
        run = cur.fetchone()
        if not run:
            return jsonify({'success': False, 'error': 'run_uuid 不存在'}), 404

        cur.execute("SELECT * FROM audit_details WHERE run_id = ? ORDER BY id", (dict(run)['id'],))
        details = [dict(d) for d in cur.fetchall()]

        return jsonify({
            'success': True,
            'run': dict(run),
            'details': details
        })
    finally:
        conn.close()


@audit_bp.route('/history/<file_id>', methods=['GET'])
def get_history(file_id: str):
    """获取某个档案的所有校验历史"""
    conn = get_db()
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT run_uuid, file_id, file_name, started_at, completed_at,
                   status, total_rules, pass_count, warn_count, fail_count
            FROM audit_runs
            WHERE file_id = ?
            ORDER BY created_at DESC
            LIMIT 20
        """, (file_id,))
        runs = [dict(r) for r in cur.fetchall()]
        return jsonify({'success': True, 'runs': runs, 'total': len(runs)})
    finally:
        conn.close()


@audit_bp.route('/rules/config', methods=['PUT'])
def update_rules_config():
    """更新规则配置（启用/禁用规则）"""
    data = request.get_json() or {}
    enabled_rules = data.get('enabled_rules', [])

    with open(RULES_CONFIG_PATH, 'r', encoding='utf-8') as f:
        rules = json.load(f)

    # 更新 enabled 状态
    for rule in rules:
        rule_id = rule.get('id', '')
        rule['enabled'] = rule_id in enabled_rules

    with open(RULES_CONFIG_PATH, 'w', encoding='utf-8') as f:
        json.dump(rules, f, ensure_ascii=False, indent=2)

    return jsonify({'success': True, 'total': len(rules)})
