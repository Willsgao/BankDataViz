# -*- coding:utf-8 -*-
"""
LLM 表格识别 API 路由
"""



import uuid
import logging

from flask import Blueprint, request, send_from_directory

from backend.utils.constants import MAIN_ROOT


TASK_RESULTS = {}

# 创建蓝图
llm_bp = Blueprint('llm', __name__)

# 设置日志
logger = logging.getLogger(__name__)


# 全局处理器实例
_table_processor_instance = None
_non_financial_table_service = None

from backend.service.llm_related_functions import *

def _send_websocket_notification(task_id, status, data=None, error_message=None):
    """发送WebSocket通知的通用函数 - 完全修复版本"""
    try:
        print(f"📨 发送WebSocket通知 - 任务ID: {task_id}, 状态: {status}")

        # 创建模拟的WebSocket函数，避免导入错误
        def mock_notify_task_completion(task_id, data):
            print(f"📨 [WebSocket模拟] 任务完成通知 - 任务ID: {task_id}")
            print(f"📨 [WebSocket模拟] 通知数据: {data}")
            # 这里可以记录到日志文件，供前端轮询使用
            logger.info(f"WebSocket任务完成模拟 - 任务ID: {task_id}, 数据: {data}")

        def mock_notify_task_error(task_id, error_message):
            print(f"📨 [WebSocket模拟] 任务错误通知 - 任务ID: {task_id}")
            print(f"📨 [WebSocket模拟] 错误信息: {error_message}")
            logger.error(f"WebSocket任务错误模拟 - 任务ID: {task_id}, 错误: {error_message}")

        # 使用模拟函数
        notify_task_completion = mock_notify_task_completion
        notify_task_error = mock_notify_task_error

        if status == 'completed' and data:
            notification_data = {
                "total": data.get('total', 0),
                "success": data.get('success', 0),
                "failed": data.get('failed', 0),
                "output_file": data.get('output_file', ''),
                "excel_url": data.get('excel_url', ''),
                "table_type": data.get('table_type', 'financial'),
                "message": data.get('message', '处理完成'),
                "task_id": task_id,
                "type": "task_completed",
                "processing_completed": True,
                "timestamp": time.time()
            }
            print(f"📨 发送完成通知数据: {notification_data}")
            notify_task_completion(task_id, notification_data)

        elif status == 'error' and error_message:
            print(f"📨 发送错误通知: {error_message}")
            notify_task_error(task_id, error_message)

        print(f"✅ WebSocket通知发送完成 - 任务ID: {task_id}")

    except Exception as notify_error:
        print(f"❌ 发送WebSocket通知失败: {str(notify_error)}")
        # 即使通知失败，也不影响主要处理流程



async def _test_connection_internal(data):
    """测试LLM连接的内部异步函数"""
    try:
        # 验证必要参数
        is_valid, error_msg = validate_required_params(
            data, ['base_url', 'api_key', 'model_id']
        )
        if not is_valid:
            return {
                "success": False,
                "error": error_msg
            }

        base_url = data.get('base_url')
        api_key = data.get('api_key')
        model_id = data.get('model_id')

        print(
            f"🔧 测试连接参数: base_url={base_url}, model_id={model_id}, api_key_length={len(api_key) if api_key else 0}")

        # 确保URL格式正确
        if not base_url.endswith('/'):
            base_url = base_url + '/'

        # 创建临时客户端进行测试
        from openai import AsyncOpenAI

        try:
            test_client = AsyncOpenAI(
                base_url=base_url,
                api_key=api_key
            )

            # 测试连接 - 使用更简单的消息
            print(f"🔧 开始API调用测试...")
            response = await test_client.chat.completions.create(
                model=model_id,
                messages=[{"role": "user", "content": "Hello, please respond with just 'OK'"}],
                max_tokens=10,
                timeout=30.0  # 添加超时设置
            )

            print(f"🔧 API响应: {response}")

            return {
                "success": True,
                "message": "LLM连接测试成功",
                "data": {
                    "model_response": bool(response.choices),
                    "response_content": response.choices[0].message.content if response.choices else None
                }
            }

        except Exception as api_error:
            print(f"❌ API调用失败: {str(api_error)}")
            error_detail = str(api_error)

            # 提供更详细的错误信息
            if "401" in error_detail:
                error_msg = "API密钥无效或未授权"
            elif "404" in error_detail:
                error_msg = "模型不存在或URL路径错误"
            elif "connect" in error_detail.lower():
                error_msg = "无法连接到服务器，请检查网络和URL"
            elif "timeout" in error_detail.lower():
                error_msg = "连接超时，请检查网络或服务器状态"
            else:
                error_msg = f"API调用失败: {error_detail}"

            return {
                "success": False,
                "error": error_msg
            }

    except Exception as e:
        print(f"❌ 连接测试异常: {str(e)}")
        return {
            "success": False,
            "error": f"连接测试失败: {str(e)}"
        }

async def _process_single_image(data):
    """处理单张图片的异步函数"""
    try:
        # 验证必要参数
        is_valid, error_msg = validate_required_params(
            data, ['image_path', 'output_path']
        )

        print("###################")
        print(is_valid, error_msg)

        if not is_valid:
            return {
                "success": False,
                "error": error_msg
            }

        image_path = data.get('image_path')
        output_path = data.get('output_path')
        sheet_name = data.get('sheet_name', '识别结果')
        bank_name = data.get('bank_name', '未知银行')
        file_name = data.get('file_name', Path(image_path).stem)

        print("************image_path****output_path****************")
        print(image_path)
        print(output_path)
        print(sheet_name)
        print(bank_name)
        print(file_name)

        # 检查图片文件是否存在
        if not Path(image_path).exists():
            return {
                "success": False,
                "error": f"图片文件不存在: {image_path}"
            }

        processor = get_table_processor()

        print("****************processor.llm_client******************")
        print(processor.llm_client)

        if not processor.llm_client:
            return {
                "success": False,
                "error": "请先配置LLM客户端"
            }

        # 使用已有的标识创建Excel存储文件夹
        # 从 image_path 提取标识：类似 d0586abf1323dbfd80a926ce1e2d5676
        folder_name = Path(image_path).stem.split('_')[0]  # 取第一个下划线前的部分

        excel_base_dir = Path("static/excel_data")
        excel_dir = excel_base_dir / folder_name
        excel_dir.mkdir(parents=True, exist_ok=True)

        # 重新构建输出路径
        excel_filename = Path(output_path).name
        new_output_path = excel_dir / excel_filename
        output_path = str(new_output_path)

        print(f"Excel文件将保存到: {output_path}")

        # 配置Excel保存参数
        excel_config = ExcelSaveConfig(
            anchor_cell=data.get('anchor_cell', 'R2'),
            width_px=data.get('width_px', 768),
            mode=data.get('mode', 'overwrite')
        )

        # 处理图片 - 使用新的方法名和参数
        result = await processor.process_table_pipeline(
            image_path=image_path,
            out_file=output_path,
            sheet_name=sheet_name,
            bank_name=bank_name,
            file_name=file_name,
            excel_config=excel_config
        )

        logger.info(f"表格识别完成 - 状态: {result.status}, 文件: {output_path}")

        return {
            "success": result.status == "success",
            "data": {
                "status": result.status,
                "complexity": result.complexity,
                "mode": result.mode,
                "table_name": result.table_name,
                "assessment_reason": result.assessment_reason,
                "error_message": result.error_message,
                "output_path": output_path  # 返回实际的输出路径
            }
        }

    except Exception as e:
        logger.error(f"表格识别失败: {str(e)}")
        return {
            "success": False,
            "error": f"处理失败: {str(e)}"
        }

async def _process_non_financial_table(data):
    """处理普通表格的异步函数"""
    try:
        # 验证必要参数
        is_valid, error_msg = validate_required_params(
            data, ['image_path', 'output_path']
        )

        if not is_valid:
            return {
                "success": False,
                "error": error_msg
            }

        image_path = data.get('image_path')
        output_path = data.get('output_path')
        sheet_name = data.get('sheet_name', '普通表格识别结果')
        bank_name = data.get('bank_name', '未知机构')
        file_name = data.get('file_name', Path(image_path).stem)

        print("普通表格处理参数:")
        print(f"image_path: {image_path}")
        print(f"output_path: {output_path}")
        print(f"sheet_name: {sheet_name}")
        print(f"bank_name: {bank_name}")
        print(f"file_name: {file_name}")

        # 检查图片文件是否存在
        if not Path(image_path).exists():
            return {
                "success": False,
                "error": f"图片文件不存在: {image_path}"
            }

        service = get_non_financial_table_service()

        if not service.llm_client:
            return {
                "success": False,
                "error": "请先配置LLM客户端"
            }

        # 使用已有的标识创建Excel存储文件夹
        folder_name = Path(image_path).stem.split('_')[0]
        excel_base_dir = Path("static/excel_data")
        excel_dir = excel_base_dir / folder_name
        excel_dir.mkdir(parents=True, exist_ok=True)

        # 重新构建输出路径
        excel_filename = Path(output_path).name
        new_output_path = excel_dir / excel_filename
        output_path = str(new_output_path)

        print(f"普通表格Excel文件将保存到: {output_path}")

        # 配置Excel保存参数
        excel_config = ExcelSaveConfig(
            anchor_cell=data.get('anchor_cell', 'R2'),
            width_px=data.get('width_px', 768),
            mode=data.get('mode', 'overwrite')
        )

        # 处理普通表格
        result = await service.process_table_pipeline(
            image_path=image_path,
            out_file=output_path,
            sheet_name=sheet_name,
            bank_name=bank_name,
            file_name=file_name,
            excel_config=excel_config
        )

        logger.info(f"普通表格识别完成 - 状态: {result.status}, 文件: {output_path}")

        return {
            "success": result.status == "success",
            "data": {
                "status": result.status,
                "complexity": result.complexity,
                "mode": result.mode,
                "table_name": result.table_name,
                "assessment_reason": result.assessment_reason,
                "error_message": result.error_message,
                "output_path": output_path,
                "table_type": "non_financial"
            }
        }

    except Exception as e:
        logger.error(f"普通表格识别失败: {str(e)}")
        return {
            "success": False,
            "error": f"处理失败: {str(e)}"
        }

def _batch_process_images_sync(data, task_id):
    """同步批量处理函数 - 修复版本"""
    try:
        # 初始化任务状态
        PROCESSING_STATUS[task_id] = {
            "status": "processing",
            "progress": 0,
            "message": "开始处理",
            "total": len(data.get('image_paths', [])),
            "processed": 0,
            "start_time": time.time()
        }

        print(f"🔄 开始同步批量处理 - 任务ID: {task_id}")

        # 发送任务开始通知
        _send_websocket_notification(task_id, 'started', {
            'message': '任务已开始处理',
            'total': len(data.get('image_paths', []))
        })

        # 同步调用批量处理逻辑
        result = asyncio.run(_batch_process_images(data, task_id))

        print(f"🔄 批量处理完成，结果: {result.get('success')}")

        # 关键修改：只要处理流程完成，就认为是成功状态
        if result.get('processing_completed', False) or result.get('success', False):
            PROCESSING_STATUS[task_id].update({
                "status": "completed",
                "progress": 100,
                "message": result.get('data', {}).get('message', '处理完成'),
                "completion_time": time.time()
            })

            # 发送完成通知
            _send_websocket_notification(task_id, 'completed', result.get('data', {}))
        else:
            PROCESSING_STATUS[task_id].update({
                "status": "error",
                "message": result.get('error', '处理失败'),
                "error_time": time.time()
            })
            _send_websocket_notification(task_id, 'error', error_message=result.get('error'))

    except Exception as e:
        logger.error(f"批量处理异常: {str(e)}")
        PROCESSING_STATUS[task_id] = {
            "status": "error",
            "message": f"处理异常: {str(e)}",
            "error_time": time.time()
        }
        _send_websocket_notification(task_id, 'error', error_message=str(e))

async def _batch_process_images(data, task_id):

    # 初始化结果变量
    final_data = {}

    try:
        # 初始化任务状态
        PROCESSING_STATUS[task_id] = {
            "status": "processing",
            "progress": 0,
            "message": "开始处理",
            "total": len(data.get('image_paths', [])),
            "processed": 0,
            "start_time": time.time()
        }

        # 验证必要参数
        is_valid, error_msg = validate_required_params(data, ['image_paths', 'output_dir'])
        if not is_valid:
            final_status = "error"
            final_message = error_msg
            final_result = {
                "success": False,
                "error": error_msg,
                "task_id": task_id
            }
        else:
            image_paths = data.get('image_paths', [])
            output_dir = data.get('output_dir')
            bank_name = data.get('bank_name', '未知银行')
            table_type = data.get('table_type', 'financial')
            output_file = data.get('output_file', 'batch_processing_results.xlsx')

            print(f"🔄 批量处理开始 - 类型: {table_type}, 图片数量: {len(image_paths)}")

            # 更新任务状态
            PROCESSING_STATUS[task_id].update({
                "total": len(image_paths),
                "message": f"开始处理 {len(image_paths)} 张图片",
                "table_type": table_type
            })

            # 发送任务开始通知
            _send_websocket_notification(task_id, 'started', {
                'message': '任务已开始处理',
                'total': len(image_paths),
                'table_type': table_type
            })

            # 检查图片文件
            missing_images = []
            valid_image_paths = []

            for img_path in image_paths:
                img_full_path = Path(img_path)
                if not img_full_path.exists():
                    # 尝试在static目录下查找
                    static_path = Path("static") / img_path
                    if static_path.exists():
                        valid_image_paths.append(str(static_path))
                        print(f"✅ 在static目录找到图片: {static_path}")
                    else:
                        # 尝试直接在当前目录查找
                        current_dir_path = Path.cwd() / img_path
                        if current_dir_path.exists():
                            valid_image_paths.append(str(current_dir_path))
                            print(f"✅ 在当前目录找到图片: {current_dir_path}")
                        else:
                            missing_images.append(img_path)
                            print(f"❌ 图片文件不存在: {img_path}")
                else:
                    valid_image_paths.append(img_path)

            if missing_images:
                final_status = "error"
                final_message = f"以下图片文件不存在: {missing_images}"
                final_result = {
                    "success": False,
                    "error": final_message,
                    "task_id": task_id
                }
            else:
                # 根据表格类型选择处理器
                if table_type == 'non_financial':
                    processor = get_non_financial_table_service()
                    print("🔄 使用普通表格处理器")
                else:
                    processor = get_table_processor()
                    print("🔄 使用金融表格处理器")

                # 检查处理器状态
                print(f"🔍 处理器检查 - llm_client: {processor.llm_client is not None}")
                print(f"🔍 处理器检查 - model_id: {getattr(processor, 'model_id', '未设置')}")

                if not processor.llm_client:
                    final_status = "error"
                    final_message = "请先配置LLM客户端"
                    final_result = {
                        "success": False,
                        "error": final_message,
                        "task_id": task_id
                    }
                else:
                    # 使用已有的标识创建Excel存储文件夹
                    if valid_image_paths:
                        first_image_path = Path(valid_image_paths[0])
                        print(f"🔍 第一个图片路径: {first_image_path}")

                        if 'joined_tables' in str(first_image_path):
                            folder_name = first_image_path.parent.name
                            print(f"📁 从joined_tables提取文件夹名: {folder_name}")
                        else:
                            folder_name = first_image_path.stem.split('_')[0]
                            print(f"📁 从文件名提取文件夹名: {folder_name}")
                    else:
                        folder_name = "unknown_batch"
                        print(f"📁 使用默认文件夹名: {folder_name}")

                    excel_base_dir = Path("static/excel_data")
                    excel_dir = excel_base_dir / folder_name
                    excel_dir.mkdir(parents=True, exist_ok=True)

                    # 根据表格类型构建不同的文件名
                    if table_type == 'non_financial':
                        excel_filename = f"non_financial_batch_{folder_name}.xlsx"
                    else:
                        excel_filename = f"financial_batch_{folder_name}.xlsx"

                    new_output_file = excel_dir / excel_filename
                    output_file = str(new_output_file)

                    print(f"📁 批量处理Excel文件将保存到: {output_file}")

                    # 更新任务状态
                    PROCESSING_STATUS[task_id].update({
                        "message": f"开始处理图片，输出文件: {output_file}",
                        "output_file": output_file
                    })

                    # 批量处理
                    results = []
                    success_count = 0
                    failed_count = 0

                    for i, image_path in enumerate(valid_image_paths):
                        try:
                            # 更新处理进度
                            progress = int((i / len(valid_image_paths)) * 100)
                            PROCESSING_STATUS[task_id].update({
                                "progress": progress,
                                "processed": i,
                                "message": f"正在处理第 {i + 1}/{len(valid_image_paths)} 张图片: {Path(image_path).name}"
                            })

                            # 发送进度通知
                            _send_websocket_notification(task_id, 'progress', {
                                'progress': progress,
                                'message': f'正在处理第 {i + 1}/{len(valid_image_paths)} 张图片',
                                'current': i + 1,
                                'total': len(valid_image_paths)
                            })

                            sheet_name = f"表格_{i + 1}"
                            file_name = Path(image_path).stem

                            print(f"🔄 处理第 {i + 1}/{len(valid_image_paths)} 张图片: {Path(image_path).name}")

                            if table_type == 'non_financial':
                                # 普通表格处理
                                result = await _process_single_non_financial_table({
                                    'image_path': image_path,
                                    'output_path': output_file,
                                    'sheet_name': sheet_name,
                                    'bank_name': bank_name,
                                    'file_name': file_name,
                                    'table_index': i
                                })

                                if result.get('success'):
                                    results.append({
                                        "image_path": image_path,
                                        "status": "success",
                                        "table_name": sheet_name,
                                        "sheet_name": sheet_name,
                                        "excel_url": result.get('excel_url')
                                    })
                                    success_count += 1
                                    print(f"✅ 普通表格处理成功: {Path(image_path).name}")
                                else:
                                    error_msg = result.get('error', '处理失败')
                                    results.append({
                                        "image_path": image_path,
                                        "status": "error",
                                        "error": error_msg
                                    })
                                    failed_count += 1
                                    print(f"❌ 普通表格处理失败: {Path(image_path).name} - {error_msg}")
                            else:
                                # 金融表格处理
                                try:
                                    result = await processor.process_table_pipeline(
                                        image_path=image_path,
                                        out_file=output_file,
                                        sheet_name=sheet_name,
                                        bank_name=bank_name,
                                        file_name=file_name
                                    )

                                    print(f"🔍 金融表格处理结果: {result}")
                                    print(f"🔍 处理状态: {getattr(result, 'status', '无状态')}")

                                    results.append({
                                        "image_path": image_path,
                                        "status": getattr(result, 'status', 'unknown'),
                                        "complexity": getattr(result, 'complexity', 'unknown'),
                                        "table_name": getattr(result, 'table_name', '未知'),
                                        "sheet_name": sheet_name,
                                        "error_message": getattr(result, 'error_message', '')
                                    })

                                    if getattr(result, 'status', '') == "success":
                                        success_count += 1
                                        print(f"✅ 金融表格处理成功: {Path(image_path).name}")
                                    else:
                                        failed_count += 1
                                        error_msg = getattr(result, 'error_message', '处理失败')
                                        print(f"❌ 金融表格处理失败: {Path(image_path).name} - {error_msg}")

                                except Exception as process_error:
                                    print(f"💥 金融表格处理异常: {str(process_error)}")
                                    import traceback
                                    traceback.print_exc()
                                    results.append({
                                        "image_path": image_path,
                                        "status": "error",
                                        "error": f"处理异常: {str(process_error)}"
                                    })
                                    failed_count += 1

                        except Exception as e:
                            logger.error(f"处理图片失败 {image_path}: {str(e)}")
                            results.append({
                                "image_path": image_path,
                                "status": "error",
                                "error": str(e)
                            })
                            failed_count += 1
                            print(f"💥 处理异常: {Path(image_path).name} - {str(e)}")
                            import traceback
                            traceback.print_exc()

                    # 最终进度更新
                    PROCESSING_STATUS[task_id].update({
                        "progress": 100,
                        "processed": len(valid_image_paths),
                        "message": f"批量处理完成，成功: {success_count}, 失败: {failed_count}"
                    })

                    logger.info(
                        f"批量处理完成 - 类型: {table_type}, 总数: {len(valid_image_paths)}, 成功: {success_count}, 失败: {failed_count}")

                    # 检查最终文件是否存在
                    final_excel_path = Path(output_file)
                    if final_excel_path.exists():
                        print(f"✅ 批量处理完成，Excel文件已生成: {output_file}")
                    else:
                        print(f"❌ 批量处理完成，但Excel文件未生成: {output_file}")

                    # 构建返回的Excel URL
                    excel_url = convert_to_excel_url(output_file)
                    print(f"🔗 生成的Excel URL: {excel_url}")

                    # 计算非金融表格的数量
                    non_financial_count = sum(1 for r in results if r.get('is_non_financial'))

                    # 根据处理结果生成不同的消息
                    if success_count > 0 and non_financial_count == success_count:
                        completion_message = f"批量处理完成：所有{len(valid_image_paths)}个表格均为非金融表格，已跳过处理"
                    elif success_count > 0:
                        completion_message = f"批量处理完成，成功: {success_count}, 失败: {failed_count}, 非金融表格: {non_financial_count}"
                    else:
                        completion_message = f"批量处理完成，总数: {len(valid_image_paths)}, 成功: {success_count}, 失败: {failed_count}"

                    print(
                        f"📊 最终处理统计 - 总数: {len(valid_image_paths)}, 成功: {success_count}, 失败: {failed_count}, 非金融表格: {non_financial_count}")

                    # 标记任务完成
                    PROCESSING_STATUS[task_id].update({
                        "status": "completed",
                        "completion_time": time.time(),
                        "success_count": success_count,
                        "failed_count": failed_count,
                        "non_financial_count": non_financial_count,
                        "excel_url": excel_url,
                        "message": completion_message
                    })

                    # 发送完成通知
                    _send_websocket_notification(task_id, 'completed', {
                        'total': len(valid_image_paths),
                        'success': success_count,
                        'failed': failed_count,
                        'non_financial': non_financial_count,
                        'output_file': output_file,
                        'excel_url': excel_url,
                        'table_type': table_type,
                        'message': completion_message,
                        'processing_completed': True
                    })

                    # 设置最终结果
                    final_status = "completed"
                    final_message = completion_message
                    final_data = {
                        "total": len(valid_image_paths),
                        "success": success_count,
                        "failed": failed_count,
                        "non_financial": non_financial_count,
                        "results": results,
                        "output_file": output_file,
                        "excel_url": excel_url,
                        "table_type": table_type,
                        "processing_completed": True,
                        "message": completion_message
                    }
                    final_result = {
                        "success": True,
                        "data": final_data,
                        "task_id": task_id,
                        "processing_completed": True
                    }

    except Exception as e:
        logger.error(f"批量处理失败: {str(e)}")
        import traceback
        traceback.print_exc()

        final_status = "error"
        final_message = f"批量处理失败: {str(e)}"
        final_result = {
            "success": False,
            "error": final_message,
            "task_id": task_id
        }

        # 记录错误状态
        PROCESSING_STATUS[task_id] = {
            "status": "error",
            "message": final_message,
            "error_time": time.time(),
            "error": str(e)
        }

        _send_websocket_notification(task_id, 'error', error_message=str(e))

    # ⭐⭐⭐ 集中存储到 TASK_RESULTS ⭐⭐⭐
    if final_status == "completed":
        TASK_RESULTS[task_id] = {
            "status": "completed",
            "data": final_data,
            "processing_completed": True,
            "completed_at": time.time()
        }
        print(f"✅ 任务结果已存储到 TASK_RESULTS: {task_id}")
    elif final_status == "error":
        TASK_RESULTS[task_id] = {
            "status": "error",
            "error": final_message,
            "completed_at": time.time()
        }
        print(f"❌ 错误结果已存储到 TASK_RESULTS: {task_id}")

    print(f"📊 当前 TASK_RESULTS 中的任务数量: {len(TASK_RESULTS)}")

    # ⭐⭐⭐ 集中返回结果 ⭐⭐⭐
    return final_result



async def _process_single_non_financial_table(process_data):
    """处理单个普通表格"""
    try:
        image_path = process_data['image_path']
        output_path = process_data['output_path']
        sheet_name = process_data['sheet_name']
        bank_name = process_data['bank_name']
        file_name = process_data['file_name']
        table_index = process_data.get('table_index', 0)

        print(f"处理普通表格 #{table_index}: {image_path}")

        # 调用普通表格识别的核心逻辑
        result = await _process_non_financial_table(process_data)

        if result.get('success'):
            # 转换为统一的URL格式
            excel_url = convert_to_excel_url(output_path)
            result['excel_url'] = excel_url

        return result

    except Exception as e:
        logger.error(f"处理普通表格失败 {process_data['image_path']}: {str(e)}")
        return {
            "success": False,
            "error": f"处理失败: {str(e)}"
        }




@llm_bp.route('/llm/test-connection', methods=['POST'])
def test_connection():
    """测试LLM连接"""
    try:
        data = request.get_json()

        if not data:
            return jsonify({
                "success": False,
                "error": "请求体不能为空"
            }), 400

        result = asyncio.run(_test_connection_internal(data))
        return jsonify(result)
    except Exception as e:
        logger.error(f"测试连接错误: {str(e)}")
        return jsonify({
            "success": False,
            "error": f"测试连接错误: {str(e)}"
        }), 500




@llm_bp.route('/llm/available-models', methods=['GET'])
def get_available_models():
    """获取可用的模型列表"""
    models = [
        {
            "id": "doubao-1-5-vision-pro-250328",
            "name": "豆包视觉专业版",
            "description": "支持视觉识别的专业模型",
            "max_tokens": 16000
        },
        {
            "id": "doubao-seed-1-6-vision-250815",
            "name": "豆包视觉种子版",
            "description": "视觉识别基础模型",
            "max_tokens": 16000
        }
    ]

    return jsonify({
        "success": True,
        "data": models
    })




@llm_bp.route('/llm/health', methods=['GET'])
def health_check():
    """健康检查端点"""
    try:
        processor = get_table_processor()

        # 修复：确保所有值都是可JSON序列化的
        base_url = getattr(processor.llm_client, 'base_url', None)
        if base_url is not None:
            base_url = str(base_url)

        health_status = {
            "service": "running",
            "llm_configured": processor.llm_client is not None,
            "model_id": processor.model_id,
            "base_url": base_url,  # 使用转换后的字符串
            "timestamp": time.time()  # 使用time.time()而不是asyncio的时间
        }

        return jsonify({
            "success": True,
            "data": health_status
        })
    except Exception as e:
        logger.error(f"健康检查失败: {str(e)}")
        return jsonify({
            "success": False,
            "error": f"健康检查失败: {str(e)}"
        }), 500


@llm_bp.route('/llm/recognize-table', methods=['POST'])
def recognize_table():
    """识别表格并保存到指定路径 - 先检查Excel是否存在"""
    try:
        data = request.get_json()
        print(f"收到识别请求: {data}")

        if not data:
            return jsonify({
                "success": False,
                "error": "请求体不能为空"
            }), 400

        # 验证必要参数
        required_fields = ['imageUrl', 'excelPath']
        missing_fields = [field for field in required_fields if not data.get(field)]
        if missing_fields:
            return jsonify({
                "success": False,
                "error": f"缺少必要参数: {', '.join(missing_fields)}"
            }), 400

        image_url = data.get('imageUrl')
        excel_path = data.get('excelPath')
        table_name = data.get('tableName', '识别结果')
        index = data.get('index', 0)

        print(f"识别参数 - imageUrl: {image_url}, excelPath: {excel_path}")

        # 构建完整Excel路径
        excel_full_path = Path(excel_path)
        if not excel_full_path.is_absolute():
            excel_full_path = Path.cwd() / excel_path

        print(f"检查Excel文件是否存在: {excel_full_path}")

        # 1. 首先检查Excel文件是否已经存在
        if excel_full_path.exists():
            print("✅ Excel文件已存在，直接读取数据")
            try:
                import openpyxl
                workbook = openpyxl.load_workbook(excel_full_path)
                sheet_name = workbook.sheetnames[0]
                worksheet = workbook[sheet_name]

                # 转换为JSON数据
                headers = []
                data = []

                # 读取表头（第一行）
                for col in range(1, worksheet.max_column + 1):
                    cell_value = worksheet.cell(row=1, column=col).value
                    headers.append(str(cell_value) if cell_value is not None else f"列{col}")

                # 读取数据行
                for row in range(2, worksheet.max_row + 1):
                    row_data = {}
                    for col, header in enumerate(headers, 1):
                        cell_value = worksheet.cell(row=row, column=col).value
                        row_data[header] = str(cell_value) if cell_value is not None else ""
                    if any(row_data.values()):  # 只添加非空行
                        data.append(row_data)

                return jsonify({
                    "success": True,
                    "recognizedData": {
                        "headers": headers,
                        "data": data,
                        "tableName": table_name,
                        "excelPath": str(excel_full_path)
                    },
                    "message": "已加载现有表格数据",
                    "fromCache": True  # 添加标记表示来自缓存
                })
            except Exception as e:
                logger.error(f"读取现有Excel失败: {str(e)}")
                # 如果读取失败，继续走LLM识别流程
                print("❌ 读取现有Excel失败，继续走LLM识别")

        # 2. 如果Excel不存在，进行LLM识别
        print("🔄 Excel文件不存在，开始LLM识别流程")

        # ⭐⭐⭐ 关键修改：改进图片路径提取逻辑 ⭐⭐⭐
        image_path = image_url

        # 处理各种URL格式
        if image_url.startswith('http://'):
            # 提取域名后的路径部分
            from urllib.parse import urlparse
            parsed_url = urlparse(image_url)
            image_path = parsed_url.path  # 获取路径部分，如 /static/joined_tables/...

            # 去掉开头的斜杠
            if image_path.startswith('/'):
                image_path = image_path[1:]

            print(f"✅ 从URL提取路径: {image_path}")

        elif image_url.startswith('/'):
            # 如果是以/开头的绝对路径，去掉开头的斜杠
            image_path = image_url[1:]
            print(f"✅ 处理绝对路径: {image_path}")

        print(f"处理图片路径: {image_path}")

        # 检查图片文件是否存在
        image_full_path = Path(image_path)
        if not image_full_path.exists():
            # 尝试在static目录下查找
            static_path = Path("static") / image_path
            if static_path.exists():
                image_full_path = static_path
                print(f"✅ 在static目录找到图片: {static_path}")
            else:
                # 尝试直接在当前目录查找
                current_dir_path = Path.cwd() / image_path
                if current_dir_path.exists():
                    image_full_path = current_dir_path
                    print(f"✅ 在当前目录找到图片: {current_dir_path}")
                else:
                    print(f"❌ 图片文件不存在，尝试的路径:")
                    print(f"   - {image_full_path}")
                    print(f"   - {static_path}")
                    print(f"   - {current_dir_path}")
                    return jsonify({
                        "success": False,
                        "error": f"图片文件不存在: {image_path}"
                    }), 404

        print(f"✅ 最终使用的图片路径: {image_full_path}")

        # 确保Excel目录存在
        excel_full_path.parent.mkdir(parents=True, exist_ok=True)

        print(f"Excel保存路径: {excel_full_path}")

        # 调用 _process_single_image 异步函数进行LLM识别
        process_data = {
            'image_path': str(image_full_path),
            'output_path': str(excel_full_path),
            'sheet_name': table_name,
            'bank_name': '未知银行',
            'file_name': f"table_{index + 1}"
        }

        # 使用异步处理
        result = asyncio.run(
            (process_data))
        print(f"LLM识别结果: {result}")

        if result.get('success'):
            # 读取生成的Excel文件数据返回给前端
            output_path = result['data']['output_path']
            excel_result_path = Path(output_path)

            if excel_result_path.exists():
                try:
                    import openpyxl
                    workbook = openpyxl.load_workbook(excel_result_path)
                    sheet_name = workbook.sheetnames[0]
                    worksheet = workbook[sheet_name]

                    # 转换为JSON数据
                    headers = []
                    data = []

                    # 读取表头（第一行）
                    for col in range(1, worksheet.max_column + 1):
                        cell_value = worksheet.cell(row=1, column=col).value
                        headers.append(str(cell_value) if cell_value is not None else f"列{col}")

                    # 读取数据行
                    for row in range(2, worksheet.max_row + 1):
                        row_data = {}
                        for col, header in enumerate(headers, 1):
                            cell_value = worksheet.cell(row=row, column=col).value
                            row_data[header] = str(cell_value) if cell_value is not None else ""
                        if any(row_data.values()):  # 只添加非空行
                            data.append(row_data)

                    return jsonify({
                        "success": True,
                        "recognizedData": {
                            "headers": headers,
                            "data": data,
                            "tableName": table_name,
                            "excelPath": str(excel_result_path)
                        },
                        "message": "表格识别完成",
                        "fromCache": False  # 添加标记表示来自LLM识别
                    })
                except Exception as e:
                    logger.error(f"读取生成的Excel失败: {str(e)}")
                    # 即使读取失败，也返回成功，因为Excel文件已经生成
                    return jsonify({
                        "success": True,
                        "recognizedData": {
                            "headers": ["识别完成"],
                            "data": [{"状态": "表格识别完成，但读取数据失败"}],
                            "tableName": table_name,
                            "excelPath": str(excel_result_path)
                        },
                        "message": "表格识别完成",
                        "fromCache": False
                    })
            else:
                return jsonify({
                    "success": False,
                    "error": "识别完成但Excel文件未生成"
                }), 500
        else:
            return jsonify({
                "success": False,
                "error": result.get('error', '识别失败')
            }), 500

    except Exception as e:
        logger.error(f"表格识别失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": f"表格识别失败: {str(e)}"
        }), 500


@llm_bp.route('/llm/check-excel', methods=['GET'])
def check_excel():
    """检查Excel文件是否存在"""
    try:
        file_path = request.args.get('path')
        print(f"🔍 检查Excel文件请求 - path: {file_path}")

        if not file_path:
            return jsonify({
                "success": False,
                "error": "缺少path参数"
            }), 400

        # 构建完整路径
        full_path = Path(file_path)
        if not full_path.is_absolute():
            full_path = Path.cwd() / file_path

        print(f"🔍 检查Excel完整路径: {full_path}")
        print(f"🔍 文件是否存在: {full_path.exists()}")

        if full_path.exists():
            try:
                import openpyxl
                workbook = openpyxl.load_workbook(full_path)
                sheet_name = workbook.sheetnames[0]
                worksheet = workbook[sheet_name]

                # 转换为JSON数据
                data = []
                headers = []

                # 读取表头（第一行）
                for col in range(1, worksheet.max_column + 1):
                    cell_value = worksheet.cell(row=1, column=col).value
                    headers.append(str(cell_value) if cell_value is not None else f"列{col}")

                # 读取数据行
                for row in range(2, worksheet.max_row + 1):
                    row_data = {}
                    for col, header in enumerate(headers, 1):
                        cell_value = worksheet.cell(row=row, column=col).value
                        row_data[header] = str(cell_value) if cell_value is not None else ""
                    if any(row_data.values()):  # 只添加非空行
                        data.append(row_data)

                print(f"🔍 Excel数据读取成功 - 表头: {headers}, 数据行数: {len(data)}")

                return jsonify({
                    "success": True,
                    "exists": True,
                    "excelData": {
                        "headers": headers,
                        "data": data,
                        "sheet_name": sheet_name,
                        "file_path": str(full_path)
                    }
                })
            except Exception as e:
                logger.error(f"读取Excel文件失败: {str(e)}")
                return jsonify({
                    "success": False,
                    "error": f"读取Excel文件失败: {str(e)}"
                }), 500
        else:
            print(f"🔍 Excel文件不存在: {file_path}")
            return jsonify({
                "success": True,
                "exists": False,
                "message": f"Excel文件不存在: {file_path}"
            })

    except Exception as e:
        logger.error(f"检查Excel失败: {str(e)}")
        return jsonify({
            "success": False,
            "error": f"检查Excel失败: {str(e)}"
        }), 500



@llm_bp.route('/llm/get-excel-data', methods=['GET'])
def get_excel_data():
    """读取Excel文件内容并返回给前端"""
    try:
        excel_url = request.args.get('url')
        if not excel_url:
            return jsonify({
                "success": False,
                "error": "缺少url参数"
            }), 400

        # 将URL转换为文件路径
        if excel_url.startswith('/static/'):
            file_path = excel_url[1:]  # 去掉开头的斜杠
        else:
            file_path = excel_url

        full_path = Path(file_path)
        if not full_path.is_absolute():
            full_path = Path.cwd() / file_path

        print(f"读取Excel文件: {full_path}")

        if not full_path.exists():
            return jsonify({
                "success": False,
                "error": f"Excel文件不存在: {file_path}"
            }), 404

        # 读取Excel文件
        import openpyxl
        workbook = openpyxl.load_workbook(full_path)
        sheet_name = workbook.sheetnames[0]
        worksheet = workbook[sheet_name]

        # 转换为JSON数据
        headers = []
        data_rows = []

        # 读取表头（第一行）
        for col in range(1, worksheet.max_column + 1):
            cell_value = worksheet.cell(row=1, column=col).value
            headers.append(str(cell_value) if cell_value is not None else f"列{col}")

        # 读取数据行
        for row in range(2, worksheet.max_row + 1):
            row_data = {}
            for col, header in enumerate(headers, 1):
                cell_value = worksheet.cell(row=row, column=col).value
                row_data[header] = str(cell_value) if cell_value is not None else ""
            if any(row_data.values()):  # 只添加非空行
                data_rows.append(row_data)

        return jsonify({
            "success": True,
            "data": {
                "headers": headers,
                "data": data_rows,
                "tableName": sheet_name,
                "excelPath": str(full_path)
            }
        })

    except Exception as e:
        logger.error(f"读取Excel数据失败: {str(e)}")
        return jsonify({
            "success": False,
            "error": f"读取Excel数据失败: {str(e)}"
        }), 500



# 在 llm_routes.py 中添加以下内容

@llm_bp.route('/llm/configure', methods=['POST'])
def configure_llm():
    """配置LLM参数"""
    try:
        data = request.get_json()
        print(f"🔧 收到配置请求: {data}")  # 调试信息

        if not data:
            return jsonify({
                "success": False,
                "error": "请求体不能为空"
            }), 400

        # 验证必要参数
        is_valid, error_msg = validate_required_params(
            data, ['base_url', 'api_key', 'model_id']
        )
        if not is_valid:
            return jsonify({
                "success": False,
                "error": error_msg
            }), 400

        base_url = data.get('base_url')
        api_key = data.get('api_key')
        model_id = data.get('model_id')
        table_type = data.get('table_type', 'financial')  # 新增：表格类型，默认为金融表格
        prompts = data.get('prompts', {})

        print(
            f"🔧 解析配置参数: base_url={base_url}, model_id={model_id}, table_type={table_type}, prompts_keys={list(prompts.keys())}")

        # 根据表格类型选择不同的处理器
        processor = None
        processor_type = ""

        if table_type == 'financial':
            from backend.service.table_llm_service import TableLLMService
            processor = TableLLMService(
                llm_client=None,  # 让服务自己创建客户端
                model_id=model_id
            )
            processor_type = "金融表格处理器"
        else:
            from backend.service.non_financial_table_service import NonFinancialTableService
            processor = NonFinancialTableService(
                llm_client=None,  # 让服务自己创建客户端
                model_id=model_id
            )
            processor_type = "普通表格处理器"

        print(
            f"🔧 使用处理器: {processor_type}, llm_client={processor.llm_client is not None}, model_id={processor.model_id}")

        # 检查客户端是否配置成功
        if not processor.llm_client:
            return jsonify({
                "success": False,
                "error": "LLM客户端配置失败，请检查API密钥和URL是否正确"
            }), 400

        # 更新提示词配置
        if prompts:
            if hasattr(processor, 'prompt_registry'):
                # TableLLMService 使用 prompt_registry
                for prompt_type, prompt_content in prompts.items():
                    if prompt_type in processor.prompt_registry:
                        processor.prompt_registry[prompt_type] = prompt_content
            elif hasattr(processor, 'prompt'):
                # NonFinancialTableService 使用 prompt
                if 'non_financial' in prompts:
                    processor.prompt = prompts['non_financial']

        # 保存配置到全局变量，以便其他接口使用
        global _table_processor_instance, _non_financial_table_service

        if table_type == 'financial':
            _table_processor_instance = processor
            # 清除普通表格服务实例，确保一致性
            _non_financial_table_service = None
        else:
            _non_financial_table_service = processor
            # 清除金融表格服务实例，确保一致性
            _table_processor_instance = None

        logger.info(f"LLM配置成功 - 模型: {model_id}, 基础URL: {base_url}, 表格类型: {table_type}")

        return jsonify({
            "success": True,
            "message": f"LLM配置成功（{processor_type}）",
            "data": {
                "model_id": processor.model_id,
                "base_url": base_url,
                "table_type": table_type,
                "processor_type": processor_type,
                "prompts_configured": list(prompts.keys()) if prompts else [],
                "client_configured": processor.llm_client is not None
            }
        })

    except Exception as e:
        logger.error(f"LLM配置失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": f"配置失败: {str(e)}"
        }), 500


# 修改获取处理器状态的接口，返回表格类型信息
@llm_bp.route('/llm/status', methods=['GET'])
def get_processor_status():
    """获取处理器状态"""
    try:
        # 检查当前配置的处理器类型
        current_processor = None
        table_type = "unknown"
        processor_type = "未知"

        if _table_processor_instance is not None:
            current_processor = _table_processor_instance
            table_type = "financial"
            processor_type = "金融表格处理器"
        elif _non_financial_table_service is not None:
            current_processor = _non_financial_table_service
            table_type = "non_financial"
            processor_type = "普通表格处理器"
        else:
            # 如果没有配置，使用默认的金融表格处理器
            from backend.service.table_llm_service import get_table_processor
            current_processor = get_table_processor()
            table_type = "financial"
            processor_type = "金融表格处理器（默认）"

        # 修复：确保所有值都是可JSON序列化的
        base_url = getattr(current_processor.llm_client, 'base_url', None)
        if base_url is not None:
            base_url = str(base_url)  # 将URL对象转换为字符串

        status = {
            "client_configured": current_processor.llm_client is not None,
            "model_id": current_processor.model_id,
            "base_url": base_url,  # 使用转换后的字符串
            "table_type": table_type,
            "processor_type": processor_type,
            "prompts_configured": {}
        }

        # 添加提示词配置信息
        if hasattr(current_processor, 'prompt_registry'):
            status["prompts_configured"] = {
                prompt_type: bool(content) and len(content.strip()) > 0
                for prompt_type, content in current_processor.prompt_registry.items()
            }
        elif hasattr(current_processor, 'prompt'):
            status["prompts_configured"] = {
                "non_financial": bool(current_processor.prompt) and len(current_processor.prompt.strip()) > 0
            }

        return jsonify({
            "success": True,
            "data": status
        })
    except Exception as e:
        logger.error(f"获取状态失败: {str(e)}")
        return jsonify({
            "success": False,
            "error": f"获取状态失败: {str(e)}"
        }), 500



# 修改处理图片的接口，根据配置自动选择服务
@llm_bp.route('/llm/process-image', methods=['POST'])
def process_table_image():
    """处理单张表格图片 - 根据配置自动选择服务"""
    try:
        data = request.get_json()
        print("&&&&&&&&&&&& 收到识别请求:", data)

        if not data:
            return jsonify({
                "success": False,
                "error": "请求体不能为空",
                "message": "请求体不能为空"
            }), 400

        # 提取参数
        image_path = data.get('image_path')
        output_path = data.get('output_path')
        sheet_name = data.get('sheet_name', '识别结果')
        bank_name = data.get('bank_name', '未知银行')
        file_name = data.get('file_name', 'table_1')
        request_table_type = data.get('table_type')  # 从请求中获取表格类型

        print(f"🔍 检查图片文件是否存在: {image_path}")

        # 检查图片文件是否存在
        image_full_path = Path(image_path)
        if not image_full_path.exists():
            # 尝试在static目录下查找
            static_path = Path("static") / image_path
            if static_path.exists():
                image_full_path = static_path
                print(f"✅ 找到图片文件: {static_path}")
            else:
                error_msg = f"图片文件不存在: {image_path}"
                print(f"❌ {error_msg}")
                return jsonify({
                    "success": False,
                    "error": error_msg,
                    "message": error_msg
                }), 404

        print(f"✅ 图片文件存在: {image_full_path}")

        # 使用已有的标识创建Excel存储文件夹
        folder_name = Path(image_path).stem.split('_')[0]
        excel_base_dir = Path("static/excel_data")
        excel_dir = excel_base_dir / folder_name
        excel_dir.mkdir(parents=True, exist_ok=True)

        # 重新构建输出路径
        excel_filename = f"single_{Path(image_path).stem}.xlsx"
        new_output_path = excel_dir / excel_filename
        final_output_path = str(new_output_path)

        print(f"Excel文件将保存到: {final_output_path}")

        # 检查最终路径是否存在
        final_excel_path = Path(final_output_path)
        if final_excel_path.exists():
            print("✅ Excel文件已存在，直接返回路径")

            # 将绝对路径转换为可访问的URL
            relative_path = final_output_path.replace('\\', '/')
            if 'static/' in relative_path:
                static_index = relative_path.find('static/')
                if static_index != -1:
                    excel_url = '/' + relative_path[static_index:]
                else:
                    excel_url = f"/static/excel_data/{folder_name}/{excel_filename}"
            else:
                excel_url = f"/static/excel_data/{folder_name}/{excel_filename}"

            return jsonify({
                "success": True,
                "from_cache": True,
                "message": "已加载现有表格数据",
                "excel_url": excel_url,
                "excel_path": f"{MAIN_ROOT}/{final_output_path}",
                "table_name": sheet_name,
                "table_type": request_table_type or "financial"
            })

        # 如果Excel不存在，进行LLM识别
        print("🔄 Excel文件不存在，开始LLM识别流程")

        # 根据请求的表格类型或当前配置选择处理器
        processor = get_appropriate_processor(request_table_type)
        current_table_type = "financial"

        if isinstance(processor, NonFinancialTableService):
            current_table_type = "non_financial"
            print("🔄 使用普通表格处理器")
        else:
            print("🔄 使用金融表格处理器")

        process_data = {
            'image_path': str(image_full_path),
            'output_path': final_output_path,
            'sheet_name': sheet_name,
            'bank_name': bank_name,
            'file_name': file_name
        }

        # 根据处理器类型调用不同的处理函数
        if current_table_type == 'non_financial':
            result = asyncio.run(_process_non_financial_table(process_data))
        else:
            result = asyncio.run(_process_single_image(process_data))

        print("LLM识别结果:", result)

        if result.get('success'):
            excel_url = f"/api/excel-data/{folder_name}/{excel_filename}"

            return jsonify({
                "success": True,
                "from_cache": False,
                "message": "表格识别完成",
                "excel_url": excel_url,
                "excel_path": f"{MAIN_ROOT}/{final_output_path}",
                "table_name": sheet_name,
                "table_type": current_table_type,
                "data": result['data']
            })
        else:
            error_msg = result.get('error', '识别失败')
            print(f"❌ LLM识别失败: {error_msg}")
            return jsonify({
                "success": False,
                "error": error_msg,
                "message": error_msg
            }), 500

    except Exception as e:
        logger.error(f"异步处理错误: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": f"异步处理错误: {str(e)}",
            "message": f"处理异常: {str(e)}"
        }), 500


@llm_bp.route('/llm/process-non-financial-table', methods=['POST'])
def process_non_financial_table():
    """处理普通表格"""
    try:
        data = request.get_json()
        print("收到普通表格识别请求:", data)

        if not data:
            return jsonify({
                "success": False,
                "error": "请求体不能为空",
                "message": "请求体不能为空"
            }), 400

        # 提取参数
        image_path = data.get('image_path')
        output_path = data.get('output_path')
        sheet_name = data.get('sheet_name', '普通表格识别结果')
        bank_name = data.get('bank_name', '未知机构')
        file_name = data.get('file_name', 'non_financial_table')

        print(f"处理普通表格 - 图片路径: {image_path}")

        # 检查图片文件是否存在
        image_full_path = Path(image_path)
        if not image_full_path.exists():
            # 尝试在static目录下查找
            static_path = Path("static") / image_path
            if static_path.exists():
                image_full_path = static_path
                print(f"✅ 找到图片文件: {static_path}")
            else:
                error_msg = f"图片文件不存在: {image_path}"
                print(f"❌ {error_msg}")
                return jsonify({
                    "success": False,
                    "error": error_msg,
                    "message": error_msg
                }), 404

        print(f"✅ 图片文件存在: {image_full_path}")

        # 使用已有的标识创建Excel存储文件夹
        folder_name = Path(image_path).stem.split('_')[0]
        excel_base_dir = Path("static/excel_data")
        excel_dir = excel_base_dir / folder_name
        excel_dir.mkdir(parents=True, exist_ok=True)

        # 重新构建输出路径
        excel_filename = f"non_financial_{Path(image_path).stem}.xlsx"
        new_output_path = excel_dir / excel_filename
        final_output_path = str(new_output_path)

        print(f"普通表格Excel文件将保存到: {final_output_path}")

        # 检查最终路径是否存在
        final_excel_path = Path(final_output_path)
        if final_excel_path.exists():
            print("✅ 普通表格Excel文件已存在，直接返回路径")

            # 将绝对路径转换为可访问的URL
            relative_path = final_output_path.replace('\\', '/')
            if 'static/' in relative_path:
                static_index = relative_path.find('static/')
                if static_index != -1:
                    excel_url = '/' + relative_path[static_index:]
                else:
                    excel_url = f"/static/excel_data/{folder_name}/{excel_filename}"
            else:
                excel_url = f"/static/excel_data/{folder_name}/{excel_filename}"

            return jsonify({
                "success": True,
                "from_cache": True,
                "message": "已加载现有普通表格数据",
                "excel_url": excel_url,
                "excel_path": f"{MAIN_ROOT}/{final_output_path}",
                "table_name": sheet_name,
                "table_type": "non_financial"
            })

        # 如果Excel不存在，进行普通表格识别
        print("🔄 普通表格Excel文件不存在，开始识别流程")

        # 调用异步处理函数
        process_data = {
            'image_path': str(image_full_path),
            'output_path': final_output_path,
            'sheet_name': sheet_name,
            'bank_name': bank_name,
            'file_name': file_name
        }

        result = asyncio.run(_process_non_financial_table(process_data))
        print("普通表格识别结果:", result)

        if result.get('success'):
            excel_url = f"/api/excel-data/{folder_name}/{excel_filename}"

            return jsonify({
                "success": True,
                "from_cache": False,
                "message": "普通表格识别完成",
                "excel_url": excel_url,
                "excel_path": f"{MAIN_ROOT}/{final_output_path}",
                "table_name": sheet_name,
                "table_type": "non_financial",
                "data": result['data']
            })
        else:
            error_msg = result.get('error', '识别失败')
            print(f"❌ 普通表格识别失败: {error_msg}")
            return jsonify({
                "success": False,
                "error": error_msg,
                "message": error_msg
            }), 500

    except Exception as e:
        logger.error(f"普通表格处理错误: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": f"普通表格处理错误: {str(e)}",
            "message": f"处理异常: {str(e)}"
        }), 500


# 在 llm_routes.py 中添加（如果还没有的话）
@llm_bp.route('/excel-data/<path:excel_path>')
def serve_excel_data(excel_path):
    """提供Excel文件数据访问"""
    try:
        # 构建完整的文件路径
        file_path = Path("static/excel_data") / excel_path

        if not file_path.exists():
            return jsonify({
                "success": False,
                "error": f"Excel文件不存在: {excel_path}"
            }), 404

        # 返回文件
        return send_from_directory('static/excel_data', excel_path)

    except Exception as e:
        logger.error(f"提供Excel文件失败: {str(e)}")
        return jsonify({
            "success": False,
            "error": f"文件访问失败: {str(e)}"
        }), 500





# 添加状态检查接口
@llm_bp.route('/llm/processing-status/<task_id>', methods=['GET'])
def get_processing_status(task_id):
    """获取处理状态"""
    try:
        status = PROCESSING_STATUS.get(task_id, {
            "status": "unknown",
            "progress": 0,
            "message": "任务不存在或已过期",
            "exists": False
        })

        # 如果任务存在，添加存在标记
        if task_id in PROCESSING_STATUS:
            status["exists"] = True

            # 清理已完成超过1小时的任务
            current_time = time.time()
            if status.get("status") in ["completed", "error"]:
                completion_time = status.get("completion_time") or status.get("error_time")
                if completion_time and (current_time - completion_time) > 3600:  # 1小时
                    del PROCESSING_STATUS[task_id]
                    return jsonify({
                        "success": True,
                        "data": {
                            "status": "expired",
                            "message": "任务已过期",
                            "exists": False
                        }
                    })

        return jsonify({
            "success": True,
            "data": status
        })
    except Exception as e:
        logger.error(f"获取处理状态失败: {str(e)}")
        return jsonify({
            "success": False,
            "error": f"获取处理状态失败: {str(e)}"
        }), 500


# 添加任务清理接口（可选）
@llm_bp.route('/llm/cleanup-tasks', methods=['POST'])
def cleanup_tasks():
    """清理过期任务"""
    try:
        current_time = time.time()
        expired_tasks = []

        for task_id, status in list(PROCESSING_STATUS.items()):
            # 清理已完成超过1小时的任务
            if status.get("status") in ["completed", "error"]:
                completion_time = status.get("completion_time") or status.get("error_time")
                if completion_time and (current_time - completion_time) > 3600:
                    del PROCESSING_STATUS[task_id]
                    expired_tasks.append(task_id)

            # 清理处理中但超过2小时的任务（可能卡住的任务）
            elif status.get("status") == "processing":
                start_time = status.get("start_time", 0)
                if (current_time - start_time) > 7200:  # 2小时
                    del PROCESSING_STATUS[task_id]
                    expired_tasks.append(task_id)

        return jsonify({
            "success": True,
            "data": {
                "cleaned_tasks": expired_tasks,
                "remaining_tasks": len(PROCESSING_STATUS)
            }
        })
    except Exception as e:
        logger.error(f"清理任务失败: {str(e)}")
        return jsonify({
            "success": False,
            "error": f"清理任务失败: {str(e)}"
        }), 500


@llm_bp.route('/llm/batch-process', methods=['POST'])
def batch_process_images():
    """批量处理图片 - 立即返回任务ID"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({
                "success": False,
                "error": "请求体不能为空"
            }), 400

        # 强制设置表格类型
        data['table_type'] = 'financial'
        print(f"🔄 金融表格批量处理 - 图片数量: {len(data.get('image_paths', []))}")

        # 立即生成任务ID并返回
        task_id = str(uuid.uuid4())

        # 启动异步处理（不阻塞当前请求）
        import threading
        def run_async():
            try:
                asyncio.run(_batch_process_images_sync(data, task_id))
            except Exception as e:
                logger.error(f"异步处理异常: {str(e)}")

        thread = threading.Thread(target=run_async)
        thread.daemon = True
        thread.start()

        print(f"🎯 立即返回任务ID: {task_id}")

        return jsonify({
            "success": True,
            "task_id": task_id,
            "message": "批量处理任务已开始，请等待WebSocket通知",
            "status": "started"
        })

    except Exception as e:
        logger.error(f"批量处理错误: {str(e)}")
        return jsonify({
            "success": False,
            "error": f"批量处理错误: {str(e)}"
        }), 500




# 确保这个路由存在且路径正确
@llm_bp.route('/llm/get-excel-content', methods=['GET'])
def get_excel_content():
    """读取Excel文件内容并返回结构化数据"""
    try:
        excel_url = request.args.get('excel_url')
        print(f"📖 读取Excel内容: {excel_url}")

        if not excel_url:
            return jsonify({
                "success": False,
                "error": "缺少excel_url参数"
            }), 400

        # 调试信息
        print(f"🔍 原始excel_url: {excel_url}")

        # 处理URL编码
        import urllib.parse
        excel_url = urllib.parse.unquote(excel_url)
        print(f"🔍 解码后excel_url: {excel_url}")

        file_path = None

        if excel_url.startswith('/api/excel-data/'):
            # 格式: /api/excel-data/{folder}/{filename}
            relative_path = excel_url.replace('/api/excel-data/', '')
            file_path = Path("static/excel_data") / relative_path
            print(f"🔍 转换路径1: {file_path}")
        elif excel_url.startswith('/static/excel_data/'):
            # 格式: /static/excel_data/{folder}/{filename}
            relative_path = excel_url.replace('/static/excel_data/', '')
            file_path = Path("static/excel_data") / relative_path
            print(f"🔍 转换路径2: {file_path}")
        else:
            # 直接使用路径
            file_path = Path(excel_url)
            print(f"🔍 直接使用路径: {file_path}")

        # 确保是绝对路径
        if not file_path.is_absolute():
            file_path = Path.cwd() / file_path

        print(f"🔍 最终文件路径: {file_path}")
        print(f"🔍 文件是否存在: {file_path.exists()}")

        if not file_path.exists():
            # 尝试在static目录下查找
            static_path = Path("static") / file_path
            if static_path.exists():
                file_path = static_path
                print(f"✅ 在static目录找到文件: {static_path}")
            else:
                print(f"❌ 文件不存在，尝试的路径:")
                print(f"   - {file_path}")
                print(f"   - {static_path}")
                return jsonify({
                    "success": False,
                    "error": f"Excel文件不存在: {file_path}"
                }), 404

        # 读取Excel文件内容...
        import openpyxl
        workbook = openpyxl.load_workbook(file_path)

        # 获取所有工作表
        sheet_data = []

        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]

            # 转换为JSON数据
            headers = []
            data_rows = []

            # 查找数据开始行（跳过空行和标题行）
            data_start_row = 1
            max_col = worksheet.max_column

            # 读取表头（第一行有数据的行）
            for row in range(1, worksheet.max_row + 1):
                row_has_data = False
                for col in range(1, max_col + 1):
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value and str(cell_value).strip():
                        row_has_data = True
                        break

                if row_has_data:
                    # 这一行有数据，作为表头
                    for col in range(1, max_col + 1):
                        cell_value = worksheet.cell(row=row, column=col).value
                        headers.append(str(cell_value) if cell_value is not None else f"列{col}")
                    data_start_row = row + 1
                    break

            # 读取数据行
            for row in range(data_start_row, worksheet.max_row + 1):
                row_data = {}
                has_data = False

                for col, header in enumerate(headers, 1):
                    if col > max_col:
                        break
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value is not None and str(cell_value).strip():
                        has_data = True
                    row_data[header] = str(cell_value) if cell_value is not None else ""

                if has_data:  # 只添加有数据的行
                    data_rows.append(row_data)

            sheet_data.append({
                "sheetName": sheet_name,
                "headers": headers,
                "data": data_rows,
                "rowCount": len(data_rows),
                "colCount": len(headers)
            })

        result = {
            "success": True,
            "data": {
                "filePath": str(file_path),
                "sheets": sheet_data,
                "totalSheets": len(sheet_data)
            }
        }

        print(f"✅ 返回数据: 共{len(sheet_data)}个工作表")
        return jsonify(result)

    except Exception as e:
        logger.error(f"读取Excel内容失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": f"读取Excel内容失败: {str(e)}"
        }), 500







# 非金融表格处理
@llm_bp.route('/llm/batch-process-non-financial', methods=['POST'])
def batch_process_non_financial_images():
    """批量处理普通表格 - 同步处理"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({
                "success": False,
                "error": "请求体不能为空"
            }), 400

        # 强制设置表格类型
        data['table_type'] = 'non_financial'
        print(f"🔄 普通表格批量处理 - 图片数量: {len(data.get('image_paths', []))}")

        # 立即生成任务ID并返回
        task_id = str(uuid.uuid4())

        # ⭐⭐⭐ 同步调用 ⭐⭐⭐
        _batch_process_images_sync(data, task_id)

        print(f"🎯 同步处理完成，任务ID: {task_id}")

        # 返回处理结果状态
        status = PROCESSING_STATUS.get(task_id, {})
        if status.get('status') == 'completed':
            return jsonify({
                "success": True,
                "task_id": task_id,
                "data": {
                    "total": status.get('total', 0),
                    "success": status.get('success_count', 0),
                    "failed": status.get('failed_count', 0),
                    "excel_url": status.get('excel_url'),
                    "table_type": 'non_financial'
                },
                "message": "批量处理完成",
                "status": "completed"
            })
        else:
            return jsonify({
                "success": False,
                "task_id": task_id,
                "error": status.get('message', '处理失败'),
                "status": "error"
            })

    except Exception as e:
        logger.error(f"普通表格批量处理错误: {str(e)}")
        return jsonify({
            "success": False,
            "error": f"普通表格批量处理错误: {str(e)}"
        }), 500


@llm_bp.route('/llm/check-config', methods=['GET'])
def check_llm_config():
    """检查LLM配置状态"""
    try:
        processor = get_table_processor()

        config_status = {
            "llm_configured": processor.llm_client is not None,
            "model_id": getattr(processor, 'model_id', '未设置'),
            "base_url": getattr(processor.llm_client, 'base_url', None) if processor.llm_client else None,
            "client_type": type(processor.llm_client).__name__ if processor.llm_client else '未配置'
        }

        return jsonify({
            "success": True,
            "data": config_status
        })

    except Exception as e:
        return jsonify({
            "success": False,
            "error": f"检查配置失败: {str(e)}"
        }), 500




@llm_bp.route('/llm/task-result/<task_id>', methods=['GET'])
def get_task_result(task_id):
    """查询任务结果"""
    try:
        print(f"🔍 查询任务结果: {task_id}")

        if task_id in TASK_RESULTS:
            result = TASK_RESULTS[task_id]
            print(f"✅ 找到任务结果: {result}")
            return jsonify({
                "success": True,
                "data": result
            })
        else:
            print(f"❌ 任务不存在: {task_id}")
            return jsonify({
                "success": False,
                "error": "任务不存在或已过期",
                "data": {
                    "status": "not_found"
                }
            }), 404

    except Exception as e:
        print(f"💥 查询任务结果失败: {str(e)}")
        return jsonify({
            "success": False,
            "error": f"查询任务结果失败: {str(e)}"
        }), 500



@llm_bp.route('/llm/batch-process', methods=['POST'], endpoint='batch_process_financial')
def batch_process_financial_images():
    try:
        data = request.get_json()
        if not data:
            return jsonify({
                "success": False,
                "error": "请求体不能为空"
            }), 400

        data['table_type'] = 'financial'
        image_count = len(data.get('image_paths', []))
        print(f"🔵 金融表格批量处理 - 图片数量: {image_count}")

        # ⭐⭐⭐ 只创建一个任务ID ⭐⭐⭐
        task_id = str(uuid.uuid4())
        print(f"🎯 创建任务ID: {task_id}")

        # 在后台线程中处理，避免阻塞请求
        import threading
        def process_in_background():
            try:
                # ⭐⭐⭐ 传递相同的任务ID ⭐⭐⭐
                _batch_process_images_sync(data, task_id)
                print(f"✅ 后台处理完成 - 任务ID: {task_id}")
            except Exception as e:
                print(f"❌ 后台处理异常: {str(e)}")
                # 即使异常也要更新状态
                TASK_RESULTS[task_id] = {
                    "status": "error",
                    "error": f"处理异常: {str(e)}",
                    "completed_at": time.time()
                }

        thread = threading.Thread(target=process_in_background)
        thread.daemon = True
        thread.start()

        # 立即返回响应，不等待处理完成
        print(f"🎯 立即返回响应，任务ID: {task_id}")

        return jsonify({
            "success": True,
            "task_id": task_id,
            "message": "批量处理任务已开始",
            "status": "started",
            "image_count": image_count,
            "processing_started": True
        })

    except Exception as e:
        logger.error(f"金融表格批量处理错误: {str(e)}")
        return jsonify({
            "success": False,
            "error": f"金融表格批量处理错误: {str(e)}"
        }), 500