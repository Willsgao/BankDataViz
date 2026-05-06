# -*- coding: UTF-8 -*-
"""
智能识别 API：触发 DeepSeek 自动化 + 保存 Excel
"""
import base64
import hashlib
import io
import json
import logging
import os
import re
import uuid
from datetime import datetime

from flask import Blueprint, request, jsonify
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell

logger = logging.getLogger(__name__)
smart_recognize_bp = Blueprint("smart_recognize", __name__)

# --------------------------------------------------------------------------- #
# 路径配置（复用现有路径）
# --------------------------------------------------------------------------- #
try:
    from backend.configs.config import config
    EXCEL_OUTPUT_ROOT = config.EXCEL_DATA_FOLDER
except Exception:
    EXCEL_OUTPUT_ROOT = os.path.join(os.getcwd(), "data/backend/static/excel_data")


# --------------------------------------------------------------------------- #
# 工具函数
# --------------------------------------------------------------------------- #

def parse_markdown_table(text: str):
    """
    从文本中提取所有 Markdown 表格。
    支持多种格式：标准 Markdown 表格、无分隔行的表格、空格/制表符分隔的数据。
    返回 list[list[list[str]]]：每个表格 → 每行 → 每列的字符串
    """
    tables = []
    lines = text.strip().split("\n")
    logger.info(f"[parse] 开始解析文本，共 {len(lines)} 行")

    # 方法1：解析标准 Markdown 表格（有 | 分隔符）
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        if "|" in line and line.count("|") >= 2:
            table_rows = []
            while i < len(lines) and "|" in lines[i]:
                row = lines[i].strip()
                if re.match(r"^\|[\s\-:|]+\|$", row):
                    i += 1
                    continue
                cells = [c.strip() for c in row.strip("|").split("|")]
                table_rows.append(cells)
                i += 1
            if table_rows:
                tables.append(table_rows)
                logger.info(f"[parse] Markdown 表格解析成功，共 {len(table_rows)} 行")
        else:
            i += 1

    # 方法2：如果没有找到表格，尝试解析空格/制表符分隔的数据
    if not tables:
        logger.info("[parse] 未找到 Markdown 表格，尝试解析空格/制表符分隔的数据")
        table_rows = []
        for line in lines:
            line = line.strip()
            if not line or line.startswith("#") or line.startswith("**"):
                if table_rows:
                    tables.append(table_rows)
                    table_rows = []
                continue
            # 尝试用制表符、多个空格、逗号分隔
            if "\t" in line:
                cells = [c.strip() for c in line.split("\t")]
            elif "," in line and line.count(",") >= 2:
                cells = [c.strip() for c in line.split(",")]
            else:
                # 多个连续空格分隔
                parts = re.split(r"\s{2,}", line)
                cells = [p.strip() for p in parts if p.strip()]
                # 如果还是只有1列，尝试用单个空格分隔（适用于"资产总额 40,571,149"格式）
                if len(cells) < 2:
                    cells = [p.strip() for p in line.split() if p.strip()]
            if len(cells) >= 2:  # 至少2列才认为是表格行
                table_rows.append(cells)
        if table_rows:
            tables.append(table_rows)
            logger.info(f"[parse] 空格/制表符表格解析成功，共 {len(table_rows)} 行")

    # 方法3：如果还是没有表格，尝试从纯文本中提取数字表格
    if not tables:
        logger.info("[parse] 尝试提取数字表格")
        table_rows = []
        for line in lines:
            line = line.strip()
            # 如果一行包含多个数字，认为是表格行
            numbers = re.findall(r"-?\d+\.?\d*", line)
            if len(numbers) >= 2:
                # 尝试按多个连续空格或制表符分隔
                parts = re.split(r"\s{2,}|\t", line)
                cells = [p.strip() for p in parts if p.strip()]
                if len(cells) < 2:
                    cells = [p.strip() for p in line.split() if p.strip()]
                if len(cells) >= 2:
                    table_rows.append(cells)
        if table_rows:
            tables.append(table_rows)
            logger.info(f"[parse] 数字表格解析成功，共 {len(table_rows)} 行")
        if table_rows:
            tables.append(table_rows)
            logger.info(f"[parse] 数字表格解析成功，共 {len(table_rows)} 行")

    if not tables:
        logger.warning("[parse] 未能解析出任何表格，将作为纯文本处理")

    return tables


# --------------------------------------------------------------------------- #
# API 路由
# --------------------------------------------------------------------------- #

@smart_recognize_bp.route("/api/smart-recognize/send", methods=["POST"])
def send_to_deepseek():
    """
    接收截图 base64，触发 DeepSeek 自动化，返回识别结果。

    Body: { image_base64: str, prompt: str, user_data_dir: str }
    """
    data = request.get_json()
    image_base64 = data.get("image_base64", "")
    prompt = data.get("prompt", "请识别这张图片中的表格内容，保持原有格式输出Markdown表格")
    user_data_dir = data.get("user_data_dir", "")

    if not user_data_dir:
        from backend.utils.constants import MAIN_ROOT
        user_data_dir = os.path.join(MAIN_ROOT, "data", "backend", "data", "chrome_profile")

    if not image_base64:
        return jsonify({"success": False, "error": "image_base64 不能为空"}), 400

    try:
        from backend.services.deepseek_automation import DeepSeekAutomation, PLAYWRIGHT_AVAILABLE
        if not PLAYWRIGHT_AVAILABLE:
            return jsonify({
                "success": False,
                "error": "playwright 未安装，请运行: pip install playwright && playwright install chromium"
            }), 500

        automation = DeepSeekAutomation(user_data_dir=user_data_dir or None)
        result = automation.recognize(image_base64, prompt)
        return jsonify(result)

    except Exception as e:
        logger.exception("DeepSeek 识别失败")
        return jsonify({"success": False, "error": str(e)}), 500


@smart_recognize_bp.route("/api/smart-recognize/save-excel", methods=["POST"])
def save_excel():
    """
    将 LLM 返回的文本保存为 Excel。
    自动识别 Markdown 表格并写入，否则直接写入纯文本。

    Body: {
        content: str,           # LLM 返回的文本内容
        filename: str,          # 保存文件名（不含扩展名）
        sheet_name: str         # Sheet 名称
    }
    """
    data = request.get_json()
    content = data.get("content", "")
    filename = data.get("filename", f"智能识别_{datetime.now().strftime('%Y%m%d_%H%M%S')}")
    sheet_name = data.get("sheet_name", "Sheet1")

    if not content:
        return jsonify({"success": False, "error": "content 不能为空"}), 400

    os.makedirs(EXCEL_OUTPUT_ROOT, exist_ok=True)
    file_path = os.path.join(EXCEL_OUTPUT_ROOT, f"{filename}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]  # Excel sheet 名最长 31 字符

    # 先过滤掉解释性文字，只保留表格数据
    lines = content.split("\n")
    filtered = []
    for line in lines:
        s = line.strip()
        if not s:
            continue
        has_table_marker = ("|" in s) or re.search(r"\d", s)
        is_noise = len(s) < 30 and any(kw in s for kw in ["您好", "谢谢", "希望", "请问", "好的", "当然", "这是"])
        if has_table_marker and not is_noise:
            filtered.append(s)
    content = "\n".join(filtered)
    logger.info(f"[save_excel] 过滤后内容长度: {len(content)}")

    tables = parse_markdown_table(content)

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    alt_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    border_thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    cell_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    current_row = 1

    if tables:
        for table_idx, table in enumerate(tables):
            for row_idx, row in enumerate(table):
                for col_idx, cell_val in enumerate(row):
                    cell = ws.cell(row=current_row, column=col_idx + 1, value=cell_val)
                    cell.alignment = cell_align
                    cell.border = border_thin

                    if row_idx == 0:
                        # 表头行
                        cell.fill = header_fill
                        cell.font = header_font
                    elif row_idx % 2 == 0:
                        cell.fill = alt_fill

                current_row += 1

            current_row += 1  # 表格之间空一行
    else:
        # 没有表格，写入纯文本（按行）
        for line in content.split("\n"):
            ws.cell(row=current_row, column=1, value=line)
            current_row += 1

    # 自动调整列宽
    for col_idx, col in enumerate(ws.columns, 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for cell in col:
            if isinstance(cell, MergedCell):
                continue
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    wb.save(file_path)
    logger.info(f"Excel 已保存: {file_path}")

    return jsonify({
        "success": True,
        "file_path": file_path,
        "filename": f"{filename}.xlsx",
        "url": f"/static/excel_data/{filename}.xlsx"
    })


# --------------------------------------------------------------------------- #
# 表格区域检测 API
# --------------------------------------------------------------------------- #

@smart_recognize_bp.route("/api/smart-recognize/detect-tables", methods=["POST"])
def detect_tables():
    """
    上传文件并检测表格区域坐标。

    Body (multipart/form-data):
        file: PDF/Excel/图片文件
        dpi: 渲染 DPI（默认 150，仅 PDF 有效）

    Returns:
        {
            "success": bool,
            "cached": bool,              # 是否命中缓存
            "file_id": str,              # 本次上传的临时文件 ID
            "file_type": "pdf" | "excel" | "image",
            "total_pages": int,
            "pages": [...],
            "total_tables": int,
            "error": str | None,
        }
    """
    try:
        from backend.services.table_region_detector import detect_tables
        from werkzeug.utils import secure_filename

        if "file" not in request.files:
            return jsonify({"success": False, "error": "未上传文件"}), 400

        file = request.files["file"]
        if not file.filename:
            return jsonify({"success": False, "error": "文件名为空"}), 400

        dpi = int(request.form.get("dpi", 150))

        # 计算文件 MD5
        file_bytes = file.read()
        file_md5 = hashlib.md5(file_bytes).hexdigest()
        file.seek(0)

        # 建立缓存目录
        from backend.utils.constants import MAIN_ROOT
        cache_dir = os.path.join(MAIN_ROOT, "data", "backend", "static", "detect_cache")
        os.makedirs(cache_dir, exist_ok=True)

        # 尝试命中缓存（key = MD5 + dpi）
        cache_key = f"{file_md5}_{dpi}"
        cache_path = os.path.join(cache_dir, f"{cache_key}.json")
        if os.path.exists(cache_path):
            with open(cache_path, "r", encoding="utf-8") as f:
                cached = json.load(f)
            cached["cached"] = True
            logger.info(f"命中缓存: {file.filename} ({file_md5})")
            return jsonify(cached)

        # 未命中缓存，正常处理
        import uuid as _uuid

        temp_dir = os.path.join(MAIN_ROOT, "data", "backend", "static", "temp_uploads")
        os.makedirs(temp_dir, exist_ok=True)

        file_id = str(_uuid.uuid4())[:8]
        ext = os.path.splitext(secure_filename(file.filename))[-1]
        temp_path = os.path.join(temp_dir, f"{file_id}{ext}")
        file.save(temp_path)

        from pathlib import Path
        result = detect_tables(Path(temp_path), dpi=dpi)
        result["file_id"] = file_id
        result["temp_path"] = temp_path
        result["cached"] = False

        # 保存到缓存（去掉不可序列化的字段）
        cacheable = {k: v for k, v in result.items()
                     if k not in ("temp_path", "error")}
        with open(cache_path, "w", encoding="utf-8") as f:
            json.dump(cacheable, f, ensure_ascii=False)

        return jsonify(result)

    except Exception as e:
        logger.exception("表格检测失败")
        return jsonify({"success": False, "error": str(e)}), 500


# --------------------------------------------------------------------------- #
# 批量 DeepSeek 自动化 API
# --------------------------------------------------------------------------- #

@smart_recognize_bp.route("/api/smart-recognize/batch-recognize", methods=["POST"])
def batch_recognize():
    """
    批量将多个截图依次发给 DeepSeek，共用同一个浏览器 session。

    Body:
        {
            "regions": [
                {
                    "id": str,
                    "image_base64": str,
                    "label": str,        # 区域标签
                },
                ...
            ],
            "prompt": str,
            "user_data_dir": str,
        }

    Returns:
        {
            "success": bool,
            "results": [...],
            "total": int,
            "succeeded": int,
            "failed": int,
        }
    """
    data = request.get_json()
    regions = data.get("regions", [])
    prompt = data.get(
        "prompt",
        "请识别这张图片中的表格内容，保持原有格式输出Markdown表格"
    )
    user_data_dir = data.get("user_data_dir", "")

    if not user_data_dir:
        from backend.utils.constants import MAIN_ROOT
        user_data_dir = os.path.join(MAIN_ROOT, "data", "backend", "data", "chrome_profile")

    if not regions:
        return jsonify({"success": False, "error": "regions 不能为空"}), 400

    try:
        from backend.services.deepseek_automation import DeepSeekAutomation, PLAYWRIGHT_AVAILABLE

        if not PLAYWRIGHT_AVAILABLE:
            return jsonify({
                "success": False,
                "error": "playwright 未安装，请运行: pip install playwright && playwright install chromium"
            }), 500

        automation = DeepSeekAutomation(user_data_dir=user_data_dir or None)
        # 复用同一个浏览器 session，只启动一次
        automation.launch(headless=False)
        automation.goto_deepseek()

        results = []
        succeeded = 0
        failed = 0

        for region in regions:
            region_id = region.get("id", "")
            label = region.get("label", region_id)
            image_b64 = region.get("image_base64", "")

            if not image_b64:
                results.append({
                    "id": region_id,
                    "label": label,
                    "success": False,
                    "result": "",
                    "error": "image_base64 为空",
                })
                failed += 1
                continue

            try:
                res = automation.recognize_one(image_b64, prompt)
                results.append({
                    "id": region_id,
                    "label": label,
                    "success": res.get("success", False),
                    "result": res.get("result", ""),
                    "error": res.get("error"),
                })
                if res.get("success"):
                    succeeded += 1
                else:
                    failed += 1
            except Exception as e:
                logger.exception(f"区域 {label} 识别失败")
                results.append({
                    "id": region_id,
                    "label": label,
                    "success": False,
                    "result": "",
                    "error": str(e),
                })
                failed += 1

        automation.close()
        logger.info(f"批量识别完成：{succeeded}/{len(results)} 成功")

        return jsonify({
            "success": True,
            "results": results,
            "total": len(results),
            "succeeded": succeeded,
            "failed": failed,
        })

    except Exception as e:
        logger.exception("批量识别失败")
        return jsonify({"success": False, "error": str(e)}), 500


# --------------------------------------------------------------------------- #
# 批量保存 Excel
# --------------------------------------------------------------------------- #

def _split_tables_by_marker(text: str):
    """
    按「表格1」「表格2」等标记拆分文本，返回 [(表名, 表格文本), ...]。
    如果没有标记，整段作为第一个表格。
    """
    pattern = re.compile(r"\*{0,2}表格\s*([0-9一二三四五六七八九十]+)\*{0,2}")
    lines = text.split("\n")
    segments = []
    current_segment = []
    current_table_name = None   # 初始为 None，遇到第一个标记才赋值

    for line in lines:
        m = pattern.search(line)
        if m:
            # 遇到"表格X"标记
            if current_table_name and current_segment:
                # 前一个表格有实质内容，保存它
                seg_text = "\n".join(current_segment)
                # 只保留包含 | 或 \t 的行（过滤掉描述文字）
                seg_lines = seg_text.split("\n")
                filtered_lines = [sl for sl in seg_lines if "|" in sl or "\t" in sl]
                if filtered_lines:
                    segments.append((current_table_name, "\n".join(filtered_lines)))
            current_segment = []
            current_table_name = f"表格{m.group(1)}"
            # 如果这一行除了标记之外还有内容，加入当前段
            rest = line.replace(m.group(0), "").strip()
            if rest:
                current_segment.append(rest)
        else:
            current_segment.append(line)

    # 保存最后一个表格
    if current_table_name and current_segment:
        seg_text = "\n".join(current_segment)
        seg_lines = seg_text.split("\n")
        filtered_lines = [sl for sl in seg_lines if "|" in sl or "\t" in sl]
        if filtered_lines:
            segments.append((current_table_name, "\n".join(filtered_lines)))
    elif not segments and current_segment:
        # 没有找到任何"表格X"标记，整段作为第一个表格
        seg_text = "\n".join(current_segment)
        seg_lines = seg_text.split("\n")
        filtered_lines = [sl for sl in seg_lines if "|" in sl or "\t" in sl]
        if filtered_lines:
            segments.append(("表格1", "\n".join(filtered_lines)))

    return segments


def _parse_single_table(text: str):
    """
    解析单个表格的文本内容，返回 list[list[str]]（每行每列）。
    只处理含 | 或 \t 的行，忽略所有其他内容（说明文字、统计结果等）。

    关键：Markdown 表格中空单元格是有意义的（表示跨列），
    必须以首行的列数为准，所有行都 pad 到相同列数。
    """
    lines = text.strip().split("\n")
    raw_rows = []  # 保留原始列数（包括空字符串）

    for line in lines:
        s = line.strip()
        if not s:
            continue
        # 跳过 Markdown 分隔行（| --- | --- |）
        if re.match(r"^[\s|:\-]+$", s):
            continue
        # 只处理含 | 或 \t 的行（真正的表格行）
        if "\t" in line:
            cells = [c.strip() for c in line.split("\t")]
            raw_rows.append(cells)
        elif "|" in line:
            # 按 | 分割，保留所有列（包括空字符串）
            # 先用 strip('|') 去掉行首行尾的 |，再按 | 分割
            cleaned = s.strip().strip("|")
            cells = [c.strip() for c in cleaned.split("|")]
            raw_rows.append(cells)
        # 不含 | 或 \t，跳过（说明文字等）

    if not raw_rows:
        return []

    # 以第一行的列数为准（Markdown 表格的所有行应有相同列数）
    expected_cols = len(raw_rows[0])

    # 所有行 pad 到 expected_cols
    table_rows = []
    for row in raw_rows:
        if len(row) < expected_cols:
            row = row + [""] * (expected_cols - len(row))
        elif len(row) > expected_cols:
            row = row[:expected_cols]
        table_rows.append(row)

    # 如果首列全部为空，删除该列（Markdown 表格首列空的常见情况）
    if table_rows and all((row[0] or "") == "" for row in table_rows):
        table_rows = [row[1:] for row in table_rows]

    return table_rows


def _try_convert_number(cell_val):
    """
    尝试将字符串转换为数字。
    支持：整数、小数、千分位逗号、括号表示负数。
    返回转换后的值；转换失败则返回原字符串。
    """
    if not isinstance(cell_val, str):
        return cell_val
    cleaned = cell_val.strip().replace(",", "")
    try:
        test_str = cleaned.replace(".", "").replace("-", "").replace("(", "").replace(")", "")
        if test_str.isdigit():
            if cleaned.startswith("(") and cleaned.endswith(")"):
                num_val = -float(cleaned[1:-1])
            else:
                num_val = float(cleaned)
            if num_val == int(num_val):
                num_val = int(num_val)
            return num_val
    except (ValueError, AttributeError):
        pass
    return cell_val


@smart_recognize_bp.route("/api/smart-recognize/batch-save-excel", methods=["POST"])
def batch_save_excel():
    """
    将识别结果按表格拆分为多个 Sheet 保存为一个 Excel 文件。

    每个 result 中可能包含多个表格（DeepSeek 返回"表格1"、"表格2"等），
    每个表格单独一个 Sheet，命名规则：第{page_num}页_表格{Y}

    Body:
        {
            "results": [
                {
                    "id": str,
                    "label": str,
                    "result": str,        # DeepSeek 返回的 Markdown 文本
                    "page_num": int,     # 页码（从1开始）
                },
                ...
            ],
            "filename": str,         # 保存文件名（不含扩展名）
        }
    """
    data = request.get_json()
    if data is None:
        return jsonify({"success": False, "error": "请求体不是有效的 JSON"}), 400
    results = data.get("results", [])
    filename = data.get(
        "filename",
        f"智能识别_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    )

    if not results:
        return jsonify({"success": False, "error": "results 不能为空"}), 400

    os.makedirs(EXCEL_OUTPUT_ROOT, exist_ok=True)
    file_path = os.path.join(EXCEL_OUTPUT_ROOT, f"{filename}.xlsx")

    # 保存原始数据到 txt 文件
    raw_txt_path = os.path.join(EXCEL_OUTPUT_ROOT, f"{filename}_raw.txt")
    try:
        with open(raw_txt_path, "w", encoding="utf-8") as f:
            for region in results:
                label = region.get("label", region.get("id", ""))
                page_num = region.get("page_num", 1)
                content = region.get("result", "")
                f.write(f"===== 第{page_num}页: {label} =====\n")
                f.write(content)
                f.write("\n\n")
        logger.info(f"原始数据已保存: {raw_txt_path}")
    except Exception as e:
        logger.warning(f"保存原始数据失败: {e}")

    wb = Workbook()
    wb.remove(wb.active)  # 删除默认 Sheet，后面按需创建

    # 样式
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    border_thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    # 不再使用统一的 cell_align，改为按列动态判断对齐方式

    for region in results:
        label = region.get("label", region.get("id", ""))
        content = region.get("result", "")
        page_num = region.get("page_num", 1)

        if not content:
            continue

        # 按"表格1""表格2"拆分
        segments = _split_tables_by_marker(content)
        logger.info(f"[batch_save] {label} 拆分出 {len(segments)} 个表格")

        for seg_idx, (table_name, table_text) in enumerate(segments, 1):
            table_rows = _parse_single_table(table_text)

            if not table_rows:
                logger.warning(f"[batch_save] {table_name} 解析结果为空，跳过")
                continue

            # Sheet 名：第{page_num}页_表格{seg_idx}（最长31字符）
            sheet_name = f"第{page_num}页_表格{seg_idx}"
            sheet_name = sheet_name[:31]

            ws = wb.create_sheet(title=sheet_name)

            # 写入表格数据（先不设置对齐，等判断列类型后统一设置）
            for row_idx, row in enumerate(table_rows):
                for col_idx, cell_val in enumerate(row):
                    val = _try_convert_number(cell_val)
                    cell = ws.cell(row=row_idx + 1, column=col_idx + 1, value=val)
                    cell.border = border_thin
                    if row_idx == 0:
                        cell.fill = header_fill
                        cell.font = header_font

            # 按列判断数据类型，决定对齐方式
            # 规则：数据行中超过一半单元格是数字 → 数字列（右对齐），否则文字列（左对齐）
            header_row_count = min(3, len(table_rows))  # 前3行视为表头
            col_types = []  # "numeric" or "text"
            for col_idx in range(ws.max_column):
                numeric_count = 0
                total_count = 0
                for row_idx in range(header_row_count, len(table_rows)):
                    cell_val = ws.cell(row=row_idx + 1, column=col_idx + 1).value
                    if cell_val is not None and cell_val != "":
                        total_count += 1
                        if isinstance(cell_val, (int, float)):
                            numeric_count += 1
                if total_count > 0 and numeric_count / total_count > 0.5:
                    col_types.append("numeric")
                else:
                    col_types.append("text")

            # 按列类型设置对齐方式
            for row_idx in range(len(table_rows)):
                for col_idx in range(ws.max_column):
                    cell = ws.cell(row=row_idx + 1, column=col_idx + 1)
                    if row_idx < header_row_count:
                        # 表头：数字列右对齐，文字列居中
                        if col_types[col_idx] == "numeric":
                            cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
                        else:
                            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    else:
                        # 数据行：数字列右对齐，文字列左对齐
                        if col_types[col_idx] == "numeric":
                            cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
                        else:
                            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

            # 自动调整列宽
            for col_idx, col in enumerate(ws.columns, 1):
                max_len = 0
                col_letter = get_column_letter(col_idx)
                for cell in col:
                    if isinstance(cell, MergedCell):
                        continue
                    try:
                        max_len = max(max_len, len(str(cell.value or "")))
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

            # 自动合并表头单元格（前3行）：将有值单元格右侧的连续空单元格合并
            header_rows = min(3, len(table_rows))
            for r in range(1, header_rows + 1):
                c = 1
                while c <= ws.max_column:
                    cell_val = ws.cell(row=r, column=c).value
                    if cell_val and c < ws.max_column:
                        # 向右查找连续为空的单元格
                        merge_end = c
                        while merge_end < ws.max_column and (ws.cell(row=r, column=merge_end + 1).value or "") == "":
                            merge_end += 1
                        if merge_end > c:
                            ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=merge_end)
                            # 合并单元格的对齐方式：数字列右对齐，文字列居中
                            if col_types[c - 1] == "numeric":
                                ws.cell(row=r, column=c).alignment = Alignment(horizontal="right", vertical="center")
                            else:
                                ws.cell(row=r, column=c).alignment = Alignment(horizontal="center", vertical="center")
                        c = merge_end + 1
                    else:
                        c += 1

            logger.info(f"[batch_save] Sheet '{sheet_name}' 写入 {len(table_rows)} 行数据")
    if len(wb.sheetnames) == 0:
        wb.create_sheet(title="无数据")

    wb.save(file_path)
    logger.info(f"批量 Excel 已保存: {file_path}，共 {len(wb.sheetnames)} 个 Sheet")

    return jsonify({
        "success": True,
        "file_path": file_path,
        "filename": f"{filename}.xlsx",
        "url": f"/static/excel_data/{filename}.xlsx",
        "sheet_count": len(wb.sheetnames),
    })
