"""
Excel数据处理器 - 重构自file.py中的Excel数据处理相关功能
"""

from pathlib import Path
import json
import time
from typing import Dict, Any, List, Tuple


class ExcelDataHandler:
    """Excel数据处理器"""

    def __init__(self, main_root: str, excel_output_root: str):
        """
        初始化处理器

        Args:
            main_root: 主根目录
            excel_output_root: Excel输出根目录
        """
        self.MAIN_ROOT = main_root
        self.EXCEL_OUTPUT_ROOT = excel_output_root

    def get_correct_pdf_id(self, pdf_id: str, db) -> str:
        """获取正确的PDF ID（UUID格式）"""
        # 如果已经是UUID格式，直接返回

        if pdf_id.endswith('.pdf'):
            pdf_id = pdf_id[:-4]

        if not pdf_id.isdigit():
            return pdf_id

        # 如果是数字ID，查询对应的UUID
        try:
            conn = db.connect()
            c = conn.cursor()

            # 🔥🔥🔥 修复：使用正确的列名 filename
            query = "SELECT filename FROM files WHERE id = ? AND deleted = 0"
            params = (int(pdf_id),)  # 确保是整数

            print(f"🔍 查询UUID: {query} 参数: {params}")
            c.execute(query, params)
            row = c.fetchone()
            conn.close()

            if row:
                uuid = row["filename"]
                print(f"✅ 数字ID {pdf_id} 对应的UUID: {uuid}")
                if uuid.endswith('.pdf'):
                    uuid = uuid[:-4]
                return uuid
            else:
                print(f"⚠️ 未找到数字ID {pdf_id} 对应的UUID，使用原ID")
                return pdf_id

        except Exception as e:
            print(f"❌ 查询UUID失败: {e}，使用原ID: {pdf_id}")
            import traceback
            traceback.print_exc()
            return pdf_id

    def save_complete_table_data(self, pdf_id: str, excel_file: str, sheet_name: str,
                                 table_data: List[List], table_type: str, db) -> Dict[str, Any]:
        """保存完整表格数据 - 只覆盖目标Sheet"""
        print("📊 保存完整表格数据（保护其他Sheet）...")

        # 🔥 在函数开头初始化变量
        protected_sheets = []

        try:
            # 1. 获取Excel文件路径
            pdf_id = self.get_correct_pdf_id(pdf_id, db)
            excel_dir = Path(self.MAIN_ROOT) / self.EXCEL_OUTPUT_ROOT / pdf_id
            excel_path = excel_dir / excel_file

            print(f"📁 Excel文件路径: {excel_path}")

            if not excel_path.exists():
                return {'success': False, 'error': f'Excel文件不存在: {excel_path}'}

            # 2. 加载工作簿
            from openpyxl import load_workbook
            workbook = load_workbook(excel_path)

            # 🔥 记录所有Sheet，确保保护其他Sheet
            all_sheets = workbook.sheetnames.copy()
            print(f"📋 工作簿包含 {len(all_sheets)} 个Sheet: {all_sheets}")
            print(f"🎯 目标Sheet: {sheet_name}")

            if sheet_name not in all_sheets:
                workbook.close()
                return {'success': False, 'error': f'Sheet不存在: {sheet_name}'}

            # 3. 🔥 只操作目标Sheet，保护其他Sheet
            worksheet = workbook[sheet_name]
            print(f"📈 目标Sheet原维度: {worksheet.max_row}行 × {worksheet.max_column}列")
            print(f"📦 新数据维度: {len(table_data)}行 × {len(table_data[0]) if table_data else 0}列")

            # 4. 🔥 只清空目标Sheet的数据（保留表头结构）
            if worksheet.max_row > 0:
                # 删除所有行（包括表头）
                worksheet.delete_rows(1, worksheet.max_row)
                print(f"🗑️ 清空目标Sheet: 删除了{worksheet.max_row}行")

            # 5. 🔥 写入完整数据到目标Sheet
            for row_idx, row_data in enumerate(table_data, 1):  # 从第1行开始
                for col_idx, cell_value in enumerate(row_data, 1):
                    worksheet.cell(row=row_idx, column=col_idx, value=cell_value)

            print(f"📝 写入目标Sheet: {len(table_data)}行 × {len(table_data[0])}列")

            # 6. 🔥 保存工作簿（其他Sheet自动保留）
            workbook.save(excel_path)
            workbook.close()

            # 7. 🔥 验证其他Sheet是否被保护
            workbook_after = load_workbook(excel_path)
            sheets_after = workbook_after.sheetnames
            workbook_after.close()

            # 🔥 计算受保护的Sheet
            protected_sheets = [s for s in all_sheets if s != sheet_name]

            print(f"✅ 保存完成，验证Sheet保护:")
            print(f"  保存前Sheet数: {len(all_sheets)}")
            print(f"  保存后Sheet数: {len(sheets_after)}")
            print(f"  Sheet保持一致: {set(all_sheets) == set(sheets_after)}")

            if protected_sheets:
                print(f"  🛡️ 受保护的Sheet: {protected_sheets}")

            # 8. 🔥 保存快照
            print("📸 开始保存数据快照...")
            snapshot_result = self.save_data_snapshot(pdf_id, excel_file, sheet_name, table_data, table_type)
            print(f"✅ 快照保存结果: {snapshot_result.get('success', False)}")

            print("✅ 完整表格数据保存成功（其他Sheet已保护）")

            return {
                'success': True,
                'saved_rows': len(table_data),
                'saved_columns': len(table_data[0]) if table_data else 0,
                'data_dimensions': f'{len(table_data)}行 × {len(table_data[0])}列',
                'excel_updated': True,
                'sheets_protected': True,
                'protected_sheets_count': len(protected_sheets),
                'snapshot_saved': snapshot_result.get('success', False),
                'snapshot_path': snapshot_result.get('path', '')
            }

        except Exception as e:
            print(f"❌ 完整表格保存失败: {e}")
            import traceback
            traceback.print_exc()
            return {'success': False, 'error': f'完整表格保存失败: {str(e)}'}

    def save_data_snapshot(self, pdf_id: str, excel_file: str, sheet_name: str,
                           data: List[List], table_type: str) -> Dict[str, Any]:
        """保存数据快照"""
        try:
            # 🔥 使用默认路径
            SNAPSHOT_ROOT = "data/backend/static/modify_data"

            # 创建快照目录
            snap_dir = Path(self.MAIN_ROOT) / SNAPSHOT_ROOT / pdf_id
            snap_dir.mkdir(parents=True, exist_ok=True)

            # 生成时间戳文件名
            ts = int(time.time())
            snap_file = f"{pdf_id}_{sheet_name}_{table_type}_{ts}.json"
            snap_path = snap_dir / snap_file

            # 准备快照数据
            snapshot_data = {
                'pdf_id': pdf_id,
                'excel_file': excel_file,
                'sheet_name': sheet_name,
                'table_type': table_type,
                'data': data,  # 保存原始数据
                'saved_at': ts,
                'data_dimensions': {
                    'rows': len(data),
                    'columns': len(data[0]) if data else 0
                }
            }

            # 保存快照
            with open(snap_path, 'w', encoding='utf-8') as f:
                json.dump(snapshot_data, f, ensure_ascii=False, indent=2)

            print(f"✅ 数据快照已保存: {snap_path}")

            return {
                'success': True,
                'path': str(snap_path),
                'file': snap_file
            }

        except Exception as e:
            print(f"⚠️ 快照保存失败: {e}")
            return {
                'success': False,
                'error': str(e)
            }


    def save_flattened_table_data(self, pdf_id: str, excel_file: str, sheet_name: str,
                                  table_data: List[List], table_type: str, db) -> Dict[str, Any]:
        """保存扁平化表格数据 - 只覆盖目标Sheet"""
        print("📊 保存扁平化表格数据（保护其他Sheet）...")

        try:
            pdf_id = self.get_correct_pdf_id(pdf_id, db)
            excel_dir = Path(self.MAIN_ROOT) / self.EXCEL_OUTPUT_ROOT / pdf_id
            excel_dir.mkdir(parents=True, exist_ok=True)
            excel_path = excel_dir / excel_file

            print(f"📁 Excel文件路径: {excel_path}")

            from openpyxl import Workbook, load_workbook

            file_exists = excel_path.exists()
            all_sheets = []

            if file_exists:
                # 🔥 文件存在，加载现有工作簿（保护其他Sheet）
                workbook = load_workbook(excel_path)
                all_sheets = workbook.sheetnames.copy()
                print(f"📄 加载现有Excel文件，包含 {len(all_sheets)} 个Sheet: {all_sheets}")
            else:
                # 文件不存在，创建新工作簿
                workbook = Workbook()
                # 删除默认Sheet
                default_sheet = workbook.active
                workbook.remove(default_sheet)
                print("📄 创建新Excel文件")

            # 🔥 处理目标Sheet（不影响其他Sheet）
            if sheet_name in workbook.sheetnames:
                print(f"📋 Sheet已存在，删除重写: {sheet_name}")
                del workbook[sheet_name]

            # 创建/重写目标Sheet
            worksheet = workbook.create_sheet(sheet_name)
            print(f"✅ 创建/重写目标Sheet: {sheet_name}")

            # 写入完整数据到目标Sheet
            for row_idx, row_data in enumerate(table_data, 1):
                for col_idx, cell_value in enumerate(row_data, 1):
                    worksheet.cell(row=row_idx, column=col_idx, value=cell_value)

            print(f"📝 写入目标Sheet: {len(table_data)}行 × {len(table_data[0])}列")

            # 保存工作簿
            workbook.save(excel_path)
            workbook.close()

            # 验证Sheet保护
            if file_exists and all_sheets:
                workbook_after = load_workbook(excel_path)
                sheets_after = workbook_after.sheetnames
                workbook_after.close()

                protected_sheets = [s for s in all_sheets if s != sheet_name]
                print(f"✅ Sheet保护验证:")
                print(f"  保存前Sheet: {len(all_sheets)}个")
                print(f"  保存后Sheet: {len(sheets_after)}个")
                print(f"  受保护Sheet: {len(protected_sheets)}个")
                print(f"  Sheet列表: {protected_sheets}")

            action = "创建" if not file_exists else "更新"
            print(f"✅ 扁平化数据{action}成功（其他Sheet已保护）")

            return {
                'success': True,
                'file_created': not file_exists,
                'saved_rows': len(table_data),
                'saved_columns': len(table_data[0]) if table_data else 0,
                'excel_updated': True,
                'sheets_protected': file_exists,  # 只有文件存在时才有其他Sheet需要保护
                'protected_sheets_count': len(all_sheets) - 1 if file_exists else 0
            }

        except Exception as e:
            print(f"❌ 扁平化数据保存失败: {e}")
            import traceback
            traceback.print_exc()
            return {'success': False, 'error': f'扁平化数据保存失败: {str(e)}'}

    def handle_frontend_data_format(self, data: Dict[str, Any], db) -> Dict[str, Any]:
        """处理前端发送的数据格式"""
        print("🎯 处理前端数据格式...")

        pdf_id = data['pdf_id']
        excel_file = data['excel_file']
        sheet_name = data['sheet_name']
        table_type = data['table_type']

        # 🔥 关键：检查前端发送的数据字段
        modifications = data.get('modifications', [])
        current_data = data.get('current_data') or data.get('data')  # 兼容两种字段名
        total_changes = data.get('total_changes', 0)

        print(f"📊 前端数据详情:")
        print(f"  📝 修改记录数: {len(modifications)}")
        print(f"  📦 当前数据行数: {len(current_data) if current_data else 0}")
        print(f"  🔢 总修改数: {total_changes}")

        # 🔥 策略1：优先使用修改记录（最精确）
        if modifications and len(modifications) > 0:
            print("🔧 策略1: 使用修改记录进行精确更新")
            result = self.save_with_modifications(pdf_id, excel_file, sheet_name, modifications, table_type, db)
            if result['success']:
                return result
            else:
                print(f"⚠️ 修改记录保存失败，回退到完整数据: {result['error']}")

        # 🔥 策略2：使用当前数据
        if current_data and len(current_data) > 0:
            print("📊 策略2: 使用当前完整数据")
            return self.save_with_current_data(pdf_id, excel_file, sheet_name, current_data, table_type, db)

        # 🔥 策略3：没有有效数据
        print("❌ 策略3: 没有接收到有效数据")
        return {'success': False, 'error': '没有接收到有效数据'}

    def save_with_modifications(self, pdf_id: str, excel_file: str, sheet_name: str,
                                modifications: List[Dict], table_type: str, db) -> Dict[str, Any]:
        """使用修改记录精确更新"""
        print(f"🔧 使用修改记录更新: {len(modifications)} 个修改")

        try:
            # 1. 获取Excel文件路径
            pdf_id = self.get_correct_pdf_id(pdf_id, db)
            excel_dir = Path(self.MAIN_ROOT) / self.EXCEL_OUTPUT_ROOT / pdf_id
            excel_path = excel_dir / excel_file

            print(f"📁 Excel文件路径: {excel_path}")

            if not excel_path.exists():
                return {'success': False, 'error': f'Excel文件不存在: {excel_path}'}

            # 2. 加载工作簿
            from openpyxl import load_workbook
            workbook = load_workbook(excel_path)

            if sheet_name not in workbook.sheetnames:
                workbook.close()
                return {'success': False, 'error': f'Sheet不存在: {sheet_name}'}

            worksheet = workbook[sheet_name]
            print(f"📊 目标工作表: {worksheet.max_row}行 × {worksheet.max_column}列")

            # 3. 🔥 精确解析前端修改记录格式
            applied_count = 0
            failed_count = 0

            for i, mod in enumerate(modifications):
                try:
                    # 🔥 前端格式：{ row, col, oldValue, newValue, saved, timestamp, tableType }
                    row_idx = mod.get('row')
                    col_idx = mod.get('col')
                    new_value = mod.get('newValue') or mod.get('new_value')  # 兼容两种字段名

                    print(f"  🔍 处理修改 {i + 1}: row={row_idx}, col={col_idx}, value={new_value}")

                    # 验证修改记录
                    if (row_idx is not None and col_idx is not None and
                            new_value is not None and
                            isinstance(row_idx, (int, float)) and
                            isinstance(col_idx, (int, float))):

                        # 转换为整数（前端从0开始，Excel从1开始）
                        excel_row = int(row_idx) + 1
                        excel_col = int(col_idx) + 1

                        # 检查行列是否在有效范围内
                        if (excel_row >= 1 and excel_row <= worksheet.max_row and
                                excel_col >= 1 and excel_col <= worksheet.max_column):

                            # 应用修改
                            worksheet.cell(row=excel_row, column=excel_col, value=new_value)
                            applied_count += 1
                            print(f"    ✅ 应用修改: [{row_idx},{col_idx}] = '{new_value}'")
                        else:
                            print(f"    ⚠️ 坐标超出范围: [{row_idx},{col_idx}] -> Excel[{excel_row},{excel_col}]")
                            failed_count += 1
                    else:
                        print(f"    ⚠️ 无效修改格式: {mod}")
                        failed_count += 1

                except Exception as e:
                    print(f"    ❌ 修改处理失败: {e}")
                    failed_count += 1
                    continue

            # 4. 保存文件
            workbook.save(excel_path)
            workbook.close()

            print(f"✅ 修改应用完成: {applied_count}成功, {failed_count}失败")

            return {
                'success': True,
                'strategy_used': 'modifications',
                'saved_count': applied_count,
                'failed_count': failed_count,
                'excel_updated': applied_count > 0,
                'data_dimensions': f'{applied_count}处修改'
            }

        except Exception as e:
            print(f"❌ 修改记录保存失败: {e}")
            return {'success': False, 'error': f'修改记录保存失败: {str(e)}'}

    def save_with_current_data(self, pdf_id: str, excel_file: str, sheet_name: str,
                               current_data: List, table_type: str, db) -> Dict[str, Any]:
        """使用当前完整数据覆盖"""
        print("📊 使用当前完整数据覆盖")
        print(f"🔍 接收数据格式: {type(current_data)}, 长度: {len(current_data)}")

        try:
            # 1. 获取Excel文件路径
            pdf_id = self.get_correct_pdf_id(pdf_id, db)
            excel_dir = Path(self.MAIN_ROOT) / self.EXCEL_OUTPUT_ROOT / pdf_id
            excel_path = excel_dir / excel_file

            print(f"📁 Excel文件路径: {excel_path}")

            if not excel_path.exists():
                return {'success': False, 'error': f'Excel文件不存在: {excel_path}'}

            # 2. 🔥 转换前端数据格式为二维数组
            print("🔄 转换前端数据格式...")
            backend_data = self.convert_frontend_to_backend_format(current_data)

            if not backend_data or len(backend_data) == 0:
                return {'success': False, 'error': '转换后数据为空'}

            print(f"✅ 转换后数据: {len(backend_data)}行 × {len(backend_data[0]) if backend_data else 0}列")

            # 3. 加载工作簿
            from openpyxl import load_workbook
            workbook = load_workbook(excel_path)

            if sheet_name not in workbook.sheetnames:
                workbook.close()
                return {'success': False, 'error': f'Sheet不存在: {sheet_name}'}

            worksheet = workbook[sheet_name]
            expected_columns = worksheet.max_column

            print(f"📈 Excel工作表列数: {expected_columns}")
            print(f"📊 转换后数据列数: {len(backend_data[0])}")

            # 4. 修复列数不匹配
            if len(backend_data[0]) != expected_columns:
                print(f"⚠️ 列数不匹配! 数据{len(backend_data[0])}列, Excel{expected_columns}列")
                backend_data = self.fix_column_mismatch(backend_data, expected_columns)
                print(f"✅ 修复后数据: {len(backend_data)}行 × {len(backend_data[0])}列")

            # 5. 清空数据行（保留表头）
            if worksheet.max_row > 1:
                rows_to_delete = worksheet.max_row - 1
                worksheet.delete_rows(2, rows_to_delete)
                print(f"🗑️ 清空数据行: 删除了{rows_to_delete}行")

            # 6. 写入新数据
            if backend_data and len(backend_data) > 0:
                for row_idx, row_data in enumerate(backend_data, 2):  # 从第2行开始（保留表头）
                    for col_idx, cell_value in enumerate(row_data, 1):
                        if col_idx <= worksheet.max_column:
                            worksheet.cell(row=row_idx, column=col_idx, value=cell_value)

                print(f"📝 写入数据: {len(backend_data)}行 × {len(backend_data[0])}列")

            # 7. 保存文件
            workbook.save(excel_path)
            workbook.close()

            print("✅ 完整数据覆盖完成")

            return {
                'success': True,
                'strategy_used': 'current_data',
                'saved_count': len(backend_data),
                'excel_updated': True,
                'data_dimensions': f'{len(backend_data)}行 × {len(backend_data[0])}列'
            }

        except Exception as e:
            print(f"❌ 完整数据保存失败: {e}")
            return {'success': False, 'error': f'完整数据保存失败: {str(e)}'}

    def convert_frontend_to_backend_format(self, frontend_data: List) -> List[List]:
        """🔥 精确转换前端数据格式为二维数组"""
        if not frontend_data:
            print("⚠️ 前端数据为空")
            return []

        print("🔄 转换前端数据格式...")
        print(f"📊 原始数据: {len(frontend_data)}行")

        backend_data = []

        for i, row in enumerate(frontend_data):
            # 🔥 处理前端可能的多种格式
            if isinstance(row, list):
                # 已经是数组格式，直接使用
                backend_data.append(row)
                if i < 2:  # 只打印前2行样本
                    print(f"  ✅ 行{i}: 数组格式 ({len(row)}列)")

            elif isinstance(row, dict):
                # 🔥 处理对象格式：{ H_1: '值1', H_2: '值2', ... }
                if any(key.startswith('H_') for key in row.keys()):
                    # 提取 H_1, H_2, H_3, ... 字段
                    row_values = []
                    col_idx = 1

                    while f'H_{col_idx}' in row:
                        value = row[f'H_{col_idx}']
                        row_values.append(value)
                        col_idx += 1

                    if row_values:
                        backend_data.append(row_values)
                        if i < 2:
                            print(f"  ✅ 行{i}: H_*对象格式 ({len(row_values)}列)")
                    else:
                        print(f"  ⚠️ 行{i}: H_*对象但无有效值")

                # 🔥 跳过元数据行
                elif row.get('__metadata') or row.get('__is_first_row'):
                    if i < 2:
                        print(f"  ⏭️ 行{i}: 跳过元数据行")
                    continue

                else:
                    # 其他对象格式，尝试提取所有值（跳过内部字段）
                    row_values = []
                    for key, value in row.items():
                        if not key.startswith('__'):  # 跳过内部字段
                            row_values.append(value)

                    if row_values:
                        backend_data.append(row_values)
                        if i < 2:
                            print(f"  ✅ 行{i}: 普通对象格式 ({len(row_values)}列)")
                    else:
                        print(f"  ⚠️ 行{i}: 无法处理的对象格式")

            else:
                # 其他格式（字符串、数字等），包装成数组
                backend_data.append([row])
                if i < 2:
                    print(f"  ✅ 行{i}: 简单值转数组")

        print(f"📈 转换完成: {len(backend_data)}行有效数据")

        if backend_data and len(backend_data) > 0:
            print(f"📏 数据维度: {len(backend_data)}行 × {len(backend_data[0])}列")
            # 显示样本数据
            for i, row in enumerate(backend_data[:2]):
                sample = row[:3] if len(row) > 3 else row
                print(f"    行{i}样本: {sample}")

        return backend_data

    def fix_column_mismatch(self, data: List[List], expected_columns: int) -> List[List]:
        """修复列数不匹配"""
        print(f"🔧 修复列数不匹配: {len(data[0])}列 -> {expected_columns}列")

        fixed_data = []
        for i, row in enumerate(data):
            if len(row) < expected_columns:
                # 补全缺失列
                fixed_row = row + [''] * (expected_columns - len(row))
                if i < 2:
                    print(f"  ✅ 行{i}: 补全{expected_columns - len(row)}列")
            elif len(row) > expected_columns:
                # 截断多余列
                fixed_row = row[:expected_columns]
                if i < 2:
                    print(f"  ✅ 行{i}: 截断{len(row) - expected_columns}列")
            else:
                fixed_row = row
            fixed_data.append(fixed_row)

        return fixed_data


    def save_to_flattened_excel(self, pdf_id: str, original_excel_file: str, flattened_excel_file: str,
                               sheet_name: str, table_data: List, db) -> Dict[str, Any]:
        """保存数据到独立的扁平化Excel文件"""
        print("📊 保存到独立扁平化Excel文件...")

        try:
            from pathlib import Path
            import os

            # 使用相同的目录结构，通过文件名区分
            pdf_id = self.get_correct_pdf_id(pdf_id, db)
            excel_dir = Path(self.MAIN_ROOT) / self.EXCEL_OUTPUT_ROOT / pdf_id
            excel_dir.mkdir(parents=True, exist_ok=True)

            flattened_excel_path = excel_dir / flattened_excel_file
            print(f"📁 扁平化文件路径: {flattened_excel_path}")

            from openpyxl import Workbook, load_workbook

            file_exists = flattened_excel_path.exists()
            file_created = not file_exists
            sheet_created = False

            if file_exists:
                # 文件存在，加载现有工作簿
                workbook = load_workbook(flattened_excel_path)
                existing_sheets = workbook.sheetnames.copy()
                print(f"📄 加载现有扁平化文件，包含 {len(existing_sheets)} 个Sheet: {existing_sheets}")
            else:
                # 文件不存在，创建新工作簿
                workbook = Workbook()
                # 删除默认Sheet
                default_sheet = workbook.active
                workbook.remove(default_sheet)
                print("📄 创建新的扁平化Excel文件")

            # 处理目标Sheet
            if sheet_name in workbook.sheetnames:
                print(f"📋 Sheet已存在，覆盖: {sheet_name}")
                # 删除现有Sheet
                del workbook[sheet_name]
                sheet_created = False
            else:
                sheet_created = True

            # 创建/重写目标Sheet
            worksheet = workbook.create_sheet(sheet_name)
            print(f"✅ 创建/覆盖Sheet: {sheet_name}")

            # 转换前端数据格式
            backend_data = self.convert_frontend_to_backend_format(table_data)
            if not backend_data or len(backend_data) == 0:
                workbook.close()
                return {'success': False, 'error': '转换后数据为空'}

            print(f"📊 转换后数据: {len(backend_data)}行 × {len(backend_data[0])}列")

            # 写入完整数据到目标Sheet
            for row_idx, row_data in enumerate(backend_data, 1):
                for col_idx, cell_value in enumerate(row_data, 1):
                    worksheet.cell(row=row_idx, column=col_idx, value=cell_value)

            print(f"📝 写入数据: {len(backend_data)}行 × {len(backend_data[0])}列")

            # 保存工作簿
            workbook.save(flattened_excel_path)
            workbook.close()

            # 验证原文件是否存在（可选，用于调试）
            original_excel_path = excel_dir / original_excel_file
            print(f"🔍 验证原文件存在: {original_excel_path.exists()}")

            print("✅ 扁平化文件保存成功")

            return {
                'success': True,
                'file_created': file_created,
                'sheet_created': sheet_created,
                'saved_rows': len(backend_data),
                'saved_columns': len(backend_data[0]) if backend_data else 0,
                'data_dimensions': f'{len(backend_data)}行 × {len(backend_data[0])}列',
                'flattened_path': str(flattened_excel_path),
                'original_path': str(original_excel_path)
            }

        except Exception as e:
            print(f"❌ 扁平化文件保存失败: {e}")
            import traceback
            traceback.print_exc()
            return {'success': False, 'error': f'扁平化文件保存失败: {str(e)}'}

    def generate_flattened_filename(self, original_excel_file: str) -> str:
        """通过固定前缀生成扁平化Excel文件名"""
        from pathlib import Path

        original_path = Path(original_excel_file)

        # 使用固定前缀 "flattened_" 来建立映射关系
        flattened_name = f"flattened_{original_path.name}"

        print(f"📝 文件名映射: {original_excel_file} -> {flattened_name}")
        return flattened_name

    def get_flattened_filename(self, original_excel_file: str) -> str:
        """根据原文件名获取对应的扁平化文件名"""
        from pathlib import Path

        original_path = Path(original_excel_file)
        return f"flattened_{original_path.name}"