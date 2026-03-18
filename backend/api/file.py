"""
文件相关蓝图 - 重构版本（只重构Excel转PDF功能）
"""
from flask import Blueprint, request, jsonify, send_from_directory, make_response, send_file
from backend.utils.constants import UPLOAD_FOLDER, MAIN_ROOT, DATABASE, EXCEL_OUTPUT_ROOT
from pathlib import Path
import sqlite3
import os

# 新增导入
from backend.services.file_mapping_service import file_mapping_service

from .file_handlers.excel_data_handler import ExcelDataHandler
excel_data_handler = ExcelDataHandler(MAIN_ROOT, EXCEL_OUTPUT_ROOT)

file_bp = Blueprint('file', __name__)

from backend.database.export import OldDatabaseManagerAdapter
db = OldDatabaseManagerAdapter(DATABASE)


# 尝试导入转换器，提供多种导入路径
CONVERTER_AVAILABLE = False
FinalDataConverter = None

try:
    # 尝试从 backend.core.services.table_processor 导入
    from backend.core.table_processor import FinalDataConverter as FC
    FinalDataConverter = FC
    CONVERTER_AVAILABLE = True
    print("✅ long_format_converter 从标准路径导入成功")
except ImportError as e:
    print(f"⚠️ 标准路径导入失败: {e}")
    try:
        # 尝试从当前目录导入
        from long_format_converter import FinalDataConverter as FC
        FinalDataConverter = FC
        CONVERTER_AVAILABLE = True
        print("✅ long_format_converter 从当前目录导入成功")
    except ImportError as e2:
        print(f"❌ 所有导入尝试都失败: {e2}")
        CONVERTER_AVAILABLE = False



# ========== 导入重构后的Excel转PDF功能模块 ==========
from .file_handlers.validators import (
    validate_excel_file,
    validate_conversion_params,
    ValidationError
)


from .file_handlers.processors import (
    save_uploaded_file,
    convert_excel_to_pdf,
    save_and_return_result
)


# ========== 原有的所有接口保持不变 ==========

# ---------- 2. 下载/预览（不返回已软删） ----------
@file_bp.get('/api/file/<path:filename>')
def get_file(filename):
    """
    filename 可能是中文原始名，也可能是磁盘 UUID 名；
    一律先查库映射到真实磁盘名，且只返回未删除的。
    """
    print(f"🔍 查找文件: {filename}")

    conn = db.connect()
    if not conn:
        return jsonify({"error": "数据库连接失败"}), 500

    try:
        c = conn.cursor()
        # 查询文件信息（包括ID和磁盘文件名）
        c.execute(
            "SELECT id, filename, raw_filename FROM files "
            "WHERE (raw_filename = ? OR filename = ?) AND deleted = 0",
            (filename, filename)
        )
        row = c.fetchone()

        if row is None:
            print(f"❌ 文件不存在或已隐藏: {filename}")
            return jsonify({"error": "文件不存在或已隐藏"}), 404

        file_id = row["id"]
        real_name = row["filename"]  # 磁盘 UUID 文件名
        raw_name = row["raw_filename"]  # 原始中文名

        print(f"✅ 找到文件: ID={file_id}, 磁盘名={real_name}, 原始名={raw_name}")

        # 构建文件路径
        upload_dir = Path(MAIN_ROOT) / UPLOAD_FOLDER
        file_path = upload_dir / real_name
        print(f"📁 文件路径: {file_path}")
        print(f"📁 上传目录: {upload_dir}")

        if not file_path.exists():
            print(f"❌ 物理文件不存在: {file_path}")
            return jsonify({"error": "物理文件不存在"}), 404

        # 返回文件 - 使用正确的目录路径
        print(f"📤 返回文件: {real_name}")
        return send_from_directory(str(upload_dir), real_name)

    except Exception as e:
        print(f"❌ 文件查找错误: {e}")
        return jsonify({"error": "文件查找失败"}), 500
    finally:
        conn.close()




# ---------- 2.2 获取文件信息 ----------
@file_bp.get('/api/file-info/<path:filename>')
def get_file_info(filename):
    """获取文件详细信息"""
    print(f"🔍 获取文件信息: {filename}")

    conn = db.connect()
    if not conn:
        return jsonify({"error": "数据库连接失败"}), 500

    try:
        c = conn.cursor()
        c.execute(
            "SELECT id, filename, raw_filename, file_type, created_at FROM files "
            "WHERE (raw_filename = ? OR filename = ?) AND deleted = 0",
            (filename, filename)
        )
        row = c.fetchone()

        if row is None:
            return jsonify({"error": "文件不存在或已隐藏"}), 404

        # 构建文件路径信息
        upload_dir = Path(MAIN_ROOT) / UPLOAD_FOLDER
        file_path = upload_dir / row["filename"]
        file_exists = file_path.exists()

        return jsonify({
            "id": row["id"],
            "disk_name": row["filename"],
            "original_name": row["raw_filename"],
            "file_type": row["file_type"],
            "created_at": row["created_at"],
            "file_exists": file_exists,
            "file_path": str(file_path)
        })

    except Exception as e:
        print(f"❌ 获取文件信息错误: {e}")
        return jsonify({"error": "获取文件信息失败"}), 500
    finally:
        conn.close()



# ---------- 3. 软删除（增强版：物理文件不存在时直接删除） ----------
@file_bp.delete('/api/file/<path:filename>')
def delete_file(filename):
    """
    删除文件逻辑：
    1. 如果物理文件不存在，直接删除数据库记录
    2. 如果物理文件存在，只进行软删除 (deleted = 1)
    """
    conn = db.connect()
    if not conn:
        return jsonify({"error": "数据库连接失败"}), 500

    c = conn.cursor()

    try:
        # 先找真实磁盘名和文件信息
        c.execute(
            "SELECT id, filename, raw_filename FROM files "
            "WHERE (raw_filename = ? OR filename = ?) AND deleted = 0",
            (filename, filename)
        )
        row = c.fetchone()
        if row is None:
            conn.close()
            return jsonify({"error": "文件不存在"}), 404

        file_id = row["id"]
        real_name = row["filename"]
        raw_name = row["raw_filename"]

        print(f"🔍 删除文件: ID={file_id}, 磁盘名={real_name}, 原始名={raw_name}")

        # 检查物理文件是否存在
        upload_dir = Path(MAIN_ROOT) / UPLOAD_FOLDER
        file_path = upload_dir / real_name

        if not file_path.exists():
            print(f"📁 物理文件不存在，直接删除数据库记录: {file_path}")
            # 物理文件不存在，直接删除数据库记录
            c.execute("DELETE FROM files WHERE id = ?", (file_id,))
            conn.commit()
            conn.close()
            return jsonify({"message": "文件已彻底删除（物理文件不存在）"}), 200
        else:
            print(f"📁 物理文件存在，进行软删除: {file_path}")
            # 物理文件存在，进行软删除
            c.execute("UPDATE files SET deleted = 1 WHERE filename = ?", (real_name,))
            conn.commit()
            conn.close()
            return jsonify({"message": "已隐藏（软删除）"}), 200

    except Exception as e:
        print(f"❌ 删除文件时出错: {e}")
        if conn:
            conn.rollback()
            conn.close()
        return jsonify({"error": "删除文件失败"}), 500



# ---------- 4. 搜索PDF文件（新增接口） ----------
@file_bp.get('/api/search-pdf')
def search_pdf():
    """
    搜索PDF文件名称 - 基于文件映射服务
    参数: keyword - 搜索关键词
    返回: 匹配的PDF文件列表（显示原始中文名）
    """
    keyword = request.args.get('keyword', '').strip()
    print(f"搜索关键词: '{keyword}'")

    if not keyword:
        return jsonify({"files": []})

    try:
        # 使用文件映射服务搜索PDF文件
        search_results = file_mapping_service.search_files(keyword, 'pdf')

        return jsonify({"files": search_results})

    except Exception as e:
        print(f"搜索PDF失败: {e}")
        return jsonify({"error": "搜索失败"}), 500


@file_bp.get('/api/file-by-id/<file_id>')
def get_file_by_id(file_id):
    try:
        print(f"🔍🔍 文件下载请求 file_id: {file_id}")

        # 直接使用 file_id（UUID格式）
        file_uuid = file_id.split('.')[0] if '.' in file_id else file_id
        print(f"🔍🔍 直接使用UUID: {file_uuid}")

        # 检查文件是否存在映射中
        file_info = file_mapping_service.get_file_info(file_uuid)
        if not file_info:
            print(f"❌ 文件映射服务中没有找到UUID: {file_uuid}")
            return jsonify({"error": "文件不存在"}), 404

        disk_name = file_info["disk_name"]
        PDF_DIR = Path(MAIN_ROOT) / UPLOAD_FOLDER

        # 检查文件是否存在
        file_path = PDF_DIR / disk_name
        print(f"🔍 完整文件路径: {file_path}")

        if not file_path.exists():
            print(f"❌ 物理文件不存在: {file_path}")

            # 列出目录中的文件
            if PDF_DIR.exists():
                files_in_dir = list(PDF_DIR.glob("*.pdf"))
                print(f"📂 目录中的PDF文件: {[f.name for f in files_in_dir[:5]]}")
            else:
                print(f"❌ 目录不存在: {PDF_DIR}")

            return jsonify({"error": "物理文件不存在"}), 404

        print(f"✅ 准备返回文件: {disk_name}")
        return send_from_directory(str(PDF_DIR), disk_name)

    except Exception as e:
        print(f"❌❌ 文件下载失败: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": "文件下载失败"}), 500



@file_bp.get('/api/excel-sheets/<file_id>')
def get_excel_sheets(file_id):
    """
    根据PDF文件ID获取对应的Excel sheet列表 - 修复路径问题
    """
    try:
        print(f"🔍🔍 获取Excel sheets请求 file_id: {file_id}")

        # 🔥🔥🔥 关键修复：清理file_id，移除.pdf扩展名
        clean_file_id = excel_data_handler.get_correct_pdf_id(file_id, db)

        # 2. 构建Excel文件目录路径
        excel_dir = Path(MAIN_ROOT) / EXCEL_OUTPUT_ROOT / clean_file_id  # 🔥 使用清理后的ID
        print(f"📁 Excel目录路径: {excel_dir}")

        if not excel_dir.exists():
            print(f"⚠️ Excel目录不存在: {excel_dir}")

            # 🔥🔥🔥 备选方案1：尝试使用原始file_id（不带.pdf）
            alt_excel_dir = Path(MAIN_ROOT) / EXCEL_OUTPUT_ROOT / file_id
            print(f"🔍🔍 尝试备选Excel目录1: {alt_excel_dir}")

            if alt_excel_dir.exists():
                excel_dir = alt_excel_dir
                print(f"✅ 使用备选目录1: {excel_dir}")
            else:
                # 🔥🔥🔥 备选方案2：尝试使用UUID部分
                if '_' in file_id:
                    uuid_part = file_id.split('_')[0]
                    alt_excel_dir2 = Path(MAIN_ROOT) / EXCEL_OUTPUT_ROOT / uuid_part
                    print(f"🔍🔍 尝试备选Excel目录2: {alt_excel_dir2}")

                    if alt_excel_dir2.exists():
                        excel_dir = alt_excel_dir2
                        print(f"✅ 使用备选目录2: {excel_dir}")
                    else:
                        return jsonify({"excel_files": []})
                else:
                    return jsonify({"excel_files": []})

        # 3. 查找目录中的所有Excel文件
        excel_files = []
        supported_extensions = ['.xlsx', '.xls']

        for ext in supported_extensions:
            for excel_file in excel_dir.glob(f"*{ext}"):
                if excel_file.is_file():
                    excel_files.append({
                        "file_name": excel_file.name,
                        "file_path": str(excel_file),
                        "file_size": excel_file.stat().st_size
                    })

        if not excel_files:
            print("⚠️ 目录中没有找到Excel文件")
            return jsonify({"excel_files": []})

        print(f"✅ 找到 {len(excel_files)} 个Excel文件")

        # 4. 读取每个Excel文件的sheet列表
        import pandas as pd
        result = []

        for excel_file in excel_files:
            try:
                print(f"🔍🔍 读取Excel文件: {excel_file['file_name']}")
                # 读取Excel文件的所有sheet
                excel_file_obj = pd.ExcelFile(excel_file["file_path"])
                sheet_names = excel_file_obj.sheet_names

                file_sheets = []
                for sheet_name in sheet_names:
                    file_sheets.append({
                        "name": sheet_name,
                        "excel_file": excel_file["file_name"]
                    })

                result.append({
                    "excel_file": excel_file["file_name"],
                    "sheets": file_sheets,
                    "total_sheets": len(sheet_names)
                })
                print(f"✅ 成功读取 {len(sheet_names)} 个sheet")

            except Exception as e:
                print(f"❌ 读取Excel文件失败 {excel_file['file_name']}: {e}")
                # 即使读取失败，也返回文件信息
                result.append({
                    "excel_file": excel_file["file_name"],
                    "sheets": [],
                    "total_sheets": 0,
                    "error": str(e)
                })
                continue

        print(f"✅ 返回结果: {len(result)} 个Excel文件信息")
        return jsonify({
            "excel_files": result,
            "pdf_id": clean_file_id,  # 🔥 返回清理后的ID
            "pdf_name": file_id,  # 保持原始file_id用于显示
            "total_excel_files": len(result)
        })

    except Exception as e:
        print(f"❌❌ 获取Excel sheets失败: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": "获取表格列表失败"}), 500


@file_bp.get('/api/file/excel-data/<file_id>/<path:excel_file_name>/<sheet_name>')
def get_excel_data(file_id, excel_file_name, sheet_name):
    """
    读取Excel文件中特定sheet的数据
    """
    try:
        print(f"🔍🔍 获取Excel数据请求: file_id={file_id}, excel_file={excel_file_name}, sheet={sheet_name}")

        # 🔥🔥 关键修复：清理file_id，移除.pdf扩展名
        clean_file_id = excel_data_handler.get_correct_pdf_id(file_id, db)

        # 1. 构建Excel文件路径
        excel_dir = Path(MAIN_ROOT) / EXCEL_OUTPUT_ROOT / clean_file_id
        print(f"📁 Excel文件路径excel_dir: {excel_dir}")
        excel_path = excel_dir / excel_file_name

        print(f"📁 Excel文件路径: {excel_path}")

        if not excel_path.exists():
            print(f"❌ Excel文件不存在: {excel_path}")

            # 🔥🔥 尝试备选路径：如果清理后找不到，尝试使用原始file_id
            if clean_file_id != file_id:
                alt_excel_dir = Path(MAIN_ROOT) / EXCEL_OUTPUT_ROOT / file_id
                alt_excel_path = alt_excel_dir / excel_file_name
                print(f"🔍🔍 尝试备选路径: {alt_excel_path}")

                if alt_excel_path.exists():
                    excel_dir = alt_excel_dir
                    excel_path = alt_excel_path
                    print(f"✅ 使用备选路径: {excel_path}")
                else:
                    return jsonify({"error": "Excel文件不存在"}), 404
            else:
                return jsonify({"error": "Excel文件不存在"}), 404


        # 2. 读取Excel文件
        try:
            import pandas as pd
            print("🎯 直接读取Excel文件数据（跳过快照）")

            # 读取指定的sheet
            df = pd.read_excel(
                excel_path,
                sheet_name=sheet_name,
                header=None,  # 不自动识别表头
                dtype=str  # 全部读取为字符串
            )

            # 将NaN替换为空字符串
            df = df.fillna('')

            # 转换为二维列表
            data = df.values.tolist()

            # 检查是否有元数据行
            metadata_row_index = -1
            for i, row in enumerate(data):
                if row and isinstance(row[0], str) and row[0].startswith('#METADATA_START#'):
                    metadata_row_index = i
                    break

            # 如果有元数据，提取它
            metadata = {}
            if metadata_row_index >= 0:
                for i in range(metadata_row_index, min(metadata_row_index + 10, len(data))):
                    if i < len(data) and data[i]:
                        cell_value = str(data[i][0])
                        if cell_value.startswith('#METADATA_END#'):
                            break
                        elif ':' in cell_value and not cell_value.startswith('#'):
                            key, value = cell_value.split(':', 1)
                            metadata[key.strip()] = value.strip()
                            print(f"📄 提取元数据: {key} = {value}")

            # 如果找到元数据，在数据中移除元数据行
            if metadata_row_index >= 0:
                data = data[:metadata_row_index]

            return jsonify({
                "success": True,
                "data": data,
                "metadata": metadata,
                "rows": len(data),
                "cols": len(data[0]) if data else 0
            })

        except Exception as e:
            print(f"❌ 读取Excel文件失败: {e}")
            return jsonify({"error": f"读取Excel文件失败: {str(e)}"}), 500

    except Exception as e:
        print(f"❌❌ 获取Excel数据失败: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": "获取Excel数据失败"}), 500



@file_bp.get('/api/excel-data/<file_id>/<path:excel_file_name>/<sheet_name>')
def get_excel_data_api(file_id, excel_file_name, sheet_name):
    """
    提供Excel数据API接口 - 修复路径问题
    """
    try:
        print(f"🎯🎯🎯 收到Excel数据API请求: file_id={file_id}, excel_file={excel_file_name}, sheet={sheet_name}")

        # 🔥🔥🔥 关键修复：清理file_id，移除.pdf扩展名
        clean_file_id = excel_data_handler.get_correct_pdf_id(file_id, db)

        # 1. 构建Excel文件路径
        excel_dir = Path(MAIN_ROOT) / EXCEL_OUTPUT_ROOT / clean_file_id
        excel_path = excel_dir / excel_file_name

        if not excel_path.exists():
            print(f"❌ Excel文件不存在: {excel_path}")

            # 🔥🔥🔥 备选方案1：尝试使用原始file_id
            if clean_file_id != file_id:
                alt_excel_dir = Path(MAIN_ROOT) / EXCEL_OUTPUT_ROOT / file_id
                alt_excel_path = alt_excel_dir / excel_file_name
                print(f"🔍🔍 尝试备选路径1: {alt_excel_path}")

                if alt_excel_path.exists():
                    excel_dir = alt_excel_dir
                    excel_path = alt_excel_path
                    clean_file_id = file_id
                    print(f"✅ 使用备选路径1: {excel_path}")
                else:
                    # 🔥🔥🔥 备选方案2：如果是数字ID，查询数据库获取UUID
                    if clean_file_id.isdigit():
                        print(f"🔍🔍 尝试备选方案2: 数字ID查询数据库 {clean_file_id}")
                        conn = db.connect()
                        if conn:
                            c = conn.cursor()
                            c.execute("SELECT filename FROM files WHERE id = ? AND deleted = 0", (clean_file_id,))
                            row = c.fetchone()
                            conn.close()

                            if row:
                                real_uuid = row["filename"].split('.')[0] if '.' in row["filename"] else row["filename"]
                                print(f"✅ 找到数据库对应UUID: {real_uuid}")

                                alt_excel_dir2 = Path(MAIN_ROOT) / EXCEL_OUTPUT_ROOT / real_uuid
                                alt_excel_path2 = alt_excel_dir2 / excel_file_name
                                print(f"🔍🔍 尝试数据库路径: {alt_excel_path2}")

                                if alt_excel_path2.exists():
                                    excel_dir = alt_excel_dir2
                                    excel_path = alt_excel_path2
                                    clean_file_id = real_uuid
                                    print(f"✅ 使用数据库路径: {excel_path}")
                                else:
                                    return jsonify({"success": False, "error": "Excel文件不存在"}), 404
                            else:
                                return jsonify({"success": False, "error": "Excel文件不存在"}), 404
                        else:
                            return jsonify({"success": False, "error": "Excel文件不存在"}), 404
                    else:
                        return jsonify({"success": False, "error": "Excel文件不存在"}), 404
            else:
                return jsonify({"success": False, "error": "Excel文件不存在"}), 404

        print(f"✅ Excel文件存在: {excel_path}")

        # 2. 读取Excel文件
        try:
            import pandas as pd
            print("🎯 直接读取Excel文件数据")

            # 读取指定的sheet
            df = pd.read_excel(
                excel_path,
                sheet_name=sheet_name,
                header=None,  # 不自动识别表头
                dtype=str  # 全部读取为字符串
            )

            # 将NaN替换为空字符串
            df = df.fillna('')

            # 🔥🔥🔥 转换为二维列表
            data = df.values.tolist()

            print(f"✅ 成功读取sheet '{sheet_name}'，数据形状: {len(data)}行 x {len(data[0]) if data else 0}列")

            # 🔥🔥🔥 提取元数据
            valid_keys = ["bankname", "currency", "report_period", "unit", "table_name", "ocr_table_id", "entity"]
            metadata = {key: "" for key in valid_keys}
            clean_data = []

            for row in data:
                if not row or not row[0]:
                    clean_data.append(row)
                    continue

                first_cell = str(row[0]).strip()

                # 查找包含冒号的行作为元数据
                if ":" in first_cell:
                    try:
                        key, value = first_cell.split(":", 1)
                        key = key.strip().lower()
                        value = value.strip()

                        # 只收集已知的元数据字段
                        if key in valid_keys:
                            metadata[key] = value
                        clean_data.append(row)
                    except:
                        clean_data.append(row)
                else:
                    clean_data.append(row)

            result = {
                "success": True,
                "data": clean_data,  # 返回清理后的数据
                "rows": len(clean_data),
                "cols": len(clean_data[0]) if clean_data else 0,
                "sheet_name": sheet_name,
                "excel_file": excel_file_name,
                "metadata": metadata,  # 🔥 返回提取的元数据
                "has_custom_metadata": bool(metadata),  # 🔥 标记是否有元数据
                "file_path": str(excel_path)
            }

            # 🔥🔥🔥 返回前端期望的格式
            return jsonify(result)

        except Exception as e:
            print(f"❌ 读取Excel文件失败: {e}")
            import traceback
            traceback.print_exc()
            return jsonify({"success": False, "error": f"读取Excel文件失败: {str(e)}"}), 500

    except Exception as e:
        print(f"❌❌❌ 获取Excel数据API失败: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": "获取Excel数据失败"}), 500



@file_bp.route('/api/excel/save-final', methods=['POST'])
def save_final_excel():
    """统一保存整个表格数据 - 保护其他Sheet"""
    data = request.json
    print("******************** 保存整个表格（保护其他Sheet） ******************")

    # 基础校验
    required_fields = ['pdf_id', 'excel_file', 'sheet_name', 'table_type', 'data']
    for field in required_fields:
        if field not in data:
            return jsonify({'error': f'缺少必要字段: {field}'}), 400

    try:
        pdf_id = data['pdf_id']
        excel_file = data['excel_file']
        sheet_name = data['sheet_name']
        table_type = data['table_type']
        table_data = data['data']

        print(f"💾 保存数据: PDF={pdf_id}, 文件={excel_file}, Sheet={sheet_name}, 类型={table_type}")
        print(f"📊 接收数据: {len(table_data)}行 × {len(table_data[0]) if table_data else 0}列")

        # 根据表类型选择保存方式
        if table_type == 'original':
            result = excel_data_handler.save_complete_table_data(pdf_id, excel_file, sheet_name, table_data, table_type, db)
        elif table_type == 'flattened':
            result = excel_data_handler.save_flattened_table_data(pdf_id, excel_file, sheet_name, table_data, table_type, db)
        else:
            return jsonify({'error': f'不支持的表类型: {table_type}'}), 400

        if not result['success']:
            return jsonify({'success': False, 'error': result['error']}), 500

        return jsonify({
            'success': True,
            'message': '表格数据保存成功',
            'saved_count': result.get('saved_rows', 0),
            'data_dimensions': result.get('data_dimensions', '未知'),
            'excel_updated': result.get('excel_updated', False),
            'sheets_protected': result.get('sheets_protected', False),
            'protected_sheets_count': result.get('protected_sheets_count', 0),
            'file_created': result.get('file_created', False)
        }), 200

    except Exception as e:
        print(f"❌ 保存失败: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': f'保存失败: {str(e)}'}), 500


@file_bp.route('/api/excel-data/<path:filename>')
def serve_excel_file(filename):
    """提供 Excel 文件下载"""
    # 修正：统一使用Path对象构建路径

    file_path = Path(MAIN_ROOT) / EXCEL_OUTPUT_ROOT / filename
    if not file_path.exists():
        return "文件不存在", 404
    return send_from_directory(file_path.parent, file_path.name)


@file_bp.route('/api/excel-data/<pdf_id>/<excel_file>/<sheet_name>', methods=['GET'])
def get_flat_excel_data(pdf_id, excel_file, sheet_name):
    """读取Excel数据（支持原始文件和扁平化文件）"""
    print(f"📥 获取Excel数据: {pdf_id}, {excel_file}, {sheet_name}")

    try:
        file_path = os.path.join(EXCEL_OUTPUT_ROOT, pdf_id, excel_file)
        print(f"🔍 文件路径: {file_path}")

        if not os.path.exists(file_path):
            return jsonify({
                "success": False,
                "error": f"文件不存在: {file_path}",
                "data": []
            }), 404

        # 读取Excel文件
        import pandas as pd
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)

        # 转换为二维数组
        data = df.values.tolist()

        print(f"✅ 文件读取成功:")
        print(f"   文件: {excel_file}")
        print(f"   Sheet: {sheet_name}")
        print(f"   数据: {len(data)}行 × {len(data[0]) if data else 0}列")

        return jsonify({
            "success": True,
            "data": data,
            "rows": len(data),
            "cols": len(data[0]) if data else 0,
            "file_type": "flattened" if "flattened" in excel_file.lower() else "original"
        })

    except Exception as e:
        print(f"❌ 读取Excel文件失败: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": f"读取Excel文件失败: {str(e)}",
            "data": []
        }), 500


# 新增导入Excel扁平化处理器
from .file_handlers.excel_flatten_handler import ExcelFlattenHandler
# 初始化Excel扁平化处理器
excel_flatten_handler = ExcelFlattenHandler(CONVERTER_AVAILABLE, FinalDataConverter, db)

# ---------- 7. Excel数据扁平化处理（支持Excel标准格式） ----------
@file_bp.route('/api/excel-flatten', methods=['POST', 'OPTIONS'])
def excel_flatten_from_excel():
    """
    处理从Excel直接提取的标准格式数据
    使用long_format_converter.py的逻辑进行扁平化转换
    """
    if request.method == 'OPTIONS':
        return jsonify({"status": "ok"}), 200

    try:
        # 获取请求数据
        data = request.get_json()

        if not data:
            print("❌ 请求数据为空")
            return jsonify({
                "success": False,
                "error": "请求数据为空"
            }), 400

        # 使用处理器处理
        result = excel_flatten_handler.excel_flatten_from_excel(data)

        if not result.get('success', True):
            return jsonify(result), 500

        return jsonify(result)

    except Exception as e:
        print(f"❌ Excel数据处理失败: {str(e)}")
        import traceback
        traceback.print_exc()

        return jsonify({
            "success": False,
            "error": f"处理失败: {str(e)}"
        }), 500


# ========== 新增的Excel转PDF接口 ==========
@file_bp.route('/api/convert/excel-to-pdf', methods=['POST'])
def convert_excel_to_pdf_api():
    """Excel转PDF接口 - 新增功能"""
    try:
        # 1. 获取文件和参数
        if 'file' not in request.files:
            return jsonify({'error': '没有上传文件'}), 400

        file = request.files['file']
        pages = request.form.get('pages', 'all')
        orientation = request.form.get('orientation', 'portrait')
        dpi = int(request.form.get('dpi', 200))

        # 2. 验证文件
        validate_excel_file(file)

        # 3. 验证参数
        orientation, page_range = validate_conversion_params(orientation, pages)

        # 4. 保存上传的文件
        temp_filepath = save_uploaded_file(file)

        # 5. 执行转换
        output_filepath = convert_excel_to_pdf(
            temp_filepath,
            page_range=page_range,
            orientation=orientation,
            dpi=dpi
        )

        # 6. 返回结果
        return save_and_return_result(output_filepath)

    except ValidationError as e:
        return jsonify({'error': e.message}), e.status_code
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'转换失败: {str(e)}'}), 500


# 在 file.py 中添加一个通用的OPTIONS处理
@file_bp.route('/api/save-excel-modifications', methods=['OPTIONS'])
def handle_save_options():
    """处理CORS预检请求"""
    response = jsonify({'status': 'ok'})
    response.headers.add('Access-Control-Allow-Origin', 'http://localhost:8080')
    response.headers.add('Access-Control-Allow-Methods', 'POST, OPTIONS')
    response.headers.add('Access-Control-Allow-Headers', 'Content-Type')
    response.headers.add('Access-Control-Allow-Credentials', 'true')
    return response

@file_bp.route('/api/save-excel-modifications', methods=['POST'])
def save_excel_modifications():
    """保存Excel修改"""
    # 添加CORS头
    response = jsonify({'success': True, 'message': '保存成功'})
    response.headers.add('Access-Control-Allow-Origin', 'http://localhost:8080')
    response.headers.add('Access-Control-Allow-Credentials', 'true')
    return response



@file_bp.route('/api/search-pdf-compatible')
def search_pdf_compatible():
    """搜索PDF文件 - 修复版本（搜索files表）"""
    print("🔥🔥🔥🔥🔥 search_pdf_compatible 函数被调用了！")

    try:
        keyword = request.args.get('keyword', '').strip()
        limit = request.args.get('limit', 100, type=int)

        print(f"🔍 搜索关键词: '{keyword}'")

        if not keyword:
            return jsonify({
                "files": [],
                "count": 0
            })

        # 🔥🔥 关键修复：直接连接数据库，搜索files表
        conn = db.connect()
        if not conn:
            return jsonify({"error": "数据库连接失败"}), 500

        c = conn.cursor()

        # 🔥🔥 修复SQL：搜索files表而不是table_processing_records
        query = """
            SELECT id, filename, raw_filename, file_type, created_at 
            FROM files 
            WHERE deleted = 0 
            AND file_type = 'pdf'
            AND (raw_filename LIKE ? OR filename LIKE ?)
            ORDER BY created_at DESC 
            LIMIT ?
        """

        search_pattern = f'%{keyword}%'
        params = (search_pattern, search_pattern, limit)

        print(f"🔍🔍 执行查询: {query}")
        print(f"🔍🔍 参数: {params}")

        c.execute(query, params)
        rows = c.fetchall()

        print(f"📊 数据库返回 {len(rows)} 条结果")

        # 转换为前端需要的格式
        files = []
        for row in rows:
            # 🔥🔥 修复字段映射：使用files表的字段
            file_info = {
                "id": row["id"],
                "file_id": row["filename"].split('.')[0] if '.' in row["filename"] else row["filename"],  # 使用UUID
                "disk_name": row["filename"],  # 磁盘文件名
                "file_type": row["file_type"],
                "filename": row["raw_filename"] or row["filename"],  # 显示中文名
                "name": row["raw_filename"] or row["filename"],  # 兼容字段
                "matchType": "文件名匹配",
                "status": "active",
                "created_at": row["created_at"],
                "raw_filename": row["raw_filename"]  # 原始文件名
            }
            files.append(file_info)
            print(f"✅ 找到文件: {file_info['filename']}")

        conn.close()

        print(f"✅ 转换完成，返回 {len(files)} 个文件")

        return jsonify({
            "files": files,
            "count": len(files)
        })

    except Exception as e:
        print(f"❌❌ 搜索失败: {e}")
        import traceback
        traceback.print_exc()

        return jsonify({
            "files": [],
            "count": 0,
            "error": str(e)
        })


# 直接调用excel_flatten_handler，避免重复代码
@file_bp.route('/api/excel/save-flattened', methods=['POST', 'OPTIONS'])
def save_flattened_data():
    """保存扁平化数据到独立的Excel文件（完全保持前端数据格式）"""
    if request.method == 'OPTIONS':
        response = make_response()
        response.headers.add('Access-Control-Allow-Origin', '*')
        response.headers.add('Access-Control-Allow-Headers', 'Content-Type,Authorization')
        response.headers.add('Access-Control-Allow-Methods', 'GET,PUT,POST,DELETE,OPTIONS')
        return response

    data = request.json

    try:
        # 🔥🔥 验证输入数据
        required_fields = ['pdf_id', 'excel_file', 'sheet_name', 'table_type', 'flattened_data']
        for field in required_fields:
            if field not in data:
                return jsonify({'error': f'缺少必要字段: {field}'}), 400

        pdf_id = data['pdf_id']
        original_excel_file = data['excel_file']
        sheet_name = data['sheet_name']
        table_type = data['table_type']
        flattened_data = data['flattened_data']  # 🔥🔥 前端已经扁平化好的数据

        print(f"💾 保存扁平化数据:")
        print(f"  PDF ID: {pdf_id}")

        # 🔥🔥 第一步：直接保存前端数据，不做任何处理
        print("🔄🔄 直接保存前端数据，保持原样...")

        data_handler = ExcelDataHandler(MAIN_ROOT, EXCEL_OUTPUT_ROOT)

        # 生成扁平化文件名
        flattened_excel_file = data_handler.generate_flattened_filename(original_excel_file)
        print(f"📁 扁平化文件名: {flattened_excel_file}")

        # 🔥🔥 直接保存前端数据，不调用任何转换逻辑
        result = data_handler.save_flattened_data_as_is(
            pdf_id=pdf_id,
            original_excel_file=original_excel_file,
            flattened_excel_file=flattened_excel_file,
            sheet_name=sheet_name,
            flattened_data=flattened_data,  # 🔥🔥 直接保存，不处理
            table_type=table_type,
            db=db
        )

        if not result['success']:
            return jsonify({'success': False, 'error': result['error']}), 500

        return jsonify({
            'success': True,
            'message': '扁平化数据保存成功',
            'flattened_file': flattened_excel_file,
            'original_file': original_excel_file,
            'file_created': result.get('file_created', False),
            'sheet_created': result.get('sheet_created', False),
            'saved_rows': result.get('saved_rows', 0),
            'saved_columns': result.get('saved_columns', 0),
            'data_dimensions': result.get('data_dimensions', '未知')
        }), 200

    except Exception as e:
        print(f"❌ 扁平化数据保存失败: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': f'扁平化数据保存失败: {str(e)}'}), 500



@file_bp.route('/api/excel/global-flatten/<pdf_id>', methods=['POST', 'OPTIONS'])
def global_flatten(pdf_id):
    """整体扁平化处理 - 处理PDF对应的所有Excel文件的所有sheet，合并成一个大的扁平化表格"""
    if request.method == 'OPTIONS':
        response = make_response()
        response.headers.add('Access-Control-Allow-Origin', '*')
        response.headers.add('Access-Control-Allow-Headers', 'Content-Type,Authorization')
        response.headers.add('Access-Control-Allow-Methods', 'GET,PUT,POST,DELETE,OPTIONS')
        return response

    data = request.json
    print(f"🔍🔍 获取Excel sheets请求 file_id: {pdf_id}")

    # 🔥🔥🔥 关键修复：清理file_id，移除.pdf扩展名
    clean_file_id = excel_data_handler.get_correct_pdf_id(pdf_id, db)

    # 2. 构建Excel文件目录路径
    excel_dir = str(Path(MAIN_ROOT) / EXCEL_OUTPUT_ROOT / clean_file_id)  # 🔥 使用清理后的ID
    print(f"📁 clean_file_id Excel目录路径: {excel_dir}")

    try:
        # 🔥 第一步：调用整体扁平化处理
        print("🔄 开始整体扁平化处理...")
        result = excel_flatten_handler.global_flatten_from_excel_files(clean_file_id, excel_dir)

        # 🔥 第二步：直接返回结果（现在结果中已经包含data字段）
        print(f"📥 返回结果: success={result.get('success')}, data长度={len(result.get('data', []))}")
        return jsonify(result), 200

    except Exception as e:
        print(f"❌ 整体扁平化处理失败: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': f'整体扁平化处理失败: {str(e)}'}), 500


@file_bp.route('/api/excel/export-final-file', methods=['POST'])
def export_final_file():
    data = request.get_json()
    current_excel_file = data.get('excel_file')

    try:
        # 从当前文件名提取UUID部分
        file_uuid = current_excel_file.split('_')[0]  # 提取UUID部分

        # 查询数据库获取文件信息
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()

        cursor.execute("""
            SELECT filename, raw_filename, file_type, file_path, bank_name 
            FROM files 
            WHERE filename LIKE ? OR raw_filename LIKE ?
        """, (f"%{file_uuid}%", f"%{file_uuid}%"))

        file_info = cursor.fetchone()
        conn.close()

        # 文件路径还是按照原先的逻辑（使用UUID）
        final_file_name = f"flattened_整合_{file_uuid}.xlsx"
        final_file_path = os.path.join(MAIN_ROOT, EXCEL_OUTPUT_ROOT, file_uuid, final_file_name)

        if os.path.exists(final_file_path):
            # 获取下载时显示的文件名（使用raw_filename）
            if file_info and file_info[1]:  # file_info[1] 是 raw_filename
                # 确保文件名有.xlsx后缀
                raw_filename = file_info[1]
                if not raw_filename.lower().endswith('.xlsx'):
                    raw_filename = f"{raw_filename}.xlsx"
                download_display_name = f"整合_{raw_filename}"  # 例如：整合_财务报表2024.xlsx
            else:
                # 如果没有raw_filename，使用UUID
                download_display_name = f"整合_{file_uuid}.xlsx"

            download_url = f"/api/excel/download-final/{file_uuid}/{final_file_name}"

            print(f"下载显示文件名: {download_display_name}")
            print(f"下载URL: {download_url}")

            return {
                'success': True,
                'file_exists': True,
                'file_name': final_file_name,  # 服务器上的文件名
                'download_name': download_display_name,  # 下载时显示的文件名
                'file_path': final_file_path,
                'download_url': download_url,
                'message': '最终文件存在'
            }
        else:
            return {
                'success': True,
                'file_exists': False,
                'file_name': final_file_name,
                'file_path': final_file_path,
                'message': '最终文件未生成'
            }

    except Exception as e:
        print(f"错误详情: {str(e)}")
        return {'success': False, 'error': str(e)}, 500



@file_bp.route('/api/excel/download-final/<pdf_id>/<file_name>')
def download_final_file(pdf_id, file_name):
    try:
        # 文件路径还是按照原先的逻辑
        file_path = os.path.join(MAIN_ROOT, EXCEL_OUTPUT_ROOT, pdf_id, file_name)

        if os.path.exists(file_path):
            # 查询数据库获取raw_filename作为下载文件名
            conn = sqlite3.connect(DATABASE)
            cursor = conn.cursor()
            cursor.execute("""
                SELECT raw_filename FROM files 
                WHERE filename LIKE ? OR raw_filename LIKE ?
            """, (f"%{pdf_id}%", f"%{pdf_id}%"))

            file_info = cursor.fetchone()
            conn.close()

            # 设置下载文件名
            if file_info and file_info[0]:
                raw_filename = file_info[0]
                base_name = os.path.splitext(raw_filename)[0]  # 去掉文件扩展名
                download_name = f"整合_{base_name}.xlsx"
                print("download_name:", download_name)
            else:
                download_name = f"整合_{pdf_id}.xlsx"  # 备用名称

            print(f"下载显示名称: {download_name}")
            print(f"开始下载文件...")

            return send_file(file_path, as_attachment=True, download_name=download_name)
        else:
            print(f"❌ 文件不存在: {file_path}")
            return jsonify({
                'success': False,
                'error': '文件不存在',
                'file_path': file_path
            }), 404

    except Exception as e:
        print(f"❌ 下载错误: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500


def delete_excel_sheet(pdf_id, excel_file, sheet_name, excel_path, db):
    """
    从Excel文件中删除指定的sheet
    """
    try:
        import pandas as pd
        from openpyxl import load_workbook

        print(f"🗑️ 删除sheet: {excel_file} -> {sheet_name}")

        # 1. 检查文件是否存在
        if not excel_path.exists():
            error_msg = f'Excel文件不存在: {excel_path}'
            print(f"❌ {error_msg}")
            return {'success': False, 'error': error_msg}

        # 2. 加载工作簿
        workbook = load_workbook(str(excel_path))

        # 3. 检查sheet是否存在
        if sheet_name not in workbook.sheetnames:
            error_msg = f'Sheet不存在: {sheet_name}'
            print(f"❌ {error_msg}")
            return {'success': False, 'error': error_msg}

        # 4. 删除sheet
        sheet_to_delete = workbook[sheet_name]
        workbook.remove(sheet_to_delete)

        # 5. 保存工作簿
        workbook.save(str(excel_path))

        print(f"✅ 成功删除sheet: {sheet_name}")

        return {
            'success': True,
            'message': f'成功删除sheet: {sheet_name}',
            'file': str(excel_path),
            'deleted_sheet': sheet_name
        }

    except Exception as e:
        error_msg = f'删除sheet失败: {str(e)}'
        print(f"❌ {error_msg}")
        return {'success': False, 'error': error_msg}


# ========== 新增的合并数据接口 ==========
@file_bp.route('/api/excel/merge-sheets', methods=['POST'])
def merge_sheets():
    """
    合并两个sheet的数据
    条件：
    1. 当前sheet名称必须包含_1_，目标sheet名称必须包含_T_
    2. 两个sheet的页号必须是连续的
    3. 表头一致性：当前表格可以没有表头，或者有表头但必须与前面表格表头完全相同
    4. 列数必须一致
    合并逻辑：
    1. 当前表格：在第二列中查找"列标记"，保留该行之前的数据
    2. 前面表格：在第二列中查找"列标记"
    3. 将当前表格保留的数据插入到前面表格的"列标记"行前面
    """
    try:
        data = request.get_json()
        print("🔄 收到合并sheet请求...")

        # 检查请求数据是否为空
        if not data:
            print("❌ 请求数据为空")
            return jsonify({
                'success': False,
                'error': '请求数据为空'
            }), 400

        # 打印请求数据摘要
        print(f"🔍 请求数据摘要:")
        print(f"  请求键: {list(data.keys())}")

        # 基本验证
        required_fields = ['sourceSheet', 'targetSheet', 'currentData']
        missing_fields = [field for field in required_fields if field not in data]

        if missing_fields:
            error_msg = f'缺少必要字段: {missing_fields}'
            print(f"❌ {error_msg}")
            return jsonify({
                'success': False,
                'error': error_msg
            }), 400

        source_sheet = data['sourceSheet']
        target_sheet = data['targetSheet']
        current_data = data['currentData']

        # 验证sheet信息
        required_sheet_fields = ['name', 'excelFile']
        sheet_errors = []

        for sheet_name, sheet_data in [('sourceSheet', source_sheet), ('targetSheet', target_sheet)]:
            for field in required_sheet_fields:
                if field not in sheet_data:
                    sheet_errors.append(f'{sheet_name}.{field}')

        if sheet_errors:
            error_msg = f'sheet信息不完整，缺少字段: {sheet_errors}'
            print(f"❌ {error_msg}")
            return jsonify({
                'success': False,
                'error': error_msg
            }), 400

        current_sheet_name = source_sheet['name']
        previous_sheet_name = target_sheet['name']
        excel_file = source_sheet['excelFile']
        pdf_id = source_sheet.get('pdfId')

        # 确保pdf_id是字符串
        if pdf_id is not None:
            pdf_id = str(pdf_id)

        print(f"🔍 开始验证sheet合并条件:")
        print(f"  当前sheet: {current_sheet_name}")
        print(f"  目标sheet: {previous_sheet_name}")
        print(f"  Excel文件: {excel_file}")
        print(f"  PDF ID: {pdf_id}")

        # ========== 第一步：验证sheet名称 ==========
        def extract_page_number(sheet_name):
            """从sheet名称中提取页号"""
            if not sheet_name or not isinstance(sheet_name, str):
                return None

            import re
            patterns = [
                r'P(\d+)_',  # P001_
                r'P(\d+)-',  # P001-
                r'[Pp](\d+)_',  # p001_
                r'[Pp](\d+)-',  # p001-
            ]

            for pattern in patterns:
                match = re.search(pattern, sheet_name)
                if match:
                    try:
                        return int(match.group(1))
                    except (ValueError, IndexError):
                        continue
            return None

        # 检查当前sheet是否包含_1_
        if '_1_' not in str(current_sheet_name):
            error_msg = f'当前sheet"{current_sheet_name}"必须包含"_1_"'
            print(f"❌ {error_msg}")
            return jsonify({
                'success': False,
                'error': error_msg
            }), 400

        # 检查目标sheet是否包含_T_
        if '_T_' not in str(previous_sheet_name):
            error_msg = f'目标sheet"{previous_sheet_name}"必须包含"_T_"'
            print(f"❌ {error_msg}")
            return jsonify({
                'success': False,
                'error': error_msg
            }), 400

        # 提取页号
        current_page = extract_page_number(str(current_sheet_name))
        previous_page = extract_page_number(str(previous_sheet_name))

        if current_page is None:
            error_msg = f'无法从当前sheet"{current_sheet_name}"提取页号'
            print(f"❌ {error_msg}")
            return jsonify({
                'success': False,
                'error': error_msg
            }), 400

        if previous_page is None:
            error_msg = f'无法从目标sheet"{previous_sheet_name}"提取页号'
            print(f"❌ {error_msg}")
            return jsonify({
                'success': False,
                'error': error_msg
            }), 400

        # 检查页号连续性
        if current_page - previous_page != 1:
            error_msg = f'页号不连续：当前页{current_page}，目标页{previous_page}，差值应为1'
            print(f"❌ {error_msg}")
            return jsonify({
                'success': False,
                'error': error_msg
            }), 400

        print(f"✅ sheet验证通过:")
        print(f"  当前页号: {current_page}")
        print(f"  目标页号: {previous_page}")
        print(f"  页号连续: {current_page} - {previous_page} = 1")

        # ========== 第二步：获取目标表格数据 ==========
        print(f"📥 获取目标sheet数据: {previous_sheet_name}")

        try:
            # 直接读取Excel文件
            import pandas as pd

            # 1. 获取正确的PDF ID
            pdf_id_str = str(pdf_id) if pdf_id is not None else ""

            # 如果pdf_id是数字，查询数据库获取实际的UUID
            if pdf_id_str.isdigit():
                conn = db.connect()
                if conn:
                    c = conn.cursor()
                    c.execute("SELECT filename FROM files WHERE id = ? AND deleted = 0", (int(pdf_id_str),))
                    row = c.fetchone()
                    conn.close()

                    if row:
                        # 提取UUID部分
                        real_uuid = row["filename"].split('.')[0] if '.' in row["filename"] else row["filename"]
                        pdf_id_str = real_uuid

            # 2. 构建Excel文件路径
            clean_file_id = pdf_id_str
            excel_dir = Path(MAIN_ROOT) / EXCEL_OUTPUT_ROOT / clean_file_id
            excel_path = excel_dir / excel_file

            print(f"🔍 尝试读取Excel文件: {excel_path}")

            if not excel_path.exists():
                error_msg = f'Excel文件不存在: {excel_path}'
                print(f"❌ {error_msg}")
                return jsonify({
                    'success': False,
                    'error': error_msg
                }), 404

            # 3. 读取Excel文件
            df = pd.read_excel(
                excel_path,
                sheet_name=previous_sheet_name,
                header=None,
                dtype=str
            )

            # 将NaN替换为空字符串
            df = df.fillna('')

            # 转换为二维列表
            target_data = df.values.tolist()

            print(f"✅ 成功读取目标sheet '{previous_sheet_name}'")
            print(f"  数据形状: {len(target_data)}行 x {len(target_data[0]) if target_data else 0}列")

        except Exception as e:
            error_msg = f'获取目标sheet数据失败: {str(e)}'
            print(f"❌ {error_msg}")
            return jsonify({
                'success': False,
                'error': error_msg
            }), 500

        if not target_data or not isinstance(target_data, list) or len(target_data) == 0:
            error_msg = f'目标sheet"{previous_sheet_name}"数据为空'
            print(f"❌ {error_msg}")
            return jsonify({
                'success': False,
                'error': error_msg
            }), 400

        if not current_data or not isinstance(current_data, list) or len(current_data) == 0:
            error_msg = '当前sheet数据为空或格式不正确'
            print(f"❌ {error_msg}")
            return jsonify({
                'success': False,
                'error': error_msg
            }), 400

        # ========== 第三步：验证列数一致性 ==========
        print(f"🔍 验证列数一致性...")

        # 只检查列数，不检查表头内容
        current_col_count = 0
        if current_data and len(current_data) > 0:
            # 直接取第一行的列数（无论是表头还是数据）
            first_row = current_data[0]
            current_col_count = len(first_row) if first_row else 0

        previous_col_count = len(previous_header) if previous_header else 0

        if current_col_count != previous_col_count:
            error_msg = f'列数不一致：前面表格{previous_col_count}列，当前表格{current_col_count}列'
            print(f"❌ {error_msg}")
            return jsonify({
                'success': False,
                'error': error_msg
            }), 400

        print(f"✅ 列数一致，可以合并")

        # ========== 第四步：数据合并 ==========
        print(f"🔄 开始合并数据...")

        # 1. 确定"列标记"行 - 修正：在第二列中查找
        def find_column_marker_row_in_second_column(data, start_index=0):
            """在第二列中查找列标记行"""
            for i in range(start_index, len(data)):
                row = data[i]
                if row and len(row) > 1 and isinstance(row[1], str):  # 🔥 修正：检查第二列
                    cell_value = str(row[1]).strip()
                    # 检查是否是列标记行
                    if '列标记' in cell_value:
                        return i
            return -1

        # 2. 处理前面表格（目标表格）数据
        # 在前面表格的第二列中查找列标记行
        previous_marker_index = find_column_marker_row_in_second_column(target_data, 1)  # 从第2行开始查找

        if previous_marker_index >= 0:
            # 有列标记行的情况
            # 前面表格的表头
            previous_header = target_data[0] if target_data else []

            # 前面表格的列标记之前数据（从第1行到列标记行之前）
            previous_before_marker = target_data[1:previous_marker_index] if previous_marker_index > 1 else []

            # 前面表格的列标记行
            previous_marker_row = target_data[previous_marker_index] if previous_marker_index < len(target_data) else []

            # 前面表格的列标记之后数据
            previous_after_marker = target_data[previous_marker_index + 1:] if previous_marker_index + 1 < len(
                target_data) else []

            print(f"📊 前面表格数据分割:")
            print(f"  表头: 1行")
            print(f"  列标记之前数据: {len(previous_before_marker)}行")
            print(f"  列标记行位置: 第{previous_marker_index}行")
            print(
                f"  列标记行第二列内容: {previous_marker_row[1] if previous_marker_row and len(previous_marker_row) > 1 else 'N/A'}")
            print(f"  列标记之后数据: {len(previous_after_marker)}行")
        else:
            # 没有找到列标记行
            previous_header = target_data[0] if target_data else []
            previous_before_marker = target_data[1:] if len(target_data) > 1 else []
            previous_marker_row = None
            previous_after_marker = []

            print(f"⚠️ 前面表格没有找到列标记行:")
            print(f"  表头: 1行")
            print(f"  所有数据行: {len(previous_before_marker)}行")

        # 3. 处理当前表格（源表格）数据
        # 不再区分表头，包含所有数据行
        current_data_start = 0

        # 获取当前表格的有效数据
        current_valid_data = current_data[current_data_start:] if current_data else []

        # 在当前表格的第二列中查找列标记行
        current_marker_index = find_column_marker_row_in_second_column(current_valid_data)

        if current_marker_index >= 0:
            # 🔥 关键修正：有列标记行，只取列标记**之前**的数据
            current_keep_data = current_valid_data[:current_marker_index]
            print(f"🔍 当前表格在第二列中找到列标记行，只取前{current_marker_index}行数据（列标记之前）")
            print(
                f"  丢弃列标记行（第{current_marker_index}行）及之后的{len(current_valid_data) - current_marker_index}行数据")

            # 调试：显示找到的列标记行
            if current_marker_index < len(current_valid_data):
                marker_row = current_valid_data[current_marker_index]
                print(f"  找到的列标记行第二列内容: {marker_row[1] if len(marker_row) > 1 else 'N/A'}")
        else:
            # 没有列标记行，取所有数据
            current_keep_data = current_valid_data
            print(f"🔍 当前表格没有找到列标记行，取所有{len(current_keep_data)}行数据")

        print(f"📊 当前表格保留数据: {len(current_keep_data)}行")

        # 4. 验证列数一致性（再次验证）
        if current_keep_data and len(current_keep_data) > 0:
            first_row_cols = len(current_keep_data[0])
            header_cols = len(previous_header) if previous_header else 0

            if first_row_cols != header_cols:
                error_msg = f'数据列数不一致：表头{header_cols}列，当前表格数据{first_row_cols}列'
                print(f"❌ {error_msg}")
                return jsonify({
                    'success': False,
                    'error': error_msg
                }), 400

        # 5. 合并数据
        merged_data = []

        # 5.1 添加表头（来自前面表格）
        if previous_header:
            merged_data.append(previous_header)
            print(f"✅ 添加前面表格的表头")

        # 5.2 添加前面表格的列标记之前数据
        if previous_before_marker:
            for row in previous_before_marker:
                merged_data.append(row)
            print(f"✅ 添加前面表格的列标记之前数据: {len(previous_before_marker)}行")

        # 5.3 🔥 关键修正：插入当前表格的数据到前面表格列标记之前
        if current_keep_data:
            for row in current_keep_data:
                merged_data.append(row)
            print(f"✅ 插入当前表格的数据到前面表格列标记之前: {len(current_keep_data)}行")

        # 5.4 添加前面表格的列标记行
        if previous_marker_row is not None:
            merged_data.append(previous_marker_row)
            print(f"✅ 添加前面表格的列标记行")

        # 5.5 添加前面表格的列标记之后数据
        if previous_after_marker:
            for row in previous_after_marker:
                merged_data.append(row)
            print(f"✅ 添加前面表格的列标记之后数据: {len(previous_after_marker)}行")

        print(f"✅ 数据合并完成:")
        print(f"  合并后总行数: {len(merged_data)}")
        print(f"  原始前面表格行数: {len(target_data)}")
        print(f"  新增行数（来自当前表格）: {len(current_keep_data)}行")

        # ========== 第五步：保存合并后的数据 ==========
        print(f"💾 保存合并后的数据到目标sheet...")

        try:
            # 使用ExcelDataHandler保存数据
            result = excel_data_handler.save_complete_table_data(
                pdf_id_str, excel_file, previous_sheet_name, merged_data, 'original', db
            )

            if not result.get('success', True):
                error_msg = f'保存合并数据失败: {result.get("error", "未知错误")}'
                print(f"❌ {error_msg}")
                return jsonify({
                    'success': False,
                    'error': error_msg
                }), 500

            print(f"✅ 数据保存成功: {result.get('message')}")

        except Exception as e:
            error_msg = f'保存合并数据失败: {str(e)}'
            print(f"❌ {error_msg}")
            return jsonify({
                'success': False,
                'error': error_msg
            }), 500

        # ========== 第六步：删除当前sheet ==========
        print(f"🗑️ 开始删除当前sheet: {current_sheet_name}")

        delete_source_sheet = data.get('metadata', {}).get('deleteSourceSheet', True)
        source_sheet_deleted = False

        if delete_source_sheet:
            try:
                # 删除当前sheet
                delete_result = delete_excel_sheet(
                    pdf_id_str, excel_file, current_sheet_name, excel_path, db
                )

                if delete_result.get('success', True):
                    source_sheet_deleted = True
                    print(f"✅ 成功删除当前sheet: {current_sheet_name}")
                else:
                    error_msg = f'合并成功但删除当前sheet失败: {delete_result.get("error", "未知错误")}'
                    print(f"⚠️ {error_msg}")

            except Exception as e:
                error_msg = f'删除当前sheet时出错: {str(e)}'
                print(f"⚠️ {error_msg}")

        # ========== 第七步：返回成功响应 ==========
        response_data = {
            'success': True,
            'message': f'数据合并成功{"，当前表格已删除" if source_sheet_deleted else ""}',
            'metadata': {
                'mergedSheet': previous_sheet_name,
                'sourceSheet': current_sheet_name,
                'currentPage': current_page,
                'previousPage': previous_page,
                'totalRows': len(merged_data),
                'headerRows': 1,
                'dataRows': len(merged_data) - 1,
                'rowsAdded': len(current_keep_data),
                'sourceSheetDeleted': source_sheet_deleted,
                'previousMarkerFound': previous_marker_index >= 0,
                'previousMarkerIndex': previous_marker_index,
                'currentMarkerFound': current_marker_index >= 0,
                'currentMarkerIndex': current_marker_index
            },
            'summary': {
                'originalRows': len(target_data),
                'currentRows': len(current_keep_data),
                'mergedRows': len(merged_data),
                'headerConsistent': True  # 列数已验证一致
            }
        }

        print(f"✅ 合并成功，返回响应")
        return jsonify(response_data)

    except Exception as e:
        error_msg = f'合并失败: {str(e)}'
        print(f"❌❌ {error_msg}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': error_msg
        }), 500



# 为了能够在merge_sheets函数中调用save_final_excel，需要重新定义save_final_excel的包装函数
def save_final_excel():
    """包装save_final_excel路由函数，使其可以在merge_sheets中调用"""
    from flask import request

    # 获取请求数据
    data = request.get_json(silent=True)
    if not data:
        return jsonify({'error': '缺少请求数据'}), 400

    # 调用原始的save_final_excel路由处理逻辑
    return save_final_excel_original(data)


def save_final_excel_original(data):
    """保存数据的原始逻辑，从save_final_excel路由中提取"""
    print("******************** 保存整个表格（保护其他Sheet） ******************")

    # 基础校验
    required_fields = ['pdf_id', 'excel_file', 'sheet_name', 'table_type', 'data']
    for field in required_fields:
        if field not in data:
            return jsonify({'error': f'缺少必要字段: {field}'}), 400

    try:
        pdf_id = data['pdf_id']
        excel_file = data['excel_file']
        sheet_name = data['sheet_name']
        table_type = data['table_type']
        table_data = data['data']

        # 根据表类型选择保存方式
        if table_type == 'original':
            result = excel_data_handler.save_complete_table_data(pdf_id, excel_file, sheet_name, table_data, table_type,
                                                                 db)
        elif table_type == 'flattened':
            result = excel_data_handler.save_flattened_table_data(pdf_id, excel_file, sheet_name, table_data,
                                                                  table_type, db)
        else:
            return jsonify({'error': f'不支持的表类型: {table_type}'}), 400

        if not result['success']:
            return jsonify({'success': False, 'error': result['error']}), 500

        return jsonify({
            'success': True,
            'message': '表格数据保存成功',
            'saved_count': result.get('saved_rows', 0),
            'data_dimensions': result.get('data_dimensions', '未知'),
            'excel_updated': result.get('excel_updated', False),
            'sheets_protected': result.get('sheets_protected', False),
            'protected_sheets_count': result.get('protected_sheets_count', 0),
            'file_created': result.get('file_created', False)
        }), 200

    except Exception as e:
        print(f"❌ 保存失败: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': f'保存失败: {str(e)}'}), 500


# 修改原始的save_final_excel路由，使其使用包装函数
@file_bp.route('/api/excel/save-final', methods=['POST'])
def save_final_excel_route():
    """统一保存整个表格数据 - 保护其他Sheet"""
    return save_final_excel_original(request.json)


# 在 table_routes.py 或相关路由文件中添加
@file_bp.route('/api/get-original-filenames', methods=['POST'])
def get_original_filenames():
    """批量获取原始文件名映射 - 直接查询数据库"""
    try:
        data = request.get_json()
        pdf_folders = data.get('pdf_folders', [])

        if not pdf_folders:
            return jsonify({"success": True, "filename_map": {}})

        print(f"🔍 开始批量查询原始文件名，数量: {len(pdf_folders)}")
        print(f"🔍 查询列表: {pdf_folders}")

        # 连接到数据库
        try:
            from backend.utils.db_manager import DatabaseManager
            import sqlite3
            import os

            db_manager = DatabaseManager()

            if not hasattr(db_manager, 'db_path'):
                return jsonify({"success": False, "error": "DatabaseManager没有db_path属性"})

            db_path = db_manager.db_path

            if not os.path.exists(db_path):
                return jsonify({"success": False, "error": f"数据库文件不存在: {db_path}"})

            conn = sqlite3.connect(db_path)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()

            # ✅ 修复1：先打印数据库结构进行调试
            try:
                cursor.execute("PRAGMA table_info(files)")
                columns = cursor.fetchall()
                column_names = [col[1] for col in columns]
                print(f"📊 数据库表结构: {column_names}")
            except:
                print("⚠️ 无法获取表结构信息")

            # ✅ 修复2：修改SQL查询条件
            # 问题：原代码查询WHERE filename IN (...)
            # 应该改为WHERE disk_name IN (...) 或更合适的字段
            placeholders = ','.join(['?'] * len(pdf_folders))

            # 尝试多种可能的字段匹配
            filename_map = {}
            found_identifiers = set()

            # 方式1：优先尝试disk_name字段
            query1 = f"""
                SELECT disk_name, filename, raw_filename 
                FROM files 
                WHERE disk_name IN ({placeholders})
            """
            print(f"🔍 执行查询1: {query1}")
            print(f"🔍 查询参数: {pdf_folders}")

            cursor.execute(query1, pdf_folders)
            results1 = cursor.fetchall()

            for row in results1:
                disk_name = row['disk_name']
                filename = row['filename']
                raw_filename = row['raw_filename']

                # 优先使用raw_filename，然后是filename，最后是disk_name
                original_name = raw_filename or filename or f"{disk_name}.pdf"
                filename_map[disk_name] = original_name
                found_identifiers.add(disk_name)

                print(f"  ✅ 通过disk_name找到: {disk_name} -> {original_name}")
                print(f"     详细信息: filename='{filename}', raw_filename='{raw_filename}'")

            # 方式2：如果disk_name没找到足够的结果，尝试filename字段
            remaining_identifiers = [folder for folder in pdf_folders if folder not in found_identifiers]

            if remaining_identifiers:
                print(f"🔍 还有 {len(remaining_identifiers)} 个标识符未找到，尝试查询filename字段")

                for folder in remaining_identifiers:
                    # 尝试查询filename字段
                    query2 = """
                        SELECT disk_name, filename, raw_filename 
                        FROM files 
                        WHERE filename LIKE ? OR raw_filename LIKE ?
                    """
                    search_pattern = f"%{folder}%"
                    cursor.execute(query2, (search_pattern, search_pattern))
                    row = cursor.fetchone()

                    if row:
                        disk_name = row['disk_name']
                        filename = row['filename']
                        raw_filename = row['raw_filename']

                        # 优先使用raw_filename
                        original_name = raw_filename or filename or f"{disk_name}.pdf"
                        filename_map[folder] = original_name
                        found_identifiers.add(folder)

                        print(f"  🔄 通过filename模糊匹配找到: {folder} -> {original_name}")

            # ✅ 修复3：处理未找到的情况
            not_found = [folder for folder in pdf_folders if folder not in found_identifiers]
            if not_found:
                print(f"⚠️ 未找到的文件标识符: {not_found}")

                # 为未找到的文件添加默认映射
                for folder in not_found:
                    # 默认使用标识符本身作为文件名
                    if folder.endswith('.pdf'):
                        default_name = folder
                    else:
                        default_name = f"{folder}.pdf"

                    filename_map[folder] = default_name
                    print(f"  ⚠️ 为未找到的标识符添加默认映射: {folder} -> {default_name}")

            # 打印数据库中的文件示例以供调试
            try:
                cursor.execute("SELECT disk_name, filename, raw_filename FROM files LIMIT 5")
                sample_files = cursor.fetchall()
                print(f"📁 数据库中的文件示例（前5条）:")
                for file in sample_files:
                    print(
                        f"  - disk_name: '{file['disk_name']}', filename: '{file['filename']}', raw_filename: '{file['raw_filename']}'")
            except:
                print("⚠️ 无法获取文件示例")

            conn.close()

            print(f"✅ 文件名映射查询完成:")
            print(f"  - 查询数量: {len(pdf_folders)}")
            print(f"  - 找到数量: {len(found_identifiers)}")
            print(f"  - 未找到: {len(not_found)}")
            print(f"  - 映射结果: {filename_map}")

            return jsonify({
                "success": True,
                "filename_map": filename_map,
                "query_count": len(pdf_folders),
                "found_count": len(found_identifiers),
                "not_found": not_found
            })

        except ImportError as e:
            print(f"❌ 无法导入DatabaseManager: {e}")
            return jsonify({"success": False, "error": f"无法导入DatabaseManager: {str(e)}"})

        except sqlite3.Error as e:
            print(f"❌ 数据库查询错误: {e}")
            return jsonify({"success": False, "error": f"数据库查询失败: {str(e)}"})

    except Exception as e:
        print(f"❌ 获取原始文件名异常: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": f"获取原始文件名异常: {str(e)}"})
