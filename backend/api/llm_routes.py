# -*- coding:utf-8 -*-
"""
LLM 表格识别 API 路由
"""

import time
from flask import Blueprint, request, jsonify
import logging
from backend.service.table_llm_service import get_table_processor
from backend.schemas.table_schemas import ExcelSaveConfig
from backend.utils.constants import MAIN_ROOT

from pprint import pprint

# 创建蓝图
llm_bp = Blueprint('llm', __name__)

# 设置日志
logger = logging.getLogger(__name__)

# llm_routes.py
from flask import Blueprint, request, jsonify, send_from_directory

# llm_routes.py
from flask import Blueprint, request, jsonify
import asyncio
import logging
from pathlib import Path

# 全局处理器实例
_table_processor_instance = None
_non_financial_table_service = None


def validate_required_params(data, required_fields):
    """验证必要参数"""
    missing_fields = [field for field in required_fields if not data.get(field)]
    if missing_fields:
        return False, f"缺少必要参数: {', '.join(missing_fields)}"
    return True, None


async def _process_single_image(data):
    """处理单张图片的异步函数"""
    try:
        # 验证必要参数
        is_valid, error_msg = validate_required_params(
            data, ['image_path', 'output_path']
        )

        print("###################")
        print(is_valid, error_msg)

        if not is_valid:
            return {
                "success": False,
                "error": error_msg
            }

        image_path = data.get('image_path')
        output_path = data.get('output_path')
        sheet_name = data.get('sheet_name', '识别结果')
        bank_name = data.get('bank_name', '未知银行')
        file_name = data.get('file_name', Path(image_path).stem)

        print("************image_path****output_path****************")
        print(image_path)
        print(output_path)
        print(sheet_name)
        print(bank_name)
        print(file_name)

        # 检查图片文件是否存在
        if not Path(image_path).exists():
            return {
                "success": False,
                "error": f"图片文件不存在: {image_path}"
            }

        processor = get_table_processor()

        print("****************processor.llm_client******************")
        print(processor.llm_client)

        if not processor.llm_client:
            return {
                "success": False,
                "error": "请先配置LLM客户端"
            }

        # 使用已有的标识创建Excel存储文件夹
        # 从 image_path 提取标识：类似 d0586abf1323dbfd80a926ce1e2d5676
        folder_name = Path(image_path).stem.split('_')[0]  # 取第一个下划线前的部分

        excel_base_dir = Path("static/excel_data")
        excel_dir = excel_base_dir / folder_name
        excel_dir.mkdir(parents=True, exist_ok=True)

        # 重新构建输出路径
        excel_filename = Path(output_path).name
        new_output_path = excel_dir / excel_filename
        output_path = str(new_output_path)

        print(f"Excel文件将保存到: {output_path}")

        # 配置Excel保存参数
        excel_config = ExcelSaveConfig(
            anchor_cell=data.get('anchor_cell', 'R2'),
            width_px=data.get('width_px', 768),
            mode=data.get('mode', 'overwrite')
        )

        # 处理图片 - 使用新的方法名和参数
        result = await processor.process_table_pipeline(
            image_path=image_path,
            out_file=output_path,
            sheet_name=sheet_name,
            bank_name=bank_name,
            file_name=file_name,
            excel_config=excel_config
        )

        logger.info(f"表格识别完成 - 状态: {result.status}, 文件: {output_path}")

        return {
            "success": result.status == "success",
            "data": {
                "status": result.status,
                "complexity": result.complexity,
                "mode": result.mode,
                "table_name": result.table_name,
                "assessment_reason": result.assessment_reason,
                "error_message": result.error_message,
                "output_path": output_path  # 返回实际的输出路径
            }
        }

    except Exception as e:
        logger.error(f"表格识别失败: {str(e)}")
        return {
            "success": False,
            "error": f"处理失败: {str(e)}"
        }




async def _test_connection_internal(data):
    """测试LLM连接的内部异步函数"""
    try:
        # 验证必要参数
        is_valid, error_msg = validate_required_params(
            data, ['base_url', 'api_key', 'model_id']
        )
        if not is_valid:
            return {
                "success": False,
                "error": error_msg
            }

        base_url = data.get('base_url')
        api_key = data.get('api_key')
        model_id = data.get('model_id')

        print(
            f"🔧 测试连接参数: base_url={base_url}, model_id={model_id}, api_key_length={len(api_key) if api_key else 0}")

        # 确保URL格式正确
        if not base_url.endswith('/'):
            base_url = base_url + '/'

        # 创建临时客户端进行测试
        from openai import AsyncOpenAI

        try:
            test_client = AsyncOpenAI(
                base_url=base_url,
                api_key=api_key
            )

            # 测试连接 - 使用更简单的消息
            print(f"🔧 开始API调用测试...")
            response = await test_client.chat.completions.create(
                model=model_id,
                messages=[{"role": "user", "content": "Hello, please respond with just 'OK'"}],
                max_tokens=10,
                timeout=30.0  # 添加超时设置
            )

            print(f"🔧 API响应: {response}")

            return {
                "success": True,
                "message": "LLM连接测试成功",
                "data": {
                    "model_response": bool(response.choices),
                    "response_content": response.choices[0].message.content if response.choices else None
                }
            }

        except Exception as api_error:
            print(f"❌ API调用失败: {str(api_error)}")
            error_detail = str(api_error)

            # 提供更详细的错误信息
            if "401" in error_detail:
                error_msg = "API密钥无效或未授权"
            elif "404" in error_detail:
                error_msg = "模型不存在或URL路径错误"
            elif "connect" in error_detail.lower():
                error_msg = "无法连接到服务器，请检查网络和URL"
            elif "timeout" in error_detail.lower():
                error_msg = "连接超时，请检查网络或服务器状态"
            else:
                error_msg = f"API调用失败: {error_detail}"

            return {
                "success": False,
                "error": error_msg
            }

    except Exception as e:
        print(f"❌ 连接测试异常: {str(e)}")
        return {
            "success": False,
            "error": f"连接测试失败: {str(e)}"
        }


@llm_bp.route('/llm/test-connection', methods=['POST'])
def test_connection():
    """测试LLM连接"""
    try:
        data = request.get_json()

        if not data:
            return jsonify({
                "success": False,
                "error": "请求体不能为空"
            }), 400

        result = asyncio.run(_test_connection_internal(data))
        return jsonify(result)
    except Exception as e:
        logger.error(f"测试连接错误: {str(e)}")
        return jsonify({
            "success": False,
            "error": f"测试连接错误: {str(e)}"
        }), 500





@llm_bp.route('/llm/available-models', methods=['GET'])
def get_available_models():
    """获取可用的模型列表"""
    models = [
        {
            "id": "doubao-1-5-vision-pro-250328",
            "name": "豆包视觉专业版",
            "description": "支持视觉识别的专业模型",
            "max_tokens": 16000
        },
        {
            "id": "doubao-seed-1-6-vision-250815",
            "name": "豆包视觉种子版",
            "description": "视觉识别基础模型",
            "max_tokens": 16000
        }
    ]

    return jsonify({
        "success": True,
        "data": models
    })




@llm_bp.route('/llm/health', methods=['GET'])
def health_check():
    """健康检查端点"""
    try:
        processor = get_table_processor()

        # 修复：确保所有值都是可JSON序列化的
        base_url = getattr(processor.llm_client, 'base_url', None)
        if base_url is not None:
            base_url = str(base_url)

        health_status = {
            "service": "running",
            "llm_configured": processor.llm_client is not None,
            "model_id": processor.model_id,
            "base_url": base_url,  # 使用转换后的字符串
            "timestamp": time.time()  # 使用time.time()而不是asyncio的时间
        }

        return jsonify({
            "success": True,
            "data": health_status
        })
    except Exception as e:
        logger.error(f"健康检查失败: {str(e)}")
        return jsonify({
            "success": False,
            "error": f"健康检查失败: {str(e)}"
        }), 500


@llm_bp.route('/llm/recognize-table', methods=['POST'])
def recognize_table():
    """识别表格并保存到指定路径 - 先检查Excel是否存在"""
    try:
        data = request.get_json()
        print(f"收到识别请求: {data}")

        if not data:
            return jsonify({
                "success": False,
                "error": "请求体不能为空"
            }), 400

        # 验证必要参数
        required_fields = ['imageUrl', 'excelPath']
        missing_fields = [field for field in required_fields if not data.get(field)]
        if missing_fields:
            return jsonify({
                "success": False,
                "error": f"缺少必要参数: {', '.join(missing_fields)}"
            }), 400

        image_url = data.get('imageUrl')
        excel_path = data.get('excelPath')
        table_name = data.get('tableName', '识别结果')
        index = data.get('index', 0)

        print(f"识别参数 - imageUrl: {image_url}, excelPath: {excel_path}")

        # 构建完整Excel路径
        excel_full_path = Path(excel_path)
        if not excel_full_path.is_absolute():
            excel_full_path = Path.cwd() / excel_path

        print(f"检查Excel文件是否存在: {excel_full_path}")

        # 1. 首先检查Excel文件是否已经存在
        if excel_full_path.exists():
            print("✅ Excel文件已存在，直接读取数据")
            try:
                import openpyxl
                workbook = openpyxl.load_workbook(excel_full_path)
                sheet_name = workbook.sheetnames[0]
                worksheet = workbook[sheet_name]

                # 转换为JSON数据
                headers = []
                data = []

                # 读取表头（第一行）
                for col in range(1, worksheet.max_column + 1):
                    cell_value = worksheet.cell(row=1, column=col).value
                    headers.append(str(cell_value) if cell_value is not None else f"列{col}")

                # 读取数据行
                for row in range(2, worksheet.max_row + 1):
                    row_data = {}
                    for col, header in enumerate(headers, 1):
                        cell_value = worksheet.cell(row=row, column=col).value
                        row_data[header] = str(cell_value) if cell_value is not None else ""
                    if any(row_data.values()):  # 只添加非空行
                        data.append(row_data)

                return jsonify({
                    "success": True,
                    "recognizedData": {
                        "headers": headers,
                        "data": data,
                        "tableName": table_name,
                        "excelPath": str(excel_full_path)
                    },
                    "message": "已加载现有表格数据",
                    "fromCache": True  # 添加标记表示来自缓存
                })
            except Exception as e:
                logger.error(f"读取现有Excel失败: {str(e)}")
                # 如果读取失败，继续走LLM识别流程
                print("❌ 读取现有Excel失败，继续走LLM识别")

        # 2. 如果Excel不存在，进行LLM识别
        print("🔄 Excel文件不存在，开始LLM识别流程")

        # ⭐⭐⭐ 关键修改：改进图片路径提取逻辑 ⭐⭐⭐
        image_path = image_url

        # 处理各种URL格式
        if image_url.startswith('http://'):
            # 提取域名后的路径部分
            from urllib.parse import urlparse
            parsed_url = urlparse(image_url)
            image_path = parsed_url.path  # 获取路径部分，如 /static/joined_tables/...

            # 去掉开头的斜杠
            if image_path.startswith('/'):
                image_path = image_path[1:]

            print(f"✅ 从URL提取路径: {image_path}")

        elif image_url.startswith('/'):
            # 如果是以/开头的绝对路径，去掉开头的斜杠
            image_path = image_url[1:]
            print(f"✅ 处理绝对路径: {image_path}")

        print(f"处理图片路径: {image_path}")

        # 检查图片文件是否存在
        image_full_path = Path(image_path)
        if not image_full_path.exists():
            # 尝试在static目录下查找
            static_path = Path("static") / image_path
            if static_path.exists():
                image_full_path = static_path
                print(f"✅ 在static目录找到图片: {static_path}")
            else:
                # 尝试直接在当前目录查找
                current_dir_path = Path.cwd() / image_path
                if current_dir_path.exists():
                    image_full_path = current_dir_path
                    print(f"✅ 在当前目录找到图片: {current_dir_path}")
                else:
                    print(f"❌ 图片文件不存在，尝试的路径:")
                    print(f"   - {image_full_path}")
                    print(f"   - {static_path}")
                    print(f"   - {current_dir_path}")
                    return jsonify({
                        "success": False,
                        "error": f"图片文件不存在: {image_path}"
                    }), 404

        print(f"✅ 最终使用的图片路径: {image_full_path}")

        # 确保Excel目录存在
        excel_full_path.parent.mkdir(parents=True, exist_ok=True)

        print(f"Excel保存路径: {excel_full_path}")

        # 调用 _process_single_image 异步函数进行LLM识别
        process_data = {
            'image_path': str(image_full_path),
            'output_path': str(excel_full_path),
            'sheet_name': table_name,
            'bank_name': '未知银行',
            'file_name': f"table_{index + 1}"
        }

        # 使用异步处理
        result = asyncio.run(_process_single_image(process_data))
        print(f"LLM识别结果: {result}")

        if result.get('success'):
            # 读取生成的Excel文件数据返回给前端
            output_path = result['data']['output_path']
            excel_result_path = Path(output_path)

            if excel_result_path.exists():
                try:
                    import openpyxl
                    workbook = openpyxl.load_workbook(excel_result_path)
                    sheet_name = workbook.sheetnames[0]
                    worksheet = workbook[sheet_name]

                    # 转换为JSON数据
                    headers = []
                    data = []

                    # 读取表头（第一行）
                    for col in range(1, worksheet.max_column + 1):
                        cell_value = worksheet.cell(row=1, column=col).value
                        headers.append(str(cell_value) if cell_value is not None else f"列{col}")

                    # 读取数据行
                    for row in range(2, worksheet.max_row + 1):
                        row_data = {}
                        for col, header in enumerate(headers, 1):
                            cell_value = worksheet.cell(row=row, column=col).value
                            row_data[header] = str(cell_value) if cell_value is not None else ""
                        if any(row_data.values()):  # 只添加非空行
                            data.append(row_data)

                    return jsonify({
                        "success": True,
                        "recognizedData": {
                            "headers": headers,
                            "data": data,
                            "tableName": table_name,
                            "excelPath": str(excel_result_path)
                        },
                        "message": "表格识别完成",
                        "fromCache": False  # 添加标记表示来自LLM识别
                    })
                except Exception as e:
                    logger.error(f"读取生成的Excel失败: {str(e)}")
                    # 即使读取失败，也返回成功，因为Excel文件已经生成
                    return jsonify({
                        "success": True,
                        "recognizedData": {
                            "headers": ["识别完成"],
                            "data": [{"状态": "表格识别完成，但读取数据失败"}],
                            "tableName": table_name,
                            "excelPath": str(excel_result_path)
                        },
                        "message": "表格识别完成",
                        "fromCache": False
                    })
            else:
                return jsonify({
                    "success": False,
                    "error": "识别完成但Excel文件未生成"
                }), 500
        else:
            return jsonify({
                "success": False,
                "error": result.get('error', '识别失败')
            }), 500

    except Exception as e:
        logger.error(f"表格识别失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": f"表格识别失败: {str(e)}"
        }), 500


@llm_bp.route('/llm/check-excel', methods=['GET'])
def check_excel():
    """检查Excel文件是否存在"""
    try:
        file_path = request.args.get('path')
        print(f"🔍 检查Excel文件请求 - path: {file_path}")

        if not file_path:
            return jsonify({
                "success": False,
                "error": "缺少path参数"
            }), 400

        # 构建完整路径
        full_path = Path(file_path)
        if not full_path.is_absolute():
            full_path = Path.cwd() / file_path

        print(f"🔍 检查Excel完整路径: {full_path}")
        print(f"🔍 文件是否存在: {full_path.exists()}")

        if full_path.exists():
            try:
                import openpyxl
                workbook = openpyxl.load_workbook(full_path)
                sheet_name = workbook.sheetnames[0]
                worksheet = workbook[sheet_name]

                # 转换为JSON数据
                data = []
                headers = []

                # 读取表头（第一行）
                for col in range(1, worksheet.max_column + 1):
                    cell_value = worksheet.cell(row=1, column=col).value
                    headers.append(str(cell_value) if cell_value is not None else f"列{col}")

                # 读取数据行
                for row in range(2, worksheet.max_row + 1):
                    row_data = {}
                    for col, header in enumerate(headers, 1):
                        cell_value = worksheet.cell(row=row, column=col).value
                        row_data[header] = str(cell_value) if cell_value is not None else ""
                    if any(row_data.values()):  # 只添加非空行
                        data.append(row_data)

                print(f"🔍 Excel数据读取成功 - 表头: {headers}, 数据行数: {len(data)}")

                return jsonify({
                    "success": True,
                    "exists": True,
                    "excelData": {
                        "headers": headers,
                        "data": data,
                        "sheet_name": sheet_name,
                        "file_path": str(full_path)
                    }
                })
            except Exception as e:
                logger.error(f"读取Excel文件失败: {str(e)}")
                return jsonify({
                    "success": False,
                    "error": f"读取Excel文件失败: {str(e)}"
                }), 500
        else:
            print(f"🔍 Excel文件不存在: {file_path}")
            return jsonify({
                "success": True,
                "exists": False,
                "message": f"Excel文件不存在: {file_path}"
            })

    except Exception as e:
        logger.error(f"检查Excel失败: {str(e)}")
        return jsonify({
            "success": False,
            "error": f"检查Excel失败: {str(e)}"
        }), 500



@llm_bp.route('/llm/get-excel-data', methods=['GET'])
def get_excel_data():
    """读取Excel文件内容并返回给前端"""
    try:
        excel_url = request.args.get('url')
        if not excel_url:
            return jsonify({
                "success": False,
                "error": "缺少url参数"
            }), 400

        # 将URL转换为文件路径
        if excel_url.startswith('/static/'):
            file_path = excel_url[1:]  # 去掉开头的斜杠
        else:
            file_path = excel_url

        full_path = Path(file_path)
        if not full_path.is_absolute():
            full_path = Path.cwd() / file_path

        print(f"读取Excel文件: {full_path}")

        if not full_path.exists():
            return jsonify({
                "success": False,
                "error": f"Excel文件不存在: {file_path}"
            }), 404

        # 读取Excel文件
        import openpyxl
        workbook = openpyxl.load_workbook(full_path)
        sheet_name = workbook.sheetnames[0]
        worksheet = workbook[sheet_name]

        # 转换为JSON数据
        headers = []
        data_rows = []

        # 读取表头（第一行）
        for col in range(1, worksheet.max_column + 1):
            cell_value = worksheet.cell(row=1, column=col).value
            headers.append(str(cell_value) if cell_value is not None else f"列{col}")

        # 读取数据行
        for row in range(2, worksheet.max_row + 1):
            row_data = {}
            for col, header in enumerate(headers, 1):
                cell_value = worksheet.cell(row=row, column=col).value
                row_data[header] = str(cell_value) if cell_value is not None else ""
            if any(row_data.values()):  # 只添加非空行
                data_rows.append(row_data)

        return jsonify({
            "success": True,
            "data": {
                "headers": headers,
                "data": data_rows,
                "tableName": sheet_name,
                "excelPath": str(full_path)
            }
        })

    except Exception as e:
        logger.error(f"读取Excel数据失败: {str(e)}")
        return jsonify({
            "success": False,
            "error": f"读取Excel数据失败: {str(e)}"
        }), 500



# 在 llm_routes.py 中添加以下内容

from backend.service.non_financial_table_service import NonFinancialTableService

# 单例实例
_non_financial_table_service = None

def get_non_financial_table_service():
    """获取普通表格服务单例实例"""
    global _non_financial_table_service

    if _non_financial_table_service is None:
        _non_financial_table_service = NonFinancialTableService()

    return _non_financial_table_service



async def _process_non_financial_table(data):
    """处理普通表格的异步函数"""
    try:
        # 验证必要参数
        is_valid, error_msg = validate_required_params(
            data, ['image_path', 'output_path']
        )

        if not is_valid:
            return {
                "success": False,
                "error": error_msg
            }

        image_path = data.get('image_path')
        output_path = data.get('output_path')
        sheet_name = data.get('sheet_name', '普通表格识别结果')
        bank_name = data.get('bank_name', '未知机构')
        file_name = data.get('file_name', Path(image_path).stem)

        print("普通表格处理参数:")
        print(f"image_path: {image_path}")
        print(f"output_path: {output_path}")
        print(f"sheet_name: {sheet_name}")
        print(f"bank_name: {bank_name}")
        print(f"file_name: {file_name}")

        # 检查图片文件是否存在
        if not Path(image_path).exists():
            return {
                "success": False,
                "error": f"图片文件不存在: {image_path}"
            }

        service = get_non_financial_table_service()

        if not service.llm_client:
            return {
                "success": False,
                "error": "请先配置LLM客户端"
            }

        # 使用已有的标识创建Excel存储文件夹
        folder_name = Path(image_path).stem.split('_')[0]
        excel_base_dir = Path("static/excel_data")
        excel_dir = excel_base_dir / folder_name
        excel_dir.mkdir(parents=True, exist_ok=True)

        # 重新构建输出路径
        excel_filename = Path(output_path).name
        new_output_path = excel_dir / excel_filename
        output_path = str(new_output_path)

        print(f"普通表格Excel文件将保存到: {output_path}")

        # 配置Excel保存参数
        excel_config = ExcelSaveConfig(
            anchor_cell=data.get('anchor_cell', 'R2'),
            width_px=data.get('width_px', 768),
            mode=data.get('mode', 'overwrite')
        )

        # 处理普通表格
        result = await service.process_table_pipeline(
            image_path=image_path,
            out_file=output_path,
            sheet_name=sheet_name,
            bank_name=bank_name,
            file_name=file_name,
            excel_config=excel_config
        )

        logger.info(f"普通表格识别完成 - 状态: {result.status}, 文件: {output_path}")

        return {
            "success": result.status == "success",
            "data": {
                "status": result.status,
                "complexity": result.complexity,
                "mode": result.mode,
                "table_name": result.table_name,
                "assessment_reason": result.assessment_reason,
                "error_message": result.error_message,
                "output_path": output_path,
                "table_type": "non_financial"
            }
        }

    except Exception as e:
        logger.error(f"普通表格识别失败: {str(e)}")
        return {
            "success": False,
            "error": f"处理失败: {str(e)}"
        }


@llm_bp.route('/llm/configure', methods=['POST'])
def configure_llm():
    """配置LLM参数"""
    try:
        data = request.get_json()
        print(f"🔧 收到配置请求: {data}")  # 调试信息

        if not data:
            return jsonify({
                "success": False,
                "error": "请求体不能为空"
            }), 400

        # 验证必要参数
        is_valid, error_msg = validate_required_params(
            data, ['base_url', 'api_key', 'model_id']
        )
        if not is_valid:
            return jsonify({
                "success": False,
                "error": error_msg
            }), 400

        base_url = data.get('base_url')
        api_key = data.get('api_key')
        model_id = data.get('model_id')
        table_type = data.get('table_type', 'financial')  # 新增：表格类型，默认为金融表格
        prompts = data.get('prompts', {})

        print(
            f"🔧 解析配置参数: base_url={base_url}, model_id={model_id}, table_type={table_type}, prompts_keys={list(prompts.keys())}")

        # 根据表格类型选择不同的处理器
        processor = None
        processor_type = ""

        if table_type == 'financial':
            from backend.service.table_llm_service import TableLLMService
            processor = TableLLMService(
                llm_client=None,  # 让服务自己创建客户端
                model_id=model_id
            )
            processor_type = "金融表格处理器"
        else:
            from backend.service.non_financial_table_service import NonFinancialTableService
            processor = NonFinancialTableService(
                llm_client=None,  # 让服务自己创建客户端
                model_id=model_id
            )
            processor_type = "普通表格处理器"

        print(
            f"🔧 使用处理器: {processor_type}, llm_client={processor.llm_client is not None}, model_id={processor.model_id}")

        # 检查客户端是否配置成功
        if not processor.llm_client:
            return jsonify({
                "success": False,
                "error": "LLM客户端配置失败，请检查API密钥和URL是否正确"
            }), 400

        # 更新提示词配置
        if prompts:
            if hasattr(processor, 'prompt_registry'):
                # TableLLMService 使用 prompt_registry
                for prompt_type, prompt_content in prompts.items():
                    if prompt_type in processor.prompt_registry:
                        processor.prompt_registry[prompt_type] = prompt_content
            elif hasattr(processor, 'prompt'):
                # NonFinancialTableService 使用 prompt
                if 'non_financial' in prompts:
                    processor.prompt = prompts['non_financial']

        # 保存配置到全局变量，以便其他接口使用
        global _table_processor_instance, _non_financial_table_service

        if table_type == 'financial':
            _table_processor_instance = processor
            # 清除普通表格服务实例，确保一致性
            _non_financial_table_service = None
        else:
            _non_financial_table_service = processor
            # 清除金融表格服务实例，确保一致性
            _table_processor_instance = None

        logger.info(f"LLM配置成功 - 模型: {model_id}, 基础URL: {base_url}, 表格类型: {table_type}")

        return jsonify({
            "success": True,
            "message": f"LLM配置成功（{processor_type}）",
            "data": {
                "model_id": processor.model_id,
                "base_url": base_url,
                "table_type": table_type,
                "processor_type": processor_type,
                "prompts_configured": list(prompts.keys()) if prompts else [],
                "client_configured": processor.llm_client is not None
            }
        })

    except Exception as e:
        logger.error(f"LLM配置失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": f"配置失败: {str(e)}"
        }), 500


# 修改获取处理器状态的接口，返回表格类型信息
@llm_bp.route('/llm/status', methods=['GET'])
def get_processor_status():
    """获取处理器状态"""
    try:
        # 检查当前配置的处理器类型
        current_processor = None
        table_type = "unknown"
        processor_type = "未知"

        if _table_processor_instance is not None:
            current_processor = _table_processor_instance
            table_type = "financial"
            processor_type = "金融表格处理器"
        elif _non_financial_table_service is not None:
            current_processor = _non_financial_table_service
            table_type = "non_financial"
            processor_type = "普通表格处理器"
        else:
            # 如果没有配置，使用默认的金融表格处理器
            from backend.service.table_llm_service import get_table_processor
            current_processor = get_table_processor()
            table_type = "financial"
            processor_type = "金融表格处理器（默认）"

        # 修复：确保所有值都是可JSON序列化的
        base_url = getattr(current_processor.llm_client, 'base_url', None)
        if base_url is not None:
            base_url = str(base_url)  # 将URL对象转换为字符串

        status = {
            "client_configured": current_processor.llm_client is not None,
            "model_id": current_processor.model_id,
            "base_url": base_url,  # 使用转换后的字符串
            "table_type": table_type,
            "processor_type": processor_type,
            "prompts_configured": {}
        }

        # 添加提示词配置信息
        if hasattr(current_processor, 'prompt_registry'):
            status["prompts_configured"] = {
                prompt_type: bool(content) and len(content.strip()) > 0
                for prompt_type, content in current_processor.prompt_registry.items()
            }
        elif hasattr(current_processor, 'prompt'):
            status["prompts_configured"] = {
                "non_financial": bool(current_processor.prompt) and len(current_processor.prompt.strip()) > 0
            }

        return jsonify({
            "success": True,
            "data": status
        })
    except Exception as e:
        logger.error(f"获取状态失败: {str(e)}")
        return jsonify({
            "success": False,
            "error": f"获取状态失败: {str(e)}"
        }), 500


# 添加一个工具函数，根据表格类型获取合适的处理器
def get_appropriate_processor(table_type=None):
    """根据表格类型获取合适的处理器"""
    if table_type == 'non_financial' or _non_financial_table_service is not None:
        if _non_financial_table_service is not None:
            return _non_financial_table_service
        else:
            from backend.service.non_financial_table_service import get_non_financial_table_service
            return get_non_financial_table_service()
    else:
        if _table_processor_instance is not None:
            return _table_processor_instance
        else:
            from backend.service.table_llm_service import get_table_processor
            return get_table_processor()


# 修改处理图片的接口，根据配置自动选择服务
@llm_bp.route('/llm/process-image', methods=['POST'])
def process_table_image():
    """处理单张表格图片 - 根据配置自动选择服务"""
    try:
        data = request.get_json()
        print("&&&&&&&&&&&& 收到识别请求:", data)

        if not data:
            return jsonify({
                "success": False,
                "error": "请求体不能为空",
                "message": "请求体不能为空"
            }), 400

        # 提取参数
        image_path = data.get('image_path')
        output_path = data.get('output_path')
        sheet_name = data.get('sheet_name', '识别结果')
        bank_name = data.get('bank_name', '未知银行')
        file_name = data.get('file_name', 'table_1')
        request_table_type = data.get('table_type')  # 从请求中获取表格类型

        print(f"🔍 检查图片文件是否存在: {image_path}")

        # 检查图片文件是否存在
        image_full_path = Path(image_path)
        if not image_full_path.exists():
            # 尝试在static目录下查找
            static_path = Path("static") / image_path
            if static_path.exists():
                image_full_path = static_path
                print(f"✅ 找到图片文件: {static_path}")
            else:
                error_msg = f"图片文件不存在: {image_path}"
                print(f"❌ {error_msg}")
                return jsonify({
                    "success": False,
                    "error": error_msg,
                    "message": error_msg
                }), 404

        print(f"✅ 图片文件存在: {image_full_path}")

        # 使用已有的标识创建Excel存储文件夹
        folder_name = Path(image_path).stem.split('_')[0]
        excel_base_dir = Path("static/excel_data")
        excel_dir = excel_base_dir / folder_name
        excel_dir.mkdir(parents=True, exist_ok=True)

        # 重新构建输出路径
        excel_filename = f"single_{Path(image_path).stem}.xlsx"
        new_output_path = excel_dir / excel_filename
        final_output_path = str(new_output_path)

        print(f"Excel文件将保存到: {final_output_path}")

        # 检查最终路径是否存在
        final_excel_path = Path(final_output_path)
        if final_excel_path.exists():
            print("✅ Excel文件已存在，直接返回路径")

            # 将绝对路径转换为可访问的URL
            relative_path = final_output_path.replace('\\', '/')
            if 'static/' in relative_path:
                static_index = relative_path.find('static/')
                if static_index != -1:
                    excel_url = '/' + relative_path[static_index:]
                else:
                    excel_url = f"/static/excel_data/{folder_name}/{excel_filename}"
            else:
                excel_url = f"/static/excel_data/{folder_name}/{excel_filename}"

            return jsonify({
                "success": True,
                "from_cache": True,
                "message": "已加载现有表格数据",
                "excel_url": excel_url,
                "excel_path": f"{MAIN_ROOT}/{final_output_path}",
                "table_name": sheet_name,
                "table_type": request_table_type or "financial"
            })

        # 如果Excel不存在，进行LLM识别
        print("🔄 Excel文件不存在，开始LLM识别流程")

        # 根据请求的表格类型或当前配置选择处理器
        processor = get_appropriate_processor(request_table_type)
        current_table_type = "financial"

        if isinstance(processor, NonFinancialTableService):
            current_table_type = "non_financial"
            print("🔄 使用普通表格处理器")
        else:
            print("🔄 使用金融表格处理器")

        process_data = {
            'image_path': str(image_full_path),
            'output_path': final_output_path,
            'sheet_name': sheet_name,
            'bank_name': bank_name,
            'file_name': file_name
        }

        # 根据处理器类型调用不同的处理函数
        if current_table_type == 'non_financial':
            result = asyncio.run(_process_non_financial_table(process_data))
        else:
            result = asyncio.run(_process_single_image(process_data))

        print("LLM识别结果:", result)

        if result.get('success'):
            excel_url = f"/api/excel-data/{folder_name}/{excel_filename}"

            return jsonify({
                "success": True,
                "from_cache": False,
                "message": "表格识别完成",
                "excel_url": excel_url,
                "excel_path": f"{MAIN_ROOT}/{final_output_path}",
                "table_name": sheet_name,
                "table_type": current_table_type,
                "data": result['data']
            })
        else:
            error_msg = result.get('error', '识别失败')
            print(f"❌ LLM识别失败: {error_msg}")
            return jsonify({
                "success": False,
                "error": error_msg,
                "message": error_msg
            }), 500

    except Exception as e:
        logger.error(f"异步处理错误: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": f"异步处理错误: {str(e)}",
            "message": f"处理异常: {str(e)}"
        }), 500


@llm_bp.route('/llm/process-non-financial-table', methods=['POST'])
def process_non_financial_table():
    """处理普通表格"""
    try:
        data = request.get_json()
        print("收到普通表格识别请求:", data)

        if not data:
            return jsonify({
                "success": False,
                "error": "请求体不能为空",
                "message": "请求体不能为空"
            }), 400

        # 提取参数
        image_path = data.get('image_path')
        output_path = data.get('output_path')
        sheet_name = data.get('sheet_name', '普通表格识别结果')
        bank_name = data.get('bank_name', '未知机构')
        file_name = data.get('file_name', 'non_financial_table')

        print(f"处理普通表格 - 图片路径: {image_path}")

        # 检查图片文件是否存在
        image_full_path = Path(image_path)
        if not image_full_path.exists():
            # 尝试在static目录下查找
            static_path = Path("static") / image_path
            if static_path.exists():
                image_full_path = static_path
                print(f"✅ 找到图片文件: {static_path}")
            else:
                error_msg = f"图片文件不存在: {image_path}"
                print(f"❌ {error_msg}")
                return jsonify({
                    "success": False,
                    "error": error_msg,
                    "message": error_msg
                }), 404

        print(f"✅ 图片文件存在: {image_full_path}")

        # 使用已有的标识创建Excel存储文件夹
        folder_name = Path(image_path).stem.split('_')[0]
        excel_base_dir = Path("static/excel_data")
        excel_dir = excel_base_dir / folder_name
        excel_dir.mkdir(parents=True, exist_ok=True)

        # 重新构建输出路径
        excel_filename = f"non_financial_{Path(image_path).stem}.xlsx"
        new_output_path = excel_dir / excel_filename
        final_output_path = str(new_output_path)

        print(f"普通表格Excel文件将保存到: {final_output_path}")

        # 检查最终路径是否存在
        final_excel_path = Path(final_output_path)
        if final_excel_path.exists():
            print("✅ 普通表格Excel文件已存在，直接返回路径")

            # 将绝对路径转换为可访问的URL
            relative_path = final_output_path.replace('\\', '/')
            if 'static/' in relative_path:
                static_index = relative_path.find('static/')
                if static_index != -1:
                    excel_url = '/' + relative_path[static_index:]
                else:
                    excel_url = f"/static/excel_data/{folder_name}/{excel_filename}"
            else:
                excel_url = f"/static/excel_data/{folder_name}/{excel_filename}"

            return jsonify({
                "success": True,
                "from_cache": True,
                "message": "已加载现有普通表格数据",
                "excel_url": excel_url,
                "excel_path": f"{MAIN_ROOT}/{final_output_path}",
                "table_name": sheet_name,
                "table_type": "non_financial"
            })

        # 如果Excel不存在，进行普通表格识别
        print("🔄 普通表格Excel文件不存在，开始识别流程")

        # 调用异步处理函数
        process_data = {
            'image_path': str(image_full_path),
            'output_path': final_output_path,
            'sheet_name': sheet_name,
            'bank_name': bank_name,
            'file_name': file_name
        }

        result = asyncio.run(_process_non_financial_table(process_data))
        print("普通表格识别结果:", result)

        if result.get('success'):
            excel_url = f"/api/excel-data/{folder_name}/{excel_filename}"

            return jsonify({
                "success": True,
                "from_cache": False,
                "message": "普通表格识别完成",
                "excel_url": excel_url,
                "excel_path": f"{MAIN_ROOT}/{final_output_path}",
                "table_name": sheet_name,
                "table_type": "non_financial",
                "data": result['data']
            })
        else:
            error_msg = result.get('error', '识别失败')
            print(f"❌ 普通表格识别失败: {error_msg}")
            return jsonify({
                "success": False,
                "error": error_msg,
                "message": error_msg
            }), 500

    except Exception as e:
        logger.error(f"普通表格处理错误: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": f"普通表格处理错误: {str(e)}",
            "message": f"处理异常: {str(e)}"
        }), 500


# 在 llm_routes.py 中添加（如果还没有的话）
@llm_bp.route('/excel-data/<path:excel_path>')
def serve_excel_data(excel_path):
    """提供Excel文件数据访问"""
    try:
        # 构建完整的文件路径
        file_path = Path("static/excel_data") / excel_path

        if not file_path.exists():
            return jsonify({
                "success": False,
                "error": f"Excel文件不存在: {excel_path}"
            }), 404

        # 返回文件
        return send_from_directory('static/excel_data', excel_path)

    except Exception as e:
        logger.error(f"提供Excel文件失败: {str(e)}")
        return jsonify({
            "success": False,
            "error": f"文件访问失败: {str(e)}"
        }), 500




@llm_bp.route('/llm/get-excel-content', methods=['GET'])
def get_excel_content():
    """读取Excel文件内容并返回结构化数据"""
    try:
        excel_url = request.args.get('excel_url')
        print(f"📖 读取Excel内容: {excel_url}")

        if not excel_url:
            return jsonify({
                "success": False,
                "error": "缺少excel_url参数"
            }), 400

        # 只接受相对路径格式
        if excel_url.startswith('http'):
            return jsonify({
                "success": False,
                "error": f"无效的Excel URL格式，请使用相对路径: {excel_url}"
            }), 400

        file_path = None

        if excel_url.startswith('/api/excel-data/'):
            # 格式: /api/excel-data/{folder}/{filename}
            relative_path = excel_url.replace('/api/excel-data/', '')
            file_path = Path("static/excel_data") / relative_path
        elif excel_url.startswith('/static/excel_data/'):
            # 格式: /static/excel_data/{folder}/{filename}
            relative_path = excel_url.replace('/static/excel_data/', '')
            file_path = Path("static/excel_data") / relative_path
        else:
            return jsonify({
                "success": False,
                "error": f"不支持的Excel URL格式: {excel_url}"
            }), 400

        full_path = Path(file_path)
        if not full_path.is_absolute():
            full_path = Path.cwd() / file_path

        print(f"Excel文件完整路径: {full_path}")

        if not full_path.exists():
            return jsonify({
                "success": False,
                "error": f"Excel文件不存在: {file_path}"
            }), 404

        # 读取Excel文件内容...
        # 读取Excel文件
        import openpyxl
        workbook = openpyxl.load_workbook(full_path)

        # 获取所有工作表
        sheet_data = []

        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]

            # 转换为JSON数据
            headers = []
            data_rows = []

            # 查找数据开始行（跳过空行和标题行）
            data_start_row = 1
            max_col = worksheet.max_column

            # 读取表头（第一行有数据的行）
            for row in range(1, worksheet.max_row + 1):
                row_has_data = False
                for col in range(1, max_col + 1):
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value and str(cell_value).strip():
                        row_has_data = True
                        break

                if row_has_data:
                    # 这一行有数据，作为表头
                    for col in range(1, max_col + 1):
                        cell_value = worksheet.cell(row=row, column=col).value
                        headers.append(str(cell_value) if cell_value is not None else f"列{col}")
                    data_start_row = row + 1
                    break

            # 读取数据行
            for row in range(data_start_row, worksheet.max_row + 1):
                row_data = {}
                has_data = False

                for col, header in enumerate(headers, 1):
                    if col > max_col:
                        break
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value is not None and str(cell_value).strip():
                        has_data = True
                    row_data[header] = str(cell_value) if cell_value is not None else ""

                if has_data:  # 只添加有数据的行
                    data_rows.append(row_data)

            sheet_data.append({
                "sheetName": sheet_name,
                "headers": headers,
                "data": data_rows,
                "rowCount": len(data_rows),
                "colCount": len(headers)
            })

        # 确保返回标准结构
        result = {
            "success": True,
            "data": {
                "filePath": str(full_path),
                "sheets": sheet_data,
                "totalSheets": len(sheet_data)
            }
        }

        print(f"✅ 返回数据: 共{len(sheet_data)}个工作表")
        for sheet in sheet_data:
            print(f"  工作表 '{sheet['sheetName']}': {sheet['rowCount']}行 {sheet['colCount']}列")

        return jsonify(result)

    except Exception as e:
        logger.error(f"读取Excel内容失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            "success": False,
            "error": f"读取Excel内容失败: {str(e)}"
        }), 500


@llm_bp.route('/llm/batch-process', methods=['POST'])
def batch_process_images():
    """批量处理图片 - 金融表格"""
    try:
        data = request.get_json()
        data['table_type'] = 'financial'  # 明确指定表格类型
        return _handle_batch_process(data)
    except Exception as e:
        logger.error(f"金融表格批量处理错误: {str(e)}")
        return jsonify({
            "success": False,
            "error": f"金融表格批量处理错误: {str(e)}"
        }), 500




async def _batch_process_images(data):
    """批量处理图片的异步函数"""
    try:
        # 验证必要参数
        is_valid, error_msg = validate_required_params(
            data, ['image_paths', 'output_dir']
        )
        if not is_valid:
            return {
                "success": False,
                "error": error_msg
            }

        image_paths = data.get('image_paths', [])
        output_dir = data.get('output_dir')
        bank_name = data.get('bank_name', '未知银行')
        table_type = data.get('table_type', 'financial')  # 获取表格类型
        output_file = data.get('output_file', 'batch_processing_results.xlsx')


        print("table_typetable_type:", table_type)


        # 检查图片文件
        missing_images = []
        for img_path in image_paths:
            if not Path(img_path).exists():
                missing_images.append(img_path)

        if missing_images:
            return {
                "success": False,
                "error": f"以下图片文件不存在: {missing_images}"
            }

        processor = get_table_processor()

        if not processor.llm_client:
            return {
                "success": False,
                "error": "请先配置LLM客户端"
            }

        # 使用已有的标识创建Excel存储文件夹
        if image_paths:
            folder_name = Path(image_paths[0]).stem.split('_')[0]
        else:
            folder_name = "unknown_batch"

        excel_base_dir = Path("static/excel_data")
        excel_dir = excel_base_dir / folder_name
        excel_dir.mkdir(parents=True, exist_ok=True)

        # 根据表格类型构建不同的文件名
        if table_type == 'non_financial':
            excel_filename = f"non_financial_batch_{folder_name}.xlsx"
        else:
            excel_filename = f"financial_batch_{folder_name}.xlsx"

        new_output_file = excel_dir / excel_filename
        output_file = str(new_output_file)

        print(f"批量处理Excel文件将保存到: {output_file}")
        print(f"表格类型: {table_type}")

        # 批量处理 - 根据表格类型选择不同的处理方法
        results = []
        success_count = 0

        for i, image_path in enumerate(image_paths):
            try:
                sheet_name = f"表格_{i + 1}"
                file_name = Path(image_path).stem

                print(f"🔄 处理第 {i + 1}/{len(image_paths)} 张图片，类型: {table_type}")

                if table_type == 'non_financial':
                    # 调用普通表格处理方法
                    print(f"🔄 调用普通表格处理: {image_path}")
                    result = await _process_single_non_financial_table({
                        'image_path': image_path,
                        'output_path': output_file,
                        'sheet_name': sheet_name,
                        'bank_name': bank_name,
                        'file_name': file_name,
                        'table_index': i
                    })
                else:
                    # 调用金融表格处理方法
                    print(f"🔄 调用金融表格处理: {image_path}")
                    result = await processor.process_table_pipeline(
                        image_path=image_path,
                        out_file=output_file,
                        sheet_name=sheet_name,
                        bank_name=bank_name,
                        file_name=file_name
                    )

                if table_type == 'non_financial':
                    # 普通表格的结果处理
                    if result.get('success'):
                        results.append({
                            "image_path": image_path,
                            "status": "success",
                            "table_name": sheet_name,
                            "sheet_name": sheet_name,
                            "excel_url": result.get('excel_url')
                        })
                        success_count += 1
                        print(f"✅ 普通表格处理成功: {image_path}")
                    else:
                        results.append({
                            "image_path": image_path,
                            "status": "error",
                            "error": result.get('error', '处理失败')
                        })
                        print(f"❌ 普通表格处理失败: {image_path} - {result.get('error')}")
                else:
                    # 金融表格的结果处理
                    results.append({
                        "image_path": image_path,
                        "status": result.status,
                        "complexity": result.complexity,
                        "table_name": result.table_name,
                        "sheet_name": sheet_name
                    })

                    if result.status == "success":
                        success_count += 1
                        print(f"✅ 金融表格处理成功: {image_path}")
                    else:
                        print(f"❌ 金融表格处理失败: {image_path}")

            except Exception as e:
                logger.error(f"处理图片失败 {image_path}: {str(e)}")
                results.append({
                    "image_path": image_path,
                    "status": "error",
                    "error": str(e)
                })
                print(f"💥 处理异常: {image_path} - {str(e)}")

        logger.info(f"批量处理完成 - 类型: {table_type}, 总数: {len(image_paths)}, 成功: {success_count}")

        # 构建返回的Excel URL
        excel_url = convert_to_excel_url(output_file)

        return {
            "success": True,
            "data": {
                "total": len(image_paths),
                "success": success_count,
                "failed": len(image_paths) - success_count,
                "results": results,
                "output_file": output_file,
                "excel_url": excel_url,
                "table_type": table_type
            }
        }

    except Exception as e:
        logger.error(f"批量处理失败: {str(e)}")
        return {
            "success": False,
            "error": f"批量处理失败: {str(e)}"
        }


async def _process_single_non_financial_table(process_data):
    """处理单个普通表格"""
    try:
        image_path = process_data['image_path']
        output_path = process_data['output_path']
        sheet_name = process_data['sheet_name']
        bank_name = process_data['bank_name']
        file_name = process_data['file_name']
        table_index = process_data.get('table_index', 0)

        print(f"处理普通表格 #{table_index}: {image_path}")

        # 调用普通表格识别的核心逻辑
        result = await _process_non_financial_table(process_data)

        if result.get('success'):
            # 转换为统一的URL格式
            excel_url = convert_to_excel_url(output_path)
            result['excel_url'] = excel_url

        return result

    except Exception as e:
        logger.error(f"处理普通表格失败 {process_data['image_path']}: {str(e)}")
        return {
            "success": False,
            "error": f"处理失败: {str(e)}"
        }


def convert_to_excel_url(file_path):
    """将文件路径转换为Excel URL"""
    file_path = file_path.replace('\\', '/')

    if 'static/excel_data/' in file_path:
        # 提取相对路径部分
        parts = file_path.split('static/excel_data/')
        if len(parts) > 1:
            return f"/api/excel-data/{parts[1]}"

    # 回退方案
    file_name = Path(file_path).name
    folder_name = Path(file_path).parent.name
    return f"/api/excel-data/{folder_name}/{file_name}"



@llm_bp.route('/llm/batch-process', methods=['POST'], endpoint='batch_process_financial')
def batch_process_financial_images():
    """批量处理图片 - 金融表格"""
    try:
        data = request.get_json()
        # 明确设置为金融表格，不依赖前端传递的table_type
        data['table_type'] = 'financial'
        print(f"🔵 金融表格批量处理 - 强制设置表格类型: financial")
        return _handle_batch_process(data)
    except Exception as e:
        logger.error(f"金融表格批量处理错误: {str(e)}")
        return jsonify({
            "success": False,
            "error": f"金融表格批量处理错误: {str(e)}"
        }), 500


@llm_bp.route('/llm/batch-process-non-financial', methods=['POST'], endpoint='batch_process_non_financial')
def batch_process_non_financial_images():
    """批量处理图片 - 普通表格"""
    try:
        data = request.get_json()
        # 明确设置为普通表格，不依赖前端传递的table_type
        data['table_type'] = 'non_financial'
        print(f"🟢 普通表格批量处理 - 强制设置表格类型: non_financial")
        return _handle_batch_process(data)
    except Exception as e:
        logger.error(f"普通表格批量处理错误: {str(e)}")
        return jsonify({
            "success": False,
            "error": f"普通表格批量处理错误: {str(e)}"
        }), 500


def _handle_batch_process(data):
    """统一的批量处理入口"""
    try:
        if not data:
            return jsonify({
                "success": False,
                "error": "请求体不能为空"
            }), 400

        # 添加调试信息
        table_type = data.get('table_type', 'unknown')
        print(f"🔄 批量处理入口 - 表格类型: {table_type}")

        result = asyncio.run(_batch_process_images(data))
        return jsonify(result)
    except Exception as e:
        logger.error(f"批量处理错误: {str(e)}")
        return jsonify({
            "success": False,
            "error": f"批量处理错误: {str(e)}"
        }), 500

