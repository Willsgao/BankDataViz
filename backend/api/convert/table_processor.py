"""
表格处理模块 - 业务逻辑层
职责：封装表格处理业务逻辑，不包含API响应
"""

import os
import sqlite3
from flask import jsonify
import threading
import time
import uuid
from typing import Dict, List, Any, Optional, Tuple
from datetime import datetime

from backend.utils.constants import DATABASE_PATH, FILTERED_TABLES_DIR, EXCEL_DATA_DIR, DATABASE
from backend.models.unified_db import NewDatabaseManager
from backend.core.table_processor import TableReconstructor
from backend.core.table_processor import EnhancedFinancialTableAnalyzer

# from backend.excel_service.baidu_table_ocr_llm import TableOCRService
from backend.core.table_processor.ocr_gateway import TableOCRService
from backend.core.incremental_processor import incremental_processor
from pathlib import Path


# ========== 1. 导入表格处理管道 ==========
from backend.configs.config import tableconfig
PIPELINE_AVAILABLE = True


# PDF数据聚合器
class PDFDataAggregator:
    """
    PDF数据聚合器 - 在内存中聚合所有表格数据
    用于最后一次性写入Excel，避免中间文件
    """

    def __init__(self, pdf_folder, bank_name=""):
        """
        初始化PDF数据聚合器

        Args:
            pdf_folder: PDF文件夹名称
            bank_name: 银行名称（用于最终文件名）
        """
        self.pdf_folder = pdf_folder
        self.bank_name = bank_name
        self.tables_data = []  # 存储所有表格的二维数组数据
        self.table_names = []  # 存储每个表格的Sheet名称
        self.image_refs = []  # 存储每个表格对应的图片信息
        self.lock = threading.RLock()  # 线程安全锁
        self.created_at = datetime.now()
        self.last_updated = self.created_at


    # 在 PDFDataAggregator 类的 add_table 方法中修复
    def add_table(self, image_name, table_data, sheet_name=None, image_path=None, metadata=None):
        """
        添加一个表格数据到聚合器（增加metadata参数）
        """
        with self.lock:
            if not table_data:
                print(f"⚠️ 空表格数据，跳过: {image_name}")
                return False

            # 验证表格数据
            if not isinstance(table_data, list) or len(table_data) == 0:
                print(f"⚠️ 无效表格数据格式: {image_name}")
                return False

            # 生成Sheet名称
            if not sheet_name:
                sheet_name = self._generate_sheet_name(image_name, len(self.tables_data))

            # 清理Sheet名称 - 直接使用简单清理逻辑，避免导入问题
            cleaned_sheet_name = self._simple_clean_sheet_name(sheet_name)

            # 存储数据（现在包含metadata）
            self.tables_data.append(table_data)
            self.table_names.append(cleaned_sheet_name)
            self.image_refs.append({
                'image_name': image_name,
                'image_path': image_path,
                'added_at': datetime.now().isoformat(),
                'table_shape': f"{len(table_data)}行×{len(table_data[0]) if table_data else 0}列",
                'metadata': metadata or {}  # 新增metadata字段
            })

            self.last_updated = datetime.now()

            print(f"📊📊 聚合表格: {image_name} -> Sheet: '{cleaned_sheet_name}' "
                  f"({len(table_data)}行×{len(table_data[0]) if table_data else 0}列)")

            # 打印元数据信息
            if metadata:
                print(f"   元数据: 币种={metadata.get('default_currency', 'N/A')}, "
                      f"报告期={metadata.get('default_report_period', 'N/A')}, "
                      f"单位={metadata.get('default_unit', 'N/A')}")

            return True

    def _simple_clean_sheet_name(self, name, max_length=31):
        """
        简单的Sheet名称清理逻辑
        """
        if not name:
            return "Sheet1"

        import re

        # 替换非法字符
        illegal_chars = r'[\\/*?\[\]:]'
        cleaned = re.sub(illegal_chars, '_', str(name))

        # 移除开头和结尾的空格
        cleaned = cleaned.strip()

        # 如果以'开头，移除
        if cleaned.startswith("'"):
            cleaned = cleaned[1:]

        # 截断到最大长度
        if len(cleaned) > max_length:
            cleaned = cleaned[:max_length]

        # 确保不为空
        if not cleaned:
            cleaned = "Sheet1"

        # 检查重复并去重
        base_name = cleaned
        counter = 1
        while cleaned in self.table_names:
            suffix = f"_{counter}"
            if len(base_name) + len(suffix) > max_length:
                truncated_base = base_name[:max_length - len(suffix)]
                cleaned = f"{truncated_base}{suffix}"
            else:
                cleaned = f"{base_name}{suffix}"
            counter += 1

        return cleaned

    def _generate_sheet_name(self, image_name, table_index):
        """
        生成Sheet名称

        Args:
            image_name: 图片文件名
            table_index: 表格索引

        Returns:
            str: Sheet名称
        """
        from pathlib import Path

        # 移除扩展名
        stem = Path(image_name).stem

        # 尝试提取页码
        page_num = ""
        for i, char in enumerate(stem):
            if char.isdigit():
                # 找到连续数字作为页码
                j = i
                while j < len(stem) and stem[j].isdigit():
                    page_num += stem[j]
                    j += 1
                if len(page_num) >= 2:  # 至少2位数字才认为是页码
                    page_num = f"P{page_num.zfill(3)}"
                    break
                page_num = ""

        # 构建Sheet名称
        if page_num:
            return f"{page_num}_表{table_index + 1}"
        else:
            return f"表{table_index + 1}_{stem}"

    def get_statistics(self):
        """
        获取聚合器统计信息

        Returns:
            dict: 统计信息
        """
        with self.lock:
            total_tables = len(self.tables_data)
            total_rows = sum(len(table) for table in self.tables_data)
            total_cells = sum(len(table) * len(table[0]) if table and len(table[0]) > 0 else 0
                              for table in self.tables_data)

            return {
                'pdf_folder': self.pdf_folder,
                'bank_name': self.bank_name,
                'total_tables': total_tables,
                'total_rows': total_rows,
                'total_cells': total_cells,
                'table_names': self.table_names.copy(),
                'created_at': self.created_at.isoformat(),
                'last_updated': self.last_updated.isoformat(),
                'image_refs': self.image_refs.copy()
            }

    def save_to_excel(self, output_path, metadata_list=None):
        """
        将所有表格数据一次性写入Excel

        Args:
            output_path: 输出Excel文件路径
            metadata_list: 元数据列表（新增）
        """
        with self.lock:  # 修改这里：从 _lock 改为 lock
            if not self.tables_data:
                print(f"⚠️ 聚合器中没有表格数据: {self.pdf_folder}")
                return False

            print(f"🔄🔄 开始写入Excel: {self.pdf_folder}")
            print(f"  表格数量: {len(self.tables_data)}")
            print(f"  输出路径: {output_path}")
            print(f"  元数据数量: {len(metadata_list) if metadata_list else 0}")

            try:
                # 导入并调用TableReconstructor的保存方法
                # 尝试不同的导入路径
                try:
                    from backend.core.table_processor import TableReconstructor
                    reconstructor = TableReconstructor()
                except ImportError:
                    try:
                        from backend.api.convert.table_reconstructor import TableReconstructor
                        reconstructor = TableReconstructor()
                    except ImportError:
                        try:
                            from ..table_reconstructor import TableReconstructor
                            reconstructor = TableReconstructor()
                        except ImportError as e:
                            print(f"❌❌ 无法导入TableReconstructor: {e}")
                            return False

                # 传递元数据到保存方法
                success = reconstructor.step9_save_to_excel_optimized(
                    tables_data=self.tables_data,
                    output_file=output_path,
                    table_names=self.table_names,
                    metadata_list=metadata_list  # 新增参数
                )

                if success:
                    print(f"✅ Excel写入成功: {output_path}")
                    stats = self.get_statistics()
                    print(f"📊📊 统计: {stats['total_tables']}个表格, "
                          f"{stats['total_rows']}行, {stats['total_cells']}个单元格")
                else:
                    print(f"❌❌ Excel写入失败: {output_path}")

                return success

            except Exception as e:
                print(f"❌❌ 保存Excel异常: {e}")
                import traceback
                traceback.print_exc()
                return False

    def clear(self):
        """清空聚合器数据（释放内存）"""
        with self.lock:
            table_count = len(self.tables_data)
            self.tables_data.clear()
            self.table_names.clear()
            self.image_refs.clear()
            print(f"🗑️ 清空聚合器: {self.pdf_folder}, 释放 {table_count} 个表格")

    def __len__(self):
        """获取表格数量"""
        with self.lock:
            return len(self.tables_data)

    def __str__(self):
        """字符串表示"""
        stats = self.get_statistics()
        return (f"PDFDataAggregator(pdf_folder={self.pdf_folder}, "
                f"tables={stats['total_tables']}, "
                f"rows={stats['total_rows']})")


# PDF聚合器管理器
class PDFAggregatorManager:
    """
    PDF聚合器管理器 - 全局管理所有PDF的数据聚合器
    单例模式，确保全局只有一个管理器实例
    """

    _instance = None
    _lock = threading.RLock()

    def __new__(cls):
        with cls._lock:
            if cls._instance is None:
                cls._instance = super().__new__(cls)
                cls._instance._initialized = False
            return cls._instance

    def __init__(self):
        if not self._initialized:
            with self._lock:
                if not self._initialized:
                    self._aggregators = {}  # pdf_folder -> PDFDataAggregator
                    self._aggregator_locks = {}  # 每个聚合器的独立锁
                    self._processing_status = {}  # PDF处理状态跟踪
                    self._initialized = True
                    self._image_processing_tracking = {}  # pdf_folder -> set(image_names)
                    print("✅ PDF聚合器管理器初始化完成")

    def __contains__(self, pdf_folder):
        """检查PDF是否在管理中"""
        with self._lock:
            return pdf_folder in self._aggregators

    def __len__(self):
        """获取管理的PDF数量"""
        with self._lock:
            return len(self._aggregators)

    def _generate_expected_sheet_name(self, image_name, existing_sheets):
        """
        生成预期的Sheet名称（与PDFDataAggregator逻辑一致）
        """
        from pathlib import Path

        stem = Path(image_name).stem

        # 提取页码逻辑（与PDFDataAggregator保持一致）
        page_num = ""
        for i, char in enumerate(stem):
            if char.isdigit():
                j = i
                while j < len(stem) and stem[j].isdigit():
                    page_num += stem[j]
                    j += 1
                if len(page_num) >= 2:
                    page_num = f"P{page_num.zfill(3)}"
                    break
                page_num = ""

        # 基础名称（与PDFDataAggregator逻辑一致）
        if page_num:
            base_name = f"{page_num}_表"
        else:
            base_name = f"表_{stem}"

        # 检查是否已存在，如果存在则添加后缀
        counter = 1
        expected_name = f"{base_name}1"  # 默认第一个表
        while expected_name in existing_sheets:
            counter += 1
            expected_name = f"{base_name}{counter}"

        return expected_name

    def get_aggregator(self, pdf_folder, bank_name=""):
        """
        获取或创建PDF数据聚合器

        Args:
            pdf_folder: PDF文件夹名称
            bank_name: 银行名称

        Returns:
            PDFDataAggregator: 数据聚合器实例
        """
        with self._lock:
            if pdf_folder not in self._aggregators:
                self._aggregators[pdf_folder] = PDFDataAggregator(pdf_folder, bank_name)

                # 初始化处理状态
                self._processing_status[pdf_folder] = {
                    'status': 'processing',  # processing, merging, completed, failed
                    'total_images': 0,
                    'processed_images': 0,
                    'failed_images': 0,
                    'start_time': datetime.now().isoformat(),
                    'last_update': datetime.now().isoformat(),
                    'expected_images': 0  # 可选：预期的总图片数
                }

                print(f"📁 创建PDF聚合器: {pdf_folder} (银行: {bank_name})")

            return self._aggregators[pdf_folder]

    def register_processing_job(self, pdf_folder, total_images, bank_name=""):
        """
        注册PDF处理任务

        Args:
            pdf_folder: PDF文件夹名称
            total_images: 总图片数
            bank_name: 银行名称
        """
        with self._lock:
            # 确保聚合器存在
            self.get_aggregator(pdf_folder, bank_name)

            # 更新状态
            self._processing_status[pdf_folder].update({
                'total_images': total_images,
                'expected_images': total_images,
                'status': 'processing',
                'start_time': datetime.now().isoformat(),
                'last_update': datetime.now().isoformat()
            })

            print(f"📝 注册PDF处理任务: {pdf_folder}, 总图片数: {total_images}")

    def update_processing_status(self, pdf_folder,
                                 processed_images=None,
                                 failed_images=None,
                                 status=None):
        """
        更新PDF处理状态

        Args:
            pdf_folder: PDF文件夹名称
            processed_images: 已处理的图片数（可选）
            failed_images: 失败的图片数（可选）
            status: 状态更新（可选）
        """
        with self._lock:
            if pdf_folder not in self._processing_status:
                return False

            if processed_images is not None:
                self._processing_status[pdf_folder]['processed_images'] = processed_images

            if failed_images is not None:
                self._processing_status[pdf_folder]['failed_images'] = failed_images

            if status is not None:
                self._processing_status[pdf_folder]['status'] = status

            self._processing_status[pdf_folder]['last_update'] = datetime.now().isoformat()
            return True

    def get_processing_status(self, pdf_folder):
        """
        获取PDF处理状态

        Args:
            pdf_folder: PDF文件夹名称

        Returns:
            dict: 处理状态信息
        """
        with self._lock:
            if pdf_folder not in self._processing_status:
                return None

            status = self._processing_status[pdf_folder].copy()

            # 添加聚合器信息
            if pdf_folder in self._aggregators:
                aggregator = self._aggregators[pdf_folder]
                status['aggregator_stats'] = aggregator.get_statistics()
                status['tables_count'] = len(aggregator)

            return status

    def get_all_status(self):
        """
        获取所有PDF的处理状态

        Returns:
            dict: 所有PDF的状态
        """
        with self._lock:
            result = {}
            for pdf_folder in self._processing_status:
                result[pdf_folder] = self.get_processing_status(pdf_folder)
            return result

    def finalize_pdf_old(self, pdf_folder, output_dir=None, force=False, metadata_list=None):
        """
        最终化PDF处理，生成Excel文件（支持增量更新）
        """
        with self._lock:
            if pdf_folder not in self._aggregators:
                return False, None, f"PDF聚合器不存在: {pdf_folder}"

            aggregator = self._aggregators[pdf_folder]
            status = self._processing_status.get(pdf_folder, {})

            # 检查是否可以合并
            if not force:
                total_images = status.get('total_images', 0)
                processed_images = status.get('processed_images', 0)
                if total_images > 0 and processed_images < total_images:
                    return False, None, f"图片未全部完成 ({processed_images}/{total_images})"

            # 生成输出路径
            try:
                if output_dir is None:
                    output_dir = Path(tableconfig.output_dir) / pdf_folder
                else:
                    output_dir = Path(output_dir) / pdf_folder

                output_dir.mkdir(parents=True, exist_ok=True)

                # 检查是否已存在Excel文件（增量更新）
                existing_excel_path = None
                existing_files = list(output_dir.glob("*.xlsx"))

                # 如果有现有文件且不是强制模式，使用增量更新
                if existing_files and not force:
                    existing_excel_path = existing_files[0]  # 取第一个文件
                    print(f"🔍🔍 发现现有Excel文件: {existing_excel_path}")
                    print(f"🔄🔄 进入增量更新模式，将新表格追加到现有文件")

                if existing_excel_path and existing_excel_path.exists():
                    # 增量更新模式 - 使用现有文件名
                    filename = existing_excel_path.name
                    output_path = existing_excel_path
                    print(f"✅ 使用增量更新模式，文件: {filename}")
                else:
                    # 全新创建模式
                    if aggregator.bank_name:
                        filename = f"{aggregator.bank_name}_{pdf_folder}.xlsx"
                    else:
                        filename = f"{pdf_folder}_合并.xlsx"
                    output_path = output_dir / filename
                    print(f"🆕 使用全新创建模式，文件: {filename}")

                # 更新状态为合并中
                self.update_processing_status(pdf_folder, status='merging')

                print(f"🔄🔄🔄🔄 开始最终化PDF: {pdf_folder}")
                print(f"  输出文件: {output_path}")
                print(f"  新表格数量: {len(aggregator)}")
                print(f"  增量更新模式: {existing_excel_path is not None}")

                # 保存Excel（TableReconstructor需要支持增量保存）
                success = aggregator.save_to_excel(str(output_path), metadata_list)

                if success:
                    # 更新状态为完成
                    self.update_processing_status(pdf_folder, status='completed')
                    status['final_excel'] = str(output_path)
                    status['finalized_at'] = datetime.now().isoformat()
                    status['incremental_update'] = existing_excel_path is not None

                    if existing_excel_path:
                        print(f"✅ PDF增量更新完成: {pdf_folder} -> {output_path}")
                    else:
                        print(f"✅ PDF全新创建完成: {pdf_folder} -> {output_path}")

                    return True, str(output_path), None
                else:
                    self.update_processing_status(pdf_folder, status='failed')
                    return False, None, "Excel保存失败"

            except Exception as e:
                import traceback
                error_msg = f"最终化失败: {str(e)}"
                print(f"❌❌❌❌ {error_msg}")
                traceback.print_exc()
                self.update_processing_status(pdf_folder, status='failed')
                return False, None, error_msg

    def finalize_pdf(self, pdf_folder, output_dir=None, force=False, metadata_list=None):
        """
        第一步修复：简化逻辑，确保文件被保存
        移除所有可能导致提前返回的检查
        """
        with self._lock:
            print(f"\n{'=' * 60}")
            print(f"🔍 第一步修复：finalize_pdf 简化版")
            print(f"{'=' * 60}")

            if pdf_folder not in self._aggregators:
                error_msg = f"PDF聚合器不存在: {pdf_folder}"
                print(f"❌ {error_msg}")
                return False, None, error_msg

            aggregator = self._aggregators[pdf_folder]

            print(f"📊 聚合器表格数量: {len(aggregator)}")

            try:
                from pathlib import Path
                from datetime import datetime

                # 使用EXCEL_DATA_DIR
                if output_dir is None:
                    output_dir = Path(EXCEL_DATA_DIR) / pdf_folder
                    print(f"📁 使用EXCEL_DATA_DIR: {EXCEL_DATA_DIR}")
                else:
                    output_dir = Path(output_dir) / pdf_folder

                output_dir.mkdir(parents=True, exist_ok=True)

                print(f"📁 输出目录: {output_dir}")
                print(f"📁 绝对路径: {output_dir.absolute()}")
                print(f"📁 目录存在: {output_dir.exists()}")

                # 创建文件名
                if aggregator.bank_name:
                    filename = f"{aggregator.bank_name}_{pdf_folder}.xlsx"
                else:
                    filename = f"{pdf_folder}_合并.xlsx"

                output_path = output_dir / filename
                print(f"🆕 保存到文件: {filename}")
                print(f"📁 完整路径: {output_path.absolute()}")

                # 第一步：直接保存，不进行任何检查
                print(f"\n🔄 第一步：直接调用 save_to_excel")

                # 保存Excel
                success = aggregator.save_to_excel(str(output_path), metadata_list)

                print(f"📊 save_to_excel 返回结果: {success}")

                # 检查文件是否被创建
                if output_path.exists():
                    file_size = output_path.stat().st_size
                    print(f"✅ 文件已创建: {output_path}")
                    print(f"📁 文件大小: {file_size} 字节")

                    if file_size > 0:
                        print(f"✅ 文件保存成功: {output_path} ({file_size} 字节)")
                        return True, str(output_path), None
                    else:
                        print(f"⚠️ 文件大小为0，可能保存失败")
                        return False, None, "文件保存失败（文件大小为0）"
                else:
                    print(f"❌ 文件不存在: {output_path}")
                    return False, None, "文件保存失败（文件不存在）"

            except Exception as e:
                import traceback
                error_msg = f"保存失败: {str(e)}"
                print(f"❌ {error_msg}")
                traceback.print_exc()
                return False, None, error_msg


    def cleanup(self, pdf_folder=None):
        """
        清理聚合器

        Args:
            pdf_folder: 要清理的PDF文件夹名称，如果为None则清理所有

        Returns:
            int: 清理的聚合器数量
        """
        with self._lock:
            if pdf_folder:
                # 清理单个PDF
                if pdf_folder in self._aggregators:
                    self._aggregators[pdf_folder].clear()
                    del self._aggregators[pdf_folder]

                if pdf_folder in self._processing_status:
                    del self._processing_status[pdf_folder]

                print(f"🧹 清理PDF聚合器: {pdf_folder}")
                return 1
            else:
                # 清理所有
                count = len(self._aggregators)
                for agg in self._aggregators.values():
                    agg.clear()

                self._aggregators.clear()
                self._processing_status.clear()

                print(f"🧹 清理所有PDF聚合器: {count} 个")
                return count

    def get_active_pdfs(self):
        """
        获取正在处理的PDF列表

        Returns:
            list: PDF文件夹名称列表
        """
        with self._lock:
            return list(self._processing_status.keys())

    def get_existing_sheets(self, pdf_folder, output_dir=None):
        """
        获取已存在的Sheet名称

        Args:
            pdf_folder: PDF文件夹名称
            output_dir: 输出目录

        Returns:
            set: 已存在的Sheet名称集合
        """
        try:
            if output_dir is None:
                output_dir = Path(tableconfig.output_dir) / pdf_folder
            else:
                output_dir = Path(output_dir) / pdf_folder

            if not output_dir.exists():
                return set()

            existing_sheets = set()

            # 查找所有Excel文件
            for excel_file in output_dir.glob("*.xlsx"):
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(excel_file, data_only=True)
                    existing_sheets.update(wb.sheetnames)
                    wb.close()
                    print(f"🔍🔍 从文件 {excel_file.name} 读取到 {len(wb.sheetnames)} 个Sheet")
                except Exception as e:
                    print(f"⚠️ 读取Excel文件失败 {excel_file}: {e}")
                    continue

            print(f"📊📊 总共发现 {len(existing_sheets)} 个已存在的Sheet")
            return existing_sheets

        except Exception as e:
            print(f"❌❌ 获取已存在Sheet失败: {e}")
            return set()

    def filter_processed_images(self, pdf_folder, image_names, output_dir=None):
        """
        过滤已处理的图片 - 使用极简增量处理器（完全替换版）

        Args:
            pdf_folder: PDF文件夹名称
            image_names: 图片名称列表
            output_dir: 输出目录（兼容参数，实际不使用）

        Returns:
            tuple: (需要处理的图片列表, 跳过的图片列表, 已处理的图片集合)
        """
        print("=" * 60)
        print(f"🔍 增量处理检查: {pdf_folder}")
        print("=" * 60)

        # 1. 使用增量处理器过滤图片
        images_to_process = incremental_processor.filter_processed_images(pdf_folder, image_names)

        # 2. 计算跳过的图片
        skipped_images = [img for img in image_names if img not in images_to_process]

        # 3. 获取处理统计
        stats = incremental_processor.get_processing_stats(pdf_folder, image_names)

        # 4. 输出详细结果
        print("\n📊 增量处理统计:")
        print(f"  ├─ 总图片数: {stats['total_images']}")
        print(f"  ├─ 已处理: {stats['processed_images']}")
        print(f"  ├─ 未处理: {stats['unprocessed_images']}")
        print(f"  └─ 进度: {stats['progress_percentage']:.1f}%")

        if stats['is_completed']:
            print(f"  🎯 处理完成: 所有图片都已处理")

        if skipped_images:
            print(f"\n⏭️ 跳过的图片 ({len(skipped_images)}张):")
            for i, img in enumerate(skipped_images[:5]):  # 只显示前5个
                print(f"  {i + 1}. {img}")
            if len(skipped_images) > 5:
                print(f"  ... 等 {len(skipped_images) - 5} 张")

        if images_to_process:
            print(f"\n🆕 需要处理的图片 ({len(images_to_process)}张):")
            for i, img in enumerate(images_to_process[:5]):  # 只显示前5个
                print(f"  {i + 1}. {img}")
            if len(images_to_process) > 5:
                print(f"  ... 等 {len(images_to_process) - 5} 张")

        print("=" * 60)

        # 返回格式保持兼容
        processed_set = set(incremental_processor.records.get(pdf_folder, []))
        return images_to_process, skipped_images, processed_set

    # 修改 mark_image_completed 方法，确保记录
    def mark_image_completed(self, pdf_folder, image_name, table_count=1):
        """
        标记图片处理完成 - 同时记录到增量处理器
        """
        with self._lock:
            if pdf_folder not in self._processing_status:
                return False, 0

            status = self._processing_status[pdf_folder]
            status['processed_images'] = status.get('processed_images', 0) + 1

            # ✅ 关键：记录到增量处理器
            print(f"📝 记录处理完成: {pdf_folder}/{image_name}")
            incremental_processor.mark_images_processed(pdf_folder, [image_name])

            # 原有逻辑...
            if 'completed_images' not in status:
                status['completed_images'] = []
            status['completed_images'].append({
                'name': image_name,
                'table_count': table_count,
                'completed_at': datetime.now().isoformat()
            })

            # 进度计算...
            total_images = status.get('total_images', 0)
            processed = status['processed_images']

            if total_images > 0:
                progress = (processed / total_images) * 100
            else:
                progress = 0

            all_completed = (total_images > 0 and processed >= total_images)

            if all_completed:
                status['status'] = 'ready_to_merge'
                status['completion_time'] = datetime.now().isoformat()
                print(f"🎉 所有图片处理完成: {pdf_folder}")

            return all_completed, progress

    def is_image_being_processed(self, pdf_folder: str, image_name: str) -> bool:
        """
        检查图片是否正在处理中（线程安全）

        Args:
            pdf_folder: PDF文件夹名称
            image_name: 图片文件名

        Returns:
            bool: 是否正在处理中
        """
        with self._lock:
            if pdf_folder not in self._image_processing_tracking:
                self._image_processing_tracking[pdf_folder] = set()
            return image_name in self._image_processing_tracking[pdf_folder]

    def mark_image_processing(self, pdf_folder: str, image_name: str, processing: bool = True):
        """
        标记图片处理状态

        Args:
            pdf_folder: PDF文件夹名称
            image_name: 图片文件名
            processing: 是否正在处理
        """
        with self._lock:
            if pdf_folder not in self._image_processing_tracking:
                self._image_processing_tracking[pdf_folder] = set()

            if processing:
                self._image_processing_tracking[pdf_folder].add(image_name)
                print(f"🔒🔒 标记图片处理中: {pdf_folder}/{image_name}")
            else:
                self._image_processing_tracking[pdf_folder].discard(image_name)
                print(f"🔓🔓 标记图片处理完成: {pdf_folder}/{image_name}")

    def get_processing_images(self, pdf_folder: str) -> set:
        """
        获取正在处理的图片集合

        Args:
            pdf_folder: PDF文件夹名称

        Returns:
            set: 正在处理的图片名称集合
        """
        with self._lock:
            return self._image_processing_tracking.get(pdf_folder, set()).copy()

    def cleanup_processing_tracking(self, pdf_folder: str = None):
        """
        清理处理状态跟踪

        Args:
            pdf_folder: 要清理的PDF文件夹名称，如果为None则清理所有
        """
        with self._lock:
            if pdf_folder:
                if pdf_folder in self._image_processing_tracking:
                    tracking_count = len(self._image_processing_tracking[pdf_folder])
                    del self._image_processing_tracking[pdf_folder]
                    print(f"🧹🧹 清理处理状态跟踪: {pdf_folder}, 释放 {tracking_count} 个跟踪项")
            else:
                total_count = sum(len(images) for images in self._image_processing_tracking.values())
                self._image_processing_tracking.clear()
                print(f"🧹🧹 清理所有处理状态跟踪: 共 {total_count} 个跟踪项")



# 创建全局管理器实例
pdf_aggregator_manager = PDFAggregatorManager()


class TableProcessingService:
    """表格处理服务类 - 纯业务逻辑"""

    def __init__(self):
        """初始化服务"""
        self.output_base_dir = self._get_output_dir()
        self._reconstructor = TableReconstructor()

        self._memory_table_storage = {}  # pdf_folder -> table_data_list
        self._memory_storage_lock = threading.RLock()  # 线程安全锁
        print("[TableProcessingService] 初始化服务")

    def _get_output_dir(self) -> Path:
        """获取输出目录"""
        try:
            output_dir = Path(tableconfig.output_dir)
            output_dir.mkdir(parents=True, exist_ok=True)
            return output_dir
        except:
            # 回退到默认目录
            default_dir = Path("data/backend/outputs")
            default_dir.mkdir(parents=True, exist_ok=True)
            return default_dir

    def _get_pdf_uuid_from_database(self, pdf_folder: str) -> str:
        """
        从数据库files表获取PDF的UUID（filename字段去掉扩展名）

        Args:
            pdf_folder: PDF文件夹名称（对应raw_filename字段）

        Returns:
            str: PDF的UUID（filename字段去掉扩展名）
        """
        try:
            import sqlite3
            from backend.utils.constants import DATABASE_PATH

            conn = sqlite3.connect(DATABASE_PATH)
            cursor = conn.cursor()

            # 查询files表，根据raw_filename匹配pdf_folder
            cursor.execute("""
                SELECT filename FROM files 
                WHERE raw_filename LIKE ? AND deleted = 0
                ORDER BY created_at DESC 
                LIMIT 1
            """, (f'%{pdf_folder}%',))

            result = cursor.fetchone()
            conn.close()

            if result and result[0]:
                # filename字段格式：54fa94e5-531c-ac39-e.pdf，去掉扩展名
                filename = result[0]
                pdf_uuid = Path(filename).stem  # 去掉扩展名
                print(f"📊📊 从数据库获取到PDF UUID: {pdf_uuid} (原始filename: {filename})")
                return pdf_uuid
            else:
                print(f"⚠️⚠️ 数据库中未找到PDF记录: {pdf_folder}")
                # 回退到使用文件夹名称
                return pdf_folder

        except Exception as e:
            print(f"❌❌ 查询数据库失败: {e}")
            # 回退到使用文件夹名称
            return pdf_folder

    def _get_uuid_based_dir(self, pdf_uuid: str) -> Path:
        """获取基于UUID的输出目录"""
        try:
            output_dir = Path(tableconfig.output_dir) / pdf_uuid
            output_dir.mkdir(parents=True, exist_ok=True)
            return output_dir
        except:
            # 回退到默认目录
            default_dir = Path("data/backend/outputs") / pdf_uuid
            default_dir.mkdir(parents=True, exist_ok=True)
            return default_dir

    def _run_ocr_llm_memory_pipeline(self, image_path: str, bank_name: str = ""):
        """
        纯算法流水线：OCR → LLM → 内存表格数据
        返回 (tables_data, sheet_names, metadata_list)
        """

        try:
            # 1. OCR识别
            print(f"🔍🔍 OCR识别: {Path(image_path).name}")
            ocr_service = TableOCRService()
            ocr_result = ocr_service.recognize_table(image_path)

            if not ocr_result.get('tables_result'):
                print(f"⚠️ OCR未识别到表格: {Path(image_path).name}")
                return [], [], []

            # 2. LLM分析
            print(f"🤖🤖 LLM分析: {Path(image_path).name}")
            analyzer = EnhancedFinancialTableAnalyzer()
            llm_result = analyzer.analyze_image(image_path, ocr_result)

            print("TTTTTTTTTTTTTllm_resultTTTTTTTTTTTTTT")
            print(llm_result)


            if not llm_result.get('tables_structure', {}).get('tables'):
                print(f"⚠️ LLM未分析出表格结构: {Path(image_path).name}")
                return [], [], []

            # 3. 表格重构到内存
            print(f"🔄🔄 表格重构: {Path(image_path).name}")
            tables_data, sheet_names, metadata_list = self._reconstructor.process_all_tables_to_memory(
                ocr_result=ocr_result,
                llm_result=llm_result,
                image_path=image_path,
                bank_name=bank_name
            )

            print(f"✅ 内存流水线完成: {Path(image_path).name} -> {len(tables_data)}个表格")
            return tables_data, sheet_names, metadata_list

        except Exception as e:
            print(f"❌❌ 内存流水线失败 {Path(image_path).name}: {e}")
            import traceback
            traceback.print_exc()
            return [], [], []

    def get_table_data(self, pdf_folder: str, image_name: str = None, table_index: int = None):
        """
        获取处理后的表格数据（内存模式专用）
        """
        with self._memory_storage_lock:
            if pdf_folder not in self._memory_table_storage:
                return {
                    "success": False,
                    "error": f"未找到文件夹 {pdf_folder} 的表格数据",
                    "table_data": []
                }

            tables_data = self._memory_table_storage[pdf_folder]

            # 筛选数据
            if image_name:
                tables_data = [table for table in tables_data if table["image_name"] == image_name]

            if table_index is not None:
                tables_data = [table for table in tables_data if table["table_index"] == table_index]

            return {
                "success": True,
                "pdf_folder": pdf_folder,
                "table_data": tables_data,
                "total_tables": len(tables_data),
                "filtered_by_image": image_name is not None,
                "filtered_by_index": table_index is not None
            }

    def clear_table_data(self, pdf_folder: str = None):
        """
        清理内存中的表格数据
        """
        with self._memory_storage_lock:
            if pdf_folder is None:
                # 清理所有数据
                cleared_count = sum(len(data) for data in self._memory_table_storage.values())
                self._memory_table_storage.clear()
                return {
                    "success": True,
                    "cleared_folders": list(self._memory_table_storage.keys()),
                    "cleared_tables": cleared_count,
                    "message": f"清理所有内存表格数据，共{cleared_count}个表格"
                }
            else:
                # 清理指定文件夹数据
                if pdf_folder in self._memory_table_storage:
                    cleared_count = len(self._memory_table_storage[pdf_folder])
                    del self._memory_table_storage[pdf_folder]
                    return {
                        "success": True,
                        "cleared_folders": [pdf_folder],
                        "cleared_tables": cleared_count,
                        "message": f"清理文件夹 {pdf_folder} 的内存表格数据，共{cleared_count}个表格"
                    }
                else:
                    return {
                        "success": False,
                        "error": f"文件夹 {pdf_folder} 没有内存表格数据",
                        "cleared_tables": 0
                    }

    def _parse_processing_result(self, raw_result: Dict[str, Any],
                                 valid_images: List[str]) -> Dict[str, Any]:
        """解析处理结果 - 原有方法保持不变"""
        if not raw_result.get('success', False):
            return {
                "success": False,
                "error": raw_result.get('error', '未知错误'),
                "total_images": len(valid_images)
            }

        # 提取成功的结果和Excel文件
        results = []
        excel_files = []
        success_count = 0

        for res in raw_result.get('results', []):
            if res.get('success'):
                success_count += 1
                results.append({
                    "image_path": Path(res.get('image_path', '')).name,
                    "success": True,
                    "output_file": res.get('output_file', ''),
                    "processing_time": res.get('processing_time', 0)
                })
                if res.get('output_file'):
                    excel_files.append(res['output_file'])

        # 统计信息
        stats = raw_result.get('stats', {})

        return {
            "success": True,
            "total_images": len(valid_images),
            "success_count": success_count,
            "failed_count": len(valid_images) - success_count,
            "processing_time": stats.get('processing_time', 0),
            "excel_files": excel_files,
            "raw_results": results,
            "raw_stats": stats
        }

    def get_excel_files(self, pdf_folder: str) -> List[Dict[str, Any]]:
        """获取指定文件夹的Excel文件列表 - 自动从数据库获取UUID"""
        # 从数据库获取UUID
        pdf_uuid = self._get_pdf_uuid_from_database(pdf_folder)

        output_dir = self._get_uuid_based_dir(pdf_uuid) / pdf_folder
        excel_files = []

        if output_dir.exists():
            for excel_file in output_dir.glob("*.xlsx"):
                excel_files.append({
                    "filename": excel_file.name,
                    "path": str(excel_file),
                    "size": excel_file.stat().st_size,
                    "modified_time": datetime.fromtimestamp(
                        excel_file.stat().st_mtime
                    ).isoformat(),
                    "relative_download_path": f"{pdf_folder}/{excel_file.name}",
                    "pdf_uuid": pdf_uuid  # 添加UUID信息
                })

        return excel_files

    def get_pdf_info(self, pdf_folder: str):
        """获取PDF信息 - 自动从数据库获取UUID"""
        pdf_uuid = self._get_pdf_uuid_from_database(pdf_folder)

        return {
            "pdf_folder": pdf_folder,
            "pdf_uuid": pdf_uuid,
            "output_dir": str(self._get_uuid_based_dir(pdf_uuid))
        }

    def create_memory_processing_task(self, pdf_folder: str, valid_images: List[str],
                                      bank_name: str = "") -> Dict[str, Any]:
        """
        创建内存表格处理任务（不保存Excel文件）

        Args:
            pdf_folder: PDF文件夹名称
            valid_images: 有效图片路径列表
            bank_name: 银行名称

        Returns:
            dict: 任务信息，包含内存处理标记
        """
        job_id = str(uuid.uuid4())

        # 自动从数据库获取UUID
        try:
            pdf_uuid = self._get_pdf_uuid_from_database(pdf_folder)
            print(f"📊📊📊📊 自动获取PDF UUID: {pdf_folder} -> {pdf_uuid}")
        except Exception as e:
            print(f"⚠️⚠️ 自动获取UUID失败: {e}, 使用文件夹名称作为UUID")
            pdf_uuid = pdf_folder

        return {
            "job_id": job_id,
            "pdf_folder": pdf_folder,
            "pdf_uuid": pdf_uuid,
            "valid_images": valid_images,
            "bank_name": bank_name,
            "total_images": len(valid_images),
            "created_at": datetime.now().isoformat(),
            "processing_mode": "memory"  # 新增：标记为内存处理模式
        }

    def execute_memory_processing(self, task_info: Dict[str, Any]) -> Dict[str, Any]:
        """
        执行内存表格处理（不保存Excel文件）

        Args:
            task_info: 任务信息，包含内存处理标记

        Returns:
            dict: 处理结果，不包含磁盘文件信息
        """

        # 从任务信息中获取参数
        pdf_folder = task_info["pdf_folder"]
        valid_images = task_info["valid_images"]
        bank_name = task_info.get("bank_name", "")

        print(f"🚀🚀🚀🚀 执行内存表格处理任务 - 文件夹: {pdf_folder}")

        # 调用内存处理函数
        result = self.process_images_to_memory(
            pdf_folder=pdf_folder,
            valid_images=valid_images,
            bank_name=bank_name
        )

        # 合并任务信息和处理结果
        final_result = {
            **task_info,
            **result,
            "completed_at": datetime.now().isoformat(),
            "processing_mode": "memory"  # 确保包含处理模式标记
        }

        return final_result

    def process_images_to_memory(self, pdf_folder: str, valid_images: List[str],
                                 bank_name: str = "") -> Dict[str, Any]:
        """处理图片表格到内存 - 统一使用内存处理模式"""
        print(f"📊📊📊📊 开始处理表格到内存 - 文件夹: {pdf_folder}, 图片数: {len(valid_images)}")

        if not PIPELINE_AVAILABLE:
            return {
                "success": False,
                "error": "表格处理管道不可用",
                "total_images": len(valid_images)
            }

        try:
            # ✅ 统一使用内存处理模式，不调用批量处理
            results = []
            all_tables_data = []
            success_count = 0
            failed_count = 0
            total_processing_time = 0

            # 逐张处理图片（与 process_table_images_real 保持一致）
            for i, image_path in enumerate(valid_images):
                image_name = Path(image_path).name
                print(f"🖼🖼 处理图片 {i + 1}/{len(valid_images)}: {image_name}")

                try:
                    start_time = time.time()

                    # ✅ 使用统一的内存处理流水线
                    tables_data, sheet_names, metadata_list = self._run_ocr_llm_memory_pipeline(
                        image_path=image_path,
                        bank_name=bank_name
                    )

                    processing_time = time.time() - start_time
                    total_processing_time += processing_time

                    if tables_data:
                        success_count += 1
                        # 收集表格数据
                        for table_idx, (table_data, sheet_name) in enumerate(zip(tables_data, sheet_names)):
                            table_info = {
                                "image_name": image_name,
                                "sheet_name": sheet_name,
                                "table_data": table_data,
                                "table_index": table_idx,
                                "metadata": metadata_list[table_idx] if table_idx < len(metadata_list) else {},
                                "processing_time": processing_time
                            }
                            all_tables_data.append(table_info)

                        results.append({
                            "image_path": image_name,
                            "success": True,
                            "tables_extracted": len(tables_data),
                            "processing_time": processing_time,
                            "table_data_available": True
                        })
                        print(f"✅ 图片处理成功: {image_name}, 提取 {len(tables_data)} 个表格")
                    else:
                        failed_count += 1
                        results.append({
                            "image_path": image_name,
                            "success": False,
                            "error": "未提取到表格数据",
                            "processing_time": processing_time
                        })
                        print(f"⚠️ 图片处理未提取到表格: {image_name}")

                except Exception as img_error:
                    failed_count += 1
                    processing_time = time.time() - start_time if 'start_time' in locals() else 0
                    results.append({
                        "image_path": image_name,
                        "success": False,
                        "error": str(img_error),
                        "processing_time": processing_time
                    })
                    print(f"❌❌ 图片处理失败 {image_name}: {img_error}")

            # 将表格数据存储到内存存储中
            with self._memory_storage_lock:
                if pdf_folder not in self._memory_table_storage:
                    self._memory_table_storage[pdf_folder] = []
                self._memory_table_storage[pdf_folder].extend(all_tables_data)

            # 构建返回结果
            return {
                "success": True,
                "total_images": len(valid_images),
                "success_count": success_count,
                "failed_count": failed_count,
                "processing_time": total_processing_time,
                "excel_files": [],  # 空数组，不返回磁盘文件
                "raw_results": results,
                "raw_stats": {
                    "processing_time": total_processing_time,
                    "average_time_per_image": total_processing_time / len(valid_images) if valid_images else 0,
                    "total_tables_extracted": len(all_tables_data),
                    "memory_mode": True
                },
                "pdf_uuid": self._get_pdf_uuid_from_database(pdf_folder),
                "table_data_available": len(all_tables_data) > 0,
                "table_data_count": len(all_tables_data),
                "stored_in_memory": True,
                "memory_storage_key": pdf_folder
            }

        except Exception as e:
            print(f"❌❌❌❌ 表格内存处理失败: {e}")
            import traceback
            traceback.print_exc()

            return {
                "success": False,
                "error": str(e),
                "total_images": len(valid_images),
                "pdf_uuid": self._get_pdf_uuid_from_database(pdf_folder)
            }

    def process_images(self, pdf_folder: str, valid_images: List[str],
                       bank_name: str = "") -> Dict[str, Any]:
        """处理图片表格 - 删除批量处理分支，统一使用内存处理"""
        print(f"📊📊 开始处理表格 - 文件夹: {pdf_folder}, 图片数: {len(valid_images)}")

        if not PIPELINE_AVAILABLE:
            return {
                "success": False,
                "error": "表格处理管道不可用",
                "total_images": len(valid_images)
            }

        try:
            # ❌ 删除或注释掉会导致重复处理的分支
            # 不再使用批量处理模式，统一使用内存处理

            # ✅ 统一使用内存处理模式
            return self.process_images_to_memory(
                pdf_folder=pdf_folder,
                valid_images=valid_images,
                bank_name=bank_name
            )

        except Exception as e:
            print(f"❌❌ 表格处理失败: {e}")
            import traceback
            traceback.print_exc()

            return {
                "success": False,
                "error": str(e),
                "total_images": len(valid_images)
            }




# ========== 2. 异步处理包装器 ==========
def create_table_processing_task(pdf_folder: str, valid_images: List[str],
                                 bank_name: str = "") -> Dict[str, Any]:
    """
    创建表格处理任务（供异步调用）- 添加UUID支持

    Args:
        pdf_folder: PDF文件夹名称
        valid_images: 有效图片路径列表
        bank_name: 银行名称

    Returns:
        dict: 任务信息，包含UUID
    """
    job_id = str(uuid.uuid4())

    # 自动从数据库获取UUID
    try:
        # 创建临时服务实例来获取UUID
        temp_service = TableProcessingService()
        pdf_uuid = temp_service._get_pdf_uuid_from_database(pdf_folder)
        print(f"📊📊 自动获取PDF UUID: {pdf_folder} -> {pdf_uuid}")
    except Exception as e:
        print(f"⚠️⚠️ 自动获取UUID失败: {e}, 使用文件夹名称作为UUID")
        pdf_uuid = pdf_folder  # 回退到文件夹名称

    return {
        "job_id": job_id,
        "pdf_folder": pdf_folder,
        "pdf_uuid": pdf_uuid,  # 新增：UUID信息
        "valid_images": valid_images,
        "bank_name": bank_name,
        "total_images": len(valid_images),
        "created_at": datetime.now().isoformat()
    }

def execute_table_processing(task_info: Dict[str, Any]) -> Dict[str, Any]:
    """执行表格处理 - 确保使用统一处理方式"""
    service = TableProcessingService()

    # 从任务信息中获取参数
    pdf_folder = task_info["pdf_folder"]
    valid_images = task_info["valid_images"]
    bank_name = task_info.get("bank_name", "")

    print(f"🚀🚀 执行表格处理任务 - 文件夹: {pdf_folder}")

    # ✅ 统一使用内存处理模式
    result = service.process_images_to_memory(
        pdf_folder=pdf_folder,
        valid_images=valid_images,
        bank_name=bank_name
    )

    # 合并任务信息和处理结果
    final_result = {
        **task_info,
        **result,
        "completed_at": datetime.now().isoformat()
    }

    return final_result


# ========== 3. 创建全局服务实例 ==========
table_processing_service = TableProcessingService()


# ========== 4. API 接口函数 ==========
def get_bank_name_from_database(pdf_folder):
    """从数据库中根据pdf_folder获取银行名称"""
    try:
        conn = sqlite3.connect(DATABASE)
        c = conn.cursor()

        print(f"🔍 查询银行名称 - pdf_folder: {pdf_folder}")

        # 🔥 修复：查询 filename 字段（不是 raw_filename）
        # SELECT bank_name FROM files
        c.execute("""
            SELECT bank_name FROM files 
            WHERE filename = ? AND deleted = 0 AND bank_name IS NOT NULL AND bank_name != ''
            ORDER BY created_at DESC 
            LIMIT 1
        """, (pdf_folder,))  # 直接精确匹配，不需要 LIKE

        result = c.fetchone()
        conn.close()

        if result and result[0]:
            print(f"🏦 从数据库获取到银行名称: {result[0]}")
            return result[0]
        else:
            print(f"ℹ️ 数据库中未找到银行名称: {pdf_folder}")
            return ""

    except Exception as e:
        print(f"⚠️ 查询数据库银行名称失败: {e}")
        import traceback
        traceback.print_exc()
        return ""


def check_existing_table_task(pdf_folder: str) -> Dict[str, Any]:
    """
    检查同一 pdf_folder 是否已有解析任务
    
    返回:
        {
            "has_existing": True/False,
            "status": "processing/completed/failed/None",
            "job_id": "xxx",
            "message": "提示信息",
            "can_rerun": True/False
        }
    """
    try:
        import redis
        from backend.configs.config import REDIS_CONFIG
        
        redis_client = redis.Redis(
            host=REDIS_CONFIG.get('host', 'localhost'),
            port=REDIS_CONFIG.get('port', 6379),
            db=REDIS_CONFIG.get('db', 0),
            decode_responses=True
        )
        
        # 查找该 pdf_folder 相关的所有任务
        task_keys = redis_client.keys("table:job:*")
        
        existing_task = None
        for key in task_keys:
            task_data = redis_client.hgetall(key)
            if task_data and task_data.get('pdf_folder') == pdf_folder:
                job_id = key.replace("table:job:", "")
                status = task_data.get('status', 'unknown')
                
                # 找到最近的任务
                if existing_task is None:
                    existing_task = {
                        'job_id': job_id,
                        'status': status,
                        'progress': task_data.get('progress', '0'),
                        'completed_at': task_data.get('completed_at', ''),
                        'original_filename': task_data.get('original_filename', '')
                    }
                else:
                    # 比较时间，取最新的
                    existing_time = existing_task.get('created_at', '')
                    new_time = task_data.get('created_at', '')
                    if new_time > existing_time:
                        existing_task = {
                            'job_id': job_id,
                            'status': status,
                            'progress': task_data.get('progress', '0'),
                            'completed_at': task_data.get('completed_at', ''),
                            'original_filename': task_data.get('original_filename', '')
                        }
        
        if existing_task:
            status = existing_task['status']
            is_processing = status in ['pending', 'queued', 'processing', 'running', 'starting', 'generating_excel']
            is_completed = status in ['completed', 'success']
            is_failed = status in ['failed', 'exception']
            
            if is_processing:
                return {
                    "has_existing": True,
                    "status": status,
                    "job_id": existing_task['job_id'],
                    "message": f"该文件已有任务正在处理中 (状态: {status})",
                    "can_rerun": True
                }
            elif is_completed:
                return {
                    "has_existing": True,
                    "status": status,
                    "job_id": existing_task['job_id'],
                    "message": f"该文件已解析完成 (任务ID: {existing_task['job_id']})",
                    "can_rerun": True
                }
            elif is_failed:
                return {
                    "has_existing": True,
                    "status": status,
                    "job_id": existing_task['job_id'],
                    "message": f"该文件上次解析失败 (任务ID: {existing_task['job_id']})",
                    "can_rerun": True
                }
        
        return {
            "has_existing": False,
            "status": None,
            "job_id": None,
            "message": "没有找到已有的任务",
            "can_rerun": False
        }
        
    except Exception as e:
        print(f"⚠️ 检查已有任务失败: {e}")
        import traceback
        traceback.print_exc()
        return {
            "has_existing": False,
            "status": None,
            "job_id": None,
            "message": f"检查任务时出错: {str(e)}",
            "can_rerun": False
        }


def submit_table_processing_task_old(pdf_folder, filtered_tables_dir, request, progress_tracker):
    """提交表格处理任务 - 更新调用方式"""
    try:
        print(f"📥📥 提交表格处理任务: pdf_folder={pdf_folder}")

        # 获取请求数据
        if request.content_type and 'application/json' in request.content_type:
            data = request.get_json() or {}
        else:
            data = request.form.to_dict() or {}
            if not data and request.data:
                try:
                    import json
                    data = json.loads(request.data.decode('utf-8'))
                except:
                    data = {}

        table_type = data.get('table_type', 'financial')
        use_ocr = data.get('use_ocr', True)

        # 先从数据库中读取银行名称，如果没有再从请求参数获取
        bank_name = get_bank_name_from_database(pdf_folder)
        print("&&&&&&&&&&&&&&&&&&&&&&pdf_folder:", pdf_folder, bank_name)
        if not bank_name:
            bank_name = data.get('bank_name', '')

        print(f"  配置参数: table_type={table_type}, use_ocr={use_ocr}, bank_name={bank_name} (from_db: {bool(bank_name)})")

        # 可选的png_names参数，如果未提供则从目录获取
        png_names = data.get('png_names', [])

        print(f"  可选png_names: {len(png_names)}个")

        # 如果未提供png_names，自动从筛选目录获取
        tables_dir = Path(filtered_tables_dir) / pdf_folder / "tables"
        if not png_names:
            if tables_dir.exists():
                png_names = [f.name for f in tables_dir.glob("*.png")]
                print(f"  自动从目录获取 {len(png_names)} 张表格图片")
            else:
                return jsonify({
                    "success": False,
                    "error": f"表格目录不存在: {tables_dir}",
                    "suggestion": "请先完成图片筛选"
                }), 400

        if not png_names:
            return jsonify({
                "success": False,
                "error": "没有找到表格图片",
                "suggestion": "筛选后没有发现包含表格的图片"
            }), 400

        # 生成作业ID
        job_id = f"table_{int(time.time())}_{uuid.uuid4().hex[:8]}"

        # 构建完整的图片路径列表
        image_paths = []
        for png_name in png_names:
            img_path = tables_dir / png_name
            if img_path.exists():
                image_paths.append(str(img_path))
            else:
                print(f"⚠️ 图片不存在: {img_path}")

        print("排序前图片：", image_paths)
        image_paths = sorted(image_paths)
        print("排序后图片：", image_paths)

        if not image_paths:
            return jsonify({
                "success": False,
                "error": "没有找到有效的图片文件",
                "suggestion": "请检查筛选后的图片文件是否存在"
            }), 400

        print(f"📸📸 找到 {len(image_paths)} 张有效图片")

        # 创建进度记录
        progress_tracker.init_table_job(
            job_id=job_id,
            job_info={
                "pdf_folder": pdf_folder,
                "table_type": table_type,
                "bank_name": bank_name,
                "total_images": len(image_paths),
                "image_paths": image_paths,
                "status": "pending",
                "stage": "pending",
                "progress": 0,
                "start_time": datetime.now().isoformat(),
                "processed_images": 0,
                "results": [],
                "error": None
            }
        )

        print(f"✅ 创建作业成功: job_id={job_id}")

        def async_process_table():
            """异步处理表格的线程函数"""
            try:
                print(f"🚀🚀 开始异步处理表格任务: {job_id}")

                # 更新状态为处理中
                progress_tracker.update_table_job(job_id, {
                    "status": "processing",
                    "stage": "starting",
                    "progress": 5,
                    "message": "开始处理表格图片..."
                })

                # ✅ 修复：使用统一的内存处理函数，传递正确的参数
                process_table_images_real(
                    job_id=job_id,
                    pdf_folder=pdf_folder,
                    image_paths=image_paths,  # 参数名统一为 image_paths
                    table_type=table_type,
                    bank_name=bank_name,
                    progress_tracker=progress_tracker,
                    skipped_images=[],  # 新增参数
                    existing_sheets=None  # 新增参数
                )

                print(f"🎉🎉 表格处理任务完成: {job_id}")

            except Exception as e:
                print(f"❌❌ 异步处理异常: {e}")
                import traceback
                traceback.print_exc()

                progress_tracker.update_table_job(job_id, {
                    "status": "failed",
                    "stage": "failed",
                    "progress": 0,
                    "error": str(e),
                    "end_time": datetime.now().isoformat(),
                    "message": f"处理失败: {str(e)}"
                })

        # 启动异步线程
        thread = threading.Thread(target=async_process_table, daemon=True)
        thread.start()

        print(f"🎯🎯 异步处理线程已启动")

        return jsonify({
            "success": True,
            "job_id": job_id,
            "message": "表格解析任务已提交",
            "pdf_folder": pdf_folder,
            "table_type": table_type,
            "bank_name": bank_name,
            "total_images": len(image_paths),
            "auto_detected_images": data.get('png_names') is None
        })

    except Exception as e:
        print(f"💥💥 提交表格处理任务失败: {e}")
        import traceback
        traceback.print_exc()

        return jsonify({
            "success": False,
            "error": f"提交任务失败: {str(e)}"
        }), 500


def submit_table_processing_task(pdf_folder, filtered_tables_dir, request, progress_tracker):
    """提交表格处理任务 - Redis队列化版本，保留原有防重逻辑"""

    # ========== 添加入口日志 ==========
    print("\n" + "=" * 80)
    print("🔄 submit_table_processing_task 函数被调用")

    try:
        print(f"📥📥 提交表格处理任务: pdf_folder={pdf_folder}")

        # 获取请求数据
        if request.content_type and 'application/json' in request.content_type:
            data = request.get_json() or {}
            print("📦 从JSON获取请求数据")
        else:
            data = request.form.to_dict() or {}
            if not data and request.data:
                try:
                    import json
                    data = json.loads(request.data.decode('utf-8'))
                    print("📦 从原始数据解析JSON")
                except:
                    data = {}
                    print("⚠️ 无法解析请求数据")
            else:
                print("📦 从表单获取请求数据")

        print(f"📊 请求数据: {data}")

        # ========== 检查是否有已有任务 ==========
        rerun = data.get('rerun', False)
        if not rerun:
            existing_check = check_existing_table_task(pdf_folder)
            if existing_check['has_existing']:
                status = existing_check['status']
                if status in ['pending', 'queued', 'processing', 'running', 'starting', 'generating_excel']:
                    # 进行中
                    print(f"⚠️ 该文件已有任务进行中: {existing_check['job_id']}")
                    return jsonify({
                        "success": False,
                        "error": existing_check['message'],
                        "existing_job_id": existing_check['job_id'],
                        "existing_status": status,
                        "action": "waiting",
                        "can_rerun": True
                    }), 200
                elif status in ['completed', 'success']:
                    # 已完成
                    print(f"ℹ️ 该文件已解析完成: {existing_check['job_id']}")
                    return jsonify({
                        "success": True,
                        "message": existing_check['message'],
                        "existing_job_id": existing_check['job_id'],
                        "existing_status": status,
                        "action": "already_completed",
                        "can_rerun": True
                    }), 200
                elif status in ['failed', 'exception']:
                    # 失败过，允许重新执行
                    print(f"ℹ️ 该文件上次解析失败，允许重新执行: {existing_check['job_id']}")
        
        table_type = data.get('table_type', 'financial')
        use_ocr = data.get('use_ocr', True)

        # 先从数据库中读取银行名称，如果没有再从请求参数获取
        print(f"🔍 查询数据库获取银行名称: {pdf_folder}")
        bank_name = get_bank_name_from_database(pdf_folder)
        print("&&&&&&&&&&&&&&&&&&&&&&pdf_folder:", pdf_folder, bank_name)
        if not bank_name:
            bank_name = data.get('bank_name', '')
            print(f"🔍 从请求参数获取银行名称: {bank_name}")

        print(
            f"  配置参数: table_type={table_type}, use_ocr={use_ocr}, bank_name={bank_name} (from_db: {bool(bank_name)})")

        # 可选的png_names参数，如果未提供则从目录获取
        png_names = data.get('png_names', [])
        print(f"  前端提供的png_names: {png_names} (数量: {len(png_names)})")

        # 如果未提供png_names，自动从筛选目录获取
        tables_dir = Path(filtered_tables_dir) / pdf_folder / "tables"
        print(f"🔍 检查表格目录: {tables_dir}")
        print(f"📁 目录是否存在: {tables_dir.exists()}")

        if not png_names:
            if tables_dir.exists():
                png_names = [f.name for f in tables_dir.glob("*.png")]
                print(f"  自动从目录获取 {len(png_names)} 张表格图片: {png_names}")
            else:
                print(f"❌ 表格目录不存在: {tables_dir}")
                return jsonify({
                    "success": False,
                    "error": f"表格目录不存在: {tables_dir}",
                    "suggestion": "请先完成图片筛选"
                }), 400

        if not png_names:
            print("❌ 没有找到表格图片")
            return jsonify({
                "success": False,
                "error": "没有找到表格图片",
                "suggestion": "筛选后没有发现包含表格的图片"
            }), 400

        # 生成作业ID
        import time
        import uuid
        job_id = f"table_{int(time.time())}_{uuid.uuid4().hex[:8]}"
        print(f"🆔 生成的作业ID: {job_id}")

        # 构建完整的图片路径列表
        image_paths = []
        for png_name in png_names:
            img_path = tables_dir / png_name
            if img_path.exists():
                image_paths.append(str(img_path))
            else:
                print(f"⚠️ 图片不存在: {img_path}")

        print("排序前图片：", image_paths)
        image_paths = sorted(image_paths)
        print("排序后图片：", image_paths)

        if not image_paths:
            print("❌ 没有找到有效的图片文件")
            return jsonify({
                "success": False,
                "error": "没有找到有效的图片文件",
                "suggestion": "请检查筛选后的图片文件是否存在"
            }), 400

        print(f"📸📸 找到 {len(image_paths)} 张有效图片")

        # 创建任务信息
        from datetime import datetime
        job_info = {
            "job_id": job_id,
            "pdf_folder": pdf_folder,
            "table_type": table_type,
            "bank_name": bank_name,
            "total_images": len(image_paths),
            "image_paths": image_paths,
            "status": "pending",
            "stage": "pending",
            "progress": 0,
            "start_time": datetime.now().isoformat(),
            "processed_images": 0,
            "results": [],
            "error": None
        }

        # ========== 初始化进度跟踪 ==========
        print(f"\n📊 初始化进度跟踪: job_id={job_id}")
        progress_tracker.init_table_job(
            job_id=job_id,
            job_info=job_info
        )

        print(f"✅ 创建作业成功: job_id={job_id}")

        # ========== 核心优化：将任务推送到Redis队列 ==========
        print(f"\n🚀 开始Redis队列推送流程...")
        queue_mode = "redis"  # 默认使用Redis队列模式

        try:
            import redis
            import json as json_module
            from backend.utils.redis_util import redis_hset_compatible

            # 连接到Redis
            print(f"🔌 尝试连接Redis: host=localhost, port=6379, db=0")
            redis_client = redis.Redis(
                host='localhost',
                port=6379,
                db=0,
                decode_responses=False,
                socket_connect_timeout=10,
                socket_timeout=10
            )

            # 测试Redis连接
            print("🔌 测试Redis连接...")
            result = redis_client.ping()
            print(f"✅ Redis连接测试成功: {result}")

            # 构建队列任务数据
            print("📦 构建队列任务数据...")
            queue_task = {
                "job_id": job_id,
                "pdf_folder": pdf_folder,
                "filtered_tables_dir": str(filtered_tables_dir),
                "table_type": table_type,
                "bank_name": bank_name,
                "image_paths": image_paths,
                "png_names": png_names,
                "use_ocr": use_ocr,
                "created_at": time.time(),
                "request_data": {
                    "table_type": table_type,
                    "use_ocr": use_ocr,
                    "bank_name": bank_name
                }
            }

            # 推送到Redis队列
            print(f"📤 推送到Redis队列: table_parse_queue")
            redis_client.lpush("table_parse_queue", json_module.dumps(queue_task, ensure_ascii=False).encode('utf-8'))
            current_queue_length = redis_client.llen("table_parse_queue")

            # 在Redis中存储任务元数据
            print(f"💾 存储任务元数据到Redis: table:job:{job_id}")
            redis_hset_compatible(redis_client, f"table:job:{job_id}", {
                "status": "queued",
                "progress": "0",
                "message": f"任务已加入队列，位置: {current_queue_length}",
                "created_at": datetime.now().isoformat(),
                "total_images": str(len(image_paths)),
                "queue_position": str(current_queue_length),
                "queue_mode": "redis"
            })

            # 设置过期时间
            redis_client.expire(f"table:job:{job_id}", 24 * 60 * 60)

            print(f"📤 任务已推送到Redis队列: {job_id}")
            print(f"📊 当前队列长度: {current_queue_length}")
            print(f"✅ Redis队列推送成功")

        except redis.exceptions.ConnectionError as e:
            print(f"🔌❌ Redis连接错误: {e}")
            queue_mode = "fallback"
        except redis.exceptions.TimeoutError as e:
            print(f"⏱️❌ Redis超时错误: {e}")
            queue_mode = "fallback"
        except Exception as redis_error:
            print(f"⚠️❌ Redis队列推送失败: {redis_error}")
            import traceback
            traceback.print_exc()

            # ========== 回退到原有异步线程模式 ==========
            print(f"\n🔄 切换到回退模式（异步线程）...")

            def async_process_table_fallback():
                """修复版本：原有的异步处理函数（回退方案）- 包含增量处理逻辑"""
                try:
                    print(f"🚀🚀 开始异步处理表格任务（回退模式）: {job_id}")

                    # 更新状态为处理中
                    progress_tracker.update_table_job(job_id, {
                        "status": "processing",
                        "stage": "starting",
                        "progress": 5,
                        "message": "开始处理表格图片..."
                    })

                    # ========== ✅ 修复点：添加增量处理逻辑 ==========
                    print(f"\n{'=' * 60}")
                    print(f"🔍 回退模式：增量处理检查")
                    print(f"{'=' * 60}")

                    image_names = [os.path.basename(img_path) for img_path in image_paths]

                    images_to_process = image_paths
                    skipped_images = []

                    try:
                        # # 尝试导入增量处理器
                        # from backend.core.incremental_processor import incremental_processor

                        # 过滤已处理的图片
                        images_to_process_names = incremental_processor.filter_processed_images(
                            pdf_folder, image_names
                        )

                        # 计算跳过的图片
                        skipped_images_names = [img for img in image_names if img not in images_to_process_names]

                        # 筛选出需要处理的图片路径
                        images_to_process = [
                            img_path for img_path in image_paths
                            if os.path.basename(img_path) in images_to_process_names
                        ]

                        skipped_images = [
                            img_path for img_path in image_paths
                            if os.path.basename(img_path) in skipped_images_names
                        ]

                        print(f"📊 增量处理结果:")
                        print(f"  - 总图片: {len(image_paths)}")
                        print(f"  - 已处理: {len(skipped_images)} (跳过)")
                        print(f"  - 待处理: {len(images_to_process)}")

                        if skipped_images:
                            print(f"⏭️ 跳过的图片:")
                            for i, img_path in enumerate(skipped_images[:3]):
                                print(f"    {i + 1}. {os.path.basename(img_path)}")
                            if len(skipped_images) > 3:
                                print(f"    ... 等 {len(skipped_images) - 3} 张")

                        if not images_to_process:
                            print(f"ℹ️ 没有新图片需要处理，所有图片都已处理过")

                    except ImportError as e:
                        print(f"⚠️ 无法导入增量处理器: {e}")
                        print(f"ℹ️ 跳过增量处理，处理所有图片")
                    except Exception as e:
                        print(f"⚠️ 增量处理异常: {e}")
                        print(f"ℹ️ 跳过增量处理，处理所有图片")

                    # 如果没有新图片需要处理，直接生成Excel文件
                    if not images_to_process and skipped_images:
                        print(f"\n{'=' * 60}")
                        print(f"🔄 回退模式：没有新图片，直接生成/检查Excel文件")
                        print(f"{'=' * 60}")

                        try:
                            excel_dir = os.path.join(EXCEL_DATA_DIR, pdf_folder)
                            if os.path.exists(excel_dir):
                                import glob
                                excel_files = glob.glob(os.path.join(excel_dir, "*.xlsx"))
                                if excel_files:
                                    existing_excel_path = excel_files[0]
                                    print(f"✅ 找到现有Excel文件: {existing_excel_path}")

                                    # 更新任务状态
                                    progress_tracker.update_table_job(job_id, {
                                        "status": "completed",
                                        "stage": "completed",
                                        "progress": 100,
                                        "message": f"任务完成。跳过 {len(skipped_images)} 张已处理图片，使用现有Excel文件",
                                        "completed_at": datetime.now().isoformat(),
                                        "total_processed": len(skipped_images),
                                        "existing_excel_used": "true",
                                        "excel_path": existing_excel_path
                                    })

                                    print(f"🎉🎉 表格处理任务完成（使用现有文件）: {job_id}")
                                    return
                        except Exception as e:
                            print(f"⚠️ 检查现有Excel文件失败: {e}")

                    # ========== ✅ 修复点：传递正确的参数 ==========
                    process_table_images_real(
                        job_id=job_id,
                        pdf_folder=pdf_folder,
                        image_paths=images_to_process,  # ✅ 使用过滤后的图片列表
                        table_type=table_type,
                        bank_name=bank_name,
                        progress_tracker=progress_tracker,
                        skipped_images=skipped_images,  # ✅ 传递跳过的图片列表
                        existing_sheets=None
                    )

                    print(f"🎉🎉 表格处理任务完成: {job_id}")

                except Exception as e:
                    print(f"❌❌ 异步处理异常: {e}")
                    import traceback
                    traceback.print_exc()

                    progress_tracker.update_table_job(job_id, {
                        "status": "failed",
                        "stage": "failed",
                        "progress": 0,
                        "error": str(e),
                        "end_time": datetime.now().isoformat(),
                        "message": f"处理失败: {str(e)}"
                    })

            # 启动异步线程（回退模式）
            import threading
            thread = threading.Thread(target=async_process_table_fallback, daemon=True)
            thread.start()

            queue_mode = "fallback"
            print(f"🎯🎯 异步处理线程已启动（回退模式）")

        # ========== 返回响应 ==========
        print(f"\n📤 返回响应: queue_mode={queue_mode}")
        print("=" * 80)

        return jsonify({
            "success": True,
            "job_id": job_id,  # 保持字段名不变
            "message": f"表格解析任务已提交 ({'Redis队列' if queue_mode == 'redis' else '异步线程'})",
            "pdf_folder": pdf_folder,
            "table_type": table_type,
            "bank_name": bank_name,
            "total_images": len(image_paths),
            "auto_detected_images": data.get('png_names') is None,
            "queue_mode": queue_mode
        })

    except Exception as e:
        print(f"\n💥💥 submit_table_processing_task 函数异常: {e}")
        import traceback
        traceback.print_exc()

        return jsonify({
            "success": False,
            "error": f"提交任务失败: {str(e)}"
        }), 500



def process_table_images_real(job_id, pdf_folder, image_paths, table_type, bank_name,
                              progress_tracker, skipped_images=None, existing_sheets=None):
    """真实的表格处理函数 - 修复重复处理问题"""
    try:
        table_service = TableProcessingService()
        processing_manager = pdf_aggregator_manager  # 获取管理器实例

        # 🔥 关键修复：过滤掉正在处理的图片
        images_to_process = []
        skipped_processing = []

        for image_path in image_paths:
            image_name = Path(image_path).name

            # 检查是否正在处理中
            if processing_manager.is_image_being_processed(pdf_folder, image_name):
                skipped_processing.append(image_name)
                print(f"⏭️⏭️ 跳过正在处理的图片: {image_name}")
            else:
                images_to_process.append(image_path)
                # 立即标记为处理中（防止并发重复）
                processing_manager.mark_image_processing(pdf_folder, image_name, True)

        total_images = len(images_to_process)
        total_original_images = total_images + (len(skipped_images) if skipped_images else 0)

        print(f"📊📊📊📊 图片处理过滤结果:")
        print(f"  - 总图片数: {len(image_paths)}")
        print(f"  - 跳过正在处理: {len(skipped_processing)}")
        print(f"  - 实际处理: {len(images_to_process)}")
        print("************bank_name*********", bank_name)

        if skipped_processing:
            print(f"  - 跳过的图片: {skipped_processing}")

        # ========== 第1步：注册PDF处理任务 ==========
        pdf_aggregator_manager.register_processing_job(pdf_folder, total_original_images, bank_name)

        # ========== 第2步：获取PDF聚合器 ==========
        aggregator = pdf_aggregator_manager.get_aggregator(pdf_folder, bank_name)

        results = []
        success_count = 0
        failed_count = 0
        total_tables_extracted = 0
        all_metadata_list = []

        # 更新进度信息
        progress_tracker.update_table_job(job_id, {
            "stage": "processing_start",
            "progress": 10,
            "message": f"开始处理 {len(images_to_process)} 张新图片",
            "skipped_processing_count": len(skipped_processing),
            "actual_processing_count": len(images_to_process)
        })

        # ========== 第3步：处理图片 ==========
        try:
            for i, image_path in enumerate(images_to_process):
                image_name = Path(image_path).name

                # 更新进度
                progress = 10 + (i / len(images_to_process) * 70)
                current_stage = "ocr" if i < len(images_to_process) * 0.3 else "llm" if i < len(
                    images_to_process) * 0.6 else "reconstruction"

                progress_tracker.update_table_job(job_id, {
                    "stage": current_stage,
                    "progress": int(progress),
                    "processed_images": i,
                    "current_image": image_name,
                    "message": f"正在处理第 {i + 1}/{len(images_to_process)} 张图片 ({current_stage})"
                })

                print(f"🖼🖼🖼🖼 处理图片 {i + 1}/{len(images_to_process)}: {image_name}")

                try:
                    # ✅ 调用内存处理流水线
                    tables_data, sheet_names, metadata_list = table_service._run_ocr_llm_memory_pipeline(
                        image_path=image_path,
                        bank_name=bank_name
                    )

                    if tables_data:
                        # 将表格数据添加到聚合器
                        for table_idx, (table_data, sheet_name) in enumerate(zip(tables_data, sheet_names)):
                            table_metadata = metadata_list[table_idx] if table_idx < len(metadata_list) else {}
                            success = aggregator.add_table(
                                image_name=image_name,
                                table_data=table_data,
                                sheet_name=sheet_name,
                                image_path=image_path,
                                metadata=table_metadata
                            )
                            if success:
                                total_tables_extracted += 1
                                all_metadata_list.append(table_metadata)

                        success_count += 1

                        results.append({
                            "image_path": image_name,
                            "success": True,
                            "tables_extracted": len(tables_data),
                            "processing_time": 0
                        })

                        print(f"✅ 图片处理成功: {image_name}, 提取 {len(tables_data)} 个表格")
                    else:
                        failed_count += 1
                        results.append({
                            "image_path": image_name,
                            "success": False,
                            "error": "未提取到表格数据"
                        })
                        print(f"⚠️ 图片处理未提取到表格: {image_name}")

                    # 标记图片完成
                    pdf_aggregator_manager.mark_image_completed(pdf_folder, image_name,
                                                                len(sheet_names) if tables_data else 0)

                except Exception as img_error:
                    print(f"❌❌ 图片处理失败 {image_name}: {img_error}")
                    import traceback
                    traceback.print_exc()
                    failed_count += 1
                    results.append({
                        "image_path": image_name,
                        "success": False,
                        "error": str(img_error)
                    })
                    pdf_aggregator_manager.mark_image_completed(pdf_folder, image_name, 0)

                finally:
                    # 🔥 关键：无论成功失败，都要标记处理完成
                    processing_manager.mark_image_processing(pdf_folder, image_name, False)

                # 更新总进度
                progress_tracker.update_table_job(job_id, {
                    "progress": int(10 + ((i + 1) / len(images_to_process) * 70)),
                    "processed_images": i + 1,
                    "success_count": success_count
                })

        except Exception as batch_error:
            print(f"❌❌ 批量处理异常: {batch_error}")
            # 确保异常时清理所有处理状态
            for image_path in images_to_process:
                image_name = Path(image_path).name
                processing_manager.mark_image_processing(pdf_folder, image_name, False)
            raise batch_error

        # ========== 第4步：最终合并 ==========
        print(f"🔄🔄🔄🔄 所有图片处理完成，开始最终合并: {pdf_folder}")

        progress_tracker.update_table_job(job_id, {
            "stage": "merging",
            "progress": 85,
            "message": f"正在合并 {len(aggregator)} 个表格到Excel..."
        })

        try:
            # 最终化PDF，生成Excel
            success, excel_path, error_msg = pdf_aggregator_manager.finalize_pdf(
                pdf_folder,
                EXCEL_DATA_DIR,
                force=False,
                metadata_list=all_metadata_list
            )

            excel_files = [excel_path] if success and excel_path else []

            # ========== 第5步：处理完成 ==========
            progress_tracker.update_table_job(job_id, {
                "status": "completed" if success else "failed",
                "stage": "completed",
                "progress": 100,
                "processed_images": len(images_to_process),
                "success_count": success_count,
                "failed_count": failed_count,
                "total_tables_extracted": total_tables_extracted,
                "final_excel": excel_path if success else None,
                "results": results,
                "excel_files": excel_files,
                "end_time": datetime.now().isoformat(),
                "message": f"处理完成: 成功 {success_count}/{len(images_to_process)} 张图片, 提取 {total_tables_extracted} 个表格",
                "summary": {
                    "total_original_images": total_original_images,
                    "skipped_processing": len(skipped_processing),
                    "actual_processed": len(images_to_process),
                    "success_rate": f"{(success_count / len(images_to_process) * 100):.1f}%" if images_to_process else "0%"
                }
            })

            # 清理聚合器数据
            if success:
                aggregator.clear()
                # 🔥 清理处理状态跟踪
                processing_manager.cleanup_processing_tracking(pdf_folder)

            print(f"✅ PDF处理完成: {pdf_folder}")

        except Exception as merge_error:
            print(f"❌❌ 最终合并异常: {merge_error}")
            raise merge_error

    except Exception as e:
        print(f"❌❌❌❌ 表格处理失败: {e}")
        import traceback
        traceback.print_exc()

        # 🔥 异常时确保清理处理状态
        processing_manager.cleanup_processing_tracking(pdf_folder)

        progress_tracker.update_table_job(job_id, {
            "status": "failed",
            "stage": "failed",
            "error": str(e),
            "end_time": datetime.now().isoformat(),
            "message": f"处理失败: {str(e)}"
        })


def _check_existing_excel(pdf_folder: str, bank_name: str = "") -> Tuple[bool, Optional[str]]:
    """
    独立的辅助函数，检查是否已存在Excel文件且包含有效数据

    返回: (是否有效, 文件路径)
    """
    try:
        from pathlib import Path
        import openpyxl

        # 使用导入的EXCEL_DATA_DIR常量
        excel_dir = Path(EXCEL_DATA_DIR) / pdf_folder
        print(f"🔍 检查目录: {excel_dir}")

        if excel_dir.exists():
            # 查找Excel文件
            excel_files = list(excel_dir.glob("*.xlsx"))
            if excel_files:
                print(f"  📁 找到 {len(excel_files)} 个Excel文件")

                # 按修改时间排序，取最新的
                excel_files.sort(key=lambda x: x.stat().st_mtime, reverse=True)
                excel_path = excel_files[0]

                # 验证文件大小
                file_size = excel_path.stat().st_size
                if file_size <= 0:
                    print(f"  ⚠️ 文件为空: {excel_path.name}")
                    return False, str(excel_path)

                # 验证Excel文件内容
                print(f"  🔍 验证Excel文件内容: {excel_path.name}")

                try:
                    # 尝试打开Excel文件
                    wb = openpyxl.load_workbook(str(excel_path), data_only=True, read_only=True)

                    sheet_names = wb.sheetnames
                    print(f"  📄 Sheet数量: {len(sheet_names)}")
                    print(f"  📄 Sheet名称: {sheet_names}")

                    # 检查是否有实际数据
                    has_real_data = False
                    table_count = 0

                    for sheet_name in sheet_names:
                        ws = wb[sheet_name]
                        max_row = ws.max_row
                        max_col = ws.max_column

                        # 排除空表或只有表头的表
                        if max_row > 1 and max_col > 1:  # 至少2行2列才认为是有效表格
                            has_real_data = True
                            table_count += 1
                            print(f"  ✅ Sheet '{sheet_name}': {max_row}行×{max_col}列")
                        else:
                            print(f"  ⚠️ Sheet '{sheet_name}' 可能为空")

                    wb.close()

                    if has_real_data:
                        print(f"  ✅ Excel文件包含 {table_count} 个有效表格，大小: {file_size} 字节")
                        return True, str(excel_path)
                    else:
                        print(f"  ⚠️ Excel文件没有有效表格数据")
                        return False, str(excel_path)

                except Exception as e:
                    print(f"  ⚠️ 读取Excel文件失败: {e}")
                    return False, str(excel_path)

        print(f"📄 未发现有效的Excel文件")
        return False, None

    except Exception as e:
        print(f"⚠️ 检查现有Excel文件失败: {e}")
        return False, None


def process_images_with_real_time_updates(
        job_id: str,
        pdf_folder: str,
        image_paths: List[str],
        table_type: str = "financial",
        bank_name: str = "",
        progress_tracker=None
) -> Dict[str, Any]:
    """
    重构版本：职责单一、状态清晰的图片处理函数

    职责：只处理图片，将表格数据添加到聚合器
    不负责生成Excel文件，Excel生成由调用方统一处理

    设计原则：
    1. 单一职责：只处理图片，不生成Excel
    2. 状态透明：明确返回处理结果和下一步建议
    3. 错误隔离：处理失败不影响整体流程
    4. 增量友好：正确处理已处理和未处理图片

    返回结果结构：
    {
        "success": bool,                # 整体处理是否成功
        "operation": str,               # 执行的操作类型
        "images_processed": int,        # 实际处理的图片数量
        "tables_added": int,            # 添加到聚合器的表格数量
        "aggregator_has_data": bool,    # 聚合器当前是否有数据
        "next_action": str,             # 建议的下一步操作
        "need_generate_excel": bool,    # 是否需要生成Excel
        "excel_exists": bool,           # Excel文件是否已存在
        "excel_path": Optional[str],    # 已存在的Excel路径（如果存在）
        "details": Dict,                # 详细处理信息
        "errors": List[str]             # 错误信息列表
    }
    """

    start_time = time.time()

    # 初始化结果结构
    result = {
        "success": False,
        "job_id": job_id,
        "pdf_folder": pdf_folder,
        "operation": "unknown",
        "total_images_received": len(image_paths),
        "images_processed": 0,
        "tables_added": 0,
        "aggregator_has_data": False,
        "need_generate_excel": False,
        "excel_exists": False,
        "excel_path": None,
        "next_action": "unknown",
        "execution_time": 0,
        "details": {
            "images_filtered": 0,
            "images_to_process": 0,
            "images_skipped": 0,
            "success_count": 0,
            "failed_count": 0,
            "aggregator_tables_before": 0,
            "aggregator_tables_after": 0
        },
        "errors": []
    }

    try:
        # ========== 阶段1：状态检查和初始化 ==========
        print(f"\n🔍 阶段1：状态检查和初始化")

        # 1.1 检查图片路径有效性
        valid_image_paths = []
        invalid_images = []

        for img_path in image_paths:
            if os.path.exists(img_path):
                valid_image_paths.append(img_path)
            else:
                invalid_images.append(img_path)

        if invalid_images:
            print(f"⚠️ 发现 {len(invalid_images)} 个无效图片路径")
            for invalid_img in invalid_images[:3]:  # 只显示前3个
                print(f"  - {invalid_img}")
            if len(invalid_images) > 3:
                print(f"  ... 等 {len(invalid_images) - 3} 个")

            result["errors"].append(f"发现 {len(invalid_images)} 个无效图片路径")

        if not valid_image_paths:
            result["operation"] = "no_valid_images"
            result["next_action"] = "check_existing_excel"
            result["execution_time"] = time.time() - start_time
            print(f"❌ 没有有效的图片路径，跳过处理")
            return result

        # 1.2 获取图片名称列表
        image_names = [os.path.basename(img_path) for img_path in valid_image_paths]

        # 1.3 检查增量处理状态
        print(f"\n🔍 检查增量处理状态...")
        try:
            # 过滤已处理的图片
            images_to_process_names = incremental_processor.filter_processed_images(pdf_folder, image_names)
            skipped_images_names = [img for img in image_names if img not in images_to_process_names]

            result["details"]["images_filtered"] = len(image_names)
            result["details"]["images_to_process"] = len(images_to_process_names)
            result["details"]["images_skipped"] = len(skipped_images_names)

        except Exception as e:
            print(f"⚠️ 增量处理器检查失败: {e}")
            result["errors"].append(f"增量处理器检查失败: {str(e)}")
            # 回退：处理所有图片
            images_to_process_names = image_names
            skipped_images_names = []

        # 1.4 检查Excel文件是否已存在
        print(f"\n🔍 检查现有Excel文件...")
        excel_exists, existing_excel_path = _check_existing_excel(pdf_folder, bank_name)

        result["excel_exists"] = excel_exists
        result["excel_path"] = existing_excel_path

        if excel_exists and existing_excel_path:
            print(f"✅ 发现有效的现有Excel文件: {existing_excel_path}")
        else:
            print(f"📄 未发现有效的Excel文件，需要创建新文件")

        # 1.5 获取聚合器并记录初始状态
        print(f"\n🔍 初始化聚合器...")
        try:
            # 注册处理任务
            pdf_aggregator_manager.register_processing_job(pdf_folder, len(image_names), bank_name)

            # 获取聚合器实例
            aggregator = pdf_aggregator_manager.get_aggregator(pdf_folder, bank_name)

            # 记录初始状态
            initial_tables_count = len(aggregator)
            result["details"]["aggregator_tables_before"] = initial_tables_count
            result["aggregator_has_data"] = initial_tables_count > 0

            print(f"📊 聚合器初始状态:")
            print(f"  - 表格数量: {initial_tables_count}")
            print(f"  - 是否有数据: {result['aggregator_has_data']}")

        except Exception as e:
            print(f"❌ 聚合器初始化失败: {e}")
            result["errors"].append(f"聚合器初始化失败: {str(e)}")
            result["execution_time"] = time.time() - start_time
            return result

        # ========== 阶段2：决策逻辑 ==========
        print(f"\n🤔 阶段2：处理决策")

        # 决策树：
        # 1. 如果没有图片需要处理
        # 2. 如果聚合器已有数据
        # 3. 如果有新图片需要处理

        if not images_to_process_names:
            # 情况1：没有新图片需要处理
            result["operation"] = "no_new_images"
            result["images_processed"] = 0

            if result["aggregator_has_data"]:
                # 聚合器有数据，需要生成Excel
                result["need_generate_excel"] = True
                result["next_action"] = "generate_excel_from_aggregator"
                print(f"📊 决策: 无新图片，但聚合器有 {initial_tables_count} 个表格，需要生成Excel")
            elif excel_exists:
                # 有有效Excel文件，无需处理
                result["need_generate_excel"] = False
                result["next_action"] = "use_existing_excel"
                print(f"📊 决策: 无新图片，有有效Excel，无需处理")
            else:
                # 既无新图片，聚合器也无数据，也没有有效Excel
                result["need_generate_excel"] = True
                result["next_action"] = "generate_excel"
                print(f"📊 决策: 无新图片，无聚合器数据，无有效Excel，需要创建新Excel")

            result["success"] = True
            result["execution_time"] = time.time() - start_time
            return result

        # 情况2：有新图片需要处理
        result["operation"] = "process_new_images"

        # 筛选出需要处理的图片路径
        images_to_process = [
            img_path for img_path in valid_image_paths
            if os.path.basename(img_path) in images_to_process_names
        ]

        print(f"🔄 决策: 有 {len(images_to_process)} 张新图片需要处理")

        # ========== 阶段3：图片处理 ==========
        print(f"\n🖼️ 阶段3：处理 {len(images_to_process)} 张新图片")

        # 3.1 初始化表格处理服务
        try:
            table_service = TableProcessingService()
        except Exception as e:
            print(f"❌ 表格处理服务初始化失败: {e}")
            result["errors"].append(f"表格处理服务初始化失败: {str(e)}")
            result["execution_time"] = time.time() - start_time
            return result

        # 3.2 逐张处理图片
        processed_count = 0
        success_count = 0
        failed_count = 0
        tables_added = 0
        all_metadata_list = []

        for i, image_path in enumerate(images_to_process):
            image_name = os.path.basename(image_path)

            try:
                # 3.2.1 更新进度
                if progress_tracker:
                    progress_percent = int((i / len(images_to_process)) * 80)
                    progress_tracker.update_table_job(job_id, {
                        "status": "processing",
                        "progress": progress_percent,
                        "message": f"处理第 {i + 1}/{len(images_to_process)} 张图片: {image_name}",
                        "current_image": image_name,
                        "processed_images": i
                    })

                print(f"\n{'─' * 40}")
                print(f"🖼 处理图片 {i + 1}/{len(images_to_process)}: {image_name}")
                print(f"{'─' * 40}")

                # 3.2.2 调用OCR-LLM流水线
                tables_data, sheet_names, metadata_list = table_service._run_ocr_llm_memory_pipeline(
                    image_path=image_path,
                    bank_name=bank_name
                )

                if tables_data:
                    # 3.2.3 将表格数据添加到聚合器
                    for table_idx, (table_data, sheet_name) in enumerate(zip(tables_data, sheet_names)):
                        table_metadata = metadata_list[table_idx] if table_idx < len(metadata_list) else {}

                        # 添加到聚合器
                        add_success = aggregator.add_table(
                            image_name=image_name,
                            table_data=table_data,
                            sheet_name=sheet_name,
                            image_path=image_path,
                            metadata=table_metadata
                        )

                        if add_success:
                            tables_added += 1
                            all_metadata_list.append(table_metadata)
                            print(f"  ✅ 添加表格 {table_idx + 1}: '{sheet_name}'")

                    success_count += 1
                    print("555555555555555555555555555")
                    print(f"✅ 图片处理成功: {image_name}, 添加 {len(tables_data)} 个表格")
                else:
                    failed_count += 1
                    print(f"⚠️ 图片未提取到表格: {image_name}")

                # 3.2.4 标记图片为已处理
                try:
                    pdf_aggregator_manager.mark_image_completed(
                        pdf_folder,
                        image_name,
                        len(tables_data) if tables_data else 0
                    )
                except Exception as e:
                    print(f"⚠️ 标记图片完成失败: {e}")
                    result["errors"].append(f"标记图片完成失败 {image_name}: {str(e)}")

            except Exception as img_error:
                failed_count += 1
                print(f"❌ 处理图片失败 {image_name}: {img_error}")
                result["errors"].append(f"图片处理失败 {image_name}: {str(img_error)}")

                # 标记图片完成（即使失败）
                try:
                    pdf_aggregator_manager.mark_image_completed(pdf_folder, image_name, 0)
                except Exception as e:
                    print(f"⚠️ 标记失败图片完成失败: {e}")

            finally:
                processed_count += 1

        # ========== 阶段4：结果汇总 ==========
        print(f"\n📊 阶段4：处理结果汇总")

        # 4.1 更新聚合器状态
        final_tables_count = len(aggregator)
        result["details"]["aggregator_tables_after"] = final_tables_count
        result["details"]["success_count"] = success_count
        result["details"]["failed_count"] = failed_count

        result["images_processed"] = processed_count
        result["tables_added"] = tables_added
        result["aggregator_has_data"] = final_tables_count > 0

        print(f"📊 处理统计:")
        print(f"  - 处理图片: {processed_count} 张")
        print(f"  - 成功: {success_count} 张")
        print(f"  - 失败: {failed_count} 张")
        print(f"  - 添加表格: {tables_added} 个")
        print(f"  - 聚合器总数: {final_tables_count} 个表格")

        # 4.2 决策下一步操作
        if result["aggregator_has_data"]:
            result["need_generate_excel"] = True
            result["next_action"] = "generate_excel_from_aggregator"
            print(f"📊 决策: 聚合器有 {final_tables_count} 个表格，需要生成Excel")
        elif excel_exists:
            result["need_generate_excel"] = False
            result["next_action"] = "use_existing_excel"
            print(f"📊 决策: 使用现有有效Excel文件")
        else:
            result["need_generate_excel"] = True
            result["next_action"] = "generate_excel"
            print(f"📊 决策: 无表格数据，无有效Excel，需要创建新Excel")

        # 4.3 更新进度
        if progress_tracker:
            progress_tracker.update_table_job(job_id, {
                "status": "processing_completed",
                "progress": 90,
                "message": f"图片处理完成: 成功 {success_count} 张，失败 {failed_count} 张，添加 {tables_added} 个表格",
                "processed_images": processed_count,
                "success_count": success_count,
                "failed_count": failed_count,
                "tables_added": tables_added
            })

        result["success"] = True
        result["execution_time"] = time.time() - start_time

        print(f"\n✅ 图片处理函数完成")
        print(f"⏱️ 执行时间: {result['execution_time']:.2f}秒")
        print(f"📤 返回结果: {result['operation']}, 下一步: {result['next_action']}")

        return result

    except Exception as e:
        print(f"❌ 图片处理函数异常: {e}")
        import traceback
        traceback.print_exc()

        result["errors"].append(f"函数异常: {str(e)}")
        result["execution_time"] = time.time() - start_time

        if progress_tracker:
            progress_tracker.update_table_job(job_id, {
                "status": "failed",
                "error": f"处理函数异常: {str(e)}"
            })

        return result



# 保留get_table_results函数
def get_table_results(pdf_folder, progress_tracker):
    """API: 查询表格处理结果"""
    from flask import jsonify

    try:
        # 1. 获取内存中的任务
        folder_tasks = progress_tracker.get_folder_tasks(pdf_folder)

        # 2. 获取Excel文件
        service = TableProcessingService()
        excel_files = service.get_excel_files(pdf_folder)

        # 3. 返回结果
        return jsonify({
            "success": True,
            "pdf_folder": pdf_folder,
            "data": {
                "tasks": folder_tasks,
                "excel_files": excel_files,
                "task_count": len(folder_tasks),
                "excel_count": len(excel_files)
            }
        })

    except Exception as e:
        return jsonify({
            "success": False,
            "error": f"查询失败: {str(e)}"
        }), 500


def download_excel_file(pdf_folder, filename):
    """API: 下载Excel文件"""
    from flask import send_from_directory, jsonify

    try:
        service = TableProcessingService()
        file_path = service.get_excel_file_path(pdf_folder, filename)

        if not file_path or not file_path.exists():
            return jsonify({
                "success": False,
                "error": "文件不存在"
            }), 404

        return send_from_directory(
            directory=str(file_path.parent),
            path=filename,
            as_attachment=True,
            download_name=f"表格处理结果_{pdf_folder}_{filename}"
        )

    except Exception as e:
        return jsonify({
            "success": False,
            "error": f"下载失败: {str(e)}"
        }), 500



def get_available_steps():
    """API: 获取可用步骤"""
    from flask import jsonify
    try:
        from backend.services.layout_service import processing_pipeline
        steps = processing_pipeline.get_available_steps()
        return jsonify({
            "success": True,
            "data": {
                "steps": steps,
                "count": len(steps)
            }
        })
    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        }), 500


def process_single_table_image(pdf_folder, image_path, table_type, bank_name):
    """
    处理单张表格图片 - 提供给 table_worker.py 调用的兼容函数

    参数：
        pdf_folder: PDF文件夹名称
        image_path: 图片路径
        table_type: 表格类型
        bank_name: 银行名称

    返回：
        Dict: 处理结果
    """
    try:
        print(f"🔍 处理单张表格图片: {os.path.basename(image_path)}")

        # 创建处理服务
        service = TableProcessingService()

        # 运行OCR-LLM流水线
        tables_data, sheet_names, metadata_list = service._run_ocr_llm_memory_pipeline(
            image_path=image_path,
            bank_name=bank_name
        )

        if tables_data:
            # 获取聚合器
            aggregator = pdf_aggregator_manager.get_aggregator(pdf_folder, bank_name)

            image_name = os.path.basename(image_path)
            tables_added = 0

            # 将表格添加到聚合器
            for table_idx, (table_data, sheet_name) in enumerate(zip(tables_data, sheet_names)):
                table_metadata = metadata_list[table_idx] if table_idx < len(metadata_list) else {}

                add_success = aggregator.add_table(
                    image_name=image_name,
                    table_data=table_data,
                    sheet_name=sheet_name,
                    image_path=image_path,
                    metadata=table_metadata
                )

                if add_success:
                    tables_added += 1

            return {
                "success": True,
                "image_path": image_path,
                "tables": len(tables_data),
                "tables_added": tables_added,
                "message": f"成功处理 {os.path.basename(image_path)}，添加 {tables_added} 个表格"
            }
        else:
            return {
                "success": False,
                "error": f"未提取到表格数据: {os.path.basename(image_path)}"
            }

    except Exception as e:
        import traceback
        traceback.print_exc()
        return {
            "success": False,
            "error": f"处理失败: {str(e)}"
        }


import concurrent.futures
from threading import Semaphore
from typing import List, Dict, Any


class HighVolumeTableProcessor:
    """高容量表格处理器 - 处理上百张图片"""

    def __init__(self, config: Dict[str, Any] = None):
        self.config = config or {
            'max_ocr_workers': 4,  # OCR线程数（I/O密集型）
            'max_llm_workers': 2,  # LLM线程数（GPU限制）
            'max_reconstruct_workers': 3,  # 重构线程数
            'batch_size': 10,  # 批次大小，控制内存
            'queue_size': 50  # 队列缓冲
        }

        # 资源限制信号量
        self.gpu_semaphore = Semaphore(self.config['max_llm_workers'])

    def process_hundred_images(self, image_paths: List[str], bank_name: str = "") -> Dict[str, Any]:
        """
        处理上百张图片的优化方案
        """
        total_images = len(image_paths)
        print(f"🚀 开始处理 {total_images} 张图片")

        # 1. 分批处理，避免内存爆炸
        batches = self._create_batches(image_paths, self.config['batch_size'])

        all_results = []

        with concurrent.futures.ThreadPoolExecutor(
                max_workers=self.config['max_ocr_workers'] +
                            self.config['max_llm_workers'] +
                            self.config['max_reconstruct_workers']
        ) as executor:

            # 提交批次任务
            future_to_batch = {}
            for batch_idx, batch_images in enumerate(batches):
                future = executor.submit(
                    self._process_batch_pipeline,
                    batch_images, batch_idx, bank_name
                )
                future_to_batch[future] = batch_idx

            # 收集结果
            for future in concurrent.futures.as_completed(future_to_batch):
                batch_idx = future_to_batch[future]
                try:
                    batch_results = future.result()
                    all_results.extend(batch_results)
                    print(f"✅ 批次 {batch_idx + 1}/{len(batches)} 处理完成")
                except Exception as e:
                    print(f"❌ 批次 {batch_idx} 处理失败: {e}")

        # 汇总统计
        return self._aggregate_results(all_results, total_images)

    def _create_batches(self, image_paths: List[str], batch_size: int) -> List[List[str]]:
        """创建批次"""
        return [image_paths[i:i + batch_size]
                for i in range(0, len(image_paths), batch_size)]

    def _process_batch_pipeline(self, batch_images: List[str],
                                batch_idx: int, bank_name: str) -> List[Dict[str, Any]]:
        """
        批次内的流水线处理
        """
        batch_results = []

        # 阶段1: OCR识别（并行）
        ocr_results = []
        with concurrent.futures.ThreadPoolExecutor(
                max_workers=min(self.config['max_ocr_workers'], len(batch_images))
        ) as ocr_executor:
            ocr_futures = {
                ocr_executor.submit(self._ocr_recognize, img_path): img_path
                for img_path in batch_images
            }

            for future in concurrent.futures.as_completed(ocr_futures):
                img_path = ocr_futures[future]
                try:
                    ocr_result = future.result()
                    ocr_results.append((img_path, ocr_result))
                except Exception as e:
                    print(f"❌ OCR失败 {img_path}: {e}")
                    ocr_results.append((img_path, {'success': False, 'error': str(e)}))

        # 阶段2: LLM分析（受GPU限制）
        llm_results = []
        for img_path, ocr_result in ocr_results:
            if ocr_result.get('success'):
                # 使用信号量限制并发
                with self.gpu_semaphore:
                    llm_result = self._llm_analyze(img_path, ocr_result)
                    llm_results.append((img_path, ocr_result, llm_result))
            else:
                llm_results.append((img_path, ocr_result, {'success': False, 'error': 'OCR失败'}))

        # 阶段3: 表格重构（并行）
        for img_path, ocr_result, llm_result in llm_results:
            if llm_result.get('success'):
                try:
                    reconstruct_result = self._table_reconstruct(
                        ocr_result, llm_result, img_path, bank_name
                    )
                    batch_results.append({
                        'image_path': img_path,
                        'success': reconstruct_result.get('success', False),
                        'output_file': reconstruct_result.get('output_file'),
                        'processing_time': reconstruct_result.get('processing_time', 0)
                    })
                except Exception as e:
                    batch_results.append({
                        'image_path': img_path,
                        'success': False,
                        'error': str(e)
                    })
            else:
                batch_results.append({
                    'image_path': img_path,
                    'success': False,
                    'error': llm_result.get('error', 'LLM分析失败')
                })

        return batch_results

    def _ocr_recognize(self, image_path: str) -> Dict[str, Any]:
        """OCR识别"""
        ocr_service = TableOCRService()
        return ocr_service.recognize_table(image_path)

    def _llm_analyze(self, image_path: str, ocr_result: Dict[str, Any]) -> Dict[str, Any]:
        """LLM分析（GPU密集型）"""
        analyzer = EnhancedFinancialTableAnalyzer()
        return analyzer.analyze_image(image_path, ocr_result)

    def _table_reconstruct(self, ocr_result: Dict[str, Any],
                           llm_result: Dict[str, Any],
                           image_path: str, bank_name: str) -> Dict[str, Any]:
        """表格重构"""
        reconstructor = TableReconstructor()

        # 生成输出文件路径
        from pathlib import Path
        image_name = Path(image_path).stem
        output_dir = Path("data/backend/outputs/large_batch")
        output_dir.mkdir(parents=True, exist_ok=True)
        output_file = str(output_dir / f"{image_name}_reconstructed.xlsx")

        success = reconstructor.process_all_tables(
            ocr_result=ocr_result,
            llm_result=llm_result,
            output_file=output_file,
            image_path=image_path,
            bank_name=bank_name
        )

        return {
            'success': success,
            'output_file': output_file if success else None
        }

    def _aggregate_results(self, all_results: List[Dict], total_images: int) -> Dict[str, Any]:
        """汇总结果"""
        successful = sum(1 for r in all_results if r.get('success'))
        failed = total_images - successful

        return {
            'success': failed == 0,
            'total_images': total_images,
            'successful': successful,
            'failed': failed,
            'results': all_results,
            'summary': {
                'success_rate': f"{(successful / total_images * 100):.1f}%",
                'failed_images': [
                    r['image_path'] for r in all_results
                    if not r.get('success')
                ]
            }
        }


def process_large_batch(pdf_folder: str, image_paths: List[str],
                        bank_name: str = "") -> Dict[str, Any]:
    """
    处理大批量图片的接口
    """
    processor = HighVolumeTableProcessor({
        'max_ocr_workers': 6,
        'max_llm_workers': 2,  # 根据GPU显存调整
        'max_reconstruct_workers': 4,
        'batch_size': 15,
        'queue_size': 30
    })

    return processor.process_hundred_images(image_paths, bank_name)

def _ensure_table_processing_db():
    """确保表格处理数据库表存在"""
    global _table_processing_db_initialized

    if _table_processing_db_initialized:
        return True

    try:

        db_handler = NewDatabaseManager(DATABASE_PATH)
        db_handler.init_table_processing_db()
        _table_processing_db_initialized = True
        return True
    except Exception as e:
        print(f"⚠️ 数据库表初始化失败，但继续处理: {e}")
        return False


# def update_job_progress(job_id, updates):
def update_job_progress(job_id, updates, progress_tracker):
    """更新任务进度"""
    if job_id in progress_tracker.TABLE_PROCESSING_JOBS:
        progress_tracker.TABLE_PROCESSING_JOBS[job_id].update(updates)

        # 保存到数据库
        try:
            from backend.utils.constants import DATABASE_PATH
            db_handler = NewDatabaseManager(DATABASE_PATH)
            job_info = progress_tracker.TABLE_PROCESSING_JOBS[job_id].copy()
            job_info['job_id'] = job_id
            db_handler.save_table_processing_record(job_info)
        except Exception as e:
            print(f"⚠️ 保存进度到数据库失败: {e}")

_table_processing_db_initialized = False



def process_tables_async(job_id, pdf_folder, valid_images, bank_name):
    """异步处理表格的完整实现 - 更新处理逻辑"""
    print(f"🚀🚀 开始异步处理表格 - Job ID: {job_id}")

    try:
        # 确保数据库表存在
        _ensure_table_processing_db()

        # 更新进度为开始
        update_job_progress(job_id, {
            "status": "processing",
            "stage": "starting",
            "progress": 5,
            "total_images": len(valid_images)
        })

        # ========== 统一使用内存处理模式 ==========
        service = TableProcessingService()

        # ✅ 直接调用内存处理方法（避免重复处理）
        result = service.process_images_to_memory(
            pdf_folder=pdf_folder,
            valid_images=valid_images,
            bank_name=bank_name
        )

        # 更新进度为完成
        update_job_progress(job_id, {
            "status": "completed" if result.get("success") else "failed",
            "stage": "completed",
            "progress": 100,
            "end_time": datetime.now().isoformat(),
            **result  # 合并结果
        })

        print(f"✅ 表格处理任务完成 - Job ID: {job_id}")

    except Exception as e:
        print(f"❌❌ 表格处理任务失败: {e}")
        import traceback
        traceback.print_exc()

        update_job_progress(job_id, {
            "status": "failed",
            "stage": "failed",
            "error": str(e),
            "end_time": datetime.now().isoformat(),
            "message": f"处理失败: {str(e)}"
        })

def execute_single_step_handler(step_name, output_dir, request):
    """分步执行表格处理 - 真正的分步实现"""

    if request.method == 'OPTIONS':
        return jsonify({"status": "ok"}), 200

    try:
        data = request.get_json()
        pdf_folder = data.get('pdf_folder')
        png_names = data.get('png_names', [])
        previous_context = data.get('previous_context', {})

        if not pdf_folder or not png_names:
            return jsonify({
                "success": False,
                "error": "参数错误：需提供pdf_folder和png_names"
            }), 400


        # 根据步骤名称执行不同的逻辑
        if step_name == "ocr":
            result = execute_ocr_step(pdf_folder, png_names, output_dir)
        elif step_name == "llm":
            result = execute_llm_step(pdf_folder, png_names, previous_context, output_dir)
        elif step_name == "reconstruct":
            result = execute_reconstruct_step(pdf_folder, png_names, previous_context, output_dir)
        elif step_name == "export":
            result = execute_export_step(pdf_folder, png_names, previous_context, output_dir)
        else:
            return jsonify({
                "success": False,
                "error": f"不支持的步骤: {step_name}"
            }), 400


        print("****************>>>>step_name:", step_name)
        print("result:::::>", result)


        return jsonify(result)

    except Exception as e:
        return jsonify({
            "success": False,
            "error": f"分步执行失败: {str(e)}"
        }), 500


def execute_ocr_step(pdf_folder, png_names, output_dir):
    """执行OCR步骤"""
    try:
        results = {}
        ocr_service = TableOCRService()

        for png_name in png_names:
            # 构建图片路径
            from backend.utils.constants import JOINED_TABLES_DIR
            image_path = JOINED_TABLES_DIR / pdf_folder / png_name

            if not image_path.exists():
                results[png_name] = {"success": False, "error": "图片不存在"}
                continue

            # 执行OCR
            ocr_result = ocr_service.recognize_table(str(image_path))
            results[png_name] = {
                "success": True,
                "result": ocr_result,
                "tables_count": len(ocr_result.get('tables_result', []))
            }

        return {
            "success": True,
            "step": "ocr",
            "results": results,
            "message": f"OCR识别完成，处理{len(png_names)}张图片"
        }

    except Exception as e:
        return {
            "success": False,
            "step": "ocr",
            "error": str(e)
        }


def execute_llm_step(pdf_folder, png_names, previous_context, output_dir):
    """执行LLM分析步骤"""
    try:

        # 检查是否有OCR结果
        ocr_results = previous_context.get('ocr_results', {})
        if not ocr_results:
            return {
                "success": False,
                "step": "llm",
                "error": "需要先执行OCR步骤"
            }

        results = {}
        analyzer = EnhancedFinancialTableAnalyzer()

        for png_name in png_names:
            # 获取OCR结果
            ocr_result = ocr_results.get(png_name, {}).get('result')
            if not ocr_result:
                results[png_name] = {"success": False, "error": "没有OCR结果"}
                continue

            # 构建图片路径
            from backend.utils.constants import JOINED_TABLES_DIR
            image_path = JOINED_TABLES_DIR / pdf_folder / png_name

            # 执行LLM分析
            llm_result = analyzer.analyze_image(str(image_path), ocr_result)

            if llm_result.get('success'):
                results[png_name] = {
                    "success": True,
                    "result": llm_result,
                    "tables_count": llm_result['processing_stats']['visual_tables_count']
                }
            else:
                results[png_name] = {
                    "success": False,
                    "error": llm_result.get('error', 'LLM分析失败')
                }

        return {
            "success": True,
            "step": "llm",
            "results": results,
            "message": f"LLM分析完成，处理{len(png_names)}张图片"
        }

    except Exception as e:
        return {
            "success": False,
            "step": "llm",
            "error": str(e)
        }


def execute_reconstruct_step(pdf_folder, png_names, previous_context, output_dir):
    """执行表格重构步骤"""
    try:

        # 检查前置结果
        ocr_results = previous_context.get('ocr_results', {})
        llm_results = previous_context.get('llm_results', {})

        if not ocr_results or not llm_results:
            return {
                "success": False,
                "step": "reconstruct",
                "error": "需要先执行OCR和LLM步骤"
            }

        results = {}
        reconstructor = TableReconstructor()

        for png_name in png_names:
            # 获取前置结果
            ocr_result = ocr_results.get(png_name, {}).get('result')
            llm_result = llm_results.get(png_name, {}).get('result')

            if not ocr_result or not llm_result:
                results[png_name] = {"success": False, "error": "缺少前置结果"}
                continue

            # 生成输出文件路径
            from pathlib import Path
            output_path = Path(output_dir) / pdf_folder
            output_path.mkdir(parents=True, exist_ok=True)
            excel_file = str(output_path / f"{Path(png_name).stem}_reconstructed.xlsx")

            effect_png_dir = FILTERED_TABLES_DIR / pdf_folder / png_name
            print("effect_png_dir::::", effect_png_dir)
            # 执行表格重构
            success = reconstructor.process_all_tables(
                ocr_result=ocr_result,
                llm_result=llm_result,
                output_file=excel_file,
                final_output_file=excel_file,
                image_path=str(effect_png_dir),
                bank_name=""
            )

            if success:
                results[png_name] = {
                    "success": True,
                    "output_file": excel_file,
                    "message": "表格重构成功"
                }
            else:
                results[png_name] = {
                    "success": False,
                    "error": "表格重构失败"
                }

        return {
            "success": True,
            "step": "reconstruct",
            "results": results,
            "message": f"表格重构完成"
        }

    except Exception as e:
        return {
            "success": False,
            "step": "reconstruct",
            "error": str(e)
        }


def execute_export_step(pdf_folder, png_names, previous_context, output_dir):
    """执行数据导出步骤"""
    # 在实际中，重构步骤通常已经生成Excel文件
    # 所以导出步骤可能只是文件整理或格式转换
    reconstruct_results = previous_context.get('reconstruct_results', {})

    excel_files = []
    for png_name in png_names:
        result = reconstruct_results.get(png_name, {})
        if result.get('success') and result.get('output_file'):
            excel_files.append(result['output_file'])

    return {
        "success": True,
        "step": "export",
        "excel_files": excel_files,
        "message": f"导出完成，生成{len(excel_files)}个Excel文件"
    }

