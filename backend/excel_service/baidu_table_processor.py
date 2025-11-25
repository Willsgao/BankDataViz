import json
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment


class BaiduTableProcessor:
    """百度表格处理器，用于将百度OCR表格数据转换为Excel格式"""

    def __init__(self):
        self.wb = None
        self.current_sheet = None

    def process_baidu_table(self, bd_json, out_file='baidu_table.xlsx'):
        """
        处理百度OCR表格数据并生成Excel文件

        Args:
            bd_json: 百度OCR表格JSON数据
            out_file: 输出Excel文件路径

        Returns:
            str: 生成的Excel文件路径
        """
        try:
            self.wb = openpyxl.Workbook()

            for idx, tbl in enumerate(bd_json.get('tables_result', [])):
                self._process_single_table(idx, tbl)

            self.wb.save(out_file)
            print(f'已生成 {out_file}')
            return out_file

        except Exception as e:
            print(f"处理表格时出错: {e}")
            raise

    def _process_single_table(self, idx, table_data):
        """处理单个表格"""
        print(f"处理表格 {idx}")

        # 创建或获取工作表
        if idx == 0:
            self.current_sheet = self.wb.active
        else:
            self.current_sheet = self.wb.create_sheet(title=f'Table{idx + 1}')
        self.current_sheet.title = f'Table{idx + 1}'

        # 分析层次关系
        hierarchy_levels = self._analyze_vertical_hierarchy(table_data['body'])

        # 创建内容映射
        content_map = self._create_content_map(table_data['body'], hierarchy_levels)

        # 修复重复问题
        content_map = self._fix_duplicate_issues(content_map)

        # 写入Excel
        self._write_to_excel(content_map)

        # 设置列宽
        self._set_column_widths(table_data['body'])

    def _analyze_vertical_hierarchy(self, body_cells, thresh=20):
        """
        基于最左侧列的「左边界 x 坐标」计算缩进级别

        Args:
            body_cells: 表格体单元格数据
            thresh: 缩进阈值像素

        Returns:
            dict: {(row, col): level} 层级映射
        """
        if not body_cells:
            return {}

        min_col = min(c['col_start'] for c in body_cells)
        left_col = [c for c in body_cells if c['col_start'] == min_col]
        left_col.sort(key=lambda c: c['row_start'])

        lvl_map = {}
        base_x = min(c['cell_location'][0]['x'] for c in left_col)
        last_x, last_lvl = base_x, 0
        stack = []

        for c in left_col:
            cur_x = c['cell_location'][0]['x']
            # 右移 ≥ thresh 像素 → 加深一级
            if cur_x - last_x >= thresh:
                last_lvl += 1
                last_x = cur_x
            elif cur_x < last_x - thresh:  # 回退到外层
                last_lvl = max(last_lvl - 1, 0)
                last_x = cur_x
            # 维护栈
            stack = stack[:last_lvl]
            stack.append(c['words'].strip())
            lvl_map[(c['row_start'], min_col)] = last_lvl

        return lvl_map

    def _create_content_map(self, body_cells, hierarchy_levels):
        """创建内容映射字典"""
        content_map = {}
        for cell in body_cells:
            r_s, c_s, r_e, c_e = cell['row_start'], cell['col_start'], cell['row_end'], cell['col_end']
            words = cell['words']

            # 最左列加缩进
            if c_s == 0 and (r_s, c_s) in hierarchy_levels:
                level = hierarchy_levels[(r_s, c_s)]
                words = "  " * level + words

            # 为合并区域的所有位置写入内容
            for row in range(r_s, r_e + 1):
                for col in range(c_s, c_e + 1):
                    content_map[(row, col)] = words

        return content_map

    def _fix_duplicate_issues(self, content_map):
        """
        修复重复问题：
        1. 第一行重复
        2. 最后一列重复
        """
        if not content_map:
            return content_map

        # 创建副本以避免修改迭代中的字典
        fixed_map = content_map.copy()

        # 修复1：删除第一行的重复（第0行和第1行内容相同）
        rows_to_remove = []
        for (row, col), content in content_map.items():
            if row == 1:  # 检查第1行（Excel中的第2行）
                # 如果第0行和第1行在相同列有相同内容，标记第1行要删除
                if (0, col) in content_map and content_map[(0, col)] == content:
                    rows_to_remove.append((1, col))

        for pos in rows_to_remove:
            if pos in fixed_map:
                del fixed_map[pos]

        # 修复2：删除最后一列的重复
        max_col = max(col for row, col in fixed_map.keys())
        last_col_duplicates = []
        for (row, col), content in fixed_map.items():
            if col == max_col:
                # 检查前一列是否有相同内容
                if (row, max_col - 1) in fixed_map and fixed_map[(row, max_col - 1)] == content:
                    last_col_duplicates.append((row, max_col))

        for pos in last_col_duplicates:
            if pos in fixed_map:
                del fixed_map[pos]

        return fixed_map

    def _write_to_excel(self, content_map):
        """将内容映射写入Excel"""
        for (row, col), content in content_map.items():
            cell = self.current_sheet.cell(row=row + 1, column=col + 1, value=content)
            if col == 0:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def _set_column_widths(self, body_cells):
        """设置列宽"""
        if not body_cells:
            return

        max_col = max(c['col_end'] for c in body_cells) + 1

        for col in range(max_col):
            if col == 0:  # 纵向表头列设置较宽
                self.current_sheet.column_dimensions[get_column_letter(col + 1)].width = 30
            else:
                self.current_sheet.column_dimensions[get_column_letter(col + 1)].width = 15


# 使用示例
if __name__ == '__main__':
    try:
        # 实例化处理器
        processor = BaiduTableProcessor()

        # 读取数据
        with open('data1.json', 'r', encoding='utf-8') as f:
            json_data = json.load(f)

        # 处理表格
        save_file = "百度表格-简化层级版.xlsx"
        result_file = processor.process_baidu_table(json_data, save_file)
        print(f"处理完成，文件保存至: {result_file}")

    except Exception as e:
        print(f"程序执行出错: {e}")
        import traceback

        traceback.print_exc()