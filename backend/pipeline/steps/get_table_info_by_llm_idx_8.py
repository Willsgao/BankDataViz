## 2. 代码实现层


import logging
from typing import Dict, Any
import asyncio
import pandas as pd
import traceback
from openpyxl import load_workbook, Workbook


# 第一层：轻量级评估提示词 (约150 tokens)
ASSESSMENT_PROMPT = """
【角色】财务表格复杂度五级评估专家

# 核心任务
基于四维要素体系，精准评估财务表格复杂度等级，新增中等表格数据量细分维度


【前置判定】
1、判定整页 PDF 截图中是否存在银行财务表格，财务表格中必须要有表示财务金额或变化比率的数值列；
存在财务表格 → 输出“是”；否则输出“否”并结束。

# 四维评估要素（新增"数据量"子维度）
## 1. 横向维度复杂度
- 指标数量：单一指标 → 多指标混合
- 业务维度：无业务分类 → 多业务维度交叉
- 维度类型：纯数值 → 时间+业务+指标混合

## 2. 纵向层级深度
- 层级数量：无层级 → 多级嵌套层级
- 结构类型：平铺列表 → 树形层级结构
- 汇总深度：无汇总 → 多层交叉汇总

## 3. 结构复杂度
- 合并单元格：无合并 → 复杂交叉合并
- 汇总结构：无汇总 → 动态矩阵汇总
- 非结构化元素：无备注 → 特殊说明影响

## 4. 数据量规模（新增子维度）
- 单元格总数：行数×列数（含汇总行、所有列）
- 明细项占比：明细数据行占总行数比例


# 五级评估体系（Level 3 细分为两子类）

## Level 1：极简单表格
**核心特征**：单维度基础结构
- 横向：1列数据（单一指标）
- 纵向：无层级平铺列表
- 结构：无合并单元格，无汇总行
- 数据量：<30（行数×列数）
- 示例：现金日记账（日期、摘要、金额三列，10行×3列=30）

## Level 2：简单表格
**核心特征**：单一维度扩展
- 横向：≤2列同类指标（本期/上期）
- 纵向：单层层级（一级科目-明细科目，≤10个子项）
- 结构：无合并单元格，1级汇总（末尾合计）
- 数据量：30-50（行数×列数）
- 示例：管理费用明细表（8行明细+1行合计=9行，6列=54，因接近50仍属Level 2）

## Level 3：中等表格（细分为两子类）
### Level 3A（中等-紧凑型，数据量<80）
**核心特征**：双维度交叉+数据量较小
- 横向：2种业务维度（如"风险分类+阶段"，≤5列）
- 纵向：2级嵌套层级（大类-子类，子项≤10个）
- 结构：表头简单合并（横向维度标题合并），2级汇总（子项小计+大类合计，汇总行≤3行）
- 数据量：50-79（行数×列数）
- 明细项占比：≥60%（明细行>汇总行）
- 示例："贷款五级分类简表"（横向"正常类、关注类"2维度4列，纵向"公司贷款-制造业/服务业"2级层级8行明细+2行小计+1行合计=11行，11×7=77<80）

### Level 3B（中等-扩展型，数据量≥80）
**核心特征**：双维度交叉+数据量较大
- 横向：2-3种业务维度（如"风险分类+阶段+期限"，6-8列）
- 纵向：2级嵌套层级（大类-子类，子项11-20个）
- 结构：表头简单合并，2级汇总（子项小计+大类合计，汇总行4-5行）
- 数据量：80-120（行数×列数）
- 明细项占比：≥50%（明细行仍为主）
- 示例："贷款五级分类详表"（横向"正常/关注/次级"3维度6列，纵向"公司贷款-制造业/服务业/批发业"等15行明细+3行小计+1行合计=19行，19×6=114≥80）

## Level 4：复杂表格
**核心特征**：多维度混合深度结构
- 横向：≥3种业务维度（时间+业务+指标，≥9列）
- 纵向：3级及以上嵌套层级（大类-子类-明细项，子项>20个）
- 结构：跨行列合并单元格（纵向层级标题跨2行），多层交叉汇总（明细→小计→合计→总合计，汇总行≥6行）
- 数据量：>120（行数×列数）
- 示例：预期信用减值准备表（横向"第一/二/三阶段+合计"4维度8列，纵向3级层级25行明细+5行汇总=30行，30×8=240>120）

## Level 5：极复杂表格
**核心特征**：多维动态关联非结构化
- 横向：≥4种混合维度（时间+业务+指标+主体，≥12列）
- 纵向：≥4级深度嵌套层级（子项>30个）
- 结构：复杂嵌套合并（跨3行3列合并），动态矩阵汇总，含非结构化说明（备注列影响数据逻辑）
- 数据量：>200（行数×列数）
- 示例：集团合并金融资产减值表（30行×10列=300>200）


# 输出格式
<财务表格>是|否</财务表格>
<complexity>极简单|简单|中等-紧凑型|中等-扩展型|复杂|极复杂</complexity>
<complexity_reason>
```json
{
    "横向维度": {
        "级别": "2种业务维度",
        "特征": "风险分类+阶段（4列）",
        "得分": 3
    },
    "纵向层级": {
        "级别": "2级嵌套层级",
        "特征": "大类-子类（10个子项）",
        "得分": 3
    },
    "结构复杂度": {
        "级别": "2级汇总",
        "特征": "子项小计+大类合计（2行汇总）",
        "得分": 3
    },
    "数据量评估": {
        "行数×列数": "12行×6列=72",
        "数据量级别": "<80",
        "明细项占比": "75%（9行明细/12总行数）"
    },
    "综合评估": "Level 3A - 中等-紧凑型表格"
}
"""

# 第二层：简单表格处理提示词 (约200 tokens)
SIMPLE_PROMPT = """
【角色】简单表格快速提取器

# 处理规则
1. 直接提取所有单元格数据
2. 平级输出，不构建层级
3. 基础数值转换

# 输出格式
<simple_data>
```csv
序号|项目|数值列1|数值列2|报告期
1|项目A|1000|1500|2023-12-31
2|项目B|2000|2500|2023-12-31
3|合计|3000|4000|2023-12-31
</simple_data>
"""


# 第三层：标准表格处理提示词 (约400 tokens)
STANDARD_PROMPT = """
【角色】银行财报多维度汇总表格提取专家

# 核心任务
按要求准确识别给定的图片表格中的文本和数据，有效处理表格的纵向与横向多维度汇总关系，特别注意跨页拼接导致的表格错位和表头重复问题。

### 1. 维度解耦  
提取纵向科目树（含层级深度、汇总关系）、横向指标集（分类维度及具体值）、原始数据矩阵  


### 2. 结构建模  
{  
  "纵向科目树": {"类型": "树形结构", "示例路径": "资产/流动资产/货币资金", "最大深度": "自动识别"},  
  "横向指标集": {"类型": "多维度组合", "分类": ["时间维度", "业务维度", "指标类型"], "示例组合": "授信类别/正常类"}
}  

# 输出格式
<start1>表名（含报告期+核心指标，如"2024年末预期信用减值准备表"）</start1>

<start4>4、表格数据（多维汇总格式，11列）：
```csv
序号|主体|纵向层级路径|横向层级路径|数据类型|币种|单位|报告期|数值|汇总标记|维度类型
2|本集团|金融投资/评级分布/AA-到AA+|金融资产类别/交易性金融资产|发生额|人民币|元|2023-12|645174.43|0|明细
3|本集团|金融投资/评级分布/AA-到AA+|金融资产类别/债权投资|发生额|人民币|元|2023-12|3922154.72|0|明细
6|本集团|金融投资/评级分布/AA-到AA+|金融资产类别/合计|发生额|人民币|元|2023-12|148967891.20|1|横向汇总
```</start4>


【表格数据-字段规则】
表格数据的csv数据里，每一行必须要11列，10个“|”，不能多也不能少。
第1列 数据的序号
第2列 主体类型：本行 / 本集团；缺省值为'-'
第3列 纵向层级路径：反映表格左侧的行科目层级关系。
第4列 横向层级路径：反映表格上方的列指标层级关系；
第5列 数据类型：余额|发生额|本期增加|本期减少|占比|同比|环比|减值准备|其他（原文出现即照录）
第6列 币种：人民币|美元|港币|欧元|其他|-
第7列 单位：原文照录，可以为%，无填 “-”
第8列 报告期：YYYY-MM-DD 或 YYYY 年；无法判断填 “-”
第9列 数值：纯数字，无值填0。去掉千位符','：1,234->1234；有括号则变负号：“(1 234)”→“-1234”；有“%”需拆分数字和%："(1,234.56%)" → 单位：%，数值：-1234.56
第10列 数据汇总标记：字段名含“合计/小计/总计/Total/汇总”字样，或该科目在层级树中拥有直属子级 → 1；其余 → 0
第11列 维度类型：表示该数据属于“明细|横向汇总|纵向汇总|全局汇总”

# 处理规则

## 1. 路径构建规则
**纵向层级路径**：反映表格左侧的行科目层级关系
- 格式：{主科目}[/{维度1}/{具体分类1}][/{维度2}/{具体分类2}]...
- 示例：`金融投资/评级分布/AA-到AA+`

**横向层级路径**：反映表格上方的列指标层级关系  
- 格式：{主指标}[/{维度}/{具体分类}]
- 示例：`金融资产类别/交易性金融资产`

## 2. 汇总标记规则
- **明细数据**（最细粒度）：标记为0，维度类型=明细
- **纵向汇总**（某纵向路径的合计）：标记为1，维度类型=纵向汇总
- **横向汇总**（某横向路径的合计）：标记为1，维度类型=横向汇总  
- **全局汇总**（总计）：标记为1，维度类型=全局汇总
- **减值准备**：标记为0，维度类型=减值准备

## 3. 跨页拼接处理
- 识别并移除重复表头行
- 校正列对齐偏差
- 检查数据连续性，修复断裂序列

## 4. 数据完整性规则
- 禁止缺失字段，无法识别填"-"
- 单位必须统一
- 数值转换：去除千分位，括号转负号
- 表格数据的csv中“……|纵向层级路径|横向层级路径|数据类型|币种|单位|报告期|数值|……”这一行是标题行，必须要有
- **数据类型**取值要求：余额|发生额|本期增加|本期减少|占比|同比|环比|减值准备|其他（原文出现即照录）
-  数值列转换规则：
   - 千位符“,”删除：1,234->1234；
   - 括号变负号：“(1 234)”→“-1234”；
   - 数值中有“%”的，需拆分数字和%，并分别存储："(1,234.56%)" → 单位：%，数值：-1234.56；

## 5. 维度识别原则
- **纵向维度**：从左侧行标题识别（如"类型""期限"等）
- **横向维度**：从上方列标题识别（如"金融资产类别"各列, 横向维度不要出现报告期）
- 动态适应不同表格结构

# 完整性检查
✓ 所有数据点完整提取
✓ 汇总关系数学正确  
✓ 跨页问题已处理
✓ 字段完整性达标
✓ 数值转换准确
"""

STANDARD_PROMPT = """
【角色】银行财报多维度汇总表格提取专家  

# 核心任务  
按要求准确识别给定图片表格中的文本和数据，处理表格分区块结构、列边界对齐及遮挡干扰问题，输出结构化CSV数据。  


### 1. 表格结构预定义（优先于自动检测）  
#### 1.1 分区块表格处理（如资产负债表）  
- 若表格含“资产”“负债及所有者权益”双区块（左右分列），按以下规则拆分识别：  
  - **左侧区块（资产）**：含3列 → 行标题列（资产类科目）、2023年12月31日数值列、2022年12月31日数值列；  
  - **右侧区块（负债及所有者权益）**：含3列 → 行标题列（负债/权益类科目）、2023年12月31日数值列、2022年12月31日数值列；  
  - 两区块独立提取后合并，行标题列分别构建纵向层级路径，数值列按“报告期”对应，禁止跨区块合并单元格。  

#### 1.2 单列组表格处理（如利润表）  
- 非分区块表格默认含：行标题列 + N个数值列（N=列标题中“时间/指标”数量），按列标题顺序提取数值列，禁止遗漏/合并列。  


### 2. 核心提取规则（维度解耦+数据转换）  
#### 2.1 维度解耦  
- **纵向层级路径**：反映左侧行标题层级关系，格式：{父级科目}/{子级科目}/{明细科目}  
  - 规则：无数值的层级标题（如“金融资产：”）需作为父级保留，子科目路径必须包含父级（如“资产/金融资产/交易性金融资产”）；  
  - 示例：行标题“一、资产→1.金融资产→交易性金融资产” → 路径“资产/金融资产/交易性金融资产”。  

- **横向层级路径**：反映上方列标题层级关系，格式：{指标类型}/{时间维度}  
  - 规则：列标题含“时间（2023/2022）”“指标（余额/发生额）”时，按“指标类型/时间维度”组合（如“期末余额/2023年12月31日”）；  
  - 示例：列标题“2023年12月31日（期末余额）” → 路径“期末余额/2023年12月31日”。  

#### 2.2 数据转换规则  
- **数值处理**：  
  - 千位符“,”删除：“1,234.56”→“1234.56”；  
  - 括号转负号：“(567)”→“-567”；  
  - 百分比拆分：“3.2%”→ 单位“%”，数值“3.2”；  
  - 遮挡/无法识别：被印章/水印完全遮挡的单元格 → 数值“-”，单位/币种继承同列上方有效值。  

- **汇总标记**：  
  - 汇总行判定：行标题含“合计/小计/总计/Total”或为树形层级中的父级标题（有子项）→ 汇总标记“1”，维度类型“纵向汇总/横向汇总”；  
  - 明细行判定：非汇总行 → 汇总标记“0”，维度类型“明细”。  


### 3. 列边界与对齐校验（解决列丢失/错位）  
#### 3.1 列边界定位三规则（优先级从高到低）  
1. **垂直分隔线规则**：表格中黑色实线垂直分隔线为列边界（即使被轻微遮挡，按可见部分延伸定位）；  
2. **标题列宽度规则**：行标题列宽度=表格左侧行标题文本最大宽度（如“资产总计”占10字符宽度，则标题列宽10字符，右侧为数值列）；  
3. **数值列对齐规则**：数值列文本右对齐（金额/比率通常靠右），左对齐为标题列，以此区分标题列与数值列。  

#### 3.2 列对齐强制校验  
- **列标题数量=数值列数量**：提取表格所有列标题（如“2023年12月31日”“2022年12月31日”共2个），则数值列必须提取2列，缺失列填充“-”；  
- **行对齐校验**：若某行数值列数量<列标题数量，按“从左到右顺序对齐”，右侧缺失列填充“-”（如标题列2列，某行仅1列数值，则第2列填“-”）。  


### 4. 输出格式（严格按CSV字段规则）  
<start1>表名（含报告期+核心指标，如"2023年末资产负债表"）</start1>  

<start4>4、表格数据（多维汇总格式，11列）：  
```csv  
序号|主体|纵向层级路径|横向层级路径|数据类型|币种|单位|报告期|数值|汇总标记|维度类型  
1|本行|资产/现金及存放中央银行款项|期末余额/2023年12月31日|余额|人民币|元|2023-12-31|421238158071|0|明细  
2|本行|资产/现金及存放中央银行款项|期末余额/2022年12月31日|余额|人民币|元|2022-12-31|435945663917|0|明细  
6|本集团|金融投资/评级分布/AA-到AA+|金融资产类别/合计|发生额|人民币|元|2023-12|148967891.20|1|横向汇总
```
</start4>

【表格数据 - 字段规则（必须严格遵守）】

11 列字段，10 个 “|”，禁止增减列；
主体：“本行”/“本集团”，无则填 “-”；
报告期：数值列对应列标题中的时间（如 “2023 年 12 月 31 日”→“2023-12-31”）；
数值：无法识别（如遮挡）填 “-”，禁止空值或删除行 / 列；
维度类型：明细 / 横向汇总 / 纵向汇总 / 全局汇总（按汇总标记 + 层级关系判定）。

"""

STANDARD_PROMPT = """
【角色】银行财报多维度汇总表格提取专家  

# 核心任务  
按要求准确识别给定图片表格中的文本和数据，处理表格分区块结构、**行/列边界对齐**及遮挡干扰问题，输出结构化CSV数据。  


### 1. 表格结构预定义（优先于自动检测）  
#### 1.1 分区块表格处理（如资产负债表）  
- 若表格含“资产”“负债及所有者权益”双区块（左右分列），按以下规则拆分识别：  
  - **左侧区块（资产）**：含3列 → 行标题列（资产类科目）、2023年12月31日数值列、2022年12月31日数值列；  
  - **右侧区块（负债及所有者权益）**：含3列 → 行标题列（负债/权益类科目）、2023年12月31日数值列、2022年12月31日数值列；  
  - 两区块独立提取后合并，**行标题列需按文本块独立分行**（每个行标题对应一行，即使数值列空白），数值列按“报告期”对应，禁止跨区块合并单元格。  

#### 1.2 单列组表格处理（如利润表）  
- 非分区块表格默认含：行标题列 + N个数值列（N=列标题中“时间/指标”数量），**行标题列文本块需独立分行**，按列标题顺序提取数值列，禁止遗漏/合并列。  


### 2. 核心提取规则（维度解耦+数据转换）  
#### 2.1 维度解耦  
- **纵向层级路径**：反映左侧行标题层级关系，格式：{父级科目}/{子级科目}/{明细科目}  
  - 规则：无数值的层级标题（如“金融资产：”）需作为父级保留，子科目路径必须包含父级；**行标题文本块独立成行，即使数值列全空白也需保留行结构**。  
  - 示例：行标题“一、资产→1.金融资产→交易性金融资产” → 路径“资产/金融资产/交易性金融资产”。  

- **横向层级路径**：反映上方列标题层级关系，格式：{指标类型}/{时间维度}  
  - 规则：列标题含“时间（2023/2022）”“指标（余额/发生额）”时，按“指标类型/时间维度”组合；**数值列与行标题严格按行对应，禁止跨行列匹配数据**。  

#### 2.2 数据转换规则  
- **数值处理**：  
  - 千位符“,”删除：“1,234.56”→“1234.56”；  
  - 括号转负号：“(567)”→“-567”；  
  - 百分比拆分：“3.2%”→ 单位“%”，数值“3.2”；  
  - **遮挡/空白处理**：  
    - 被印章/水印完全遮挡的单元格 → 数值“-”，单位/币种继承同列上方有效值；  
    - **数值列空白（无数据）的单元格 → 强制填“-”，禁止从其他行/列调取数据填充（即使相邻行有数据）**。  

- **汇总标记**：  
  - 汇总行判定：行标题含“合计/小计/总计/Total”或为树形层级中的父级标题（有子项）→ 汇总标记“1”，维度类型“纵向汇总/横向汇总”；  
  - 明细行判定：非汇总行 → 汇总标记“0”，维度类型“明细”。  


### 3. 行/列边界与对齐校验（解决行合并+列错位）  
#### 3.1 列边界定位三规则（优先级从高到低）  
1. **垂直分隔线规则**：表格中黑色实线垂直分隔线为列边界（即使被轻微遮挡，按可见部分延伸定位）；  
2. **标题列宽度规则**：行标题列宽度=表格左侧行标题文本最大宽度（如“资产总计”占10字符宽度，则标题列宽10字符，右侧为数值列）；  
3. **数值列对齐规则**：数值列文本右对齐（金额/比率通常靠右），左对齐为标题列，以此区分标题列与数值列。  

#### 3.2 行边界保护三规则（新增，解决行合并/错位）  
1. **行标题文本块独立规则**：左侧行标题列中，每个独立文本块（如“现金及存放中央银行款项”“存放联行款项”）对应一行，**禁止因数值列空白/遮挡将多行标题合并为单行**；  
2. **水平分隔线延伸规则**：若水平分隔线被印章/干扰元素遮挡，按未遮挡部分的斜率延伸补全行边界（如从表格左侧未遮挡的分隔线端点向右延伸至右侧边界）；  
3. **行高一致性规则**：计算表格中无干扰行的平均行高（如“存放同业款项”“拆出资金”行），**若某行高度超过平均行高的1.5倍，强制按平均行高拆分为多行**，避免两行合并为一行。  

#### 3.3 行列对齐强制校验  
- **列对齐**：列标题数量=数值列数量，缺失列填充“-”；  
- **行对齐**：行标题数量=数值行行数，**每个行标题必须对应一行数值（即使数值全为“-”）**，禁止行合并导致的“一行多标题”或“多行一标题”；  
- **数据归属校验**：数值列数据的纵坐标范围必须落在对应行标题文本块的纵坐标范围内，禁止跨纵坐标范围分配数据（如下行数值归属到上行）。  


### 4. 输出格式（严格按CSV字段规则）  
<start1>表名（含报告期+核心指标，如"2023年末资产负债表"）</start1>  

<start4>4、表格数据（多维汇总格式，11列）：  
```csv  
序号|主体|纵向层级路径|横向层级路径|数据类型|币种|单位|报告期|数值|汇总标记|维度类型  
1|本行|资产/现金及存放中央银行款项|期末余额/2023年12月31日|余额|人民币|元|2023-12-31|421238158071|0|明细  
2|本行|资产/现金及存放中央银行款项|期末余额/2022年12月31日|余额|人民币|元|2022-12-31|435945663917|0|明细  
3|本行|资产/存放联行款项|期末余额/2023年12月31日|余额|人民币|元|2023-12-31|-|0|明细  
4|本行|资产/存放联行款项|期末余额/2022年12月31日|余额|人民币|元|2022-12-31|20875712.48|0|明细  
6|本集团|金融投资/评级分布/AA-到AA+|金融资产类别/合计|发生额|人民币|元|2023-12|148967891.20|1|横向汇总
```
</start4>

【表格数据 - 字段规则（必须严格遵守）】
- 11 列字段，10 个 “|”，禁止增减列；
- 数值列空白 / 遮挡必须填 “-”，禁止跨行列填充数据；
- 主体：“本行”/“本集团”，无则填 “-”；
- 报告期：数值列对应列标题中的时间（如 “2023 年 12 月 31 日”→“2023-12-31”）；
- 维度类型：明细 / 横向汇总 / 纵向汇总 / 全局汇总（按汇总标记 + 层级关系判定）。

"""

COMPLEX_PROMPT = """
【角色】专业财务表格解析引擎，采用三维解析框架精准提取复杂表格数据

【解析框架：三维维度分解】
## 维度一：纵向科目维度（Y-axis）  
财务科目层级树（如：资产→流动资产→货币资金），含父子层级关系  

## 维度二：横向指标维度（X-axis）  
多维度指标组合，包含：  
- 时间维度（如2023年、2024年）  
- 业务维度（如正常类、关注类）  
- 指标类型（如余额、比例、增长率）  

## 维度三：数据值维度（Z-axis）  
具体数值+元数据（单位、币种、数据类型）  


【解析流程】
### 1. 维度解耦  
提取纵向科目树（含层级深度、汇总关系）、横向指标集（分类维度及具体值）、原始数据矩阵  


### 2. 结构建模  
{  
  "纵向科目树": {"类型": "树形结构", "示例路径": "资产/流动资产/货币资金", "最大深度": "自动识别"},  
  "横向指标集": {"类型": "多维度组合", "分类": ["时间维度", "业务维度", "指标类型"], "示例组合": "授信类别/正常类"},  
  "数据矩阵": {"维度": ["row_id", "col_id", "value"], "元数据": ["单位", "币种", "数据类型"]}  
}  

### 3. 数据网格生成  
基于纵向科目×横向指标的笛卡尔积生成完整数据点位，规则：  
1. 每个点位分配唯一坐标：(纵向路径, 横向路径)  
2. 提取对应数值，空值填"0"  
3. 验证网格完整性（预期单元格数=实际提取数）  

### 4. 智能映射规则  
- 汇总行识别：含"合计/小计/总计"→标记汇总标记=1，建立父子关系  
- 层级识别：缩进增加→层级深度+1（如父项缩进2字符→子项缩进4字符）  
- 维度分类：含"YYYY年"→时间维度；含"%"→比例类型，分离数值与单位  
- 表格数据的csv中“……|纵向层级路径|横向层级路径|数值|单位|报告期|……”这一行是标题行，必须要有
- **数据类型**取值要求：余额|发生额|本期增加|本期减少|占比|同比|环比|减值准备|其他（原文出现即照录）
- 完整数据网格中的数据要完整输出，不能省略；
-  数值列转换规则：
   - 千位符“,”删除：1,234->1234；
   - 括号变负号：“(1 234)”→“-1234”；
   - 数值中有“%”的，需拆分数字和%，并分别存储："(1,234.56%)" → 单位：%，数值：-1234.56；


### 5. 完整性验证  
- 网格完整性：预期单元格数=实际提取数  
- 数值一致性：汇总值≈∑明细值（误差≤0.01元或0.01%）  
- 层级合理性：子项缩进>父项缩进  
- 维度正交性：同一维度内元素互斥（如时间维度不含业务分类）  


【输出格式】
<start1>表名（含报告期+核心指标，如"2024年末预期信用减值准备表"）</start1>

<start2>{"纵向维度":{"类型":"科目层级树","根节点":"自动识别","最大深度":"自动识别","节点总数":"自动识别"},"横向维度":{"时间指标":["自动提取"],"业务指标":["自动提取"],"计量指标":["自动提取"]}}</start2>


<start3>3、表格数据（多维汇总格式，11列）：
```csv
序号|主体|纵向层级路径|横向层级路径|数据类型|币种|单位|报告期|数值|汇总标记|维度类型
2|本集团|金融投资/评级分布/AA-到AA+|金融资产类别/交易性金融资产|余额|人民币|元|2023-12-31|64585174.43|0|明细
3|本集团|金融投资/评级分布/AA-到AA+|金融资产类别/-|余额|人民币|元|2023-12-31|392822154.72|0|明细
4|本集团|金融投资/评级分布/AA-到AA+|金融资产类别/其他债权投资|发生额|人民币|元|2023-12-31|14256841827.22|0|明细
5|本集团|金融投资/评级分布/AA-到AA+|金融资产类别/-|占比|人民币|%|2023-12-31|2.83|0|明细
6|本集团|金融投资/评级分布/AA-到AA+|金融资产类别/合计|余额|人民币|元|2023-12-31|14892567891.20|1|横向汇总
```</start3>


<start4>{"提取完整性":"100% (预期/实际)","数值一致性":"通过","层级合理性":"通过","维度正交性":"通过"}</start4>

【表格数据-字段规则】
表格数据的csv数据里，每一行必须要11列，10个“|”，不能多也不能少。
第1列 数据的序号
第2列 主体类型：本行 / 本集团；缺省值为'-'
第3列 纵向层级路径：反映表格左侧的行科目层级关系；路径中的序号数字可省略。
第4列 横向层级路径：反映表格上方的列指标层级关系；路径中的序号数字可省略；
第5列 数据类型：余额|发生额|本期增加|本期减少|占比|同比|环比|减值准备|其他（原文出现即照录）
第6列 币种：人民币|美元|港币|欧元|其他|-
第7列 单位：原文照录，可以为%，无填 “-”
第8列 报告期：YYYY-MM-DD 或 YYYY 年；无法判断填 “-”
第9列 数值：纯数字，无值填0。去掉千位符','：1,234->1234；有括号则变负号：“(1 234)”→“-1234”；有“%”需拆分数字和%："(1,234.56%)" → 单位：%，数值：-1234.56
第10列 数据汇总标记：字段名含“合计/小计/总计/Total/汇总”字样，或该科目在层级树中拥有直属子级 → 1；其余 → 0
第11列 维度类型：表示该数据属于“明细|横向汇总|纵向汇总|全局汇总”


##  汇总标记规则
- **明细数据**（最细粒度）：标记为0，维度类型=明细
- **纵向汇总**（某纵向路径的合计）：标记为1，维度类型=纵向汇总
- **横向汇总**（某横向路径的合计）：标记为1，维度类型=横向汇总  
- **全局汇总**（总计）：标记为1，维度类型=全局汇总
- **减值准备**：标记为0，维度类型=减值准备


## 数据完整性规则
- 禁止缺失字段，无法识别填"-"
- 单位必须统一
- 数值转换：去除千分位，括号转负号
-  数值列转换规则：
   - 千位符“,”删除：1,234->1234；
   - 括号变负号：“(1 234)”→“-1234”；
   - 数值中有“%”的，需拆分数字和%，并分别存储："(1,234.56%)" → 单位：%，数值：-1234.56；
"""

COMPLEX_PROMPT = """
【角色】专业财务表格解析引擎，采用三维解析框架精准提取复杂表格数据，**强化行/列边界抗干扰能力**  

【解析框架：三维维度分解】  
## 维度一：纵向科目维度（Y-axis）  
财务科目层级树（如：资产→流动资产→货币资金），含父子层级关系；**行标题文本块独立成行，禁止因数值列空白/遮挡合并多行**  

## 维度二：横向指标维度（X-axis）  
多维度指标组合，包含：  
- 时间维度（如2023年、2024年）  
- 业务维度（如正常类、关注类）  
- 指标类型（如余额、比例、增长率）；**列标题与数值列严格对齐，禁止跨列合并**  

## 维度三：数据值维度（Z-axis）  
具体数值+元数据（单位、币种、数据类型）；**空白/遮挡单元格明确标记，禁止跨行列填充数据**  


【表格结构预定义（新增，优先于自动检测）】  
#### 1. 分区块表格处理（如资产负债表）  
- 若表格含“资产”“负债及所有者权益”等多区块（左右/上下分列），按区块独立提取：  
  - 每个区块含“行标题列+N个数值列”（N=区块内列标题数量）；  
  - 独立提取后合并，**行标题按文本块独立分行，数值列按“横向指标维度”对应，禁止跨区块合并单元格**。  

#### 2. 单列组表格处理（如利润表）  
- 非分区块表格默认含：行标题列 + N个数值列（N=横向指标维度数量），**行标题文本块独立成行，按列标题顺序提取数值列，禁止遗漏/合并列**。  


【解析流程】  
### 1. 维度解耦  
提取纵向科目树（含层级深度、汇总关系）、横向指标集（分类维度及具体值）、原始数据矩阵；**行标题列文本块需独立分行，即使数值列全空白也保留行结构**。  


### 2. 结构建模  
{  
  "纵向科目树": {"类型": "树形结构", "示例路径": "资产/流动资产/货币资金", "最大深度": "自动识别", "分行规则": "每个行标题文本块对应一行"},  
  "横向指标集": {"类型": "多维度组合", "分类": ["时间维度", "业务维度", "指标类型"], "示例组合": "授信类别/正常类", "对齐规则": "列标题数量=数值列数量"},  
  "数据矩阵": {"维度": ["row_id", "col_id", "value"], "元数据": ["单位", "币种", "数据类型"], "空值标记": "-"}  
}  

### 3. 数据网格生成  
基于纵向科目×横向指标的笛卡尔积生成完整数据点位，规则：  
1. 每个点位分配唯一坐标：(纵向路径, 横向路径)  
2. **提取对应数值，空值/遮挡填"-"（区分于数值0）**：  
   - 空白单元格（无数据）→ "-"  
   - 被印章/水印遮挡 → "-"（元数据继承同列上方有效值）  
   - 明确数值0 → "0"  
3. 验证网格完整性（预期单元格数=实际提取数，含"-"占位）  

### 4. 智能映射规则  
- 汇总行识别：含"合计/小计/总计/Total/汇总"→标记汇总标记=1，建立父子关系  
- 层级识别：缩进增加→层级深度+1（如父项缩进2字符→子项缩进4字符）  
- 维度分类：含"YYYY年"→时间维度；含"%"→比例类型，分离数值与单位  
- **行边界保护**（新增）：  
  - 行标题文本块独立规则：每个独立行标题（如“现金及存放中央银行款项”）对应一行，禁止因数值列空白合并多行；  
  - 水平分隔线延伸规则：遮挡的水平分隔线按未遮挡部分斜率延伸补全；  
  - 行高一致性规则：异常高行（>平均行高1.5倍）强制拆分多行  
- **列边界定位**（新增）：  
  - 垂直分隔线优先：实线分隔线为列边界，遮挡部分延伸定位；  
  - 标题列宽度规则：行标题列宽度=最长行标题文本宽度，右侧为数值列；  
  - 数值列对齐规则：数值列右对齐，左对齐为标题列  


### 5. 完整性验证  
- 网格完整性：预期单元格数=实际提取数（含"-"占位）  
- 数值一致性：汇总值≈∑明细值（误差≤0.01元或0.01%）  
- 层级合理性：子项缩进>父项缩进，路径包含完整层级  
- 维度正交性：同一维度内元素互斥（如时间维度不含业务分类）  
- **行列对齐验证**（新增）：  
  - 列对齐：横向指标维度数量=数值列数量，缺失列填"-"；  
  - 行对齐：纵向科目数量=数值行行数，每个行标题对应一行数据（即使全为"-"）  


【输出格式】  
<start1>表名（含报告期+核心指标，如"2023年末资产负债表"）</start1>  

<start2>{"纵向维度":{"类型":"科目层级树","根节点":"自动识别","最大深度":"自动识别","行标题数":"X行"},"横向维度":{"时间指标":["自动提取"],"业务指标":["自动提取"],"计量指标":["自动提取"],"列标题数":"Y列"}}</start2>  


<start3>3、表格数据（多维汇总格式，11列）：  
```csv  
序号|主体|纵向层级路径|横向层级路径|数据类型|币种|单位|报告期|数值|汇总标记|维度类型  
1|本行|资产/现金及存放中央银行款项|期末余额/2023年12月31日|余额|人民币|元|2023-12-31|421238158071|0|明细  
2|本行|资产/现金及存放中央银行款项|期末余额/2022年12月31日|余额|人民币|元|2022-12-31|435945663917|0|明细  
3|本行|资产/存放联行款项|期末余额/2023年12月31日|余额|人民币|元|2023-12-31|-|0|明细  // 空白数值列填"-"  
4|本行|资产/存放联行款项|期末余额/2022年12月31日|余额|人民币|元|2022-12-31|20875712.48|0|明细  
5|本行|资产/金融资产/交易性金融资产|期末余额/2023年12月31日|余额|人民币|元|2023-12-31|2994610206.80|0|明细  
6|本行|资产/资产总计|期末余额/2023年12月31日|余额|人民币|元|2023-12-31|88495635648.41|1|全局汇总
```
</start3>
<start4>{"提取完整性":"100% (预期 X×Y / 实际 X×Y)","数值一致性":"通过","层级合理性":"通过","维度正交性":"通过","行边界识别":"通过","列边界识别":"通过"}</start4>

【表格数据 - 字段规则（严格遵守）】
表格数据的 csv 数据里，每一行必须要 11 列，10 个 “|”，不能多也不能少。
第 1 列 数据的序号
第 2 列 主体类型：本行 / 本集团；缺省值为 "-"
第 3 列 纵向层级路径：反映表格左侧行科目层级关系，行标题文本块独立成行，路径包含完整父子层级；路径中的序号数字可省略。
第 4 列 横向层级路径：反映表格上方列指标层级关系，列标题与数值列严格对齐；路径中的序号数字可省略；
第 5 列 数据类型：余额 | 发生额 | 本期增加 | 本期减少 | 占比 | 同比 | 环比 | 减值准备 | 其他（原文出现即照录）
第 6 列 币种：人民币 | 美元 | 港币 | 欧元 | 其他 |-
第 7 列 单位：原文照录，可以为 %，无填 "-"
第 8 列 报告期：YYYY-MM-DD 或 YYYY 年；无法判断填 "-"
第 9 列 数值：纯数字，无数据 / 遮挡填 "-"，数值为 0 填 "0"；处理规则：

千位符 "," 删除：1,234→1234；
括号变负号：“(1 234)”→“-1234”；
百分比拆分："(1,234.56%)"→单位：%，数值：-1234.56
第 10 列 数据汇总标记：字段名含 “合计 / 小计 / 总计 / Total / 汇总” 或为层级树父节点→1；其余→0
第 11 列 维度类型：明细 | 横向汇总 | 纵向汇总 | 全局汇总

【汇总标记与维度类型规则】
明细数据（最细粒度）：标记 0，维度类型 = 明细
纵向汇总（某纵向路径合计）：标记 1，维度类型 = 纵向汇总
横向汇总（某横向路径合计）：标记 1，维度类型 = 横向汇总
全局汇总（表格总合计）：标记 1，维度类型 = 全局汇总
减值准备：标记 0，维度类型 = 减值准备

【干扰处理规则】
印章 / 水印遮挡：被遮挡单元格数值填 "-"，单位 / 币种继承同列上方有效值；
行合并预防：每个行标题文本块独立成行，禁止因数值列空白 / 遮挡合并多行；
列错位预防：垂直分隔线延伸定位列边界，确保列标题数量 = 数值列数量。
"""


import pandas as pd

import os
import re
import time
import csv
import io
from io import StringIO
import base64
from pathlib import Path
from openai import OpenAI
from PIL import Image
import logging

from io import BytesIO
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage   # 避免与PIL同名冲突
from PIL import Image as PILImage                     # 用于 bytes→图片对象


ARK_API_KEY = "90b9c47f-815c-4216-913a-3d1a567e35ac"

# client = OpenAI(
#     base_url="https://ark.cn-beijing.volces.com/api/v3",
#     api_key=ARK_API_KEY
# )

from openai import AsyncOpenAI
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

import os
import errno
from openpyxl import Workbook, load_workbook
from openpyxl.utils import quote_sheetname
from openpyxl.styles import Font
from openpyxl.worksheet.hyperlink import Hyperlink


client = AsyncOpenAI(
    base_url="https://ark.cn-beijing.volces.com/api/v3",
    api_key=ARK_API_KEY
)


class SmartTableProcessor:
    def __init__(self, llm_client, model_id):
        self.llm_client = llm_client
        self.model_id=model_id
        self.prompt_registry = {
            "assessment": ASSESSMENT_PROMPT,
            # "simple": SIMPLE_PROMPT,
            "simple": STANDARD_PROMPT,
            "standard": STANDARD_PROMPT,
            "complex": COMPLEX_PROMPT
        }


    def encode_image(self, image_path: str) -> tuple[str, int]:
        """
        返回：(base64编码字符串, 像素总数)
        不压缩原始图片，支持PNG和JPEG等格式
        """
        # 打开图片，不转换颜色模式以保留原始信息
        img = Image.open(image_path)
        w, h = img.size
        pixel_count = w * h  # 计算像素总数

        # 不进行缩放，保留原始尺寸
        buffer = io.BytesIO()

        # 获取图片原始格式，如果无法获取则默认使用PNG
        img_format = img.format or "PNG"
        # 统一转为大写格式标识
        img_format = img_format.upper()

        # 根据图片格式保存，使用最佳质量参数
        if img_format == "JPEG":
            img.save(buffer, format=img_format, quality=100, optimize=True)
        else:  # 对PNG等其他格式不进行压缩
            img.save(buffer, format=img_format)

        # 编码为base64字符串
        image_b64 = base64.b64encode(buffer.getvalue()).decode("utf-8")

        return image_b64, pixel_count


    def save_df_to_sheet_2(self,
                         df: pd.DataFrame,
                         excel_path: str,
                         sheet_name: str,
                         map_name: str,  # 新增：表格业务名称
                         mode: str = 'overwrite',
                         image_data=None,
                         anchor_cell: str = 'P2',
                         width_px: int = 768):
        """
        将 DataFrame 写入指定 Excel 的指定 Sheet，并可同时插入图片；
        同时维护一张“目录”工作表，记录 sheet_name ↔ table_name 映射，目录始终位于第一页。
        """
        # ---------- 1. 确保文件存在 ----------
        if not os.path.exists(excel_path):
            df.to_excel(excel_path, sheet_name=sheet_name, index=False)
            wb = load_workbook(excel_path)
        else:
            wb = load_workbook(excel_path)

        # ---------- 2. 写目标 sheet ----------
        if mode == 'overwrite':
            if sheet_name in wb.sheetnames:
                wb.remove(wb[sheet_name])
            with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a',
                                if_sheet_exists='replace') as w:
                df.to_excel(w, sheet_name=sheet_name, index=False)
        else:  # append
            with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a',
                                if_sheet_exists='overlay') as w:
                start_row = w.sheets[sheet_name].max_row if sheet_name in w.sheets else 0
                df.to_excel(w, sheet_name=sheet_name, index=False,
                            header=not bool(start_row), startrow=start_row)

        # ---------- 3. 维护“目录”sheet ----------
        wb = load_workbook(excel_path)  # 重新打开，保证拿到最新结构
        catalog_title = '目录'
        # 如果不存在就创建
        if catalog_title not in wb.sheetnames:
            catalog_ws = wb.create_sheet(catalog_title, 0)  # 0 表示插入到最前
            catalog_ws.append(['sheet_name', 'table_name'])  # 表头
        else:
            catalog_ws = wb[catalog_title]

        # 查找是否已有该行
        for row in catalog_ws.iter_rows(min_row=2, max_col=2, values_only=False):
            if row[0].value == sheet_name:
                row[1].value = map_name  # 更新
                break
        else:  # 未找到则追加
            catalog_ws.append([sheet_name, map_name])

        # 确保目录在第一页（保险做法）
        wb.move_sheet(catalog_title, offset=-len(wb.sheetnames))

        # ---------- 4. 插入图片（同之前逻辑） ----------
        if image_data is not None:
            ws = wb[sheet_name]
            # 统一转 PIL.Image
            if isinstance(image_data, bytes):
                pil_img = PILImage.open(BytesIO(image_data))
            elif isinstance(image_data, str) and os.path.isfile(image_data):
                pil_img = PILImage.open(image_data)
            elif isinstance(image_data, PILImage.Image):
                pil_img = image_data
            else:
                raise TypeError('image_data 必须是文件路径、bytes 或 PIL.Image')

            if width_px:
                ratio = pil_img.height / pil_img.width
                pil_img = pil_img.resize((width_px, int(width_px * ratio)), PILImage.LANCZOS)

            tmp = BytesIO()
            pil_img.save(tmp, format='PNG')
            tmp.seek(0)
            xl_img = XLImage(tmp)
            xl_img.anchor = anchor_cell
            ws.add_image(xl_img)

        wb.save(excel_path)
        print(f'已按 "{mode}" 模式写入 "{sheet_name}" 表 → {excel_path}')
        if image_data:
            print(f'  并插入图片 @ {anchor_cell}')
        print(f'  目录已更新：{sheet_name} -> {map_name}')

    def save_df_to_sheet_3(self, df, excel_path, sheet_name, map_name,
                         mode='overwrite', image_data=None, anchor_cell='P2', width_px=768):
        # 1. 统一在一次打开里完成所有事
        if os.path.exists(excel_path):
            wb = load_workbook(excel_path)
        else:
            wb = Workbook()
            wb.remove(wb.active)  # 去掉默认空白 sheet

        # 2. 写/覆盖数据 sheet
        if mode == 'overwrite' and sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
        with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a',
                            if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        # 3. 重新打开（writer 已关闭），继续改目录、插图
        wb = load_workbook(excel_path)

        # ---------- 目录：用内部超链接（优化WPS兼容性） ----------
        from openpyxl.utils import quote_sheetname
        from openpyxl.styles import Font
        from openpyxl.worksheet.hyperlink import Hyperlink

        catalog = '目录'
        if catalog not in wb.sheetnames:
            cat_ws = wb.create_sheet(catalog, 0)
            cat_ws.append(['sheet_name', 'table_name'])
        else:
            cat_ws = wb[catalog]

        # 优化：为WPS创建兼容的超链接格式
        # 1. 生成内部地址：使用更明确的格式
        quoted_sheet = quote_sheetname(sheet_name)
        # 对于WPS，有时需要在地址前添加#号
        loc = f"#{quoted_sheet}!A1"

        # 查找并更新现有行
        updated = False
        for r in cat_ws.iter_rows(min_row=2, max_col=2, values_only=False):
            if r[0].value == sheet_name:
                r[1].value = map_name
                # 2. 优化：直接设置单元格的hyperlink属性和值
                r[1].hyperlink = Hyperlink(ref=r[1].coordinate, location=loc)
                r[1].font = Font(underline='single', color='0563C1')
                updated = True
                break

        # 如果没找到则添加新行
        if not updated:
            new_row = cat_ws.max_row + 1
            cat_ws.cell(row=new_row, column=1, value=sheet_name)
            # 为WPS优化的超链接设置方式
            cell = cat_ws.cell(row=new_row, column=2)
            cell.value = map_name
            cell.hyperlink = Hyperlink(ref=cell.coordinate, location=loc)
            cell.font = Font(underline='single', color='0563C1')

        wb.move_sheet(catalog, offset=-len(wb.sheetnames))

        # ---------- 插入图片（保持原逻辑） ----------
        if image_data is not None:
            try:
                import PIL.Image as PILImage
                from io import BytesIO
                from openpyxl.drawing.image import Image as XLImage
            except ImportError as e:
                print(f"插入图片失败：缺少依赖库 - {e}")
                return

            ws = wb[sheet_name]
            if isinstance(image_data, bytes):
                pil_img = PILImage.open(BytesIO(image_data))
            elif isinstance(image_data, str) and os.path.isfile(image_data):
                pil_img = PILImage.open(image_data)
            elif isinstance(image_data, PILImage.Image):
                pil_img = image_data
            else:
                raise TypeError('image_data 必须是文件路径、bytes 或 PIL.Image')

            if width_px:
                ratio = pil_img.height / pil_img.width
                pil_img = pil_img.resize((width_px, int(width_px * ratio)), PILImage.LANCZOS)

            tmp = BytesIO()
            pil_img.save(tmp, format='PNG')
            tmp.seek(0)
            xl_img = XLImage(tmp)
            xl_img.anchor = anchor_cell
            ws.add_image(xl_img)

        wb.save(excel_path)
        print(f'已按 "{mode}" 模式写入 "{sheet_name}" 表 → {excel_path}')
        if image_data:
            print(f'  并插入图片 @ {anchor_cell}')
        print(f'  目录已更新：{sheet_name} -> {map_name}（内部超链接，WPS 可跳）')

    def save_df_to_sheet_4(self, df, excel_path, sheet_name, map_name,
                         mode='overwrite', image_data=None, anchor_cell='P2', width_px=768):
        file_exists = os.path.exists(excel_path)
        # 统一使用 openpyxl 来创建和操作工作簿
        if not file_exists:
            wb = Workbook()
            wb.remove(wb.active)
        else:
            wb = load_workbook(excel_path)

        # 处理工作表（覆盖或新增）
        if sheet_name in wb.sheetnames:
            if mode == 'overwrite':
                del wb[sheet_name]
        ws = wb.create_sheet(sheet_name)
        # 将 DataFrame 数据写入工作表
        for r_idx, row in enumerate(df.itertuples(index=False), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # 处理目录（与之前逻辑类似）
        catalog = '目录'
        if catalog not in wb.sheetnames:
            cat_ws = wb.create_sheet(catalog, 0)
            cat_ws.append(['sheet_name', 'table_name'])
        else:
            cat_ws = wb[catalog]
        # 目录超链接等逻辑（省略，与原代码一致）

        # 插入图片（与之前逻辑类似）
        if image_data is not None:
            try:
                import PIL.Image as PILImage
                from io import BytesIO
                from openpyxl.drawing.image import Image as XLImage
            except ImportError as e:
                print(f"插入图片失败：缺少依赖库 - {e}")
                return
            # 图片处理逻辑（省略，与原代码一致）

        # 最后统一保存工作簿
        wb.save(excel_path)
        print(f'已按 "{mode}" 模式写入 "{sheet_name}" 表 → {excel_path}')



    def save_df_to_sheet(self, df, excel_path, sheet_name, map_name,
                         mode='overwrite', image_data=None, anchor_cell='R2', width_px=768):
        # 确保文件路径为.xlsx格式
        if not excel_path.endswith('.xlsx'):
            raise ValueError("文件路径必须以.xlsx结尾（openpyxl仅支持xlsx格式）")

        # 1. 加载或创建工作簿（关键修复：统一用openpyxl处理，避免格式冲突）
        try:
            if os.path.exists(excel_path) and os.path.getsize(excel_path) > 0:
                # 只加载有效的非空文件
                wb = load_workbook(excel_path)
            else:
                # 创建新工作簿时保留默认工作表（避免空文件结构错误）
                wb = Workbook()
                # 仅当需要时删除默认表（确保至少有一个工作表）
                if len(wb.sheetnames) > 0:
                    default_sheet = wb.active.title
                    if default_sheet != sheet_name and default_sheet != '目录':
                        del wb[default_sheet]
        except Exception as e:
            # 处理损坏文件：删除并重创建
            print(f"文件损坏或格式错误，将创建新文件：{e}")
            wb = Workbook()
            if len(wb.sheetnames) > 0:
                del wb[wb.active.title]

        # 2. 处理工作表（覆盖或新增）
        if mode == 'overwrite' and sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name)  # 新增工作表

        # 3. 写入DataFrame数据（用openpyxl直接写入，避免pandas的ExcelWriter冲突）
        # 写入表头
        for col_idx, col_name in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)
        # 写入数据行
        for row_idx, row_data in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # 4. 处理目录和超链接（保持原逻辑）
        catalog = '目录'
        if catalog not in wb.sheetnames:
            cat_ws = wb.create_sheet(catalog, 0)  # 目录放最前面
            cat_ws.append(['sheet_name', 'table_name'])
        else:
            cat_ws = wb[catalog]

        # 生成超链接（兼容WPS）
        quoted_sheet = quote_sheetname(sheet_name)
        loc = f"#{quoted_sheet}!A1"  # 链接到工作表的A1单元格

        # 更新或新增目录行
        updated = False
        for row in cat_ws.iter_rows(min_row=2, max_col=2, values_only=False):
            if row[0].value == sheet_name:
                row[1].value = map_name
                row[1].hyperlink = Hyperlink(ref=row[1].coordinate, location=loc)
                row[1].font = Font(underline='single', color='0563C1')
                updated = True
                break
        if not updated:
            new_row = cat_ws.max_row + 1
            cat_ws.cell(row=new_row, column=1, value=sheet_name)
            cell = cat_ws.cell(row=new_row, column=2)
            cell.value = map_name
            cell.hyperlink = Hyperlink(ref=cell.coordinate, location=loc)
            cell.font = Font(underline='single', color='0563C1')

        # 5. 插入图片（保持原逻辑）
        if image_data is not None:
            try:
                import PIL.Image as PILImage
                from io import BytesIO
                from openpyxl.drawing.image import Image as XLImage
            except ImportError as e:
                print(f"插入图片失败：缺少依赖 - {e}")
                return

            try:
                # 处理不同类型的image_data
                if isinstance(image_data, bytes):
                    pil_img = PILImage.open(BytesIO(image_data))
                elif isinstance(image_data, str) and os.path.isfile(image_data):
                    pil_img = PILImage.open(image_data)
                elif isinstance(image_data, PILImage.Image):
                    pil_img = image_data
                else:
                    raise TypeError("image_data必须是文件路径、bytes或PIL.Image对象")

                # 调整图片尺寸
                if width_px:
                    ratio = pil_img.height / pil_img.width
                    pil_img = pil_img.resize((width_px, int(width_px * ratio)), PILImage.LANCZOS)

                # 保存到临时流并插入Excel
                tmp = BytesIO()
                pil_img.save(tmp, format='PNG')
                tmp.seek(0)
                xl_img = XLImage(tmp)
                xl_img.anchor = anchor_cell
                ws.add_image(xl_img)
            except Exception as e:
                print(f"图片处理失败：{e}")

        # 6. 保存工作簿（关键：确保只保存一次，避免结构损坏）
        try:
            wb.save(excel_path)
            print(f'已按"{mode}"模式写入"{sheet_name}"表 → {excel_path}')
            if image_data:
                print(f'  并插入图片 @ {anchor_cell}')
            print(f'  目录已更新：{sheet_name} -> {map_name}（内部超链接，WPS可跳）')
        except Exception as e:
            print(f"保存文件失败：{e}")


    def fallback_df(self, exc: Exception) -> pd.DataFrame:
        """
        把异常对象转成 1 行 1 列的 DataFrame，列名固定为 error_msg，
        内容就是完整 traceback，方便后续写 Excel 留存。
        """
        tb_str = traceback.format_exception(type(exc), exc, exc.__traceback__)
        # 合并成一整段字符串
        error_text = ''.join(tb_str)
        return pd.DataFrame({'error_msg': [error_text]})


    async def process_table(self, image_name: str, out_file:str, sheet_name: str, bank_name: str, file_name: str) -> Dict[str, Any]:
        """
        智能处理表格的主流程
        """
        try:

            image_data, img_size = self.encode_image(image_name)
            print("img_size:, img_size", img_size)
            t0 = time.time()

            # 第一步：复杂度评估
            complexity_result = await self._assess_complexity(image_data)
            complexity_name = complexity_result["complexity"]
            print("complexity_result:", complexity_result)
            t1 = time.time()
            print("************>:", t1 - t0)

            if complexity_result.get("complexity") == "否":
                return {"status": "skip", "reason": "非财务表格"}

            # 第二步：选择处理模式
            complexity_level = complexity_result["complexity"]
            processing_mode = self._select_processing_mode(complexity_level)

            print("processing_mode:", processing_mode)


            # 第三步：执行具体处理
            res_df, table_name = await self._process_with_mode(image_data, img_size, processing_mode, complexity_level, bank_name, complexity_name)

        except Exception as e:
            res_df = self.fallback_df(e)
            table_name = f"此表报错_{table_name}"

        map_name = "{}_{}".format(table_name, processing_mode)

        self.save_df_to_sheet(res_df, out_file, sheet_name, map_name, image_data=image_name)

        return {
            "status": "success",
            "complexity": complexity_level,
            "mode": processing_mode,
            # "data": result,
            "assessment_reason": complexity_result.get("reason", "")
        }

        # except Exception as e:
        #     print(f"处理表格失败: {e}")
        #     return {"status": "error", "message": str(e)}

    async def _assess_complexity(self, image_data: bytes) -> Dict[str, str]:
        """发送评估请求"""

        response = await self.llm_client.chat.completions.create(
            model=self.model_id,  # 替换为你的模型名
            messages=[
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": self.prompt_registry["assessment"]},
                        {
                            "type": "image_url",
                            "image_url": {"url": f"data:image/png;base64,{image_data}"}
                            # "image_url": {"url": f"data:image/jpeg;base64,{image_data}"}
                        }
                    ]
                }
            ],
            max_tokens=200
        )

        usage_msg = response.usage
        # 统一打印
        print("1111usage_msg:", usage_msg)
        print(f"prompt_tokens    : {usage_msg.prompt_tokens}")
        print(f"completion_tokens: {usage_msg.completion_tokens}")
        print(f"total_tokens     : {usage_msg.total_tokens}")

        # 解析评估结果
        return self._parse_assessment_response(response)

    def _select_processing_mode(self, complexity: str) -> str:
        """选择处理模式"""
        mode_mapping = {
            "简单": "simple",
            "标准": "standard",
            "复杂": "complex"
        }
        return mode_mapping.get(complexity, "standard")

    async def _process_with_mode(self, image_data: bytes, img_size:int, mode: str, complexity: str, banke_name:str, complexity_name:str) -> Any:
        """根据模式处理表格"""

        prompt = self.prompt_registry[mode]

        print("*****************mode*******************", mode, img_size,  complexity)
        max_tokens = 6000

        if mode in ["standard"]:
            max_tokens = 8000
            if complexity in ["中等-扩展型"]:
                max_tokens = 10000
        elif mode in ["complex"]:
            max_tokens = 13000

        if img_size > 3000000:
            max_tokens += 5000
        if img_size > 2000000:
            max_tokens += 3000
        elif img_size > 1500000:
            max_tokens += 2000
        elif img_size > 1000000:
            max_tokens += 1000

        max_tokens = min(max_tokens, 16000)

        pass_state = 0
        print("max_tokens:::", max_tokens)


        t0 = time.time()
        response = "图片太大，太复杂！"
        if not pass_state:
            response = await self.llm_client.chat.completions.create(
                model=self.model_id,  # 替换为你的模型名
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {"type": "text", "text": prompt},
                            {
                                "type": "image_url",
                                "image_url": {"url": f"data:image/png;base64,{image_data}"}
                                # "image_url": {"url": f"data:image/jpeg;base64,{image_data}"}
                            }
                        ]
                    }
                ],
                max_tokens=max_tokens,
                temperature=0.01,
            )

            print("response:", response.choices[0].message.content)

            usage_msg = response.usage
            # 统一打印
            print("2222usage_msg:", usage_msg)
            print(f"prompt_tokens    : {usage_msg.prompt_tokens}")
            print(f"completion_tokens: {usage_msg.completion_tokens}")
            print(f"total_tokens     : {usage_msg.total_tokens}")

            t1 = time.time()
            print("************>:", t1 - t0)
        else:
            print("response:", response)

        return self._parse_processing_response(response, banke_name, complexity_name)

    def _parse_assessment_response(self, response: str) -> Dict[str, str]:
        """解析评估结果"""

        all_content = response.choices[0].message.content
        is_res = re.findall(r"<财务表格>(.*?)</财务表格>", all_content)
        content = ''
        if is_res:
            is_cont = is_res[0]
            if "是" in is_cont:
                cont_res = re.findall("<complexity.*?</complexity>", all_content)
                count_res = re.findall("<table_count.*?=(\d+).*?.*?</table_count>", all_content)

                if cont_res:
                    content = cont_res[0]
                if count_res:
                    print("count_res:", count_res)

        # 简化实现，实际需要更复杂的解析逻辑
        if "极简单" in content:
            return {"complexity": "极简单", "reason": "表格结构简单"}
        elif "简单" in content:
            return {"complexity": "简单", "reason": "表格结构简单"}
        elif "中等" in content:
            if '紧凑' in content:
                return {"complexity": "中等-紧凑型", "reason": "中等复杂度表格"}
            return {"complexity": "中等-扩展型", "reason": "中等复杂度表格"}
        elif "极复杂" in content:
            return {"complexity": "极复杂", "reason": "复杂结构表格"}
        elif "复杂" in content:
            return {"complexity": "复杂", "reason": "复杂结构表格"}
        else:
            return {"complexity": "否", "reason": "非财务表格"}

    def _parse_processing_response_1(self, response: str, bank_name: str, complexity_name: str) -> Any:
        """解析处理结果"""

        def safe_float(x):
            try:
                # 去掉千分位、去空格、转负号
                x = str(x).replace(',', '').strip()
                if x.startswith('(') and x.endswith(')'):
                    x = '-' + x[1:-1]
                return float(x) if x else 0.0
            except Exception:
                return 0.0

        def count_decimal_places(x):
            """计算数值的小数点位数"""
            try:
                # 转换为字符串
                x_str = str(x)
                # 检查是否包含小数点
                if '.' in x_str:
                    # 返回小数点后的位数
                    return len(x_str.split('.')[1])
                else:
                    # 没有小数点，返回0
                    return 0
            except Exception:
                return 0

        cont_pat1 = r"<start4>[\s\S]*?(序号\|主体\|[\s\S]*?)```[\s\S]*?</start4>"
        cont_pat2 = r"<start3>[\s\S]*?(序号\|主体\|[\s\S]*?)```[\s\S]*?</start3>"

        table_pat = r"<start1>(.*?)</start1>"

        content = response.choices[0].message.content

        table_res = re.findall(table_pat, content)
        table_name = ""
        if table_res:
            table_name = table_res[0]

        cont_res = re.search(cont_pat1, content)
        if not cont_res:
            cont_res = re.search(cont_pat2, content)
        print("cont_res:", cont_res)
        raw_txt = cont_res.group(1) if cont_res else ""
        print("raw_txt:", raw_txt)

        ct_res = raw_txt.strip().split("\n")
        # 调整列名，添加"小数点位数"
        name_ct = "银行名|表名|" + ct_res[0] + "|表格复杂度|小数点位数"
        val_res = ct_res[1:]
        # 在每行数据末尾添加一个占位符，后续会被实际计算值替换
        val_cts = ["{}|{}|{}|{}|".format(bank_name, table_name, val_ct, complexity_name) for val_ct in val_res]
        res_cts = [name_ct] + val_cts
        final_cont = '\n'.join(res_cts)

        print("final_cont:", final_cont)

        df = pd.read_csv(StringIO(final_cont), sep='|', dtype=str)  # 先全部当字符串读，避免编号被截断

        # 转换数值列并计算小数点位数
        df['数值'] = df['数值'].apply(safe_float)
        df['小数点位数'] = df['数值'].apply(count_decimal_places)

        return df, table_name

    def _parse_processing_response(self, response: str, bank_name: str, complexity_name: str) -> Any:
        """解析处理结果"""

        def safe_float(x):
            try:
                # 去掉千分位、去空格、转负号
                x = str(x).replace(',', '').strip()
                if x.startswith('(') and x.endswith(')'):
                    x = '-' + x[1:-1]
                # 新增：判断是否为有效数值
                if x.replace('.', '', 1).replace('-', '', 1).isdigit():
                    return float(x) if x else 0.0
                else:
                    return None  # 非数值返回None
            except Exception:
                return None  # 解析失败返回None

        def count_decimal_places(x):
            """计算数值的有效小数点位数（末尾0不计入）"""
            if x is None:  # 非数值返回0
                return 0

            try:
                # 转换为字符串，处理科学计数法
                x_str = str(x).split('e')[0].split('E')[0]

                # 检查是否包含小数点
                if '.' in x_str:
                    # 分割小数部分并去除末尾的0
                    decimal_part = x_str.split('.')[1].rstrip('0')
                    # 如果去除后为空，说明都是0，返回0
                    return len(decimal_part) if decimal_part else 0
                else:
                    # 没有小数点，返回0
                    return 0
            except Exception:
                return 0

        def is_numeric_type(x):
            """判断是否为有效数值类型，是则返回1，否则返回0"""
            return 1 if x is not None else 0

        cont_pat1 = r"<start4>[\s\S]*?(序号\|主体\|[\s\S]*?)```[\s\S]*?</start4>"
        cont_pat2 = r"<start3>[\s\S]*?(序号\|主体\|[\s\S]*?)```[\s\S]*?</start3>"

        table_pat = r"<start1>(.*?)</start1>"

        content = response.choices[0].message.content

        table_res = re.findall(table_pat, content)
        table_name = ""
        if table_res:
            table_name = table_res[0]

        cont_res = re.search(cont_pat1, content)
        if not cont_res:
            cont_res = re.search(cont_pat2, content)
        print("cont_res:", cont_res)
        raw_txt = cont_res.group(1) if cont_res else ""
        print("raw_txt:", raw_txt)

        ct_res = raw_txt.strip().split("\n")
        # 调整列名，添加"小数点位数"和"数值类型"
        name_ct = "银行名|表名|" + ct_res[0] + "|表格复杂度|小数点位数|数值类型"
        val_res = ct_res[1:]
        # 在每行数据末尾添加占位符
        val_num = len(val_res)
        val_state = "少"
        if val_num > 100:
            val_state = "少"
        elif val_num > 50:
            val_state = "中"
        complexity_state = complexity_name + '_' + val_state
        val_cts = ["{}|{}|{}|{}||".format(bank_name, table_name, val_ct, complexity_state) for val_ct in val_res]
        res_cts = [name_ct] + val_cts
        final_cont = '\n'.join(res_cts)

        print("final_cont:", final_cont)

        df = pd.read_csv(StringIO(final_cont), sep='|', dtype=str)  # 先全部当字符串读，避免编号被截断

        # 转换数值列并计算小数点位数和数值类型
        df['数值'] = df['数值'].apply(safe_float)
        df['小数点位数'] = df['数值'].apply(count_decimal_places)
        df['数值类型'] = df['数值'].apply(is_numeric_type)

        return df, table_name

    def insert_pic_to_sheet(self, excel_path, sheet_name, pic_path, anchor_cell='B2', width_px=None, height_px=None):
        """
        excel_path : 目标工作簿
        sheet_name : 目标工作表（不存在则新建）
        pic_path   : 图片文件（png/jpg/gif/bmp...）
        anchor_cell: 图片左上角锚定单元格，如 'B2'
        width_px   : 可选，强制宽度（像素），None 则保持原比例
        height_px  : 可选，强制高度（像素），None 则保持原比例
        """
        wb = load_workbook(excel_path)
        # 1. 获取或创建工作表
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)

        # 2. 加载图片
        img = Image(pic_path)

        # 3. 比例锁定
        if width_px and height_px:
            img.width, img.height = width_px, height_px
        elif width_px:  # 按宽等比缩放
            ratio = img.height / img.width
            img.width, img.height = width_px, int(width_px * ratio)
        elif height_px:  # 按高等比缩放
            ratio = img.width / img.height
            img.width, img.height = int(height_px * ratio), height_px

        # 4. 锚定到单元格
        img.anchor = anchor_cell  # 左上角顶点对齐该单元格

        # 5. 添加并保存
        ws.add_image(img)
        wb.save(excel_path)
        print(f'已插入 {pic_path} 到 {excel_path} 的 {sheet_name}!{anchor_cell}')




import asyncio

async def main():
    model_id = "doubao-1-5-vision-pro-250328"
    # model_id =  "doubao-seed-1-6-vision-250815"
    model = SmartTableProcessor(client, model_id)
    # img_path = r"F:\wills\codes\bankdata\images\re_sub_imgs\514001\514001_128_3_table_last.png"
    img_path = r"F:\wills\codes\bankdata\images\re_sub_imgs\514001\514001_167_4_table_last.png"
    dir_idx = "514010"
    img_dir = r"F:\wills\codes\bankdata\images\re_sub_imgs\{}".format(dir_idx)
    out_dir = r"F:\wills\codes\bankdata\images\out_dir\{}".format(dir_idx)
    if not os.path.exists(out_dir):
        os.makedirs(out_dir)


    imgs = [r"F:\wills\codes\bankdata\images\re_sub_imgs\514001\514001_126_2_table.png",
            r"F:\wills\codes\bankdata\images\re_sub_imgs\514001\514001_105_1_table.png"]

    t001 = time.time()
    for root,_,files in os.walk(img_dir):
        i = 0
        for file in files:
            i += 1
            # if i < 40:
            #     continue
            t1 = time.time()
            try:

                filename = os.path.join(root,file)
                print("filename:", filename)
                stem = Path(filename).stem
                sh_nms = stem.split("_")[1:3]
                sheet_name = '_'.join(sh_nms)
                print("stem:", stem)
                sub_stems = stem.split("_")
                check_stem = sub_stems[1]
                # if check_stem not in ["238", "193", "244"]:
                #     continue



                # if filename not in imgs:
                #     continue


                out_file = out_dir + "/" + "财务数据表_2.xlsx"
                res = await model.process_table(filename, out_file, sheet_name,"中国建设银行", "")

                print("res:", res)
            except Exception as e:
                print(">>>>>>>>>>>>>>>>>>e:", e)

            # if i > 1:
            #     break
            t2 = time.time()

            print("當前耗時：", t2 - t1)

    t002 = time.time()
    print("總耗時：", t002 - t001)

if __name__ == "__main__":
    asyncio.run(main())

    # 序号|主体|纵向层级路径|横向层级路径|数值|单位|报告期|数据类型|汇总标记|维度类型
    # 网格ID|主体|纵向层级路径|横向层级路径|数值|单位|报告期|数据类型|汇总标记|维度类型