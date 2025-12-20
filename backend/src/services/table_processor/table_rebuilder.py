
import re
import os
from backend.src.services.table_processor.long_format_converter import FinalDataConverter
from backend.src.services.table_processor.marked_table_processor import MarkedTableProcessor


class TableReconstructor:
    """表格重构器：整合7步流程"""

    def __init__(self):
        self.warnings = []
        self.issues = []
        # 新增：记忆前一个表格的列结构
        self.prev_table_header_structure = None
        self.final_data_converter = FinalDataConverter()

        # 新增：独立的标记表格处理器
        self.marked_table_processor = MarkedTableProcessor()

    def _unify_headers_across_tables(self, llm_tables):
        """
        统一相似表格的列标题结构 - 简单实现版
        对于结构相似的相邻表格，如果存在列标题缺失，用前面表格的结构填充
        """
        print("\n=== 列标题统一化处理 ===")

        if len(llm_tables) <= 1:
            print("表格数量≤1，无需统一处理")
            return llm_tables

        for i in range(1, len(llm_tables)):
            current_table = llm_tables[i]
            prev_table = llm_tables[i - 1]

            # 获取列标题
            current_cols = current_table.get('headers', {}).get('cols', [])
            prev_cols = prev_table.get('headers', {}).get('cols', [])

            # 只有列数相同时才考虑统一
            if len(current_cols) == len(prev_cols):
                # 检查是否存在空列标题
                new_cols = []
                need_unify = False

                for idx, (prev_col, curr_col) in enumerate(zip(prev_cols, current_cols)):
                    # 如果当前列标题为空，而前一表格的对应列有内容
                    if (not curr_col or str(curr_col).strip() == '') and prev_col and str(prev_col).strip() != '':
                        new_cols.append(prev_col)
                        need_unify = True
                        print(f"  表格{i}第{idx}列: 空白 → 填充为 '{prev_col}'")
                    else:
                        new_cols.append(curr_col)

                # 如果需要统一，更新表格
                if need_unify:
                    llm_tables[i]['headers']['cols'] = new_cols
                    print(f"  表格{i}统一化完成")

        return llm_tables

    def _analyze_column_structure(self, col_headers):
        """
        分析列标题的结构特征
        返回: 结构特征字典
        """
        if not col_headers:
            return {"empty": True}

        structure = {
            "count": len(col_headers),
            "first_empty": col_headers[0] == '' if col_headers else False,
            "year_count": sum(
                1 for col in col_headers if col and any(year in str(col) for year in ['年', '202', '201'])),
            "change_col": any(col and '变化' in str(col) for col in col_headers),
            "pattern": []
        }

        # 分析列标题模式
        for col in col_headers:
            if not col or str(col).strip() == '':
                structure["pattern"].append("empty")
            elif any(year in str(col) for year in ['2024', '2023', '2022']):
                structure["pattern"].append("year")
            elif '变化' in str(col) or 'chang' in str(col).lower():
                structure["pattern"].append("change")
            else:
                structure["pattern"].append("other")

        return structure

    def _are_similar_column_structures(self, struct1, struct2):
        """
        判断两个列结构是否相似
        """
        if not struct1 or not struct2:
            return False

        # 如果都是空结构
        if struct1.get("empty") and struct2.get("empty"):
            return False

        # 列数相同是关键
        if struct1.get("count") != struct2.get("count"):
            return False

        # 年份列数量相似
        if abs(struct1.get("year_count", 0) - struct2.get("year_count", 0)) > 1:
            return False

        # 是否都有变化列
        if struct1.get("change_col") != struct2.get("change_col"):
            return False

        # 模式相似度（至少60%相同）
        pattern1 = struct1.get("pattern", [])
        pattern2 = struct2.get("pattern", [])

        if len(pattern1) != len(pattern2):
            return False

        same_count = sum(1 for p1, p2 in zip(pattern1, pattern2) if p1 == p2)
        similarity = same_count / len(pattern1) if pattern1 else 0

        return similarity >= 0.6

    # ========== 工具函数 ==========
    def log_warning(self, message):
        """记录警告"""
        self.warnings.append(message)
        print(f"⚠️ {message}")

    def log_issue(self, message):
        """记录问题"""
        self.issues.append(message)
        print(f"❓ {message}")

    def clean_text_for_matching(self, text):
        """清理文本用于匹配"""
        if not text:
            return ""
        text = str(text)
        # 替换换行、空格
        text = text.replace('\n', '').replace('\r', '').replace(' ', '')
        # 去掉特殊符号，保留中文、数字、字母
        cleaned = ''.join(c for c in text if c.isalnum() or '\u4e00-\u9fff' in c)
        return cleaned

    def calculate_similarity(self, text1, text2):
        """计算两个文本的相似度"""
        t1 = self.clean_text_for_matching(text1)
        t2 = self.clean_text_for_matching(text2)

        if not t1 or not t2:
            return 0

        # 完全相等
        if t1 == t2:
            return 1.0

        # 包含关系
        if t1 in t2 or t2 in t1:
            return 0.9

        # 公共字符比例
        common_chars = set(t1) & set(t2)
        if not common_chars:
            return 0

        # 计算Jaccard相似度
        union_chars = set(t1) | set(t2)
        return len(common_chars) / len(union_chars)

    def calculate_similarity_v2(self, text1, text2):
        """增强版相似度计算（考虑包含关系的覆盖率）"""
        t1 = self.clean_text_for_matching(text1)
        t2 = self.clean_text_for_matching(text2)

        if not t1 or not t2:
            return 0

        # 1. 完全相等
        if t1 == t2:
            return 1.0

        # 2. 包含关系，但要求高覆盖率
        if t1 in t2:
            coverage = len(t1) / len(t2)
            # 覆盖度>80%才给高分，否则给低分
            # 避免"经营活动"匹配"每股经营活动"
            return 0.9 if coverage > 0.8 else 0.2

        if t2 in t1:
            coverage = len(t2) / len(t1)
            return 0.9 if coverage > 0.8 else 0.2

        # 3. 公共字符比例（原有逻辑）
        common_chars = set(t1) & set(t2)
        if not common_chars:
            return 0

        # 计算Jaccard相似度
        union_chars = set(t1) | set(t2)
        similarity = len(common_chars) / len(union_chars)

        return similarity

    def _detect_tables_to_merge(self, llm_tables):
        """
        检测哪些LLM表格应该合并处理
        返回：需要合并的表格组列表，每个组包含应该合并的表格索引
        """
        print("\n=== 检测可合并的表格结构 ===")

        if len(llm_tables) <= 1:
            print("表格数量≤1，无需合并检测")
            return []

        # 检查每个表格的列结构
        table_col_structures = []

        for i, table in enumerate(llm_tables):
            col_headers = table.get('headers', {}).get('cols', [])
            # 清理列标题（去掉空值和特殊字符）
            cleaned_cols = []
            for col in col_headers:
                if col and str(col).strip():
                    # 提取文本内容（去掉>>等层级标记）
                    text = str(col).split('>>')[-1].strip()
                    if text:
                        cleaned_cols.append(text)

            table_info = {
                'index': i,
                'col_count': len(col_headers),
                'cleaned_cols': cleaned_cols,
                'original_cols': col_headers,
                'row_count': len(table.get('headers', {}).get('rows', []))
            }
            table_col_structures.append(table_info)

            print(f"表格{i}: {table_info['row_count']}行 × {table_info['col_count']}列")
            print(f"  列标题: {cleaned_cols[:5]}...")

        # 检测相似的列结构（用于合并）
        merge_groups = []
        processed_indices = set()

        for i in range(len(table_col_structures)):
            if i in processed_indices:
                continue

            current_table = table_col_structures[i]
            merge_group = [i]

            # 寻找与当前表格列结构相似的表格
            for j in range(i + 1, len(table_col_structures)):
                if j in processed_indices:
                    continue

                next_table = table_col_structures[j]

                # 合并条件1：列数相同
                if current_table['col_count'] != next_table['col_count']:
                    continue

                # 合并条件2：列标题相似度检查（简化版）
                is_similar = self._check_columns_similarity(
                    current_table['cleaned_cols'],
                    next_table['cleaned_cols']
                )

                if is_similar:
                    merge_group.append(j)
                    processed_indices.add(j)

            if len(merge_group) > 1:
                merge_groups.append(merge_group)
                processed_indices.add(i)

        # 输出检测结果
        if merge_groups:
            print("\n检测到可合并的表格组:")
            for group_idx, group in enumerate(merge_groups):
                print(f"  合并组{group_idx}: 表格索引 {group}")
                print(f"    包含表格: ", end="")
                for table_idx in group:
                    table_info = table_col_structures[table_idx]
                    print(f"表格{table_idx}({table_info['row_count']}行) ", end="")
                print()
        else:
            print("未检测到需要合并的表格")

        return merge_groups

    def _check_columns_similarity(self, cols1, cols2):
        """
        检查两个列标题列表是否相似
        简化版：检查前几列是否包含相同的年份/数字模式
        """
        if not cols1 or not cols2:
            return False

        # 如果列数不同，肯定不相似
        if len(cols1) != len(cols2):
            return False

        # 检查是否包含典型的年份模式
        year_patterns = ['年', '202', '201', '月', '季度', 'Q']

        def contains_year_pattern(text):
            if not text:
                return False
            text_str = str(text)
            for pattern in year_patterns:
                if pattern in text_str:
                    return True
            return False

        # 统计两个列表中包含年份模式的列数
        cols1_year_count = sum(1 for col in cols1 if contains_year_pattern(col))
        cols2_year_count = sum(1 for col in cols2 if contains_year_pattern(col))

        # 如果都有年份列，认为可能是相似的表格结构
        if cols1_year_count > 0 and cols2_year_count > 0:
            return True

        # 检查前两列是否相似（对于财务报表，通常第一列是项目，后面是年份）
        similarity_threshold = 0.7
        similar_count = 0

        for idx in range(min(len(cols1), len(cols2))):
            if idx == 0:
                # 第一列通常是项目名称，允许不同
                continue

            sim = self.calculate_similarity(cols1[idx], cols2[idx])
            if sim > similarity_threshold:
                similar_count += 1

        # 如果有超过一半的非首列相似，认为表格结构相似
        if len(cols1) > 1:
            return similar_count / (len(cols1) - 1) > 0.5

        return False

    def _match_with_table_boundaries(self, table, row_headers, table_boundaries):
        """基于表格边界的匹配算法"""
        print(f"表格: {len(table)}行 × {len(table[0])}列")
        print(f"LLM行表头数: {len(row_headers)}")

        # 1. 为每个OCR表格建立搜索空间
        search_spaces = []
        for boundary in table_boundaries:
            start_row = boundary['start_row']
            end_row = boundary['end_row']
            # 转换为1-based索引（表格数据是0-based）
            search_space = list(range(start_row, end_row + 1))
            search_spaces.append({
                'table_idx': boundary['table_idx'],
                'rows': search_space,
                'next_row_idx': 0,  # 在这个表格内下一个可用的行索引
                'used_rows': set()  # 已使用的行
            })
            print(f"  表格{boundary['table_idx']}: 行{start_row}-{end_row}")

        # 2. 确定搜索列（与原有逻辑一致）
        left_empty_cols = 0
        for title in table[0]:
            if not title or str(title).strip() == '':
                left_empty_cols += 1
            else:
                break

        search_columns = list(range(min(3, len(table[0]))))
        if left_empty_cols > 0:
            search_columns = list(range(left_empty_cols))

        print(f"搜索列: {search_columns}")

        # 3. 按LLM顺序匹配
        assignments = {}

        for header_idx, llm_header in enumerate(row_headers):
            target_text = llm_header.split('>>')[-1] if '>>' in llm_header else llm_header

            print(f"  表头[{header_idx}]: '{target_text[:30]}...'")

            best_match = None
            best_score = 0
            best_table_idx = -1

            # 在每个表格的搜索空间中寻找最佳匹配
            for table_info in search_spaces:
                for row_idx in table_info['rows']:
                    if row_idx in table_info['used_rows']:
                        continue

                    # 计算匹配分数
                    row_score = 0
                    for col_idx in search_columns:
                        if col_idx < len(table[row_idx]):
                            cell_val = table[row_idx][col_idx]
                            if cell_val:
                                score = self.calculate_similarity_v2(target_text, str(cell_val))
                                if score > row_score:
                                    row_score = score

                    if row_score > best_score:
                        best_score = row_score
                        best_match = row_idx
                        best_table_idx = table_info['table_idx']

            # 4. 如果找到高质量匹配，分配该行
            if best_match and best_score > 0.6:
                assignments[header_idx] = best_match
                # 标记该行为已使用
                for table_info in search_spaces:
                    if best_match in table_info['rows']:
                        table_info['used_rows'].add(best_match)
                print(f"    → 匹配到表格{best_table_idx}的行{best_match} (分数:{best_score:.2f})")
            else:
                # 5. 按表格顺序分配下一个可用行
                assigned = False
                for table_info in search_spaces:
                    # 在这个表格内按顺序找下一个可用行
                    while table_info['next_row_idx'] < len(table_info['rows']):
                        row_idx = table_info['rows'][table_info['next_row_idx']]
                        if row_idx not in table_info['used_rows']:
                            assignments[header_idx] = row_idx
                            table_info['used_rows'].add(row_idx)
                            table_info['next_row_idx'] += 1
                            assigned = True
                            print(f"    → 顺序分配到表格{table_info['table_idx']}的行{row_idx}")
                            break
                        table_info['next_row_idx'] += 1

                    if assigned:
                        break

                if not assigned:
                    print(f"    → 无法分配")
                    assignments[header_idx] = -1

        # 6. 填充表格（与原有逻辑一致）
        for row_idx in range(1, len(table)):
            table[row_idx][0] = None

        filled_count = 0
        for header_idx, row_idx in assignments.items():
            if row_idx != -1 and row_idx < len(table):
                table[row_idx][0] = row_headers[header_idx]
                filled_count += 1

        print(f"\n匹配结果: {filled_count}/{len(row_headers)} 个行表头已填充")
        return table


    def _analyze_cell_type(self, cell_value):
        """
        分析单元格类型
        返回: "blank", "text", "std_num", "minor_num", "error_num"
        """
        if cell_value is None:
            return "blank"

        text = str(cell_value).strip()

        # 1. 空白单元格
        if text == "":
            return "blank"

        # 2. 检查是否为纯文本（不包含数字）
        if not any(c.isdigit() for c in text):
            return "text"

        # ========== 新增：检查%号前是否有逗号 ==========
        if '%' in text:
            # 找到%号的位置
            percent_pos = text.find('%')
            # 检查%号前面是否有逗号
            if percent_pos > 0 and ',' in text[:percent_pos]:
                # %号前面有逗号，这是错误格式（如"0,82%"）
                # 但是要排除类似"1,234.56%"的情况
                # 检查逗号是否在正确的位置
                before_percent = text[:percent_pos]
                if '.' in before_percent:
                    # 有小数点，检查逗号和小数点的关系
                    dot_pos = before_percent.find('.')
                    for i, char in enumerate(before_percent):
                        if char == ',' and i > dot_pos:
                            # 逗号在小数点后面，这是错误格式
                            return "error_num"
                else:
                    # 没有小数点，%号前不应该有逗号
                    return "error_num"

        # 3. 检查负值格式
        cleaned = text

        # 处理括号负数
        is_parenthesis_negative = (
                cleaned.startswith('(') and
                cleaned.endswith(')') and
                '(' not in cleaned[1:-1] and
                ')' not in cleaned[1:-1]
        )

        if is_parenthesis_negative:
            # 检查括号内是否还有负号
            inside = cleaned[1:-1].strip()
            if inside.startswith('-') or inside.endswith('-'):
                return "error_num"  # 如"(-450)"或"(450-)"，错误格式
            cleaned = '-' + inside

        # 4. 检查多个负号或负号位置错误
        if cleaned.count('-') > 1:
            return "error_num"  # 如"--450"

        if '-' in cleaned and not cleaned.startswith('-'):
            return "error_num"  # 如"450-"

        # 5. 移除负号、逗号、小数点、%号用于数值转换
        test_str = cleaned
        if '-' in test_str:
            test_str = test_str.replace('-', '', 1)

        # 记录原始符号用于格式检查
        has_percent = '%' in test_str
        dot_count = test_str.count('.')
        comma_count = test_str.count(',')

        # 用于数值转换的字符串
        numeric_test = test_str.replace(',', '').replace('.', '').replace('%', '')

        # 6. 尝试转换为数值
        try:
            # 如果转换失败，说明不是有效数值
            float(numeric_test)
        except ValueError:
            return "text"  # 包含数字但不是有效数值

        # 7. 格式检查
        # 检查%号
        if has_percent:
            if not cleaned.endswith('%'):
                return "error_num"  # %不在末尾
            if cleaned.count('%') > 1:
                return "error_num"  # 多个%

        # 检查小数点
        if dot_count > 1:
            return "error_num"  # 多个小数点

        # 检查逗号位置（如果有小数点）
        if dot_count == 1 and comma_count > 0:
            dot_pos = cleaned.find('.')
            # 检查逗号是否都在小数点前面
            for i, char in enumerate(cleaned):
                if char == ',' and i > dot_pos:
                    return "error_num"  # 逗号在小数点后面

        # 8. 检查是否为标准格式
        # 标准格式：有逗号时必须在正确位置，%在末尾等
        if self._is_standard_format(cleaned):
            return "std_num"
        else:
            return "minor_num"

    def _is_standard_format(self, text):
        """
        判断是否为标准数值格式
        标准格式要求：
        1. 千分位逗号格式正确（每3位一个逗号）
        2. 如果有小数点，后面通常有数字
        3. 没有多余的空格或其他字符
        """
        if not text:
            return False

        # 移除负号和%号
        test_text = text
        if test_text.startswith('-'):
            test_text = test_text[1:]
        if test_text.endswith('%'):
            test_text = test_text[:-1]

        # 检查逗号格式
        if ',' in test_text:
            # 分割整数和小数部分
            if '.' in test_text:
                int_part, dec_part = test_text.split('.', 1)
            else:
                int_part, dec_part = test_text, ""

            # 检查整数部分的逗号格式
            groups = int_part.split(',')
            # 第一组可以是1-3位
            if not (1 <= len(groups[0]) <= 3):
                return False

            # 后面的每组必须正好是3位
            for group in groups[1:]:
                if len(group) != 3:
                    return False

        return True

    def _fill_blank_cells(self, cell_types):
        """
        填充空白单元格的类型（根据同行/同列的众数类型）
        """
        if not cell_types:
            return

        num_rows = len(cell_types)
        num_cols = len(cell_types[0]) if num_rows > 0 else 0

        for r in range(num_rows):
            for c in range(num_cols):
                if cell_types[r][c] == "blank":
                    # 查找同行和同列的非空白类型
                    row_types = []
                    col_types = []

                    # 同行
                    for cc in range(num_cols):
                        if cc != c and cell_types[r][cc] != "blank":
                            row_types.append(cell_types[r][cc])

                    # 同列
                    for rr in range(num_rows):
                        if rr != r and cell_types[rr][c] != "blank":
                            col_types.append(cell_types[rr][c])

                    # 合并所有类型
                    all_types = row_types + col_types

                    if all_types:
                        # 找出众数类型
                        from collections import Counter
                        type_counts = Counter(all_types)
                        most_common = type_counts.most_common(1)

                        if most_common:
                            cell_types[r][c] = most_common[0][0]

    def _determine_row_column_mark(self, cell_types, label):
        """
        根据单元格类型列表确定行或列的标记
        """
        # 过滤掉空白单元格
        non_blank_types = [t for t in cell_types if t != "blank"]

        if not non_blank_types:
            return 0  # 全空白，视为纯文本

        # 检查是否有文本
        has_text = "text" in non_blank_types
        has_std_num = "std_num" in non_blank_types
        has_minor_num = "minor_num" in non_blank_types
        has_error_num = "error_num" in non_blank_types

        # 4. 混合类型（包含文本和任何数值）
        if has_text and (has_std_num or has_minor_num or has_error_num):
            print(f"  {label}: 标记4 - 混合类型 (文本+数值)")
            return 4

        # 3. 很可能错误的数值
        if has_error_num:
            print(f"  {label}: 标记3 - 可能错误的数值")
            return 3

        # 现在只包含数值类型
        # 2. 有小问题的数值（有minor_num，可能也有std_num）
        if has_minor_num:
            print(f"  {label}: 标记2 - 格式问题数值")
            return 2

        # 1. 完全正确的数值（全是std_num）
        if has_std_num and not has_minor_num and not has_error_num:
            print(f"  {label}: 标记1 - 标准数值")
            return 1

        # 0. 纯文本
        if not has_std_num and not has_minor_num and not has_error_num:
            print(f"  {label}: 标记0 - 纯文本")
            return 0

        # 默认返回0
        return 0

    def _count_marks(self, marks_list):
        """
        统计标记数量
        """
        from collections import Counter
        counter = Counter(marks_list)

        result = []
        for mark in range(5):
            count = counter.get(mark, 0)
            if count > 0:
                result.append(f"标记{mark}:{count}个")

        return ", ".join(result)

    def _fix_llm_table_references(self, ocr_result, llm_result):
        """修正LLM引用的表格索引，确保引用有数据的表格"""
        ocr_tables = ocr_result.get('tables_result', [])

        # 找出有数据的OCR表格
        valid_ocr_indices = []
        for i, table in enumerate(ocr_tables):
            if table.get('body') and len(table['body']) > 0:
                valid_ocr_indices.append(i)

        print(f"有效OCR表格索引: {valid_ocr_indices}")

        if not valid_ocr_indices:
            print("⚠️ 没有找到有数据的OCR表格")
            return

        # 修正每个LLM表格的引用
        llm_tables = llm_result.get('tables_structure', {}).get('tables', [])
        for llm_table in llm_tables:
            original_refs = llm_table.get('ocr_tables', [])

            # 检查引用的表格是否有数据
            valid_refs = []
            for ref_idx in original_refs:
                if ref_idx in valid_ocr_indices:
                    valid_refs.append(ref_idx)

            # 如果没有有效的引用，使用第一个有效表格
            if not valid_refs and valid_ocr_indices:
                valid_refs = [valid_ocr_indices[0]]
                print(f"⚠️ 表格'{llm_table.get('name')}'的引用无效，修正为: {valid_refs}")

            llm_table['ocr_tables'] = valid_refs

    # ========== 核心7步 ==========
    def step1_prepare_data(self, ocr_result, llm_result):
        """第1步：准备数据"""
        # 直接返回传入的数据
        return ocr_result, llm_result

    def step2_extract_table_data(self, ocr_result, llm_result):
        """
        从数据中提取表格信息
        """
        # 提取OCR表格
        ocr_tables = ocr_result.get('tables_result', [])
        print(f"OCR表格数: {len(ocr_tables)}")

        # 提取LLM表格结构
        llm_tables = []
        if 'tables_structure' in llm_result:
            llm_tables = llm_result['tables_structure'].get('tables', [])
        elif 'tables' in llm_result:
            llm_tables = llm_result['tables']
        print(f"LLM表格结构数: {len(llm_tables)}")

        return ocr_tables, llm_tables

    def step3_merge_ocr_tables(self, ocr_tables, llm_table_info):
        """合并OCR表格 - 修正版"""
        ocr_table_indices = llm_table_info.get('ocr_tables', [])
        print(f"要合并的OCR表格索引: {ocr_table_indices}")

        # 🔥 添加调试：查看整个OCR数据结构
        print(f"\n🔍 OCR表格数据结构调试:")
        for i, table in enumerate(ocr_tables):
            print(f"OCR表格{i} 类型: {type(table)}")
            if isinstance(table, dict):
                print(f"  键: {list(table.keys())}")
                if 'body' in table:
                    print(f"  body类型: {type(table['body'])}, 长度: {len(table['body']) if table['body'] else 0}")
                if 'cell_set' in table:
                    print(f"  cell_set长度: {len(table['cell_set']) if table['cell_set'] else 0}")



        all_cells = []
        row_offset = 0
        table_boundaries = []

        for idx in ocr_table_indices:
            if idx >= len(ocr_tables):
                print(f"警告: OCR表格索引 {idx} 超出范围")
                continue

            table = ocr_tables[idx]
            cells = table.get('body', [])

            print(f"处理OCR表格{idx}: {len(cells)}个单元格")

            if not cells:
                continue

            # 找出这个表格的实际最大行号
            max_row_in_table = 0
            for cell in cells:
                max_row_in_table = max(max_row_in_table, cell['row_end'])

            # 记录表格边界（在调整行号之前）
            table_boundaries.append({
                'table_idx': idx,
                'start_row': row_offset,  # 调整后的起始行
                'end_row': row_offset + max_row_in_table,  # 调整后的结束行
                'row_count': max_row_in_table + 1,
                'original_max_row': max_row_in_table
            })

            print(f"  表格{idx}: 原始行0-{max_row_in_table} → 合并后行{row_offset}-{row_offset + max_row_in_table}")

            # 调整行号并收集单元格
            for cell in cells:
                adjusted_cell = cell.copy()
                adjusted_cell['row_start'] += row_offset
                adjusted_cell['row_end'] += row_offset
                adjusted_cell['source_table_idx'] = idx
                adjusted_cell['original_row'] = cell['row_start']
                all_cells.append(adjusted_cell)

            # 更新行偏移量：当前表格的最大行号 + 分隔空行
            row_offset += (max_row_in_table + 2)  # +2留一个空行分隔

        print(f"\n表格边界信息（修正后）:")
        for boundary in table_boundaries:
            print(
                f"  表格{boundary['table_idx']}: 行{boundary['start_row']}-{boundary['end_row']} ({boundary['row_count']}行)")

        return {
            'cells': all_cells,
            'headers': llm_table_info.get('headers', {}),
            'table_id': llm_table_info.get('id', '1'),
            'name': llm_table_info.get('name', ''),
            'table_boundaries': table_boundaries
        }

    def step4_create_base_data_table(self, merged_data):
        """
            创建基础数据表格
            检查每行列数是否一致
            """
        cells = merged_data['cells']

        if not cells:
            return []

        # 找出最大行列
        max_row = 0
        max_col = 0
        for cell in cells:
            max_row = max(max_row, cell['row_end'])
            max_col = max(max_col, cell['col_end'])

        num_rows = max_row
        num_cols = max_col

        # 检查每行的列数
        print("检查每行列数:")
        row_col_counts = {}
        for cell in cells:
            row = cell['row_start']
            col_end = cell['col_end']

            if row not in row_col_counts:
                row_col_counts[row] = []
            row_col_counts[row].append(col_end)

        # 找出每行的最大列号
        row_max_cols = {}
        for row, col_ends in row_col_counts.items():
            row_max_cols[row] = max(col_ends)
            print(f"  第{row}行: 最大列索引={row_max_cols[row]}")

        # 检查是否所有行的最大列索引一致
        all_max_cols = list(row_max_cols.values())
        if len(set(all_max_cols)) > 1:
            print(f"警告: 不同行的列数不一致: {all_max_cols}")

        print(f"表格: {num_rows}行 × {num_cols}列")

        # 创建表格
        table = []
        for r in range(num_rows):
            table.append([None] * num_cols)

        # 填充数据
        for cell in cells:
            value = cell['words']
            row_idx = cell['row_start']
            col_idx = cell['col_start']

            if row_idx < num_rows and col_idx < num_cols:
                table[row_idx][col_idx] = value

        return table


    def step5_add_column_headers(self, table, col_headers):
        """
        优化版：清理和验证列标题，调整列数匹配
        """
        if not col_headers:
            print("无列标题")
            return table

        if not table:
            print("表格为空")
            return table

        current_cols = len(table[0])  # OCR表格的列数
        target_cols = len(col_headers)  # LLM的最终列数

        print(f"OCR列数: {current_cols}, LLM列数: {target_cols}")

        # 1. 清理列标题
        cleaned_headers = []
        for header in col_headers:
            # 移除空白和空值
            if not header or str(header).strip() == '':
                cleaned_headers.append("")
                continue

            header_str = str(header).strip()

            # 标准化分隔符
            header_str = header_str.replace('><', '>>')
            header_str = header_str.replace('＞＞', '>>')

            # 修复可能的分隔符错误
            if '>' in header_str and '>' in header_str[1:]:
                # 处理类似 "b>2024年12月31日" 的情况
                parts = header_str.split('>')
                if len(parts) == 2:
                    header_str = f"{parts[0]}>>{parts[1]}"

            # 移除多余空格
            header_str = '>>'.join([part.strip() for part in header_str.split('>>')])

            # 特别处理：如果标题就是">>"，设为空
            if header_str == '>>' or header_str == '>':
                header_str = ""

            cleaned_headers.append(header_str)

        # 2. 验证列标题有效性
        self._validate_cleaned_headers(cleaned_headers)

        # 3. 调整列数匹配
        # 情况1：OCR列数 > LLM列数（需要删除左侧多余的列）
        if current_cols > target_cols:
            excess_cols = current_cols - target_cols
            print(f"需要删除左侧{excess_cols}列")

            for i in range(len(table)):
                table[i] = table[i][excess_cols:]

            current_cols = target_cols

        # 情况2：OCR列数 < LLM列数（需要补充左侧空列）
        elif current_cols < target_cols:
            needed_cols = target_cols - current_cols
            print(f"需要补充左侧{needed_cols}列")

            for i in range(len(table)):
                table[i] = [None] * needed_cols + table[i]

            current_cols = target_cols
        else:
            print("列数已匹配，无需调整")

        print(f"调整后表格列数: {current_cols}")

        # 4. 在顶部添加一行用于列标题
        table.insert(0, [None] * current_cols)

        # 5. 从右向左填充清理后的列标题
        cleaned_headers_copy = cleaned_headers.copy()
        for col in range(current_cols - 1, -1, -1):
            if cleaned_headers_copy:
                table[0][col] = cleaned_headers_copy.pop()

        print("列标题填充完成")
        print(f"第0行（列标题）: {table[0]}")

        return table

    def _validate_cleaned_headers(self, cleaned_headers):
        """
        验证清理后的列标题
        """
        print(f"清理后的列标题: {cleaned_headers}")

        # 统计空列标题
        empty_count = sum(1 for h in cleaned_headers if not h)
        if empty_count > 0:
            print(f"⚠️ 有{empty_count}个空列标题")

        # 检查是否有非空的列标题
        non_empty_count = len(cleaned_headers) - empty_count
        if non_empty_count == 0:
            print("⚠️ 所有列标题都为空，请检查LLM输出")

        # 检查列标题格式（可选，仅用于信息提示）
        for i, header in enumerate(cleaned_headers):
            if header:
                # 检查是否有分隔符但格式异常
                if '>>' in header:
                    parts = header.split('>>')
                    if len(parts) > 2:
                        print(f"  列{i}: 包含多层分隔符 -> {header}")
                elif '>' in header:
                    print(f"  列{i}: 使用单层分隔符 -> {header}")

    def _looks_like_numeric_data(self, text):
        """判断文本是否看起来像数值型数据"""
        if not text:
            return False

        text_str = str(text).strip()
        clean_text = text_str.replace(',', '').replace(' ', '').replace('¥', '').replace('$', '').replace('€', '')

        if clean_text.startswith('(') and clean_text.endswith(')'):
            clean_text = '-' + clean_text[1:-1]

        if not any(c.isdigit() for c in clean_text):
            return False

        try:
            if '%' in clean_text:
                clean_text = clean_text.replace('%', '')
                float(clean_text)
                return True
            else:
                float(clean_text)
                return True
        except ValueError:
            digit_count = sum(1 for c in clean_text if c.isdigit())
            total_len = len(clean_text)
            if digit_count / total_len > 0.7:
                return True

        return False

    def step6_add_row_headers_intelligent(self, table, row_headers, table_boundaries=None):
        """第6步：优化版 - 利用层级关系和相邻位置插入"""

        # [1] 首先尝试匹配所有LLM表头
        llm_to_ocr_match = {}
        ocr_to_llm_match = {}

        for llm_idx, llm_header in enumerate(row_headers):
            header_text = str(llm_header)

            # 使用更灵活的匹配策略
            best_match_row = -1
            best_match_score = 0.0

            for ocr_row in range(1, len(table)):
                if ocr_row in ocr_to_llm_match:
                    continue

                # 计算匹配分数
                row_score = self._calculate_header_match_score(header_text, table[ocr_row])

                if row_score > best_match_score:
                    best_match_score = row_score
                    best_match_row = ocr_row

            # 动态调整匹配阈值
            threshold = self._get_dynamic_threshold(header_text)
            if best_match_score >= threshold:
                llm_to_ocr_match[llm_idx] = (best_match_row, best_match_score)
                ocr_to_llm_match[best_match_row] = llm_idx

        # [2] 分析LLM表头的层级关系
        header_hierarchy = self._analyze_header_hierarchy(row_headers)

        print("1111111111111111111111111row_headers:", row_headers)
        print("2222222222222header_hierarchy2222222222222")
        print(header_hierarchy)


        # [3] 为每个LLM表头确定插入位置
        llm_insert_positions = {}

        for llm_idx, llm_header in enumerate(row_headers):
            if llm_idx in llm_to_ocr_match:
                # 有精确匹配
                ocr_row, score = llm_to_ocr_match[llm_idx]


                llm_insert_positions[llm_idx] = ('exact', ocr_row)
            else:
                # 无匹配，需要找插入位置
                insert_info = self._find_best_insert_position(
                    llm_idx, llm_header, row_headers,
                    llm_to_ocr_match, header_hierarchy, table
                )
                llm_insert_positions[llm_idx] = insert_info

        # [4] 按LLM顺序构建表格
        new_table = []
        if len(table) > 0:
            new_table.append(table[0].copy())

        # 记录已处理的OCR行
        processed_ocr_rows = set()

        # 按LLM顺序处理
        for llm_idx in range(len(row_headers)):
            insert_type, ref_info = llm_insert_positions[llm_idx]

            if insert_type == 'exact':
                # 处理精确匹配
                ocr_row = ref_info
                if ocr_row not in processed_ocr_rows:
                    new_row = table[ocr_row].copy()
                    new_row[0] = row_headers[llm_idx]
                    new_table.append(new_row)
                    processed_ocr_rows.add(ocr_row)
            elif insert_type == 'before':
                # 在参考行之前插入
                ref_row = ref_info
                if ref_row in processed_ocr_rows:
                    # 找到参考行在新表格中的位置
                    ref_position = self._find_row_in_new_table(new_table, ref_row)
                    if ref_position != -1:
                        blank_row = [None] * len(table[0]) if table else []
                        blank_row[0] = row_headers[llm_idx]
                        new_table.insert(ref_position, blank_row)
            elif insert_type == 'after':
                # 在参考行之后插入
                ref_row = ref_info
                if ref_row in processed_ocr_rows:
                    ref_position = self._find_row_in_new_table(new_table, ref_row)
                    if ref_position != -1:
                        blank_row = [None] * len(table[0]) if table else []
                        blank_row[0] = row_headers[llm_idx]
                        new_table.insert(ref_position + 1, blank_row)

        # 假设new_table中的行顺序和原始table中匹配的行顺序一致
        complete_table = []

        # 添加表头行
        if len(new_table) > 0:
            complete_table.append(new_table[0].copy())

        # 遍历原始table的行
        new_table_index = 1  # 跳过表头
        for ocr_row in range(1, len(table)):
            if ocr_row in ocr_to_llm_match:
                # 这行应该有LLM匹配，应该在new_table中
                if new_table_index < len(new_table):
                    complete_table.append(new_table[new_table_index].copy())
                    new_table_index += 1
            else:
                # 这行没有LLM匹配，直接使用原始行
                complete_table.append(table[ocr_row].copy())

        # 处理new_table中可能剩余的行
        while new_table_index < len(new_table):
            complete_table.append(new_table[new_table_index].copy())
            new_table_index += 1

        table[:] = complete_table

        return table

    def _find_best_insert_position(self, llm_idx, llm_header, row_headers,
                                   llm_to_ocr_match, header_hierarchy, table):
        """
        为无匹配的LLM表头找到最佳插入位置
        """
        header_text = str(llm_header)

        # 1. 检查层级关系：如果是分类标题，应该在其子项之前
        if llm_idx in header_hierarchy['parents']:
            children = header_hierarchy['parents'][llm_idx]
            # 找到第一个有匹配的子项
            for child_idx in children:
                if child_idx in llm_to_ocr_match:
                    ocr_row, score = llm_to_ocr_match[child_idx]
                    return ('before', ocr_row)  # 在子项之前插入

        # 2. 向后查找相邻的有匹配表头
        next_idx = llm_idx + 1
        while next_idx < len(row_headers):
            if next_idx in llm_to_ocr_match:
                ocr_row, score = llm_to_ocr_match[next_idx]
                return ('before', ocr_row)  # 在后一个匹配表头之前插入
            next_idx += 1

        # 3. 向前查找相邻的有匹配表头
        prev_idx = llm_idx - 1
        while prev_idx >= 0:
            if prev_idx in llm_to_ocr_match:
                ocr_row, score = llm_to_ocr_match[prev_idx]
                return ('after', ocr_row)  # 在前一个匹配表头之后插入
            prev_idx -= 1

        # 4. 都没有找到，使用默认位置
        return ('end', -1)

    def _calculate_header_match_score(self, header_text, table_row):
        """
        计算表头与表格行的匹配分数
        不使用硬编码的关键词
        """
        # 提取表头关键词（最后一个>>后的部分）
        if '>>' in header_text:
            match_text = header_text.split('>>')[-1].strip()
        else:
            match_text = header_text.strip()

        if not match_text:
            return 0.0

        best_score = 0.0
        for cell in table_row:
            if not cell:
                continue

            cell_text = str(cell).strip()

            # 使用模糊匹配
            similarity = self.calculate_similarity_v2(match_text, cell_text)

            # 考虑长度比例：短文本完全匹配更重要
            len_ratio = min(len(match_text), len(cell_text)) / max(len(match_text), len(cell_text), 1)
            adjusted_score = similarity * len_ratio

            if adjusted_score > best_score:
                best_score = adjusted_score

        return best_score

    def _analyze_header_hierarchy(self, row_headers):
        """
        分析表头的层级关系
        """
        hierarchy = {
            'parents': {},  # parent_idx -> [child_idx1, child_idx2, ...]
            'children': {},  # child_idx -> parent_idx
            'depths': {}  # idx -> depth
        }

        for i, header in enumerate(row_headers):
            header_text = str(header)

            # 计算深度（>>的数量）
            depth = header_text.count('>>') + 1
            hierarchy['depths'][i] = depth

            # 如果是子项，寻找父级
            if '>>' in header_text:
                # 提取父级路径
                parts = header_text.split('>>')
                parent_path = '>>'.join(parts[:-1])

                # 寻找父级索引
                for j in range(i - 1, -1, -1):
                    if str(row_headers[j]) == parent_path:
                        hierarchy['children'][i] = j
                        if j not in hierarchy['parents']:
                            hierarchy['parents'][j] = []
                        hierarchy['parents'][j].append(i)
                        break

        return hierarchy

    def _get_dynamic_threshold(self, header_text):
        """
        动态调整匹配阈值
        """
        # 短文本需要更高的匹配度
        text_len = len(header_text)

        if text_len <= 3:
            return 0.9  # 短文本要求高匹配度
        elif text_len <= 10:
            return 0.7
        else:
            return 0.6  # 长文本允许较低的匹配度

    def _find_row_in_new_table(self, new_table, ocr_row):
        """
        在新表格中查找原始OCR行的位置
        """
        for i, row in enumerate(new_table):
            # 这里需要根据实际情况实现
            # 可能需要额外的标记来追踪原始行号
            pass
        return -1

    def _merge_and_remove_columns(self, table):
        """
        合并行表头并删除列：
        1. 如果列标题行左侧有2个及以上连续空表头
        2. 对每一行：
           - 如果第一列不为空，则将后面连续空表头列的数据合并到第一列
           - 如果第一列为空，则不合并
        3. 删除后面连续的空表头列（保留第一列）
        """

        if not table or len(table) < 2:
            return table

        # 获取列标题行（第0行）
        header_row = table[0]

        # 统计左侧连续空表头的数量
        empty_header_count = 0
        for header in header_row:
            if not header or str(header).strip() == "":
                empty_header_count += 1
            else:
                break

        # 如果左侧连续空表头少于2个，不处理
        if empty_header_count < 2:
            print(f"左侧连续空表头数: {empty_header_count}，少于2个，不处理")
            return table

        # 第一步：逐行处理，条件性合并
        for i in range(1, len(table)):
            row = table[i]

            # 跳过完全空的行
            if not any(cell for cell in row if cell and str(cell).strip() != ""):
                continue

            # 如果第一列不为空，才进行合并
            if row[0] and str(row[0]).strip() != "":
                print(f"行{i}: 第一列='{row[0]}'，进行合并")

                # 构建合并后的内容
                merged_parts = [str(row[0]).strip()]

                # 添加后续连续空表头列的内容
                for j in range(1, empty_header_count):
                    if j < len(row) and row[j] and str(row[j]).strip() != "":
                        merged_parts.append(str(row[j]).strip())

                # 如果有多部分内容，用>>连接
                if len(merged_parts) > 1:
                    table[i][0] = ">>".join(merged_parts)
                    print(f"  合并为: '{table[i][0]}'")
            else:
                print(f"行{i}: 第一列为空，不合并")

        # 第二步：无论如何都删除后面连续的空表头列（保留第一列）
        # 要删除的列索引：1 到 empty_header_count-1
        columns_to_remove = list(range(1, empty_header_count))

        if columns_to_remove:
            print(f"删除列索引: {columns_to_remove}")

            # 对每一行，删除指定列
            for i in range(len(table)):
                # 从右向左删除，避免索引错位
                for col_idx in sorted(columns_to_remove, reverse=True):
                    if col_idx < len(table[i]):
                        del table[i][col_idx]

        return table

    # 新增：标记表格的代理方法
    def step7_create_marked_table(self, table, row_checks=None, col_checks=None):
        """代理到独立的标记表格处理器"""
        return self.marked_table_processor.create_marked_table(table, row_checks, col_checks)

    def step8_add_feature_marks(self, validated_table, validation_marks):
        """代理到独立的标记表格处理器"""
        return self.marked_table_processor.add_feature_marks(validated_table, validation_marks)

    def is_numeric_value(self, cell_value):
        """代理到独立的标记表格处理器"""
        return self.marked_table_processor.is_numeric_value(cell_value)

    def _is_pure_numeric(self, cell):
        """代理到独立的标记表格处理器"""
        return self.marked_table_processor._is_pure_numeric(cell)


    # ========== 完整流程 ==========
    def process_single_table(self, ocr_tables, llm_table_info):
        """
        处理单个表格的完整流程
        """
        print(f"\n{'=' * 60}")
        print(f"开始处理表格")
        print(f"{'=' * 60}")

        # 第3步：合并OCR表格
        merged_data = self.step3_merge_ocr_tables(ocr_tables, llm_table_info)
        if not merged_data:
            self.log_issue("表格合并失败")
            return None

        # 第4步：创建基础表格
        base_table = self.step4_create_base_data_table(merged_data)
        if not base_table:
            self.log_issue("创建基础表格失败")
            return None

        # 第5步：添加列标题
        col_headers = merged_data.get('headers', {}).get('cols', [])
        table_with_cols = self.step5_add_column_headers(base_table, col_headers)

        # 第6步：添加行标题
        row_headers = merged_data.get('headers', {}).get('rows', [])
        final_table = self.step6_add_row_headers_intelligent(
            table_with_cols,
            row_headers,
            table_boundaries=merged_data.get('table_boundaries')
        )

        # 第7步：合并行表头并删除列
        final_table = self._merge_and_remove_columns(final_table)

        # 第8步：添加行列标记（根据数据类型）
        marked_table = self.step7_create_marked_table(final_table)

        return marked_table


    def _clean_sheet_name(self, name):
        """
        清理Sheet名称，确保符合Excel要求
        Excel Sheet名称限制：
        1. 不超过31个字符
        2. 不能包含字符：: \ / ? * [ ]
        3. 不能为空
        4. 不能以'开头
        """
        if not name or not isinstance(name, str):
            return "未命名表格"

        # 替换非法字符
        illegal_chars = [':', '\\', '/', '?', '*', '[', ']']
        for char in illegal_chars:
            name = name.replace(char, '_')

        # 移除首尾空格
        name = name.strip()

        # 如果以'开头，移除
        if name.startswith("'"):
            name = name[1:]

        # 截断到31个字符
        if len(name) > 31:
            name = name[:31]

        # 确保不为空
        if not name:
            name = "未命名表格"

        return name

    def step9_save_to_excel(self, tables_data, output_file, table_names=None):
        """
        将多个表格保存到Excel，每个表格一个Sheet
        tables_data: 列表，每个元素是一个表格的完整数据
        output_file: 输出Excel文件路径
        table_names: 可选，表格名称列表，用于Sheet名称
        """
        import openpyxl
        from openpyxl.styles import Alignment, Font, Border, Side
        from openpyxl.utils import get_column_letter

        print(f"\n=== 第7步：保存到Excel ===")
        print(f"要保存{len(tables_data)}个表格到: {output_file}")

        # 如果有表格名称，显示它们
        if table_names:
            print(f"使用的表格名称列表: {table_names}")

        # 创建新的工作簿
        wb = openpyxl.Workbook()

        # 删除默认创建的Sheet
        if 'Sheet' in wb.sheetnames:
            default_sheet = wb['Sheet']
            wb.remove(default_sheet)

        # 处理每个表格
        for table_idx, table in enumerate(tables_data):
            if not table:
                print(f"表格{table_idx}为空，跳过")
                continue

            # 创建Sheet名称（Excel限制31字符）
            # 如果提供了table_names，使用表格名称，否则使用默认名称
            if table_names and table_idx < len(table_names):
                table_name = table_names[table_idx]
                # 清理Sheet名称
                sheet_name = self._clean_sheet_name(table_name)
            else:
                sheet_name = f"Table{table_idx + 1}"

            print(f"  表格{table_idx + 1}: {len(table)}行 × {len(table[0]) if table else 0}列 -> Sheet: '{sheet_name}'")

            # 检查Sheet名称是否重复
            original_name = sheet_name
            counter = 1
            while sheet_name in wb.sheetnames:
                sheet_name = f"{original_name}_{counter}"
                counter += 1

            # 创建工作表
            ws = wb.create_sheet(title=sheet_name)

            # 获取表格尺寸
            num_rows = len(table)
            num_cols = len(table[0]) if num_rows > 0 else 0

            # 填充数据到Excel
            for r in range(num_rows):
                for c in range(num_cols):
                    cell_value = table[r][c]

                    # Excel行号从1开始，列号从1开始
                    excel_row = r + 1
                    excel_col = c + 1

                    # 写入值
                    ws.cell(row=excel_row, column=excel_col, value=cell_value)

                    # 获取单元格对象设置样式
                    cell_obj = ws.cell(row=excel_row, column=excel_col)

                    # 设置对齐方式
                    if r == 0:  # 第0行是表头
                        cell_obj.font = Font(bold=True)
                        cell_obj.alignment = Alignment(horizontal='center', vertical='center')
                        # 浅灰色背景
                        from openpyxl.styles import PatternFill
                        cell_obj.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                    elif c == 0:  # 第0列是行表头
                        cell_obj.alignment = Alignment(horizontal='left', vertical='center')
                    else:  # 数据单元格
                        cell_obj.alignment = Alignment(horizontal='center', vertical='center')

            # 设置列宽
            for col in range(1, num_cols + 1):
                col_letter = get_column_letter(col)

                # 根据内容调整列宽
                max_length = 0
                for row in range(1, min(num_rows + 1, 50)):  # 只检查前50行
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value:
                        # 计算文本长度（中文算2个字符）
                        text_len = 0
                        for char in str(cell_value):
                            if '\u4e00' <= char <= '\u9fff':  # 中文字符
                                text_len += 2
                            else:
                                text_len += 1
                        max_length = max(max_length, text_len)

                # 设置列宽（最小8，最大50）
                width = min(max(max_length + 2, 8), 50)
                ws.column_dimensions[col_letter].width = width

            # 添加边框
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # 为整个表格添加边框
            for row in ws.iter_rows(min_row=1, max_row=num_rows, min_col=1, max_col=num_cols):
                for cell in row:
                    cell.border = thin_border

            # 冻结首行和首列（方便查看）
            ws.freeze_panes = 'B2'  # 冻结第1行和第1列

        # 如果没有表格，创建一个空Sheet
        if len(wb.sheetnames) == 0:
            ws = wb.create_sheet(title="空表格")
            ws['A1'] = "无表格数据"

        # 保存文件
        try:
            wb.save(output_file)
            print(f"\n✅ Excel文件保存成功: {output_file}")
            print(f"   共保存了 {len(wb.sheetnames)} 个工作表")

            # 显示Sheet列表
            print("   Sheet列表:")
            for i, sheet_name in enumerate(wb.sheetnames, 1):
                ws = wb[sheet_name]
                max_row = ws.max_row
                max_column = ws.max_column
                print(f"     {i}. {sheet_name}: {max_row}行 × {max_column}列")

            return True

        except Exception as e:
            print(f"\n❌ 保存Excel文件失败: {str(e)}")
            return False

    def _extract_page_number_from_image_path(self, image_path):
        """
        从图片路径中提取页码信息
        例如: "XXX_015.png" -> "P015"
              "514001_152.png" -> "P152"  # 假设152是页码
              "document_123_page_045.jpg" -> "P045"
              "img_001.png" -> "P001"
        """
        if not image_path:
            print("  图片路径为空")
            return ""

        # 获取文件名（不含路径和扩展名）
        file_name = os.path.basename(image_path)
        base_name = os.path.splitext(file_name)[0]

        print(f"解析图片文件名: '{base_name}'")

        # 先尝试从文件名中直接提取可能的页码（通常是最后一段数字）
        # 对于类似"514001_152"的文件名，152可能是页码
        parts = base_name.split('_')
        if len(parts) >= 2:
            last_part = parts[-1]
            if last_part.isdigit() and len(last_part) >= 2:
                # 假设最后一部分数字是页码
                page_num = last_part.zfill(3)  # 格式化为3位
                result = f"P{page_num}"
                print(f"  从最后一部分提取页码: {last_part} -> {result}")
                return result

        # 尝试匹配各种页码格式
        patterns = [
            r'_(\d{3})$',  # 匹配_015（3位数字结尾）
            r'_(\d{2})$',  # 匹配_15（2位数字结尾）
            r'[pP]age_?(\d+)',  # 匹配page_15, Page45等
            r'p(\d+)',  # 匹配p15, P123等
            r'(\d{3})[._]',  # 匹配数字后跟点或下划线
            r'\D(\d{3})\D',  # 匹配被非数字包围的3位数字
        ]

        for pattern in patterns:
            match = re.search(pattern, base_name)
            if match:
                page_num = match.group(1)
                # 格式化页码，确保有前导零
                page_num_formatted = page_num.zfill(3)  # 至少3位，如015
                result = f"P{page_num_formatted}"
                print(f"  正则匹配到页码: {page_num} -> {result} (模式: {pattern})")
                return result

        # 如果没有匹配到，尝试提取所有数字
        all_digits = re.findall(r'\d+', base_name)
        if all_digits:
            # 取最长的一组数字（通常页码较长）
            longest_digits = max(all_digits, key=len)
            if len(longest_digits) >= 2:  # 至少2位才认为是页码
                page_num_formatted = longest_digits.zfill(3)
                result = f"P{page_num_formatted}"
                print(f"  提取最长数字作为页码: {longest_digits} -> {result}")
                return result

        print(f"  未找到页码信息，返回空字符串")
        return ""

    def process_all_tables(self, ocr_result, llm_result, output_file="output.xlsx", final_output_file="final.xlsx",
                           image_path=None, bank_name="未知银行"):
        """
        完整处理流程：第1-7步 - 增强版（支持表格合并）
        主要改进：提取LLM表格的name字段作为Excel Sheet名称，并添加页码前缀
        image_path: 新增参数，原始图片路径，用于提取页码信息
        """
        print("开始表格重构流程...", image_path)

        # ========== 新增：提取图片页码信息 ==========
        page_prefix = ""
        if image_path:
            page_prefix = self._extract_page_number_from_image_path(image_path)
            print(f"从图片路径提取页码前缀: {page_prefix}")

        # ========== 第0步：数据预处理和验证 ==========
        # 🔥 修正：检查并修正LLM引用的表格索引
        self._fix_llm_table_references(ocr_result, llm_result)

        # ========== 第1步：准备数据 ==========
        ocr_result, llm_result = self.step1_prepare_data(ocr_result, llm_result)

        # ========== 第2步：提取表格数据 ==========
        ocr_tables, llm_tables = self.step2_extract_table_data(ocr_result, llm_result)

        # 调试信息
        print(f"OCR表格数量: {len(ocr_tables)}")
        print(f"LLM表格结构数量: {len(llm_tables)}")

        if not ocr_tables or not llm_tables:
            self.log_issue("提取表格数据失败")
            return False

        # ========== 第3步：列标题统一化处理 ==========
        print("\n=== 开始列标题统一化处理 ===")
        llm_tables = self._unify_headers_across_tables(llm_tables)

        # ========== 第4步：检测需要合并的表格 ==========
        merge_groups = self._detect_tables_to_merge(llm_tables)

        # 分离需要合并的表格和独立处理的表格
        all_indices = set(range(len(llm_tables)))
        merged_indices = set()

        # 收集所有合并组的索引
        for group in merge_groups:
            for idx in group:
                merged_indices.add(idx)

        independent_indices = all_indices - merged_indices

        # ========== 第5步：处理所有表格并收集结果 ==========
        all_final_tables = []  # 存储表格数据
        all_table_names = []  # 存储表格名称（用于Sheet名称）
        all_table_full_names = []  # 存储完整Sheet名称（带页码前缀）

        # 1. 处理合并的表格组
        for group_idx, group in enumerate(merge_groups):
            print(f"\n{'=' * 60}")
            print(f"处理合并表格组 {group_idx + 1}/{len(merge_groups)}")
            print(f"包含表格索引: {group}")
            print(f"{'=' * 60}")

            # 先按独立表格处理（保持原有行为）
            for table_idx in group:
                llm_table_info = llm_tables[table_idx]

                # 🔥 提取表格名称
                table_name = llm_table_info.get('name', f'表格{table_idx + 1}')

                # 🔥 构建完整Sheet名称（带页码前缀）
                full_table_name = f"{page_prefix}_{table_name}" if page_prefix else table_name

                all_table_names.append(table_name)  # 原始表格名称
                all_table_full_names.append(full_table_name)  # 完整Sheet名称

                print(f"处理表格 {table_idx + 1}/{len(llm_tables)} (合并组)")
                print(f"表格名称: '{table_name}'")
                print(f"完整Sheet名称: '{full_table_name}'")

                # 原有独立处理逻辑
                final_table = self.process_single_table(ocr_tables, llm_table_info)

                if final_table:
                    all_final_tables.append(final_table)

        # 2. 处理独立的表格
        for table_idx in independent_indices:
            print(f"\n处理表格 {table_idx + 1}/{len(llm_tables)} (独立)")

            llm_table_info = llm_tables[table_idx]

            print("XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXllm_table_info")
            print(llm_table_info)

            # 🔥 提取表格名称
            table_name = llm_table_info.get('name', f'表格{table_idx + 1}')

            # 🔥 构建完整Sheet名称（带页码前缀）
            full_table_name = f"{page_prefix}_{table_name}" if page_prefix else table_name

            all_table_names.append(table_name)  # 原始表格名称
            all_table_full_names.append(full_table_name)  # 完整Sheet名称

            print(f"表格名称: '{table_name}'")
            print(f"完整Sheet名称: '{full_table_name}'")

            # 原有独立处理逻辑
            final_table = self.process_single_table(ocr_tables, llm_table_info)

            if final_table:
                all_final_tables.append(final_table)

        # 检查是否生成了表格
        if not all_final_tables:
            self.log_issue("无表格数据生成")
            return False

        # 验证表格名称和表格数据数量一致
        if len(all_final_tables) != len(all_table_names):
            print(f"⚠️ 警告：表格数据({len(all_final_tables)})和表格名称({len(all_table_names)})数量不一致")
            # 补齐名称列表
            while len(all_table_names) < len(all_final_tables):
                default_name = f"表格{len(all_table_names) + 1}"
                all_table_names.append(default_name)
                full_name = f"{page_prefix}_{default_name}" if page_prefix else default_name
                all_table_full_names.append(full_name)

        print(f"\n所有表格名称列表:")
        for i, (name, full_name) in enumerate(zip(all_table_names, all_table_full_names)):
            print(f"  表格{i + 1}: '{name}' -> Sheet: '{full_name}'")

        # ========== 第6步：保存到Excel（使用带页码前缀的完整Sheet名） ==========
        print(f"\n=== 开始保存到Excel ===")
        print(f"输出文件: {output_file}")

        # 🔥 关键修改：传入带页码前缀的完整表格名称
        success = self.step9_save_to_excel(
            tables_data=all_final_tables,
            output_file=output_file,
            table_names=all_table_full_names  # 🔥 传入带页码前缀的完整表格名称
        )

        if not success:
            self.log_issue("保存原始Excel失败")
            return False

        # ========== 第7步：生成最终数据Excel ==========
        try:
            print(f"\n开始生成最终数据Excel...")
            print(f"最终数据输出路径: {final_output_file}")

            # 确保输出路径不为空
            if not final_output_file or final_output_file == "final.xlsx":
                # 自动生成最终数据文件名
                import os
                base_name = os.path.splitext(output_file)[0]
                final_output_file = f"{base_name}_final_data.xlsx"
                print(f"自动生成最终数据文件名: {final_output_file}")

            # 获取LLM表格数据
            llm_table_list = []
            if 'tables_structure' in llm_result:
                llm_table_list = llm_result['tables_structure'].get('tables', [])
            elif 'tables' in llm_result:
                llm_table_list = llm_result['tables']

            print(f"要转换的表格数: {len(all_final_tables)}")
            print(f"LLM表格元数据数: {len(llm_table_list)}")

            # 调用转换器（不使用table_names参数，因为FinalDataConverter不支持）
            print(f"注意：FinalDataConverter不支持自定义Sheet名称，使用默认Sheet名称")
            final_success = self.final_data_converter.batch_convert_tables(
                all_tables_data=all_final_tables,
                all_llm_tables=llm_table_list,
                output_excel_path=final_output_file,
                bank_name=bank_name
            )

            if final_success:
                print(f"✅ 最终数据Excel生成成功: {final_output_file}")
            else:
                print(f"⚠️ 最终数据Excel生成失败")

            # 只要原始表格保存成功就返回True
            return success

        except Exception as e:
            print(f"⚠️ 最终数据转换异常: {str(e)}")
            print(f"原始表格已保存: {output_file}")
            import traceback
            traceback.print_exc()
            return success  # 即使最终数据转换失败，也返回原始表格保存成功

        # ========== 第8步：输出统计信息 ==========
        print(f"\n{'=' * 60}")
        print(f"处理完成统计:")
        print(f"  成功处理表格: {len(all_final_tables)}个")
        print(f"  表格合并组: {len(merge_groups)}组")
        print(f"  警告: {len(self.warnings)}个")
        print(f"  问题: {len(self.issues)}个")
        print(f"  原始输出文件: {output_file}")
        if 'final_success' in locals():
            print(f"  最终数据文件: {final_output_file}")
        print(f"{'=' * 60}")

        return success


# ====================================
# 使用示例
# ====================================
def main():
    # 你的数据
    ocr_result = {}
    llm_result= {'success': True,
                 'image_info': {'image_path': 'E:\\Datas\\base_pros\\DocuVista\\test_codes\\pngs\\123.png',
                                'image_id': 'img_0e6447019e10'}, 'tables_structure': {'tables': [
            {'id': '1', 'ocr_tables': [0],
             'headers': {'cols': ['', 'b>>2024年12月31日', 'c>>2024年9月30日', 'd>>2024年6月30日', 'e>>2024年3月31日'],
                         'rows': ['可用资本（数额）>>1核心一级资本净额', '可用资本（数额）>>2一级资本净额',
                                  '可用资本（数额）>>3资本净额', '风险加权资产（数额）>>4风险加权资产合计',
                                  '风险加权资产（数额）>>4a风险加权资产合计（应用资本底线前）',
                                  '资本充足率>>5核心一级资本充足率（%）',
                                  '资本充足率>>5a核心一级资本充足率（%）（应用资本底线前）',
                                  '资本充足率>>6一级资本充足率（%）', '资本充足率>>6a一级资本充足率（%）（应用资本底线前）',
                                  '资本充足率>>7资本充足率（%）', '资本充足率>>7a资本充足率（%）（应用资本底线前）',
                                  '其他资本要求>>8储备资本要求（%）', '其他资本要求>>9逆周期资本要求（%）',
                                  '其他资本要求>>10系统重要性银行附加资本要求（%）',
                                  '其他资本要求>>11其他各级资本要求（%）（8+9+10）',
                                  '其他资本要求>>12满足最低资本要求后的可用核心一级资本净额占风险加权资产的比例（%）']}}]},
                 'processing_stats': {'analysis_time_sec': 13.35, 'ocr_tables_count': 1, 'visual_tables_count': 1,
                                      'token_usage': {'prompt_tokens': 1155, 'completion_tokens': 421,
                                                      'total_tokens': 1576}}}

    # 创建重构器并处理
    reconstructor = TableReconstructor()
    success = reconstructor.process_all_tables(
        ocr_result=ocr_result,
        llm_result=llm_result,
        output_file="../../../../test_codes/enhanced_table_analyzer/reconstructed_tables2.xlsx",
        bank_name="中国建设银行"
    )

    if success:
        print("✅ 表格重构成功！")
    else:
        print("❌ 表格重构失败！")


if __name__ == "__main__":
    main()